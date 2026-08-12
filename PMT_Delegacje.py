# =============================================================================
#  PMT DELEGACJE — EDYCJA NAWIGATOR
#  - SILNIK: oryginalny rdzeń obliczeniowy v72 (matematyka nietknięta).
#  - STAWKA: wybór pojemności silnika (≤900 cm³ → 0,89 zł | >900 cm³ → 1,15 zł).
#  - UI: duże logo PMT, gradientowe tytuły sekcji, boczna nawigacja z funkcjami,
#        szklany kalendarz, overlay WOW (orbita + rysujący się checkmark).
# =============================================================================

import sys
import os

# --- Bezpiecznik aktualizacji: DOWÓD, że Python naprawdę wystartował -------
# Skrypt aktualizacji (updater.bat) sprawdza ISTNIENIE i ŚWIEŻOŚĆ tego pliku,
# żeby odróżnić "program działa" od "proces istnieje, ale utknął na oknie
# błędu ładowania biblioteki" (np. "Failed to load Python DLL" — to okno
# NIE kończy procesu, tylko czeka na kliknięcie OK). Samo sprawdzenie przez
# tasklist tego nie odróżnia — proces bootloadera wciąż tam widnieje.
# Pisane MOŻLIWIE najwcześniej: jeśli ten kod się wykonał, Python na pewno
# wystartował, niezależnie od tego, co się stanie później w programie.
try:
    import tempfile as _tmp_rozruch
    with open(os.path.join(_tmp_rozruch.gettempdir(), "pmt_zyje.flag"), "w") as _f_rozruch:
        _f_rozruch.write(str(os.getpid()))
except Exception:
    pass

import re
import json
import math
import random
import calendar
import datetime
import traceback
import itertools
import webbrowser
import subprocess
import zipfile
import urllib.request
import urllib.parse
import tempfile
import shutil
import hashlib
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple, Set

from fpdf import FPDF
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QComboBox, QPushButton, QFrame,
    QGraphicsDropShadowEffect, QGraphicsOpacityEffect, QSizePolicy, QSizeGrip, QCalendarWidget,
    QProxyStyle, QStyle, QScrollArea, QFileDialog, QPlainTextEdit, QCheckBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QDialog, QSlider
)
from PyQt6.QtCore import (
    Qt, QThread, pyqtSignal, QPropertyAnimation, QEasingCurve, 
    QVariantAnimation, QTimer, QPoint, QRectF, QRect, QEvent, QPointF,
    pyqtProperty, QSequentialAnimationGroup, QParallelAnimationGroup
)
from PyQt6.QtGui import (
    QPixmap, QColor, QCursor, QPainter, QPainterPath, QPen, QBrush, 
    QFont, QRadialGradient, QLinearGradient, QConicalGradient, QIcon, QFontMetrics
)

os.environ["QT_ENABLE_HIGHDPI_SCALING"]   = "1"
os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"


class ThickCaretStyle(QProxyStyle):
    """Pogrubiony kursor tekstowy (caret) — 5px zamiast systemowego 1px,
    wyraźnie widoczny we wszystkich polach formularza."""
    def pixelMetric(self, metric, option=None, widget=None):
        if metric == QStyle.PixelMetric.PM_TextCursorWidth:
            return 5
        return super().pixelMetric(metric, option, widget)

# ============================================================================= 
#  KONFIGURACJA BIZNESOWA I FIZYCZNA (Kopia 1:1 z v72)
# ============================================================================= 

STAWKA_ZA_KM        = 0.60 
SREDNIA_PREDKOSC    = 65.0 
LIMIT_CZASU_MINUTY  = 8 * 60   # sufit dnia; ustaw_tryb_pracy() zmienia go w trybie wieczornym
MAX_KWOTA_DELEGACJI = 587.19 
MIN_KWOTA           = 50.0 

PROCENT_WLASNE    = 0.50 
PROCENT_SASIEDNIE = 0.50 

KARA_DYSTANSOWA       = 500.0   
MIN_SIECI             = 1       
MIN_DYSTANS_LINIA     = 10.0    

COOLDOWN_DNI = 35   

# --- TRYB PRACY (ustawiany z interfejsu przed generowaniem) -------------------
#  "tydzien"  — cykl tygodniowy: pon-pt, dzień pracy o ustawionej długości.
#  "wieczory" — wieczory i weekendy: start ok. 16:30, ostatnia placówka do
#               22:00 plus powrót do bazy, czyli 5,5 h na trasę. Sobota liczy
#               się jak zwykły dzień, niedziela tylko przy niedzielach handlowych.
TRYB_PRACY          = "tydzien"
NIEDZIELE_HANDLOWE  = False
GODZINA_STARTU_WIECZOR = "16:30"
GODZINA_KONCA_WIECZOR  = "22:00"

def ustaw_tryb_pracy(tryb: str = "tydzien", dlugosc_dnia_h: float = None,
                     niedziele_handlowe: bool = None):
    """Przestawia silnik na wybrany tryb. Wywoływane PRZED generowaniem tras.

    dlugosc_dnia_h dotyczy trybu tygodniowego (domyślnie 3 h — tyle ustaliliśmy
    jako realny czas trasy). W trybie wieczornym długość wynika z okna
    16:30-22:00 i nie jest konfigurowalna.
    """
    global TRYB_PRACY, NIEDZIELE_HANDLOWE, LIMIT_CZASU_MINUTY, PRZERWA_JEDZENIE_MIN
    TRYB_PRACY = "wieczory" if str(tryb).startswith("wiecz") else "tydzien"
    if TRYB_PRACY == "wieczory":
        # Okno 16:30-22:00 to sufit dnia; wewnątrz niego silnik dalej miksuje
        # dni typu gniazdo / po drodze / daleki — tak jak w cyklu tygodniowym.
        LIMIT_CZASU_MINUTY = 5 * 60 + 30
        PRZERWA_JEDZENIE_MIN = 0              # krótkie okno, bez przerwy obiadowej
        NIEDZIELE_HANDLOWE = True             # jeśli w miesiącu są — silnik je wykorzysta
    else:
        LIMIT_CZASU_MINUTY = 8 * 60           # sufit dnia; realna długość wynika z typu dnia
        PRZERWA_JEDZENIE_MIN = 30
        NIEDZIELE_HANDLOWE = False
    return TRYB_PRACY, LIMIT_CZASU_MINUTY
TEST_MNOZNIK_TRASY = 1.28 

# --- Realia dnia pracy merchandisera (trasa "po drodze", nie nabijanie km) ---
PRZERWA_JEDZENIE_MIN   = 30      # posiłek/odpoczynek wliczony w 8h
POSTOJ_MIN_MIN         = 12      # min. czas w miejscowości
POSTOJ_MAX_MIN         = 23      # max. czas w miejscowości
MAX_MIEJSCOWOSCI_DZIEN = 7       # górny limit punktów dziennie
MIN_MIEJSCOWOSCI_DZIEN = 4       # dolny cel punktów dziennie
# Kolejny przystanek musi leżeć "po drodze": maksymalny skok między sąsiednimi
# punktami trasy. Trzyma pętlę zwartą — koniec z Ciechanów→Małdyty→Strzegowo→Iława.
MAX_SKOK_MIEDZY_PUNKTAMI_KM = 35.0

def max_skok_bazowy(promien):
    """Bazowy maksymalny skok między punktami — nieco rośnie w większym promieniu."""
    return MAX_SKOK_MIEDZY_PUNKTAMI_KM if promien <= 70 else min(MAX_SKOK_MIEDZY_PUNKTAMI_KM * (promien/70.0), 55.0)
# Maksymalny promień pętli od bazy — dzienna trasa nie ucieka na drugi kraniec woj.
MAX_PROMIEN_PETLI_KM        = 70.0

def zasob_sciezka(nazwa: str) -> str:
    """Zwraca ścieżkę do zasobu (grafiki tła, logo) działającą ZARÓWNO przy
    uruchomieniu ze źródeł, JAK I ze spakowanego .exe (PyInstaller --onefile).
    PyInstaller rozpakowuje dołączone pliki do katalogu tymczasowego _MEIPASS;
    gdy go nie ma, szukamy obok pliku programu."""
    if hasattr(sys, "_MEIPASS"):
        p = os.path.join(sys._MEIPASS, nazwa)
        if os.path.exists(p):
            return p
    # obok skryptu/exe
    baza = os.path.dirname(os.path.abspath(sys.argv[0]))
    return os.path.join(baza, nazwa)

# Logi i cache trzymamy OBOK programu / w katalogu użytkownika (zapisywalne
# miejsca — _MEIPASS jest tylko do odczytu i znika po zamknięciu .exe).
BASE_DIR  = os.path.dirname(os.path.abspath(sys.argv[0])) 
LOGS_DIR  = os.path.join(BASE_DIR, "logs") 
LOG_FILE  = os.path.join(LOGS_DIR, "error.log") 
GEO_CACHE = os.path.join(os.path.expanduser("~"), ".pmt_geo_cache.json")
NOTATKI_DNI_STORE = os.path.join(os.path.expanduser("~"), ".pmt_notatki_dni.json")
# Notatka dnia i "dzień wolny" — OGÓLNE, powiązane wyłącznie z datą (nie z
# imieniem/PESEL jak w starym kalendarzu delegacji). Dzięki temu: (1) nie
# giną przy zmianie danych osobowych, (2) planer wizyt może je uwzględniać
# i realnie omijać dni wolne przy układaniu tras. 


# ~~~~~~~~~~~~~~~~~~~~ POCZATEK SEKCJI PMT-ONLINE ~~~~~~~~~~~~~~~~~~~~
# =============================================================================
#  PMT ONLINE — łączność z arkuszem Google (sekcja do wklejenia do
#  PMT_Delegacje.py albo trzymana jako osobny moduł pmt_online.py)
#
#  ZASADA DZIAŁANIA (offline-first, zgodnie z ustaleniem):
#   • Program NIGDY nie czeka na internet. Wszystkie zdarzenia (uruchomienie,
#     utworzony dokument, minuty pracy) dopisują się do lokalnej kolejki
#     (~/.pmt_kolejka.json).
#   • Przy starcie i co SYNC_CO_MINUT program próbuje w tle wysłać kolejkę
#     jednym "pulsem". Udało się → kolejka się zeruje, a odpowiedź serwera
#     (do kiedy sesja, imię, rejon, nieobecności) ląduje w lokalnej pamięci
#     podręcznej (~/.pmt_status.json). Nie udało się → nic się nie dzieje,
#     spróbujemy później; dane NIE giną.
#   • Ważność sesji offline ocenia się po ostatnim ZNANYM "wazne_do"
#     z pamięci podręcznej — czyli przedłużenie w arkuszu dotrze do
#     użytkownika przy pierwszym kontakcie z internetem.
#
#  INTEGRACJA — trzy punkty zaczepienia w PMT_Delegacje.py:
#   1. Start programu (main, po utworzeniu QApplication):
#        online_zdarzenie(uruchomienia=1)
#        online_synchronizuj_w_tle()
#        # + jeśli online_kod_uzytkownika() is None → pokaż dialog
#        #   z pytaniem o 5-cyfrowy kod i zapisz go online_zapisz_kod(kod)
#   2. Po wygenerowaniu PDF-a delegacji:
#        online_zdarzenie(dokumenty=1)
#   3. Przy zamykaniu programu (closeEvent):
#        online_zdarzenie(minuty=<czas_sesji_w_minutach>)
#        online_synchronizuj()   # ostatnia próba, bez czekania na wynik
#
#  Nieobecności dla planera tras: online_nieobecnosci() → lista słowników
#  {kod, od, do, typ, zastepuje} — bieżące i przyszłe.
# =============================================================================
import time
import threading   # (pozostale importy sa juz na gorze programu)

# Adres wdrożenia Apps Script (kończy się na /exec) — wklej swój:
URL_BACKENDU = "https://script.google.com/macros/s/AKfycbyKUur4Fhs90_I8w5qqB39V5nqSQgQYixkItuxvT3gu4HokQlaMz3yXbiL06S2jZgoYAw/exec"

PLIK_KOLEJKI = os.path.join(os.path.expanduser("~"), ".pmt_kolejka.json")
PLIK_STATUSU = os.path.join(os.path.expanduser("~"), ".pmt_status.json")
SYNC_CO_MINUT = 15
_blokada = threading.Lock()

# --- pliki pomocnicze --------------------------------------------------------
def _wczytaj(plik, domyslne):
    try:
        with open(plik, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return dict(domyslne)

def _zapisz(plik, dane):
    try:
        tmp = plik + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(dane, f, ensure_ascii=False)
        os.replace(tmp, plik)          # zapis atomowy — brak "urwanych" plików
    except Exception:
        pass

# --- kod użytkownika ---------------------------------------------------------
def online_kod_uzytkownika():
    """Zwraca zapisany 5-cyfrowy kod albo None (trzeba zapytać w dialogu)."""
    kod = str(_wczytaj(PLIK_STATUSU, {}).get("kod", "")).strip()
    return kod if kod.isdigit() and len(kod) == 5 else None

def online_zapisz_kod(kod: str):
    st = _wczytaj(PLIK_STATUSU, {})
    st["kod"] = str(kod).strip()
    _zapisz(PLIK_STATUSU, st)

# --- kolejka zdarzeń (offline-first) -----------------------------------------
def online_zdarzenie(uruchomienia=0, dokumenty=0, minuty=0):
    """Dopisuje zdarzenia do lokalnej kolejki. Błyskawiczne, bez internetu."""
    with _blokada:
        q = _wczytaj(PLIK_KOLEJKI, {"uruchomienia": 0, "dokumenty": 0, "minuty": 0})
        q["uruchomienia"] = int(q.get("uruchomienia", 0)) + uruchomienia
        q["dokumenty"]    = int(q.get("dokumenty", 0))    + dokumenty
        q["minuty"]       = int(q.get("minuty", 0))       + round(minuty)
        _zapisz(PLIK_KOLEJKI, q)

# --- synchronizacja ----------------------------------------------------------
def online_synchronizuj() -> bool:
    """Jedna próba wysłania pulsu. True = sukces (kolejka wyzerowana,
    status odświeżony). Ciche niepowodzenie = zostajemy przy danych
    z pamięci podręcznej. NIE wywoływać z wątku interfejsu — patrz
    online_synchronizuj_w_tle()."""
    kod = online_kod_uzytkownika()
    if not kod or "TU_WKLEJ" in URL_BACKENDU:
        return False
    with _blokada:
        q = _wczytaj(PLIK_KOLEJKI, {"uruchomienia": 0, "dokumenty": 0, "minuty": 0})
    try:
        cialo = json.dumps({
            "akcja": "puls", "kod": kod,
            "wersja": globals().get("WERSJA_PROGRAMU", "?"),
            "system": {"nt": "Windows"}.get(os.name,
                      "macOS" if __import__("sys").platform == "darwin" else "Linux"),
            "uruchomienia": q.get("uruchomienia", 0),
            "dokumenty":    q.get("dokumenty", 0),
            "minuty":       q.get("minuty", 0),
        }).encode("utf-8")
        req = urllib.request.Request(
            URL_BACKENDU, data=cialo,
            headers={"Content-Type": "application/json", "User-Agent": "PMT-Planer"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            odp = json.loads(resp.read().decode("utf-8", errors="ignore"))
    except Exception:
        return False

    if odp.get("status") not in ("ok", "wygasla"):
        return False           # np. zly_kod — kolejki nie zerujemy

    with _blokada:
        _zapisz(PLIK_KOLEJKI, {"uruchomienia": 0, "dokumenty": 0, "minuty": 0})
        st = _wczytaj(PLIK_STATUSU, {})
        st.update({
            "kod": kod,
            "status": odp.get("status"),
            "imie": odp.get("imie", ""),
            "rejon": odp.get("rejon", ""),
            "wazne_do": odp.get("wazne_do", ""),
            "nieobecnosci": odp.get("nieobecnosci", []),
            "ostatnia_synchronizacja": datetime.date.today().isoformat(),
        })
        _zapisz(PLIK_STATUSU, st)
    return True

def online_synchronizuj_w_tle():
    """Odpala synchronizację w osobnym wątku — start programu nic nie czeka."""
    threading.Thread(target=online_synchronizuj, daemon=True).start()

def online_petla_synchronizacji():
    """Wywołać RAZ przy starcie: ponawia synchronizację co SYNC_CO_MINUT,
    więc dane 'wypuszczają się' same, gdy tylko wróci internet."""
    def _petla():
        while True:
            time.sleep(SYNC_CO_MINUT * 60)
            online_synchronizuj()
    threading.Thread(target=_petla, daemon=True).start()

# --- status sesji i nieobecności (czytane z pamięci podręcznej) --------------
def online_status_sesji():
    """Zwraca (czy_wazna, dni_pozostalo, imie). Działa w pełni offline —
    ocenia po ostatnim znanym 'wazne_do'. Brak jakiegokolwiek kontaktu
    z serwerem w historii → (None, None, '') i wtedy program stosuje
    dotychczasową lokalną zasadę 30 dni (demo_status)."""
    st = _wczytaj(PLIK_STATUSU, {})
    wazne_do = st.get("wazne_do", "")
    if not wazne_do:
        return None, None, st.get("imie", "")
    try:
        data = datetime.date.fromisoformat(wazne_do)
    except Exception:
        return None, None, st.get("imie", "")
    dni = (data - datetime.date.today()).days
    return dni >= 0, max(0, dni), st.get("imie", "")

def online_nieobecnosci():
    """Lista nieobecności (bieżące i przyszłe) z ostatniej synchronizacji:
    [{kod, od, do, typ, zastepuje}, ...] — do użycia w planerze tras."""
    return _wczytaj(PLIK_STATUSU, {}).get("nieobecnosci", [])


PLIK_LOGOWAN = os.path.join(os.path.expanduser("~"), ".pmt_logowania.json")


def historia_logowan(limit: int = 3):
    """Ostatnie konta używane na TYM komputerze — do podpowiedzi w oknie
    logowania. Lista słowników: kod, imie, ostatnio (ISO). Bez haseł."""
    dane = _wczytaj(PLIK_LOGOWAN, {})
    wpisy = []
    for kod, w in dane.items():
        if not (str(kod).isdigit() and len(str(kod)) == 5):
            continue
        wpisy.append({"kod": str(kod), "imie": str(w.get("imie", "")),
                      "ostatnio": str(w.get("ostatnio", ""))})
    wpisy.sort(key=lambda w: w["ostatnio"], reverse=True)
    return wpisy[:limit]


def _zapisz_logowanie(kod: str, imie: str, skrot: str):
    """Dopisuje konto do historii urządzenia. Skrót hasła pozwala zalogować
    się później BEZ internetu — samo hasło nigdy nie jest zapisywane."""
    dane = _wczytaj(PLIK_LOGOWAN, {})
    dane[str(kod)] = {"imie": str(imie or ""), "skrot": str(skrot),
                      "ostatnio": datetime.datetime.now().isoformat(timespec="seconds")}
    _zapisz(PLIK_LOGOWAN, dane)


def _hash_hasla(kod: str, haslo: str) -> str:
    """Skrót SHA-256 z solą (kodem). Hasło NIGDY nie jest zapisywane jawnie —
    ani na dysku, ani w arkuszu. Ten sam wzór stosuje backend, więc skróty
    lokalne i te z arkusza są porównywalne."""
    return hashlib.sha256(("PMT|" + str(kod) + "|" + str(haslo)).encode("utf-8")).hexdigest()


def online_zaloguj(kod: str, haslo: str):
    """Weryfikuje logowanie. Zwraca (czy_ok, imie, komunikat).

    ONLINE  — pyta arkusz (akcja "logowanie"): sprawdzane jest hasło albo,
              gdy użytkownik go jeszcze nie ustawił, numer telefonu.
              Po udanym logowaniu zapisujemy lokalnie SKRÓT hasła, żeby
              następnym razem dało się wejść bez zasięgu.
    OFFLINE — porównuje podane hasło ze skrótem zapisanym przy poprzednim
              udanym logowaniu na tym komputerze. Pierwsze logowanie na
              danym sprzęcie zawsze wymaga internetu (świadoma decyzja:
              inaczej blokada konta byłaby nie do wyegzekwowania).
    """
    kod = str(kod).strip()
    haslo = str(haslo)
    if not (kod.isdigit() and len(kod) == 5):
        return False, "", "Login to dokładnie 5 cyfr."
    if len(haslo) < 4:
        return False, "", "Podaj hasło (lub numer telefonu)."
    try:
        cialo = json.dumps({"akcja": "logowanie", "kod": kod,
                            "haslo": haslo, "telefon": haslo,
                            "zrodlo": "program"}).encode("utf-8")
        req = urllib.request.Request(
            URL_BACKENDU, data=cialo,
            headers={"Content-Type": "application/json", "User-Agent": "PMT-Planer"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            odp = json.loads(resp.read().decode("utf-8", errors="ignore"))
    except Exception as _blad:
        # BEZ SIECI: porównujemy skrót hasła z zapisanym przy poprzednim udanym
        # logowaniu — osobno dla każdego konta używanego na tym komputerze.
        znane = _wczytaj(PLIK_LOGOWAN, {}).get(kod, {})
        st = _wczytaj(PLIK_STATUSU, {})
        skrot = _hash_hasla(kod, haslo)
        if znane.get("skrot") == skrot:
            return True, str(znane.get("imie", "")), "offline"
        if st.get("kod") == kod and st.get("skrot") == skrot:
            return True, str(st.get("imie", "")), "offline"
        # Nie zgadujemy przyczyny: pokazujemy, co DOKŁADNIE zawiodło. Wcześniej
        # każdy błąd wyglądał jak brak internetu — a bywa nim blokada firmowa,
        # certyfikat proxy albo przeciążony serwer.
        _powod = _opisz_blad_sieci(_blad)
        try:
            with open(os.path.join(os.path.expanduser("~"), ".pmt_logowanie_bledy.txt"),
                      "a", encoding="utf-8") as _f:
                _f.write(f"{datetime.datetime.now().isoformat(timespec='seconds')} "
                         f"kod={kod} {type(_blad).__name__}: {_blad}\n")
        except Exception:
            pass
        return False, "", _powod
    if odp.get("status") != "ok":
        return False, "", str(odp.get("opis") or "Logowanie nie powiodło się.")
    st = _wczytaj(PLIK_STATUSU, {})
    st["kod"] = kod
    st["skrot"] = _hash_hasla(kod, haslo)
    st["imie"] = str(odp.get("imie", ""))
    _zapisz(PLIK_STATUSU, st)
    _zapisz_logowanie(kod, st["imie"], st["skrot"])
    return True, st["imie"], ""


def _opisz_blad_sieci(blad) -> str:
    """Zamienia wyjątek sieciowy na wskazówkę, co realnie zrobić."""
    tekst = f"{type(blad).__name__}: {blad}".lower()
    if "certificate" in tekst or "ssl" in tekst:
        return ("Połączenie zablokowane przez zabezpieczenia sieci (certyfikat). "
                "Zwykle to firmowy filtr lub VPN — spróbuj na innej sieci "
                "(np. hotspot z telefonu) albo poproś IT o odblokowanie "
                "adresu script.google.com.")
    if "timed out" in tekst or "timeout" in tekst:
        return ("Serwer nie odpowiedział w wyznaczonym czasie. Sprawdź połączenie "
                "i spróbuj ponownie — przy słabym zasięgu wystarczy druga próba.")
    if "name or service not known" in tekst or "getaddrinfo" in tekst or "dns" in tekst:
        return ("Komputer nie potrafi znaleźć adresu serwera (DNS). Sprawdź, czy "
                "internet działa w przeglądarce, i spróbuj ponownie.")
    if "forbidden" in tekst or "403" in tekst:
        return ("Serwer odrzucił połączenie (403). Najczęściej blokuje je sieć "
                "firmowa lub VPN — spróbuj na innej sieci.")
    if "proxy" in tekst:
        return ("Połączenie idzie przez serwer proxy, który je blokuje. Poproś IT "
                "o dostęp do script.google.com.")
    return (f"Nie udało się połączyć z serwerem ({type(blad).__name__}). "
            "Sprawdź internet i spróbuj ponownie; jeśli błąd wraca, prześlij "
            "administratorowi plik .pmt_logowanie_bledy.txt z katalogu domowego.")


def test_polaczenia() -> str:
    """Sprawdza po kolei: internet, dostęp do serwera i odpowiedź backendu.
    Zwraca gotowy tekst do pokazania użytkownikowi."""
    wyniki = []
    for opis, url in (("internet", "https://www.google.com/generate_204"),
                      ("serwer Google Apps Script", "https://script.google.com")):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "PMT-Planer"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                wyniki.append(f"✔ {opis}: odpowiada ({resp.status})")
        except Exception as e:
            wyniki.append(f"✘ {opis}: {_opisz_blad_sieci(e).split('.')[0]}")
    try:
        cialo = json.dumps({"akcja": "puls", "kod": "00000"}).encode("utf-8")
        req = urllib.request.Request(URL_BACKENDU, data=cialo,
              headers={"Content-Type": "application/json", "User-Agent": "PMT-Planer"})
        with urllib.request.urlopen(req, timeout=20) as resp:
            resp.read()
            wyniki.append("✔ nasz serwer PMT: odpowiada")
    except Exception as e:
        wyniki.append(f"✘ nasz serwer PMT: {_opisz_blad_sieci(e).split('.')[0]}")
    return "\n".join(wyniki)


def online_imie_uzytkownika():
    """Imię i nazwisko przypisane do zalogowanego kodu (z arkusza, zapisane
    lokalnie przy logowaniu). Pusty napis, gdy nieznane."""
    return str(_wczytaj(PLIK_STATUSU, {}).get("imie", "")).strip()


def online_wyloguj():
    """Usuwa zapamiętany kod użytkownika (plik statusu zostaje — bez pola 'kod')."""
    try:
        dane = {}
        if os.path.exists(PLIK_STATUSU):
            with open(PLIK_STATUSU, "r", encoding="utf-8") as f:
                dane = json.load(f)
        dane.pop("kod", None)
        dane.pop("skrot", None)
        dane.pop("wazne_do", None)
        dane.pop("imie", None)
        with open(PLIK_STATUSU, "w", encoding="utf-8") as f:
            json.dump(dane, f, ensure_ascii=False)
        return True
    except Exception:
        return False


def dialog_logowania():
    """Okno logowania w stylu programu: LOGIN (5-cyfrowy kod) + HASŁO.
    Zwraca (kod, imie) albo (None, "") gdy użytkownik zrezygnował.

    Logowanie odbywa się przy KAŻDYM uruchomieniu programu — dzięki temu
    administrator widzi w arkuszu każde wejście do systemu, a dostęp można
    odebrać zdalnie (zmiana hasła albo daty w kolumnie "Wazne do")."""
    from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel,
                                 QLineEdit, QPushButton)
    from PyQt6.QtCore import Qt, QRegularExpression, QTimer
    from PyQt6.QtGui import QRegularExpressionValidator
    d = QDialog()
    d.setWindowTitle("PMT Planer \u2014 logowanie")
    d.setModal(True)
    d.setWindowFlags(Qt.WindowType.Dialog | Qt.WindowType.FramelessWindowHint)
    d.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
    d.setFixedSize(470, 410)
    d.setObjectName("PmtLogowanie")
    d.setStyleSheet('''
        #PmtKarta { background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                      stop:0 #0F172A, stop:1 #04121A);
                    border: 1px solid rgba(0,240,255,0.30); border-radius: 14px; }
        QLabel#tytul { color:#F8FAFC; font-family:'Segoe UI'; font-size:18px; font-weight:800;
                       background: transparent; }
        QLabel#pod   { color:#94A3B8; font-family:'Segoe UI'; font-size:12px; background: transparent; }
        QLabel#etyk  { color:#94A3B8; font-family:'Segoe UI'; font-size:10px; font-weight:700;
                       letter-spacing:1px; background: transparent; }
        QLabel#blad  { color:#F87171; font-family:'Segoe UI'; font-size:11px; font-weight:600;
                       background: transparent; }
        QLineEdit { background:#0B1320; color:#F8FAFC; border:1px solid rgba(255,255,255,0.20);
                    border-radius:8px; padding:9px; font-family:'Segoe UI'; }
        QLineEdit#kod { font-size:20px; font-weight:700; letter-spacing:8px; }
        QLineEdit#haslo { font-size:15px; letter-spacing:2px; }
        QLineEdit:focus { border:1px solid #00E4A1; }
        QPushButton#ok { background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                           stop:0 #00F0FF, stop:1 #00E4A1);
                         color:#050B14; font-family:'Segoe UI'; font-size:14px; font-weight:800;
                         border:none; border-radius:8px; padding:10px 24px; }
        QPushButton#ok:hover { background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                                 stop:0 #33F5FF, stop:1 #33EAB7); }
        QPushButton#ok:disabled { background: rgba(30,41,59,0.6); color:#475569; }
        QPushButton#anuluj { background:transparent; color:#94A3B8; font-family:'Segoe UI';
                             font-size:13px; font-weight:600;
                             border:1px solid rgba(255,255,255,0.20);
                             border-radius:8px; padding:10px 18px; }
        QPushButton#anuluj:hover { color:#E2E8F0; border-color: rgba(255,255,255,0.40); }
        QPushButton#chip { color:#94A3B8; background:rgba(255,255,255,0.05);
                           border:1px solid rgba(255,255,255,0.18); border-radius:12px;
                           font-family:'Segoe UI'; font-size:11px; font-weight:700; padding:5px 9px; }
        QPushButton#chip:hover { color:#00E4A1; border-color:#00E4A1; }
    ''')
    from PyQt6.QtWidgets import QFrame, QGraphicsDropShadowEffect
    zewn = QVBoxLayout(d); zewn.setContentsMargins(10, 10, 10, 10)
    karta = QFrame(); karta.setObjectName("PmtKarta")
    cien = QGraphicsDropShadowEffect(karta)
    cien.setBlurRadius(28); cien.setOffset(0, 6); cien.setColor(QColor(0, 0, 0, 180))
    karta.setGraphicsEffect(cien)
    zewn.addWidget(karta)
    ukl = QVBoxLayout(karta); ukl.setContentsMargins(26, 22, 26, 18); ukl.setSpacing(6)
    t = QLabel("Zaloguj si\u0119"); t.setObjectName("tytul"); ukl.addWidget(t)
    pod = QLabel("Login to Tw\u00f3j 5-cyfrowy kod, has\u0142o otrzymasz od administratora.")
    pod.setObjectName("pod"); pod.setWordWrap(True); ukl.addWidget(pod)
    ukl.addSpacing(8)

    # Podpowiedzi: konta używane wcześniej na tym komputerze — jedno kliknięcie
    # wstawia login i przeskakuje do hasła.
    ostatni_uzytkownicy = historia_logowan(3)
    if ostatni_uzytkownicy:
        e0 = QLabel("OSTATNIO NA TYM KOMPUTERZE"); e0.setObjectName("etyk")
        ukl.addWidget(e0)
        rzad_hist = QHBoxLayout(); rzad_hist.setSpacing(6)
        for w in ostatni_uzytkownicy:
            _imie = w["imie"].split()[0] if w["imie"] else ""
            if len(_imie) > 8:
                _imie = _imie[:7] + "."
            etykieta = w["kod"] + ((" · " + _imie) if _imie else "")
            chip = QPushButton(etykieta); chip.setObjectName("chip")
            chip.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            chip.setToolTip("Zaloguj jako " + (w["imie"] or w["kod"]))
            # bez tego Qt ściska przyciski i ucina tekst bez wielokropka
            chip.setMinimumWidth(chip.fontMetrics().horizontalAdvance(etykieta) + 30)
            rzad_hist.addWidget(chip)
            chip._kod = w["kod"]
        rzad_hist.addStretch(1)
        ukl.addLayout(rzad_hist)
        ukl.addSpacing(4)

    e1 = QLabel("LOGIN"); e1.setObjectName("etyk"); ukl.addWidget(e1)
    pole_kod = QLineEdit(); pole_kod.setObjectName("kod"); pole_kod.setMaxLength(5)
    pole_kod.setAlignment(Qt.AlignmentFlag.AlignCenter)
    pole_kod.setValidator(QRegularExpressionValidator(QRegularExpression(r"\d{0,5}")))
    pole_kod.setPlaceholderText("\u2022\u2022\u2022\u2022\u2022")
    ukl.addWidget(pole_kod)

    ukl.addSpacing(6)
    e2 = QLabel("HAS\u0141O"); e2.setObjectName("etyk"); ukl.addWidget(e2)
    pole_haslo = QLineEdit(); pole_haslo.setObjectName("haslo")
    pole_haslo.setEchoMode(QLineEdit.EchoMode.Password)
    pole_haslo.setAlignment(Qt.AlignmentFlag.AlignCenter)
    pole_haslo.setPlaceholderText("has\u0142o lub numer telefonu")
    ukl.addWidget(pole_haslo)

    blad = QLabel(" "); blad.setObjectName("blad"); blad.setWordWrap(True)
    blad.setMinimumHeight(30); ukl.addWidget(blad)

    rzad = QHBoxLayout()
    b_test = QPushButton("Sprawdź połączenie"); b_test.setObjectName("anuluj")
    b_test.setToolTip("Sprawdza internet i dostęp do serwera — przydatne, gdy logowanie nie przechodzi")
    rzad.addWidget(b_test)
    rzad.addStretch(1)
    b_anuluj = QPushButton("Zamknij"); b_anuluj.setObjectName("anuluj")
    b_ok = QPushButton("Zaloguj"); b_ok.setObjectName("ok"); b_ok.setEnabled(False)
    rzad.addWidget(b_anuluj); rzad.addSpacing(8); rzad.addWidget(b_ok)
    ukl.addLayout(rzad)

    wynik = {"kod": None, "imie": ""}

    def _sprawdz_pola():
        b_ok.setEnabled(len(pole_kod.text()) == 5 and len(pole_haslo.text()) >= 4)

    def _zatwierdz():
        if not b_ok.isEnabled():
            return
        b_ok.setEnabled(False); b_ok.setText("Sprawdzam\u2026")
        blad.setText(" ")
        QApplication.processEvents()
        ok, imie, komunikat = online_zaloguj(pole_kod.text(), pole_haslo.text())
        b_ok.setText("Zaloguj"); _sprawdz_pola()
        if ok:
            wynik["kod"] = pole_kod.text().strip()
            wynik["imie"] = imie
            d.accept()
        else:
            blad.setText(komunikat)
            pole_haslo.selectAll(); pole_haslo.setFocus()

    for _chip in d.findChildren(QPushButton):
        if _chip.objectName() == "chip":
            def _uzyj(_=False, kod=_chip._kod):
                pole_kod.setText(kod); pole_haslo.clear(); pole_haslo.setFocus()
            _chip.clicked.connect(_uzyj)

    pole_kod.textChanged.connect(lambda _: _sprawdz_pola())
    pole_haslo.textChanged.connect(lambda _: _sprawdz_pola())
    pole_kod.returnPressed.connect(lambda: pole_haslo.setFocus())
    pole_haslo.returnPressed.connect(_zatwierdz)
    def _test():
        b_test.setEnabled(False); b_test.setText("Sprawdzam…")
        blad.setText("Sprawdzam połączenie…")
        QApplication.processEvents()
        wynik = test_polaczenia()
        b_test.setEnabled(True); b_test.setText("Sprawdź połączenie")
        from PyQt6.QtWidgets import QMessageBox
        mb = QMessageBox(d)
        mb.setWindowTitle("Test połączenia")
        mb.setText("Wynik sprawdzenia:")
        mb.setInformativeText(wynik + "\n\nJeśli nasz serwer nie odpowiada, a internet "
                              "działa, blokuje go najprawdopodobniej sieć firmowa lub VPN.")
        mb.exec()
        blad.setText(" ")
    b_test.clicked.connect(_test)
    b_ok.clicked.connect(_zatwierdz)
    b_anuluj.clicked.connect(d.reject)

    # wygodne wznowienie pracy: login podpowiadamy, hasło zawsze od nowa
    ostatni = str(_wczytaj(PLIK_STATUSU, {}).get("kod", "")).strip()
    if ostatni.isdigit() and len(ostatni) == 5:
        pole_kod.setText(ostatni)
        QTimer.singleShot(0, pole_haslo.setFocus)
    else:
        QTimer.singleShot(0, pole_kod.setFocus)
    _sprawdz_pola()
    d.exec()
    return wynik["kod"], wynik["imie"]


# ~~~~~~~~~~~~~~~~~~~~ KONIEC SEKCJI PMT-ONLINE ~~~~~~~~~~~~~~~~~~~~~~

# =============================================================================
#  WERSJA DEMO — limit 30 dni od pierwszego uruchomienia
#  Przy pierwszym starcie zapisujemy ukryty, lekko zaszyfrowany plik z datą.
#  Przy każdym starcie sprawdzamy, czy nie minęło DEMO_DNI. To bariera
#  grzecznościowa (da się obejść cofnięciem zegara/usunięciem pliku) — ma
#  zniechęcać, nie stanowić twardego zabezpieczenia.
#
#  ODBLOKOWANIE NA STAŁE: kod aktywacyjny wpisany w panelu administratora
#  (pole na samym dole, bez fanfar) zapisuje w tym samym pliku znacznik
#  "ODBLOKOWANE" — od tej chwili ta instalacja jest ważna bezterminowo,
#  niezależnie od tego, ile dni minęło. WAŻNE: aktywuj kodem PRZED upływem
#  30 dni na własnym komputerze — po tym terminie program blokuje się
#  jeszcze zanim zdąży pokazać panel administratora (awaryjnie: usunięcie
#  pliku ~/.pmt_licencja.dat resetuje licznik od zera).
# =============================================================================
DEMO_DNI = 30
DEMO_PLIK = os.path.join(os.path.expanduser("~"), ".pmt_licencja.dat")
_DEMO_KLUCZ = 0x5A            # prosty XOR — zaciemnienie, nie kryptografia bankowa
KOD_AKTYWACYJNY = "PMT-KOMPAS-STALY"   # zmień na własny w dowolnym momencie

def _demo_szyfruj(tekst: str) -> bytes:
    surowe = tekst.encode("utf-8")
    return bytes(b ^ _DEMO_KLUCZ for b in surowe)

def _demo_odszyfruj(dane: bytes) -> str:
    return bytes(b ^ _DEMO_KLUCZ for b in dane).decode("utf-8", errors="ignore")

def _demo_zapisz(data_inst, odblokowane: bool):
    """Zapisuje plik licencji: data pierwszego uruchomienia + opcjonalny
    znacznik trwałego odblokowania. Próbuje też ukryć plik na Windows."""
    try:
        tresc = f"PMT|{data_inst.isoformat()}"
        if odblokowane:
            tresc += "|ODBLOKOWANE"
        with open(DEMO_PLIK, "wb") as f:
            f.write(_demo_szyfruj(tresc))
        if os.name == "nt":
            try:
                import ctypes
                ctypes.windll.kernel32.SetFileAttributesW(DEMO_PLIK, 0x02)  # HIDDEN
            except Exception:
                pass
    except Exception:
        pass

def _demo_wczytaj():
    """Zwraca (data_pierwszego_uruchomienia, czy_odblokowane) albo (None, False)."""
    if not os.path.exists(DEMO_PLIK):
        return None, False
    try:
        with open(DEMO_PLIK, "rb") as f:
            tresc = _demo_odszyfruj(f.read())
        czesci = tresc.split("|")
        if czesci[0] != "PMT" or len(czesci) < 2:
            return None, False
        data_inst = datetime.date.fromisoformat(czesci[1].strip())
        odblokowane = len(czesci) >= 3 and czesci[2].strip() == "ODBLOKOWANE"
        return data_inst, odblokowane
    except Exception:
        return None, False

def demo_status():
    """Zwraca (czy_wazna, pozostalo_dni). Przy pierwszym uruchomieniu zakłada
    plik licencji. Gdy plik zniknie/uszkodzony — traktujemy jak nową instalację.
    Jeśli wpisano poprawny kod aktywacyjny (panel administratora), instalacja
    jest ważna BEZTERMINOWO — ten warunek sprawdzamy jako pierwszy."""
    dzis = datetime.date.today()
    data_inst, odblokowane = _demo_wczytaj()
    if data_inst is None:
        data_inst = dzis
        _demo_zapisz(data_inst, False)
    if odblokowane:
        return (True, DEMO_DNI)
    minelo = (dzis - data_inst).days
    pozostalo = max(0, DEMO_DNI - minelo)
    return (minelo < DEMO_DNI, pozostalo)

def odblokuj_licencje_na_stale():
    """Wywoływane po wpisaniu poprawnego kodu aktywacyjnego w panelu
    administratora — instalacja staje się ważna bezterminowo."""
    data_inst, _ = _demo_wczytaj()
    if data_inst is None:
        data_inst = datetime.date.today()
    _demo_zapisz(data_inst, True)

# =============================================================================
#  AKTUALIZACJE (Wariant A — tylko powiadomienie, bez auto-podmiany .exe)
#  Program po cichu czyta numer najnowszej wersji z pliku na GitHubie (raw).
#  Jeśli jest nowsza niż WERSJA_PROGRAMU, w pasku pojawia się ikona prezentu.
#  Pobranie nowego .exe użytkownik wykonuje SAM (klik → strona Releases).
#
#  ▸ JAK WŁĄCZYĆ (gdy założysz repo na GitHubie):
#    1. Wrzuć do repo plik tekstowy, np. "wersja.txt", z jedną linią: 1.1.0
#       (opcjonalnie druga linia = krótki opis nowości).
#    2. Ustaw URL_WERSJI na jego adres RAW, np.:
#       https://raw.githubusercontent.com/TWOJ_LOGIN/TWOJE_REPO/main/wersja.txt
#    3. Ustaw URL_POBIERANIA na stronę Releases:
#       https://github.com/TWOJ_LOGIN/TWOJE_REPO/releases/latest
#  Dopóki URL_WERSJI jest puste, sprawdzanie jest wyłączone (nic się nie dzieje).
# =============================================================================
WERSJA_PROGRAMU = "3.15.0"
# Sygnatura silnika — zmieniana przy każdej istotnej poprawce logiki tras.
# Pozwala jednoznacznie sprawdzić w aplikacji (ekran "O programie"), czy
# uruchomiony .exe zawiera aktualny silnik, czy stary build z cache.
SYGNATURA_SILNIKA = "S11-2026-adresy-wiejskie | PLAN-1-planer-wizyt"
URL_WERSJI     = "https://raw.githubusercontent.com/anowicki0205-coder/pmt-planer/main/wersja.txt"
URL_POBIERANIA = "https://github.com/anowicki0205-coder/pmt-planer/releases/latest"
URL_API_RELEASE = "https://api.github.com/repos/anowicki0205-coder/pmt-planer/releases/latest"
# Skrypt podmiany pobierany ŚWIEŻO z GitHuba przy każdej aktualizacji — NIE jest
# zaszyty w .exe. Dzięki temu błąd w logice podmiany można naprawić edytując
# ten plik na GitHubie, bez wydawania nowej wersji programu — poprawka działa
# natychmiast u WSZYSTKICH użytkowników, niezależnie jaką wersję .exe mają.
URL_UPDATERA = "https://raw.githubusercontent.com/anowicki0205-coder/pmt-planer/main/updater.bat"
# Wersja folderowa programu (bez rozpakowywania przy starcie) uzywa innego skryptu:
URL_UPDATERA_FOLDER = "https://raw.githubusercontent.com/anowicki0205-coder/pmt-planer/main/updater_folder.bat"
# odpowiednik dla macOS i Linuksa — ta sama idea: poprawka na GitHubie
# działa natychmiast u wszystkich, bez wydawania nowej wersji programu
URL_UPDATERA_SH = "https://raw.githubusercontent.com/anowicki0205-coder/pmt-planer/main/updater.sh"


def czy_zamrozony() -> bool:
    """Czy działamy jako .exe (PyInstaller)? Tylko wtedy da się podmienić plik."""
    return bool(getattr(sys, "frozen", False))


def czy_wersja_folderowa() -> bool:
    """Wersja folderowa ma obok pliku .exe katalog "_internal" z bibliotekami.
    Wersja jednoplikowa rozpakowuje sie do katalogu tymczasowego (sys._MEIPASS
    wskazuje wtedy poza folder programu)."""
    try:
        if not getattr(sys, "frozen", False):
            return False
        katalog = os.path.dirname(os.path.abspath(sys.executable))
        return os.path.isdir(os.path.join(katalog, "_internal"))
    except Exception:
        return False


def sciezka_programu() -> str:
    """Pełna ścieżka do aktualnie działającego pliku .exe."""
    return os.path.abspath(sys.executable if czy_zamrozony() else __file__)


# --- Warstwa międzyplatformowa (Windows / macOS / Linux) --------------------
CZY_WINDOWS = (os.name == "nt")
CZY_MAC = (sys.platform == "darwin")
CZY_LINUX = (not CZY_WINDOWS and not CZY_MAC)


def otworz_w_systemie(sciezka: str):
    """Otwiera plik lub folder w domyślnym programie systemowym.
    Odpowiednik os.startfile, ale działający także na macOS i Linuksie."""
    try:
        if CZY_WINDOWS:
            os.startfile(sciezka)
        elif CZY_MAC:
            subprocess.Popen(["open", sciezka])
        else:
            subprocess.Popen(["xdg-open", sciezka])
    except Exception:
        pass


def sciezka_pulpitu() -> str:
    """Ścieżka do Pulpitu użytkownika, niezależnie od systemu i języka.
    Na polskim Linuksie katalog nazywa się „Pulpit”, nie „Desktop” —
    pytamy więc Qt (czyta konfigurację XDG), a dopiero potem zgadujemy.
    Gdy Pulpit nie istnieje (np. system bez środowiska graficznego),
    zwraca katalog domowy — program zawsze ma gdzie zapisywać."""
    try:
        from PyQt6.QtCore import QStandardPaths
        p = QStandardPaths.writableLocation(QStandardPaths.StandardLocation.DesktopLocation)
        if p and os.path.isdir(p):
            return p
    except Exception:
        pass
    for nazwa in ("Desktop", "Pulpit"):
        p = os.path.join(os.path.expanduser("~"), nazwa)
        if os.path.isdir(p):
            return p
    return os.path.expanduser("~")


def znajdz_plik_wydania() -> tuple:
    """Pyta GitHub API o najnowsze wydanie. Zwraca (url, czy_zip).

    PREFERUJE plik .zip nad gołym .exe. Powód: przeglądarki (Chrome, Edge)
    traktują pobieranie nieznanego, niepodpisanego pliku .exe jako podejrzane
    — potrafią zostawić pobieranie zawieszone jako „.crdownload" w
    nieskończoność albo po cichu je skasować po „skanowaniu". Plik .zip nie
    wywołuje tej blokady, bo sam w sobie nie jest programem wykonywalnym.
    """
    try:
        req = urllib.request.Request(URL_API_RELEASE,
                                     headers={"User-Agent": "PMT-Planer",
                                              "Accept": "application/vnd.github+json"})
        with urllib.request.urlopen(req, timeout=8) as resp:
            dane = json.loads(resp.read().decode("utf-8", errors="ignore"))
        assets = dane.get("assets", [])
        nazwy = [(a, str(a.get("name", "")).lower()) for a in assets]

        def _url(a):
            return a.get("browser_download_url", "")

        if CZY_WINDOWS:
            # 1) NAJPIERW archiwum wprost oznaczone jako windows — odporne na
            #    obecność innych .zip w wydaniu (np. paczek źródeł czy kompletu).
            #    Bez tego program mógł pobrać przypadkowy .zip bez pliku .exe.
            for a, n in nazwy:
                if n.endswith(".zip") and ("windows" in n or "win32" in n or "win64" in n):
                    return _url(a), True
            # 2) zapasowo: dowolny .zip poza linuksowym/macowym (stare wydania)
            for a, n in nazwy:
                if n.endswith(".zip") and "linux" not in n and "mac" not in n \
                        and "zrodl" not in n and "komplet" not in n and "source" not in n:
                    return _url(a), True
            # zapasowo: starsze wydania mogły mieć dołączony goły .exe
            for a, n in nazwy:
                if n.endswith(".exe"):
                    return _url(a), False
        elif CZY_MAC:
            for a, n in nazwy:
                if n.endswith(".zip") and ("mac" in n or "darwin" in n or "osx" in n):
                    return _url(a), True
        else:  # Linux
            for a, n in nazwy:
                if n.endswith(".zip") and "linux" in n and "komplet" not in n:
                    return _url(a), True
            for a, n in nazwy:
                if n.endswith(".appimage"):
                    return _url(a), False
    except Exception:
        pass
    return "", False


def znajdz_plik_exe() -> str:
    """Zachowane dla zgodności — zwraca sam URL (bez informacji o .zip)."""
    url, _ = znajdz_plik_wydania()
    return url


def pobierz_tekst_url(url: str, timeout=6) -> str:
    """Pobiera zawartość tekstową spod URL. Pusty string przy błędzie."""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "PMT-Planer"})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return resp.read().decode("utf-8", errors="ignore")
    except Exception:
        return ""


def przygotuj_skrypt_aktualizacji(nowy_plik: str, docelowy: str, pid: int):
    """Zwraca (ścieżka_do_bat, argumenty) gotowe do uruchomienia.

    NAJPIERW próbuje pobrać świeży skrypt z GitHuba (URL_UPDATERA) — to
    pozwala naprawiać błędy w logice podmiany bez wydawania nowej wersji
    programu. Jeśli pobieranie zawiedzie (brak sieci, plik jeszcze nie
    istnieje w repozytorium), używa wbudowanej, zapasowej wersji — program
    działa więc w każdych warunkach, tylko czasem bez najnowszych poprawek.
    """
    if not CZY_WINDOWS:
        # macOS / Linux — skrypt powłoki zamiast .bat, ta sama filozofia:
        # najpierw świeży z GitHuba, zapasowo wbudowany generator
        tekst = pobierz_tekst_url(URL_UPDATERA_SH)
        if tekst and tekst.startswith("#!/") and "PMT" in tekst and len(tekst) < 20000:
            sh = os.path.join(tempfile.gettempdir(), "pmt_updater_zdalny.sh")
            try:
                with open(sh, "w", encoding="utf-8") as f:
                    f.write(tekst)
                os.chmod(sh, 0o755)
                return sh, [nowy_plik, docelowy, str(pid)]
            except Exception:
                pass
        return zbuduj_skrypt_podmiany_sh(nowy_plik, docelowy, pid), []

    # Wersja folderowa ma wlasny skrypt — podmienia caly katalog programu,
    # a nie pojedynczy plik .exe.
    _url_skryptu = URL_UPDATERA_FOLDER if czy_wersja_folderowa() else URL_UPDATERA
    tekst = pobierz_tekst_url(_url_skryptu)
    # prosta weryfikacja, że to nasz skrypt, a nie np. strona błędu GitHuba
    if tekst and "setlocal enabledelayedexpansion" in tekst and len(tekst) < 20000:
        bat = os.path.join(tempfile.gettempdir(), "pmt_updater_zdalny.bat")
        try:
            with open(bat, "w", encoding="cp1250", errors="replace") as f:
                f.write(tekst)
            return bat, [nowy_plik, docelowy, str(pid)]
        except Exception:
            pass
    # zapasowo: wbudowany generator (zawsze dostępny, nawet offline)
    return zbuduj_skrypt_podmiany(nowy_plik, docelowy, pid), []


def zbuduj_skrypt_podmiany(nowy_plik: str, docelowy: str, pid: int) -> str:
    r"""Skrypt .bat, który PO ZAMKNIĘCIU programu podmienia plik .exe
    i uruchamia nową wersję.

    Windows blokuje działający plik .exe — program nie nadpisze sam siebie
    w locie. Skrypt czeka, aż proces zniknie, dopiero wtedy podmienia plik.

    ZABEZPIECZENIE: przed podmianą robi kopię STAREJ wersji. Po uruchomieniu
    nowej wersji SPRAWDZA, czy proces faktycznie wystartował i przeżył
    kilka sekund — jeśli nowa wersja nie potrafi się uruchomić (np. brakuje
    jej jakiejś biblioteki systemowej), skrypt SAM przywraca poprzednią,
    działającą wersję i o tym informuje. Użytkownik nigdy nie zostaje bez
    działającego programu.

    Skrypt PISZE DZIENNIK do %TEMP%\pmt_aktualizacja.log — gdyby coś poszło
    nie tak, w tym pliku widać dokładnie, na którym kroku i dlaczego.
    """
    log = os.path.join(tempfile.gettempdir(), "pmt_aktualizacja.log")
    bat = os.path.join(tempfile.gettempdir(), "pmt_aktualizacja.bat")
    tresc = rf"""@echo off
setlocal enabledelayedexpansion
set "NOWY={nowy_plik}"
set "CEL={docelowy}"
set "PID={pid}"
set "LOG={log}"
set "KOPIA=!CEL!.poprzednia"
for %%F in ("!CEL!") do set "NAZWA_EXE=%%~nxF"

title Aktualizacja PMT Planer
echo ============================== > "!LOG!"
echo Start: %DATE% %TIME%          >> "!LOG!"
echo Nowy plik: !NOWY!             >> "!LOG!"
echo Plik docelowy: !CEL!          >> "!LOG!"
echo PID starego programu: !PID!   >> "!LOG!"

echo.
echo   Aktualizuje PMT Planer...
echo   Program uruchomi sie sam za chwile.
echo.

rem === 1) Czekaj, az stary program naprawde sie zamknie ===
echo [1] Czekam na zamkniecie programu... >> "!LOG!"
for /l %%i in (1,1,60) do (
    tasklist /FI "PID eq !PID!" 2>nul | find "!PID!" >nul
    if errorlevel 1 goto zamkniety
    ping -n 3 127.0.0.1 >nul
)
echo [1] UWAGA: program nadal dziala po 2 minutach. >> "!LOG!"
:zamkniety
echo [1] Program zamkniety. >> "!LOG!"

rem === 2) Kopia zapasowa STAREJ, dzialajacej wersji — na wypadek gdyby ===
rem === nowa sie nie uruchomila. Bez tego nie byloby do czego wracac.   ===
echo [2] Zapisuje kopie poprzedniej wersji... >> "!LOG!"
if exist "!CEL!" (
    copy /y "!CEL!" "!KOPIA!" >nul 2>>"!LOG!"
)

rem === 3) Podmien plik (bywa jeszcze chwile zablokowany - ponawiamy) ===
echo [3] Podmieniam plik... >> "!LOG!"
for /l %%i in (1,1,20) do (
    copy /y "!NOWY!" "!CEL!" >nul 2>>"!LOG!"
    if not errorlevel 1 goto podmieniono
    echo     proba %%i nieudana, czekam... >> "!LOG!"
    ping -n 3 127.0.0.1 >nul
)

echo [3] BLAD: nie udalo sie podmienic pliku. >> "!LOG!"
echo.
echo   ================================================
echo    Nie udalo sie podmienic pliku programu.
echo    Stara wersja pozostala nienaruszona.
echo.
echo    Najczestsza przyczyna: program lezy w katalogu
echo    wymagajacym uprawnien administratora
echo    (np. C:\\Program Files).
echo.
echo    Nowa wersja czeka tutaj:
echo    !NOWY!
echo    Mozesz podmienic plik recznie.
echo.
echo    Szczegoly: !LOG!
echo   ================================================
echo.
pause
exit /b 1

:podmieniono
echo [3] Plik podmieniony. >> "!LOG!"
del /q "!NOWY!" >nul 2>&1

rem === 4) Usun stary znacznik "program zyje" (z ewentualnego poprzedniego ===
rem === uruchomienia) - zeby nie dac falszywego pozytywnego wyniku niżej.  ===
set "FLAGA=%TEMP%\pmt_zyje.flag"
if exist "!FLAGA!" del /q "!FLAGA!" >nul 2>&1

rem === 5) Krotka przerwa PRZED pierwszym uruchomieniem — realny przypadek  ===
rem === pokazal, ze odpalenie NATYCHMIAST po kopiowaniu czasem konczylo    ===
rem === sie bledem brakujacej biblioteki (antywirus/system jeszcze         ===
rem === "trzymal" swiezo skopiowany plik) — mimo ze plik byl w 100% OK:    ===
rem === reczne uruchomienie chwile pozniej dzialalo bez zarzutu.           ===
echo [4] Czekam chwile przed uruchomieniem (antywirus bywa jeszcze zajety plikiem)... >> "!LOG!"
ping -n 3 127.0.0.1 >nul
echo [4] Uruchamiam nowa wersje... >> "!LOG!"
start "" "!CEL!"

rem === 6) WERYFIKACJA PRZEZ ZNACZNIK — nie samo istnienie procesu!        ===
rem === Bootloader PyInstallera, ktoremu nie udalo sie wczytac biblioteki  ===
rem === Pythona, pokazuje natywne okno bledu Windows i CZEKA na klikniecie ===
rem === OK — proces w tym czasie WCIAZ ISTNIEJE (tasklist by go znalazl!). ===
rem === Prawdziwy dowod, ze program dziala: sam program zapisal znacznik  ===
rem === zaraz po starcie (patrz poczatek pliku .py). Brak znacznika =      ===
rem === program sie nie uruchomil, niezaleznie od tego, co widac w tasklist.
set "WYSTARTOWALA=0"
set "PROBA_2=0"
for /l %%i in (1,1,8) do (
    ping -n 2 127.0.0.1 >nul
    if exist "!FLAGA!" (
        set "WYSTARTOWALA=1"
        goto po_weryfikacji
    )
    if "!PROBA_2!"=="0" if %%i GEQ 3 (
        echo [4] Znacznik sie nie pojawil - probuje uruchomic ponownie (drugie podejscie)... >> "!LOG!"
        start "" "!CEL!"
        set "PROBA_2=1"
    )
)
:po_weryfikacji

if "!WYSTARTOWALA!"=="1" goto sukces
goto porazka

:sukces
echo [4] Nowa wersja dziala poprawnie. >> "!LOG!"
if exist "!KOPIA!" del /q "!KOPIA!" >nul 2>&1
echo [OK] Gotowe: %DATE% %TIME% >> "!LOG!"
(goto) 2>nul & del "%~f0"

:porazka
rem === 7) Nowa wersja NIE wystartowala — automatyczny odwrot ===
echo [4] BLAD: nowa wersja sie nie uruchomila. Przywracam poprzednia... >> "!LOG!"
if not exist "!KOPIA!" goto brak_kopii

copy /y "!KOPIA!" "!CEL!" >nul 2>>"!LOG!"
del /q "!KOPIA!" >nul 2>&1
start "" "!CEL!"
echo.
echo   ================================================================
echo    Nowa wersja programu nie uruchomila sie poprawnie.
echo    Automatycznie przywrocono poprzednia, dzialajaca wersje —
echo    powinna wlasnie sie otworzyc.
echo.
echo    Sprobuj zaktualizowac ponownie pozniej. Jesli problem
echo    sie powtorzy, skontaktuj sie z autorem programu.
echo    Szczegoly: !LOG!
echo   ================================================================
echo.
pause
exit /b 1

:brak_kopii
echo.
echo   ================================================================
echo    Nowa wersja programu nie uruchomila sie, a nie znalazlem
echo    kopii poprzedniej wersji do przywrocenia.
echo    Szczegoly: !LOG!
echo   ================================================================
echo.
pause
exit /b 1
"""
    with open(bat, "w", encoding="cp1250", errors="replace") as f:
        f.write(tresc)
    return bat


def zbuduj_skrypt_podmiany_sh(nowy_plik: str, docelowy: str, pid: int) -> str:
    r"""Odpowiednik zbuduj_skrypt_podmiany dla macOS i Linuksa.

    Ta sama logika co w .bat: poczekaj na zamknięcie starego procesu, zrób
    kopię działającej wersji, podmień plik, uruchom nową wersję i sprawdź
    znacznik pmt_zyje.flag — jeśli nowa wersja nie wystartowała, przywróć
    poprzednią. Dziennik trafia do katalogu tymczasowego, do pliku
    pmt_aktualizacja.log (na uniksach skrypt działa w tle, bez okna —
    log jest więc jedynym śladem przebiegu)."""
    log = os.path.join(tempfile.gettempdir(), "pmt_aktualizacja.log")
    flaga = os.path.join(tempfile.gettempdir(), "pmt_zyje.flag")
    sh = os.path.join(tempfile.gettempdir(), "pmt_aktualizacja.sh")
    tresc = f'''#!/bin/bash
# PMT Planer - skrypt aktualizacji (macOS / Linux)
NOWY="{nowy_plik}"
CEL="{docelowy}"
PID="{pid}"
LOG="{log}"
FLAGA="{flaga}"
KOPIA="$CEL.poprzednia"

echo "==============================" >  "$LOG"
echo "Start: $(date)"                 >> "$LOG"
echo "Nowy plik: $NOWY"               >> "$LOG"
echo "Plik docelowy: $CEL"            >> "$LOG"
echo "PID starego programu: $PID"     >> "$LOG"

[ -f "$NOWY" ] || {{ echo "[BLAD] Brak pobranego pliku" >> "$LOG"; exit 1; }}

# 1) czekaj, az stary proces sie zamknie (do ~2 min)
echo "[1] Czekam na zamkniecie programu..." >> "$LOG"
for i in $(seq 1 60); do
    kill -0 "$PID" 2>/dev/null || break
    sleep 2
done
echo "[1] OK." >> "$LOG"

# 2) kopia zapasowa starej, dzialajacej wersji
[ -f "$CEL" ] && cp -f "$CEL" "$KOPIA" 2>>"$LOG"

# 3) podmiana pliku (z ponawianiem - bywa chwile zablokowany)
echo "[3] Podmieniam plik..." >> "$LOG"
OK=0
for i in $(seq 1 15); do
    if cp -f "$NOWY" "$CEL" 2>>"$LOG"; then OK=1; break; fi
    sleep 1
done
if [ "$OK" != "1" ]; then
    echo "[3] BLAD: nie udalo sie podmienic pliku (uprawnienia?)." >> "$LOG"
    exit 1
fi
chmod +x "$CEL" 2>>"$LOG"
rm -f "$NOWY" 2>/dev/null

# 4) usun stary znacznik "program zyje" i uruchom nowa wersje
rm -f "$FLAGA" 2>/dev/null
sleep 2
echo "[4] Uruchamiam nowa wersje..." >> "$LOG"
nohup "$CEL" >/dev/null 2>&1 &

# 5) weryfikacja przez znacznik - tak samo jak w wersji Windows:
# istnienie procesu nic nie dowodzi, znacznik zapisany przez program - tak
WYSTARTOWALA=0
PROBA2=0
for i in $(seq 1 8); do
    sleep 2
    if [ -f "$FLAGA" ]; then WYSTARTOWALA=1; break; fi
    if [ "$PROBA2" = "0" ] && [ "$i" -ge 3 ]; then
        echo "[4] Znacznik sie nie pojawil - drugie podejscie..." >> "$LOG"
        nohup "$CEL" >/dev/null 2>&1 &
        PROBA2=1
    fi
done

if [ "$WYSTARTOWALA" = "1" ]; then
    echo "[OK] Nowa wersja dziala. $(date)" >> "$LOG"
    rm -f "$KOPIA" 2>/dev/null
    rm -f "$0" 2>/dev/null
    exit 0
fi

# 6) rollback - przywroc poprzednia, dzialajaca wersje
echo "[4] BLAD: nowa wersja sie nie uruchomila. Przywracam poprzednia..." >> "$LOG"
if [ -f "$KOPIA" ]; then
    cp -f "$KOPIA" "$CEL" 2>>"$LOG"
    chmod +x "$CEL" 2>>"$LOG"
    rm -f "$KOPIA" 2>/dev/null
    nohup "$CEL" >/dev/null 2>&1 &
    echo "[OK] Przywrocono poprzednia wersje." >> "$LOG"
fi
exit 1
'''
    with open(sh, "w", encoding="utf-8") as f:
        f.write(tresc)
    try:
        os.chmod(sh, 0o755)
    except Exception:
        pass
    return sh


def oblicz_statystyki_osobiste() -> dict:
    """Analizuje DZIENNIK (prawdziwe znaczniki czasu odhaczeń) i PLAN, żeby
    zbudować osobisty obraz pracy: ile wizyt, w jakim rytmie, które sieci
    dominują, jaki jest aktualny 'streak' ukończonych dni. To, co faktycznie
    się wydarzyło — nie prognoza, tylko fakty z przeszłości."""
    dz = _wczytaj_dziennik()
    dzis = datetime.date.today()

    # zbierz realne (adres, data) dla każdego wpisu z prawdziwym znacznikiem czasu
    wpisy = []
    for klucz, wpis in dz.items():
        if not wpis.get("czas") or "|" not in klucz:
            continue
        iso_data, adres = klucz.split("|", 1)
        try:
            data = datetime.date.fromisoformat(iso_data)
        except Exception:
            continue
        wpisy.append((data, adres))

    lacznie = len(wpisy)
    w_tym_miesiacu = sum(1 for d, _ in wpisy if d.year == dzis.year and d.month == dzis.month)
    pop_m = dzis.month - 1 or 12
    pop_r = dzis.year if dzis.month > 1 else dzis.year - 1
    poprzedni_miesiac = sum(1 for d, _ in wpisy if d.year == pop_r and d.month == pop_m)

    # wizyty per miesiąc — ostatnie 6 miesięcy (z zerami dla pustych, żeby wykres był porównywalny)
    miesiace_lista = []
    mm, rr = dzis.month, dzis.year
    for _ in range(6):
        miesiace_lista.append((rr, mm))
        mm -= 1
        if mm < 1:
            mm = 12; rr -= 1
    miesiace_lista.reverse()
    per_miesiac = []
    nazwy_mies = ["", "Sty", "Lut", "Mar", "Kwi", "Maj", "Cze", "Lip", "Sie", "Wrz", "Paź", "Lis", "Gru"]
    for rok, mies in miesiace_lista:
        n = sum(1 for d, _ in wpisy if d.year == rok and d.month == mies)
        per_miesiac.append((nazwy_mies[mies], n))

    # podział na sieci — dopasowanie adres -> sieć z aktualnej listy punktów
    punkty = wczytaj_punkty()
    mapa_siec = {(p.get("adres", "") or "").strip().lower(): (p.get("siec", "") or "Inne") for p in punkty}
    per_siec = {}
    for _, adres in wpisy:
        siec = mapa_siec.get(adres.strip().lower(), "Inne") or "Inne"
        per_siec[siec] = per_siec.get(siec, 0) + 1

    # rekord — najwięcej wizyt jednego dnia
    per_dzien = {}
    for d, _ in wpisy:
        per_dzien[d] = per_dzien.get(d, 0) + 1
    rekord_dzien = max(per_dzien.items(), key=lambda kv: kv[1]) if per_dzien else None

    # passa — ile dni z rzędu w 100% ukończone, licząc wstecz od dziś (z zapisanego planu)
    passa = 0
    plan = wczytaj_plan()
    suma_km_ukonczone = 0.0
    if plan and plan.get("dni"):
        dni_min = sorted((d for d in plan["dni"] if d.data <= dzis), key=lambda d: d.data)
        for d in dni_min:
            zrobione = sum(1 for wz in d.wizyty if czy_odwiedzona(d.data, wz.adres or wz.nazwa))
            if d.wizyty and zrobione >= len(d.wizyty):
                suma_km_ukonczone += d.km
        for d in reversed(dni_min):
            zrobione = sum(1 for wz in d.wizyty if czy_odwiedzona(d.data, wz.adres or wz.nazwa))
            if d.wizyty and zrobione >= len(d.wizyty):
                passa += 1
            else:
                break

    return {
        "lacznie_wizyt": lacznie,
        "w_tym_miesiacu": w_tym_miesiacu,
        "poprzedni_miesiac": poprzedni_miesiac,
        "passa_dni": passa,
        "suma_km_ukonczone": round(suma_km_ukonczone, 0),
        "per_miesiac": per_miesiac,
        "per_siec": dict(sorted(per_siec.items(), key=lambda kv: -kv[1])),
        "rekord_dzien": rekord_dzien,
    }


def wykryj_sugestie_cyklu(min_wizyt=3, tolerancja=0.45) -> list:
    """Analizuje DZIENNIK (prawdziwe znaczniki czasu odhaczeń — nie plan, tylko
    to, co faktycznie się wydarzyło) i wykrywa naturalny rytm powrotu do
    każdego punktu. Sugeruje dopasowanie do najbliższej standardowej opcji
    cyklu (1,2,3,4,6,8,12 tygodni), ale TYLKO gdy rytm jest wystarczająco
    spójny (mało rozrzutu) i różni się od obecnie ustawionego cyklu.

    Zero sztucznej inteligencji — czysta mediana odstępów między wizytami,
    ale sprawia wrażenie, że program uczy się Twojego rytmu pracy.
    """
    dz = _wczytaj_dziennik()
    wg_adresu = {}
    for klucz, wpis in dz.items():
        if not wpis.get("czas"):        # wpis bez znacznika = niekompletny/stary format
            continue
        if "|" not in klucz:
            continue
        iso_data, adres = klucz.split("|", 1)
        try:
            data = datetime.date.fromisoformat(iso_data)
        except Exception:
            continue
        wg_adresu.setdefault(adres, []).append(data)

    punkty = wczytaj_punkty()
    mapa_punktow = {(p.get("adres", "") or "").strip().lower(): p for p in punkty}

    OPCJE_TYG = [1, 2, 3, 4, 6, 8, 12]
    cykl_domyslny = int(ustawienie("cykl_domyslny", 4) or 4)
    cykle_sieci = ustawienie("cykle_sieci", {}) or {}
    cykle_sieci_duze = ustawienie("cykle_sieci_duze", {}) or {}
    cykle_punktow = ustawienie("cykle_punktow", {}) or {}

    sugestie = []
    for adres, daty in wg_adresu.items():
        daty = sorted(set(daty))
        if len(daty) < min_wizyt:
            continue
        odstepy = sorted((daty[i + 1] - daty[i]).days for i in range(len(daty) - 1))
        odstepy = [o for o in odstepy if o > 0]
        if not odstepy:
            continue
        n = len(odstepy)
        mediana = odstepy[n // 2] if n % 2 == 1 else (odstepy[n // 2 - 1] + odstepy[n // 2]) / 2
        if mediana <= 0:
            continue
        # spójność rytmu: jeśli odstępy skaczą bardzo nieregularnie, nie zgadujemy
        rozrzut = max(abs(o - mediana) for o in odstepy) / mediana
        if rozrzut > tolerancja:
            continue

        tyg_rzeczywiste = mediana / 7.0
        sugerowany = min(OPCJE_TYG, key=lambda t: abs(t - tyg_rzeczywiste))

        punkt = mapa_punktow.get(adres)
        siec = (punkt or {}).get("siec", "")
        miasto = (punkt or {}).get("miasto", "")
        nazwa = f"{siec} — {miasto}" if siec and miasto else (miasto or adres)

        if punkt:
            obecny = cykl_dla_punktu(_PunktLekki(punkt), cykl_domyslny,
                                     cykle_sieci, cykle_sieci_duze, cykle_punktow)
        else:
            obecny = cykl_domyslny

        if sugerowany == obecny:
            continue      # już tak ustawione — nic do zaproponowania

        sugestie.append({
            "adres": adres, "nazwa": nazwa, "siec": siec, "miasto": miasto,
            "liczba_wizyt": len(daty), "sredni_odstep_dni": round(mediana, 1),
            "obecny_cykl_tyg": obecny, "sugerowany_cykl_tyg": sugerowany,
        })

    sugestie.sort(key=lambda s: -s["liczba_wizyt"])   # najpewniejsze (najwięcej danych) pierwsze
    return sugestie


def eksportuj_plan_excel(plan: dict, sciezka: str):
    """Eksportuje plan wizyt WRAZ Z DZIENNIKIEM do jednego pliku Excel —
    jeden wiersz na wizytę, plus arkusz podsumowania. To, co faktycznie się
    wydarzyło (status, znacznik czasu, notatka) idzie obok tego, co było
    zaplanowane, żeby plik był kompletnym zapisem pracy, nie tylko planem."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Plan i dziennik"
    naglowki = ["Data", "Dzień tygodnia", "Sieć", "Nazwa", "Adres", "Miasto",
               "Cykl (tyg.)", "Status", "Odwiedzono (data i godz.)", "Notatka"]
    ws.append(naglowki)
    naglowek_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    naglowek_font = Font(bold=True, color="FFFFFF", name="Arial")
    for col in range(1, len(naglowki) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = naglowek_fill; c.font = naglowek_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    dni_pl = ["poniedziałek", "wtorek", "środa", "czwartek", "piątek", "sobota", "niedziela"]
    cykliczny = bool(plan.get("cykliczny"))
    cykl_domyslny = int(plan.get("cykl_domyslny", 4) or 4)
    cykle_sieci = plan.get("cykle_sieci", {}) or {}
    cykle_sieci_duze = plan.get("cykle_sieci_duze", {}) or {}
    cykle_punktow = plan.get("cykle_punktow", {}) or {}

    zwykla_czcionka = Font(name="Arial", size=10)
    zrobiona_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    for d in plan.get("dni", []):
        for wz in d.wizyty:
            adres_klucz = wz.adres or wz.nazwa
            wpis = wpis_wizyty(d.data, adres_klucz)
            odwiedzona = czy_odwiedzona(d.data, adres_klucz)
            tyg = cykl_dla_punktu(wz, cykl_domyslny, cykle_sieci, cykle_sieci_duze, cykle_punktow) if cykliczny else ""
            wiersz = [
                d.data.strftime("%d.%m.%Y"),
                dni_pl[d.data.weekday()].capitalize(),
                wz.siec or "",
                wz.nazwa or "",
                wz.adres or "",
                wz.miasto or "",
                tyg,
                "Zrobione ✓" if odwiedzona else "Zaplanowane",
                wpis.get("czas", ""),
                wpis.get("notatka", ""),
            ]
            ws.append(wiersz)
            nr = ws.max_row
            for col in range(1, len(naglowki) + 1):
                cell = ws.cell(row=nr, column=col)
                cell.font = zwykla_czcionka
                if odwiedzona:
                    cell.fill = zrobiona_fill

    szerokosci = [12, 14, 14, 26, 32, 16, 11, 14, 20, 34]
    for i, sz in enumerate(szerokosci, start=1):
        ws.column_dimensions[get_column_letter(i)].width = sz
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(naglowki))}{ws.max_row}"

    # --- arkusz podsumowania ---
    ws2 = wb.create_sheet("Podsumowanie")
    zrobione = sum(1 for d in plan.get("dni", []) for wz in d.wizyty
                   if czy_odwiedzona(d.data, wz.adres or wz.nazwa))
    dane_podsum = [
        ("Program", "PMT Planer"),
        ("Wersja", WERSJA_PROGRAMU),
        ("Wygenerowano", datetime.datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("Liczba wizyt (plan)", plan.get("suma_wizyt", 0)),
        ("Wizyty odhaczone", zrobione),
        ("Suma kilometrów", f"{plan.get('suma_km', 0):.0f} km"),
        ("Liczba miesięcy", plan.get("liczba_miesiecy", 0)),
        ("Tryb", "Cykliczny" if cykliczny else "Jednorazowy"),
    ]
    for etyk, wart in dane_podsum:
        ws2.append([etyk, wart])
    for row in ws2.iter_rows(min_row=1, max_row=len(dane_podsum), min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True, name="Arial")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 26

    wb.save(sciezka)


def eksportuj_plan_pdf(plan: dict, sciezka: str):
    """Eksportuje plan wizyt do PDF — czytelny, drukowalny wykaz dzień po
    dniu. Używa wbudowanego w Qt mechanizmu PDF (QPrinter) — bez dodatkowych
    zależności, bo cały program jest już zbudowany na Qt."""
    from PyQt6.QtPrintSupport import QPrinter
    from PyQt6.QtGui import QPageSize
    from PyQt6.QtCore import QMarginsF, QSizeF

    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
    printer.setOutputFileName(sciezka)
    printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
    # KLUCZOWE: 72 DPI sprawia, że 1 jednostka rysowania = 1 punkt (1/72 cala).
    # Bez tego QPainter rysuje w pikselach urządzenia przy rozdzielczości
    # druku (nawet >1000 DPI) — te same współrzędne co niżej byłyby wtedy
    # niewidocznymi kropkami w rogu strony (dokładnie taki błąd złapałem w teście).
    printer.setResolution(72)
    printer.setPageMargins(QMarginsF(18, 16, 18, 16), printer.pageLayout().units())

    painter = QPainter(printer)
    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
    strona = printer.pageRect(QPrinter.Unit.Point)
    W, H = strona.width(), strona.height()
    margines = 34
    y = [margines]     # w liście, żeby domykać w zagnieżdżonych funkcjach

    kol_txt = QColor("#0F172A")
    kol_mut = QColor("#64748B")
    kol_akc = QColor("#0D9488")
    kol_zie = QColor("#059669")

    def naglowek_strony():
        painter.setPen(kol_akc)
        painter.setFont(QFont("Arial", 13, QFont.Weight.Bold))
        painter.drawText(QRectF(margines, y[0], W - 2 * margines, 22), "PMT Planer — Plan Wizyt")
        painter.setPen(kol_mut)
        painter.setFont(QFont("Arial", 8))
        painter.drawText(QRectF(margines, y[0] + 20, W - 2 * margines, 16),
                         f"Wygenerowano {datetime.datetime.now().strftime('%d.%m.%Y %H:%M')}  •  "
                         f"{plan.get('suma_wizyt', 0)} wizyt  •  {plan.get('suma_km', 0):.0f} km łącznie")
        y[0] += 46
        painter.setPen(QPen(QColor("#CBD5E1"), 1))
        painter.drawLine(QPointF(margines, y[0]), QPointF(W - margines, y[0]))
        y[0] += 14

    dni_pl = ["poniedziałek", "wtorek", "środa", "czwartek", "piątek", "sobota", "niedziela"]
    naglowek_strony()
    dzien_w_toku = [None]     # etykieta dnia aktualnie renderowanego (do "cd.")

    def nowa_strona():
        printer.newPage()
        y[0] = margines
        naglowek_strony()
        if dzien_w_toku[0]:
            painter.setPen(kol_mut)
            painter.setFont(QFont("Arial", 9, QFont.Weight.Bold))
            painter.drawText(QRectF(margines, y[0], W - 2 * margines, 16), f"(cd.)  {dzien_w_toku[0]}")
            y[0] += 20

    def potrzebne_miejsce(wys):
        if y[0] + wys > H - margines:
            nowa_strona()

    for d in plan.get("dni", []):
        dzien_w_toku[0] = None    # nowy dzień — jeśli akurat tu wypadnie łamanie strony, NIE pokazuj "(cd.)" poprzedniego
        potrzebne_miejsce(30)
        painter.setPen(kol_akc)
        painter.setFont(QFont("Arial", 11, QFont.Weight.Bold))
        etyk_dnia = f"{dni_pl[d.data.weekday()].capitalize()}, {d.data.strftime('%d.%m.%Y')}"
        dzien_w_toku[0] = etyk_dnia
        painter.drawText(QRectF(margines, y[0], W - 2 * margines, 18), etyk_dnia)
        painter.setPen(kol_mut)
        painter.setFont(QFont("Arial", 8))
        painter.drawText(QRectF(margines, y[0], W - 2 * margines, 18),
                         Qt.AlignmentFlag.AlignRight,
                         f"{len(d.wizyty)} wizyt  •  {d.km:.0f} km")
        y[0] += 20

        for i, wz in enumerate(d.wizyty, 1):
            potrzebne_miejsce(16)
            adres_klucz = wz.adres or wz.nazwa
            odwiedzona = czy_odwiedzona(d.data, adres_klucz)
            znak = "✓" if odwiedzona else "○"
            painter.setPen(kol_zie if odwiedzona else kol_mut)
            painter.setFont(QFont("Arial", 9, QFont.Weight.Bold))
            painter.drawText(QRectF(margines, y[0], 16, 14), znak)
            painter.setPen(kol_txt)
            painter.setFont(QFont("Arial", 9))
            tekst = f"{i}.  "
            nazwa_pkt = wz.nazwa or ""
            if wz.siec and wz.siec.lower() not in nazwa_pkt.lower():
                tekst += f"[{wz.siec}]  "
            tekst += nazwa_pkt
            if wz.adres:
                tekst += f"  —  {wz.adres}"
            painter.drawText(QRectF(margines + 18, y[0], W - 2 * margines - 18, 14), tekst)
            y[0] += 15
        y[0] += 10

    painter.end()


def _wersja_na_liczbe(txt):
    """'1.2.10' -> (1,2,10) do porównywania. Odporne na śmieci."""
    czesci = []
    for kawalek in str(txt).strip().split("."):
        cyfry = "".join(ch for ch in kawalek if ch.isdigit())
        czesci.append(int(cyfry) if cyfry else 0)
    while len(czesci) < 3:
        czesci.append(0)
    return tuple(czesci[:3])

def sprawdz_aktualizacje():
    """Zwraca (jest_nowsza, nowa_wersja, opis) albo (False, '', '').
    Działa cicho: każdy błąd (brak sieci, zły URL) = po prostu nic nie pokazujemy.
    Wywoływane w osobnym wątku, żeby nie blokować startu programu."""
    if not URL_WERSJI:
        return (False, "", "")
    try:
        req = urllib.request.Request(URL_WERSJI, headers={"User-Agent": "PMT-Planer"})
        with urllib.request.urlopen(req, timeout=4) as resp:
            tresc = resp.read().decode("utf-8", errors="ignore").strip()
        linie = [l.strip() for l in tresc.splitlines() if l.strip()]
        if not linie:
            return (False, "", "")
        nowa = linie[0]
        opis = linie[1] if len(linie) > 1 else "Dostępna jest nowsza wersja programu."
        if _wersja_na_liczbe(nowa) > _wersja_na_liczbe(WERSJA_PROGRAMU):
            return (True, nowa, opis)
    except Exception:
        pass
    return (False, "", "")

# =============================================================================
#  MAGAZYN DANYCH UŻYTKOWNIKA  (profil + historia rozliczeń)
#  Każdy pracownik ma osobny "sejf" kluczowany haszem (nazwisko+PESEL).
#  Dzięki temu na współdzielonym komputerze nikt nie widzi cudzej historii —
#  podpowiedzi i historia pojawiają się dopiero po wpisaniu WŁASNEGO PESEL-u.
# =============================================================================
USER_STORE = os.path.join(os.path.expanduser("~"), ".pmt_uzytkownicy.json")

# --- DZIENNIK WIZYT (rzetelna dokumentacja własnej pracy) ------------------
# Każda odwiedzona wizyta zapisuje: KIEDY ją odhaczono (znacznik czasu) oraz
# NOTATKĘ z wizyty. Dzięki temu dziennik odzwierciedla faktyczny przebieg pracy,
# a nie tylko "ptaszek". Nie da się odhaczyć wizyty z przyszłości (blokada).
# Klucz wizyty = "RRRR-MM-DD|adres".
WIZYTY_STORE = os.path.join(os.path.expanduser("~"), ".pmt_wizyty.json")
_dziennik = None      # {klucz: {"czas": "...", "notatka": "..."}}

def _wczytaj_dziennik():
    """Wczytuje dziennik z dysku (raz)."""
    global _dziennik
    if _dziennik is not None:
        return _dziennik
    _dziennik = {}
    try:
        if os.path.exists(WIZYTY_STORE):
            with open(WIZYTY_STORE, "r", encoding="utf-8") as f:
                dane = json.load(f)
            if isinstance(dane, dict):
                _dziennik = dane
            elif isinstance(dane, list):
                # zgodność wstecz: stary format (lista kluczy, bez czasu)
                _dziennik = {k: {"czas": "", "notatka": ""} for k in dane}
    except Exception:
        _dziennik = {}
    return _dziennik

def _zapisz_dziennik():
    try:
        with open(WIZYTY_STORE, "w", encoding="utf-8") as f:
            json.dump(_wczytaj_dziennik(), f, ensure_ascii=False, indent=1)
    except Exception:
        pass

def klucz_wizyty(data, adres) -> str:
    """Jednoznaczny klucz wizyty: dzień + adres punktu."""
    try:
        d = data.isoformat()
    except Exception:
        d = str(data)
    return f"{d}|{str(adres).strip().lower()}"

def czy_odwiedzona(data, adres) -> bool:
    return klucz_wizyty(data, adres) in _wczytaj_dziennik()

def wpis_wizyty(data, adres) -> dict:
    """Zwraca wpis dziennika dla wizyty ({"czas","notatka"}) albo pusty dict."""
    return _wczytaj_dziennik().get(klucz_wizyty(data, adres), {})

def mozna_odhaczyc(data) -> bool:
    """Nie da się odhaczyć wizyty zaplanowanej w PRZYSZŁOŚCI — dziennik ma
    dokumentować to, co się faktycznie wydarzyło."""
    try:
        return data <= datetime.date.today()
    except Exception:
        return True

def ustaw_odwiedzona(data, adres, stan: bool, notatka: str = None):
    """Odhacza/cofa wizytę. Przy odhaczeniu zapisuje ZNACZNIK CZASU (kiedy
    faktycznie kliknięto) — to czyni dziennik wiarygodnym."""
    dz = _wczytaj_dziennik()
    k = klucz_wizyty(data, adres)
    if stan:
        if not mozna_odhaczyc(data):
            return False        # blokada: wizyta z przyszłości
        stary = dz.get(k, {})
        dz[k] = {
            "czas": stary.get("czas") or datetime.datetime.now().strftime("%d.%m.%Y %H:%M"),
            "notatka": notatka if notatka is not None else stary.get("notatka", ""),
        }
    else:
        dz.pop(k, None)
    _zapisz_dziennik()
    return True

def ustaw_notatke(data, adres, notatka: str):
    """Zapisuje/aktualizuje notatkę z wizyty (tylko dla odhaczonych)."""
    dz = _wczytaj_dziennik()
    k = klucz_wizyty(data, adres)
    if k in dz:
        dz[k]["notatka"] = notatka
        _zapisz_dziennik()


# --- TRWAŁY PLAN WIZYT ----------------------------------------------------
# Plan zapisujemy na dysk, żeby przetrwał zamknięcie programu. Bez tego
# dziennik (odhaczenia) nie miałby do czego się przypiąć po restarcie.
# --- TRWAŁA LISTA PUNKTÓW (baza sklepów użytkownika) ----------------------
# Bez tego po zamknięciu programu trzeba było importować Excel od nowa.
PUNKTY_STORE = os.path.join(os.path.expanduser("~"), ".pmt_punkty.json")

def zapisz_punkty(pozycje):
    """Zapisuje listę punktów (adres/sieć/miasto) na dysk."""
    try:
        with open(PUNKTY_STORE, "w", encoding="utf-8") as f:
            json.dump(list(pozycje), f, ensure_ascii=False)
    except Exception:
        pass

def wczytaj_punkty():
    """Odczytuje listę punktów z dysku (albo pustą listę)."""
    try:
        if os.path.exists(PUNKTY_STORE):
            with open(PUNKTY_STORE, "r", encoding="utf-8") as f:
                dane = json.load(f)
            if isinstance(dane, list):
                out = []
                for p in dane:
                    if isinstance(p, dict) and p.get("adres"):
                        out.append({"adres": str(p.get("adres", "")),
                                    "siec": str(p.get("siec", "") or ""),
                                    "miasto": str(p.get("miasto", "") or "")})
                return out
    except Exception:
        pass
    return []


PLAN_STORE = os.path.join(os.path.expanduser("~"), ".pmt_plan.json")

# --- USTAWIENIA użytkownika (adres startowy itp.) --------------------------
USTAWIENIA_STORE = os.path.join(os.path.expanduser("~"), ".pmt_ustawienia.json")
_ustawienia = None

def _wczytaj_ustawienia():
    global _ustawienia
    if _ustawienia is not None:
        return _ustawienia
    _ustawienia = {}
    try:
        if os.path.exists(USTAWIENIA_STORE):
            with open(USTAWIENIA_STORE, "r", encoding="utf-8") as f:
                d = json.load(f)
            if isinstance(d, dict):
                _ustawienia = d
    except Exception:
        _ustawienia = {}
    return _ustawienia

def ustawienie(klucz, domyslne=""):
    return _wczytaj_ustawienia().get(klucz, domyslne)

def _ustawienia_reset():
    """Wymusza ponowny odczyt ustawień z dysku (po zmianie w dialogu)."""
    global _ustawienia
    _ustawienia = None


def zapisz_ustawienie(klucz, wartosc):
    u = _wczytaj_ustawienia()
    u[klucz] = wartosc
    try:
        with open(USTAWIENIA_STORE, "w", encoding="utf-8") as f:
            json.dump(u, f, ensure_ascii=False, indent=1)
    except Exception:
        pass

def zapisz_plan(plan: dict):
    """Serializuje plan wizyt do pliku."""
    try:
        dane = {
            "rok": plan.get("rok"), "miesiac": plan.get("miesiac"),
            "suma_km": plan.get("suma_km", 0), "suma_wizyt": plan.get("suma_wizyt", 0),
            "zapisano": datetime.datetime.now().strftime("%d.%m.%Y %H:%M"),
            "cykliczny": bool(plan.get("cykliczny")),
            "cykl_domyslny": plan.get("cykl_domyslny", 4),
            "cykle_sieci": plan.get("cykle_sieci", {}) or {},
            "adres_bazy": plan.get("adres_bazy", ""),
            "miesiace": [],
        }
        for m in plan.get("miesiace", []):
            dane["miesiace"].append({
                "rok": m["rok"], "miesiac": m["miesiac"],
                "suma_km": m.get("suma_km", 0), "suma_wizyt": m.get("suma_wizyt", 0),
                "dni": [{
                    "data": d.data.isoformat(), "km": d.km, "minuty": d.minuty,
                    "wizyty": [{"nazwa": w.nazwa, "adres": w.adres, "siec": w.siec,
                                "lat": w.lat, "lng": w.lng, "miasto": w.miasto}
                               for w in d.wizyty],
                } for d in m["dni"]],
            })
        with open(PLAN_STORE, "w", encoding="utf-8") as f:
            json.dump(dane, f, ensure_ascii=False)
    except Exception:
        pass

def wczytaj_plan():
    """Odtwarza zapisany plan z dysku albo zwraca None."""
    try:
        if not os.path.exists(PLAN_STORE):
            return None
        with open(PLAN_STORE, "r", encoding="utf-8") as f:
            dane = json.load(f)
        miesiace = []; wszystkie_dni = []
        for m in dane.get("miesiace", []):
            dni = []
            for d in m.get("dni", []):
                dz = DzienPlanu(datetime.date.fromisoformat(d["data"]))
                dz.km = d.get("km", 0); dz.minuty = d.get("minuty", 0)
                for w in d.get("wizyty", []):
                    dz.wizyty.append(PunktWizyty(
                        w.get("nazwa", ""), w.get("adres", ""), siec=w.get("siec", ""),
                        lat=w.get("lat"), lng=w.get("lng"), miasto=w.get("miasto", "")))
                dni.append(dz); wszystkie_dni.append(dz)
            if dni:
                miesiace.append({"rok": m["rok"], "miesiac": m["miesiac"], "dni": dni,
                                 "suma_km": m.get("suma_km", 0), "suma_wizyt": m.get("suma_wizyt", 0)})
        if not miesiace:
            return None
        return {
            "dni": wszystkie_dni, "miesiace": miesiace,
            "liczba_miesiecy": len(miesiace),
            "bez_pozycji": [], "nierozplanowane": [],
            "suma_km": dane.get("suma_km", 0), "suma_wizyt": dane.get("suma_wizyt", 0),
            "rok": dane.get("rok"), "miesiac": dane.get("miesiac"),
            "zapisano": dane.get("zapisano", ""),
            "cykliczny": bool(dane.get("cykliczny")),
            "cykl_domyslny": dane.get("cykl_domyslny", 4),
            "cykle_sieci": dane.get("cykle_sieci", {}) or {},
            "adres_bazy": dane.get("adres_bazy", ""),
        }
    except Exception:
        return None

def _klucz_uzytkownika(imie: str, pesel: str) -> str:
    surowy = f"{imie.strip().lower()}|{pesel.strip()}"
    return hashlib.sha256(surowy.encode("utf-8")).hexdigest()[:16]

def _wczytaj_store() -> dict:
    if os.path.exists(USER_STORE):
        try:
            with open(USER_STORE, "r", encoding="utf-8") as f: return json.load(f)
        except Exception: return {}
    return {}

def _zapisz_store(store: dict):
    try:
        with open(USER_STORE, "w", encoding="utf-8") as f:
            json.dump(store, f, ensure_ascii=False, indent=2)
    except Exception as e:
        log_error(e)

def zapisz_profil(imie, pesel, adres, stanowisko, silnik_idx):
    """Zapamiętuje ostatnie dane pracownika (podpowiedzi przy następnym wejściu)."""
    if not (imie and pesel): return
    store = _wczytaj_store()
    k = _klucz_uzytkownika(imie, pesel)
    wpis = store.get(k, {})
    wpis["profil"] = {"imie": imie, "pesel": pesel, "adres": adres,
                      "stanowisko": stanowisko, "silnik_idx": silnik_idx}
    store[k] = wpis
    _zapisz_store(store)

def wczytaj_profil(imie, pesel) -> Optional[dict]:
    store = _wczytaj_store()
    return store.get(_klucz_uzytkownika(imie, pesel), {}).get("profil")

def szukaj_profilu_po_nazwisku(imie) -> Optional[dict]:
    """Podpowiedź po samym nazwisku (zanim jeszcze wpiszą PESEL)."""
    if not imie or len(imie.strip()) < 3: return None
    store = _wczytaj_store()
    imie_low = imie.strip().lower()
    for wpis in store.values():
        prof = wpis.get("profil", {})
        if prof.get("imie", "").strip().lower() == imie_low:
            return prof
    return None

def dodaj_do_historii(imie, pesel, wpis_historii: dict):
    if not (imie and pesel): return
    store = _wczytaj_store()
    k = _klucz_uzytkownika(imie, pesel)
    wpis = store.get(k, {})
    # data pierwszej aktywności (do panelu administratora — "od kiedy korzysta")
    if not wpis.get("pierwsza_aktywnosc"):
        wpis["pierwsza_aktywnosc"] = wpis_historii.get("data") or datetime.datetime.now().strftime("%d.%m.%Y %H:%M")
    hist = wpis.get("historia", [])
    hist.insert(0, wpis_historii)      # najnowsze na górze
    wpis["historia"] = hist[:20]        # trzymamy max 20 ostatnich
    store[k] = wpis
    _zapisz_store(store)

def wczytaj_historie(imie, pesel) -> list:
    store = _wczytaj_store()
    return store.get(_klucz_uzytkownika(imie, pesel), {}).get("historia", [])


# --- Inteligentny import plików (Excel/CSV) --------------------------------
# Kolumny w plikach klientów bywają różnie nazwane i w różnej kolejności.
# Rozpoznajemy je po nagłówku, dopasowując do zestawu typowych wariantów.
NAGLOWKI_KOLUMN = {
    "miasto": ["miasto", "miejscowość", "miejscowosc", "miejsc", "city", "town",
               "m.", "msc", "lokalizacja", "location", "poczta"],
    "ulica": ["ulica", "ul", "ul.", "adres", "address", "street", "str", "droga"],
    "numer": ["numer", "nr", "nr.", "nr bud", "nr budynku", "no", "building",
              "dom", "nr domu", "l.dz."],
    "kod": ["kod", "kod pocztowy", "kod poczt", "zip", "postal", "post code",
            "kod-pocztowy", "pna"],
    "siec": ["sieć", "siec", "sklep", "network", "chain", "brand", "marka",
             "klient", "punkt", "nazwa sklepu", "shop", "store"],
}


def _dopasuj_kolumny(naglowki: list) -> dict:
    """Zwraca mapę {rola: indeks_kolumny} na podstawie nagłówków pliku.
    Dopasowanie odporne na wielkość liter, spacje i drobne różnice nazw."""
    mapa = {}
    norm = [(_i, str(h or "").strip().lower()) for _i, h in enumerate(naglowki)]
    for rola, warianty in NAGLOWKI_KOLUMN.items():
        for idx, h in norm:
            if not h:
                continue
            # dokładne dopasowanie albo nagłówek zaczyna się od wariantu
            if h in warianty or any(h == w or h.startswith(w + " ") or h.startswith(w + ".") for w in warianty):
                mapa[rola] = idx
                break
        else:
            # druga runda: wariant zawiera się w nagłówku (np. "adres klienta").
            # Tylko dłuższe warianty (≥4 znaki), by uniknąć fałszywych trafień
            # typu "dom" w "Radom" albo "ul" w "Tytuł".
            for idx, h in norm:
                if h and any(len(w) >= 4 and w in h for w in warianty):
                    mapa[rola] = idx
                    break
    return mapa


def _wiersz_na_adres(wiersz, mapa: dict) -> tuple:
    """Z wiersza danych + mapy kolumn buduje (etykieta_do_planera, opis_sieci).
    Etykieta to możliwie pełny adres: 'ul. Nazwa Nr, KOD Miasto' lub sama
    miejscowość, jeśli tylko ona jest dostępna."""
    def pole(rola):
        i = mapa.get(rola)
        if i is None or i >= len(wiersz):
            return ""
        return str(wiersz[i] if wiersz[i] is not None else "").strip()

    miasto = pole("miasto"); ulica = pole("ulica")
    numer = pole("numer"); kod = pole("kod"); siec = pole("siec")

    # zbuduj adres z dostępnych części
    czesc_ulica = ""
    if ulica:
        # nie dubluj "ul." jeśli już jest
        u = ulica if ulica.lower().startswith(("ul", "al", "pl", "os")) else f"ul. {ulica}"
        czesc_ulica = f"{u} {numer}".strip() if numer else u
    lewa = czesc_ulica
    prawa = f"{kod} {miasto}".strip() if kod else miasto
    if lewa and prawa:
        adres = f"{lewa}, {prawa}"
    else:
        adres = lewa or prawa
    return adres.strip(), siec


def statystyki_administratora() -> dict:
    """Zbiera zbiorcze statystyki WSZYSTKICH użytkowników programu na tym
    komputerze (do panelu administratora). Zwraca podsumowanie + listę
    użytkowników z ich metrykami."""
    store = _wczytaj_store()
    uzytkownicy = []
    suma_dok = 0; suma_wypraw = 0; suma_km = 0.0; suma_kwota = 0.0
    for wpis in store.values():
        profil = wpis.get("profil", {})
        historia = wpis.get("historia", [])
        if not profil and not historia:
            continue
        imie = profil.get("imie") or (historia[0].get("imie") if historia else "—")
        pesel = profil.get("pesel", "")
        # metryki z historii
        u_dok = 0; u_km = 0.0; u_kwota = 0.0
        daty = []
        for h in historia:
            u_dok += int(h.get("dokumenty", 0) or 0)
            u_km += float(h.get("km", 0) or 0)
            u_kwota += _parsuj_kwote(h.get("kwota", 0))
            d = h.get("data", "")
            if d: daty.append(d)
        u_wypraw = len(historia)
        suma_dok += u_dok; suma_wypraw += u_wypraw
        suma_km += u_km; suma_kwota += u_kwota
        # pierwsza/ostatnia aktywność (daty w formacie "dd.mm.rrrr gg:mm")
        pierwsza = wpis.get("pierwsza_aktywnosc", "")
        ostatnia = daty[0] if daty else pierwsza      # historia: najnowsze na górze
        uzytkownicy.append({
            "imie": imie, "pesel": pesel,
            "stanowisko": profil.get("stanowisko", ""),
            "dokumenty": u_dok, "wyprawy": u_wypraw,
            "km": round(u_km), "kwota": round(u_kwota, 2),
            "pierwsza": pierwsza, "ostatnia": ostatnia,
        })
    # sortuj po liczbie wypraw malejąco
    uzytkownicy.sort(key=lambda u: -u["wyprawy"])
    return {
        "liczba_uzytkownikow": len(uzytkownicy),
        "suma_dokumentow": suma_dok,
        "suma_wypraw": suma_wypraw,
        "suma_km": round(suma_km),
        "suma_kwota": round(suma_kwota, 2),
        "uzytkownicy": uzytkownicy,
    }

def _parsuj_kwote(wartosc) -> float:
    """Zamienia zapis kwoty na liczbę, odporny na różne formaty:
    '2254.50', '2254,50', '2 254,50', '2,254.50'. Nie psuje wartości przez
    naiwne replace(',', '.') — to powodowało błędne 10× zawyżenie sum."""
    if isinstance(wartosc, (int, float)):
        return float(wartosc)
    s = str(wartosc).strip()
    if not s:
        return 0.0
    # usuń spacje (separator tysięcy) i znaki waluty
    s = s.replace(" ", "").replace("zł", "").replace("PLN", "")
    # jeśli są i przecinek, i kropka — ostatni z nich jest separatorem dziesiętnym
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")   # 2.254,50 -> 2254.50
        else:
            s = s.replace(",", "")                       # 2,254.50 -> 2254.50
    elif "," in s:
        # sam przecinek: jeśli po nim dokładnie 2 cyfry → separator dziesiętny
        czesci = s.split(",")
        if len(czesci) == 2 and len(czesci[1]) == 2:
            s = s.replace(",", ".")                       # 2254,50 -> 2254.50
        else:
            s = s.replace(",", "")                        # 2,254 -> 2254
    try:
        return float(s)
    except Exception:
        return 0.0

# --- KALENDARZ: wpisy per pracownik, kluczowane datą "YYYY-MM-DD" ---
# Każdy wpis: {"notatka": str, "wolne": bool, "plan": bool}
def wczytaj_kalendarz(imie, pesel) -> dict:
    if not (imie and pesel): return {}
    store = _wczytaj_store()
    return store.get(_klucz_uzytkownika(imie, pesel), {}).get("kalendarz", {})

def zapisz_wpis_kalendarza(imie, pesel, data_iso: str, notatka: str, wolne: bool, plan: bool):
    if not (imie and pesel): return
    store = _wczytaj_store()
    k = _klucz_uzytkownika(imie, pesel)
    wpis = store.get(k, {})
    kal = wpis.get("kalendarz", {})
    if not notatka and not wolne and not plan:
        kal.pop(data_iso, None)               # pusty wpis = usuń
    else:
        kal[data_iso] = {"notatka": notatka, "wolne": wolne, "plan": plan}
    wpis["kalendarz"] = kal
    store[k] = wpis
    _zapisz_store(store)

def kalendarz_z_historii(imie, pesel) -> dict:
    """Zwraca daty (YYYY-MM-DD) dni wyjazdowych z wygenerowanych delegacji —
    do oznaczenia w kalendarzu. Każda delegacja obejmuje wiele dni jazdy;
    oznaczamy KAŻDY z nich (pole 'dni_daty'). Dla starszych wpisów (bez
    'dni_daty') używamy dnia utworzenia dokumentu jako przybliżenia."""
    out = {}
    for h in wczytaj_historie(imie, pesel):
        dni_daty = h.get("dni_daty")
        if isinstance(dni_daty, list) and dni_daty:
            # nowy format: konkretne dni wyjazdowe
            for iso in dni_daty:
                out[iso] = h
        else:
            # stary wpis: tylko dzień utworzenia ("dd.mm.YYYY HH:MM")
            d = h.get("data", "")
            try:
                czesc = d.split(" ")[0]
                dd, mm, yyyy = czesc.split(".")
                out[f"{yyyy}-{mm}-{dd}"] = h
            except Exception:
                continue
    return out

# Inicjalizacja cache geolokalizacji (przywrócona z v72 — jej brak powodował
# NameError: name '_geo_cache' is not defined przy pierwszym generowaniu)
_geo_cache = {}
_notatki_dni = None


def _wczytaj_notatki_dni():
    """Wczytuje notatki dnia z dysku (raz, potem cache w pamięci)."""
    global _notatki_dni
    if _notatki_dni is not None:
        return _notatki_dni
    _notatki_dni = {}
    try:
        if os.path.exists(NOTATKI_DNI_STORE):
            with open(NOTATKI_DNI_STORE, "r", encoding="utf-8") as f:
                dane = json.load(f)
            if isinstance(dane, dict):
                _notatki_dni = dane
    except Exception:
        pass
    return _notatki_dni


def notatka_dnia(iso_data: str) -> dict:
    """Zwraca {'notatka': str, 'wolne': bool} dla danego dnia (klucz: ISO)."""
    return _wczytaj_notatki_dni().get(iso_data, {"notatka": "", "wolne": False})


def ustaw_notatke_dnia(iso_data: str, notatka: str, wolne: bool):
    """Zapisuje notatkę/dzień wolny. Pusty wpis (brak notatki i nie wolne)
    jest usuwany z magazynu, żeby plik nie puchł niepotrzebnie."""
    n = _wczytaj_notatki_dni()
    notatka = (notatka or "").strip()
    if not notatka and not wolne:
        n.pop(iso_data, None)
    else:
        n[iso_data] = {"notatka": notatka, "wolne": bool(wolne)}
    try:
        with open(NOTATKI_DNI_STORE, "w", encoding="utf-8") as f:
            json.dump(n, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def czy_dzien_wolny(data: datetime.date) -> bool:
    """Używane przez planer wizyt — dzień wolny jest realnie omijany przy
    układaniu tras, nie tylko oznaczony wizualnie w kalendarzu."""
    return bool(notatka_dnia(data.isoformat()).get("wolne"))


# --- KOPIA ZAPASOWA -- pakuje/rozpakowuje WSZYSTKIE magazyny danych --------
# Chroni przed utratą punktów, planu, dziennika i ustawień przy awarii dysku
# albo przeniesieniu programu na inny komputer.
WSZYSTKIE_MAGAZYNY = {
    "punkty": PUNKTY_STORE,
    "plan": PLAN_STORE,
    "wizyty": WIZYTY_STORE,
    "ustawienia": USTAWIENIA_STORE,
    "geo_cache": GEO_CACHE,
    "uzytkownicy": USER_STORE,
    "notatki_dni": NOTATKI_DNI_STORE,
}


def eksportuj_kopie_zapasowa(sciezka_zip: str) -> int:
    """Pakuje wszystkie istniejące magazyny danych do jednego pliku .zip.
    Zwraca liczbę spakowanych plików (0 = nic do zapisania)."""
    manifest = {
        "program": "PMT Planer",
        "wersja": WERSJA_PROGRAMU,
        "data_eksportu": datetime.datetime.now().isoformat(),
        "pliki": [],
    }
    ile = 0
    with zipfile.ZipFile(sciezka_zip, "w", zipfile.ZIP_DEFLATED) as z:
        for klucz, sciezka in WSZYSTKIE_MAGAZYNY.items():
            if os.path.exists(sciezka):
                z.write(sciezka, arcname=os.path.basename(sciezka))
                manifest["pliki"].append(klucz)
                ile += 1
        z.writestr("_manifest_pmt.json", json.dumps(manifest, ensure_ascii=False, indent=2))
    return ile


def sprawdz_kopie_zapasowa(sciezka_zip: str) -> dict:
    """Waliduje plik przed przywróceniem — zwraca manifest albo {} gdy to
    nie jest prawidłowa kopia PMT Planer (zabezpieczenie przed nadpisaniem
    danych przypadkowym/obcym plikiem .zip)."""
    try:
        with zipfile.ZipFile(sciezka_zip, "r") as z:
            if "_manifest_pmt.json" not in z.namelist():
                return {}
            return json.loads(z.read("_manifest_pmt.json").decode("utf-8", errors="ignore"))
    except Exception:
        return {}


def przywroc_z_kopii(sciezka_zip: str) -> int:
    """Rozpakowuje kopię zapasową, NADPISUJĄC obecne magazyny danych.
    Zwraca liczbę przywróconych plików. Czyści też cache w pamięci (dziennik,
    ustawienia, geokodowanie), żeby program od razu widział świeże dane —
    bez tego stare wartości zostałyby w pamięci aż do restartu."""
    global _dziennik, _ustawienia, _geo_cache, _notatki_dni
    ile = 0
    with zipfile.ZipFile(sciezka_zip, "r") as z:
        nazwy_w_zip = set(z.namelist())
        for klucz, sciezka in WSZYSTKIE_MAGAZYNY.items():
            nazwa_pliku = os.path.basename(sciezka)
            if nazwa_pliku in nazwy_w_zip:
                with z.open(nazwa_pliku) as src, open(sciezka, "wb") as dst:
                    dst.write(src.read())
                ile += 1
    # unieważnij cache w pamięci — kolejne odczyty wezmą świeże dane z dysku
    _dziennik = None
    _ustawienia = None
    _geo_cache = {}
    _notatki_dni = None
    return ile


if os.path.exists(GEO_CACHE):
    try:
        with open(GEO_CACHE, 'r', encoding='utf-8') as f: _geo_cache = json.load(f)
    except Exception: pass

MIESIACE_PL = ["styczeń", "luty", "marzec", "kwiecień", "maj", "czerwiec", "lipiec", "sierpień", "wrzesień", "październik", "listopad", "grudzień"] 
WSZYSTKIE_WOJEWODZTWA = ["dolnośląskie", "kujawsko-pomorskie", "lubelskie", "lubuskie", "łódzkie", "małopolskie", "mazowieckie", "opolskie", "podkarpackie", "podlaskie", "pomorskie", "śląskie", "świętokrzyskie", "warmińsko-mazurskie", "wielkopolskie", "zachodniopomorskie"] 

SASIEDZI_WOJ = {
    'dolnośląskie': ['dolnośląskie', 'lubuskie', 'wielkopolskie', 'opolskie'],
    'kujawsko-pomorskie': ['kujawsko-pomorskie', 'pomorskie', 'warmińsko-mazurskie', 'mazowieckie', 'łódzkie', 'wielkopolskie'],
    'lubelskie': ['lubelskie', 'podlaskie', 'mazowieckie', 'świętokrzyskie', 'podkarpackie'],
    'lubuskie': ['lubuskie', 'zachodniopomorskie', 'wielkopolskie', 'dolnośląskie'],
    'łódzkie': ['łódzkie', 'wielkopolskie', 'kujawsko-pomorskie', 'mazowieckie', 'świętokrzyskie', 'śląskie', 'opolskie'],
    'małopolskie': ['małopolskie', 'śląskie', 'świętokrzyskie', 'podkarpackie'],
    'mazowieckie': ['mazowieckie', 'warmińsko-mazurskie', 'podlaskie', 'lubelskie', 'świętokrzyskie', 'łódzkie', 'kujawsko-pomorskie'],
    'opolskie': ['opolskie', 'dolnośląskie', 'wielkopolskie', 'łódzkie', 'śląskie'],
    'podkarpackie': ['podkarpackie', 'lubelskie', 'świętokrzyskie', 'małopolskie'],
    'podlaskie': ['podlaskie', 'warmińsko-mazurskie', 'mazowieckie', 'lubelskie'],
    'pomorskie': ['pomorskie', 'zachodniopomorskie', 'wielkopolskie', 'kujawsko-pomorskie', 'warmińsko-mazurskie'],
    'śląskie': ['śląskie', 'opolskie', 'łódzkie', 'świętokrzyskie', 'małopolskie'],
    'świętokrzyskie': ['świętokrzyskie', 'łódzkie', 'mazowieckie', 'lubelskie', 'podkarpackie', 'małopolskie', 'śląskie'],
    'warmińsko-mazurskie': ['warmińsko-mazurskie', 'pomorskie', 'kujawsko-pomorskie', 'mazowieckie', 'podlaskie'],
    'wielkopolskie': ['wielkopolskie', 'zachodniopomorskie', 'pomorskie', 'kujawsko-pomorskie', 'łódzkie', 'opolskie', 'dolnośląskie', 'lubuskie'],
    'zachodniopomorskie': ['zachodniopomorskie', 'pomorskie', 'wielkopolskie', 'lubuskie']
}

AGLOMERACJE_LISTA = [(52.23, 21.01), (52.40, 16.92), (50.06, 19.94), (51.75, 19.45), (51.10, 17.03), (54.35, 18.64), (51.24, 22.56), (53.01, 18.59), (53.12, 18.01), (53.42, 14.55), (50.87, 20.62), (50.04, 21.99), (53.13, 23.16), (53.77, 20.48), (50.26, 19.02), (50.67, 17.92), (51.93, 15.50), (52.73, 15.24), (50.80, 19.11), (51.40, 21.14), (50.01, 20.98), (54.19, 16.18)]
STOLICE = {'mazowieckie': (52.23, 21.01), 'wielkopolskie': (52.40, 16.92), 'małopolskie': (50.06, 19.94), 'łódzkie': (51.75, 19.45), 'dolnośląskie': (51.10, 17.03), 'pomorskie': (54.35, 18.64), 'lubelskie': (51.24, 22.56), 'kujawsko-pomorskie': (53.01, 18.59), 'zachodniopomorskie': (53.42, 14.55), 'świętokrzyskie': (50.87, 20.62), 'podkarpackie': (50.04, 21.99), 'podlaskie': (53.13, 23.16), 'warmińsko-mazurskie': (53.77, 20.48), 'śląskie': (50.26, 19.02), 'opolskie': (50.67, 17.92), 'lubuskie': (51.93, 15.50)}
SEKTORY_KOLEJNOSC = ["N", "NE", "E", "SE", "S", "SW", "W", "NW"] 

# ============================================================================= 
#  OGÓLNOPOLSKA BAZA GMIN (Skompresowana objętościowo)
# ============================================================================= 
MIASTA_RAW = { 
  "mazowieckie": [{"n":"Goszczyn","lat":51.74,"lng":20.85,"typ":"gmina","sieci":3},{"n":"Błędów","lat":51.78,"lng":20.70,"typ":"gmina","sieci":4},{"n":"Mogielnica","lat":51.69,"lng":20.72,"typ":"gmina","sieci":3},{"n":"Promna","lat":51.68,"lng":20.95,"typ":"gmina","sieci":2},{"n":"Wyśmierzyce","lat":51.61,"lng":20.81,"typ":"gmina","sieci":2},{"n":"Stromiec","lat":51.64,"lng":21.08,"typ":"gmina","sieci":2},{"n":"Jedlińsk","lat":51.52,"lng":21.11,"typ":"gmina","sieci":3},{"n":"Zakrzew","lat":51.46,"lng":21.02,"typ":"gmina","sieci":2},{"n":"Przytyk","lat":51.46,"lng":20.82,"typ":"gmina","sieci":3},{"n":"Potworów","lat":51.52,"lng":20.72,"typ":"gmina","sieci":2},{"n":"Rusinów","lat":51.45,"lng":20.52,"typ":"gmina","sieci":2},{"n":"Odrzywół","lat":51.40,"lng":20.55,"typ":"gmina","sieci":2},{"n":"Maciejowice","lat":51.71,"lng":21.56,"typ":"gmina","sieci":3},{"n":"Sobolew","lat":51.74,"lng":21.67,"typ":"gmina","sieci":2},{"n":"Wilga","lat":51.85,"lng":21.38,"typ":"gmina","sieci":2},{"n":"Trojanów","lat":51.68,"lng":21.81,"typ":"gmina","sieci":2},{"n":"Wodynie","lat":52.05,"lng":22.02,"typ":"gmina","sieci":2},{"n":"Repki","lat":52.39,"lng":22.42,"typ":"gmina","sieci":3},{"n":"Sabnie","lat":52.51,"lng":22.30,"typ":"gmina","sieci":2},{"n":"Sterdyń","lat":52.58,"lng":22.29,"typ":"gmina","sieci":2},{"n":"Korczew","lat":52.35,"lng":22.61,"typ":"gmina","sieci":2},{"n":"Przesmyki","lat":52.26,"lng":22.58,"typ":"gmina","sieci":2},{"n":"Paprotnia","lat":52.26,"lng":22.46,"typ":"gmina","sieci":2},{"n":"Wiśniew","lat":52.09,"lng":22.28,"typ":"gmina","sieci":3},{"n":"Kotuń","lat":52.17,"lng":22.05,"typ":"gmina","sieci":2},{"n":"Mokobody","lat":52.26,"lng":22.11,"typ":"gmina","sieci":2},{"n":"Skórzec","lat":52.11,"lng":22.12,"typ":"gmina","sieci":2},{"n":"Korytnica","lat":52.41,"lng":21.82,"typ":"gmina","sieci":3},{"n":"Jadów","lat":52.47,"lng":21.62,"typ":"gmina","sieci":3},{"n":"Strachówka","lat":52.43,"lng":21.68,"typ":"gmina","sieci":2},{"n":"Poświętne","lat":52.33,"lng":21.43,"typ":"gmina","sieci":2},{"n":"Osieck","lat":51.97,"lng":21.44,"typ":"gmina","sieci":2},{"n":"Celestynów","lat":52.06,"lng":21.39,"typ":"gmina","sieci":3},{"n":"Kołbiel","lat":52.06,"lng":21.48,"typ":"gmina","sieci":3},{"n":"Sobienie-Jeziory","lat":51.93,"lng":21.31,"typ":"gmina","sieci":2},{"n":"Latowicz","lat":52.03,"lng":21.80,"typ":"gmina","sieci":2},{"n":"Parysów","lat":51.98,"lng":21.68,"typ":"gmina","sieci":2},{"n":"Borowie","lat":51.94,"lng":21.75,"typ":"gmina","sieci":2},{"n":"Górzno","lat":51.86,"lng":21.66,"typ":"gmina","sieci":2},{"n":"Miastków","lat":51.86,"lng":21.81,"typ":"gmina","sieci":2},{"n":"Siennica","lat":52.09,"lng":21.61,"typ":"gmina","sieci":3},{"n":"Cegłów","lat":52.14,"lng":21.71,"typ":"gmina","sieci":3},{"n":"Dębe Wielkie","lat":52.20,"lng":21.46,"typ":"gmina","sieci":3},{"n":"Halinów","lat":52.23,"lng":21.35,"typ":"gmina","sieci":3},{"n":"Zwoleń","lat":51.35,"lng":21.58,"typ":"powiat","sieci":4},{"n":"Kazanów","lat":51.27,"lng":21.46,"typ":"gmina","sieci":2},{"n":"Głowaczów","lat":51.61,"lng":21.30,"typ":"gmina","sieci":2},{"n":"Magnuszew","lat":51.75,"lng":21.38,"typ":"gmina","sieci":3},{"n":"Mniszew","lat":51.81,"lng":21.28,"typ":"gmina","sieci":2},{"n":"Tarczyn","lat":51.98,"lng":20.83,"typ":"gmina","sieci":3},{"n":"Chynów","lat":51.90,"lng":21.08,"typ":"gmina","sieci":2},{"n":"Pniewy","lat":51.90,"lng":20.73,"typ":"gmina","sieci":1},{"n":"Mszczonów","lat":51.97,"lng":20.52,"typ":"gmina","sieci":4},{"n":"Radziejowice","lat":51.99,"lng":20.53,"typ":"gmina","sieci":2},{"n":"Żabia Wola","lat":52.05,"lng":20.66,"typ":"gmina","sieci":2},{"n":"Nadarzyn","lat":52.09,"lng":20.80,"typ":"gmina","sieci":3},{"n":"Lesznowola","lat":52.08,"lng":20.93,"typ":"gmina","sieci":3},{"n":"Błonie","lat":52.19,"lng":20.61,"typ":"gmina","sieci":4},{"n":"Teresin","lat":52.19,"lng":20.41,"typ":"gmina","sieci":3},{"n":"Kampinos","lat":52.26,"lng":20.46,"typ":"gmina","sieci":2},{"n":"Leszno","lat":52.26,"lng":20.59,"typ":"gmina","sieci":2},{"n":"Izabelin","lat":52.29,"lng":20.81,"typ":"gmina","sieci":2},{"n":"Czosnów","lat":52.39,"lng":20.72,"typ":"gmina","sieci":3},{"n":"Leoncin","lat":52.40,"lng":20.52,"typ":"gmina","sieci":2},{"n":"Pomiechówek","lat":52.48,"lng":20.73,"typ":"gmina","sieci":3},{"n":"Zakroczym","lat":52.43,"lng":20.61,"typ":"gmina","sieci":2},{"n":"Nasielsk","lat":52.59,"lng":20.79,"typ":"gmina","sieci":4},{"n":"Winnica","lat":52.64,"lng":20.95,"typ":"gmina","sieci":2},{"n":"Serock","lat":52.51,"lng":21.06,"typ":"gmina","sieci":4},{"n":"Nieporęt","lat":52.42,"lng":21.03,"typ":"gmina","sieci":3},{"n":"Wieliszew","lat":52.45,"lng":20.95,"typ":"gmina","sieci":3},{"n":"Radzymin","lat":52.41,"lng":21.18,"typ":"gmina","sieci":4},{"n":"Klembów","lat":52.40,"lng":21.33,"typ":"gmina","sieci":2},{"n":"Stanisławów","lat":52.28,"lng":21.56,"typ":"gmina","sieci":2},{"n":"Grójec","lat":51.86,"lng":20.87,"typ":"powiat","sieci":5},{"n":"Warka","lat":51.78,"lng":21.20,"typ":"powiat","sieci":5},{"n":"Mińsk Maz.","lat":52.18,"lng":21.56,"typ":"powiat","sieci":5},{"n":"Góra Kalwaria","lat":51.98,"lng":21.22,"typ":"powiat","sieci":5},{"n":"Glinojeck","lat":52.82,"lng":20.29,"typ":"gmina","sieci":3},{"n":"Strzegowo","lat":52.90,"lng":20.28,"typ":"gmina","sieci":2},{"n":"Raciąż","lat":52.78,"lng":20.12,"typ":"gmina","sieci":3},{"n":"Baboszewo","lat":52.69,"lng":20.21,"typ":"gmina","sieci":2},{"n":"Ojrzeń","lat":52.78,"lng":20.53,"typ":"gmina","sieci":1},{"n":"Sochocin","lat":52.72,"lng":20.45,"typ":"gmina","sieci":2},{"n":"Nowe Miasto","lat":52.65,"lng":20.52,"typ":"gmina","sieci":2},{"n":"Joniec","lat":52.61,"lng":20.55,"typ":"gmina","sieci":1},{"n":"Załuski","lat":52.54,"lng":20.53,"typ":"gmina","sieci":1},{"n":"Drobin","lat":52.74,"lng":19.98,"typ":"gmina","sieci":2},{"n":"Gąbin","lat":52.39,"lng":19.73,"typ":"gmina","sieci":2},{"n":"Czerwińsk","lat":52.39,"lng":20.30,"typ":"gmina","sieci":2},{"n":"Wyszogród","lat":52.39,"lng":20.20,"typ":"gmina","sieci":2},{"n":"Sierpc","lat":52.85,"lng":19.66,"typ":"gmina","sieci":4},{"n":"Gozdowo","lat":52.73,"lng":19.83,"typ":"gmina","sieci":1},{"n":"Rościszewo","lat":52.90,"lng":19.77,"typ":"gmina","sieci":1},{"n":"Mochowo","lat":52.75,"lng":19.64,"typ":"gmina","sieci":1},{"n":"Szczutowo","lat":52.96,"lng":19.60,"typ":"gmina","sieci":1},{"n":"Żuromin","lat":53.06,"lng":19.90,"typ":"gmina","sieci":3},{"n":"Bieżuń","lat":52.96,"lng":19.89,"typ":"gmina","sieci":2},{"n":"Kuczbork","lat":53.07,"lng":19.99,"typ":"gmina","sieci":1},{"n":"Lipowiec K.","lat":53.11,"lng":20.06,"typ":"gmina","sieci":1},{"n":"Lubowidz","lat":53.08,"lng":19.82,"typ":"gmina","sieci":1},{"n":"Brudzeń Duży","lat":52.66,"lng":19.83,"typ":"gmina","sieci":1},{"n":"Bielsk","lat":52.65,"lng":19.77,"typ":"gmina","sieci":2},{"n":"Zawidz","lat":52.80,"lng":19.90,"typ":"gmina","sieci":2},{"n":"Sanniki","lat":52.33,"lng":19.86,"typ":"gmina","sieci":2},{"n":"Gostynin","lat":52.42,"lng":19.46,"typ":"gmina","sieci":4},{"n":"Szczawin","lat":52.36,"lng":19.79,"typ":"gmina","sieci":1},{"n":"Młodzieszyn","lat":52.81,"lng":20.19,"typ":"gmina","sieci":1},{"n":"Stara Biała","lat":52.66,"lng":19.69,"typ":"gmina","sieci":1},{"n":"Radzanowo","lat":52.57,"lng":19.98,"typ":"gmina","sieci":1},{"n":"Bulkowo","lat":52.54,"lng":20.06,"typ":"gmina","sieci":1},{"n":"Dąbrówka","lat":52.30,"lng":21.57,"typ":"gmina","sieci":1},{"n":"Małkinia G.","lat":52.69,"lng":21.89,"typ":"gmina","sieci":3},{"n":"Lochów","lat":52.63,"lng":21.71,"typ":"gmina","sieci":3},{"n":"Brok","lat":52.69,"lng":21.84,"typ":"gmina","sieci":2},{"n":"Jabłonna","lat":52.37,"lng":20.92,"typ":"gmina","sieci":3},{"n":"Tłuszcz","lat":52.43,"lng":21.44,"typ":"gmina","sieci":3},{"n":"Wołomin","lat":52.34,"lng":21.24,"typ":"gmina","sieci":4},{"n":"Kobyłka","lat":52.34,"lng":21.2,"typ":"gmina","sieci":3},{"n":"Zielonka","lat":52.3,"lng":21.15,"typ":"gmina","sieci":3},{"n":"Ząbki","lat":52.29,"lng":21.18,"typ":"gmina","sieci":4},{"n":"Sulejówek","lat":52.25,"lng":21.27,"typ":"gmina","sieci":3},{"n":"Wiązowna","lat":52.15,"lng":21.28,"typ":"gmina","sieci":2},{"n":"Karczew","lat":52.08,"lng":21.25,"typ":"gmina","sieci":3},{"n":"Otwock","lat":52.11,"lng":21.26,"typ":"gmina","sieci":4},{"n":"Piaseczno","lat":52.08,"lng":21.02,"typ":"gmina","sieci":5},{"n":"Konstancin","lat":51.99,"lng":21.11,"typ":"gmina","sieci":3},{"n":"Prażmów","lat":51.95,"lng":21.0,"typ":"gmina","sieci":2},{"n":"Grabów","lat":51.62,"lng":21.6,"typ":"gmina","sieci":1},{"n":"Iłża","lat":51.16,"lng":21.24,"typ":"gmina","sieci":3},{"n":"Skaryszew","lat":51.31,"lng":21.25,"typ":"gmina","sieci":2},{"n":"Pionki","lat":51.48,"lng":21.45,"typ":"gmina","sieci":4},{"n":"Jastrzębia","lat":51.42,"lng":21.3,"typ":"gmina","sieci":1},{"n":"Wierzbica","lat":51.2,"lng":21.08,"typ":"gmina","sieci":2},{"n":"Wolanów","lat":51.36,"lng":21.0,"typ":"gmina","sieci":1},{"n":"Gózd","lat":51.4,"lng":21.33,"typ":"gmina","sieci":1},{"n":"Kowala","lat":51.3,"lng":21.1,"typ":"gmina","sieci":1},{"n":"Marki","lat":52.32,"lng":21.1,"typ":"gmina","sieci":3},{"n":"Ostrów Maz.","lat":52.8,"lng":21.89,"typ":"gmina","sieci":3},{"n":"Wyszków","lat":52.59,"lng":21.46,"typ":"gmina","sieci":3},{"n":"Maków Maz.","lat":52.86,"lng":21.1,"typ":"gmina","sieci":2},{"n":"Różan","lat":52.88,"lng":21.39,"typ":"gmina","sieci":1},{"n":"Pułtusk","lat":52.7,"lng":21.08,"typ":"gmina","sieci":3},{"n":"Płońsk","lat":52.62,"lng":20.38,"typ":"gmina","sieci":3},{"n":"Mława","lat":53.11,"lng":20.38,"typ":"gmina","sieci":3},{"n":"Przasnysz","lat":53.02,"lng":20.88,"typ":"gmina","sieci":3},{"n":"Chorzele","lat":53.27,"lng":20.9,"typ":"gmina","sieci":1},{"n":"Krasnosielc","lat":53.02,"lng":21.15,"typ":"gmina","sieci":1},{"n":"Białobrzegi","lat":51.64,"lng":20.95,"typ":"gmina","sieci":2},{"n":"Kozienice","lat":51.58,"lng":21.55,"typ":"gmina","sieci":3},{"n":"Lipsko","lat":51.16,"lng":21.65,"typ":"gmina","sieci":1},{"n":"Przysucha","lat":51.36,"lng":20.63,"typ":"gmina","sieci":2},{"n":"Szydłowiec","lat":51.23,"lng":20.85,"typ":"gmina","sieci":2}],
  "wielkopolskie": [{"n":"Luboń","lat":52.33,"lng":16.88,"typ":"gmina","sieci":4},{"n":"Komorniki","lat":52.33,"lng":16.80,"typ":"gmina","sieci":3},{"n":"Tarnowo Podg.","lat":52.46,"lng":16.66,"typ":"gmina","sieci":4},{"n":"Dopiewo","lat":52.35,"lng":16.67,"typ":"gmina","sieci":3},{"n":"Rokietnica","lat":52.50,"lng":16.75,"typ":"gmina","sieci":2},{"n":"Suchy Las","lat":52.48,"lng":16.86,"typ":"gmina","sieci":3},{"n":"Czerwonak","lat":52.46,"lng":16.98,"typ":"gmina","sieci":2},{"n":"Swarzędz","lat":52.41,"lng":17.07,"typ":"gmina","sieci":5},{"n":"Kostrzyn","lat":52.39,"lng":17.22,"typ":"gmina","sieci":3},{"n":"Kórnik","lat":52.24,"lng":17.09,"typ":"gmina","sieci":3},{"n":"Mosina","lat":52.24,"lng":16.84,"typ":"gmina","sieci":3},{"n":"Puszczykowo","lat":52.27,"lng":16.93,"typ":"gmina","sieci":2},{"n":"Stęszew","lat":52.28,"lng":16.70,"typ":"gmina","sieci":2},{"n":"Pobiedziska","lat":52.47,"lng":17.28,"typ":"gmina","sieci":2},{"n":"Mur. Goślina","lat":52.57,"lng":17.00,"typ":"gmina","sieci":2},{"n":"Opalenica","lat":52.30,"lng":16.41,"typ":"gmina","sieci":3},{"n":"Buk","lat":52.35,"lng":16.52,"typ":"gmina","sieci":2},{"n":"Kleszczewo","lat":52.33,"lng":17.17,"typ":"gmina","sieci":2},{"n":"Nekla","lat":52.36,"lng":17.41,"typ":"gmina","sieci":2},{"n":"Środa Wlkp.","lat":52.23,"lng":17.28,"typ":"gmina","sieci":4},{"n":"Zaniemyśl","lat":52.15,"lng":17.16,"typ":"gmina","sieci":2},{"n":"Pniewy","lat":52.51,"lng":16.26,"typ":"gmina","sieci":3},{"n":"Szamotuły","lat":52.61,"lng":16.58,"typ":"gmina","sieci":4},{"n":"Oborniki","lat":52.65,"lng":16.81,"typ":"gmina","sieci":4},{"n":"Rogoźno","lat":52.75,"lng":16.99,"typ":"gmina","sieci":3},{"n":"Skoki","lat":52.67,"lng":17.16,"typ":"gmina","sieci":2},{"n":"Kłecko","lat":52.63,"lng":17.43,"typ":"gmina","sieci":2},{"n":"Czerniejewo","lat":52.55,"lng":17.49,"typ":"gmina","sieci":2},{"n":"Trzemeszno","lat":52.56,"lng":17.82,"typ":"gmina","sieci":3},{"n":"Witkowo","lat":52.54,"lng":17.77,"typ":"gmina","sieci":3},{"n":"Miłosław","lat":52.10,"lng":17.48,"typ":"gmina","sieci":2},{"n":"N. M. nad Wartą","lat":52.09,"lng":17.41,"typ":"gmina","sieci":2},{"n":"Jarocin","lat":51.65,"lng":17.85,"typ":"gmina","sieci":4},{"n":"Pleszew","lat":51.90,"lng":17.78,"typ":"gmina","sieci":4},{"n":"Koźmin Wlkp.","lat":51.83,"lng":17.46,"typ":"gmina","sieci":3},{"n":"Krotoszyn","lat":51.69,"lng":17.44,"typ":"gmina","sieci":4},{"n":"Bojanowo","lat":52.01,"lng":16.54,"typ":"gmina","sieci":2},{"n":"Śmigiel","lat":52.01,"lng":16.52,"typ":"gmina","sieci":2},{"n":"Kościan","lat":52.08,"lng":16.65,"typ":"gmina","sieci":4},{"n":"Wolsztyn","lat":52.11,"lng":16.11,"typ":"gmina","sieci":4},{"n":"Rakoniewice","lat":52.14,"lng":16.27,"typ":"gmina","sieci":2},{"n":"Granowo","lat":52.11,"lng":16.22,"typ":"gmina","sieci":2},{"n":"Grodzisk Wlkp.","lat":52.22,"lng":16.36,"typ":"gmina","sieci":3},{"n":"Wronki","lat":52.71,"lng":16.38,"typ":"gmina","sieci":3},{"n":"Krzyż Wlkp.","lat":52.88,"lng":16.01,"typ":"gmina","sieci":2},{"n":"Trzcianka","lat":53.04,"lng":16.45,"typ":"gmina","sieci":3},{"n":"Ujście","lat":53.05,"lng":16.73,"typ":"gmina","sieci":2},{"n":"Wyrzysk","lat":53.09,"lng":16.74,"typ":"gmina","sieci":2},{"n":"Kaczory","lat":53.06,"lng":16.88,"typ":"gmina","sieci":1},{"n":"Wysoka","lat":53.14,"lng":17.08,"typ":"gmina","sieci":2},{"n":"Łobżenica","lat":53.27,"lng":17.26,"typ":"gmina","sieci":2},{"n":"Wągrowiec","lat":52.80,"lng":17.20,"typ":"gmina","sieci":4},{"n":"Margonin","lat":52.97,"lng":17.09,"typ":"gmina","sieci":2},{"n":"Mieścisko","lat":52.19,"lng":17.32,"typ":"gmina","sieci":1},{"n":"Zagorów","lat":52.01,"lng":15.72,"typ":"gmina","sieci":2},{"n":"Pyzdry","lat":52.16,"lng":17.67,"typ":"gmina","sieci":2},{"n":"Rychwał","lat":52.06,"lng":18.02,"typ":"gmina","sieci":2},{"n":"Czempiń","lat":52.16,"lng":16.76,"typ":"gmina","sieci":1},{"n":"Krzywiń","lat":51.96,"lng":16.83,"typ":"gmina","sieci":1},{"n":"Gostyń","lat":51.88,"lng":17.01,"typ":"gmina","sieci":3},{"n":"Poniec","lat":51.79,"lng":16.81,"typ":"gmina","sieci":1},{"n":"Krobia","lat":51.79,"lng":17.0,"typ":"gmina","sieci":1},{"n":"Borek Wlkp.","lat":51.91,"lng":17.22,"typ":"gmina","sieci":1},{"n":"Pogorzela","lat":51.82,"lng":17.2,"typ":"gmina","sieci":1},{"n":"Rozdrażew","lat":51.79,"lng":17.55,"typ":"gmina","sieci":1}],
  "łódzkie": [{"n":"Rzgów","lat":51.66,"lng":19.49,"typ":"gmina","sieci":3},{"n":"Tuszyn","lat":51.60,"lng":19.53,"typ":"gmina","sieci":2},{"n":"Stryków","lat":51.90,"lng":19.61,"typ":"gmina","sieci":3},{"n":"Głowno","lat":51.96,"lng":19.71,"typ":"gmina","sieci":3},{"n":"Ozorków","lat":51.96,"lng":19.29,"typ":"gmina","sieci":4},{"n":"Poddębice","lat":51.89,"lng":18.96,"typ":"gmina","sieci":3},{"n":"Dobroń","lat":51.63,"lng":19.21,"typ":"gmina","sieci":2},{"n":"Szadek","lat":51.69,"lng":18.98,"typ":"gmina","sieci":2},{"n":"Wodzierady","lat":51.72,"lng":19.13,"typ":"gmina","sieci":1},{"n":"Zelów","lat":51.46,"lng":19.22,"typ":"gmina","sieci":3},{"n":"Lutomiersk","lat":51.75,"lng":19.21,"typ":"gmina","sieci":2},{"n":"Dalików","lat":51.88,"lng":19.11,"typ":"gmina","sieci":1},{"n":"Parzęczew","lat":51.95,"lng":19.20,"typ":"gmina","sieci":2},{"n":"Dmosin","lat":51.91,"lng":19.75,"typ":"gmina","sieci":2},{"n":"Piątek","lat":52.06,"lng":19.48,"typ":"gmina","sieci":2},{"n":"Zduńska Wola","lat":51.60,"lng":18.94,"typ":"gmina","sieci":4},{"n":"Łask","lat":51.59,"lng":19.13,"typ":"gmina","sieci":4},{"n":"Aleksandrów Ł.","lat":51.81,"lng":19.30,"typ":"gmina","sieci":4},{"n":"Ksawerów","lat":51.67,"lng":19.46,"typ":"gmina","sieci":2},{"n":"Koluszki","lat":51.74,"lng":19.81,"typ":"gmina","sieci":3},{"n":"Brzeziny","lat":51.80,"lng":19.74,"typ":"gmina","sieci":3},{"n":"Widawa","lat":51.53,"lng":18.94,"typ":"gmina","sieci":2},{"n":"Złoczew","lat":51.41,"lng":18.60,"typ":"gmina","sieci":2},{"n":"Warta","lat":51.47,"lng":18.62,"typ":"gmina","sieci":2},{"n":"Błaszki","lat":51.62,"lng":18.43,"typ":"gmina","sieci":2},{"n":"Wieruszów","lat":51.29,"lng":18.15,"typ":"gmina","sieci":3},{"n":"Działoszyn","lat":51.11,"lng":18.86,"typ":"gmina","sieci":2},{"n":"Pajęczno","lat":51.14,"lng":18.99,"typ":"gmina","sieci":2},{"n":"Dębiak","lat":51.35,"lng":19.15,"typ":"gmina","sieci":1},{"n":"Rusiec","lat":51.32,"lng":18.98,"typ":"gmina","sieci":1},{"n":"Kleszczów","lat":51.22,"lng":19.30,"typ":"gmina","sieci":2},{"n":"Sulmierzyce","lat":51.18,"lng":19.19,"typ":"gmina","sieci":1},{"n":"Rząśnia","lat":51.24,"lng":19.03,"typ":"gmina","sieci":1},{"n":"Kiełczygłów","lat":51.23,"lng":18.97,"typ":"gmina","sieci":1},{"n":"Szczerców","lat":51.23,"lng":19.11,"typ":"gmina","sieci":1},{"n":"Rozprza","lat":51.30,"lng":19.00,"typ":"gmina","sieci":1},{"n":"Zgierz","lat":51.86,"lng":19.41,"typ":"gmina","sieci":3},{"n":"Konstantynów","lat":51.75,"lng":19.33,"typ":"gmina","sieci":2},{"n":"Pabianice","lat":51.66,"lng":19.35,"typ":"gmina","sieci":3},{"n":"Rogów","lat":51.81,"lng":19.9,"typ":"gmina","sieci":1},{"n":"Łęczyca","lat":52.06,"lng":19.2,"typ":"gmina","sieci":2},{"n":"Góra św.Małg.","lat":52.03,"lng":19.55,"typ":"gmina","sieci":1},{"n":"Sędziejowice","lat":51.52,"lng":19.0,"typ":"gmina","sieci":1},{"n":"Wieluń","lat":51.22,"lng":18.57,"typ":"gmina","sieci":3}],
  "śląskie": [{"n":"Pszczyna","lat":49.98,"lng":18.94,"typ":"gmina","sieci":4},{"n":"Łaziska Górne","lat":50.15,"lng":18.84,"typ":"gmina","sieci":3},{"n":"Orzesze","lat":50.14,"lng":18.77,"typ":"gmina","sieci":2},{"n":"Czerwionka","lat":50.15,"lng":18.67,"typ":"gmina","sieci":3},{"n":"Knułów","lat":50.22,"lng":18.67,"typ":"gmina","sieci":3},{"n":"Pyskowice","lat":50.39,"lng":18.62,"typ":"gmina","sieci":2},{"n":"Toszek","lat":50.45,"lng":18.52,"typ":"gmina","sieci":2},{"n":"Radzionków","lat":50.39,"lng":18.90,"typ":"gmina","sieci":3},{"n":"Siewierz","lat":50.46,"lng":19.20,"typ":"gmina","sieci":2},{"n":"Poręba","lat":50.49,"lng":19.33,"typ":"gmina","sieci":2},{"n":"Lędziny","lat":50.13,"lng":19.11,"typ":"gmina","sieci":2},{"n":"Imielin","lat":50.14,"lng":19.17,"typ":"gmina","sieci":2},{"n":"Czechowice","lat":49.91,"lng":19.00,"typ":"gmina","sieci":3},{"n":"Bojszowy","lat":50.04,"lng":19.10,"typ":"gmina","sieci":2},{"n":"Chełm Śląski","lat":50.11,"lng":19.18,"typ":"gmina","sieci":2},{"n":"Suszec","lat":50.03,"lng":18.82,"typ":"gmina","sieci":2},{"n":"Kobiór","lat":50.05,"lng":18.94,"typ":"gmina","sieci":2},{"n":"Pawłowice","lat":49.96,"lng":18.71,"typ":"gmina","sieci":3},{"n":"Goczałkowice","lat":49.94,"lng":18.97,"typ":"gmina","sieci":2},{"n":"Wyry","lat":50.13,"lng":18.89,"typ":"gmina","sieci":2},{"n":"Ornontowice","lat":50.18,"lng":18.75,"typ":"gmina","sieci":2},{"n":"Wielowieś","lat":50.50,"lng":18.62,"typ":"gmina","sieci":1},{"n":"Rudziniec","lat":50.35,"lng":18.41,"typ":"gmina","sieci":2},{"n":"Tworóg","lat":50.53,"lng":18.71,"typ":"gmina","sieci":2},{"n":"Krupski Młyn","lat":50.57,"lng":18.62,"typ":"gmina","sieci":1},{"n":"Zbrosławice","lat":50.41,"lng":18.76,"typ":"gmina","sieci":2},{"n":"Miasteczko Śl.","lat":50.49,"lng":18.92,"typ":"gmina","sieci":2},{"n":"Ożarowice","lat":50.47,"lng":19.04,"typ":"gmina","sieci":2},{"n":"Świerklaniec","lat":50.43,"lng":18.95,"typ":"gmina","sieci":2},{"n":"Bieruń","lat":50.08,"lng":19.09,"typ":"powiat","sieci":3},{"n":"Kłobuck","lat":50.90,"lng":18.93,"typ":"gmina","sieci":4},{"n":"Krzepice","lat":50.96,"lng":18.82,"typ":"gmina","sieci":2},{"n":"Wręczyca Wlk.","lat":50.85,"lng":18.93,"typ":"gmina","sieci":2},{"n":"Przystajń","lat":50.85,"lng":18.78,"typ":"gmina","sieci":1},{"n":"Opatów","lat":50.95,"lng":18.81,"typ":"gmina","sieci":1},{"n":"Lipie","lat":51.02,"lng":18.80,"typ":"gmina","sieci":1},{"n":"Popów","lat":51.05,"lng":18.95,"typ":"gmina","sieci":1},{"n":"Miedźno","lat":50.96,"lng":18.96,"typ":"gmina","sieci":2},{"n":"Mykanów","lat":50.83,"lng":19.18,"typ":"gmina","sieci":2},{"n":"Rędziny","lat":50.86,"lng":19.22,"typ":"gmina","sieci":2},{"n":"Kłomnice","lat":50.90,"lng":19.25,"typ":"gmina","sieci":2},{"n":"Kruszyna","lat":50.93,"lng":19.28,"typ":"gmina","sieci":1},{"n":"Lubliniec","lat":50.66,"lng":18.68,"typ":"gmina","sieci":4},{"n":"Koszęcin","lat":50.63,"lng":18.84,"typ":"gmina","sieci":2},{"n":"Boronów","lat":50.67,"lng":18.90,"typ":"gmina","sieci":1},{"n":"Woźniki","lat":50.43,"lng":18.99,"typ":"gmina","sieci":2},{"n":"Wilamowice","lat":49.91,"lng":19.15,"typ":"gmina","sieci":2},{"n":"Kozy","lat":49.86,"lng":19.14,"typ":"gmina","sieci":2},{"n":"Wilkowice","lat":49.75,"lng":19.05,"typ":"gmina","sieci":2},{"n":"Buczkowice","lat":49.72,"lng":19.06,"typ":"gmina","sieci":1},{"n":"Szczyrk","lat":49.72,"lng":19.02,"typ":"gmina","sieci":2},{"n":"Jaworze","lat":49.81,"lng":18.94,"typ":"gmina","sieci":1},{"n":"Jasienica","lat":49.81,"lng":18.9,"typ":"gmina","sieci":2},{"n":"Jaworzynka","lat":49.53,"lng":18.85,"typ":"gmina","sieci":1},{"n":"Istebna","lat":49.55,"lng":18.9,"typ":"gmina","sieci":2},{"n":"Wisła","lat":49.66,"lng":18.86,"typ":"gmina","sieci":3},{"n":"Ustroń","lat":49.72,"lng":18.81,"typ":"gmina","sieci":3},{"n":"Skoczów","lat":49.8,"lng":18.79,"typ":"gmina","sieci":3},{"n":"Strumień","lat":49.94,"lng":18.76,"typ":"gmina","sieci":2},{"n":"Zebrzydowice","lat":49.87,"lng":18.6,"typ":"gmina","sieci":2},{"n":"Hażlach","lat":49.86,"lng":18.66,"typ":"gmina","sieci":1},{"n":"Dębowiec","lat":49.85,"lng":18.75,"typ":"gmina","sieci":1},{"n":"Chybie","lat":49.9,"lng":18.81,"typ":"gmina","sieci":1},{"n":"Brenna","lat":49.72,"lng":18.9,"typ":"gmina","sieci":1},{"n":"Miedźna","lat":49.98,"lng":19.03,"typ":"gmina","sieci":1},{"n":"Chełm Śl.","lat":50.1,"lng":19.2,"typ":"gmina","sieci":1},{"n":"Mysłowice","lat":50.24,"lng":19.13,"typ":"gmina","sieci":3},{"n":"Jaworzno","lat":50.2,"lng":19.27,"typ":"gmina","sieci":3},{"n":"Sosnowiec","lat":50.28,"lng":19.13,"typ":"gmina","sieci":4},{"n":"Będzin","lat":50.33,"lng":19.13,"typ":"gmina","sieci":3},{"n":"Czeladź","lat":50.32,"lng":19.08,"typ":"gmina","sieci":2},{"n":"Wojkowice","lat":50.36,"lng":19.02,"typ":"gmina","sieci":1},{"n":"Łazy","lat":50.43,"lng":19.4,"typ":"gmina","sieci":1},{"n":"Ogrodzieniec","lat":50.45,"lng":19.52,"typ":"gmina","sieci":1},{"n":"Zawiercie","lat":50.49,"lng":19.43,"typ":"gmina","sieci":3},{"n":"Kroczyce","lat":50.53,"lng":19.6,"typ":"gmina","sieci":1},{"n":"Włodowice","lat":50.51,"lng":19.5,"typ":"gmina","sieci":1},{"n":"Myszków","lat":50.58,"lng":19.32,"typ":"gmina","sieci":3},{"n":"Koziegłowy","lat":50.61,"lng":19.14,"typ":"gmina","sieci":1},{"n":"Kalety","lat":50.61,"lng":18.9,"typ":"gmina","sieci":1}],
  "małopolskie": [{"n":"Skawina","lat":49.97,"lng":19.82,"typ":"gmina","sieci":4},{"n":"Niepołomice","lat":50.03,"lng":20.21,"typ":"gmina","sieci":3},{"n":"Krzeszowice","lat":50.13,"lng":19.63,"typ":"gmina","sieci":3},{"n":"Zabierzów","lat":50.11,"lng":19.79,"typ":"gmina","sieci":2},{"n":"Liszki","lat":50.03,"lng":19.78,"typ":"gmina","sieci":2},{"n":"Zielonki","lat":50.11,"lng":19.93,"typ":"gmina","sieci":3},{"n":"Michałowice","lat":50.13,"lng":19.99,"typ":"gmina","sieci":2},{"n":"Słomniki","lat":50.24,"lng":20.08,"typ":"gmina","sieci":2},{"n":"Kocmyrzów","lat":50.12,"lng":20.10,"typ":"gmina","sieci":2},{"n":"Dobczyce","lat":50.24,"lng":20.09,"typ":"gmina","sieci":2},{"n":"Kłaj","lat":50.00,"lng":20.26,"typ":"gmina","sieci":2},{"n":"Biskupice","lat":49.96,"lng":20.12,"typ":"gmina","sieci":2},{"n":"Drwinia","lat":50.11,"lng":20.43,"typ":"gmina","sieci":1},{"n":"Rzezawa","lat":49.98,"lng":20.50,"typ":"gmina","sieci":2},{"n":"Koszyce","lat":50.16,"lng":20.57,"typ":"gmina","sieci":2},{"n":"Koniusza","lat":50.16,"lng":20.20,"typ":"gmina","sieci":1},{"n":"Igołomia","lat":50.08,"lng":20.21,"typ":"gmina","sieci":2},{"n":"Czernichów","lat":49.98,"lng":19.68,"typ":"gmina","sieci":2},{"n":"Świątniki Górne","lat":49.93,"lng":19.95,"typ":"gmina","sieci":2},{"n":"Gdów","lat":49.90,"lng":20.19,"typ":"gmina","sieci":3},{"n":"Sułkowice","lat":49.83,"lng":19.79,"typ":"gmina","sieci":3},{"n":"Pcim","lat":49.75,"lng":19.97,"typ":"gmina","sieci":2},{"n":"Wieliczka","lat":49.98,"lng":20.06,"typ":"gmina","sieci":5},{"n":"Proszowice","lat":50.19,"lng":20.28,"typ":"gmina","sieci":3},{"n":"Tuchów","lat":49.89,"lng":21.05,"typ":"gmina","sieci":2},{"n":"Ryglice","lat":49.87,"lng":21.13,"typ":"gmina","sieci":2},{"n":"Gromnik","lat":49.83,"lng":20.96,"typ":"gmina","sieci":2},{"n":"Żabno","lat":50.13,"lng":20.88,"typ":"gmina","sieci":2},{"n":"Wojnicz","lat":49.96,"lng":20.84,"typ":"gmina","sieci":2},{"n":"Radłów","lat":50.08,"lng":20.85,"typ":"gmina","sieci":2},{"n":"Wietrzychowice","lat":50.10,"lng":20.76,"typ":"gmina","sieci":1},{"n":"Bolesław","lat":50.23,"lng":20.87,"typ":"gmina","sieci":1},{"n":"Olesno","lat":50.21,"lng":20.95,"typ":"gmina","sieci":2},{"n":"Szczucin","lat":50.31,"lng":21.07,"typ":"gmina","sieci":2},{"n":"Radgoszcz","lat":50.11,"lng":20.80,"typ":"gmina","sieci":1},{"n":"Dąbrowa Tarn.","lat":50.17,"lng":20.98,"typ":"gmina","sieci":4},{"n":"Skrzyszów","lat":49.90,"lng":21.05,"typ":"gmina","sieci":1},{"n":"Pleśna","lat":49.93,"lng":20.94,"typ":"gmina","sieci":2},{"n":"Tarnów Opol.","lat":49.95,"lng":21.00,"typ":"gmina","sieci":2},{"n":"Ciężkowice","lat":49.78,"lng":20.97,"typ":"gmina","sieci":2},{"n":"Świątniki","lat":49.92,"lng":19.95,"typ":"gmina","sieci":2},{"n":"Mogilany","lat":49.94,"lng":19.9,"typ":"gmina","sieci":2},{"n":"Myślenice","lat":49.83,"lng":19.94,"typ":"gmina","sieci":4},{"n":"Kalwaria Z.","lat":49.85,"lng":19.68,"typ":"gmina","sieci":3},{"n":"Wadowice","lat":49.88,"lng":19.49,"typ":"gmina","sieci":4},{"n":"Andrychów","lat":49.86,"lng":19.34,"typ":"gmina","sieci":4},{"n":"Kęty","lat":49.88,"lng":19.22,"typ":"gmina","sieci":4},{"n":"Zator","lat":49.99,"lng":19.43,"typ":"gmina","sieci":2},{"n":"Brzeźnica","lat":49.98,"lng":19.63,"typ":"gmina","sieci":2},{"n":"Alwernia","lat":50.06,"lng":19.54,"typ":"gmina","sieci":2},{"n":"Wielka Wieś","lat":50.15,"lng":19.83,"typ":"gmina","sieci":2},{"n":"Iwanowice","lat":50.2,"lng":19.93,"typ":"gmina","sieci":1},{"n":"Trzebinia","lat":50.16,"lng":19.47,"typ":"gmina","sieci":3},{"n":"Chrzanów","lat":50.14,"lng":19.4,"typ":"gmina","sieci":3},{"n":"Libiąż","lat":50.1,"lng":19.32,"typ":"gmina","sieci":2},{"n":"Olkusz","lat":50.28,"lng":19.56,"typ":"gmina","sieci":3},{"n":"Bukowno","lat":50.27,"lng":19.46,"typ":"gmina","sieci":2},{"n":"Wolbrom","lat":50.38,"lng":19.76,"typ":"gmina","sieci":2},{"n":"Miechów","lat":50.36,"lng":20.03,"typ":"gmina","sieci":2},{"n":"Książ Wielki","lat":50.45,"lng":20.13,"typ":"gmina","sieci":1},{"n":"Charsznica","lat":50.42,"lng":19.93,"typ":"gmina","sieci":1},{"n":"Gołcza","lat":50.32,"lng":19.95,"typ":"gmina","sieci":1},{"n":"Skała","lat":50.22,"lng":19.86,"typ":"gmina","sieci":1},{"n":"Sułoszowa","lat":50.25,"lng":19.75,"typ":"gmina","sieci":1},{"n":"Jerzmanowice","lat":50.2,"lng":19.75,"typ":"gmina","sieci":1},{"n":"Chełmek","lat":50.1,"lng":19.25,"typ":"gmina","sieci":1},{"n":"Oświęcim","lat":50.04,"lng":19.22,"typ":"gmina","sieci":3},{"n":"Brzeszcze","lat":49.98,"lng":19.15,"typ":"gmina","sieci":2},{"n":"Polanka W.","lat":49.99,"lng":19.35,"typ":"gmina","sieci":1},{"n":"Spytkowice","lat":49.99,"lng":19.53,"typ":"gmina","sieci":1},{"n":"Tomice","lat":49.9,"lng":19.55,"typ":"gmina","sieci":1},{"n":"Wieprz","lat":49.87,"lng":19.45,"typ":"gmina","sieci":1},{"n":"Stryszów","lat":49.82,"lng":19.6,"typ":"gmina","sieci":1},{"n":"Lanckorona","lat":49.85,"lng":19.72,"typ":"gmina","sieci":1},{"n":"Budzów","lat":49.78,"lng":19.7,"typ":"gmina","sieci":1},{"n":"Zembrzyce","lat":49.75,"lng":19.62,"typ":"gmina","sieci":1}],
  "dolnośląskie": [{"n":"Kąty Wrocławskie","lat":51.03,"lng":16.77,"typ":"gmina","sieci":3},{"n":"Kobierzyce","lat":50.96,"lng":16.93,"typ":"gmina","sieci":3},{"n":"Żórawina","lat":50.98,"lng":17.04,"typ":"gmina","sieci":2},{"n":"Siechnice","lat":51.03,"lng":17.15,"typ":"gmina","sieci":3},{"n":"Długołęka","lat":51.17,"lng":17.18,"typ":"gmina","sieci":3},{"n":"Czernica","lat":51.06,"lng":17.25,"typ":"gmina","sieci":2},{"n":"Miękinia","lat":51.16,"lng":16.76,"typ":"gmina","sieci":2},{"n":"Sobótka","lat":50.90,"lng":16.74,"typ":"gmina","sieci":3},{"n":"Oborniki Śląskie","lat":51.30,"lng":16.91,"typ":"gmina","sieci":3},{"n":"Jelcz-Laskowice","lat":51.03,"lng":17.33,"typ":"gmina","sieci":3},{"n":"Brzeg Dolny","lat":51.26,"lng":16.72,"typ":"gmina","sieci":2},{"n":"Wołów","lat":51.34,"lng":16.64,"typ":"gmina","sieci":3},{"n":"Żmigród","lat":51.47,"lng":16.90,"typ":"gmina","sieci":3},{"n":"Prusice","lat":51.37,"lng":16.96,"typ":"gmina","sieci":2},{"n":"Milicz","lat":51.53,"lng":17.28,"typ":"gmina","sieci":3},{"n":"Twardogóra","lat":51.36,"lng":17.47,"typ":"gmina","sieci":2},{"n":"Bierutów","lat":51.74,"lng":16.54,"typ":"gmina","sieci":2},{"n":"Strzelin","lat":50.78,"lng":17.06,"typ":"gmina","sieci":4},{"n":"Wiązów","lat":50.81,"lng":17.20,"typ":"gmina","sieci":2},{"n":"Borów","lat":50.79,"lng":17.02,"typ":"gmina","sieci":1},{"n":"Kondratowice","lat":50.77,"lng":16.93,"typ":"gmina","sieci":1},{"n":"Jordanów Śląski","lat":50.86,"lng":16.87,"typ":"gmina","sieci":2},{"n":"Marcinowice","lat":50.95,"lng":16.53,"typ":"gmina","sieci":2},{"n":"Żarów","lat":50.94,"lng":16.50,"typ":"gmina","sieci":3},{"n":"Jaworzyna Śl.","lat":50.91,"lng":16.44,"typ":"gmina","sieci":2},{"n":"Strzegom","lat":50.96,"lng":16.35,"typ":"gmina","sieci":4},{"n":"Udanin","lat":50.99,"lng":16.45,"typ":"gmina","sieci":2},{"n":"Kostomłoty","lat":51.01,"lng":16.27,"typ":"gmina","sieci":2},{"n":"Ziębice","lat":50.60,"lng":17.04,"typ":"gmina","sieci":3},{"n":"Ząbkowice Śl.","lat":50.59,"lng":16.81,"typ":"gmina","sieci":4},{"n":"Ciepłowody","lat":50.67,"lng":16.90,"typ":"gmina","sieci":1},{"n":"Kamieniec Ząb.","lat":50.52,"lng":16.88,"typ":"gmina","sieci":2},{"n":"Przeworno","lat":50.68,"lng":17.15,"typ":"gmina","sieci":1},{"n":"Grodków","lat":50.69,"lng":17.38,"typ":"gmina","sieci":3},{"n":"Łagiewniki","lat":50.79,"lng":16.84,"typ":"gmina","sieci":2},{"n":"Niemcza","lat":50.72,"lng":16.83,"typ":"gmina","sieci":2},{"n":"Piława Górna","lat":50.68,"lng":16.74,"typ":"gmina","sieci":2},{"n":"Dzierżoniów","lat":50.73,"lng":16.65,"typ":"gmina","sieci":4},{"n":"Pieszyce","lat":50.71,"lng":16.76,"typ":"gmina","sieci":2},{"n":"Bielawa","lat":50.68,"lng":16.61,"typ":"gmina","sieci":3},{"n":"Stoszowice","lat":50.60,"lng":16.74,"typ":"gmina","sieci":1},{"n":"Złoty Stok","lat":50.44,"lng":16.87,"typ":"gmina","sieci":2},{"n":"Kąty Wr.","lat":51.03,"lng":16.77,"typ":"gmina","sieci":3},{"n":"Wisznia Mała","lat":51.24,"lng":17.08,"typ":"gmina","sieci":2},{"n":"Oborniki Śl.","lat":51.3,"lng":16.92,"typ":"gmina","sieci":3},{"n":"Trzebnica","lat":51.31,"lng":17.06,"typ":"gmina","sieci":4},{"n":"Środa Śl.","lat":51.16,"lng":16.6,"typ":"gmina","sieci":3},{"n":"Jordanów","lat":50.94,"lng":16.83,"typ":"gmina","sieci":2},{"n":"Mietków","lat":50.98,"lng":16.62,"typ":"gmina","sieci":1},{"n":"Oleśnica","lat":51.21,"lng":17.38,"typ":"gmina","sieci":3},{"n":"Dobroszyce","lat":51.28,"lng":17.36,"typ":"gmina","sieci":1},{"n":"Syców","lat":51.3,"lng":17.72,"typ":"gmina","sieci":2},{"n":"Międzybórz","lat":51.42,"lng":17.62,"typ":"gmina","sieci":1},{"n":"Oława","lat":50.94,"lng":17.3,"typ":"gmina","sieci":3},{"n":"Jelcz","lat":51.02,"lng":17.32,"typ":"gmina","sieci":2},{"n":"Domaniów","lat":50.9,"lng":17.28,"typ":"gmina","sieci":1},{"n":"Bardo","lat":50.5,"lng":16.75,"typ":"gmina","sieci":1},{"n":"Kamieniec Z.","lat":50.52,"lng":16.88,"typ":"gmina","sieci":1}],
  "pomorskie": [{"n":"Żukowo","lat":54.34,"lng":18.36,"typ":"gmina","sieci":3},{"n":"Skarszewy","lat":54.06,"lng":18.45,"typ":"gmina","sieci":2},{"n":"Rumia","lat":54.57,"lng":18.39,"typ":"gmina","sieci":4},{"n":"Reda","lat":54.60,"lng":18.35,"typ":"gmina","sieci":3},{"n":"Luzino","lat":54.56,"lng":18.10,"typ":"gmina","sieci":2},{"n":"Szemud","lat":54.49,"lng":18.22,"typ":"gmina","sieci":2},{"n":"Kosakowo","lat":54.59,"lng":18.48,"typ":"gmina","sieci":2},{"n":"Przodkowo","lat":54.38,"lng":18.26,"typ":"gmina","sieci":2},{"n":"Kolbudy","lat":54.27,"lng":18.46,"typ":"gmina","sieci":2},{"n":"Trąbki Wielkie","lat":54.17,"lng":18.54,"typ":"gmina","sieci":2},{"n":"Pszczółki","lat":54.17,"lng":18.70,"typ":"gmina","sieci":2},{"n":"Przywidz","lat":54.19,"lng":18.32,"typ":"gmina","sieci":2},{"n":"Somonino","lat":54.28,"lng":18.19,"typ":"gmina","sieci":2},{"n":"Stężyca","lat":54.21,"lng":17.96,"typ":"gmina","sieci":2},{"n":"Chmielno","lat":54.32,"lng":18.10,"typ":"gmina","sieci":2},{"n":"Sierakowice","lat":54.34,"lng":17.89,"typ":"gmina","sieci":3},{"n":"Sulęczyno","lat":54.23,"lng":17.77,"typ":"gmina","sieci":2},{"n":"Linia","lat":54.45,"lng":17.98,"typ":"gmina","sieci":2},{"n":"Krokowa","lat":54.78,"lng":18.16,"typ":"gmina","sieci":2},{"n":"Władysławowo","lat":54.79,"lng":18.40,"typ":"gmina","sieci":4},{"n":"Gniewino","lat":54.71,"lng":18.01,"typ":"gmina","sieci":2},{"n":"Łęczyce","lat":54.60,"lng":17.86,"typ":"gmina","sieci":2},{"n":"Kartuzy","lat":54.33,"lng":18.19,"typ":"gmina","sieci":4},{"n":"Pruszcz Gdań.","lat":54.26,"lng":18.63,"typ":"gmina","sieci":4},{"n":"Puck","lat":54.71,"lng":18.40,"typ":"gmina","sieci":3},{"n":"Kępice","lat":54.24,"lng":16.88,"typ":"gmina","sieci":2},{"n":"Trzebielino","lat":54.33,"lng":16.81,"typ":"gmina","sieci":1},{"n":"Kołczygłowy","lat":54.24,"lng":17.23,"typ":"gmina","sieci":1},{"n":"Dębnica Kasz.","lat":54.26,"lng":17.16,"typ":"gmina","sieci":2},{"n":"Czarna Dąbr.","lat":54.35,"lng":17.56,"typ":"gmina","sieci":2},{"n":"Borzytuchom","lat":54.19,"lng":17.37,"typ":"gmina","sieci":1},{"n":"Tuchomie","lat":54.12,"lng":17.33,"typ":"gmina","sieci":1},{"n":"Lipnica","lat":54.00,"lng":17.40,"typ":"gmina","sieci":2},{"n":"Parchowo","lat":54.20,"lng":17.66,"typ":"gmina","sieci":2},{"n":"Studzienice","lat":54.09,"lng":17.57,"typ":"gmina","sieci":1},{"n":"Bytów","lat":54.17,"lng":17.49,"typ":"gmina","sieci":4},{"n":"Miastko","lat":54.00,"lng":16.98,"typ":"gmina","sieci":4},{"n":"Koczała","lat":54.05,"lng":17.06,"typ":"gmina","sieci":1},{"n":"Przechlewo","lat":53.80,"lng":17.18,"typ":"gmina","sieci":2},{"n":"Pruszcz Gd.","lat":54.26,"lng":18.63,"typ":"gmina","sieci":4},{"n":"Trąbki Wlk.","lat":54.19,"lng":18.62,"typ":"gmina","sieci":1},{"n":"Cedry Wlk.","lat":54.24,"lng":18.79,"typ":"gmina","sieci":1},{"n":"Suchy Dąb","lat":54.18,"lng":18.75,"typ":"gmina","sieci":1},{"n":"Tczew okolice","lat":54.09,"lng":18.78,"typ":"gmina","sieci":2},{"n":"Subkowy","lat":54.02,"lng":18.76,"typ":"gmina","sieci":1},{"n":"Gniew","lat":53.84,"lng":18.82,"typ":"gmina","sieci":2},{"n":"Starogard okolice","lat":53.96,"lng":18.53,"typ":"gmina","sieci":2},{"n":"Zblewo","lat":53.93,"lng":18.32,"typ":"gmina","sieci":1},{"n":"Kaliska","lat":53.89,"lng":18.24,"typ":"gmina","sieci":1},{"n":"Lubichowo","lat":53.86,"lng":18.42,"typ":"gmina","sieci":1},{"n":"Osieczna","lat":53.9,"lng":18.13,"typ":"gmina","sieci":1},{"n":"Czarna Woda","lat":53.87,"lng":18.09,"typ":"gmina","sieci":1},{"n":"Osiek","lat":53.79,"lng":18.35,"typ":"gmina","sieci":1},{"n":"Wejherowo","lat":54.6,"lng":18.24,"typ":"gmina","sieci":3},{"n":"Jastarnia","lat":54.7,"lng":18.68,"typ":"gmina","sieci":1},{"n":"Hel","lat":54.61,"lng":18.8,"typ":"gmina","sieci":1},{"n":"Choczewo","lat":54.65,"lng":17.9,"typ":"gmina","sieci":1},{"n":"Czarne","lat":53.68,"lng":16.93,"typ":"gmina","sieci":1},{"n":"Człuchów","lat":53.66,"lng":17.36,"typ":"gmina","sieci":3},{"n":"Debrzno","lat":53.54,"lng":17.24,"typ":"gmina","sieci":1}],
  "lubelskie": [{"n":"Niemce","lat":51.35,"lng":22.63,"typ":"gmina","sieci":2},{"n":"Poniatowa","lat":51.17,"lng":22.06,"typ":"gmina","sieci":2},{"n":"Bychawa","lat":51.01,"lng":22.53,"typ":"gmina","sieci":2},{"n":"Bełżyce","lat":51.17,"lng":22.28,"typ":"gmina","sieci":2},{"n":"Niedrzwica D.","lat":51.11,"lng":22.38,"typ":"gmina","sieci":2},{"n":"Piaski","lat":51.13,"lng":22.84,"typ":"gmina","sieci":2},{"n":"Nałęczów","lat":51.28,"lng":22.21,"typ":"gmina","sieci":2},{"n":"Garbów","lat":51.35,"lng":22.33,"typ":"gmina","sieci":2},{"n":"Jastków","lat":51.30,"lng":22.44,"typ":"gmina","sieci":2},{"n":"Wąwolnica","lat":51.29,"lng":22.14,"typ":"gmina","sieci":2},{"n":"Kurów","lat":51.39,"lng":22.18,"typ":"gmina","sieci":2},{"n":"Końskowola","lat":51.41,"lng":22.05,"typ":"gmina","sieci":2},{"n":"Kock","lat":51.64,"lng":22.44,"typ":"gmina","sieci":2},{"n":"Kamionka","lat":51.47,"lng":22.46,"typ":"gmina","sieci":2},{"n":"Firlej","lat":51.55,"lng":22.50,"typ":"gmina","sieci":2},{"n":"Spiczyn","lat":51.28,"lng":22.77,"typ":"gmina","sieci":2},{"n":"Milejów","lat":51.22,"lng":22.92,"typ":"gmina","sieci":2},{"n":"Rybczewice","lat":51.06,"lng":22.86,"typ":"gmina","sieci":1},{"n":"Świdnik","lat":51.21,"lng":22.69,"typ":"gmina","sieci":4},{"n":"Łęczna","lat":51.30,"lng":22.88,"typ":"gmina","sieci":3},{"n":"Lubartów","lat":51.46,"lng":22.60,"typ":"gmina","sieci":4},{"n":"Urzędów","lat":50.99,"lng":22.14,"typ":"gmina","sieci":2},{"n":"Zakrzówek","lat":50.93,"lng":22.39,"typ":"gmina","sieci":1},{"n":"Batorz","lat":50.81,"lng":22.49,"typ":"gmina","sieci":1},{"n":"Chodel","lat":51.11,"lng":22.13,"typ":"gmina","sieci":2},{"n":"Opole Lub.","lat":51.14,"lng":21.97,"typ":"gmina","sieci":3},{"n":"Karczmiska","lat":51.22,"lng":22.01,"typ":"gmina","sieci":2},{"n":"Wojciechów","lat":51.23,"lng":22.03,"typ":"gmina","sieci":1},{"n":"Puławy","lat":51.41,"lng":21.96,"typ":"gmina","sieci":5},{"n":"Gołąb","lat":51.35,"lng":22.33,"typ":"gmina","sieci":1},{"n":"Markuszów","lat":51.37,"lng":22.26,"typ":"gmina","sieci":1},{"n":"Baranów","lat":51.55,"lng":22.13,"typ":"gmina","sieci":2},{"n":"Michów","lat":51.52,"lng":22.31,"typ":"gmina","sieci":2},{"n":"Jeziorzany","lat":51.59,"lng":22.27,"typ":"gmina","sieci":1},{"n":"Abramów","lat":51.46,"lng":22.31,"typ":"gmina","sieci":1},{"n":"Ryki","lat":51.62,"lng":21.93,"typ":"gmina","sieci":3},{"n":"Dęblin","lat":51.56,"lng":21.84,"typ":"gmina","sieci":3},{"n":"Stężyca","lat":51.56,"lng":21.78,"typ":"gmina","sieci":2},{"n":"Mełgiew","lat":51.2,"lng":22.77,"typ":"gmina","sieci":2},{"n":"Wólka","lat":51.24,"lng":22.66,"typ":"gmina","sieci":2},{"n":"Konopnica","lat":51.22,"lng":22.44,"typ":"gmina","sieci":2},{"n":"Głusk","lat":51.18,"lng":22.6,"typ":"gmina","sieci":2},{"n":"Strzyżewice","lat":51.1,"lng":22.5,"typ":"gmina","sieci":1},{"n":"Krzczonów","lat":51.05,"lng":22.72,"typ":"gmina","sieci":1},{"n":"Borzechów","lat":51.1,"lng":22.3,"typ":"gmina","sieci":1},{"n":"Lubartów okolice","lat":51.46,"lng":22.61,"typ":"gmina","sieci":2},{"n":"Kraśnik","lat":50.92,"lng":22.22,"typ":"gmina","sieci":3},{"n":"Annopol","lat":50.88,"lng":21.85,"typ":"gmina","sieci":1},{"n":"Zaklików","lat":50.79,"lng":22.1,"typ":"gmina","sieci":1},{"n":"Modliborzyce","lat":50.75,"lng":22.33,"typ":"gmina","sieci":1},{"n":"Janów Lub.","lat":50.71,"lng":22.41,"typ":"gmina","sieci":2},{"n":"Frampol","lat":50.67,"lng":22.67,"typ":"gmina","sieci":1},{"n":"Goraj","lat":50.72,"lng":22.66,"typ":"gmina","sieci":1},{"n":"Turobin","lat":50.83,"lng":22.74,"typ":"gmina","sieci":1},{"n":"Szczebrzeszyn","lat":50.7,"lng":22.98,"typ":"gmina","sieci":2},{"n":"Zwierzyniec","lat":50.61,"lng":22.97,"typ":"gmina","sieci":1},{"n":"Józefów","lat":50.48,"lng":23.05,"typ":"gmina","sieci":1},{"n":"Krasnobród","lat":50.55,"lng":23.21,"typ":"gmina","sieci":1},{"n":"Tomaszów Lub.","lat":50.45,"lng":23.42,"typ":"gmina","sieci":3},{"n":"Tarnawatka","lat":50.53,"lng":23.4,"typ":"gmina","sieci":1},{"n":"Łaszczów","lat":50.53,"lng":23.72,"typ":"gmina","sieci":1},{"n":"Tyszowce","lat":50.62,"lng":23.7,"typ":"gmina","sieci":1},{"n":"Komarów","lat":50.65,"lng":23.45,"typ":"gmina","sieci":1},{"n":"Zamość okolice","lat":50.72,"lng":23.25,"typ":"gmina","sieci":3},{"n":"Sitno","lat":50.68,"lng":23.38,"typ":"gmina","sieci":1},{"n":"Skierbieszów","lat":50.85,"lng":23.3,"typ":"gmina","sieci":1},{"n":"Izbica","lat":50.88,"lng":23.16,"typ":"gmina","sieci":1},{"n":"Krasnystaw","lat":50.98,"lng":23.17,"typ":"gmina","sieci":3},{"n":"Fajsławice","lat":51.04,"lng":23.0,"typ":"gmina","sieci":1},{"n":"Łopiennik","lat":51.06,"lng":23.13,"typ":"gmina","sieci":1},{"n":"Siennica Róż.","lat":51.07,"lng":23.25,"typ":"gmina","sieci":1}],
  "kujawsko-pomorskie": [{"n":"Solec Kujawski","lat":53.08,"lng":18.22,"typ":"gmina","sieci":3},{"n":"Koronowo","lat":53.31,"lng":17.93,"typ":"gmina","sieci":3},{"n":"Chełmża","lat":53.18,"lng":18.60,"typ":"gmina","sieci":3},{"n":"Szubin","lat":53.00,"lng":17.74,"typ":"gmina","sieci":3},{"n":"Barcin","lat":52.85,"lng":17.95,"typ":"gmina","sieci":2},{"n":"Łabiszyn","lat":52.95,"lng":17.91,"typ":"gmina","sieci":2},{"n":"Złotniki Kuj.","lat":52.89,"lng":18.14,"typ":"gmina","sieci":2},{"n":"Pakość","lat":52.80,"lng":18.08,"typ":"gmina","sieci":2},{"n":"Ciechocinek","lat":52.88,"lng":18.79,"typ":"gmina","sieci":3},{"n":"Gniewkowo","lat":52.89,"lng":18.41,"typ":"gmina","sieci":2},{"n":"Janikowo","lat":52.75,"lng":18.11,"typ":"gmina","sieci":2},{"n":"Kruszwica","lat":52.68,"lng":18.32,"typ":"gmina","sieci":3},{"n":"Strzelno","lat":52.63,"lng":18.17,"typ":"gmina","sieci":2},{"n":"Radziejów","lat":52.62,"lng":18.52,"typ":"gmina","sieci":3},{"n":"Piotrków Kuj.","lat":52.55,"lng":18.49,"typ":"gmina","sieci":2},{"n":"Zakrzewo","lat":52.76,"lng":18.63,"typ":"gmina","sieci":1},{"n":"Koneck","lat":52.78,"lng":18.71,"typ":"gmina","sieci":1},{"n":"Aleksandrów Kuj.","lat":52.87,"lng":18.70,"typ":"gmina","sieci":3},{"n":"Raciążek","lat":52.82,"lng":18.81,"typ":"gmina","sieci":2},{"n":"Lubanie","lat":52.77,"lng":18.93,"typ":"gmina","sieci":2},{"n":"Waganiec","lat":52.80,"lng":18.85,"typ":"gmina","sieci":1},{"n":"Skępe","lat":52.86,"lng":19.34,"typ":"gmina","sieci":2},{"n":"Tłuchowo","lat":52.74,"lng":19.46,"typ":"gmina","sieci":2},{"n":"Dobrzyń n. Wisłą","lat":52.63,"lng":19.32,"typ":"gmina","sieci":2},{"n":"Nakło n. Notecią","lat":53.14,"lng":17.60,"typ":"gmina","sieci":3},{"n":"Sadki","lat":53.15,"lng":17.44,"typ":"gmina","sieci":2},{"n":"Mrocza","lat":53.24,"lng":17.60,"typ":"gmina","sieci":2},{"n":"Kcynia","lat":53.30,"lng":17.48,"typ":"gmina","sieci":2},{"n":"Sicienko","lat":53.20,"lng":17.81,"typ":"gmina","sieci":1},{"n":"Witosław","lat":53.18,"lng":17.40,"typ":"gmina","sieci":1},{"n":"Śmielin","lat":53.19,"lng":17.45,"typ":"gmina","sieci":1},{"n":"Dziewierzewo","lat":53.21,"lng":17.51,"typ":"gmina","sieci":1},{"n":"Gołańcz","lat":52.94,"lng":17.29,"typ":"gmina","sieci":2},{"n":"Damasławek","lat":52.84,"lng":17.48,"typ":"gmina","sieci":2},{"n":"Wapno","lat":53.07,"lng":17.24,"typ":"gmina","sieci":1},{"n":"Żnin","lat":52.98,"lng":17.67,"typ":"gmina","sieci":3},{"n":"Janowiec Wlkp.","lat":52.75,"lng":17.49,"typ":"gmina","sieci":2},{"n":"Rogowo","lat":52.72,"lng":17.64,"typ":"gmina","sieci":2},{"n":"Gąsawa","lat":52.76,"lng":17.74,"typ":"gmina","sieci":2},{"n":"Solec Kuj.","lat":53.08,"lng":18.22,"typ":"gmina","sieci":3},{"n":"Białe Błota","lat":53.08,"lng":17.93,"typ":"gmina","sieci":2},{"n":"Nowa Wieś Wlk.","lat":53.14,"lng":17.86,"typ":"gmina","sieci":2},{"n":"Osielsko","lat":53.18,"lng":18.05,"typ":"gmina","sieci":2},{"n":"Dobrcz","lat":53.23,"lng":18.19,"typ":"gmina","sieci":1},{"n":"Dąbrowa Chełm.","lat":53.13,"lng":18.42,"typ":"gmina","sieci":1},{"n":"Unisław","lat":53.18,"lng":18.38,"typ":"gmina","sieci":1},{"n":"Kowalewo Pom.","lat":53.17,"lng":18.86,"typ":"gmina","sieci":2},{"n":"Łysomice","lat":53.09,"lng":18.63,"typ":"gmina","sieci":1},{"n":"Lubicz","lat":53.02,"lng":18.78,"typ":"gmina","sieci":2},{"n":"Obrowo","lat":52.96,"lng":18.87,"typ":"gmina","sieci":1},{"n":"Czernikowo","lat":52.94,"lng":18.92,"typ":"gmina","sieci":1},{"n":"Wielka Nieszawka","lat":53.0,"lng":18.5,"typ":"gmina","sieci":1},{"n":"Zławieś Wlk.","lat":53.08,"lng":18.42,"typ":"gmina","sieci":1},{"n":"Nakło","lat":53.14,"lng":17.6,"typ":"gmina","sieci":3},{"n":"Więcbork","lat":53.35,"lng":17.51,"typ":"gmina","sieci":1},{"n":"Sępólno Kr.","lat":53.45,"lng":17.53,"typ":"gmina","sieci":2},{"n":"Tuchola","lat":53.59,"lng":17.86,"typ":"gmina","sieci":2},{"n":"Śliwice","lat":53.7,"lng":18.19,"typ":"gmina","sieci":1},{"n":"Cekcyn","lat":53.6,"lng":18.1,"typ":"gmina","sieci":1},{"n":"Kęsowo","lat":53.52,"lng":17.72,"typ":"gmina","sieci":1},{"n":"Gostycyn","lat":53.51,"lng":17.86,"typ":"gmina","sieci":1},{"n":"Świecie","lat":53.41,"lng":18.44,"typ":"gmina","sieci":3},{"n":"Nowe","lat":53.65,"lng":18.73,"typ":"gmina","sieci":1},{"n":"Warlubie","lat":53.55,"lng":18.65,"typ":"gmina","sieci":1},{"n":"Jeżewo","lat":53.51,"lng":18.55,"typ":"gmina","sieci":1},{"n":"Bukowiec","lat":53.44,"lng":18.34,"typ":"gmina","sieci":1},{"n":"Pruszcz","lat":53.36,"lng":18.32,"typ":"gmina","sieci":1},{"n":"Świekatowo","lat":53.42,"lng":18.16,"typ":"gmina","sieci":1},{"n":"Lniano","lat":53.48,"lng":18.3,"typ":"gmina","sieci":1},{"n":"Drzycim","lat":53.51,"lng":18.42,"typ":"gmina","sieci":1},{"n":"Grudziądz","lat":53.48,"lng":18.75,"typ":"gmina","sieci":4},{"n":"Łasin","lat":53.52,"lng":19.09,"typ":"gmina","sieci":1},{"n":"Radzyń Chełm.","lat":53.38,"lng":18.93,"typ":"gmina","sieci":1},{"n":"Gruta","lat":53.46,"lng":18.98,"typ":"gmina","sieci":1},{"n":"Rogóźno","lat":53.43,"lng":18.9,"typ":"gmina","sieci":1},{"n":"Świecie n.Osą","lat":53.45,"lng":19.03,"typ":"gmina","sieci":1}],
  "zachodniopomorskie": [{"n":"Chojna","lat":52.96,"lng":14.42,"typ":"gmina","sieci":2},{"n":"Nowogard","lat":53.66,"lng":15.11,"typ":"gmina","sieci":3},{"n":"Maszewo","lat":53.49,"lng":15.05,"typ":"gmina","sieci":2},{"n":"Stepnica","lat":53.65,"lng":14.62,"typ":"gmina","sieci":2},{"n":"Wolin","lat":53.84,"lng":14.61,"typ":"gmina","sieci":3},{"n":"Międzyzdroje","lat":53.92,"lng":14.44,"typ":"gmina","sieci":3},{"n":"Kołbaskowo","lat":53.33,"lng":14.43,"typ":"gmina","sieci":2},{"n":"Nowe Warpno","lat":53.72,"lng":14.28,"typ":"gmina","sieci":1},{"n":"Banie","lat":53.10,"lng":14.66,"typ":"gmina","sieci":2},{"n":"Widuchowa","lat":53.12,"lng":14.38,"typ":"gmina","sieci":2},{"n":"Cedynia","lat":52.88,"lng":14.20,"typ":"gmina","sieci":2},{"n":"Mieszkowice","lat":52.78,"lng":14.49,"typ":"gmina","sieci":2},{"n":"Moryń","lat":52.85,"lng":14.39,"typ":"gmina","sieci":2},{"n":"Trzcińsko-Zdrój","lat":52.96,"lng":14.61,"typ":"gmina","sieci":2},{"n":"Kozielice","lat":53.07,"lng":14.80,"typ":"gmina","sieci":1},{"n":"Warnice","lat":53.25,"lng":14.99,"typ":"gmina","sieci":1},{"n":"Bielice","lat":53.16,"lng":14.65,"typ":"gmina","sieci":2},{"n":"Przelewice","lat":53.10,"lng":15.08,"typ":"gmina","sieci":2},{"n":"Stare Czarnowo","lat":53.27,"lng":14.77,"typ":"gmina","sieci":2},{"n":"Kobylanka","lat":53.34,"lng":14.88,"typ":"gmina","sieci":2},{"n":"Stara Dąbrowa","lat":53.44,"lng":15.13,"typ":"gmina","sieci":1},{"n":"Dziwnów","lat":53.94,"lng":14.75,"typ":"gmina","sieci":3},{"n":"Kamień Pomorski","lat":53.97,"lng":14.76,"typ":"gmina","sieci":3},{"n":"Świerzno","lat":54.12,"lng":16.03,"typ":"gmina","sieci":1},{"n":"Gryfice","lat":53.91,"lng":15.19,"typ":"gmina","sieci":4},{"n":"Karnice","lat":54.02,"lng":15.02,"typ":"gmina","sieci":1},{"n":"Rewal","lat":54.08,"lng":15.01,"typ":"gmina","sieci":3},{"n":"Trzebiatów","lat":54.06,"lng":15.26,"typ":"gmina","sieci":3},{"n":"Brojce","lat":53.98,"lng":15.34,"typ":"gmina","sieci":1},{"n":"Płoty","lat":53.27,"lng":17.02,"typ":"gmina","sieci":2},{"n":"Resko","lat":53.64,"lng":15.40,"typ":"gmina","sieci":2},{"n":"Radowo Małe","lat":53.63,"lng":15.43,"typ":"gmina","sieci":1},{"n":"Węgorzyno","lat":53.54,"lng":15.55,"typ":"gmina","sieci":2},{"n":"Dobra","lat":53.58,"lng":15.30,"typ":"gmina","sieci":2},{"n":"Marianowo","lat":53.29,"lng":15.15,"typ":"gmina","sieci":1},{"n":"Suchań","lat":53.27,"lng":15.32,"typ":"gmina","sieci":1},{"n":"Dobrzany","lat":53.35,"lng":15.42,"typ":"gmina","sieci":2},{"n":"Chociwel","lat":53.27,"lng":15.33,"typ":"gmina","sieci":2},{"n":"Dobra Szcz.","lat":53.46,"lng":14.44,"typ":"gmina","sieci":2},{"n":"Police","lat":53.55,"lng":14.57,"typ":"gmina","sieci":4},{"n":"Goleniów","lat":53.56,"lng":14.83,"typ":"gmina","sieci":4},{"n":"Gryfino","lat":53.25,"lng":14.49,"typ":"gmina","sieci":3},{"n":"Stargard okolice","lat":53.34,"lng":15.05,"typ":"gmina","sieci":2},{"n":"Pyrzyce","lat":53.15,"lng":14.89,"typ":"gmina","sieci":3},{"n":"Lipiany","lat":53.01,"lng":14.97,"typ":"gmina","sieci":1},{"n":"Barlinek","lat":52.99,"lng":15.22,"typ":"gmina","sieci":3},{"n":"Myślibórz","lat":52.92,"lng":14.87,"typ":"gmina","sieci":3},{"n":"Dębno","lat":52.74,"lng":14.7,"typ":"gmina","sieci":3},{"n":"Boleszkowice","lat":52.72,"lng":14.57,"typ":"gmina","sieci":1},{"n":"Świnoujście","lat":53.91,"lng":14.25,"typ":"gmina","sieci":3},{"n":"Kamień Pom.","lat":53.97,"lng":14.77,"typ":"gmina","sieci":2},{"n":"Golczewo","lat":53.82,"lng":14.98,"typ":"gmina","sieci":1},{"n":"Świdwin","lat":53.77,"lng":15.78,"typ":"gmina","sieci":2},{"n":"Połczyn-Zdrój","lat":53.76,"lng":16.1,"typ":"gmina","sieci":2},{"n":"Białogard","lat":54.01,"lng":15.99,"typ":"gmina","sieci":3},{"n":"Karlino","lat":54.04,"lng":15.87,"typ":"gmina","sieci":1},{"n":"Tychowo","lat":53.9,"lng":16.25,"typ":"gmina","sieci":1},{"n":"Szczecinek","lat":53.71,"lng":16.7,"typ":"gmina","sieci":3},{"n":"Borne Sulinowo","lat":53.58,"lng":16.53,"typ":"gmina","sieci":1},{"n":"Barwice","lat":53.75,"lng":16.35,"typ":"gmina","sieci":1},{"n":"Grzmiąca","lat":53.83,"lng":16.42,"typ":"gmina","sieci":1},{"n":"Czaplinek","lat":53.55,"lng":16.23,"typ":"gmina","sieci":2},{"n":"Drawsko Pom.","lat":53.53,"lng":15.81,"typ":"gmina","sieci":2},{"n":"Złocieniec","lat":53.53,"lng":16.01,"typ":"gmina","sieci":2},{"n":"Kalisz Pom.","lat":53.3,"lng":15.9,"typ":"gmina","sieci":1},{"n":"Wałcz","lat":53.27,"lng":16.47,"typ":"gmina","sieci":3},{"n":"Mirosławiec","lat":53.35,"lng":16.09,"typ":"gmina","sieci":1}],
  "świętokrzyskie": [{"n":"Chęciny","lat":50.79,"lng":20.46,"typ":"gmina","sieci":2},{"n":"Morawica","lat":50.75,"lng":20.61,"typ":"gmina","sieci":2},{"n":"Suchedniów","lat":51.04,"lng":20.83,"typ":"gmina","sieci":2},{"n":"Piekoszów","lat":50.88,"lng":20.46,"typ":"gmina","sieci":2},{"n":"Daleszyce","lat":50.80,"lng":20.80,"typ":"gmina","sieci":2},{"n":"Górno","lat":50.84,"lng":20.74,"typ":"gmina","sieci":1},{"n":"Masłów","lat":50.90,"lng":20.72,"typ":"gmina","sieci":1},{"n":"Zagnańsk","lat":50.98,"lng":20.66,"typ":"gmina","sieci":2},{"n":"Bodzentyn","lat":50.94,"lng":20.95,"typ":"gmina","sieci":2},{"n":"Miedziana Góra","lat":50.94,"lng":20.58,"typ":"gmina","sieci":2},{"n":"Strawczyn","lat":50.93,"lng":20.42,"typ":"gmina","sieci":2},{"n":"Łopuszno","lat":50.98,"lng":20.26,"typ":"gmina","sieci":2},{"n":"Sobków","lat":50.69,"lng":20.46,"typ":"gmina","sieci":2},{"n":"Małogoszcz","lat":50.81,"lng":20.26,"typ":"gmina","sieci":2},{"n":"Mniów","lat":51.02,"lng":20.48,"typ":"gmina","sieci":2},{"n":"Smyków","lat":51.05,"lng":20.40,"typ":"gmina","sieci":1},{"n":"Stąporków","lat":51.14,"lng":20.55,"typ":"gmina","sieci":3},{"n":"Bliżyn","lat":51.10,"lng":20.76,"typ":"gmina","sieci":2},{"n":"Wąchock","lat":51.07,"lng":21.01,"typ":"gmina","sieci":2},{"n":"Bieliny","lat":50.84,"lng":20.94,"typ":"gmina","sieci":2},{"n":"Łagów","lat":50.77,"lng":21.08,"typ":"gmina","sieci":2},{"n":"Raków","lat":50.67,"lng":21.04,"typ":"gmina","sieci":2},{"n":"Nowa Słupia","lat":50.86,"lng":21.08,"typ":"gmina","sieci":2},{"n":"Krasocin","lat":50.96,"lng":20.11,"typ":"gmina","sieci":2},{"n":"Ruda Maleniecka","lat":51.01,"lng":20.30,"typ":"gmina","sieci":1},{"n":"Fałków","lat":51.07,"lng":20.10,"typ":"gmina","sieci":1},{"n":"Gowarczów","lat":51.17,"lng":20.21,"typ":"gmina","sieci":2},{"n":"Słupia Konecka","lat":51.00,"lng":20.00,"typ":"gmina","sieci":1},{"n":"Secemin","lat":50.58,"lng":19.06,"typ":"gmina","sieci":2},{"n":"Radków","lat":50.71,"lng":19.98,"typ":"gmina","sieci":1},{"n":"Moskorzew","lat":50.66,"lng":19.95,"typ":"gmina","sieci":1},{"n":"Słupia","lat":50.61,"lng":20.01,"typ":"gmina","sieci":2},{"n":"Kluczewsko","lat":50.86,"lng":19.92,"typ":"gmina","sieci":2},{"n":"Włoszczowa","lat":50.85,"lng":19.96,"typ":"gmina","sieci":3},{"n":"Buczek","lat":51.48,"lng":19.16,"typ":"gmina","sieci":1},{"n":"Złotniki","lat":50.65,"lng":20.28,"typ":"gmina","sieci":1},{"n":"Nowy Korczyn","lat":50.31,"lng":20.80,"typ":"gmina","sieci":2},{"n":"Wiślica","lat":50.34,"lng":20.67,"typ":"gmina","sieci":2},{"n":"Opatowiec","lat":50.24,"lng":20.72,"typ":"gmina","sieci":1},{"n":"Solec-Zdrój","lat":50.30,"lng":20.88,"typ":"gmina","sieci":2},{"n":"Sitkówka","lat":50.83,"lng":20.53,"typ":"gmina","sieci":2},{"n":"Pierzchnica","lat":50.68,"lng":20.9,"typ":"gmina","sieci":1},{"n":"Busko-Zdrój","lat":50.47,"lng":20.72,"typ":"gmina","sieci":4},{"n":"Stopnica","lat":50.45,"lng":20.94,"typ":"gmina","sieci":1},{"n":"Jędrzejów","lat":50.64,"lng":20.3,"typ":"gmina","sieci":3},{"n":"Sędziszów","lat":50.57,"lng":20.06,"typ":"gmina","sieci":2},{"n":"Końskie","lat":51.19,"lng":20.41,"typ":"gmina","sieci":3},{"n":"Radoszyce","lat":51.07,"lng":20.25,"typ":"gmina","sieci":1},{"n":"Ruda Malen.","lat":51.15,"lng":20.48,"typ":"gmina","sieci":1},{"n":"Skarżysko","lat":51.11,"lng":20.87,"typ":"gmina","sieci":3},{"n":"Ostrowiec Św.","lat":50.93,"lng":21.39,"typ":"gmina","sieci":4},{"n":"Kunów","lat":50.97,"lng":21.25,"typ":"gmina","sieci":1},{"n":"Ćmielów","lat":50.9,"lng":21.51,"typ":"gmina","sieci":1},{"n":"Bodzechów","lat":50.93,"lng":21.44,"typ":"gmina","sieci":1},{"n":"Waśniów","lat":50.86,"lng":21.28,"typ":"gmina","sieci":1},{"n":"Ożarów","lat":50.88,"lng":21.66,"typ":"gmina","sieci":1},{"n":"Sandomierz","lat":50.68,"lng":21.75,"typ":"gmina","sieci":3},{"n":"Zawichost","lat":50.8,"lng":21.85,"typ":"gmina","sieci":1},{"n":"Klimontów","lat":50.68,"lng":21.42,"typ":"gmina","sieci":1},{"n":"Koprzywnica","lat":50.61,"lng":21.58,"typ":"gmina","sieci":1},{"n":"Łoniów","lat":50.56,"lng":21.55,"typ":"gmina","sieci":1}],
  "podkarpackie": [{"n":"Boguchwała","lat":49.98,"lng":21.93,"typ":"gmina","sieci":3},{"n":"Głogów Młp.","lat":50.15,"lng":21.96,"typ":"gmina","sieci":3},{"n":"Tyczyn","lat":49.96,"lng":22.02,"typ":"gmina","sieci":2},{"n":"Trzebownisko","lat":50.08,"lng":22.05,"typ":"gmina","sieci":2},{"n":"Sędziszów Młp.","lat":50.07,"lng":21.70,"typ":"gmina","sieci":3},{"n":"Dynów","lat":49.82,"lng":22.23,"typ":"gmina","sieci":2},{"n":"Czudec","lat":49.94,"lng":21.83,"typ":"gmina","sieci":2},{"n":"Świlcza","lat":50.06,"lng":21.91,"typ":"gmina","sieci":2},{"n":"Krasne","lat":50.04,"lng":22.06,"typ":"gmina","sieci":2},{"n":"Chmielnik","lat":49.95,"lng":22.14,"typ":"gmina","sieci":2},{"n":"Błażowa","lat":49.88,"lng":22.09,"typ":"gmina","sieci":2},{"n":"Niebylec","lat":49.85,"lng":21.89,"typ":"gmina","sieci":2},{"n":"Wiśniowa","lat":49.87,"lng":21.65,"typ":"gmina","sieci":2},{"n":"Wielopole","lat":49.95,"lng":21.61,"typ":"gmina","sieci":2},{"n":"Iwierzyce","lat":50.03,"lng":21.78,"typ":"gmina","sieci":2},{"n":"Sokołów Młp.","lat":50.23,"lng":22.11,"typ":"gmina","sieci":2},{"n":"Raniżów","lat":50.26,"lng":21.98,"typ":"gmina","sieci":1},{"n":"Dzikowiec","lat":50.27,"lng":21.84,"typ":"gmina","sieci":1},{"n":"Kolbuszowa","lat":50.24,"lng":21.76,"typ":"gmina","sieci":4},{"n":"Niwiska","lat":50.21,"lng":21.75,"typ":"gmina","sieci":1},{"n":"Majdan Król.","lat":50.36,"lng":21.76,"typ":"gmina","sieci":2},{"n":"Cmolas","lat":50.23,"lng":21.67,"typ":"gmina","sieci":1},{"n":"Tuszów Nar.","lat":50.37,"lng":21.46,"typ":"gmina","sieci":2},{"n":"Mielec","lat":50.28,"lng":21.42,"typ":"gmina","sieci":5},{"n":"Przecław","lat":50.19,"lng":21.48,"typ":"gmina","sieci":2},{"n":"Radomyśl Wlk.","lat":50.19,"lng":21.27,"typ":"gmina","sieci":3},{"n":"Wadowice G.","lat":50.20,"lng":21.24,"typ":"gmina","sieci":1},{"n":"Żyraków","lat":50.07,"lng":21.36,"typ":"gmina","sieci":2},{"n":"Czarna","lat":50.06,"lng":21.24,"typ":"gmina","sieci":2},{"n":"Pilzno","lat":49.97,"lng":21.29,"typ":"gmina","sieci":2},{"n":"Jodłowa","lat":49.88,"lng":21.29,"typ":"gmina","sieci":2},{"n":"Brzostek","lat":49.88,"lng":21.44,"typ":"gmina","sieci":2},{"n":"Frysztak","lat":49.83,"lng":21.61,"typ":"gmina","sieci":2},{"n":"Kołaczyce","lat":49.81,"lng":21.43,"typ":"gmina","sieci":2},{"n":"Brzyska","lat":49.81,"lng":21.46,"typ":"gmina","sieci":1},{"n":"Jasło","lat":49.74,"lng":21.47,"typ":"gmina","sieci":4},{"n":"Hyżne","lat":49.9,"lng":22.2,"typ":"gmina","sieci":1},{"n":"Kańczuga","lat":49.99,"lng":22.41,"typ":"gmina","sieci":2},{"n":"Pruchnik","lat":49.9,"lng":22.51,"typ":"gmina","sieci":1},{"n":"Roźwienica","lat":49.96,"lng":22.55,"typ":"gmina","sieci":1},{"n":"Jawornik","lat":49.89,"lng":22.3,"typ":"gmina","sieci":1},{"n":"Kamień","lat":50.19,"lng":22.05,"typ":"gmina","sieci":1},{"n":"Łańcut","lat":50.07,"lng":22.23,"typ":"gmina","sieci":3},{"n":"Leżajsk","lat":50.26,"lng":22.42,"typ":"gmina","sieci":3},{"n":"Nowa Sarzyna","lat":50.32,"lng":22.32,"typ":"gmina","sieci":1},{"n":"Nisko","lat":50.52,"lng":22.14,"typ":"gmina","sieci":3},{"n":"Stalowa Wola","lat":50.58,"lng":22.05,"typ":"gmina","sieci":4},{"n":"Rudnik","lat":50.44,"lng":22.25,"typ":"gmina","sieci":1},{"n":"Ulanów","lat":50.49,"lng":22.27,"typ":"gmina","sieci":1},{"n":"Krzeszów","lat":50.42,"lng":22.34,"typ":"gmina","sieci":1},{"n":"Tarnobrzeg","lat":50.57,"lng":21.68,"typ":"gmina","sieci":3},{"n":"Baranów Sand.","lat":50.5,"lng":21.53,"typ":"gmina","sieci":1},{"n":"Nowa Dęba","lat":50.43,"lng":21.77,"typ":"gmina","sieci":2},{"n":"Majdan Kr.","lat":50.36,"lng":21.9,"typ":"gmina","sieci":1},{"n":"Ropczyce","lat":50.05,"lng":21.61,"typ":"gmina","sieci":3},{"n":"Dębica","lat":50.05,"lng":21.41,"typ":"gmina","sieci":3},{"n":"Strzyżów","lat":49.87,"lng":21.79,"typ":"gmina","sieci":2}],
  "podlaskie": [{"n":"Choroszcz","lat":53.14,"lng":22.98,"typ":"gmina","sieci":2},{"n":"Wasilków","lat":53.20,"lng":23.20,"typ":"gmina","sieci":3},{"n":"Supraśl","lat":53.21,"lng":23.33,"typ":"gmina","sieci":2},{"n":"Czarna Białost.","lat":53.30,"lng":23.28,"typ":"gmina","sieci":2},{"n":"Zabłudów","lat":53.01,"lng":23.34,"typ":"gmina","sieci":2},{"n":"Łapy","lat":52.99,"lng":22.88,"typ":"gmina","sieci":3},{"n":"Michałowo","lat":53.03,"lng":23.60,"typ":"gmina","sieci":2},{"n":"Juchnowiec Kośc.","lat":53.01,"lng":23.13,"typ":"gmina","sieci":2},{"n":"Dobrzyniewo","lat":53.20,"lng":23.01,"typ":"gmina","sieci":2},{"n":"Turośń Kościelna","lat":52.98,"lng":23.06,"typ":"gmina","sieci":2},{"n":"Suraż","lat":52.95,"lng":22.95,"typ":"gmina","sieci":1},{"n":"Narew","lat":52.91,"lng":23.52,"typ":"gmina","sieci":2},{"n":"Narewka","lat":52.83,"lng":23.76,"typ":"gmina","sieci":2},{"n":"Czyże","lat":52.78,"lng":23.42,"typ":"gmina","sieci":1},{"n":"Krynki","lat":53.26,"lng":23.77,"typ":"gmina","sieci":2},{"n":"Janów","lat":53.47,"lng":23.22,"typ":"gmina","sieci":2},{"n":"Dąbrowa Białost.","lat":53.65,"lng":23.35,"typ":"gmina","sieci":2},{"n":"Suchowola","lat":53.58,"lng":23.10,"typ":"gmina","sieci":2},{"n":"Korycin","lat":53.44,"lng":23.15,"typ":"gmina","sieci":2},{"n":"Jaświły","lat":53.48,"lng":23.00,"typ":"gmina","sieci":1},{"n":"Knyszyn","lat":53.31,"lng":22.92,"typ":"gmina","sieci":2},{"n":"Trzcianne","lat":53.44,"lng":22.68,"typ":"gmina","sieci":1},{"n":"Mońki","lat":53.40,"lng":22.79,"typ":"gmina","sieci":3},{"n":"Goniądz","lat":53.48,"lng":22.73,"typ":"gmina","sieci":2},{"n":"Krypno","lat":53.27,"lng":22.87,"typ":"gmina","sieci":1},{"n":"Sokółka","lat":53.40,"lng":23.49,"typ":"gmina","sieci":4},{"n":"Szudziałowo","lat":53.29,"lng":23.65,"typ":"gmina","sieci":1},{"n":"Kuznica","lat":53.50,"lng":23.64,"typ":"gmina","sieci":1},{"n":"Sidra","lat":53.54,"lng":23.46,"typ":"gmina","sieci":1},{"n":"Lipsk","lat":53.73,"lng":23.39,"typ":"gmina","sieci":2},{"n":"Sztabin","lat":54.14,"lng":23.11,"typ":"gmina","sieci":1},{"n":"Augustów","lat":53.84,"lng":22.97,"typ":"gmina","sieci":4},{"n":"Bargłów Kośc.","lat":53.76,"lng":22.82,"typ":"gmina","sieci":1},{"n":"Grajewo","lat":53.64,"lng":22.45,"typ":"gmina","sieci":4},{"n":"Szczuczyn","lat":53.56,"lng":22.28,"typ":"gmina","sieci":2},{"n":"Wąsosz","lat":53.52,"lng":22.31,"typ":"gmina","sieci":1},{"n":"Radziłów","lat":53.46,"lng":22.40,"typ":"gmina","sieci":1},{"n":"Zawady","lat":53.15,"lng":22.66,"typ":"gmina","sieci":1},{"n":"Kobylin-Borzymy","lat":53.10,"lng":22.66,"typ":"gmina","sieci":1},{"n":"Czarna Białostocka","lat":53.3,"lng":23.29,"typ":"gmina","sieci":2},{"n":"Turośń","lat":52.98,"lng":23.05,"typ":"gmina","sieci":1},{"n":"Juchnowiec","lat":53.06,"lng":23.24,"typ":"gmina","sieci":2},{"n":"Sokoły","lat":52.98,"lng":22.68,"typ":"gmina","sieci":1},{"n":"Kobylin","lat":52.9,"lng":22.71,"typ":"gmina","sieci":1},{"n":"Kulesze","lat":52.92,"lng":22.6,"typ":"gmina","sieci":1},{"n":"Wysokie Maz.","lat":52.92,"lng":22.51,"typ":"gmina","sieci":3},{"n":"Czyżew","lat":52.8,"lng":22.3,"typ":"gmina","sieci":2},{"n":"Szepietowo","lat":52.86,"lng":22.55,"typ":"gmina","sieci":2},{"n":"Ciechanowiec","lat":52.68,"lng":22.5,"typ":"gmina","sieci":2},{"n":"Brańsk","lat":52.74,"lng":22.84,"typ":"gmina","sieci":2},{"n":"Rajgród","lat":53.73,"lng":22.7,"typ":"gmina","sieci":1},{"n":"Dąbrowa Biał.","lat":53.65,"lng":23.35,"typ":"gmina","sieci":1},{"n":"Kuźnica","lat":53.51,"lng":23.65,"typ":"gmina","sieci":1},{"n":"Gródek","lat":53.1,"lng":23.66,"typ":"gmina","sieci":1},{"n":"Bielsk Podl.","lat":52.77,"lng":23.19,"typ":"gmina","sieci":3},{"n":"Boćki","lat":52.66,"lng":23.04,"typ":"gmina","sieci":1},{"n":"Orla","lat":52.71,"lng":23.33,"typ":"gmina","sieci":1},{"n":"Kleszczele","lat":52.58,"lng":23.32,"typ":"gmina","sieci":1},{"n":"Hajnówka","lat":52.74,"lng":23.58,"typ":"gmina","sieci":3},{"n":"Siemiatycze","lat":52.43,"lng":22.86,"typ":"gmina","sieci":3},{"n":"Drohiczyn","lat":52.4,"lng":22.66,"typ":"gmina","sieci":1},{"n":"Mielnik","lat":52.33,"lng":23.05,"typ":"gmina","sieci":1},{"n":"Nurzec","lat":52.44,"lng":23.06,"typ":"gmina","sieci":1},{"n":"Milejczyce","lat":52.53,"lng":23.13,"typ":"gmina","sieci":1},{"n":"Dziadkowice","lat":52.55,"lng":22.9,"typ":"gmina","sieci":1},{"n":"Grodzisk","lat":52.62,"lng":22.75,"typ":"gmina","sieci":1}],
  "warmińsko-mazurskie": [{"n":"Barczewo","lat":53.82,"lng":20.69,"typ":"gmina","sieci":2},{"n":"Dobre Miasto","lat":53.98,"lng":20.39,"typ":"gmina","sieci":3},{"n":"Olsztynek","lat":53.58,"lng":20.28,"typ":"gmina","sieci":3},{"n":"Biskupiec","lat":53.86,"lng":20.95,"typ":"gmina","sieci":3},{"n":"Jeziorany","lat":53.97,"lng":20.74,"typ":"gmina","sieci":2},{"n":"Pasym","lat":53.63,"lng":20.79,"typ":"gmina","sieci":2},{"n":"Stawiguda","lat":53.66,"lng":20.39,"typ":"gmina","sieci":2},{"n":"Jonkowo","lat":53.83,"lng":20.31,"typ":"gmina","sieci":2},{"n":"Purda","lat":53.71,"lng":20.70,"typ":"gmina","sieci":2},{"n":"Morąg","lat":53.91,"lng":19.92,"typ":"gmina","sieci":3},{"n":"Gietrzwałd","lat":53.73,"lng":20.23,"typ":"gmina","sieci":2},{"n":"Dywity","lat":53.83,"lng":20.47,"typ":"gmina","sieci":2},{"n":"Świątki","lat":53.92,"lng":20.24,"typ":"gmina","sieci":2},{"n":"Lubomino","lat":54.06,"lng":20.24,"typ":"gmina","sieci":1},{"n":"Miłakowo","lat":54.00,"lng":20.07,"typ":"gmina","sieci":2},{"n":"Małdyty","lat":53.92,"lng":19.74,"typ":"gmina","sieci":2},{"n":"Miłomłyn","lat":53.76,"lng":19.84,"typ":"gmina","sieci":2},{"n":"Łukta","lat":53.80,"lng":20.08,"typ":"gmina","sieci":2},{"n":"Grunwald","lat":53.48,"lng":20.12,"typ":"gmina","sieci":1},{"n":"Kozłowo","lat":53.30,"lng":20.30,"typ":"gmina","sieci":1},{"n":"Janowiec Kośc.","lat":53.28,"lng":20.52,"typ":"gmina","sieci":1},{"n":"Jedwabno","lat":53.47,"lng":20.73,"typ":"gmina","sieci":2},{"n":"Dźwierzuty","lat":53.70,"lng":20.95,"typ":"gmina","sieci":2},{"n":"Działdowo","lat":53.23,"lng":20.18,"typ":"gmina","sieci":4},{"n":"Iłowo-Osada","lat":53.16,"lng":20.29,"typ":"gmina","sieci":2},{"n":"Nidzica","lat":53.36,"lng":20.42,"typ":"gmina","sieci":3},{"n":"Rybno","lat":53.38,"lng":19.98,"typ":"gmina","sieci":2},{"n":"Lidzbark","lat":53.26,"lng":19.82,"typ":"gmina","sieci":3},{"n":"Płośnica","lat":53.28,"lng":20.04,"typ":"gmina","sieci":1},{"n":"Kurzętnik","lat":53.39,"lng":19.58,"typ":"gmina","sieci":2},{"n":"Biskupiec Pom.","lat":53.48,"lng":19.33,"typ":"gmina","sieci":2},{"n":"Iława","lat":53.59,"lng":19.56,"typ":"gmina","sieci":4},{"n":"Kisielice","lat":53.60,"lng":19.26,"typ":"gmina","sieci":2},{"n":"Susz","lat":53.72,"lng":19.34,"typ":"gmina","sieci":2},{"n":"Zalewo","lat":53.84,"lng":19.60,"typ":"gmina","sieci":2},{"n":"Lubawa","lat":53.72,"lng":19.74,"typ":"gmina","sieci":3},{"n":"Nowe Miasto L.","lat":53.42,"lng":19.59,"typ":"gmina","sieci":3},{"n":"Grodziczno","lat":53.37,"lng":19.75,"typ":"gmina","sieci":1},{"n":"Kolno","lat":53.92,"lng":20.9,"typ":"gmina","sieci":1},{"n":"Reszel","lat":54.05,"lng":21.15,"typ":"gmina","sieci":2},{"n":"Bisztynek","lat":54.09,"lng":20.9,"typ":"gmina","sieci":2},{"n":"Lidzbark W.","lat":54.13,"lng":20.58,"typ":"gmina","sieci":3},{"n":"Orneta","lat":54.12,"lng":20.13,"typ":"gmina","sieci":2},{"n":"Ostróda","lat":53.7,"lng":19.97,"typ":"gmina","sieci":3},{"n":"Iłowo","lat":53.14,"lng":20.3,"typ":"gmina","sieci":1},{"n":"Dąbrówno","lat":53.42,"lng":20.03,"typ":"gmina","sieci":1},{"n":"Janowo","lat":53.3,"lng":20.51,"typ":"gmina","sieci":1},{"n":"Pasłęk","lat":54.06,"lng":19.66,"typ":"gmina","sieci":2},{"n":"Młynary","lat":54.18,"lng":19.75,"typ":"gmina","sieci":1},{"n":"Godkowo","lat":54.11,"lng":19.86,"typ":"gmina","sieci":1},{"n":"Elbląg","lat":54.16,"lng":19.4,"typ":"gmina","sieci":4},{"n":"Tolkmicko","lat":54.32,"lng":19.53,"typ":"gmina","sieci":1},{"n":"Frombork","lat":54.36,"lng":19.68,"typ":"gmina","sieci":1},{"n":"Braniewo","lat":54.38,"lng":19.82,"typ":"gmina","sieci":3},{"n":"Pieniężno","lat":54.24,"lng":20.13,"typ":"gmina","sieci":1},{"n":"Górowo Iław.","lat":54.28,"lng":20.49,"typ":"gmina","sieci":1},{"n":"Bartoszyce","lat":54.25,"lng":20.81,"typ":"gmina","sieci":3},{"n":"Sępopol","lat":54.27,"lng":21.15,"typ":"gmina","sieci":1},{"n":"Korsze","lat":54.17,"lng":21.15,"typ":"gmina","sieci":1},{"n":"Kętrzyn","lat":54.08,"lng":21.38,"typ":"gmina","sieci":3},{"n":"Mrągowo","lat":53.87,"lng":21.3,"typ":"gmina","sieci":3}],
  "opolskie": [{"n":"Ozimek","lat":50.67,"lng":18.21,"typ":"gmina","sieci":3},{"n":"Gogolin","lat":50.48,"lng":18.02,"typ":"gmina","sieci":2},{"n":"Zdzieszowice","lat":50.42,"lng":18.12,"typ":"gmina","sieci":2},{"n":"Grodków","lat":50.69,"lng":17.38,"typ":"gmina","sieci":3},{"n":"Prószków","lat":50.57,"lng":17.87,"typ":"gmina","sieci":2},{"n":"Niemodlin","lat":50.64,"lng":17.62,"typ":"gmina","sieci":2},{"n":"Dobrzeń Wielki","lat":50.75,"lng":17.84,"typ":"gmina","sieci":2},{"n":"Komprachcice","lat":50.63,"lng":17.81,"typ":"gmina","sieci":2},{"n":"Tarnów Opolski","lat":50.57,"lng":18.08,"typ":"gmina","sieci":2},{"n":"Chrząstowice","lat":50.66,"lng":18.08,"typ":"gmina","sieci":2},{"n":"Turawa","lat":50.73,"lng":18.08,"typ":"gmina","sieci":2},{"n":"Murów","lat":50.86,"lng":17.93,"typ":"gmina","sieci":2},{"n":"Łubniany","lat":50.81,"lng":18.01,"typ":"gmina","sieci":2},{"n":"Popielów","lat":50.82,"lng":17.74,"typ":"gmina","sieci":2},{"n":"Pokój","lat":50.90,"lng":17.83,"typ":"gmina","sieci":2},{"n":"Lewin Brzeski","lat":50.75,"lng":17.61,"typ":"gmina","sieci":2},{"n":"Tułowice","lat":50.59,"lng":17.65,"typ":"gmina","sieci":2},{"n":"Zawadzkie","lat":50.61,"lng":18.47,"typ":"gmina","sieci":3},{"n":"Ujazd","lat":50.39,"lng":18.35,"typ":"gmina","sieci":2},{"n":"Krapkowice","lat":50.47,"lng":17.96,"typ":"gmina","sieci":3},{"n":"Głuchołazy","lat":50.31,"lng":17.38,"typ":"gmina","sieci":3},{"n":"Biała","lat":50.38,"lng":17.66,"typ":"gmina","sieci":2},{"n":"Strzeleczki","lat":50.46,"lng":17.86,"typ":"gmina","sieci":1},{"n":"Walce","lat":50.39,"lng":17.94,"typ":"gmina","sieci":1},{"n":"Głogówek","lat":50.35,"lng":17.86,"typ":"gmina","sieci":2},{"n":"Otmuchów","lat":50.58,"lng":17.58,"typ":"gmina","sieci":2},{"n":"Paczków","lat":50.46,"lng":17.00,"typ":"gmina","sieci":2},{"n":"Kamiennik","lat":50.56,"lng":17.48,"typ":"gmina","sieci":1},{"n":"Lubsza","lat":50.80,"lng":17.55,"typ":"gmina","sieci":1},{"n":"Skarbimierz","lat":50.90,"lng":17.46,"typ":"gmina","sieci":1},{"n":"Namysłów","lat":51.07,"lng":17.71,"typ":"gmina","sieci":4},{"n":"Wilków","lat":51.01,"lng":17.65,"typ":"gmina","sieci":1},{"n":"Domaszowice","lat":50.99,"lng":17.92,"typ":"gmina","sieci":1},{"n":"Kluczbork","lat":50.97,"lng":18.21,"typ":"gmina","sieci":4},{"n":"Wołczyn","lat":51.01,"lng":18.05,"typ":"gmina","sieci":2},{"n":"Byczyna","lat":51.11,"lng":18.21,"typ":"gmina","sieci":2},{"n":"Gorzów Śl.","lat":50.29,"lng":18.52,"typ":"gmina","sieci":2},{"n":"Praszka","lat":51.02,"lng":18.44,"typ":"gmina","sieci":2},{"n":"Olesno","lat":50.87,"lng":18.42,"typ":"gmina","sieci":3},{"n":"Dobrzeń Wlk.","lat":50.75,"lng":17.86,"typ":"gmina","sieci":2},{"n":"Kolonowskie","lat":50.63,"lng":18.35,"typ":"gmina","sieci":1},{"n":"Dobrodzień","lat":50.72,"lng":18.44,"typ":"gmina","sieci":2},{"n":"Lubrza","lat":50.32,"lng":17.68,"typ":"gmina","sieci":1},{"n":"Prudnik","lat":50.32,"lng":17.58,"typ":"gmina","sieci":4},{"n":"Nysa okolice","lat":50.47,"lng":17.33,"typ":"gmina","sieci":3},{"n":"Lasowice","lat":50.78,"lng":18.28,"typ":"gmina","sieci":1},{"n":"Rudniki","lat":51.06,"lng":18.42,"typ":"gmina","sieci":1},{"n":"Świerczów","lat":50.94,"lng":17.75,"typ":"gmina","sieci":1},{"n":"Namysłów ok.","lat":51.05,"lng":17.75,"typ":"gmina","sieci":1},{"n":"Brzeg","lat":50.86,"lng":17.47,"typ":"gmina","sieci":3},{"n":"Lewin Brz.","lat":50.75,"lng":17.61,"typ":"gmina","sieci":1},{"n":"Skoroszyce","lat":50.55,"lng":17.36,"typ":"gmina","sieci":1},{"n":"Pakosławice","lat":50.52,"lng":17.3,"typ":"gmina","sieci":1},{"n":"Nysa","lat":50.47,"lng":17.33,"typ":"gmina","sieci":3},{"n":"Korfantów","lat":50.48,"lng":17.57,"typ":"gmina","sieci":1},{"n":"Łambinowice","lat":50.51,"lng":17.55,"typ":"gmina","sieci":1}],
  "lubuskie": [{"n":"Sulechów","lat":52.08,"lng":15.62,"typ":"gmina","sieci":4},{"n":"Czerwieńsk","lat":52.01,"lng":15.42,"typ":"gmina","sieci":2},{"n":"Nowogród Bobrz.","lat":51.79,"lng":15.23,"typ":"gmina","sieci":2},{"n":"Kargowa","lat":52.03,"lng":15.86,"typ":"gmina","sieci":2},{"n":"Babimost","lat":52.16,"lng":15.82,"typ":"gmina","sieci":2},{"n":"Zbąszynek","lat":52.24,"lng":15.81,"typ":"gmina","sieci":2},{"n":"Rzepin","lat":52.34,"lng":14.83,"typ":"gmina","sieci":3},{"n":"Skwierzyna","lat":52.59,"lng":15.50,"typ":"gmina","sieci":3},{"n":"Ośno Lubuskie","lat":52.45,"lng":14.87,"typ":"gmina","sieci":2},{"n":"Cybinka","lat":52.19,"lng":14.79,"typ":"gmina","sieci":2},{"n":"Torzym","lat":52.31,"lng":15.08,"typ":"gmina","sieci":2},{"n":"Lubniewice","lat":52.51,"lng":15.24,"typ":"gmina","sieci":2},{"n":"Krzeszyce","lat":52.58,"lng":15.11,"typ":"gmina","sieci":2},{"n":"Bytnica","lat":52.14,"lng":15.16,"typ":"gmina","sieci":1},{"n":"Trzebiechów","lat":52.02,"lng":15.73,"typ":"gmina","sieci":2},{"n":"Bojadła","lat":51.95,"lng":15.81,"typ":"gmina","sieci":1},{"n":"Świdnica","lat":51.88,"lng":15.39,"typ":"gmina","sieci":2},{"n":"Zabór","lat":51.95,"lng":15.71,"typ":"gmina","sieci":1},{"n":"Sława","lat":51.87,"lng":16.08,"typ":"gmina","sieci":3},{"n":"Wschowa","lat":51.80,"lng":16.31,"typ":"gmina","sieci":4},{"n":"Szlichtyngowa","lat":51.71,"lng":16.24,"typ":"gmina","sieci":2},{"n":"Kolsko","lat":51.96,"lng":15.93,"typ":"gmina","sieci":1},{"n":"Trzciel","lat":52.37,"lng":15.88,"typ":"gmina","sieci":2},{"n":"Bledzew","lat":52.51,"lng":15.41,"typ":"gmina","sieci":1},{"n":"Deszczno","lat":52.67,"lng":15.32,"typ":"gmina","sieci":2},{"n":"Santok","lat":52.73,"lng":15.40,"typ":"gmina","sieci":2},{"n":"Bogdaniec","lat":52.66,"lng":15.18,"typ":"gmina","sieci":2},{"n":"Lubiszyn","lat":52.66,"lng":14.98,"typ":"gmina","sieci":1},{"n":"Witnica","lat":52.67,"lng":14.90,"typ":"gmina","sieci":2},{"n":"Kostrzyn n. Odrą","lat":52.58,"lng":14.66,"typ":"gmina","sieci":4},{"n":"Górzyca","lat":52.49,"lng":14.76,"typ":"gmina","sieci":2},{"n":"Słońsk","lat":51.95,"lng":16.08,"typ":"gmina","sieci":2},{"n":"Nowa Sól","lat":51.80,"lng":15.71,"typ":"gmina","sieci":5},{"n":"Otyń","lat":51.84,"lng":15.71,"typ":"gmina","sieci":2},{"n":"Kozuchów","lat":51.76,"lng":15.53,"typ":"gmina","sieci":3},{"n":"Brodce","lat":51.64,"lng":15.33,"typ":"gmina","sieci":2},{"n":"Lubsko","lat":51.79,"lng":14.97,"typ":"gmina","sieci":3},{"n":"Tuplice","lat":52.18,"lng":14.83,"typ":"gmina","sieci":1},{"n":"Żary","lat":51.64,"lng":15.13,"typ":"gmina","sieci":5},{"n":"Jasień","lat":51.75,"lng":15.01,"typ":"gmina","sieci":2},{"n":"Międzyrzecz","lat":52.44,"lng":15.58,"typ":"gmina","sieci":3},{"n":"Przytoczna","lat":52.55,"lng":15.6,"typ":"gmina","sieci":1},{"n":"Kłodawa","lat":52.78,"lng":15.3,"typ":"gmina","sieci":1},{"n":"Sulęcin","lat":52.44,"lng":15.12,"typ":"gmina","sieci":2},{"n":"Zielona Góra ok.","lat":51.94,"lng":15.51,"typ":"gmina","sieci":3},{"n":"Nowogród Bob.","lat":51.79,"lng":15.24,"typ":"gmina","sieci":1},{"n":"Żagań","lat":51.62,"lng":15.32,"typ":"gmina","sieci":3},{"n":"Iłowa","lat":51.5,"lng":15.21,"typ":"gmina","sieci":1},{"n":"Gozdnica","lat":51.44,"lng":15.1,"typ":"gmina","sieci":1},{"n":"Wymiarki","lat":51.48,"lng":15.05,"typ":"gmina","sieci":1},{"n":"Szprotawa","lat":51.56,"lng":15.54,"typ":"gmina","sieci":2},{"n":"Małomice","lat":51.57,"lng":15.44,"typ":"gmina","sieci":1},{"n":"Niegosławice","lat":51.62,"lng":15.66,"typ":"gmina","sieci":1},{"n":"Kożuchów","lat":51.75,"lng":15.59,"typ":"gmina","sieci":2},{"n":"Bytom Odrz.","lat":51.73,"lng":15.83,"typ":"gmina","sieci":1},{"n":"Nowe Miasteczko","lat":51.68,"lng":15.72,"typ":"gmina","sieci":1},{"n":"Krosno Odrz.","lat":52.05,"lng":15.09,"typ":"gmina","sieci":2},{"n":"Gubin","lat":51.95,"lng":14.72,"typ":"gmina","sieci":2},{"n":"Brody","lat":51.79,"lng":14.77,"typ":"gmina","sieci":1},{"n":"Łęknica","lat":51.55,"lng":14.74,"typ":"gmina","sieci":1},{"n":"Trzebiel","lat":51.63,"lng":14.82,"typ":"gmina","sieci":1},{"n":"Przewóz","lat":51.48,"lng":14.94,"typ":"gmina","sieci":1}]
}

@dataclass
class Miasto:
    n: str
    lat: float
    lng: float
    typ: str
    sieci: int
    woj: str = ""
    sektor: str = ""

@dataclass
class RawEtap:
    skad: str
    dokad: str
    data_str: str
    d_line: float
    czas_w_sklepie: float
    dokad_woj: str
    skad_lat: float
    skad_lng: float
    dokad_lat: float
    dokad_lng: float
    kwota: float = 0.0
    czas_jazdy_minuty: float = 0.0

@dataclass
class Etap:
    skad: str
    dokad: str
    data: str
    godz_wyj: str
    godz_przyj: str
    kwota: float
    dokad_woj: str = ""

@dataclass
class DzienTrasy:
    data: datetime.date
    etapy_surowe: List[RawEtap] = field(default_factory=list)
    etapy: List[Etap] = field(default_factory=list)

    @property
    def suma(self) -> float:
        return sum(e.kwota for e in self.etapy)

@dataclass
class DanePracownika:
    imie: str
    pesel: str
    adres: str
    stanowisko: str
    kod_pocztowy: str
    baza_miasto: str
    baza_lat: float = 52.23
    baza_lng: float = 21.01
    wojewodztwo: str = "mazowieckie"

_baza_miast: Dict[str, List[Miasto]] = {}

def oblicz_dystans(lat1, lon1, lat2, lon2) -> float:
    R = 6371
    dLat = math.radians(lat2 - lat1)
    dLon = math.radians(lon2 - lon1)
    a = (math.sin(dLat / 2) ** 2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dLon / 2) ** 2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# --- Realne odległości drogowe (OSRM) z cache dyskowym i fallbackiem offline ---
ROAD_CACHE_FILE = os.path.join(os.path.expanduser("~"), ".pmt_road_cache.json")
_road_cache = {}
_osrm_dostepny = None    # None=niesprawdzone, True/False po pierwszej próbie

def _wczytaj_road_cache():
    global _road_cache
    if os.path.exists(ROAD_CACHE_FILE):
        try:
            with open(ROAD_CACHE_FILE, "r", encoding="utf-8") as f: _road_cache = json.load(f)
        except Exception: _road_cache = {}
_wczytaj_road_cache()

def _zapisz_road_cache():
    try:
        with open(ROAD_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(_road_cache, f)
    except Exception: pass

def dystans_drogowy(lat1, lon1, lat2, lon2, tylko_cache=False) -> float:
    """Zwraca realną odległość DROGOWĄ w km (nie w linii prostej).
    Kolejność źródeł:
      1) cache dyskowy (natychmiast),
      2) OSRM — publiczny serwer routingu (gdy jest internet),
      3) fallback: linia prosta × współczynnik krętości (offline / błąd).
    Wynik z OSRM jest zapisywany w cache, więc kolejne generowania są szybkie."""
    global _osrm_dostepny
    klucz = f"{lat1:.4f},{lon1:.4f};{lat2:.4f},{lon2:.4f}"
    if klucz in _road_cache:
        return _road_cache[klucz]
    if tylko_cache or _osrm_dostepny is False:
        return oblicz_dystans(lat1, lon1, lat2, lon2) * TEST_MNOZNIK_TRASY
    # zapytanie do OSRM
    try:
        url = (f"https://router.project-osrm.org/route/v1/driving/"
               f"{lon1},{lat1};{lon2},{lat2}?overview=false")
        req = urllib.request.Request(url, headers={"User-Agent": "PMT-Delegacje/1.0"})
        with urllib.request.urlopen(req, timeout=4) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        if data.get("code") == "Ok" and data.get("routes"):
            km = data["routes"][0]["distance"] / 1000.0
            _osrm_dostepny = True
            _road_cache[klucz] = round(km, 2)
            return _road_cache[klucz]
    except Exception:
        _osrm_dostepny = False
    # fallback
    return oblicz_dystans(lat1, lon1, lat2, lon2) * TEST_MNOZNIK_TRASY

def dystans_odcinek_punkt(lat1, lon1, lat2, lon2, p_lat, p_lon):
    kx = 111.0 * math.cos(math.radians((lat1+lat2)/2.0))
    ky = 111.0
    x1, y1 = lon1 * kx, lat1 * ky
    x2, y2 = lon2 * kx, lat2 * ky
    px, py = p_lon * kx, p_lat * ky

    l2 = (x2 - x1)**2 + (y2 - y1)**2
    if l2 == 0: return math.sqrt((px - x1)**2 + (py - y1)**2)
    t = max(0, min(1, ((px - x1) * (x2 - x1) + (py - y1) * (y2 - y1)) / l2))
    return math.sqrt((px - (x1 + t * (x2 - x1)))**2 + (py - (y1 + t * (y2 - y1)))**2)

def wyznacz_sektor(lat_cel, lng_cel, lat_baza, lng_baza) -> str:
    kat = math.degrees(math.atan2(lng_cel - lng_baza, lat_cel - lat_baza)) % 360
    if kat < 22.5 or kat >= 337.5: return "N"
    elif kat < 67.5: return "NE"
    elif kat < 112.5: return "E"
    elif kat < 157.5: return "SE"
    elif kat < 202.5: return "S"
    elif kat < 247.5: return "SW"
    elif kat < 292.5: return "W"
    else: return "NW"

def odsuwa_sie_od_stolicy(baza_lat, baza_lng, m_lat, m_lng, woj):
    if woj not in STOLICE: return True
    w_lat, w_lng = STOLICE[woj]
    d_baza = oblicz_dystans(baza_lat, baza_lng, w_lat, w_lng)
    if d_baza < 20.0: return True
    d_cel = oblicz_dystans(m_lat, m_lng, w_lat, w_lng)
    if d_cel <= d_baza: return False  
    return True

def pobierz_zablokowane_sektory(baza_lat, baza_lng, woj) -> List[str]:
    if woj not in STOLICE: return []
    w_lat, w_lng = STOLICE[woj]
    d = oblicz_dystans(baza_lat, baza_lng, w_lat, w_lng)
    if d < 12.0: return []
    sek = wyznacz_sektor(w_lat, w_lng, baza_lat, baza_lng)
    idx = SEKTORY_KOLEJNOSC.index(sek)
    if d <= 90.0: 
        return [SEKTORY_KOLEJNOSC[(idx - 2) % 8], SEKTORY_KOLEJNOSC[(idx - 1) % 8], sek, SEKTORY_KOLEJNOSC[(idx + 1) % 8], SEKTORY_KOLEJNOSC[(idx + 2) % 8]]
    else: 
        return [SEKTORY_KOLEJNOSC[(idx - 1) % 8], sek, SEKTORY_KOLEJNOSC[(idx + 1) % 8]]

_baza_miast_coords = None   # (lat, lng) bazy dla której policzono sektory w cache

def zaladuj_baze(baza_lat, baza_lng) -> Dict[str, List[Miasto]]:
    """Ładuje miasta i wylicza ich sektory (Płn/Płd/...) WZGLĘDEM PODANEJ BAZY.
    Cache jest ważny tylko dla tej samej bazy — gdy kolejny pracownik ma inny
    adres domowy, sektory są przeliczane od nowa (inaczej Marek z Krakowa
    dostałby kierunki liczone względem Warszawy Jana)."""
    global _baza_miast, _baza_miast_coords
    biezaca = (round(baza_lat, 4), round(baza_lng, 4))
    if _baza_miast and _baza_miast_coords == biezaca:
        return _baza_miast

    # Baza się zmieniła (lub pierwszy raz) — przeliczamy sektory od zera
    _baza_miast = {}
    _baza_miast_coords = biezaca
    for woj, lista in MIASTA_RAW.items():
        _baza_miast[woj] = []
        for d in lista:
            if len(d['n']) > 16: continue 
            sektor = wyznacz_sektor(d['lat'], d['lng'], baza_lat, baza_lng)
            _baza_miast[woj].append(Miasto(n=d['n'], lat=d['lat'], lng=d['lng'], typ=d.get('typ', 'gmina'), sieci=d.get('sieci', 2), woj=woj, sektor=sektor))
    
    for woj in WSZYSTKIE_WOJEWODZTWA:
        if woj not in _baza_miast: _baza_miast[woj] = []
        obecne = {m.n for m in _baza_miast[woj]}
        if len(_baza_miast[woj]) < 15:
            for s in SASIEDZI_WOJ.get(woj, []):
                if s == woj: continue
                for m in _baza_miast.get(s, []):
                    if m.n not in obecne:
                        sekt = wyznacz_sektor(m.lat, m.lng, baza_lat, baza_lng)
                        _baza_miast[woj].append(Miasto(n=m.n, lat=m.lat, lng=m.lng, typ=m.typ, sieci=m.sieci, woj=s, sektor=sekt))
                        obecne.add(m.n)
    return _baza_miast

def _wielkanoc(rok: int) -> datetime.date:
    """Data Wielkanocy (algorytm gaussowski) — potrzebna też do ustalenia
    niedzieli handlowej poprzedzającej święta."""
    a = rok % 19; b = rok // 100; c = rok % 100; d = b // 4; e = b % 4; f = (b + 8) // 25
    g = (b - f + 1) // 3; h = (19 * a + b - d - g + 15) % 30; i = c // 4; k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7; m = (a + 11 * h + 22 * l) // 451
    miesiac = (h + l - 7 * m + 114) // 31; dzien = ((h + l - 7 * m + 114) % 31) + 1
    return datetime.date(rok, miesiac, dzien)


def swieta_w_roku(rok: int) -> Set[datetime.date]:
    a = rok % 19; b = rok // 100; c = rok % 100; d = b // 4; e = b % 4; f = (b + 8) // 25
    g = (b - f + 1) // 3; h = (19 * a + b - d - g + 15) % 30; i = c // 4; k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7; m = (a + 11 * h + 22 * l) // 451
    miesiac = (h + l - 7 * m + 114) // 31; dzien = ((h + l - 7 * m + 114) % 31) + 1
    wielkanoc = datetime.date(rok, miesiac, dzien)
    stale = [
        datetime.date(rok, 1, 1), datetime.date(rok, 1, 6),
        datetime.date(rok, 5, 1), datetime.date(rok, 5, 3),
        datetime.date(rok, 8, 15), datetime.date(rok, 11, 1),
        datetime.date(rok, 11, 11), datetime.date(rok, 12, 25), datetime.date(rok, 12, 26),
        wielkanoc + datetime.timedelta(days=1), wielkanoc + datetime.timedelta(days=60)
    ]
    return set(stale)

def pobierz_dni_robocze(rok: int, miesiac: int) -> List[datetime.date]:
    """Dni, w które można rozpisać pracę.

    Tryb tygodniowy: poniedziałek-piątek, bez świąt (jak dotychczas).
    Tryb wieczorny: dodatkowo SOBOTY. Niedziele wchodzą wyłącznie wtedy, gdy
    włączono niedziele handlowe — i tylko te, które nimi są.
    """
    swieta = swieta_w_roku(rok)
    _, n = calendar.monthrange(rok, miesiac)
    dni = []
    for d in range(1, n + 1):
        data = datetime.date(rok, miesiac, d)
        if data in swieta:
            continue
        dow = data.weekday()                     # 0=pon ... 5=sob, 6=nie
        if dow < 5:
            dni.append(data)
        elif dow == 5 and TRYB_PRACY == "wieczory":
            dni.append(data)                     # sobota jak zwykły dzień
        elif dow == 6 and TRYB_PRACY == "wieczory" and NIEDZIELE_HANDLOWE \
                and czy_niedziela_handlowa(data):
            dni.append(data)
    return dni


def czy_niedziela_handlowa(data: datetime.date) -> bool:
    """Niedziele handlowe: ostatnia niedziela stycznia, kwietnia, czerwca
    i sierpnia, dwie niedziele przed Bożym Narodzeniem oraz niedziela przed
    Wielkanocą. Zasada z ustawy o ograniczeniu handlu w niedziele —
    gdyby przepisy się zmieniły, wystarczy poprawić tę jedną funkcję."""
    if data.weekday() != 6:
        return False
    _, dni_mies = calendar.monthrange(data.year, data.month)
    ostatnia_niedziela = max(d for d in range(1, dni_mies + 1)
                             if datetime.date(data.year, data.month, d).weekday() == 6)
    if data.month in (1, 4, 6, 8) and data.day == ostatnia_niedziela:
        return True
    if data.month == 12:                          # dwie niedziele przed 24.12
        przed = [d for d in range(1, 25)
                 if datetime.date(data.year, 12, d).weekday() == 6]
        if data.day in przed[-2:]:
            return True
    try:
        wielkanoc = _wielkanoc(data.year)          # niedziela przed Wielkanocą
        if data == wielkanoc - datetime.timedelta(days=7):
            return True
    except Exception:
        pass
    return False


# ==========================================================================
# ||                                                                      ||
# ||   NOWY SILNIK: PLANER WIZYT (PLAN-1)                                  ||
# ||   Niezależny od starego generatora delegacji (kwota→trasy, S11).      ||
# ||   Wejście: lista punktów z adresami. Wyjście: logiczny plan wizyt     ||
# ||   rozłożony na dni robocze miesiąca (pon-pt, bez świąt), z trasami    ||
# ||   grupowanymi geograficznie i limitem czasu dnia.                     ||
# ||                                                                      ||
# ==========================================================================

# Parametry planowania (realistyczny dzień przedstawiciela)
PLAN_LIMIT_MINUT_DNIA = 8 * 60      # 8h pracy dziennie (dojazdy + wizyty)
PLAN_CZAS_WIZYTY_MIN = 20           # średni czas jednej wizyty w punkcie
PLAN_SREDNIA_PREDKOSC_KMH = 50      # średnia prędkość w trasie (miasto+trasa)


class PunktWizyty:
    """Pojedynczy punkt do odwiedzenia (sklep/klient)."""
    __slots__ = ("nazwa", "adres", "siec", "lat", "lng", "miasto")
    def __init__(self, nazwa, adres, siec="", lat=None, lng=None, miasto=""):
        self.nazwa = nazwa          # etykieta wyświetlana
        self.adres = adres          # pełny adres (do geokodowania)
        self.siec = siec            # sieć/marka (opcjonalnie)
        self.lat = lat; self.lng = lng
        self.miasto = miasto

    def __repr__(self):
        return f"PunktWizyty({self.nazwa!r})"


class DzienPlanu:
    """Jeden dzień roboczy planu z listą wizyt w kolejności przejazdu."""
    __slots__ = ("data", "wizyty", "km", "minuty")
    def __init__(self, data):
        self.data = data            # datetime.date
        self.wizyty = []            # lista PunktWizyty w kolejności
        self.km = 0.0               # łączny dystans dnia
        self.minuty = 0.0           # łączny czas dnia (dojazdy + wizyty)


def _czas_przejazdu_min(km: float, predkosc: float = None) -> float:
    """Ile minut zajmie przejazd danego dystansu (średnia prędkość)."""
    v = predkosc or PLAN_SREDNIA_PREDKOSC_KMH
    if not v or v <= 0:
        v = PLAN_SREDNIA_PREDKOSC_KMH
    return (km / v) * 60.0


def ustawienia_planowania() -> dict:
    """Parametry planowania wg ustawień użytkownika (z sensownymi domyślnymi).
    Wcześniej były to sztywne stałe — a od nich zależy CAŁY plan: ile wizyt
    zmieści się w dniu, ile godzin zajmie, ile km."""
    def _num(klucz, dom, mini, maks):
        try:
            v = float(ustawienie(klucz, dom))
        except (TypeError, ValueError):
            v = float(dom)
        return max(mini, min(maks, v))
    return {
        "limit_minut": _num("plan_godziny", 8.0, 1.0, 16.0) * 60.0,
        "czas_wizyty": _num("plan_minuty_wizyta", PLAN_CZAS_WIZYTY_MIN, 1.0, 240.0),
        "predkosc": _num("plan_predkosc", PLAN_SREDNIA_PREDKOSC_KMH, 10.0, 130.0),
        "powrot": bool(ustawienie("plan_powrot_do_bazy", False)),
    }


def _dyst(a: PunktWizyty, b: PunktWizyty, offline=False) -> float:
    """Dystans drogowy między dwoma punktami (km)."""
    if a.lat is None or b.lat is None:
        return 0.0
    return dystans_drogowy(a.lat, a.lng, b.lat, b.lng, tylko_cache=offline)


def _dyst_prosty(a: PunktWizyty, b: PunktWizyty) -> float:
    """Szybki dystans liniowy (haversine) × współczynnik krętości dróg.
    Używany w planowaniu — NIE odpytuje sieci, więc działa błyskawicznie
    nawet dla tysięcy punktów."""
    if a.lat is None or b.lat is None:
        return 0.0
    return oblicz_dystans(a.lat, a.lng, b.lat, b.lng) * TEST_MNOZNIK_TRASY


# --- ROZSZERZENIE BAZY MIEJSCOWOSCI (2026-08) --------------------------------
# Doliczone gminy i miasta powiatowe dla szesciu wojewodztw: mazowieckie,
# lodzkie, kujawsko-pomorskie, lubelskie, podlaskie, warminsko-mazurskie.
# Wspolrzedne sprawdzone wzgledem granic wojewodztw (dokladnosc ~1 km — do
# planowania tras wystarcza, bo realne odleglosci liczy OSRM po drogach).
# Wpisy o nazwach juz obecnych w MIASTA_RAW zostaly pominiete przy tworzeniu.
MIASTA_ROZSZERZENIE = {
  "mazowieckie": [{"n":"Ciechanów","lat":52.88,"lng":20.62,"typ":"gmina","sieci":3},{"n":"Sochaczew","lat":52.23,"lng":20.24,"typ":"gmina","sieci":3},{"n":"Żyrardów","lat":52.05,"lng":20.45,"typ":"gmina","sieci":3},{"n":"Grodzisk Mazowiecki","lat":52.11,"lng":20.63,"typ":"gmina","sieci":4},{"n":"Pruszków","lat":52.17,"lng":20.81,"typ":"gmina","sieci":4},{"n":"Mińsk Mazowiecki","lat":52.18,"lng":21.57,"typ":"gmina","sieci":3},{"n":"Legionowo","lat":52.40,"lng":20.93,"typ":"gmina","sieci":4},{"n":"Nowy Dwór Mazowiecki","lat":52.43,"lng":20.72,"typ":"gmina","sieci":3},{"n":"Ostrów Mazowiecka","lat":52.80,"lng":21.90,"typ":"gmina","sieci":3},{"n":"Ostrołęka","lat":53.09,"lng":21.57,"typ":"gmina","sieci":3},{"n":"Maków Mazowiecki","lat":52.86,"lng":21.10,"typ":"gmina","sieci":2},{"n":"Płock","lat":52.55,"lng":19.71,"typ":"gmina","sieci":4},{"n":"Radom","lat":51.40,"lng":21.15,"typ":"gmina","sieci":4},{"n":"Konstancin-Jeziorna","lat":52.09,"lng":21.11,"typ":"gmina","sieci":3},{"n":"Garwolin","lat":51.90,"lng":21.61,"typ":"gmina","sieci":2},{"n":"Łaskarzew","lat":51.79,"lng":21.60,"typ":"gmina","sieci":2},{"n":"Pilawa","lat":51.87,"lng":21.55,"typ":"gmina","sieci":2},{"n":"Siedlce","lat":52.17,"lng":22.29,"typ":"gmina","sieci":4},{"n":"Sokołów Podlaski","lat":52.41,"lng":22.25,"typ":"gmina","sieci":3},{"n":"Węgrów","lat":52.40,"lng":22.02,"typ":"gmina","sieci":2},{"n":"Łochów","lat":52.53,"lng":21.68,"typ":"gmina","sieci":2},{"n":"Kosów Lacki","lat":52.60,"lng":22.15,"typ":"gmina","sieci":2},{"n":"Mordy","lat":52.21,"lng":22.52,"typ":"gmina","sieci":2},{"n":"Łosice","lat":52.21,"lng":22.72,"typ":"gmina","sieci":2},{"n":"Żelechów","lat":51.81,"lng":21.90,"typ":"gmina","sieci":2},{"n":"Raszyn","lat":52.14,"lng":20.93,"typ":"gmina","sieci":3},{"n":"Brwinów","lat":52.14,"lng":20.72,"typ":"gmina","sieci":3},{"n":"Milanówek","lat":52.12,"lng":20.67,"typ":"gmina","sieci":3},{"n":"Podkowa Leśna","lat":52.12,"lng":20.69,"typ":"gmina","sieci":2},{"n":"Ożarów Mazowiecki","lat":52.21,"lng":20.81,"typ":"gmina","sieci":3},{"n":"Łomianki","lat":52.34,"lng":20.88,"typ":"gmina","sieci":3},{"n":"Józefów (otwocki)","lat":52.14,"lng":21.23,"typ":"gmina","sieci":3},{"n":"Stoczek Łukowski","lat":51.96,"lng":21.97,"typ":"gmina","sieci":2},{"n":"Wierzbica (radomska)","lat":51.22,"lng":21.09,"typ":"gmina","sieci":2},{"n":"Jedlnia-Letnisko","lat":51.42,"lng":21.30,"typ":"gmina","sieci":2},{"n":"Zakrzew (radomski)","lat":51.42,"lng":21.00,"typ":"gmina","sieci":2},{"n":"Kozłów","lat":51.58,"lng":21.28,"typ":"gmina","sieci":2},{"n":"Sienno","lat":51.24,"lng":21.70,"typ":"gmina","sieci":2},{"n":"Solec nad Wisłą","lat":51.14,"lng":21.77,"typ":"gmina","sieci":2},{"n":"Chotcza","lat":51.28,"lng":21.75,"typ":"gmina","sieci":2},{"n":"Sarnaki","lat":52.32,"lng":22.90,"typ":"gmina","sieci":2},{"n":"Platerów","lat":52.25,"lng":22.83,"typ":"gmina","sieci":2},{"n":"Huszlew","lat":52.15,"lng":22.80,"typ":"gmina","sieci":2},{"n":"Bielany","lat":52.10,"lng":22.61,"typ":"gmina","sieci":2},{"n":"Olszanka","lat":52.19,"lng":22.70,"typ":"gmina","sieci":2},{"n":"Suchożebry","lat":52.24,"lng":22.31,"typ":"gmina","sieci":2},{"n":"Zbuczyn","lat":52.05,"lng":22.45,"typ":"gmina","sieci":2},{"n":"Domanice","lat":52.02,"lng":22.10,"typ":"gmina","sieci":2}],
  "łódzkie": [{"n":"Piotrków Trybunalski","lat":51.40,"lng":19.70,"typ":"gmina","sieci":4},{"n":"Radomsko","lat":51.07,"lng":19.44,"typ":"gmina","sieci":3},{"n":"Bełchatów","lat":51.37,"lng":19.36,"typ":"gmina","sieci":3},{"n":"Tomaszów Mazowiecki","lat":51.53,"lng":20.01,"typ":"gmina","sieci":3},{"n":"Opoczno","lat":51.38,"lng":20.28,"typ":"gmina","sieci":2},{"n":"Rawa Mazowiecka","lat":51.77,"lng":20.25,"typ":"gmina","sieci":2},{"n":"Skierniewice","lat":51.96,"lng":20.15,"typ":"gmina","sieci":3},{"n":"Łowicz","lat":52.11,"lng":19.95,"typ":"gmina","sieci":3},{"n":"Kutno","lat":52.23,"lng":19.36,"typ":"gmina","sieci":3},{"n":"Aleksandrów Łódzki","lat":51.82,"lng":19.30,"typ":"gmina","sieci":3},{"n":"Konstantynów Łódzki","lat":51.75,"lng":19.32,"typ":"gmina","sieci":3},{"n":"Sieradz","lat":51.60,"lng":18.73,"typ":"gmina","sieci":3},{"n":"Sulejów","lat":51.36,"lng":19.88,"typ":"gmina","sieci":2},{"n":"Wolbórz","lat":51.51,"lng":19.83,"typ":"gmina","sieci":2},{"n":"Moszczenica","lat":51.44,"lng":19.71,"typ":"gmina","sieci":2},{"n":"Gorzkowice","lat":51.19,"lng":19.60,"typ":"gmina","sieci":2},{"n":"Kamieńsk","lat":51.20,"lng":19.42,"typ":"gmina","sieci":2},{"n":"Przedbórz","lat":51.09,"lng":19.87,"typ":"gmina","sieci":2},{"n":"Żarnów","lat":51.25,"lng":20.17,"typ":"gmina","sieci":2},{"n":"Drzewica","lat":51.45,"lng":20.45,"typ":"gmina","sieci":2},{"n":"Białaczów","lat":51.30,"lng":20.27,"typ":"gmina","sieci":2},{"n":"Sławno","lat":51.44,"lng":20.13,"typ":"gmina","sieci":2},{"n":"Inowłódz","lat":51.53,"lng":20.24,"typ":"gmina","sieci":2},{"n":"Lubochnia","lat":51.56,"lng":20.10,"typ":"gmina","sieci":2},{"n":"Czerniewice","lat":51.68,"lng":20.14,"typ":"gmina","sieci":2},{"n":"Żelechlinek","lat":51.75,"lng":20.05,"typ":"gmina","sieci":2},{"n":"Mszczonów-Puszcza","lat":51.86,"lng":20.20,"typ":"gmina","sieci":2},{"n":"Nowy Kawęczyn","lat":51.87,"lng":20.20,"typ":"gmina","sieci":2},{"n":"Bolimów","lat":52.06,"lng":20.20,"typ":"gmina","sieci":2},{"n":"Kiernozia","lat":52.25,"lng":19.90,"typ":"gmina","sieci":2},{"n":"Bielawy","lat":52.05,"lng":19.75,"typ":"gmina","sieci":2},{"n":"Domaniewice","lat":51.98,"lng":19.86,"typ":"gmina","sieci":2},{"n":"Zduny","lat":52.10,"lng":19.92,"typ":"gmina","sieci":2},{"n":"Chąśno","lat":52.13,"lng":19.86,"typ":"gmina","sieci":2},{"n":"Krośniewice","lat":52.25,"lng":19.17,"typ":"gmina","sieci":2},{"n":"Żychlin","lat":52.24,"lng":19.62,"typ":"gmina","sieci":2},{"n":"Bedlno","lat":52.20,"lng":19.55,"typ":"gmina","sieci":2},{"n":"Krzyżanów","lat":52.20,"lng":19.35,"typ":"gmina","sieci":2},{"n":"Strzelce","lat":52.16,"lng":19.42,"typ":"gmina","sieci":2},{"n":"Oporów","lat":52.15,"lng":19.55,"typ":"gmina","sieci":2},{"n":"Świnice Warckie","lat":51.94,"lng":18.92,"typ":"gmina","sieci":2},{"n":"Uniejów","lat":51.97,"lng":18.79,"typ":"gmina","sieci":2},{"n":"Dąbie","lat":52.10,"lng":18.83,"typ":"gmina","sieci":2},{"n":"Wartkowice","lat":51.94,"lng":19.06,"typ":"gmina","sieci":2},{"n":"Zadzim","lat":51.75,"lng":18.86,"typ":"gmina","sieci":2}],
  "kujawsko-pomorskie": [{"n":"Bydgoszcz","lat":53.12,"lng":18.01,"typ":"gmina","sieci":4},{"n":"Toruń","lat":53.01,"lng":18.60,"typ":"gmina","sieci":4},{"n":"Włocławek","lat":52.65,"lng":19.07,"typ":"gmina","sieci":4},{"n":"Inowrocław","lat":52.79,"lng":18.26,"typ":"gmina","sieci":3},{"n":"Brodnica","lat":53.25,"lng":19.40,"typ":"gmina","sieci":3},{"n":"Chełmno","lat":53.35,"lng":18.42,"typ":"gmina","sieci":2},{"n":"Nakło nad Notecią","lat":53.14,"lng":17.60,"typ":"gmina","sieci":2},{"n":"Sępólno Krajeńskie","lat":53.45,"lng":17.53,"typ":"gmina","sieci":2},{"n":"Janowiec Wielkopolski","lat":52.75,"lng":17.49,"typ":"gmina","sieci":2},{"n":"Piotrków Kujawski","lat":52.53,"lng":18.54,"typ":"gmina","sieci":2},{"n":"Aleksandrów Kujawski","lat":52.87,"lng":18.70,"typ":"gmina","sieci":2},{"n":"Nieszawa","lat":52.85,"lng":18.75,"typ":"gmina","sieci":2},{"n":"Kowalewo Pomorskie","lat":53.17,"lng":18.87,"typ":"gmina","sieci":2},{"n":"Golub-Dobrzyń","lat":53.11,"lng":19.05,"typ":"gmina","sieci":2},{"n":"Wąbrzeźno","lat":53.28,"lng":18.94,"typ":"gmina","sieci":2},{"n":"Rypin","lat":53.07,"lng":19.41,"typ":"gmina","sieci":2},{"n":"Lipno","lat":52.85,"lng":19.18,"typ":"gmina","sieci":2},{"n":"Dobrzyń nad Wisłą","lat":52.64,"lng":19.34,"typ":"gmina","sieci":2},{"n":"Kikoł","lat":52.79,"lng":19.20,"typ":"gmina","sieci":2},{"n":"Chrostkowo","lat":52.94,"lng":19.31,"typ":"gmina","sieci":2},{"n":"Brześć Kujawski","lat":52.61,"lng":18.90,"typ":"gmina","sieci":2},{"n":"Kowal","lat":52.53,"lng":19.15,"typ":"gmina","sieci":2},{"n":"Lubraniec","lat":52.53,"lng":18.85,"typ":"gmina","sieci":2},{"n":"Izbica Kujawska","lat":52.42,"lng":18.76,"typ":"gmina","sieci":2},{"n":"Chodecz","lat":52.40,"lng":19.03,"typ":"gmina","sieci":2},{"n":"Lubień Kujawski","lat":52.39,"lng":19.16,"typ":"gmina","sieci":2},{"n":"Baruchowo","lat":52.53,"lng":19.28,"typ":"gmina","sieci":2},{"n":"Fabianki","lat":52.72,"lng":19.10,"typ":"gmina","sieci":2},{"n":"Bobrowniki","lat":52.75,"lng":19.05,"typ":"gmina","sieci":2},{"n":"Jabłonowo Pomorskie","lat":53.39,"lng":19.15,"typ":"gmina","sieci":2},{"n":"Bartniczka","lat":53.28,"lng":19.53,"typ":"gmina","sieci":2},{"n":"Zbiczno","lat":53.31,"lng":19.35,"typ":"gmina","sieci":2},{"n":"Bobrowo","lat":53.31,"lng":19.28,"typ":"gmina","sieci":2},{"n":"Dragacz","lat":53.47,"lng":18.74,"typ":"gmina","sieci":2},{"n":"Radzyń Chełmiński","lat":53.38,"lng":18.93,"typ":"gmina","sieci":2},{"n":"Lisewo","lat":53.28,"lng":18.52,"typ":"gmina","sieci":2},{"n":"Papowo Biskupie","lat":53.24,"lng":18.60,"typ":"gmina","sieci":2},{"n":"Dąbrowa Chełmińska","lat":53.19,"lng":18.28,"typ":"gmina","sieci":2},{"n":"Zławieś Wielka","lat":53.10,"lng":18.35,"typ":"gmina","sieci":2}],
  "lubelskie": [{"n":"Lublin","lat":51.25,"lng":22.57,"typ":"gmina","sieci":4},{"n":"Zamość","lat":50.72,"lng":23.25,"typ":"gmina","sieci":3},{"n":"Chełm","lat":51.14,"lng":23.47,"typ":"gmina","sieci":3},{"n":"Biała Podlaska","lat":52.03,"lng":23.12,"typ":"gmina","sieci":3},{"n":"Łuków","lat":51.93,"lng":22.38,"typ":"gmina","sieci":2},{"n":"Radzyń Podlaski","lat":51.78,"lng":22.62,"typ":"gmina","sieci":2},{"n":"Parczew","lat":51.64,"lng":22.90,"typ":"gmina","sieci":2},{"n":"Włodawa","lat":51.55,"lng":23.55,"typ":"gmina","sieci":2},{"n":"Hrubieszów","lat":50.81,"lng":23.89,"typ":"gmina","sieci":2},{"n":"Tomaszów Lubelski","lat":50.45,"lng":23.42,"typ":"gmina","sieci":2},{"n":"Biłgoraj","lat":50.54,"lng":22.72,"typ":"gmina","sieci":2},{"n":"Janów Lubelski","lat":50.71,"lng":22.41,"typ":"gmina","sieci":2},{"n":"Opole Lubelskie","lat":51.15,"lng":21.97,"typ":"gmina","sieci":2},{"n":"Trawniki","lat":51.13,"lng":22.98,"typ":"gmina","sieci":2},{"n":"Ostrów Lubelski","lat":51.49,"lng":22.85,"typ":"gmina","sieci":2},{"n":"Kazimierz Dolny","lat":51.32,"lng":21.95,"typ":"gmina","sieci":2},{"n":"Żyrzyn","lat":51.49,"lng":22.09,"typ":"gmina","sieci":2},{"n":"Zakrzew (lubelski)","lat":51.04,"lng":22.35,"typ":"gmina","sieci":2},{"n":"Nowodwór","lat":51.72,"lng":22.03,"typ":"gmina","sieci":2},{"n":"Ułęż","lat":51.63,"lng":22.05,"typ":"gmina","sieci":2},{"n":"Kłoczew","lat":51.72,"lng":21.88,"typ":"gmina","sieci":2},{"n":"Adamów","lat":51.75,"lng":22.25,"typ":"gmina","sieci":2},{"n":"Serokomla","lat":51.72,"lng":22.20,"typ":"gmina","sieci":2},{"n":"Wojcieszków","lat":51.83,"lng":22.28,"typ":"gmina","sieci":2},{"n":"Stanin","lat":51.85,"lng":22.32,"typ":"gmina","sieci":2},{"n":"Trzebieszów","lat":51.94,"lng":22.55,"typ":"gmina","sieci":2},{"n":"Ulan-Majorat","lat":51.85,"lng":22.55,"typ":"gmina","sieci":2},{"n":"Kąkolewnica","lat":51.87,"lng":22.65,"typ":"gmina","sieci":2},{"n":"Międzyrzec Podlaski","lat":51.98,"lng":22.79,"typ":"gmina","sieci":2},{"n":"Wisznice","lat":51.79,"lng":23.20,"typ":"gmina","sieci":2},{"n":"Rossosz","lat":51.86,"lng":23.08,"typ":"gmina","sieci":2},{"n":"Łomazy","lat":51.90,"lng":23.19,"typ":"gmina","sieci":2},{"n":"Piszczac","lat":52.00,"lng":23.35,"typ":"gmina","sieci":2},{"n":"Kodeń","lat":51.91,"lng":23.61,"typ":"gmina","sieci":2},{"n":"Janów Podlaski","lat":52.19,"lng":23.21,"typ":"gmina","sieci":2},{"n":"Leśna Podlaska","lat":52.13,"lng":23.02,"typ":"gmina","sieci":2},{"n":"Rokitno","lat":52.10,"lng":23.24,"typ":"gmina","sieci":2},{"n":"Zalesie","lat":52.10,"lng":23.42,"typ":"gmina","sieci":2},{"n":"Sławatycze","lat":51.79,"lng":23.55,"typ":"gmina","sieci":2},{"n":"Hanna","lat":51.72,"lng":23.53,"typ":"gmina","sieci":2},{"n":"Dubienka","lat":51.05,"lng":23.87,"typ":"gmina","sieci":2},{"n":"Dorohusk","lat":51.15,"lng":23.79,"typ":"gmina","sieci":2},{"n":"Rejowiec","lat":51.02,"lng":23.28,"typ":"gmina","sieci":2},{"n":"Siedliszcze","lat":51.16,"lng":23.20,"typ":"gmina","sieci":2},{"n":"Sawin","lat":51.28,"lng":23.42,"typ":"gmina","sieci":2},{"n":"Ruda-Huta","lat":51.20,"lng":23.55,"typ":"gmina","sieci":2},{"n":"Wierzbica (chełmska)","lat":51.05,"lng":23.32,"typ":"gmina","sieci":2},{"n":"Żółkiewka","lat":50.86,"lng":22.86,"typ":"gmina","sieci":2},{"n":"Gorzków","lat":50.90,"lng":22.98,"typ":"gmina","sieci":2},{"n":"Grabowiec","lat":50.83,"lng":23.55,"typ":"gmina","sieci":2},{"n":"Miączyn","lat":50.75,"lng":23.50,"typ":"gmina","sieci":2},{"n":"Józefów (biłgorajski)","lat":50.48,"lng":23.05,"typ":"gmina","sieci":2},{"n":"Susiec","lat":50.42,"lng":23.20,"typ":"gmina","sieci":2},{"n":"Komarów-Osada","lat":50.61,"lng":23.48,"typ":"gmina","sieci":2},{"n":"Tarnogród","lat":50.36,"lng":22.74,"typ":"gmina","sieci":2},{"n":"Józefów Roztoczański","lat":50.48,"lng":23.05,"typ":"gmina","sieci":2},{"n":"Terespol","lat":52.07,"lng":23.61,"typ":"gmina","sieci":2}],
  "podlaskie": [{"n":"Białystok","lat":53.13,"lng":23.16,"typ":"gmina","sieci":4},{"n":"Suwałki","lat":54.10,"lng":22.93,"typ":"gmina","sieci":3},{"n":"Łomża","lat":53.18,"lng":22.06,"typ":"gmina","sieci":3},{"n":"Bielsk Podlaski","lat":52.77,"lng":23.19,"typ":"gmina","sieci":2},{"n":"Zambrów","lat":52.98,"lng":22.24,"typ":"gmina","sieci":2},{"n":"Wysokie Mazowieckie","lat":52.92,"lng":22.51,"typ":"gmina","sieci":2},{"n":"Sejny","lat":54.11,"lng":23.35,"typ":"gmina","sieci":2},{"n":"Dąbrowa Białostocka","lat":53.65,"lng":23.35,"typ":"gmina","sieci":2},{"n":"Jasionówka","lat":53.46,"lng":22.99,"typ":"gmina","sieci":2},{"n":"Stawiski","lat":53.38,"lng":22.14,"typ":"gmina","sieci":2},{"n":"Jedwabne","lat":53.29,"lng":22.30,"typ":"gmina","sieci":2},{"n":"Nowogród","lat":53.22,"lng":21.88,"typ":"gmina","sieci":2},{"n":"Piątnica","lat":53.19,"lng":22.11,"typ":"gmina","sieci":2},{"n":"Wizna","lat":53.19,"lng":22.38,"typ":"gmina","sieci":2},{"n":"Śniadowo","lat":53.03,"lng":22.02,"typ":"gmina","sieci":2},{"n":"Miastkowo","lat":53.09,"lng":21.87,"typ":"gmina","sieci":2},{"n":"Czerwone","lat":53.32,"lng":21.98,"typ":"gmina","sieci":2},{"n":"Turośl","lat":53.42,"lng":21.75,"typ":"gmina","sieci":2},{"n":"Zbójna","lat":53.30,"lng":21.75,"typ":"gmina","sieci":2},{"n":"Łomża-Wschód","lat":53.16,"lng":22.15,"typ":"gmina","sieci":2},{"n":"Rutki","lat":53.10,"lng":22.51,"typ":"gmina","sieci":2},{"n":"Kołaki","lat":53.02,"lng":22.38,"typ":"gmina","sieci":2},{"n":"Szumowo","lat":52.90,"lng":22.09,"typ":"gmina","sieci":2},{"n":"Andrzejewo","lat":52.85,"lng":22.22,"typ":"gmina","sieci":2},{"n":"Klukowo","lat":52.75,"lng":22.42,"typ":"gmina","sieci":2},{"n":"Nowe Piekuty","lat":52.83,"lng":22.62,"typ":"gmina","sieci":2},{"n":"Kulesze Kościelne","lat":52.98,"lng":22.55,"typ":"gmina","sieci":2},{"n":"Rudka","lat":52.63,"lng":22.75,"typ":"gmina","sieci":2},{"n":"Czeremcha","lat":52.51,"lng":23.34,"typ":"gmina","sieci":2},{"n":"Białowieża","lat":52.70,"lng":23.86,"typ":"gmina","sieci":2},{"n":"Dubicze Cerkiewne","lat":52.64,"lng":23.51,"typ":"gmina","sieci":2},{"n":"Nurzec-Stacja","lat":52.51,"lng":23.05,"typ":"gmina","sieci":2},{"n":"Perlejewo","lat":52.62,"lng":22.63,"typ":"gmina","sieci":2},{"n":"Filipów","lat":54.18,"lng":22.61,"typ":"gmina","sieci":2},{"n":"Bakałarzewo","lat":54.07,"lng":22.65,"typ":"gmina","sieci":2},{"n":"Raczki","lat":53.98,"lng":22.78,"typ":"gmina","sieci":2},{"n":"Nowinka","lat":53.92,"lng":23.02,"typ":"gmina","sieci":2},{"n":"Płaska","lat":53.94,"lng":23.24,"typ":"gmina","sieci":2},{"n":"Giby","lat":54.05,"lng":23.42,"typ":"gmina","sieci":2},{"n":"Krasnopol","lat":54.08,"lng":23.28,"typ":"gmina","sieci":2},{"n":"Puńsk","lat":54.24,"lng":23.19,"typ":"gmina","sieci":2},{"n":"Szypliszki","lat":54.20,"lng":23.02,"typ":"gmina","sieci":2},{"n":"Jeleniewo","lat":54.18,"lng":22.86,"typ":"gmina","sieci":2},{"n":"Suwałki-Wschód","lat":54.09,"lng":23.00,"typ":"gmina","sieci":2},{"n":"Wiżajny","lat":54.38,"lng":22.86,"typ":"gmina","sieci":2}],
  "warmińsko-mazurskie": [{"n":"Olsztyn","lat":53.78,"lng":20.49,"typ":"gmina","sieci":4},{"n":"Ełk","lat":53.83,"lng":22.36,"typ":"gmina","sieci":3},{"n":"Giżycko","lat":54.04,"lng":21.77,"typ":"gmina","sieci":2},{"n":"Szczytno","lat":53.56,"lng":21.00,"typ":"gmina","sieci":2},{"n":"Pisz","lat":53.63,"lng":21.81,"typ":"gmina","sieci":2},{"n":"Lidzbark Warmiński","lat":54.13,"lng":20.58,"typ":"gmina","sieci":2},{"n":"Olecko","lat":54.03,"lng":22.50,"typ":"gmina","sieci":2},{"n":"Gołdap","lat":54.31,"lng":22.31,"typ":"gmina","sieci":2},{"n":"Węgorzewo","lat":54.21,"lng":21.74,"typ":"gmina","sieci":2},{"n":"Nowe Miasto Lubawskie","lat":53.42,"lng":19.60,"typ":"gmina","sieci":2},{"n":"Dobrze Miasto-Wieś","lat":53.99,"lng":20.42,"typ":"gmina","sieci":2},{"n":"Wilczęta","lat":54.15,"lng":19.75,"typ":"gmina","sieci":2},{"n":"Górowo Iławeckie","lat":54.28,"lng":20.49,"typ":"gmina","sieci":2},{"n":"Srokowo","lat":54.22,"lng":21.52,"typ":"gmina","sieci":2},{"n":"Barciany","lat":54.20,"lng":21.35,"typ":"gmina","sieci":2},{"n":"Kruklanki","lat":54.11,"lng":21.94,"typ":"gmina","sieci":2},{"n":"Pozezdrze","lat":54.14,"lng":21.86,"typ":"gmina","sieci":2},{"n":"Budry","lat":54.25,"lng":21.98,"typ":"gmina","sieci":2},{"n":"Banie Mazurskie","lat":54.25,"lng":22.05,"typ":"gmina","sieci":2},{"n":"Kowale Oleckie","lat":54.14,"lng":22.44,"typ":"gmina","sieci":2},{"n":"Świętajno","lat":54.05,"lng":22.35,"typ":"gmina","sieci":2},{"n":"Wieliczki","lat":54.00,"lng":22.44,"typ":"gmina","sieci":2},{"n":"Prostki","lat":53.70,"lng":22.44,"typ":"gmina","sieci":2},{"n":"Kalinowo","lat":53.83,"lng":22.55,"typ":"gmina","sieci":2},{"n":"Stare Juchy","lat":53.94,"lng":22.16,"typ":"gmina","sieci":2},{"n":"Ryn","lat":53.94,"lng":21.55,"typ":"gmina","sieci":2},{"n":"Miłki","lat":53.94,"lng":21.80,"typ":"gmina","sieci":2},{"n":"Wydminy","lat":53.99,"lng":22.00,"typ":"gmina","sieci":2},{"n":"Orzysz","lat":53.81,"lng":21.94,"typ":"gmina","sieci":2},{"n":"Biała Piska","lat":53.61,"lng":22.05,"typ":"gmina","sieci":2},{"n":"Ruciane-Nida","lat":53.65,"lng":21.60,"typ":"gmina","sieci":2},{"n":"Świętajno-Szczytno","lat":53.65,"lng":21.14,"typ":"gmina","sieci":2},{"n":"Rozogi","lat":53.44,"lng":21.35,"typ":"gmina","sieci":2},{"n":"Wielbark","lat":53.40,"lng":20.94,"typ":"gmina","sieci":2},{"n":"Janowiec Kościelny","lat":53.20,"lng":20.50,"typ":"gmina","sieci":2},{"n":"Biskupiec Pomorski","lat":53.45,"lng":19.44,"typ":"gmina","sieci":2}],
}


# --- SCALANIE Z DEDUPLIKACJĄ -------------------------------------------------
# Ta sama miejscowość bywa zapisana pełną nazwą i skrótem ("Maków Mazowiecki"
# vs "Maków Maz."). Zwykłe porównanie napisów tego nie wyłapie, więc:
#   1) rozwijamy skróty przed porównaniem,
#   2) dodatkowo odrzucamy wpisy leżące bliżej niż 2 km od już istniejącego
#      (to na pewno ten sam punkt na mapie).
_SKROTY_NAZW = {
    "maz.": "mazowiecki", "maz": "mazowiecki", "mazowiecka": "mazowiecki",
    "mazowieckie": "mazowiecki", "wlkp.": "wielkopolski", "wlkp": "wielkopolski",
    "wielkopolska": "wielkopolski", "śl.": "śląski", "sl.": "śląski",
    "śląska": "śląski", "gd.": "gdański", "kuj.": "kujawski",
    "pom.": "pomorski", "pomorska": "pomorski", "lub.": "lubelski",
    "podl.": "podlaski", "podlaska": "podlaski", "warm.": "warmiński",
    "krak.": "krakowski", "ł.": "łódzki", "l.": "łódzki", "łódzka": "łódzki",
    "wlk.": "wielki", "wielka": "wielki", "wlk": "wielki",
    "dln.": "dolny", "dolna": "dolny", "górna": "górny", "gorna": "górny",
    "g.": "górny", "st.": "stary", "stara": "stary", "n.": "nowy", "nowa": "nowy",
}


def _normalizuj_nazwe(nazwa: str) -> str:
    """Ujednolica nazwę do porównań: małe litery, rozwinięte skróty.
    'Maków Maz.' i 'Maków Mazowiecki' dają ten sam wynik."""
    czesci = str(nazwa).strip().lower().replace("-", " ").split()
    return " ".join(_SKROTY_NAZW.get(c, c) for c in czesci)


def _blisko_km(lat1, lng1, lat2, lng2) -> float:
    """Odległość w linii prostej — do wykrywania tego samego punktu."""
    import math as _m
    dlat = _m.radians(lat2 - lat1); dlng = _m.radians(lng2 - lng1)
    a = (_m.sin(dlat / 2) ** 2 + _m.cos(_m.radians(lat1)) *
         _m.cos(_m.radians(lat2)) * _m.sin(dlng / 2) ** 2)
    return 6371.0 * 2 * _m.asin(_m.sqrt(a))


def _scal_baze_miast():
    """Dokłada rozszerzenie do bazy głównej, pomijając duplikaty. Dodatkowo
    czyści duplikaty już obecne w danych źródłowych — dzięki temu w jednej
    trasie nie pojawi się dwa razy to samo miasto pod dwiema nazwami."""
    for _woj in list(MIASTA_RAW.keys()) + [w for w in MIASTA_ROZSZERZENIE if w not in MIASTA_RAW]:
        istniejace = MIASTA_RAW.get(_woj, [])
        wynik, klucze = [], []
        for _m in list(istniejace) + list(MIASTA_ROZSZERZENIE.get(_woj, [])):
            klucz = _normalizuj_nazwe(_m.get("n", ""))
            lat, lng = float(_m.get("lat", 0)), float(_m.get("lng", 0))
            duplikat = False
            for k2, la2, ln2 in klucze:
                if klucz == k2 or _blisko_km(lat, lng, la2, ln2) < 2.0:
                    duplikat = True
                    break
            if duplikat:
                continue
            klucze.append((klucz, lat, lng))
            wynik.append(_m)
        MIASTA_RAW[_woj] = wynik


def _rozwin_skroty_w_bazie():
    """Na dokumencie i w trasach mają być pełne nazwy: 'Maków Mazowiecki',
    nie 'Maków Maz.'. Zamiana obejmuje wyłącznie widoczną nazwę."""
    _pelne = {"maz.": "Mazowiecki", "wlkp.": "Wielkopolski", "śl.": "Śląski",
              "gd.": "Gdański", "kuj.": "Kujawski", "pom.": "Pomorski",
              "lub.": "Lubelski", "podl.": "Podlaski", "warm.": "Warmiński",
              "krak.": "Krakowski", "ł.": "Łódzki", "wlk.": "Wielki",
              "g.": "Górny", "dln.": "Dolny", "st.": "Stary", "n.": "Nowy"}
    # formy żeńskie tam, gdzie nazwa miasta jest rodzaju żeńskiego
    _zenskie = {"Mazowiecki": "Mazowiecka", "Wielkopolski": "Wielkopolska",
                "Śląski": "Śląska", "Pomorski": "Pomorska", "Wielki": "Wielka"}
    _konczy_a = lambda n: n.strip().endswith("a")
    for _woj, _lista in MIASTA_RAW.items():
        for _m in _lista:
            czesci = str(_m.get("n", "")).split()
            if not czesci:
                continue
            zmiana = False
            for i, c in enumerate(czesci):
                pelna = _pelne.get(c.lower())
                if not pelna:
                    continue
                if i > 0 and _konczy_a(czesci[i - 1]):
                    pelna = _zenskie.get(pelna, pelna)
                czesci[i] = pelna
                zmiana = True
            if zmiana:
                _m["n"] = " ".join(czesci)


_scal_baze_miast()
_rozwin_skroty_w_bazie()

# --- Offline baza współrzędnych miast (z MIASTA_RAW) ---------------------
# UWAGA: MIASTA_RAW zawiera głównie mniejsze gminy — brakuje w niej WIĘKSZOŚCI
# dużych miast (Warszawa, Kraków, Radom, Kielce…). Dlatego uzupełniamy indeks
# poniższą listą największych miast, inaczej punkty w miastach nie miałyby
# lokalizacji i wypadałyby z planu.
MIASTA_DUZE = {
    "warszawa": (52.23, 21.01), "kraków": (50.06, 19.94), "łódź": (51.76, 19.46),
    "wrocław": (51.11, 17.03), "poznań": (52.41, 16.93), "gdańsk": (54.35, 18.65),
    "szczecin": (53.43, 14.55), "bydgoszcz": (53.12, 18.01), "lublin": (51.25, 22.57),
    "białystok": (53.13, 23.16), "katowice": (50.26, 19.02), "gdynia": (54.52, 18.53),
    "częstochowa": (50.81, 19.12), "radom": (51.40, 21.15), "toruń": (53.01, 18.60),
    "kielce": (50.87, 20.63), "rzeszów": (50.04, 22.00), "gliwice": (50.29, 18.67),
    "zabrze": (50.32, 18.79), "olsztyn": (53.78, 20.49), "bielsko-biała": (49.82, 19.04),
    "bytom": (50.35, 18.92), "zielona góra": (51.94, 15.51), "rybnik": (50.10, 18.55),
    "ruda śląska": (50.26, 18.86), "tychy": (50.14, 18.99), "dąbrowa górnicza": (50.32, 19.20),
    "gorzów wielkopolski": (52.73, 15.24), "płock": (52.55, 19.71), "wałbrzych": (50.77, 16.28),
    "włocławek": (52.65, 19.07), "tarnów": (50.01, 20.99), "chorzów": (50.30, 18.95),
    "koszalin": (54.19, 16.18), "legnica": (51.21, 16.16), "słupsk": (54.46, 17.03),
    "jaworzno": (50.20, 19.27), "jastrzębie-zdrój": (49.95, 18.60), "nowy sącz": (49.62, 20.70),
    "jelenia góra": (50.90, 15.73), "siedlce": (52.17, 22.29), "mysłowice": (50.21, 19.13),
    "piła": (53.15, 16.74), "konin": (52.22, 18.25), "piotrków trybunalski": (51.41, 19.70),
    "inowrocław": (52.79, 18.26), "lubin": (51.40, 16.20), "ostrów wielkopolski": (51.65, 17.81),
    "suwałki": (54.10, 22.93), "stargard": (53.34, 15.05), "gniezno": (52.53, 17.60),
    "ostrowiec świętokrzyski": (50.93, 21.39), "siemianowice śląskie": (50.33, 19.03),
    "głogów": (51.66, 16.08), "pabianice": (51.66, 19.36), "leszno": (51.84, 16.58),
    "żory": (50.05, 18.70), "zamość": (50.72, 23.25), "pruszków": (52.17, 20.81),
    "łomża": (53.18, 22.08), "ełk": (53.83, 22.36), "tarnobrzeg": (50.57, 21.68),
    "chełm": (51.14, 23.47), "mielec": (50.29, 21.42), "kędzierzyn-koźle": (50.35, 18.23),
    "piaseczno": (52.08, 21.02), "tomaszów mazowiecki": (51.53, 20.01), "świdnica": (50.84, 16.49),
    "bełchatów": (51.37, 19.36), "zgierz": (51.86, 19.41), "otwock": (52.11, 21.26),
    "legionowo": (52.40, 20.93), "wejherowo": (54.60, 18.24), "skierniewice": (51.96, 20.16),
    "starachowice": (51.05, 21.07), "racibórz": (50.09, 18.22), "świnoujście": (53.91, 14.25),
    "kutno": (52.23, 19.36), "sieradz": (51.60, 18.73), "nowa sól": (51.80, 15.72),
    "krosno": (49.69, 21.77), "szczecinek": (53.71, 16.70), "skarżysko-kamienna": (51.11, 20.88),
    "malbork": (54.04, 19.03), "sanok": (49.56, 22.21), "świętochłowice": (50.29, 18.92),
    "puławy": (51.42, 21.97), "przemyśl": (49.78, 22.77), "stalowa wola": (50.58, 22.05),
    "biała podlaska": (52.03, 23.12), "ciechanów": (52.88, 20.62), "żyrardów": (52.05, 20.44),
    "nowy targ": (49.48, 20.03), "oświęcim": (50.04, 19.22), "zawiercie": (50.49, 19.42),
    "wodzisław śląski": (50.00, 18.47), "mikołów": (50.17, 18.90), "knurów": (50.22, 18.66),
    "czechowice-dziedzice": (49.91, 19.00), "cieszyn": (49.75, 18.63), "będzin": (50.33, 19.13),
    "sopot": (54.44, 18.56), "tczew": (54.09, 18.79), "starogard gdański": (53.97, 18.53),
    "kwidzyn": (53.73, 18.93), "chojnice": (53.70, 17.56), "lębork": (54.55, 17.75),
    "rumia": (54.57, 18.39), "reda": (54.60, 18.35), "iława": (53.60, 19.57),
    "ostróda": (53.70, 19.97), "giżycko": (54.04, 21.76), "kętrzyn": (54.07, 21.38),
    "mrągowo": (53.87, 21.30), "szczytno": (53.56, 20.99), "działdowo": (53.24, 20.18),
    "augustów": (53.84, 22.98), "grudziądz": (53.48, 18.75), "elbląg": (54.16, 19.40),
    "opole": (50.67, 17.92), "kalisz": (51.76, 18.09), "sosnowiec": (50.28, 19.13),
    "nowy dwór mazowiecki": (52.43, 20.72), "wołomin": (52.34, 21.24), "mińsk mazowiecki": (52.18, 21.57),
    "grodzisk mazowiecki": (52.11, 20.63), "żyrardow": (52.05, 20.44), "sochaczew": (52.23, 20.24),
    "wyszków": (52.59, 21.46), "garwolin": (51.90, 21.62), "kozienice": (51.58, 21.55),
    "grójec": (51.86, 20.87), "pułtusk": (52.70, 21.09), "sokołów podlaski": (52.41, 22.25),
    "ostrołęka": (53.08, 21.57), "mława": (53.11, 20.38), "przasnysz": (53.02, 20.88),
    "zwoleń": (51.36, 21.59), "pionki": (51.48, 21.45), "szydłowiec": (51.23, 20.85),
    "końskie": (51.19, 20.41), "jędrzejów": (50.64, 20.30), "busko-zdrój": (50.47, 20.72),
    "sandomierz": (50.68, 21.75), "staszów": (50.56, 21.17), "opatów": (50.80, 21.42),
}

_INDEKS_MIAST = None

def _norm_miasto(nazwa: str) -> str:
    """Normalizuje nazwę miasta do porównań (małe litery, bez skrótów/spacji)."""
    s = str(nazwa or "").strip().lower()
    # usuń typowe skróty wojewódzkie/powiatowe na końcu
    for skrot in [" wlkp.", " maz.", " k.", " śl.", " lub.", " podl.", " pom."]:
        if s.endswith(skrot):
            s = s[:-len(skrot)]
    return s.strip()

def _zbuduj_indeks_miast():
    global _INDEKS_MIAST
    if _INDEKS_MIAST is not None:
        return _INDEKS_MIAST
    _INDEKS_MIAST = {}
    # 1) gminy z MIASTA_RAW
    for woj, miasta in MIASTA_RAW.items():
        for m in miasta:
            klucz = _norm_miasto(m["n"])
            if klucz and klucz not in _INDEKS_MIAST:
                _INDEKS_MIAST[klucz] = (m["lat"], m["lng"])
    # 2) duże miasta (mają pierwszeństwo — dokładniejsze centra)
    for nazwa, coords in MIASTA_DUZE.items():
        _INDEKS_MIAST[_norm_miasto(nazwa)] = coords
    return _INDEKS_MIAST

def coords_z_miasta(miasto: str):
    """Zwraca (lat, lng) z bazy offline dla nazwy miasta albo None.
    Natychmiastowe — bez internetu."""
    if not miasto:
        return None
    idx = _zbuduj_indeks_miast()
    return idx.get(_norm_miasto(miasto))


def planuj_wizyty(punkty: List[PunktWizyty], rok: int, miesiac: int,
                  baza: PunktWizyty = None, offline=False, postep_cb=None,
                  max_miesiecy=12, od_daty: datetime.date = None) -> dict:
    """Główny planer WIELOMIESIĘCZNY. Rozkłada punkty na dni robocze —
    jeśli nie mieszczą się w jednym miesiącu, przechodzą na kolejne (aż do
    max_miesiecy). Każdy punkt odwiedzany RAZ (w swoim miesiącu).
      - grupuje geograficznie (najbliższy sąsiad),
      - limit 8h dziennie (dojazdy liniowe + wizyty),
      - pon-pt, bez świąt.
    baza: punkt startowy każdego dnia (np. adres domowy). None = środek
    geograficzny punktów.
    od_daty: pomija dni robocze WCZEŚNIEJSZE niż podana data — używane przy
    przenoszeniu zaległych wizyt (nie planujemy w przeszłości).
    Zwraca dict: 'miesiace' (lista planów miesięcznych) + zbiorcze metryki.
    Dla zgodności zwraca też 'dni' = wszystkie dni ze wszystkich miesięcy."""
    z_pozycja = [p for p in punkty if p.lat is not None and p.lng is not None]
    bez_pozycji = [p for p in punkty if p.lat is None or p.lng is None]

    if not z_pozycja:
        return {"dni": [], "miesiace": [], "bez_pozycji": bez_pozycji,
                "nierozplanowane": [], "suma_km": 0.0, "suma_wizyt": 0,
                "rok": rok, "miesiac": miesiac, "liczba_miesiecy": 0}

    # baza — punkt startowy dnia. Jeśli nie podano, bierzemy środek geograficzny
    # (sztuczny). Realny adres domowy daje znacznie lepsze trasy.
    if baza is None or baza.lat is None:
        sx = sum(p.lat for p in z_pozycja) / len(z_pozycja)
        sy = sum(p.lng for p in z_pozycja) / len(z_pozycja)
        baza = PunktWizyty("Środek trasy", "", lat=sx, lng=sy)

    # parametry planowania z ustawień użytkownika (godziny/dzień, czas wizyty…)
    UP = ustawienia_planowania()
    LIMIT = UP["limit_minut"]; CZAS_W = UP["czas_wizyty"]
    PREDK = UP["predkosc"];    POWROT = UP["powrot"]

    pozostale = list(z_pozycja)
    miesiace_plan = []
    wszystkie_dni = []
    cur_rok, cur_mies = rok, miesiac
    licznik_mies = 0

    while pozostale and licznik_mies < max_miesiecy:
        dni_robocze = pobierz_dni_robocze(cur_rok, cur_mies)
        dni_robocze = [d for d in dni_robocze if not czy_dzien_wolny(d)]
        if od_daty:
            dni_robocze = [d for d in dni_robocze if d >= od_daty]
        dni_plan = []
        aktualna_pozycja = baza
        for data in dni_robocze:
            if not pozostale:
                break
            dzien = DzienPlanu(data)
            czas = 0.0; km_dnia = 0.0
            pozycja = aktualna_pozycja
            if postep_cb:
                zrobione = len(z_pozycja) - len(pozostale)
                postep_cb(f"Planowanie {cur_mies:02d}.{cur_rok} — {data.strftime('%d.%m')}...",
                          zrobione / max(len(z_pozycja), 1))
            while pozostale:
                najbl = None; najbl_km = None
                for p in pozostale:
                    d = _dyst_prosty(pozycja, p)
                    if najbl_km is None or d < najbl_km:
                        najbl_km = d; najbl = p
                czas_do = _czas_przejazdu_min(najbl_km, PREDK)
                # jeśli wliczamy powrót do bazy — rezerwujemy na niego czas
                rezerwa = _czas_przejazdu_min(_dyst_prosty(najbl, baza), PREDK) if POWROT else 0.0
                if czas + czas_do + CZAS_W + rezerwa > LIMIT and dzien.wizyty:
                    break
                dzien.wizyty.append(najbl)
                czas += czas_do + CZAS_W
                km_dnia += najbl_km
                pozycja = najbl
                pozostale.remove(najbl)
            if POWROT and dzien.wizyty:
                km_pow = _dyst_prosty(pozycja, baza)
                km_dnia += km_pow
                czas += _czas_przejazdu_min(km_pow, PREDK)
            dzien.km = round(km_dnia, 1)
            dzien.minuty = round(czas, 0)
            if dzien.wizyty:
                dni_plan.append(dzien)
                aktualna_pozycja = dzien.wizyty[-1]

        if dni_plan:
            miesiace_plan.append({
                "rok": cur_rok, "miesiac": cur_mies,
                "dni": dni_plan,
                "suma_km": round(sum(d.km for d in dni_plan), 1),
                "suma_wizyt": sum(len(d.wizyty) for d in dni_plan),
            })
            wszystkie_dni.extend(dni_plan)

        # następny miesiąc
        cur_mies += 1
        if cur_mies > 12:
            cur_mies = 1; cur_rok += 1
        licznik_mies += 1

    suma_km = round(sum(d.km for d in wszystkie_dni), 1)
    suma_wizyt = sum(len(d.wizyty) for d in wszystkie_dni)
    return {
        "dni": wszystkie_dni,           # wszystkie dni (zgodność wstecz)
        "miesiace": miesiace_plan,      # plany per miesiąc
        "liczba_miesiecy": len(miesiace_plan),
        "bez_pozycji": bez_pozycji,
        "nierozplanowane": pozostale,   # gdyby przekroczono max_miesiecy
        "suma_km": suma_km,
        "suma_wizyt": suma_wizyt,
        "rok": rok, "miesiac": miesiac,
    }


# --- DUŻE MIASTA (do reguł cyklu wizyt) -----------------------------------
# UWAGA: to NIE jest to samo co MIASTA_DUZE (tamto to baza współrzędnych,
# zawiera też mniejsze miejscowości). Tu są miasta faktycznie duże —
# powyżej ~100 tys. mieszkańców. Wg tej listy działa reguła "w dużym mieście
# odwiedzamy częściej".
MIASTA_WIELKIE = {
    "warszawa", "kraków", "łódź", "wrocław", "poznań", "gdańsk", "szczecin",
    "bydgoszcz", "lublin", "białystok", "katowice", "gdynia", "częstochowa",
    "radom", "sosnowiec", "toruń", "kielce", "rzeszów", "gliwice", "zabrze",
    "olsztyn", "bielsko-biała", "bytom", "zielona góra", "rybnik", "ruda śląska",
    "opole", "tychy", "gorzów wielkopolski", "elbląg", "dąbrowa górnicza",
    "płock", "wałbrzych", "włocławek", "tarnów", "chorzów", "koszalin",
    "kalisz", "legnica", "grudziądz", "słupsk", "jaworzno", "jastrzębie-zdrój",
    "nowy sącz", "jelenia góra", "konin", "piotrków trybunalski", "siedlce",
    "mysłowice", "piła", "ostrów wielkopolski", "lubin", "suwałki",
    "inowrocław", "gniezno", "stargard", "ostrowiec świętokrzyski",
    "siemianowice śląskie", "głogów", "pabianice", "leszno", "żory",
}

def czy_duze_miasto(miasto: str) -> bool:
    """Czy miejscowość to DUŻE miasto (>~100 tys.) — wg listy MIASTA_WIELKIE.
    Używane w regułach cyklu: np. Biedronki w miastach częściej niż poza nimi."""
    if not miasto:
        return False
    return _norm_miasto(miasto) in MIASTA_WIELKIE


def cykl_dla_punktu(punkt, cykl_domyslny=4, cykle_sieci=None,
                    cykle_sieci_duze=None, cykle_punktow=None) -> int:
    """Ustala co ile TYGODNI odwiedzać dany punkt. Hierarchia (od najsilniejszej):
      1. cykl ustawiony dla KONKRETNEGO punktu (adres) — nadpisuje wszystko,
      2. cykl sieci Z PODZIAŁEM na duże miasta / poza nimi,
      3. cykl sieci (wspólny),
      4. cykl domyślny.
    """
    cykle_sieci = cykle_sieci or {}
    cykle_sieci_duze = cykle_sieci_duze or {}
    cykle_punktow = cykle_punktow or {}

    def _int(v):
        try:
            v = int(v)
            return v if v >= 1 else None
        except (TypeError, ValueError):
            return None

    adres = (getattr(punkt, "adres", "") or getattr(punkt, "nazwa", "") or "").strip().lower()
    siec = (getattr(punkt, "siec", "") or "").strip()
    miasto = getattr(punkt, "miasto", "") or ""

    # 1) punkt indywidualny
    v = _int(cykle_punktow.get(adres))
    if v:
        return v
    # 2) sieć + klasa miasta
    if czy_duze_miasto(miasto):
        v = _int(cykle_sieci_duze.get(siec))
        if v:
            return v
    # 3) sieć (wspólny)
    v = _int(cykle_sieci.get(siec))
    if v:
        return v
    # 4) domyślny
    return max(1, _int(cykl_domyslny) or 4)


def planuj_cyklicznie(punkty: List[PunktWizyty], rok: int, miesiac: int,
                      cykl_domyslny: int = 4, cykle_sieci: dict = None,
                      cykle_sieci_duze: dict = None, cykle_punktow: dict = None,
                      miesiecy: int = 3, baza: PunktWizyty = None,
                      offline=False, od_daty: datetime.date = None,
                      postep_cb=None) -> dict:
    """PLANER CYKLICZNY (rotacja).

    Każdy punkt odwiedzany jest POWTARZALNIE — co X tygodni, a nie raz.
    Tak wygląda realna praca przedstawiciela: Biedronki co 2 tygodnie,
    Żabki co 4 itd.

    Jak działa:
      • każdy punkt ma termin następnej wizyty (na starcie: pierwszy dzień),
      • każdego dnia roboczego bierzemy punkty, którym termin już minął,
        i układamy z nich trasę metodą najbliższego sąsiada (limit 8h),
      • po odwiedzeniu punkt dostaje nowy termin: dziś + jego cykl.

    Dzięki temu plan sam się balansuje: punkty niedawno odwiedzone czekają,
    a te zaległe mają pierwszeństwo w kolejnych dniach.

    cykl_domyslny  — co ile tygodni odwiedzać punkt (gdy sieć nie ma własnego)
    cykle_sieci    — słownik {nazwa_sieci: tygodnie}, np. {"Biedronka": 2}
    miesiecy       — na ile miesięcy w przód generujemy cykl
    """
    cykle_sieci = cykle_sieci or {}
    cykle_sieci_duze = cykle_sieci_duze or {}
    cykle_punktow = cykle_punktow or {}
    z_pozycja = [p for p in punkty if p.lat is not None and p.lng is not None]
    bez_pozycji = [p for p in punkty if p.lat is None or p.lng is None]

    if not z_pozycja:
        return {"dni": [], "miesiace": [], "liczba_miesiecy": 0,
                "bez_pozycji": bez_pozycji, "nierozplanowane": [],
                "suma_km": 0.0, "suma_wizyt": 0, "rok": rok, "miesiac": miesiac,
                "cykliczny": True}

    if baza is None or baza.lat is None:
        sx = sum(p.lat for p in z_pozycja) / len(z_pozycja)
        sy = sum(p.lng for p in z_pozycja) / len(z_pozycja)
        baza = PunktWizyty("Środek trasy", "", lat=sx, lng=sy)

    def cykl_punktu(p) -> int:
        """Ile tygodni między wizytami — wg hierarchii reguł."""
        return cykl_dla_punktu(p, cykl_domyslny, cykle_sieci,
                               cykle_sieci_duze, cykle_punktow)

    # parametry planowania z ustawień użytkownika
    UP = ustawienia_planowania()
    LIMIT = UP["limit_minut"]; CZAS_W = UP["czas_wizyty"]
    PREDK = UP["predkosc"];    POWROT = UP["powrot"]

    # zbierz dni robocze w horyzoncie
    dni_robocze = []
    cur_r, cur_m = rok, miesiac
    for _ in range(max(1, miesiecy)):
        dni_robocze.extend(pobierz_dni_robocze(cur_r, cur_m))
        cur_m += 1
        if cur_m > 12:
            cur_m = 1; cur_r += 1
    if od_daty:
        dni_robocze = [d for d in dni_robocze if d >= od_daty]
    dni_robocze.sort()
    if not dni_robocze:
        return {"dni": [], "miesiace": [], "liczba_miesiecy": 0,
                "bez_pozycji": bez_pozycji, "nierozplanowane": [],
                "suma_km": 0.0, "suma_wizyt": 0, "rok": rok, "miesiac": miesiac,
                "cykliczny": True}

    start = dni_robocze[0]
    # termin następnej wizyty dla każdego punktu (na starcie: wszystkie zaległe)
    termin = {id(p): start for p in z_pozycja}

    dni_plan = []
    for nr_dnia, data in enumerate(dni_robocze):
        if postep_cb and nr_dnia % 5 == 0:
            postep_cb(f"Układanie cyklu ({nr_dnia+1}/{len(dni_robocze)})...",
                      (nr_dnia + 1) / len(dni_robocze))
        # punkty, którym termin już minął
        nalezne = [p for p in z_pozycja if termin[id(p)] <= data]
        if not nalezne:
            continue
        dzien = DzienPlanu(data)
        czas = 0.0; km_dnia = 0.0
        pozycja = baza
        dostepne = list(nalezne)
        while dostepne:
            najbl = None; najbl_km = None
            for p in dostepne:
                d = _dyst_prosty(pozycja, p)
                if najbl_km is None or d < najbl_km:
                    najbl_km = d; najbl = p
            czas_do = _czas_przejazdu_min(najbl_km, PREDK)
            rezerwa = _czas_przejazdu_min(_dyst_prosty(najbl, baza), PREDK) if POWROT else 0.0
            if czas + czas_do + CZAS_W + rezerwa > LIMIT and dzien.wizyty:
                break                      # dzień pełny
            dzien.wizyty.append(najbl)
            czas += czas_do + CZAS_W
            km_dnia += najbl_km
            pozycja = najbl
            dostepne.remove(najbl)
            # następna wizyta w tym punkcie za X tygodni
            termin[id(najbl)] = data + datetime.timedelta(weeks=cykl_punktu(najbl))
        if POWROT and dzien.wizyty:
            km_pow = _dyst_prosty(pozycja, baza)
            km_dnia += km_pow
            czas += _czas_przejazdu_min(km_pow, PREDK)
        dzien.km = round(km_dnia, 1)
        dzien.minuty = round(czas, 0)
        if dzien.wizyty:
            dni_plan.append(dzien)

    # grupowanie miesięczne
    miesiace_plan = []
    for d in dni_plan:
        klucz = (d.data.year, d.data.month)
        if not miesiace_plan or (miesiace_plan[-1]["rok"], miesiace_plan[-1]["miesiac"]) != klucz:
            miesiace_plan.append({"rok": d.data.year, "miesiac": d.data.month,
                                  "dni": [], "suma_km": 0.0, "suma_wizyt": 0})
        miesiace_plan[-1]["dni"].append(d)
    for m in miesiace_plan:
        m["suma_km"] = round(sum(x.km for x in m["dni"]), 1)
        m["suma_wizyt"] = sum(len(x.wizyty) for x in m["dni"])

    # punkty, które ani razu nie trafiły do planu (np. skrajnie oddalone)
    odwiedzone_ids = set()
    for d in dni_plan:
        for w in d.wizyty:
            odwiedzone_ids.add(id(w))
    nigdy = [p for p in z_pozycja if id(p) not in odwiedzone_ids]

    return {
        "dni": dni_plan,
        "miesiace": miesiace_plan,
        "liczba_miesiecy": len(miesiace_plan),
        "bez_pozycji": bez_pozycji,
        "nierozplanowane": nigdy,
        "suma_km": round(sum(d.km for d in dni_plan), 1),
        "suma_wizyt": sum(len(d.wizyty) for d in dni_plan),
        "rok": rok, "miesiac": miesiac,
        "cykliczny": True,
        "cykl_domyslny": cykl_domyslny,
        "cykle_sieci": dict(cykle_sieci),
        "cykle_sieci_duze": dict(cykle_sieci_duze),
        "cykle_punktow": dict(cykle_punktow),
    }


def zalegle_wizyty(plan: dict) -> list:
    """Zwraca listę (data, punkt) wizyt z MINIONYCH dni, których nie odhaczono.
    To one 'wypadły' z planu — trzeba je przenieść na kolejne dni."""
    if not plan:
        return []
    dzis = datetime.date.today()
    out = []
    for d in plan.get("dni", []):
        if d.data < dzis:
            for w in d.wizyty:
                if not czy_odwiedzona(d.data, w.adres or w.nazwa):
                    out.append((d.data, w))
    return out


def przeplanuj_zalegle(plan: dict, baza: PunktWizyty = None, offline=False) -> dict:
    """Przenosi niezrealizowane wizyty na kolejne dni robocze.

    Zasady:
      • HISTORIA JEST NIETYKALNA — dni minione zachowują tylko te wizyty, które
        faktycznie odhaczono (z ich pieczątkami i notatkami). Nic nie znika
        i nic nie jest dopisywane wstecz.
      • Wszystkie wizyty nieodhaczone (zaległe z przeszłości + przyszłe)
        układane są od nowa, poczynając od DZIŚ.
    """
    if not plan:
        return plan
    dzis = datetime.date.today()

    # 1) historia: minione dni, tylko odhaczone wizyty
    historia = []
    for d in plan.get("dni", []):
        if d.data < dzis:
            zrobione = [w for w in d.wizyty if czy_odwiedzona(d.data, w.adres or w.nazwa)]
            if zrobione:
                nd = DzienPlanu(d.data)
                nd.wizyty = zrobione
                nd.km = d.km; nd.minuty = d.minuty
                historia.append(nd)

    # 2) do przeplanowania: wszystko, co nieodhaczone (przeszłe i przyszłe)
    do_zaplanowania = []
    for d in plan.get("dni", []):
        for w in d.wizyty:
            if not czy_odwiedzona(d.data, w.adres or w.nazwa):
                do_zaplanowania.append(w)
    # dorzuć te, które wcześniej się nie zmieściły
    do_zaplanowania.extend(plan.get("nierozplanowane", []))

    if not do_zaplanowania:
        return plan     # nie ma czego przenosić

    if plan.get("cykliczny"):
        # PLAN CYKLICZNY — punkty się powtarzają, więc nie przeplanowujemy
        # pojedynczych wystąpień, tylko generujemy cykl od nowa (od dziś)
        # na unikalnych punktach. Historia zostaje nietknięta.
        unikalne = {}
        for w in do_zaplanowania:
            klucz = (w.adres or w.nazwa).strip().lower()
            if klucz not in unikalne:
                unikalne[klucz] = w
        nowy = planuj_cyklicznie(
            list(unikalne.values()), dzis.year, dzis.month,
            cykl_domyslny=int(plan.get("cykl_domyslny", 4) or 4),
            cykle_sieci=plan.get("cykle_sieci", {}) or {},
            cykle_sieci_duze=plan.get("cykle_sieci_duze", {}) or {},
            cykle_punktow=plan.get("cykle_punktow", {}) or {},
            miesiecy=max(1, plan.get("liczba_miesiecy", 3) or 3),
            baza=baza, offline=offline, od_daty=dzis)
    else:
        nowy = planuj_wizyty(do_zaplanowania, dzis.year, dzis.month,
                             baza=baza, offline=offline, od_daty=dzis)

    # 3) scal: historia + nowe dni
    wszystkie = historia + nowy.get("dni", [])
    wszystkie.sort(key=lambda d: d.data)

    # przebuduj grupowanie miesięczne
    miesiace = []
    for d in wszystkie:
        klucz = (d.data.year, d.data.month)
        if not miesiace or (miesiace[-1]["rok"], miesiace[-1]["miesiac"]) != klucz:
            miesiace.append({"rok": d.data.year, "miesiac": d.data.month,
                             "dni": [], "suma_km": 0.0, "suma_wizyt": 0})
        miesiace[-1]["dni"].append(d)
    for m in miesiace:
        m["suma_km"] = round(sum(x.km for x in m["dni"]), 1)
        m["suma_wizyt"] = sum(len(x.wizyty) for x in m["dni"])

    return {
        "dni": wszystkie, "miesiace": miesiace, "liczba_miesiecy": len(miesiace),
        "bez_pozycji": plan.get("bez_pozycji", []),
        "nierozplanowane": nowy.get("nierozplanowane", []),
        "suma_km": round(sum(d.km for d in wszystkie), 1),
        "suma_wizyt": sum(len(d.wizyty) for d in wszystkie),
        "rok": plan.get("rok"), "miesiac": plan.get("miesiac"),
        "cykliczny": bool(plan.get("cykliczny")),
        "cykl_domyslny": plan.get("cykl_domyslny", 4),
        "cykle_sieci": plan.get("cykle_sieci", {}) or {},
        "adres_bazy": plan.get("adres_bazy", ""),
    }


def _tydzien_numer(data: datetime.date, dni_robocze: List[datetime.date]) -> int:
    """Numer tygodnia w miesiącu (1..) dla grupowania planu tygodniowego."""
    if not dni_robocze:
        return 1
    pierwszy = dni_robocze[0]
    return (data - pierwszy).days // 7 + 1


def generuj_mape_planu_html(plan: dict, folder: str, is_dark: bool = True) -> str:
    """Tworzy plik HTML z podglądem planu wizyt — każdy dzień jako karta z
    trasą i przyciskiem 'Otwórz w Google Maps'. Zwraca ścieżkę pliku."""
    bg = "#0B1320" if is_dark else "#F1F5F9"
    txt = "#F8FAFC" if is_dark else "#0F172A"
    karta = "rgba(16,25,43,0.85)" if is_dark else "#FFFFFF"
    ramka = "rgba(0,240,255,0.3)" if is_dark else "rgba(13,148,136,0.3)"
    tytul_c = "#00F0FF" if is_dark else "#0D9488"
    mut = "#94A3B8" if is_dark else "#475569"
    btn_bg = "linear-gradient(to right,#00F0FF,#00E4A1)" if is_dark else "#0D9488"
    btn_txt = "#04121A" if is_dark else "#FFFFFF"

    mies_pl = ["","styczeń","luty","marzec","kwiecień","maj","czerwiec","lipiec",
               "sierpień","wrzesień","październik","listopad","grudzień"]
    naglo_mies = f"{mies_pl[plan['miesiac']].capitalize()} {plan['rok']}"

    html = f"""<!DOCTYPE html><html lang="pl"><head><meta charset="UTF-8">
<title>Plan Wizyt — {naglo_mies}</title><style>
body {{ font-family:'Segoe UI',sans-serif; background:{bg}; color:{txt}; padding:24px; }}
.container {{ max-width:820px; margin:0 auto; }}
h1 {{ color:{tytul_c}; text-align:center; margin-bottom:4px; }}
.sub {{ text-align:center; color:{mut}; margin-bottom:24px; }}
.day {{ background:{karta}; border:1px solid {ramka}; border-radius:10px; padding:16px 20px; margin-bottom:14px; }}
.day-t {{ font-size:15px; font-weight:bold; margin-bottom:6px; }}
.meta {{ font-size:12px; color:{mut}; margin-bottom:10px; }}
.route {{ font-size:13px; color:{mut}; line-height:1.6; margin-bottom:12px; }}
.btn {{ display:inline-block; background:{btn_bg}; color:{btn_txt}; padding:8px 16px; border-radius:6px; text-decoration:none; font-weight:bold; font-size:13px; }}
</style></head><body><div class="container">
<h1>Plan Wizyt</h1><div class="sub">{naglo_mies} &nbsp;•&nbsp; {plan['suma_wizyt']} wizyt &nbsp;•&nbsp; {plan['suma_km']:.0f} km</div>
"""
    dni_pl = ["Poniedziałek","Wtorek","Środa","Czwartek","Piątek","Sobota","Niedziela"]
    for d in plan["dni"]:
        if not d.wizyty:
            continue
        etykiety = []; url_punkty = []
        for w in d.wizyty:
            ety = w.nazwa or w.miasto
            etykiety.append(ety)
            cel = w.adres if w.adres else f"{w.miasto}, Polska"
            url_punkty.append(urllib.parse.quote(cel))
        h = int(d.minuty // 60); m = int(d.minuty % 60)
        dzien_tyg = dni_pl[d.data.weekday()]
        html += f"""<div class="day"><div class="day-t">{dzien_tyg}, {d.data.strftime('%d.%m.%Y')}</div>
<div class="meta">{len(d.wizyty)} wizyt • {d.km:.0f} km • ok. {h}h {m}min</div>
<div class="route">{' ➔ '.join(etykiety)}</div>
<a class="btn" target="_blank" href="https://www.google.com/maps/dir/{'/'.join(url_punkty)}">🗺 Otwórz w Google Maps</a></div>
"""
    html += "</div></body></html>"
    sciezka = os.path.join(folder, "Plan_Wizyt_Mapa.html")
    with open(sciezka, "w", encoding="utf-8") as f:
        f.write(html)
    return sciezka


# ==========================================================================
# ||   KONIEC NOWEGO SILNIKA PLANERA. Poniżej: stary generator (S11).      ||
# ==========================================================================


def waliduj_pesel(pesel: str) -> bool:
    if not re.match(r"^\d{11}$", pesel): return False
    return (10 - sum(int(pesel[i]) * w for i, w in enumerate([1, 3, 7, 9, 1, 3, 7, 9, 1, 3])) % 10) % 10 == int(pesel[10])

def _czy_znana_miejscowosc(nazwa: str) -> bool:
    """Czy nazwa występuje w naszej bazie miejscowości? Jeśli tak, adres bez
    przedrostka jest adresem wiejskim i nie trzeba o nic pytać."""
    try:
        cel = str(nazwa).strip().lower()
        if not cel:
            return False
        for _lista in MIASTA_RAW.values():
            for _m in _lista:
                if str(_m.get("n", "")).strip().lower() == cel:
                    return True
        for _n in MIASTA_DUZE:
            if str(_n).strip().lower() == cel:
                return True
    except Exception:
        pass
    return False


def _klasyfikuj_adres(nazwa: str) -> str:
    """Zwraca typ członu po 'ul.':
      'ulica'          — pewna ulica (nie pytamy użytkownika),
      'wies'           — pewna miejscowość (nie pytamy),
      'niejednoznaczne'— nie da się rozstrzygnąć z nazwy → zapytamy użytkownika.
    Nie da się w 100% odróżnić 'Zielona Ścieżka' (ulica) od 'Sarnowa Góra'
    (wieś) — obie to dwa słowa z wielkich liter. Dlatego takie przypadki
    oznaczamy jako niejednoznaczne i pytamy użytkownika."""
    n = nazwa.strip()
    if not n:
        return 'ulica'
    # Nazwa jest w bazie miejscowości → adres wiejski, bez pytania użytkownika.
    if _czy_znana_miejscowosc(n):
        return 'wies'
    slowa = n.split()
    low = [s.lower().strip(".") for s in slowa]

    # 1) PEWNA ULICA — typowe słowa uliczne albo patron z inicjałami/cyframi
    ulica_slowa = {"aleja", "aleje", "rondo", "plac", "osiedle", "os", "skwer", "bulwar"}
    if any(s in ulica_slowa for s in low):
        return 'ulica'
    if any(s in {"i", "im", "gen", "ks", "sw", "św", "dr", "mjr", "kpt", "płk", "por"} for s in low):
        return 'ulica'
    if any(any(ch.isdigit() for ch in s) for s in slowa):   # "3 Maja", "1000-lecia"
        return 'ulica'

    # Częste patronki/patroni ulic (imiona, tytuły) — jeśli pierwszy człon to
    # popularne imię, to niemal na pewno ULICA (Jana Pawła, Marii Curie itd.).
    imiona_ulic = {"jana", "marii", "adama", "stefana", "józefa", "jozefa",
                   "tadeusza", "władysława", "wladyslawa", "kazimierza",
                   "mikołaja", "mikolaja", "henryka", "stanisława", "stanislawa",
                   "bolesława", "boleslawa", "fryderyka", "ignacego", "romualda",
                   "księdza", "ksiedza", "świętego", "swietego", "generała", "generala",
                   "marszałka", "marszalka", "papieża", "papieza", "króla", "krola"}
    if low[0] in imiona_ulic:
        return 'ulica'

    # 2) Jednoczłonowa nazwa (Marszałkowska, Kowalskiego, Krzyżówki) — traktujemy
    #    jako ULICA (najczęstszy przypadek miejski). Jeśli to jednak wieś
    #    jednoczłonowa, użytkownik użyje przełącznika (rzadki przypadek).
    if len(slowa) == 1:
        return 'ulica'

    # 3) Dwu-/trzyczłonowa, każde słowo z wielkiej litery — NIEJEDNOZNACZNE
    #    (Sarnowa Góra = wieś, Zielona Ścieżka = ulica). Pytamy użytkownika.
    if 2 <= len(slowa) <= 3 and all(s[0].isupper() for s in slowa if s):
        return 'niejednoznaczne'

    # 4) Wszystko inne (np. z małą literą w środku: "Jana Pawła") — ULICA
    return 'ulica'


def _wyglada_jak_miejscowosc(nazwa: str) -> bool:
    """Zgodność wstecz: True tylko dla PEWNEJ wsi."""
    return _klasyfikuj_adres(nazwa) == 'wies'


def waliduj_adres(adres: str, wymus_typ: str = None) -> dict:
    """wymus_typ: None (auto), 'ulica' lub 'wies' — decyzja użytkownika dla
    adresów niejednoznacznych. Gdy typ jest niejednoznaczny i nie podano
    wymus_typ, zwracamy flagę 'niejednoznaczne'=True, żeby formularz zapytał."""
    # PRZEDROSTEK JEST OPCJONALNY. Adresy wiejskie zapisuje się bez "ul." —
    # np. "Sarnowa Góra 42, 06-430 Sońsk". Wcześniej taki adres był odrzucany,
    # a użytkownik był zmuszany dopisywać "ul." przed nazwą wsi.
    m = re.match(
        r"^(?:(ul\.|al\.|pl\.|ulica|aleja|plac|os\.|osiedle)\s+)?"   # opcjonalny przedrostek
        r"(.+?)\s+"                                                   # nazwa (ulicy albo wsi)
        r"(\d+[a-zA-Z]?(?:\s*[/\\]\s*\d+[a-zA-Z]?)?)"                 # numer, także 12/3 lub 12A
        r"[,\s]+\s*(\d{2}-\d{3})\s+(.+)$",                           # kod pocztowy + poczta
        adres.strip(), re.IGNORECASE)
    if not m:
        raise ValueError(
            "Błędny format adresu!\n\nPodaj adres wg jednego ze wzorów:\n"
            "  Kowalczyka 12, 03-193 Warszawa\n"
            "  ul. Kowalczyka 12, 03-193 Warszawa\n"
            "  Sarnowa Góra 42, 06-430 Sońsk        (adres wiejski)\n\n"
            "Wymagany jest numer domu, kod pocztowy i miejscowość.")
    prefiks_raw = (m.group(1) or "").lower()
    # ujednolicamy zapis przedrostka; brak przedrostka = pusty napis
    prefiks = {"ulica": "ul.", "aleja": "al.", "plac": "pl.",
               "osiedle": "os.", "os.": "os.", "ul.": "ul.",
               "al.": "al.", "pl.": "pl."}.get(prefiks_raw, prefiks_raw)
    nazwa = ' '.join(w.capitalize() for w in m.group(2).split())   # "Sarnowa Góra" lub "Kowalskiego"
    numer = m.group(3).strip()
    kod = m.group(4).strip()
    poczta = ' '.join(w.capitalize() for w in m.group(5).split())  # "Sońsk" / "Warszawa"
    # Jawny przedrostek to rozstrzygnięcie użytkownika — nie pytamy o typ.
    if prefiks and not wymus_typ:
        wymus_typ = 'ulica'

    typ = _klasyfikuj_adres(nazwa)
    niejednoznaczne = False
    if wymus_typ in ('ulica', 'wies'):
        typ = wymus_typ                    # użytkownik rozstrzygnął
    elif typ == 'niejednoznaczne':
        niejednoznaczne = True
        typ = 'ulica'                      # tymczasowo, formularz i tak zapyta

    to_wies = (typ == 'wies')

    if to_wies:
        # Miejscowością (bazą, startem/metą, celem geocodingu) jest nazwa wsi,
        # a NIE poczta. Poczta to tylko urząd, do którego wieś przynależy.
        baza_miasto = nazwa
        adres_pdf = f"{nazwa} {numer}, {kod} {poczta}"
        adres_geo = f"{nazwa}, {kod} {poczta}"
    else:
        # Klasyczny adres miejski z ulicą — start/meta to MIEJSCOWOŚĆ (poczta).
        baza_miasto = poczta
        _pre = (prefiks + " ") if prefiks else ""
        adres_pdf = f"{_pre}{nazwa} {numer}, {kod} {poczta}"
        adres_geo = f"{_pre}{nazwa} {numer}, {kod} {poczta}"

    return {
        'adres_caly': adres_pdf,        # to trafia do PDF (start/meta, nagłówek)
        'adres_geo': adres_geo,         # to idzie do geocodingu (wyznaczenie bazy)
        'kod_pocztowy': kod,
        'baza_miasto': baza_miasto,     # start/meta trasy i to, co liczy silnik
        'to_wies': to_wies,
        'niejednoznaczne': niejednoznaczne,   # True → formularz pyta użytkownika
        'nazwa_do_pytania': nazwa,      # co pokazać w pytaniu
    }

def waliduj_miesiac(s: str) -> Tuple[int, int]:
    if not re.match(r"^\d{2}\.\d{4}$", s.strip()): raise ValueError("Format daty: MM.YYYY (np. 06.2026)")
    return map(int, s.strip().split('.'))

def waliduj_kwote(s: str) -> float:
    try: k = float(s.replace(' ', '').replace(',', '.'))
    except ValueError: raise ValueError("Kwota nieprawidłowa.")
    if k < MIN_KWOTA: raise ValueError(f"Min kwota to {MIN_KWOTA:.0f} PLN.")
    return k

def pobierz_coords(adres_caly: str, miasto: str, woj: str, zapisz_cache=True) -> Tuple[float, float]:
    global _geo_cache
    klucz = adres_caly.lower().strip()
    if klucz in _geo_cache: return tuple(_geo_cache[klucz])
    for zapytanie in [adres_caly, miasto]:
        try:
            q = urllib.parse.quote(f"{zapytanie}, Polska")
            req = urllib.request.Request(f"https://nominatim.openstreetmap.org/search?q={q}&format=json&limit=1", headers={'User-Agent': 'PMT/112'})
            with urllib.request.urlopen(req, timeout=4) as resp:
                data = json.loads(resp.read())
                if data:
                    coords = (float(data[0]['lat']), float(data[0]['lon']))
                    _geo_cache[klucz] = list(coords)
                    # Zapis cache TYLKO gdy zapisz_cache=True. Przy masowym
                    # geokodowaniu (setki punktów) zapisujemy raz na końcu —
                    # inaczej zapis całego pliku przy każdym punkcie zawiesza program.
                    if zapisz_cache:
                        with open(GEO_CACHE, 'w', encoding='utf-8') as f: json.dump(_geo_cache, f, ensure_ascii=False, indent=2)
                    return coords
        except Exception: continue
    return STOLICE.get(woj, (52.23, 21.01))


def zapisz_geo_cache():
    """Zapisuje cache geolokalizacji na dysk (po masowym geokodowaniu)."""
    global _geo_cache
    try:
        with open(GEO_CACHE, 'w', encoding='utf-8') as f:
            json.dump(_geo_cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def rozpoznaj_wojewodztwo(kod: str) -> str:
    try: p = int(kod[:2])
    except: return "mazowieckie"
    if p <= 9: return "mazowieckie"
    if p <= 14: return "warmińsko-mazurskie"
    if p <= 19: return "podlaskie"
    if p <= 24: return "lubelskie"
    if p <= 29: return "świętokrzyskie"
    if p <= 34: return "małopolskie"
    if p <= 39: return "podkarpackie"
    if p <= 44: return "śląskie"
    if p <= 49: return "opolskie"
    if p <= 59: return "dolnośląskie"
    if p <= 64: return "wielkopolskie"
    if p <= 69: return "lubuskie"
    if p <= 78: return "zachodniopomorskie"
    if p <= 84: return "pomorskie"
    if p <= 89: return "kujawsko-pomorskie"
    return "łódzkie"

def szacuj_delegacje(kwota: float, stawka: float, dni_robocze: int) -> dict:
    """Szybki szacunek BEZ uruchamiania pełnego algorytmu — zasila panel
    podglądu na żywo. Zwraca przewidywany dystans, liczbę dni z wyjazdami
    i średni dzienny przebieg. To przybliżenie oparte na tej samej fizyce
    co silnik (dystans = kwota / stawka), nie na realnym doborze miast."""
    if stawka <= 0 or dni_robocze <= 0:
        return {"km": 0, "dni_wyjazdowe": 0, "km_dzien": 0, "dokumenty": 0}
    km_total = kwota / stawka
    # Realny mnożnik krzywizny dróg (jak w silniku fizyki)
    km_real = km_total
    # średnio ~120 km przejazdu liniowego na dzień wyjazdowy (limit czasowy 8h)
    sredni_dzien_km = 115.0
    dni_wyjazdowe = max(1, min(dni_robocze, round(km_real / sredni_dzien_km)))
    km_dzien = km_real / dni_wyjazdowe if dni_wyjazdowe else 0
    # dokumenty: limit ~597 zł na dokument
    dokumenty = max(1, math.ceil(kwota / 590.0))
    return {
        "km": round(km_real),
        "dni_wyjazdowe": dni_wyjazdowe,
        "km_dzien": round(km_dzien),
        "dokumenty": dokumenty,
    }

def optymalizuj_tsp(trasa: List[Miasto], baza_lat, baza_lng) -> List[Miasto]:
    if len(trasa) <= 1: return trasa
    najlepsza = trasa
    min_dyst = float('inf')
    for perm in itertools.permutations(trasa):
        d = oblicz_dystans(baza_lat, baza_lng, perm[0].lat, perm[0].lng)
        for i in range(len(perm) - 1): d += oblicz_dystans(perm[i].lat, perm[i].lng, perm[i+1].lat, perm[i+1].lng)
        d += oblicz_dystans(perm[-1].lat, perm[-1].lng, baza_lat, baza_lng)
        if d < min_dyst: min_dyst = d; najlepsza = list(perm)
    return najlepsza

def usun_duplikaty_nazw(lista: List[Miasto]) -> List[Miasto]:
    u = {}
    for m in lista:
        if m.n not in u: u[m.n] = m
    return list(u.values())

def przecina_aglomeracje(lat1, lng1, lat2, lng2, baza_lat=None, baza_lng=None):
    """Sprawdza czy odcinek trasy przechodzi przez aglomerację.
    Aglomeracje w promieniu 30 km od bazy użytkownika są POMIJANE —
    mieszkaniec stolicy województwa musi móc z niej wyjechać i do niej wrócić."""
    for w_lat, w_lng in AGLOMERACJE_LISTA:
        if baza_lat is not None and oblicz_dystans(baza_lat, baza_lng, w_lat, w_lng) < 30.0:
            continue  # macierzysta aglomeracja — nie blokuje tras z/do domu
        if dystans_odcinek_punkt(lat1, lng1, lat2, lng2, w_lat, w_lng) < 25.0: return True
    return False

def miasto_w_obcej_aglomeracji(baza_lat, baza_lng, m_lat, m_lng):
    for w_lat, w_lng in AGLOMERACJE_LISTA:
        if oblicz_dystans(m_lat, m_lng, w_lat, w_lng) < 40.0:
            if oblicz_dystans(baza_lat, baza_lng, w_lat, w_lng) >= 45.0: return True
    return False

class ListaTras(list):
    """Lista dni tras z dodatkowymi metadanymi o pokryciu kwoty. Dziedziczy po
    list, więc działa wszędzie jak zwykła lista — po prostu niesie też info,
    czy realnymi trasami udało się pokryć żądaną kwotę."""
    kwota_docelowa = 0.0
    kwota_osiagnieta = 0.0
    kwota_niepelna = False
    brak_dni = False


def generuj_trasy(kwota_calkowita, baza_nazwa, baza_lat, baza_lng, woj, dni_robocze, pesel, stawka=None, postep_cb=None) -> List[DzienTrasy]:
    if stawka is None: stawka = STAWKA_ZA_KM
    seed_val = int(pesel[-6:]) if pesel.isdigit() and len(pesel) >= 6 else 42
    rng = random.Random(seed_val)
    baza = zaladuj_baze(baza_lat, baza_lng)
    
    ostatnie_uzycie = {} 
    finalne_dni = []
    
    cel_calkowity_dystans = kwota_calkowita / stawka
    aktualny_dystans_linii = 0.0

    sasiednie = [w for w in SASIEDZI_WOJ.get(woj, []) if w != woj]
    zablokowane_sektory = pobierz_zablokowane_sektory(baza_lat, baza_lng, woj)
    dozwolone_sektory = [s for s in SEKTORY_KOLEJNOSC if s not in zablokowane_sektory]
    if not dozwolone_sektory: dozwolone_sektory = SEKTORY_KOLEJNOSC[:]

    # ADAPTACYJNY PROMIEŃ PĘTLI — bazowo 70 km, ale w rzadko zaludnionych
    # rejonach (peryferia, przygranicze) miast bywa za mało, więc rozszerzamy
    # promień, żeby dało się w ogóle ułożyć trasę z min. 3 postojami.
    _wszystkie_okoliczne = []
    for _w in [woj] + sasiednie:
        for _m in baza.get(_w, []):
            _wszystkie_okoliczne.append(oblicz_dystans(baza_lat, baza_lng, _m.lat, _m.lng))
    _blisko_70 = sum(1 for d in _wszystkie_okoliczne if d <= 70.0)
    if _blisko_70 >= 25:
        promien_petli = 70.0
    elif _blisko_70 >= 15:
        promien_petli = 95.0
    elif _blisko_70 >= 8:
        promien_petli = 120.0
    else:
        promien_petli = 150.0

    # BAZA PRZYGRANICZNA: gdy blokada sektorów (kierunki "za granicę") zostawia
    # za mało miast w zasięgu, odblokowujemy wszystkie sektory. Inaczej pracownik
    # z rogu kraju (Suwałki, Ustrzyki, Kołobrzeg) nie dostałby żadnej trasy.
    _miast_w_dozwolonych = 0
    for _w in [woj] + sasiednie:
        for _m in baza.get(_w, []):
            if _m.sektor in dozwolone_sektory and oblicz_dystans(baza_lat, baza_lng, _m.lat, _m.lng) <= promien_petli:
                _miast_w_dozwolonych += 1
    if _miast_w_dozwolonych < 12:
        dozwolone_sektory = SEKTORY_KOLEJNOSC[:]   # odblokuj wszystkie kierunki

    # ADAPTACYJNY COOLDOWN — ile dni miasto "odpoczywa" zanim może się powtórzyć.
    # Domyślnie 35 dni (żadnych powtórek w miesiącu). Ale gdy dostępnych miast
    # jest za mało, by wypełnić wszystkie dni robocze bez powtórek, skracamy go —
    # inaczej silnik wygenerowałby za mało dni i naciągał kilometry, by trafić
    # w kwotę. Realnie merchandiser i tak odwiedza te same sklepy cyklicznie.
    _pula_szac = 0
    for _w in [woj] + sasiednie:
        for _m in baza.get(_w, []):
            if oblicz_dystans(baza_lat, baza_lng, _m.lat, _m.lng) <= promien_petli:
                _pula_szac += 1
    _dni_do_pokrycia = max(1, len(dni_robocze))
    # ile miast potrzeba: ~6 na dzień. Jeśli pula wystarcza na wszystkie dni → 35.
    if _pula_szac >= _dni_do_pokrycia * 6:
        cooldown_dni = 35
    elif _pula_szac >= _dni_do_pokrycia * 3:
        cooldown_dni = 14
    else:
        cooldown_dni = 8

    # Ile dni wyjazdowych chcemy w miesiącu? Zależnie od kwoty — ale rozłożonych
    # RÓWNOMIERNIE na cały miesiąc, nie skupionych na początku. Celujemy w tyle
    # dni, ile realnie potrzeba, by trasy pokryły kwotę bez naciągania km.
    # === ILE DNI? Zasada: rozpisuj REALNE trasy, aż wyczerpiesz budżet. ===
    # Typowy dzień daje ~115 km realnych (pętla "po drodze" w granicach 8h).
    # Bierzemy tyle dni, ile potrzeba, by realnymi trasami pokryć kwotę — bez
    # napompowywania odcinków. Górny limit to liczba dni roboczych w miesiącu.
    # Szacunek km/dzień celowo OSTROŻNY (90, nie 115): realne trasy drogowe z
    # internetu (OSRM) są dłuższe niż linia prosta, więc dzień pokrywa mniej
    # budżetu, niż wynikałoby z prostego szacunku. Niższa wartość => więcej dni
    # => drobniejsze, bardziej realne trasy i ZAPAS na pełne pokrycie kwoty.
    _km_na_dzien_szac = 90.0
    _dni_z_km = math.ceil(cel_calkowity_dystans / _km_na_dzien_szac)
    # żaden dzień nie może przekroczyć limitu delegacji (~587 zł)
    _dni_z_limitu = math.ceil((kwota_calkowita / (MAX_KWOTA_DELEGACJI * 0.85)))
    _dni_potrzeba = max(3, _dni_z_km, _dni_z_limitu)
    # GÓRNY limit liczby dni: dzień musi udźwignąć co najmniej ~110 zł realnie,
    # inaczej przy zbyt wielu dniach nie da się ścisnąć tras do budżetu (dolny
    # próg mnożnika 0.5) i suma WYSZŁABY ZA WYSOKA. Ten limit trzyma dzienną
    # kwotę w realnym przedziale i chroni przed przekroczeniem budżetu.
    _dni_max_dla_kwoty = max(3, math.floor(kwota_calkowita / 110.0))
    _dni_potrzeba = min(_dni_potrzeba, _dni_max_dla_kwoty)
    # Jeśli potrzeba więcej dni niż jest w miesiącu — bierzemy WSZYSTKIE dostępne
    # i sygnalizujemy, że kwota może się nie zmieścić (komunikat po generacji).
    _brak_dni_na_kwote = _dni_potrzeba > len(dni_robocze)
    _dni_potrzeba = min(len(dni_robocze), _dni_potrzeba)
    # Wybierz dni rozłożone równomiernie przez cały miesiąc (nie po kolei z początku).
    if _dni_potrzeba >= len(dni_robocze):
        dni_wybrane = list(range(len(dni_robocze)))
    else:
        krok = len(dni_robocze) / _dni_potrzeba
        dni_wybrane = sorted(set(int(i * krok) for i in range(_dni_potrzeba)))
        _i = 0
        while len(dni_wybrane) < _dni_potrzeba and _i < len(dni_robocze):
            if _i not in dni_wybrane: dni_wybrane.append(_i)
            _i += 1
        dni_wybrane = sorted(dni_wybrane)
    dni_wybrane_set = set(dni_wybrane)

    # Budujemy dni, celując w _dni_potrzeba UDANYCH dni. Zaczynamy od równomiernie
    # rozłożonego zestawu (dni_wybrane_set) dla ładnego pokrycia miesiąca, ale
    # jeśli któryś dzień się nie zbuduje, sięgamy po KOLEJNE wolne daty — dzięki
    # temu zawsze powstaje tyle dni, ile trzeba na budżet (o ile starczy dat).
    # Najpierw preferowane daty (równomierny rozkład), potem cała reszta jako zapas.
    kolejnosc_dni = [d for d in dni_wybrane] + [i for i in range(len(dni_robocze)) if i not in dni_wybrane_set]

    for _poz, numer_dnia in enumerate(kolejnosc_dni):
        if len(finalne_dni) >= _dni_potrzeba:
            break            # zebraliśmy tyle dni, ile potrzeba na budżet
        data = dni_robocze[numer_dnia]
        if postep_cb: postep_cb(f"Klastrowanie GPS (Dzień {len(finalne_dni)+1}/{_dni_potrzeba})...", 0.30 + (_poz / max(len(kolejnosc_dni),1)) * 0.40)

        # W rzadkich rejonach (przygranicze) stare filtry "ucieczki od stolicy"
        # i "obcej aglomeracji" są zbyt restrykcyjne i zostawiają 0 miast.
        # Wykrywamy taki przypadek i wtedy je pomijamy.
        _rzadki_rejon = (promien_petli > 70)

        def dostepne_miasta(pula, wymus_ucieczke=True, spr_cooldown=True):
            dostepne = []
            for m in pula:
                if m.n.lower() == baza_nazwa.lower() or m.sieci < MIN_SIECI: continue
                if wymus_ucieczke and not _rzadki_rejon and not odsuwa_sie_od_stolicy(baza_lat, baza_lng, m.lat, m.lng, woj): continue
                if not _rzadki_rejon and miasto_w_obcej_aglomeracji(baza_lat, baza_lng, m.lat, m.lng): continue
                if spr_cooldown and m.n in ostatnie_uzycie and (data - ostatnie_uzycie[m.n]).days < cooldown_dni: continue

                d = oblicz_dystans(baza_lat, baza_lng, m.lat, m.lng) * TEST_MNOZNIK_TRASY
                # limit: dojazd do miasta i z powrotem musi się mieścić w dniu
                # (z zapasem na postoje). W rzadkim rejonie dajemy większy zapas.
                zapas = 60 if _rzadki_rejon else 90
                if (d * 2 / SREDNIA_PREDKOSC) * 60 + zapas <= LIMIT_CZASU_MINUTY:
                    if _rzadki_rejon or not przecina_aglomeracje(baza_lat, baza_lng, m.lat, m.lng, baza_lat, baza_lng):
                        dostepne.append(m)
            return dostepne

        czy_wlasne = (rng.random() <= PROCENT_WLASNE)
        def pobierz_kandydatow(ucieczka, spr_cd):
            kw = usun_duplikaty_nazw(dostepne_miasta(baza.get(woj, []), ucieczka, spr_cd))
            ko = usun_duplikaty_nazw([m for w in sasiednie for m in dostepne_miasta(baza.get(w, []), ucieczka, spr_cd)])
            if czy_wlasne and kw: return kw
            elif not czy_wlasne and ko: return ko
            elif kw: return kw
            elif ko: return ko
            return []

        kandydaci = pobierz_kandydatow(True, True)
        if not kandydaci: kandydaci = pobierz_kandydatow(False, True)
        # Gdy cooldown wyczerpał pulę miast (przy wielu dniach to naturalne),
        # NIE pomijamy dnia — próbujemy bez cooldownu. Realny merchandiser i tak
        # odwiedza te same sklepy cyklicznie. Priorytetem jest rozpisanie
        # WSZYSTKICH potrzebnych dni, żeby pokryć budżet realnymi trasami.
        if not kandydaci: kandydaci = pobierz_kandydatow(True, False)
        if not kandydaci: kandydaci = pobierz_kandydatow(False, False)
        if not kandydaci: continue

        # DETERMINIZM: sortujemy kandydatów po nazwie, ZANIM losujemy z nich
        # trasę. Bez tego kolejność zależałaby od losowego haszowania tekstu
        # w Pythonie (inne przy każdym starcie programu), więc ten sam pracownik
        # dostawałby różne trasy przy każdym uruchomieniu — a bywało, że w
        # niektórych kolejnościach budowało się dużo mniej dni.
        kandydaci = sorted(kandydaci, key=lambda m: (m.n, round(m.lat, 5), round(m.lng, 5)))

        # --- TYP DNIA (różnorodność, jak w realnej pracy merchandisera) ---
        # Trzy typy losowane w miesiącu dają naturalną zmienność zamiast
        # jednakowych dni. Deterministyczne wg PESEL+dzień, więc ten sam
        # pracownik zawsze dostaje ten sam plan.
        _typy = ["gniazdo", "po_drodze", "po_drodze", "po_drodze", "daleki", "po_drodze"]
        typ_dnia = _typy[(seed_val + numer_dnia * 3) % len(_typy)]
        if promien_petli > 70:
            typ_dnia = "po_drodze"    # w rzadkich rejonach trzymamy się bezpiecznego schematu

        if typ_dnia == "gniazdo":
            # ciasna pętla blisko bazy, DUŻO punktów, krótkie skoki
            start_min, start_max = 15.0, 45.0
            skok_dnia = 24.0
            cele_min, cele_max = 5, 7
        elif typ_dnia == "daleki":
            # dojazd w dalszy rejon (dużo km z dojazdu), potem solidna pętla
            start_min, start_max = min(45.0, promien_petli*0.6), promien_petli
            skok_dnia = min(max_skok_bazowy(promien_petli), 40.0)
            cele_min, cele_max = 5, 7
        else:  # po_drodze
            start_min, start_max = 25.0, promien_petli * 0.75
            skok_dnia = max_skok_bazowy(promien_petli)
            cele_min, cele_max = 5, 7

        sektor_dnia = dozwolone_sektory[(seed_val + numer_dnia * 7) % len(dozwolone_sektory)]
        idx_sek = SEKTORY_KOLEJNOSC.index(sektor_dnia)
        aktywne_sektory = [SEKTORY_KOLEJNOSC[(idx_sek - 1) % 8], sektor_dnia, SEKTORY_KOLEJNOSC[(idx_sek + 1) % 8]]

        # Punkt startowy dnia: miasto w wybranym sektorze, w pierścieniu zależnym
        # od TYPU dnia (gniazdo blisko, daleki wyjazd dalej).
        srodek_startu = (start_min + start_max) / 2.0
        wagi = []
        for m in kandydaci:
            d = oblicz_dystans(baza_lat, baza_lng, m.lat, m.lng)
            if d > promien_petli or d < start_min - 8:
                waga = 0.001
            else:
                waga = 1000.0 / (abs(d - srodek_startu) + 8.0)
            waga *= 50.0 if m.typ == 'gmina' else 1.0
            if m.sektor == sektor_dnia: waga *= 1000.0
            elif m.sektor in aktywne_sektory: waga *= 12.0
            else: waga *= 0.05
            wagi.append(waga)

        udana_trasa = False
        ostateczne_etapy_dnia = []

        for proba in range(15):
            punkt_startowy = rng.choices(kandydaci, weights=wagi, k=1)[0]
            ile_celow = rng.randint(cele_min, cele_max)

            # --- BUDOWA ŁAŃCUCHA "PO DRODZE" ---
            # Zaczynamy od punktu startowego, a każdy kolejny przystanek to
            # NAJBLIŻSZY nieodwiedzony sąsiad poprzedniego. Maksymalny skok
            # zależy od TYPU dnia (gniazdo = ciasno, po drodze = luźniej).
            wybrane = [punkt_startowy]
            uzyte_nazwy = {punkt_startowy.n}
            max_skok = skok_dnia if promien_petli <= 70 else min(skok_dnia * (promien_petli/70.0), 60.0)
            while len(wybrane) < ile_celow:
                ostatni = wybrane[-1]
                najlepszy = None
                najlepszy_d = max_skok
                for m in kandydaci:
                    if m.n in uzyte_nazwy: continue
                    if m.typ != 'gmina' and len(wybrane) > 1: continue
                    d_skok = oblicz_dystans(ostatni.lat, ostatni.lng, m.lat, m.lng)
                    if d_skok < 4.0: continue                    # to praktycznie ten sam punkt
                    if d_skok >= najlepszy_d: continue           # dalej niż obecny najlepszy
                    if oblicz_dystans(baza_lat, baza_lng, m.lat, m.lng) > promien_petli: continue
                    if przecina_aglomeracje(ostatni.lat, ostatni.lng, m.lat, m.lng, baza_lat, baza_lng): continue
                    najlepszy = m; najlepszy_d = d_skok
                if najlepszy is None: break
                wybrane.append(najlepszy); uzyte_nazwy.add(najlepszy.n)

            # Optymalna kolejność zwiedzania (pętla z bazy i z powrotem)
            wybrane_tsp = optymalizuj_tsp(wybrane, baza_lat, baza_lng)

            # --- WERYFIKACJA CZASOWA (8h = posiłek + jazda + postoje + POWRÓT) ---
            # Przerwa na posiłek jest wliczona w budżet z góry. Sprawdzamy, czy
            # CAŁY dzień z powrotem do bazy mieści się w limicie — jeśli nie,
            # obcinamy ostatnie punkty, aż się zmieści. Powrót do domu NIGDY nie
            # może wypchnąć dnia ponad 8h.
            czas_akt_test = PRZERWA_JEDZENIE_MIN
            etapy_test = []
            obecny_lat, obecny_lng, obecna_nazwa = baza_lat, baza_lng, baza_nazwa

            for cel in wybrane_tsp:
                d_line = oblicz_dystans(obecny_lat, obecny_lng, cel.lat, cel.lng)
                t_drive = (d_line * TEST_MNOZNIK_TRASY / SREDNIA_PREDKOSC) * 60
                t_shop = rng.randint(POSTOJ_MIN_MIN, POSTOJ_MAX_MIN)
                # powrót Z TEGO punktu prosto do bazy
                d_ret_line = oblicz_dystans(cel.lat, cel.lng, baza_lat, baza_lng)
                t_ret = (d_ret_line * TEST_MNOZNIK_TRASY / SREDNIA_PREDKOSC) * 60

                # pełny czas gdyby to był ostatni punkt: dojazd + postój + powrót do bazy
                if czas_akt_test + t_drive + t_shop + t_ret > LIMIT_CZASU_MINUTY:
                    break

                etapy_test.append(RawEtap(
                    skad=obecna_nazwa, dokad=cel.n, data_str=data.strftime("%d.%m.%Y") + "r",
                    d_line=d_line, czas_w_sklepie=t_shop, dokad_woj=cel.woj,
                    skad_lat=obecny_lat, skad_lng=obecny_lng, dokad_lat=cel.lat, dokad_lng=cel.lng
                ))
                czas_akt_test += t_drive + t_shop
                obecny_lat, obecny_lng, obecna_nazwa = cel.lat, cel.lng, cel.n

            # Dzień ważny tylko z min. 3 postojami "po drodze".
            if len(etapy_test) < 3:
                continue

            # powrót do domu — dodajemy i liczymy PEŁNY czas dnia
            d_p_line = oblicz_dystans(obecny_lat, obecny_lng, baza_lat, baza_lng)
            t_powrot = (d_p_line * TEST_MNOZNIK_TRASY / SREDNIA_PREDKOSC) * 60
            czas_pelny = czas_akt_test + t_powrot

            # TWARDY BEZPIECZNIK: jeśli mimo wszystko przekracza 8h, obcinamy
            # ostatnie punkty aż cały dzień (z powrotem) zmieści się w limicie.
            while czas_pelny > LIMIT_CZASU_MINUTY and len(etapy_test) > 3:
                usuniety = etapy_test.pop()
                # przelicz czas: cofamy do punktu przed usuniętym
                if etapy_test:
                    ost = etapy_test[-1]
                    obecny_lat, obecny_lng, obecna_nazwa = ost.dokad_lat, ost.dokad_lng, ost.dokad
                else:
                    obecny_lat, obecny_lng, obecna_nazwa = baza_lat, baza_lng, baza_nazwa
                # odejmij czas usuniętego etapu (dojazd do niego + jego postój)
                czas_akt_test -= (usuniety.d_line * TEST_MNOZNIK_TRASY / SREDNIA_PREDKOSC) * 60 + (usuniety.czas_w_sklepie or 0)
                d_p_line = oblicz_dystans(obecny_lat, obecny_lng, baza_lat, baza_lng)
                t_powrot = (d_p_line * TEST_MNOZNIK_TRASY / SREDNIA_PREDKOSC) * 60
                czas_pelny = czas_akt_test + t_powrot

            # po obcięciu dzień nadal musi mieć min. 3 postoje — inaczej pomijamy
            if len(etapy_test) < 3:
                continue
            # ostateczna kontrola — jeśli wciąż ponad limit, ten dzień odrzucamy
            if czas_pelny > LIMIT_CZASU_MINUTY:
                continue

            etapy_test.append(RawEtap(
                skad=obecna_nazwa, dokad=baza_nazwa, data_str=data.strftime("%d.%m.%Y") + "r",
                d_line=d_p_line, czas_w_sklepie=0, dokad_woj="",
                skad_lat=obecny_lat, skad_lng=obecny_lng, dokad_lat=baza_lat, dokad_lng=baza_lng
            ))

            udana_trasa = True
            ostateczne_etapy_dnia = etapy_test
            break

        if not udana_trasa: continue  

        for e in ostateczne_etapy_dnia[:-1]:
            ostatnie_uzycie[e.dokad] = data
            aktualny_dystans_linii += e.d_line
            
        aktualny_dystans_linii += ostateczne_etapy_dnia[-1].d_line
        finalne_dni.append(DzienTrasy(data=data, etapy_surowe=ostateczne_etapy_dnia))

    # Dni mogły powstać w innej kolejności niż kalendarzowa (gdy sięgaliśmy po
    # zapasowe daty) — porządkujemy chronologicznie.
    finalne_dni.sort(key=lambda d: d.data)

    if postep_cb: postep_cb("Skalowanie wektorów do budżetu...", 0.75)

    wszystkie_surowe = [e for dzien in finalne_dni for e in dzien.etapy_surowe]
    if not wszystkie_surowe: return []

    # --- REALNE ODLEGŁOŚCI DROGOWE ---
    # Do tej pory d_line to odległość w linii prostej (szybka, do doboru miast).
    # Teraz każdemu finalnemu odcinkowi nadajemy realny dystans DROGOWY z OSRM
    # (z cache + fallbackiem offline). Dzięki temu skalowanie kwot i wyliczone
    # kilometry odpowiadają faktycznym trasom po drogach, a nie liniom prostym.
    if postep_cb: postep_cb("Wyznaczanie realnych tras drogowych...", 0.78)
    for e in wszystkie_surowe:
        try:
            km_droga = dystans_drogowy(e.skad_lat, e.skad_lng, e.dokad_lat, e.dokad_lng)
            if km_droga and km_droga > 0:
                e.d_line = km_droga
        except Exception:
            pass  # zostaje linia prosta — bezpieczny fallback
    _zapisz_road_cache()

    suma_linii = max(sum(e.d_line for e in wszystkie_surowe), 1.0)
    wymagany_dystans_calkowity = kwota_calkowita / stawka
    rzeczywisty_mnoznik = wymagany_dystans_calkowity / suma_linii

    # --- TWARDY LIMIT 8h ---
    # Skalowanie km do kwoty nie może sprawić, że jakikolwiek dzień przekroczy
    # 8h (jazda + postoje + 30 min przerwy). Wyliczamy maksymalny dopuszczalny
    # mnożnik dla każdego dnia i bierzemy najmniejszy — dzięki temu ŻADEN dzień
    # nie wyjdzie ponad 8h, nawet kosztem tego, że kwota nie zostanie w całości
    # "wyjeżdżona" (nadwyżkę i tak przycinamy do realiów).
    def _czas_dnia_przy_mnozniku(dzien, mn):
        # przerwa na jedzenie wliczona w dzień pracy
        czas = PRZERWA_JEDZENIE_MIN
        for e in dzien.etapy_surowe:
            czas += (e.d_line * mn / SREDNIA_PREDKOSC) * 60      # jazda po realnym dystansie
            czas += (e.czas_w_sklepie or 0)                       # postój w miejscowości
        return czas

    max_mnoznik_globalny = rzeczywisty_mnoznik
    for dzien in finalne_dni:
        # jaki mnożnik sprawia, że ten dzień dobija DOKŁADNIE do 8h?
        baza_km = sum(e.d_line for e in dzien.etapy_surowe)
        postoje = sum((e.czas_w_sklepie or 0) for e in dzien.etapy_surowe) + PRZERWA_JEDZENIE_MIN
        dostepne_min_na_jazde = LIMIT_CZASU_MINUTY - postoje
        if dostepne_min_na_jazde <= 0 or baza_km <= 0:
            continue
        max_km_dnia = (dostepne_min_na_jazde / 60) * SREDNIA_PREDKOSC
        mn_dnia = max_km_dnia / baza_km
        max_mnoznik_globalny = min(max_mnoznik_globalny, mn_dnia)

    # mnożnik finalny: nie większy niż limit 8h. Dolny próg 0.3 (nie 0.5) pozwala
    # ścisnąć trasy dla małych kwot w rzadkich rejonach, gdzie minimalne 3 dni i
    # tak generują sporo kilometrów — bez tego suma wychodziłaby za wysoka.
    rzeczywisty_mnoznik = max(0.3, min(rzeczywisty_mnoznik, max_mnoznik_globalny))

    wszystkie_surowe = [e for dzien in finalne_dni for e in dzien.etapy_surowe]
    suma_linii = max(sum(e.d_line for e in wszystkie_surowe), 1.0)
    rzeczywisty_mnoznik = max(0.3, min(wymagany_dystans_calkowity / suma_linii, max_mnoznik_globalny))

    # Koszt każdego etapu = jego REALNY dystans (km × mnożnik) × stawka.
    # Nic nie dopisujemy. Dzięki dobraniu mnożnika suma wychodzi z natury
    # bardzo blisko celu (różnica to zwykle kilka GROSZY z zaokrągleń).
    for e in wszystkie_surowe:
        e.dystans_rzeczywisty = e.d_line * rzeczywisty_mnoznik
        e.kwota = max(round(e.dystans_rzeczywisty * stawka, 2), 0.0)
        e.czas_jazdy_minuty = (e.dystans_rzeczywisty / SREDNIA_PREDKOSC) * 60

    # Jedyna korekta: domknięcie DROBNYCH groszy z zaokrągleń, rozłożone po
    # 1 groszu na kolejne etapy. To NIE jest dopisywanie kwot do tras — to
    # tylko wyrównanie zaokrągleń (rzędu kilku groszy). Gdyby różnica była
    # duża (co nie powinno się zdarzyć), zostawiamy trasy realne — lepiej
    # kwota bliska niż sztucznie napompowany odcinek.
    reszta = round(kwota_calkowita - sum(e.kwota for e in wszystkie_surowe), 2)

    # BEZPIECZNIK PRZEKROCZENIA: jeśli suma WYSZŁA PONAD budżet (zdarza się dla
    # bardzo małych kwot w rzadkich rejonach, gdzie minimum dni i tak generuje
    # sporo km), delikatnie skalujemy WSZYSTKIE etapy w dół do budżetu. Delegacja
    # NIGDY nie może przekroczyć wpisanej kwoty.
    suma_biezaca = sum(e.kwota for e in wszystkie_surowe)
    if suma_biezaca > kwota_calkowita + 0.01 and suma_biezaca > 0:
        skala = kwota_calkowita / suma_biezaca
        for e in wszystkie_surowe:
            e.dystans_rzeczywisty *= skala
            e.kwota = max(round(e.kwota * skala, 2), 0.0)
            e.czas_jazdy_minuty = (e.dystans_rzeczywisty / SREDNIA_PREDKOSC) * 60
        reszta = round(kwota_calkowita - sum(e.kwota for e in wszystkie_surowe), 2)

    LIMIT_KOREKTY_GROSZY = max(len(wszystkie_surowe), 20)  # max ~1 gr/etap
    if 0.01 <= abs(reszta) <= LIMIT_KOREKTY_GROSZY * 0.01 and wszystkie_surowe:
        krok = 0.01 if reszta > 0 else -0.01
        idx = 0
        while abs(reszta) >= 0.01:
            e = wszystkie_surowe[idx % len(wszystkie_surowe)]
            e.kwota = round(e.kwota + krok, 2)
            reszta = round(reszta - krok, 2); idx += 1

    for dzien in finalne_dni:
        # Godzina wyjazdu zależy od TRYBU PRACY:
        #  tygodniowy — rano, między 7:00 a 8:00 (jak dotąd),
        #  wieczorny  — po pracy, między 16:15 a 16:45, tak aby ostatnia
        #               placówka zamknęła się przed 22:00 (patrz limit dnia).
        if TRYB_PRACY == "wieczory":
            czas_akt = rng.randint(16 * 60 + 15, 16 * 60 + 45)
        else:
            czas_akt = rng.randint(7 * 60, 8 * 60)
        ile_etapow = len(dzien.etapy_surowe)
        polowa = ile_etapow // 2
        for idx, e in enumerate(dzien.etapy_surowe):
            godz_wyj = f"{int(czas_akt//60)%24:02d}:{int(czas_akt%60):02d}"
            czas_akt += e.czas_jazdy_minuty
            godz_przyj = f"{int(czas_akt//60)%24:02d}:{int(czas_akt%60):02d}"
            dzien.etapy.append(Etap(skad=e.skad, dokad=e.dokad, data=e.data_str, godz_wyj=godz_wyj, godz_przyj=godz_przyj, kwota=e.kwota, dokad_woj=e.dokad_woj))
            czas_akt += e.czas_w_sklepie
            # przerwa na jedzenie mniej więcej w połowie dnia (wliczona w 8h)
            if idx == polowa:
                czas_akt += PRZERWA_JEDZENIE_MIN

    # Informacja dla warstwy UI: czy realnymi trasami udało się pokryć kwotę.
    # Jeśli mimo wykorzystania wszystkich możliwych dni suma odbiega od celu
    # o więcej niż ~2%, sygnalizujemy to (użytkownik dostanie komunikat, że
    # kwota jest za wysoka na ten miesiąc/rejon). NIE pompujemy tras na siłę.
    suma_final = sum(d.suma for d in finalne_dni)
    finalne_dni = ListaTras(finalne_dni)
    finalne_dni.kwota_docelowa = kwota_calkowita
    finalne_dni.kwota_osiagnieta = round(suma_final, 2)
    finalne_dni.kwota_niepelna = (kwota_calkowita - suma_final) > max(kwota_calkowita * 0.02, 20.0)
    finalne_dni.brak_dni = _brak_dni_na_kwote
    return finalne_dni

def generuj_mape_html(finalne_dni: List[DzienTrasy], pracownik: DanePracownika, miesiac_slownie: str, rok: int, folder: str, is_dark: bool):
    bg_color = "#0B1320" if is_dark else "#F1F5F9"
    text_color = "#F8FAFC" if is_dark else "#0F172A"
    card_bg = "rgba(16, 25, 43, 0.8)" if is_dark else "#FFFFFF"
    border_c = "rgba(0, 240, 255, 0.3)" if is_dark else "rgba(13, 148, 136, 0.3)"
    title_c = "#00F0FF" if is_dark else "#0D9488"
    text_muted = "#94A3B8" if is_dark else "#475569"
    btn_bg = "linear-gradient(to right, #00F0FF, #0093E9)" if is_dark else "#10B981"
    btn_txt = "#000000" if is_dark else "#FFFFFF"
    
    html_content = f"""
    <!DOCTYPE html>
    <html lang="pl">
    <head>
        <meta charset="UTF-8">
        <title>Trasy Delegacji - {pracownik.imie}</title>
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: {bg_color}; color: {text_color}; padding: 20px; }}
            .container {{ max-width: 800px; margin: 0 auto; }}
            .header {{ text-align: center; margin-bottom: 30px; }}
            .header h1 {{ color: {title_c}; margin-bottom: 5px; }}
            .day-card {{ background-color: {card_bg}; border: 1px solid {border_c}; border-radius: 8px; padding: 15px 20px; margin-bottom: 15px; box-shadow: 0 4px 6px rgba(0,0,0,0.5); }}
            .day-title {{ font-size: 16px; font-weight: bold; color: {text_color}; margin-bottom: 10px; }}
            .route {{ font-size: 14px; color: {text_muted}; margin-bottom: 15px; line-height: 1.5; }}
            .btn {{ display: inline-block; background: {btn_bg}; color: {btn_txt}; padding: 8px 16px; border-radius: 6px; text-decoration: none; font-weight: bold; font-size: 13px; transition: opacity 0.2s; }}
            .btn:hover {{ opacity: 0.8; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>Wykaz Tras Geograficznych - {pracownik.imie}</h1>
                <p>Miesiąc: {miesiac_slownie} {rok}</p>
            </div>
    """
    # Baza (dom pracownika) to START i KONIEC każdej trasy. Przekazujemy pełny
    # adres pracownika — jest dokładny (ulica+kod), więc Google nie podstawi
    # sąsiedniego budynku, a jednocześnie czytelnie pokazuje punkt startu/powrotu.
    # Fallback na współrzędne, gdyby adres był pusty.
    adres_bazy = (pracownik.adres or "").strip()
    if adres_bazy:
        baza_punkt = adres_bazy
    else:
        baza_punkt = f"{pracownik.baza_lat:.6f},{pracownik.baza_lng:.6f}"
    for dzien in finalne_dni:
        if not dzien.etapy: continue
        # START = baza (dom). Etykieta pokazuje miasto, ale URL używa pełnego adresu.
        punkty_wyswietlane = [f"🏠 {pracownik.baza_miasto}"]
        punkty_url = [urllib.parse.quote(baza_punkt)]
        for e in dzien.etapy:
            if e.dokad == pracownik.baza_miasto:
                punkty_wyswietlane.append(f"🏠 {pracownik.baza_miasto}")
                punkty_url.append(urllib.parse.quote(baza_punkt))
            else:
                punkty_wyswietlane.append(f"{e.dokad}")
                punkty_url.append(urllib.parse.quote(f"{e.dokad}, {e.dokad_woj}, Polska"))
        html_content += f"""
            <div class="day-card">
                <div class="day-title">Data: {dzien.data.strftime("%d.%m.%Y")}</div>
                <div class="route"><strong>Trasa:</strong> {" ➔ ".join(punkty_wyswietlane)}</div>
                <a href="https://www.google.com/maps/dir/{"/".join(punkty_url)}" target="_blank" class="btn">🗺️ Otwórz w Google Maps</a>
            </div>
        """
    html_content += "</div></body></html>"
    
    html_path = os.path.join(folder, "Trasy_Mapa.html")
    try:
        with open(html_path, "w", encoding="utf-8") as f: f.write(html_content)
    except OSError:
        raise ValueError("Plik Trasy_Mapa.html jest otwarty w przeglądarce! Zamknij go przed wygenerowaniem.")

def _znajdz_font(nazwa_win: str) -> Optional[str]:
    """Szuka czcionki TTF do PDF-ów — Windows, macOS i Linux.
    nazwa_win to windowsowa nazwa pliku (arial.ttf / arialbd.ttf);
    dla pozostałych systemów mapujemy ją na tamtejsze odpowiedniki
    (macOS: systemowy Arial; Linux: Liberation Sans lub DejaVu Sans,
    które mają pełny zestaw polskich znaków)."""
    bold = "bd" in nazwa_win.lower()
    kandydaci = [
        Path(r"C:\Windows\Fonts") / nazwa_win,
        Path("/System/Library/Fonts/Supplemental") / ("Arial Bold.ttf" if bold else "Arial.ttf"),
        Path("/Library/Fonts") / nazwa_win,
        Path(os.path.expanduser("~/Library/Fonts")) / nazwa_win,
        Path("/usr/share/fonts/truetype/msttcorefonts") / nazwa_win,
        Path("/usr/share/fonts/truetype/liberation") / ("LiberationSans-Bold.ttf" if bold else "LiberationSans-Regular.ttf"),
        Path("/usr/share/fonts/truetype/dejavu") / ("DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf"),
        Path("/usr/share/fonts/TTF") / ("DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf"),
    ]
    for p in kandydaci:
        if p.exists(): return str(p)
    return None

class PDFReport(FPDF):
    def __init__(self):
        super().__init__()
        fr, fb = _znajdz_font("arial.ttf"), _znajdz_font("arialbd.ttf")
        if fr and fb:
            try: self.add_font("Arial", "", fr); self.add_font("Arial", "B", fb)
            except: pass
    def header(self): pass

def _podziel_na_dokumenty(dni: List[DzienTrasy]) -> List[List[DzienTrasy]]:
    docs, obecny, koszt, etapy = [], [], 0.0, 0
    for d in dni:
        e = len(d.etapy)
        if (koszt + d.suma > MAX_KWOTA_DELEGACJI or etapy + e > 16) and obecny:
            docs.append(obecny); obecny, koszt, etapy = [d], d.suma, e
        else: obecny.append(d); koszt += d.suma; etapy += e
    if obecny: docs.append(obecny)
    return docs

def generuj_pdfy(finalne_dni: List[DzienTrasy], pracownik: DanePracownika, miesiac: int, rok: int, folder: str, stawka: float = None, postep_callback=None) -> List[dict]:
    if stawka is None: stawka = STAWKA_ZA_KM
    os.makedirs(folder, exist_ok=True)
    ms = MIESIACE_PL[miesiac - 1]
    # Gwarancja chronologii: dni zawsze rosnąco wg daty, ZANIM podzielimy je na
    # dokumenty. Dzięki temu daty w obrębie każdego PDF-a idą po kolei.
    finalne_dni = sorted(finalne_dni, key=lambda d: d.data)
    docs = _podziel_na_dokumenty(finalne_dni)
    podsumowanie = []
    firma_nazwa = "PMT HR SP. Z O.O. SP. K." if pracownik.stanowisko == "KR" else "PMT SERVICES SP. Z O.O. SP. K."
    nip_dlugi   = "NIP 779-24-86-133" if pracownik.stanowisko == "KR" else "NIP 779-24-86-127"
    nip_krotki  = "NIP 779-248-61-33" if pracownik.stanowisko == "KR" else "NIP 779-248-61-27"

    total = len(docs)
    for i, doc_dni in enumerate(docs, 1):
        if postep_callback: postep_callback(f"Renderowanie pliku PDF ({i}/{total})...", 0.75 + (i/total)*0.20)
        suma_doc = sum(d.suma for d in doc_dni)
        pdf = PDFReport(); pdf.add_page(); pdf.set_margins(10, 8, 10)
        pdf.set_font("Arial",'B',10); pdf.cell(0,4,firma_nazwa,new_x="LMARGIN",new_y="NEXT")
        pdf.set_font("Arial",'',9); pdf.cell(0,4,"ul. Ptasia 10, 60-319 Poznań",new_x="LMARGIN",new_y="NEXT")
        pdf.cell(0,4,nip_dlugi,new_x="LMARGIN",new_y="NEXT"); pdf.ln(3)
        pdf.set_font("Arial",'B',13); pdf.cell(100,6,"Rozliczenie delegacji",border=0)
        pdf.set_font("Arial",'B',10); pdf.cell(90,6,"6",border=0,new_x="LMARGIN",new_y="NEXT",align='R')
        pdf.set_font("Arial",'',8); pdf.set_xy(pdf.get_x()+130, pdf.get_y()-2); pdf.cell(60,3,"Centrum",border=0,new_x="LMARGIN",new_y="NEXT",align='R'); pdf.ln(3)
        pdf.set_font("Arial",'B',9); pdf.cell(95,5,f"{pracownik.imie} {pracownik.pesel}",border=1,align='C')
        pdf.cell(95,5,pracownik.stanowisko,border=1,new_x="LMARGIN",new_y="NEXT",align='C')
        pdf.set_font("Arial",'',7); pdf.cell(95,3,"imię i nazwisko oraz PESEL",border=0,align='C'); pdf.cell(95,3,"stanowisko",border=0,new_x="LMARGIN",new_y="NEXT",align='C')
        pdf.set_font("Arial",'B',9); pdf.cell(95,5,pracownik.adres,border=1,align='C'); pdf.cell(95,5,"Małgorzata Murawska",border=1,new_x="LMARGIN",new_y="NEXT",align='C')
        pdf.set_font("Arial",'',7); pdf.cell(95,3,"adres zamieszkania",border=0,align='C'); pdf.cell(95,3,"przełożony",border=0,new_x="LMARGIN",new_y="NEXT",align='C')
        pdf.set_font("Arial",'B',8); pdf.cell(95,5,"projekt: Biedronka, Dino, Eurocash, Społem, Stokrotka, Żabka",border=1,new_x="LMARGIN",new_y="NEXT",align='C')
        pdf.set_font("Arial",'',7); pdf.cell(95,3,"cel wyjazdu - Projekt",border=0,new_x="LMARGIN",new_y="NEXT",align='C'); pdf.ln(3)
        pdf.set_font("Arial",'B',10); pdf.cell(0,5,"Rozliczenie kosztów podróży",new_x="LMARGIN",new_y="NEXT")
        pdf.set_font("Arial",'B',8); pdf.cell(65,4,"WYJAZD",border=1,align='C'); pdf.cell(65,4,"PRZYJAZD",border=1,align='C')
        y_s = pdf.get_y(); pdf.cell(35,8,"Środki lokomocji",border=1,align='C'); pdf.cell(25,8,"Koszty",border=1,new_x="LMARGIN",new_y="NEXT",align='C')
        pdf.set_y(y_s+4)
        for lbl in ["miejscowość","data","godz.","miejscowość","data","godz."]:
            pdf.cell(30 if lbl=="miejscowość" else (20 if lbl=="data" else 15),4,lbl,border=1,align='C')
        pdf.set_y(y_s+8); pdf.set_x(10); pdf.set_font("Arial",'',8)
        for dzien in doc_dni:
            for etap in dzien.etapy:
                if pdf.get_y() > 250: pdf.add_page(); pdf.set_margins(10,8,10)
                pdf.cell(30,4.5,etap.skad[:18]+"..." if len(etap.skad)>18 else etap.skad,border=1)
                pdf.cell(20,4.5,etap.data,border=1,align='C'); pdf.cell(15,4.5,etap.godz_wyj,border=1,align='C')
                pdf.cell(30,4.5,etap.dokad[:18]+"..." if len(etap.dokad)>18 else etap.dokad,border=1)
                pdf.cell(20,4.5,etap.data,border=1,align='C'); pdf.cell(15,4.5,etap.godz_przyj,border=1,align='C')
                pdf.cell(35,4.5,"samochód osobowy",border=1,align='C'); pdf.cell(25,4.5,f"{etap.kwota:.2f}",border=1,new_x="LMARGIN",new_y="NEXT",align='R')
                
        pdf.ln(3); pdf.set_font("Arial",'B',8)
        pdf.cell(25,4.5,"Rodzaj",border=1,align='C'); pdf.cell(140,4.5,"Opis kosztu",border=1,align='C'); pdf.cell(25,4.5,"Koszty",border=1,new_x="LMARGIN",new_y="NEXT",align='C')
        pdf.set_font("Arial",'',8)
        for r,o,k in [("(1) Przejazdy","Prywatny samochód: kilometrówka wyliczana wg wzoru: liczba km x stawka",f"{suma_doc:.2f}"),("(2) Noclegi","Koszt hotelu lub ryczałt 67,50 zł","-"),("(3) Diety","22,50 zł lub 45,00 zł","-"),("(4) Inne","Autostrada, parking, reprezentacja","-")]:
            pdf.cell(25,4.5,r,border=1); pdf.cell(140,4.5,o,border=1); pdf.cell(25,4.5,k,border=1,new_x="LMARGIN",new_y="NEXT",align='R')
        pdf.set_font("Arial",'B',8); pdf.cell(165,4.5,"Suma wydatków",border=1,align='R'); pdf.cell(25,4.5,f"{suma_doc:.2f}",border=1,new_x="LMARGIN",new_y="NEXT",align='R')
        pdf.ln(3); ys = pdf.get_y(); pdf.rect(10,ys,190,24)
        pdf.line(18,ys,18,ys+24); pdf.line(78,ys,78,ys+18); pdf.line(138,ys,138,ys+18); pdf.line(18,ys+18,200,ys+18); pdf.line(180,ys+18,180,ys+24)
        pdf.set_font("Arial",'B',6); yl = ys+1.5
        for letter in "AKCEPTACJA": pdf.set_xy(10,yl); pdf.cell(8,2,letter,align='C'); yl+=2
        pdf.set_font("Arial",'',7)
        for x,sz,ly,txt in [(18,60,ys+13,"podpis Księgowej"), (78,60,ys+13,"podpis Dyrektora Zarządzającego"), (138,62,ys+13,"podpis pracownika")]:
            pdf.line(x+5,ly,x+sz-5,ly); pdf.set_xy(x,ly); pdf.cell(sz,3,txt,align='C')
        pdf.set_font("Arial",'B',8); pdf.set_xy(20,ys+19); pdf.cell(160,3,"Liczba załączników"); pdf.set_xy(180,ys+19); pdf.cell(20,3,"1",align='C')
        
        pdf_path = os.path.join(folder, f"delegacja_{i:02d}_{pracownik.imie.replace(' ','_')}_{ms}_{rok}r.pdf")
        try: pdf.output(pdf_path)
        except OSError: raise ValueError(f"Plik PDF jest otwarty w innym programie!\nZamknij: delegacja_{i:02d}...pdf")
        podsumowanie.append({'lp':i,'dokument':f"delegacja {i:02d} {pracownik.imie}",'opis':"Koszty dojazdów",'kwota':suma_doc})

    pdf = PDFReport(); pdf.add_page(); pdf.set_margins(10,10,10)
    pdf.set_font("Arial",'B',10); pdf.cell(0,5,firma_nazwa,new_x="LMARGIN",new_y="NEXT")
    pdf.set_font("Arial",'',10); pdf.cell(0,5,"ul. Ptasia 10, 60-319 Poznań",new_x="LMARGIN",new_y="NEXT"); pdf.cell(0,5,nip_krotki,new_x="LMARGIN",new_y="NEXT"); pdf.ln(5)
    pdf.set_font("Arial",'B',12); pdf.cell(0,8,f"Rozliczenie wydatków za miesiąc: {ms} {rok}r.",new_x="LMARGIN",new_y="NEXT"); pdf.ln(2)
    for lbl,val in [("Imię i Nazwisko:",pracownik.imie),("Projekt/Stanowisko:",pracownik.stanowisko),("MENEDŻER:","Małgorzata Murawska"),("Liczba dokumentów:",str(len(podsumowanie)))]:
        pdf.set_font("Arial",'B',9); pdf.cell(50,6,lbl,border=0); pdf.set_font("Arial",'',9); pdf.cell(140,6,val,border=0,new_x="LMARGIN",new_y="NEXT")
    pdf.ln(5); pdf.set_font("Arial",'B',9)
    for sz,t in [(10,"Lp."),(80,"Dokument"),(70,"Opis"),(30,"Kwota brutto")]: pdf.cell(sz,8,t,border=1,align='C')
    pdf.ln(); suma_g = 0.0; pdf.set_font("Arial",'',9)
    for w in podsumowanie:
        pdf.cell(10,6,f"{w['lp']}.",border=1,align='C'); pdf.cell(80,6,w['dokument'],border=1); pdf.cell(70,6,w['opis'],border=1); pdf.cell(30,6,f"{w['kwota']:.2f}",border=1,new_x="LMARGIN",new_y="NEXT",align='R'); suma_g += w['kwota']
    for idx in range(len(podsumowanie)+1,14):
        pdf.cell(10,6,f"{idx}.",border=1,align='C'); pdf.cell(80,6,"",border=1); pdf.cell(70,6,"",border=1); pdf.cell(30,6,"",border=1,new_x="LMARGIN",new_y="NEXT")
    pdf.cell(160,6,"zaliczka:",border=0,align='R'); pdf.cell(30,6,"",border=1,new_x="LMARGIN",new_y="NEXT")
    pdf.set_font("Arial",'B',9); pdf.cell(160,6,"Suma wydatków global:",border=0,align='R'); pdf.cell(30,6,f"{suma_g:.2f}",border=1,new_x="LMARGIN",new_y="NEXT",align='R')
    
    pdf_path_summary = os.path.join(folder,f"rozliczenie_wydatków_{pracownik.imie.replace(' ','_')}_{ms}_{rok}r.pdf")
    try: pdf.output(pdf_path_summary)
    except OSError: raise ValueError(f"Plik podsumowania jest otwarty w innym programie!\nZamknij: rozliczenie_wydatków...pdf")
    try:
        online_zdarzenie(dokumenty=len(podsumowanie))
        online_synchronizuj_w_tle()   # dane w arkuszu od razu, nie za 15 min
    except Exception:
        pass
    return podsumowanie

def log_error(exc: Exception):
    os.makedirs(LOGS_DIR, exist_ok=True)
    with open(LOG_FILE, 'a', encoding='utf-8') as f: f.write(f"\n{'='*60}\n{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n{traceback.format_exc()}")

class PlanerWizytThread(QThread):
    """Wątek planera wizyt: geokoduje punkty i układa plan w tle, żeby pasek
    postępu w overlayu mógł się animować. Niezależny od generatora delegacji."""
    postep = pyqtSignal(str, float)
    sukces = pyqtSignal(object)     # plan (dict)
    blad   = pyqtSignal(str)
    def __init__(self, pozycje, rok, miesiac, offline, adres_bazy="",
                 cykliczny=False, cykl_domyslny=4, cykle_sieci=None, horyzont=3,
                 cykle_sieci_duze=None, cykle_punktow=None):
        super().__init__()
        self.pozycje = pozycje; self.rok = rok; self.miesiac = miesiac; self.offline = offline
        self.adres_bazy = (adres_bazy or "").strip()
        self.cykliczny = bool(cykliczny)
        self.cykl_domyslny = int(cykl_domyslny or 4)
        self.cykle_sieci = cykle_sieci or {}
        self.cykle_sieci_duze = cykle_sieci_duze or {}
        self.cykle_punktow = cykle_punktow or {}
        self.horyzont = int(horyzont or 3)
    def run(self):
        try:
            punkty = []
            n = len(self.pozycje)
            # FAZA 1 — NATYCHMIAST: współrzędne z bazy offline miast (bez sieci)
            for i, poz in enumerate(self.pozycje):
                adres = poz.get("adres", "") if isinstance(poz, dict) else str(poz)
                siec = poz.get("siec", "") if isinstance(poz, dict) else ""
                miasto = poz.get("miasto", "") if isinstance(poz, dict) else ""
                if not miasto:
                    miasto = adres.split(",")[-1].strip() if "," in adres else adres
                    # odetnij kod pocztowy jeśli został
                    miasto = re.sub(r"^\d{2}-\d{3}\s*", "", miasto).strip()
                coords = coords_z_miasta(miasto)
                klucz = adres.lower().strip()
                if klucz in _geo_cache:
                    lat, lng = tuple(_geo_cache[klucz])
                elif coords:
                    lat, lng = coords
                else:
                    lat = lng = None
                etykieta = f"{siec} — {miasto}" if (siec and miasto) else (siec or miasto or adres)
                punkty.append(PunktWizyty(etykieta, adres, siec=siec, lat=lat, lng=lng, miasto=miasto))
                if i % 200 == 0:
                    self.postep.emit(f"Lokalizowanie z bazy miast ({i+1}/{n})...", (i+1)/max(n,1)*0.5)

            # FAZA 2 — DOGEOKODOWANIE: miasta spoza bazy pobieramy z sieci, ale
            # TYLKO RAZ NA MIASTO (nie na każdy adres!). Zwykle to garstka miast.
            braki = {}
            for p in punkty:
                if p.lat is None and p.miasto:
                    braki.setdefault(_norm_miasto(p.miasto), []).append(p)
            if braki:
                lista = list(braki.items())
                for j, (mnorm, punkty_miasta) in enumerate(lista):
                    self.postep.emit(f"Nowe miejscowości ({j+1}/{len(lista)})...",
                                     0.5 + (j+1)/max(len(lista),1) * 0.3)
                    nazwa = punkty_miasta[0].miasto
                    try:
                        lat, lng = pobierz_coords(nazwa, nazwa, "mazowieckie", zapisz_cache=False)
                    except Exception:
                        lat = lng = None
                    if lat is not None:
                        # zapamiętaj w indeksie miast — kolejne razy natychmiast
                        idx = _zbuduj_indeks_miast()
                        idx[mnorm] = (lat, lng)
                        for p in punkty_miasta:
                            p.lat, p.lng = lat, lng
                zapisz_geo_cache()

            # BAZA — punkt startowy dnia (adres domowy). Bez niego trasy startują
            # ze sztucznego środka geograficznego, co daje gorsze wyniki.
            baza = None
            if self.adres_bazy:
                self.postep.emit("Lokalizowanie punktu startowego...", 0.82)
                miasto_b = self.adres_bazy.split(",")[-1].strip() if "," in self.adres_bazy else self.adres_bazy
                miasto_b = re.sub(r"^\d{2}-\d{3}\s*", "", miasto_b).strip()
                lat_b = lng_b = None
                klucz_b = self.adres_bazy.lower().strip()
                # 1) dokładny adres z cache (najlepszy)
                if klucz_b in _geo_cache:
                    lat_b, lng_b = tuple(_geo_cache[klucz_b])
                else:
                    # 2) miasto z bazy offline — pewne i natychmiastowe
                    c = coords_z_miasta(miasto_b)
                    if c:
                        lat_b, lng_b = c
                    else:
                        # 3) dopiero teraz sieć (miasto nieznane bazie).
                        # UWAGA: pobierz_coords przy braku sieci zwraca domyślną
                        # stolicę województwa — dlatego używamy go na końcu.
                        try:
                            lat_b, lng_b = pobierz_coords(self.adres_bazy, miasto_b or self.adres_bazy, "mazowieckie")
                        except Exception:
                            lat_b = lng_b = None
                if lat_b is not None:
                    baza = PunktWizyty("Baza (start)", self.adres_bazy, lat=lat_b, lng=lng_b, miasto=miasto_b)

            self.postep.emit("Układanie tras...", 0.85)
            if self.cykliczny:
                plan = planuj_cyklicznie(
                    punkty, self.rok, self.miesiac,
                    cykl_domyslny=self.cykl_domyslny,
                    cykle_sieci=self.cykle_sieci,
                    cykle_sieci_duze=self.cykle_sieci_duze,
                    cykle_punktow=self.cykle_punktow,
                    miesiecy=self.horyzont,
                    baza=baza, offline=self.offline)
            else:
                plan = planuj_wizyty(punkty, self.rok, self.miesiac,
                                     baza=baza, offline=self.offline)
            plan["adres_bazy"] = self.adres_bazy
            self.postep.emit("Gotowe!", 1.0)
            self.sukces.emit(plan)
        except Exception as e:
            self.blad.emit(str(e))


class GeneratorThread(QThread):
    postep = pyqtSignal(str, float)
    sukces = pyqtSignal(list, object, str)
    blad   = pyqtSignal(str)
    def __init__(self, params: dict):
        super().__init__(); self.params = params
    def run(self):
        try:
            p = self.params
            self.postep.emit("Pobieranie współrzędnych GPS...", 0.10)
            # Do geocodingu używamy 'adres_geo' — dla wsi to sama miejscowość
            # (np. "Sarnowa Góra, 06-430 Sońsk"), nie fałszywa "ul." — dzięki
            # temu baza trafia we właściwe miejsce, a nie w siedzibę gminy.
            adres_do_geo = p.get('adres_geo', p['adres_caly'])
            baza_lat, baza_lng = pobierz_coords(adres_do_geo, p['baza_miasto'], p['woj'])
            pracownik = DanePracownika(imie=p['imie'], pesel=p['pesel'], adres=p['adres_caly'], stanowisko=p['stanowisko'], kod_pocztowy=p['kod_pocztowy'], baza_miasto=p['baza_miasto'], baza_lat=baza_lat, baza_lng=baza_lng, wojewodztwo=p['woj'])
            
            self.postep.emit("Generowanie tras...", 0.40)
            def p_cb(txt, val): self.postep.emit(txt, val)
            finalne_dni = generuj_trasy(p['kwota_cel'], p['baza_miasto'], baza_lat, baza_lng, p['woj'], p['dni_robocze'], p['pesel'], p.get('stawka', STAWKA_ZA_KM), p_cb)

            # Czy realnymi trasami udało się pokryć żądaną kwotę? (info dla UI)
            self._kwota_niepelna = getattr(finalne_dni, 'kwota_niepelna', False)
            self._kwota_osiagnieta = getattr(finalne_dni, 'kwota_osiagnieta', 0.0)
            self._kwota_docelowa = getattr(finalne_dni, 'kwota_docelowa', p['kwota_cel'])
            
            self.postep.emit("Rysowanie dokumentów PDF...", 0.85)
            folder = os.path.join(sciezka_pulpitu(), f"Rozliczenie_{pracownik.imie.replace(' ','_')}_{p['miesiac_slownie']}_{p['rok']}r")
            generuj_pdfy(finalne_dni, pracownik, p['miesiac'], p['rok'], folder, p.get('stawka', STAWKA_ZA_KM), p_cb)
            
            self.postep.emit("Generowanie podglądu tras HTML...", 0.95)
            generuj_mape_html(finalne_dni, pracownik, p['miesiac_slownie'], p['rok'], folder, p.get('is_dark', True))
            
            self.sukces.emit(finalne_dni, pracownik, folder)
        except Exception as e:
            log_error(e); self.blad.emit(str(e))

def znajdz_logo() -> Optional[str]:
    """Szuka pliku logo w kolejności: nowe nazwy pmt_logo.*, potem stare pmt.*.
    Przeszukuje: wnętrze .exe (_MEIPASS), katalog obok programu, Pulpit oraz
    folder PMT na Pulpicie (tam użytkownik trzyma pliki źródłowe)."""
    nazwy = ["pmt_logo.png", "pmt_logo.jpg", "pmt_logo.ico",
             "pmt.png", "pmt.jpg", "PMT.jpg"]
    # katalogi do przeszukania
    pulpit = sciezka_pulpitu()
    katalogi = [None,                                   # zasob_sciezka (exe/_MEIPASS/obok)
                pulpit,
                os.path.join(pulpit, "PMT")]
    for nazwa in nazwy:
        for kat in katalogi:
            p = zasob_sciezka(nazwa) if kat is None else os.path.join(kat, nazwa)
            if os.path.exists(p):
                return p
    return None

def znajdz_ikone() -> Optional[str]:
    """Ikona okna/aplikacji — preferuje .ico, potem .png. Te same lokalizacje."""
    pulpit = sciezka_pulpitu()
    for nazwa in ["pmt_logo.ico", "pmt_logo.png", "pmt.ico", "pmt.png"]:
        for kat in [None, pulpit, os.path.join(pulpit, "PMT")]:
            p = zasob_sciezka(nazwa) if kat is None else os.path.join(kat, nazwa)
            if os.path.exists(p):
                return p
    return None

def draw_svg_icon(painter, icon_name, rect, color, is_dark):
    painter.save()
    painter.setRenderHint(QPainter.RenderHint.Antialiasing)
    painter.translate(rect.topLeft())
    pen = QPen(color); pen.setWidthF(2.5); pen.setCapStyle(Qt.PenCapStyle.RoundCap); pen.setJoinStyle(Qt.PenJoinStyle.RoundJoin)
    painter.setPen(pen); painter.setBrush(Qt.BrushStyle.NoBrush)
    w, h = rect.width(), rect.height()
    if icon_name == "user":
        painter.drawEllipse(QRectF(w*0.3, h*0.15, w*0.4, w*0.4))
        path = QPainterPath(); path.moveTo(w*0.15, h*0.9); path.cubicTo(w*0.15, h*0.6, w*0.85, h*0.6, w*0.85, h*0.9); painter.drawPath(path)
    elif icon_name == "card":
        painter.drawRoundedRect(QRectF(w*0.1, h*0.25, w*0.8, h*0.5), 4, 4)
        painter.drawLine(QPointF(w*0.1, h*0.45), QPointF(w*0.9, h*0.45)); painter.drawPoint(QPointF(w*0.3, h*0.6))
    elif icon_name == "map":
        path = QPainterPath(); path.moveTo(w*0.5, h*0.9); path.cubicTo(w*0.1, h*0.5, w*0.1, h*0.1, w*0.5, h*0.1)
        path.cubicTo(w*0.9, h*0.1, w*0.9, h*0.5, w*0.5, h*0.9); painter.drawPath(path)
        painter.drawEllipse(QRectF(w*0.35, h*0.3, w*0.3, w*0.3))
    elif icon_name == "home":
        path = QPainterPath(); path.moveTo(w*0.1, h*0.4); path.lineTo(w*0.5, h*0.1); path.lineTo(w*0.9, h*0.4)
        path.lineTo(w*0.9, h*0.9); path.lineTo(w*0.1, h*0.9); path.closeSubpath()
        painter.drawRect(QRectF(w*0.4, h*0.6, w*0.2, h*0.3))
    elif icon_name == "briefcase":
        painter.drawRoundedRect(QRectF(w*0.15, h*0.3, w*0.7, h*0.6), 3, 3)
        path = QPainterPath(); path.moveTo(w*0.35, h*0.3); path.lineTo(w*0.35, h*0.15); path.lineTo(w*0.65, h*0.15); path.lineTo(w*0.65, h*0.3); painter.drawPath(path)
    elif icon_name == "calendar":
        painter.drawRoundedRect(QRectF(w*0.15, h*0.2, w*0.7, h*0.7), 3, 3)
        painter.drawLine(QPointF(w*0.15, h*0.4), QPointF(w*0.85, h*0.4))
        painter.drawLine(QPointF(w*0.3, h*0.1), QPointF(w*0.3, h*0.3))
        painter.drawLine(QPointF(w*0.7, h*0.1), QPointF(w*0.7, h*0.3))
    elif icon_name == "activity":
        path = QPainterPath(); path.moveTo(0, h*0.5); path.lineTo(w*0.2, h*0.5); path.lineTo(w*0.4, h*0.2)
        path.lineTo(w*0.6, h*0.8); path.lineTo(w*0.8, h*0.5); path.lineTo(w, h*0.5); painter.drawPath(path)
    elif icon_name == "download":
        path = QPainterPath(); path.moveTo(w*0.5, h*0.1); path.lineTo(w*0.5, h*0.7)
        path.moveTo(w*0.25, h*0.45); path.lineTo(w*0.5, h*0.7); path.lineTo(w*0.75, h*0.45)
        path.moveTo(w*0.2, h*0.9); path.lineTo(w*0.8, h*0.9); painter.drawPath(path)
    elif icon_name == "moon":
        big = QPainterPath(); big.addEllipse(QRectF(w*0.15, h*0.15, w*0.68, h*0.68))
        cut = QPainterPath(); cut.addEllipse(QRectF(w*0.38, h*0.02, w*0.68, h*0.68))
        painter.setPen(Qt.PenStyle.NoPen); painter.setBrush(color)
        painter.drawPath(big.subtracted(cut))
    elif icon_name == "folder":
        path = QPainterPath(); path.moveTo(w*0.1, h*0.3); path.lineTo(w*0.1, h*0.85)
        path.lineTo(w*0.9, h*0.85); path.lineTo(w*0.9, h*0.35); path.lineTo(w*0.5, h*0.35)
        path.lineTo(w*0.42, h*0.2); path.lineTo(w*0.1, h*0.2); path.closeSubpath()
        painter.drawPath(path)
    elif icon_name == "kompas":
        # Nowa Wyprawa — kompas z igłą (róża wiatrów)
        painter.drawEllipse(QRectF(w*0.12, h*0.12, w*0.76, h*0.76))
        igla = QPainterPath()
        igla.moveTo(w*0.5, h*0.22); igla.lineTo(w*0.62, h*0.5)
        igla.lineTo(w*0.5, h*0.78); igla.lineTo(w*0.38, h*0.5); igla.closeSubpath()
        painter.setBrush(color); painter.drawPath(igla); painter.setBrush(Qt.BrushStyle.NoBrush)
    elif icon_name == "szlak":
        # Moje Szlaki — kręta ścieżka z pinezkami
        sciezka = QPainterPath(); sciezka.moveTo(w*0.2, h*0.85)
        sciezka.cubicTo(w*0.1, h*0.55, w*0.9, h*0.55, w*0.8, h*0.2)
        pen2 = QPen(color); pen2.setWidthF(2.3); pen2.setStyle(Qt.PenStyle.DashLine)
        pen2.setCapStyle(Qt.PenCapStyle.RoundCap); painter.setPen(pen2); painter.drawPath(sciezka)
        painter.setPen(QPen(color, 2.5)); painter.setBrush(color)
        painter.drawEllipse(QRectF(w*0.14, h*0.79, w*0.12, w*0.12))
        painter.drawEllipse(QRectF(w*0.74, h*0.14, w*0.12, w*0.12))
        painter.setBrush(Qt.BrushStyle.NoBrush)
    elif icon_name == "kalendarz2":
        # Kalendarz Wypraw — kartka z zaznaczonym dniem
        painter.drawRoundedRect(QRectF(w*0.15, h*0.2, w*0.7, h*0.68), 4, 4)
        painter.drawLine(QPointF(w*0.15, h*0.42), QPointF(w*0.85, h*0.42))
        painter.drawLine(QPointF(w*0.32, h*0.1), QPointF(w*0.32, h*0.28))
        painter.drawLine(QPointF(w*0.68, h*0.1), QPointF(w*0.68, h*0.28))
        painter.setBrush(color); painter.drawEllipse(QRectF(w*0.58, h*0.58, w*0.16, w*0.16))
        painter.setBrush(Qt.BrushStyle.NoBrush)
    elif icon_name == "wykres":
        # Bilans Miesiąca — słupki wykresu
        painter.drawLine(QPointF(w*0.15, h*0.85), QPointF(w*0.85, h*0.85))
        for bx, bh in [(0.28, 0.35), (0.48, 0.55), (0.68, 0.28)]:
            painter.setBrush(color)
            painter.drawRoundedRect(QRectF(w*bx, h*(0.8-bh), w*0.12, h*bh), 2, 2)
        painter.setBrush(Qt.BrushStyle.NoBrush)
    elif icon_name == "zebatka":
        # Ustawienia — koło zębate
        import math as _m
        cx, cy, rz = w*0.5, h*0.5, w*0.26
        painter.drawEllipse(QRectF(cx-rz, cy-rz, rz*2, rz*2))
        painter.drawEllipse(QRectF(cx-w*0.09, cy-w*0.09, w*0.18, w*0.18))
        for i in range(8):
            a = i * _m.pi / 4
            x1 = cx + _m.cos(a)*rz; y1 = cy + _m.sin(a)*rz
            x2 = cx + _m.cos(a)*(rz+w*0.12); y2 = cy + _m.sin(a)*(rz+w*0.12)
            painter.drawLine(QPointF(x1, y1), QPointF(x2, y2))
    elif icon_name == "info":
        # O programie — okrąg z „i”
        painter.drawEllipse(QRectF(w*0.15, h*0.15, w*0.7, h*0.7))
        painter.setBrush(color)
        painter.drawEllipse(QRectF(w*0.46, h*0.28, w*0.08, w*0.08))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawLine(QPointF(w*0.5, h*0.44), QPointF(w*0.5, h*0.7))
    elif icon_name == "checklista":
        # Plan Wizyt — kartka z odhaczonymi punktami
        painter.drawRoundedRect(QRectF(w*0.18, h*0.1, w*0.64, h*0.8), 4, 4)
        for fy in (0.3, 0.5, 0.7):
            painter.setBrush(color)
            painter.drawEllipse(QRectF(w*0.28, h*fy - w*0.04, w*0.08, w*0.08))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawLine(QPointF(w*0.44, h*fy), QPointF(w*0.72, h*fy))
    elif icon_name == "tarcza":
        # Kopia zapasowa — tarcza ochronna z ptaszkiem
        tarcza = QPainterPath()
        tarcza.moveTo(w*0.5, h*0.08)
        tarcza.cubicTo(w*0.75, h*0.16, w*0.82, h*0.18, w*0.82, h*0.18)
        tarcza.lineTo(w*0.82, h*0.48)
        tarcza.cubicTo(w*0.82, h*0.72, w*0.66, h*0.86, w*0.5, h*0.92)
        tarcza.cubicTo(w*0.34, h*0.86, w*0.18, h*0.72, w*0.18, h*0.48)
        tarcza.lineTo(w*0.18, h*0.18)
        tarcza.cubicTo(w*0.18, h*0.18, w*0.25, h*0.16, w*0.5, h*0.08)
        tarcza.closeSubpath()
        painter.drawPath(tarcza)
        pen3 = QPen(color); pen3.setWidthF(2.4); pen3.setCapStyle(Qt.PenCapStyle.RoundCap)
        pen3.setJoinStyle(Qt.PenJoinStyle.RoundJoin)
        painter.setPen(pen3)
        ptaszek = QPainterPath()
        ptaszek.moveTo(w*0.36, h*0.5); ptaszek.lineTo(w*0.46, h*0.62); ptaszek.lineTo(w*0.66, h*0.36)
        painter.drawPath(ptaszek)
    elif icon_name == "trend":
        # Statystyki osobiste — linia trendu wzrostowego + kropka na szczycie
        pen4 = QPen(color); pen4.setWidthF(2.2); pen4.setCapStyle(Qt.PenCapStyle.RoundCap)
        pen4.setJoinStyle(Qt.PenJoinStyle.RoundJoin)
        painter.setPen(pen4)
        linia = QPainterPath()
        linia.moveTo(w*0.14, h*0.78)
        linia.lineTo(w*0.38, h*0.52)
        linia.lineTo(w*0.56, h*0.66)
        linia.lineTo(w*0.86, h*0.24)
        painter.drawPath(linia)
        painter.setBrush(color)
        painter.drawEllipse(QRectF(w*0.80, h*0.18, w*0.12, w*0.12))
        pen4b = QPen(color); pen4b.setWidthF(1.6)
        painter.setPen(pen4b)
        painter.drawLine(QPointF(w*0.14, h*0.86), QPointF(w*0.9, h*0.86))
    painter.restore()

class SvgIconLabel(QWidget):
    def __init__(self, icon_name, parent=None, size=20):
        super().__init__(parent); self.icon_name = icon_name; self.is_dark = True; self.color = QColor("#94A3B8"); self.setFixedSize(size, size)
    def set_theme(self, is_dark): self.is_dark = is_dark; self.color = QColor("#94A3B8") if is_dark else QColor("#475569"); self.update()
    def set_color(self, color_hex): self.color = QColor(color_hex); self.update()
    def paintEvent(self, event):
        painter = QPainter(self)
        draw_svg_icon(painter, self.icon_name, QRectF(2, 2, self.width()-4, self.height()-4), self.color, self.is_dark)
        painter.end()

class ImageBackgroundWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent); self.is_dark = True; self._orig_dark = None; self._orig_light = None; self._load_images()
    def _load_images(self):
        p_dark = zasob_sciezka("ciemny.png")
        p_light = zasob_sciezka("jasny.png")
        if os.path.exists(p_dark): self._orig_dark = QPixmap(p_dark)
        if os.path.exists(p_light): self._orig_light = QPixmap(p_light)
    def set_theme(self, is_dark): self.is_dark = is_dark; self.update()
    def _tlo_zastepcze(self, painter):
        """Gradient w barwach PMT — używany, gdy w wydaniu zabrakło plików
        ciemny.png / jasny.png. Zamiast płaskiej czerni: granat -> turkus."""
        from PyQt6.QtGui import QLinearGradient
        g = QLinearGradient(0, 0, self.width(), self.height())
        if self.is_dark:
            g.setColorAt(0.0, QColor("#0F172A")); g.setColorAt(0.55, QColor("#0B1320"))
            g.setColorAt(1.0, QColor("#04121A"))
        else:
            g.setColorAt(0.0, QColor("#F8FAFC")); g.setColorAt(1.0, QColor("#E2E8F0"))
        painter.fillRect(self.rect(), g)
        if self.is_dark:                      # subtelna poświata w rogu
            from PyQt6.QtGui import QRadialGradient
            r = QRadialGradient(self.width() * 0.82, self.height() * 0.18,
                                max(self.width(), self.height()) * 0.55)
            r.setColorAt(0.0, QColor(0, 240, 255, 34)); r.setColorAt(1.0, QColor(0, 240, 255, 0))
            painter.fillRect(self.rect(), r)

    def paintEvent(self, event):
        painter = QPainter(self); painter.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        orig = self._orig_dark if self.is_dark else self._orig_light
        if orig and not orig.isNull():
            painter.fillRect(self.rect(), QColor("#0B1320") if self.is_dark else QColor("#F1F5F9"))
        else:
            self._tlo_zastepcze(painter)
        if orig and not orig.isNull():
            scaled = orig.scaled(self.size(), Qt.AspectRatioMode.KeepAspectRatioByExpanding, Qt.TransformationMode.SmoothTransformation)
            crop_x = (scaled.width() - self.width()) // 2; crop_y = (scaled.height() - self.height()) // 2
            painter.drawPixmap(0, 0, scaled, crop_x, crop_y, self.width(), self.height())
        painter.end()

class GrubyKursorEdit(QLineEdit):
    """QLineEdit z własnoręcznie rysowanym, grubym i wyraźnie widocznym
    kursorem. Systemowy caret w PyQt6 bywa niewidoczny (ciemny motyw Windows,
    ignorowane 'caret-color', przezroczyste tło) — dlatego rysujemy go sami,
    migającego, o szerokości 3px w kolorze akcentu. To działa niezależnie od
    ustawień systemu."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._caret_on = True
        self._caret_kolor = QColor("#00F0FF")
        self._blink = QTimer(self)
        self._blink.setInterval(530)
        self._blink.timeout.connect(self._toggle_caret)
        self._blink.start()

    def set_caret_kolor(self, hex_kolor):
        self._caret_kolor = QColor(hex_kolor); self.update()

    def _toggle_caret(self):
        if self.hasFocus():
            self._caret_on = not self._caret_on
        else:
            self._caret_on = False
        self.update()

    def focusInEvent(self, e):
        super().focusInEvent(e)
        self._caret_on = True; self.update()

    def focusOutEvent(self, e):
        super().focusOutEvent(e)
        self._caret_on = False; self.update()

    def paintEvent(self, event):
        super().paintEvent(event)
        if not (self.hasFocus() and self._caret_on):
            return
        try:
            # cursorRect zwraca prostokąt systemowego kursora; jego środek to
            # dokładne miejsce między znakami. Systemowy caret ustawiamy na 1px
            # (patrz __init__), więc środek nie jest przesunięty.
            x = self.cursorRect().center().x()
        except Exception:
            return
        fm = QFontMetrics(self.font())
        gora = (self.height() - fm.height()) // 2 + 2
        dol = gora + fm.height() - 4
        p = QPainter(self)
        pen = QPen(self._caret_kolor); pen.setWidth(2); pen.setCapStyle(Qt.PenCapStyle.RoundCap)
        p.setPen(pen)
        p.drawLine(x, gora, x, dol)
        p.end()


class StyledInput(QFrame):
    def __init__(self, icon_name, widget, is_dark=True, parent=None):
        super().__init__(parent); self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True); self.setObjectName("StyledInputField")
        self.is_dark = is_dark; self.setFixedHeight(32)   # niżej: karty nie zachodzą na siebie
        self._stan_walidacji = None    # None / "ok" / "err" — pamięć podświetlenia Asystenta
        layout = QHBoxLayout(self); layout.setContentsMargins(12, 0, 12, 0); layout.setSpacing(10)
        self.icon_w = SvgIconLabel(icon_name, self, size=20); layout.addWidget(self.icon_w)
        self.widget = widget; self.widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self.widget.installEventFilter(self); layout.addWidget(self.widget)
        self.update_theme(is_dark)

    def _caret_css(self):
        return "#00F0FF" if self.is_dark else "#0D9488"

    def _text_css(self):
        return "#FFFFFF" if self.is_dark else "#000000"

    def _ustaw_kolor_kursora(self):
        """Kolor kursora (caret) w PyQt6 ustawia się przez paletę (QSS caret-color
        jest ignorowane). Ustawiamy jasny kolor tekstu = jasny, widoczny kursor."""
        from PyQt6.QtGui import QPalette
        pal = self.widget.palette()
        kolor = QColor("#FFFFFF") if self.is_dark else QColor("#0F172A")
        pal.setColor(QPalette.ColorRole.Text, kolor)
        self.widget.setPalette(pal)
        # jeśli to nasze pole z ręcznie rysowanym kursorem — ustaw jego kolor
        if hasattr(self.widget, "set_caret_kolor"):
            self.widget.set_caret_kolor("#00F0FF" if self.is_dark else "#0D9488")

    def _widget_css(self):
        txt = self._text_css()
        ph = "#64748B" if self.is_dark else "#94A3B8"
        view_bg = "#0B1320" if self.is_dark else "#FFFFFF"
        view_sel = "#00F0FF" if self.is_dark else "#0D9488"
        view_seltxt = "#000000" if self.is_dark else "#FFFFFF"
        return (f"QLineEdit, QComboBox {{ background: transparent; border: none; color: {txt}; "
                f"font-size: 13px; font-family: 'Segoe UI', sans-serif; font-weight: 500; }} "
                f"QLineEdit::placeholder {{ color: {ph}; }} "
                f"QComboBox QAbstractItemView {{ background: {view_bg}; color: {txt}; "
                f"selection-background-color: {view_sel}; selection-color: {view_seltxt}; }}")

    def _ramka_koloru(self, kolor, bg):
        self.setStyleSheet(f"#StyledInputField {{ background-color: {bg}; border: 1px solid {kolor}; border-radius: 8px; }}")
        self.icon_w.set_color(kolor)
        self.widget.setStyleSheet(self._widget_css())
        self._ustaw_kolor_kursora()

    def ustaw_walidacje(self, stan):
        """Zapamiętuje status z Asystenta: 'ok'(zielony) / 'err'(czerwony) / None.
        Podświetlenie utrzymuje się także po opuszczeniu pola myszką."""
        self._stan_walidacji = stan
        # nie nadpisuj gdy pole ma aktualnie fokus (wtedy rządzi ramka focusu)
        if not self.widget.hasFocus():
            self._przywroc_wyglad()

    def _przywroc_wyglad(self):
        if self._stan_walidacji == "ok":
            self._ramka_koloru("#10B981", "rgba(16,185,129,0.10)")
        elif self._stan_walidacji == "err":
            self._ramka_koloru("#EF4444", "rgba(239,68,68,0.10)")
        else:
            self.update_theme(self.is_dark)

    def update_theme(self, is_dark):
        self.is_dark = is_dark; self.icon_w.set_theme(is_dark)
        if self.is_dark:
            self.setStyleSheet("#StyledInputField { background-color: rgba(5, 10, 20, 0.28); border: 1px solid rgba(255, 255, 255, 0.28); border-radius: 8px; }")
        else:
            self.setStyleSheet("#StyledInputField { background-color: rgba(255, 255, 255, 0.42); border: 1px solid rgba(148, 163, 184, 0.7); border-radius: 8px; }")
        self.widget.setStyleSheet(self._widget_css())
        self._ustaw_kolor_kursora()

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Type.FocusIn:
            color = self._caret_css(); bg = "rgba(0, 0, 0, 0.75)" if self.is_dark else "rgba(255, 255, 255, 0.9)"
            self._ramka_koloru(color, bg)
        elif event.type() == QEvent.Type.FocusOut:
            # NIE kasujemy koloru walidacji — przywracamy zapamiętany status
            self._przywroc_wyglad()
        return super().eventFilter(obj, event)

def styl_zglos_blad(btn):
    """Zgłoszenie błędu ma być rozpoznawalne kolorem alarmu — czerwony obrys,
    a po najechaniu pełne wypełnienie. Celowo słabszy akcent niż Wyloguj,
    żeby dwa kolorowe przyciski obok siebie nie krzyczały jednakowo."""
    btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
    btn.setMinimumHeight(34)
    btn.setStyleSheet(
        "QPushButton { color:#F87171; background:rgba(239,68,68,0.10);"
        " border:1.5px solid #EF4444; border-radius:17px; padding:7px 18px;"
        " font-family:'Segoe UI'; font-size:12.5px; font-weight:800; }"
        "QPushButton:hover { background:#EF4444; color:#FFFFFF; }"
        "QPushButton:pressed { background:#B91C1C; color:#FFE4E6; }")


def styl_wyloguj(btn, is_dark=True):
    """Wylogowanie ma być widoczne — bursztynowy pigułkowy przycisk, wyraźnie
    większy od sąsiadów w pasku. Kolor celowo inny niż akcent systemu (cyjan),
    żeby nie mylił się z akcjami „twórczymi”."""
    btn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
    btn.setMinimumHeight(38)
    btn.setStyleSheet(
        "QPushButton { color:#1A1206;"
        " background: qlineargradient(x1:0,y1:0,x2:1,y2:0,"
        "   stop:0 #FBBF24, stop:1 #F97316);"
        " border:none; border-radius:19px; padding:9px 22px;"
        " font-family:'Segoe UI'; font-size:13px; font-weight:800; }"
        "QPushButton:hover { background: qlineargradient(x1:0,y1:0,x2:1,y2:0,"
        "   stop:0 #FCD34D, stop:1 #FB923C); }"
        "QPushButton:pressed { background:#EA580C; color:#FFF7ED; }")


class OutlineButton(QPushButton):
    def __init__(self, text, is_dark=True, parent=None):
        super().__init__(text, parent); self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor)); self.update_theme(is_dark)
    def update_theme(self, is_dark):
        if is_dark: self.setStyleSheet("QPushButton { color: #E2E8F0; background-color: rgba(255, 255, 255, 0.05); border: 1px solid rgba(255, 255, 255, 0.2); border-radius: 14px; padding: 6px 16px; font-family: 'Segoe UI', sans-serif; font-size: 12px; font-weight: 600; } QPushButton:hover { background-color: rgba(255, 255, 255, 0.15); }")
        else: self.setStyleSheet("QPushButton { color: #475569; background-color: rgba(255, 255, 255, 0.8); border: 1px solid rgba(200, 210, 220, 1.0); border-radius: 14px; padding: 6px 16px; font-family: 'Segoe UI', sans-serif; font-size: 12px; font-weight: 600; } QPushButton:hover { background-color: #F8FAFC; border: 1px solid #94A3B8; }")

class PmtOrbitWidget(QWidget):
    """Logo PMT z orbitującą kometą podczas ładowania i animowanym checkmarkiem po sukcesie.
    W 100% rysowane przez QPainter — brak zewnętrznych zasobów graficznych."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(160, 160)
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        self._scale = 0.0
        self._orbit_angle = 0.0
        self._check_t = 0.0
        self._check_alpha = 0.0
        self._pulse = 1.0
        self._show_check = False

        # wjazd: scale 0→1  (wolniejszy, dostojny)
        self._a_scale = QPropertyAnimation(self, b"_scale_prop", self)
        self._a_scale.setDuration(1100); self._a_scale.setStartValue(0.0); self._a_scale.setEndValue(1.0)
        self._a_scale.setEasingCurve(QEasingCurve.Type.OutBack)

        # orbita (timer 12ms ≈ 83fps)
        self._orbit_timer = QTimer(self)
        self._orbit_timer.setInterval(12)
        self._orbit_timer.timeout.connect(self._tick_orbit)

        # rysowanie ptaszka — powoli, tak żeby dało się śledzić kreślenie
        self._a_check = QPropertyAnimation(self, b"_check_t_prop", self)
        self._a_check.setDuration(1400); self._a_check.setStartValue(0.0); self._a_check.setEndValue(1.0)
        self._a_check.setEasingCurve(QEasingCurve.Type.InOutCubic)

        # pojawianie się zielonego tła checkmark (fade)
        self._a_check_alpha = QPropertyAnimation(self, b"_check_alpha_prop", self)
        self._a_check_alpha.setDuration(420)
        self._a_check_alpha.setStartValue(0.0); self._a_check_alpha.setEndValue(1.0)

        # pulsowanie po sukcesie — dwa spokojne oddechy
        self._a_pulse = QPropertyAnimation(self, b"_pulse_prop", self)
        self._a_pulse.setDuration(900)
        self._a_pulse.setKeyValueAt(0.0, 1.0); self._a_pulse.setKeyValueAt(0.25, 1.12)
        self._a_pulse.setKeyValueAt(0.5, 1.0);  self._a_pulse.setKeyValueAt(0.75, 1.09)
        self._a_pulse.setKeyValueAt(1.0, 1.0)
        self._a_pulse.setEasingCurve(QEasingCurve.Type.InOutSine)

    # pyqtProperty helpers
    def _get_scale(self): return self._scale
    def _set_scale(self, v): self._scale = v; self.update()
    _scale_prop = pyqtProperty(float, _get_scale, _set_scale)

    def _get_check_t(self): return self._check_t
    def _set_check_t(self, v): self._check_t = v; self.update()
    _check_t_prop = pyqtProperty(float, _get_check_t, _set_check_t)

    def _get_check_alpha(self): return self._check_alpha
    def _set_check_alpha(self, v): self._check_alpha = v; self.update()
    _check_alpha_prop = pyqtProperty(float, _get_check_alpha, _set_check_alpha)

    def _get_pulse(self): return self._pulse
    def _set_pulse(self, v): self._pulse = v; self.update()
    _pulse_prop = pyqtProperty(float, _get_pulse, _set_pulse)

    def start_enter(self):
        self._show_check = False; self._scale = 0.0; self._check_t = 0.0; self._check_alpha = 0.0
        self._orbit_angle = -90.0    # start z góry (godz. 12) — krąg domknie się wizualnie
        self._a_scale.start(); self._orbit_timer.start()

    def start_success(self):
        self._orbit_timer.stop(); self._show_check = True
        self._a_check_alpha.start()
        self._a_check.start()
        self._a_check.finished.connect(lambda: self._a_pulse.start())

    def stop_all(self):
        self._orbit_timer.stop()
        for a in [self._a_scale, self._a_check, self._a_check_alpha, self._a_pulse]: a.stop()

    def _tick_orbit(self):
        self._orbit_angle = (self._orbit_angle + 3.0) % 360.0; self.update()

    def paintEvent(self, event):
        import math
        s = self._scale * self._pulse
        if s <= 0.01: return
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        cx, cy = self.width() / 2, self.height() / 2
        p.translate(cx, cy); p.scale(s, s); p.translate(-cx, -cy)

        # tło logo — gradient teal
        r = 52.0
        grad = QLinearGradient(cx - r, cy - r, cx + r, cy + r)
        grad.setColorAt(0.0, QColor("#00F0FF")); grad.setColorAt(1.0, QColor("#00C480"))
        p.setPen(Qt.PenStyle.NoPen); p.setBrush(QBrush(grad))
        p.drawEllipse(QPointF(cx, cy), r, r)
        # napis PMT
        p.setPen(QColor("#050B14"))
        f = QFont("Segoe UI", 21, QFont.Weight.Black)
        p.setFont(f)
        p.drawText(QRectF(cx - r, cy - r, r * 2, r * 2), Qt.AlignmentFlag.AlignCenter, "PMT")

        # orbita — ogon komety
        if not self._show_check:
            r_orb = r + 17
            for i in range(32):
                t = i / 31.0
                alpha = int(t * 200)
                a_seg = self._orbit_angle - 90 * (1 - t)
                x = cx + math.cos(math.radians(a_seg)) * r_orb
                y = cy + math.sin(math.radians(a_seg)) * r_orb
                p.setPen(Qt.PenStyle.NoPen)
                p.setBrush(QColor(0, 240, 255, alpha))
                p.drawEllipse(QPointF(x, y), 1.4 + t * 3.2, 1.4 + t * 3.2)
            # głowa komety
            hx = cx + math.cos(math.radians(self._orbit_angle)) * r_orb
            hy = cy + math.sin(math.radians(self._orbit_angle)) * r_orb
            # poświata
            glow = QRadialGradient(QPointF(hx, hy), 13)
            glow.setColorAt(0, QColor(0, 240, 255, 180)); glow.setColorAt(1, QColor(0, 240, 255, 0))
            p.setBrush(QBrush(glow)); p.drawEllipse(QPointF(hx, hy), 13, 13)
            # jasny rdzeń
            p.setBrush(QColor(255, 255, 255, 240)); p.drawEllipse(QPointF(hx, hy), 4.5, 4.5)

        # checkmark
        if self._show_check and self._check_alpha > 0.001:
            alpha_int = int(self._check_alpha * 255)
            p.setPen(Qt.PenStyle.NoPen)
            p.setBrush(QColor(16, 185, 129, alpha_int))
            p.drawEllipse(QPointF(cx, cy), r, r)
            t = self._check_t
            A = QPointF(cx + (0.25 - 0.5) * r * 1.6, cy + (0.52 - 0.5) * r * 1.6)
            B = QPointF(cx + (0.43 - 0.5) * r * 1.6, cy + (0.70 - 0.5) * r * 1.6)
            C = QPointF(cx + (0.78 - 0.5) * r * 1.6, cy + (0.31 - 0.5) * r * 1.6)
            chk_pen = QPen(QColor(255, 255, 255, alpha_int))
            chk_pen.setWidthF(5.5); chk_pen.setCapStyle(Qt.PenCapStyle.RoundCap)
            chk_pen.setJoinStyle(Qt.PenJoinStyle.RoundJoin); p.setPen(chk_pen)
            if t <= 0.5:
                frac = t / 0.5
                mid = QPointF(A.x() + (B.x() - A.x()) * frac, A.y() + (B.y() - A.y()) * frac)
                p.drawLine(A, mid)
            else:
                p.drawLine(A, B)
                frac = (t - 0.5) / 0.5
                mid = QPointF(B.x() + (C.x() - B.x()) * frac, B.y() + (C.y() - B.y()) * frac)
                p.drawLine(B, mid)


class GeneratingOverlay(QWidget):
    """Pełnoekranowy overlay: ciemne tło opada → logo PMT rośnie ze środka + orbita →
    po sukcesie orbita znika, pojawia się checkmark → po 1.8s overlay znika."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self._bg_alpha = 0.0

        lay = QVBoxLayout(self); lay.setAlignment(Qt.AlignmentFlag.AlignCenter); lay.setSpacing(20)
        self.logo = PmtOrbitWidget(self)
        lay.addWidget(self.logo, alignment=Qt.AlignmentFlag.AlignCenter)

        self.lbl_status = QLabel(""); self.lbl_status.setFixedSize(320, 28)
        self.lbl_status.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_status.setStyleSheet("color:rgba(148,163,184,0); font-family:'Segoe UI',sans-serif; font-size:12px; font-weight:600; background:transparent; border:none;")
        lay.addWidget(self.lbl_status, alignment=Qt.AlignmentFlag.AlignCenter)

        self._a_bg = QPropertyAnimation(self, b"_bg_alpha_prop", self)
        self._a_bg.setDuration(340); self._a_bg.setEasingCurve(QEasingCurve.Type.OutCubic)

        self._lbl_alpha = 0.0
        self._a_lbl = QPropertyAnimation(self, b"_lbl_alpha_prop", self)
        self._a_lbl.setDuration(260); self._a_lbl.setEasingCurve(QEasingCurve.Type.OutCubic)

        self.hide()

    def _get_bg_alpha(self): return self._bg_alpha
    def _set_bg_alpha(self, v): self._bg_alpha = v; self.update()
    _bg_alpha_prop = pyqtProperty(float, _get_bg_alpha, _set_bg_alpha)

    def _get_lbl_alpha(self): return self._lbl_alpha
    def _set_lbl_alpha(self, v):
        self._lbl_alpha = v
        self.lbl_status.setStyleSheet(
            f"color:rgba(148,163,184,{int(v*255)}); font-family:'Segoe UI',sans-serif; "
            f"font-size:12px; font-weight:600; background:transparent; border:none;")
    _lbl_alpha_prop = pyqtProperty(float, _get_lbl_alpha, _set_lbl_alpha)

    def paintEvent(self, event):
        if self._bg_alpha < 0.001: return
        p = QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        p.fillRect(self.rect(), QColor(5, 11, 20, int(self._bg_alpha * 215)))

    def show_generating(self):
        # Podczas liczenia NIE pokazujemy logo — użytkownik widzi pasek postępu
        # (oś czasu) na dole. Tło jest tylko delikatnie przyciemnione.
        self.resize(self.parent().size()); self.raise_(); self.show()
        self.logo.hide()
        self._a_bg.setStartValue(0.0); self._a_bg.setEndValue(0.55); self._a_bg.start()

    def update_status(self, txt):
        self.lbl_status.setText(txt)
        self._a_lbl.stop()
        self._a_lbl.setStartValue(0.0); self._a_lbl.setEndValue(1.0); self._a_lbl.start()

    def show_success(self):
        # DOPIERO TERAZ (po przejeździe paska) pojawia się logo z rysującą się
        # "okejką" — jak gratulacja, że wszystko poszło zgodnie z planem.
        self.raise_(); self.show()
        self.logo.show()
        # pogłębiamy przyciemnienie, żeby logo wybrzmiało
        self._a_bg.stop(); self._a_bg.setStartValue(self._bg_alpha); self._a_bg.setEndValue(1.0); self._a_bg.start()
        self.logo.start_enter()                       # logo rośnie ze środka + rusza orbita
        # Kometa okrąża logo pełnym kręgiem (~1.5s) ZANIM zamieni się w checkmark —
        # dzięki temu pętla zamyka koło i wygląda logicznie, nie urywa się w pół drogi.
        QTimer.singleShot(1650, self.logo.start_success)
        self.update_status("Gotowe! ✓")
        # 1.65s orbita + 1.4s checkmark + oddechy → daj wszystkiemu wybrzmieć
        QTimer.singleShot(5200, self._fade_out)

    def _fade_out(self):
        self.logo.stop_all()
        self._a_bg.setStartValue(self._bg_alpha); self._a_bg.setEndValue(0.0)
        try: self._a_bg.finished.disconnect()
        except: pass
        self._a_bg.finished.connect(self.hide); self._a_bg.start()

    def hide_overlay(self): self._fade_out()


class SpinnerButton(QPushButton):
    def __init__(self, text, icon_name="download", is_dark=True, parent=None):
        super().__init__(parent); self.original_text = text; self.icon_name = icon_name; self.is_loading = False; self._angle = 0; self.is_dark = is_dark
        self.icon_w = SvgIconLabel(self.icon_name, parent=self); self.timer = QTimer(self); self.timer.timeout.connect(self.rotate_spinner)
        self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor)); self.update_theme(is_dark)

    def update_theme(self, is_dark):
        self.is_dark = is_dark; self.icon_w.set_theme(is_dark)
        self.icon_w.set_color("#050B14" if is_dark else "#FFFFFF")
        if is_dark:
            self.setStyleSheet("QPushButton { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #00F0FF, stop:1 #00E4A1); color: #050B14; font-family: 'Segoe UI', sans-serif; font-size: 14px; font-weight: 800; border-radius: 8px; border: none; padding-left: 20px;} QPushButton:hover { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #33F5FF, stop:1 #33EAB7); } QPushButton:disabled { background: rgba(30, 41, 59, 0.6); color: #475569; }")
        else:
            self.setStyleSheet("QPushButton { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #10B981, stop:1 #059669); color: #FFFFFF; font-family: 'Segoe UI', sans-serif; font-size: 14px; font-weight: 800; border-radius: 8px; border: none; padding-left: 20px;} QPushButton:hover { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #34D399, stop:1 #047857); } QPushButton:disabled { background: #E2E8F0; color: #94A3B8; }")

    def start_loading(self):
        self.is_loading = True; self.icon_w.hide(); self.setEnabled(False); self.timer.start(15)

    def stop_loading(self):
        self.is_loading = False; self.icon_w.show(); self.setEnabled(True); self.timer.stop(); self.update()

    def rotate_spinner(self):
        self._angle = (self._angle + 6) % 360; self.update()

    def resizeEvent(self, event): 
        super().resizeEvent(event)
        if not self.is_loading: self.icon_w.move(int(self.width()/2 - 70), int(self.height()/2 - 10))

    def paintEvent(self, event):
        super().paintEvent(event); painter = QPainter(self); painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        if self.is_loading:
            size = min(self.width(), self.height()) - 20; rect = QRectF((self.width() - size) / 2, (self.height() - size) / 2, size, size)
            pen = QPen(QColor(71, 85, 105)); pen.setWidthF(3.0); painter.setPen(pen); painter.drawEllipse(rect)
            pen.setColor(QColor("#00F0FF") if self.is_dark else QColor("#10B981")); pen.setCapStyle(Qt.PenCapStyle.RoundCap); painter.setPen(pen); painter.drawArc(rect, -self._angle * 16, 120 * 16)
        else:
            painter.setPen(QPen(QColor("#050B14") if self.is_dark else QColor("#FFFFFF"))); font = QFont("Segoe UI", 12, QFont.Weight.Bold); painter.setFont(font)
            painter.drawText(QRectF(self.icon_w.x() + 28, 0, self.width(), self.height()), Qt.AlignmentFlag.AlignVCenter, self.original_text)
        painter.end()

class GpsProgressBar(QWidget):
    def __init__(self, is_dark=True, parent=None):
        super().__init__(parent); self.setFixedHeight(30); self._progress = 0.0; self.is_dark = is_dark
        self.anim = QVariantAnimation(self); self.anim.setDuration(400); self.anim.valueChanged.connect(self._update_prog)
    def set_theme(self, is_dark): self.is_dark = is_dark; self.update()
    def set_progress(self, val): self.anim.stop(); self.anim.setStartValue(self._progress); self.anim.setEndValue(val); self.anim.start()
    def _update_prog(self, val): self._progress = val; self.update()
    def paintEvent(self, event):
        painter = QPainter(self); painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        w, h = self.width(), self.height(); cy = h / 2
        bg_color = QColor(255, 255, 255, 20) if self.is_dark else QColor(0, 0, 0, 20)
        pen_bg = QPen(bg_color, 6); pen_bg.setCapStyle(Qt.PenCapStyle.RoundCap); painter.setPen(pen_bg); painter.drawLine(10, int(cy), w - 10, int(cy))
        current_x = 10 + (w - 20) * self._progress
        if current_x > 10:
            grad = QLinearGradient(10, cy, current_x, cy)
            if self.is_dark: grad.setColorAt(0.0, QColor("#0093E9")); grad.setColorAt(1.0, QColor("#00F0FF"))
            else: grad.setColorAt(0.0, QColor("#059669")); grad.setColorAt(1.0, QColor("#10B981"))
            pen_fg = QPen(QBrush(grad), 6); pen_fg.setCapStyle(Qt.PenCapStyle.RoundCap); painter.setPen(pen_fg); painter.drawLine(10, int(cy), int(current_x), int(cy))
        painter.end()

class ToastNotification(QFrame):
    def __init__(self, parent):
        super().__init__(parent); self.setFixedSize(450, 150)
        self.setStyleSheet("QFrame { background-color: rgba(15, 23, 42, 0.95); border: 1px solid rgba(239, 68, 68, 0.5); border-left: 4px solid #EF4444; border-radius: 8px; }")
        sh = QGraphicsDropShadowEffect(self); sh.setBlurRadius(20); sh.setColor(QColor(0,0,0,150)); sh.setOffset(0, 5); self.setGraphicsEffect(sh)
        self._klik_akcja = None
        self.historia = []          # [(title, desc, success, czas), ...] — dla centrum powiadomień
        self.on_nowe_powiadomienie = None   # callback(title, desc, success) wołany przy każdym show_toast
        layout = QHBoxLayout(self); layout.setContentsMargins(16, 16, 16, 16)
        self.ic = SvgIconLabel("map", self); self.ic.set_color("#EF4444"); layout.addWidget(self.ic, alignment=Qt.AlignmentFlag.AlignTop)
        txt_l = QVBoxLayout()
        self.lbl_title = QLabel("Błąd"); self.lbl_title.setStyleSheet("color: #F8FAFC; font-weight: bold; font-size: 14px; border: none; background: transparent;")
        self.lbl_desc = QLabel("Opis"); self.lbl_desc.setStyleSheet("color: #94A3B8; font-size: 12px; border: none; background: transparent;")
        self.lbl_desc.setWordWrap(True); self.lbl_desc.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding); self.lbl_desc.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        txt_l.addWidget(self.lbl_title); txt_l.addWidget(self.lbl_desc); layout.addLayout(txt_l, 1)
        btn_close = QPushButton("✕", self); btn_close.setFixedSize(24, 24); btn_close.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        btn_close.setStyleSheet("QPushButton { color: #64748B; background: transparent; border: none; font-size: 16px; } QPushButton:hover { color: #F8FAFC; }")
        btn_close.clicked.connect(self.hide_toast); layout.addWidget(btn_close, alignment=Qt.AlignmentFlag.AlignTop)
        self.anim = QPropertyAnimation(self, b"pos"); self.anim.setDuration(400); self.anim.setEasingCurve(QEasingCurve.Type.OutBack)
        self.timer = QTimer(self); self.timer.timeout.connect(self.hide_toast); self.hide()

    def show_toast(self, title, desc, success=False, klik_akcja=None):
        self._klik_akcja = klik_akcja
        akcent = "#10B981" if success else "#EF4444"
        self.setStyleSheet(f"QFrame {{ background-color: rgba(15, 23, 42, 0.95); border: 1px solid {akcent}80; border-left: 4px solid {akcent}; border-radius: 8px; }}")
        self.ic.set_color(akcent)
        self.ic.icon_name = "check" if success else "map"; self.ic.update()
        # Podpowiedź o kliknięciu, gdy akcja dostępna
        if klik_akcja:
            desc = desc.rstrip() + "\n↳  kliknij, aby otworzyć folder"
            self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        else:
            self.setCursor(QCursor(Qt.CursorShape.ArrowCursor))
        self.lbl_title.setText(title); self.lbl_desc.setText(desc)
        p = self.parent(); start_pos = QPoint(p.width() - self.width() - 20, p.height()); end_pos = QPoint(p.width() - self.width() - 20, p.height() - self.height() - 20)
        self.move(start_pos); self.show(); self.raise_(); self.anim.setStartValue(start_pos); self.anim.setEndValue(end_pos); self.anim.start(); self.timer.start(9000)

        # Zapisz do historii (centrum powiadomień) — nawet po zniknięciu toastu
        # informacja zostaje dostępna do ponownego rozwinięcia.
        self.historia.insert(0, {
            "title": title, "desc": desc.split("\n↳")[0], "success": success,
            "czas": datetime.datetime.now().strftime("%H:%M"),
            "klik_akcja": klik_akcja,
        })
        del self.historia[20:]   # mutacja w miejscu — zachowuje referencję dla panelu
        if self.on_nowe_powiadomienie:
            self.on_nowe_powiadomienie(title, desc, success)

    def mousePressEvent(self, event):
        if self._klik_akcja and event.button() == Qt.MouseButton.LeftButton:
            try: self._klik_akcja()
            except Exception: pass
            self.hide_toast()

    def hide_toast(self):
        end_pos = QPoint(self.x(), self.parent().height()); self.anim.setStartValue(self.pos()); self.anim.setEndValue(end_pos); self.anim.start(); self.timer.stop()


class PanelPowiadomien(QFrame):
    """Rozwijana lista ostatnich powiadomień — wysuwa się spod dzwoneczka.
    Pozwala wrócić do komunikatu, który już zniknął jako toast."""
    def __init__(self, parent=None):
        super().__init__(parent); self.hide()
        self.setObjectName("PanelPowiadomien")
        self.setFixedSize(340, 420)
        self.is_dark = True
        root = QVBoxLayout(self); root.setContentsMargins(14, 12, 14, 12); root.setSpacing(8)
        gora = QHBoxLayout()
        self.tytul = QLabel("Powiadomienia")
        gora.addWidget(self.tytul); gora.addStretch()
        self.btn_wyczysc = QPushButton("wyczyść")
        self.btn_wyczysc.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_wyczysc.clicked.connect(self._wyczysc)
        gora.addWidget(self.btn_wyczysc)
        root.addLayout(gora)
        self.scroll = QScrollArea(); self.scroll.setWidgetResizable(True)
        self.scroll.setFixedHeight(360)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.wnetrze = QWidget(); self.lay = QVBoxLayout(self.wnetrze)
        self.lay.setContentsMargins(2, 2, 2, 2); self.lay.setSpacing(6); self.lay.addStretch()
        self.scroll.setWidget(self.wnetrze)
        root.addWidget(self.scroll)
        sh = QGraphicsDropShadowEffect(self); sh.setBlurRadius(28); sh.setColor(QColor(0,0,0,160)); sh.setOffset(0, 8)
        self.setGraphicsEffect(sh)
        self._historia_ref = None

    def podepnij_historie(self, lista_historii):
        self._historia_ref = lista_historii

    def _wyczysc(self):
        if self._historia_ref is not None:
            self._historia_ref.clear()
        self._przerysuj()

    def odswiez(self):
        self._przerysuj()

    def _przerysuj(self):
        while self.lay.count() > 1:
            it = self.lay.takeAt(0)
            w = it.widget()
            if w: w.deleteLater()
        hist = self._historia_ref or []
        if not hist:
            pusty = QLabel("Brak powiadomień."); pusty.setObjectName("PowPusty")
            pusty.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.lay.insertWidget(0, pusty)
        else:
            for wpis in hist:
                self.lay.insertWidget(self.lay.count() - 1, self._wiersz(wpis))
        self._zastosuj_style()

    def _wiersz(self, wpis):
        w = QFrame(); w.setObjectName("PowWiersz")
        akcent = "#10B981" if wpis["success"] else "#EF4444"
        h = QVBoxLayout(w); h.setContentsMargins(12, 8, 12, 8); h.setSpacing(2)
        gora = QHBoxLayout()
        kropka = QLabel("●"); kropka.setStyleSheet(f"color:{akcent}; background:transparent; border:none; font-size:10px;")
        tyt = QLabel(wpis["title"]); tyt.setObjectName("PowTyt")
        czas = QLabel(wpis["czas"]); czas.setObjectName("PowCzas")
        gora.addWidget(kropka); gora.addWidget(tyt, 1); gora.addWidget(czas)
        h.addLayout(gora)
        opis = QLabel(wpis["desc"]); opis.setObjectName("PowOpis"); opis.setWordWrap(True)
        h.addWidget(opis)
        if wpis.get("klik_akcja"):
            w.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            akcja = wpis["klik_akcja"]
            def _klik(event, a=akcja):
                try: a()
                except Exception: pass
            w.mousePressEvent = _klik
        return w

    def update_theme(self, is_dark):
        self.is_dark = is_dark
        self._zastosuj_style()

    def _zastosuj_style(self):
        is_dark = self.is_dark
        if is_dark:
            tlo="rgba(11,19,32,0.98)"; ramka="rgba(0,240,255,0.25)"; txt="#F8FAFC"; txt_mut="#94A3B8"
            wiersz="rgba(255,255,255,0.05)"; akc="#00F0FF"
        else:
            tlo="#FFFFFF"; ramka="rgba(13,148,136,0.30)"; txt="#0F172A"; txt_mut="#64748B"
            wiersz="rgba(15,23,42,0.04)"; akc="#0D9488"
        self.setStyleSheet(f"#PanelPowiadomien {{ background:{tlo}; border:1px solid {ramka}; border-radius:14px; }}")
        self.tytul.setStyleSheet(f"color:{txt}; font-family:'Segoe UI'; font-size:13px; font-weight:800; background:transparent; border:none;")
        self.btn_wyczysc.setStyleSheet(f"QPushButton {{ color:{txt_mut}; background:transparent; border:none; font-size:11px; }} QPushButton:hover {{ color:{akc}; }}")
        self.scroll.setStyleSheet("background:transparent; border:none;")
        self.wnetrze.setStyleSheet("background:transparent;")
        for w in self.wnetrze.findChildren(QFrame, "PowWiersz"):
            w.setStyleSheet(f"#PowWiersz {{ background:{wiersz}; border-radius:9px; }}")
        for w in self.wnetrze.findChildren(QLabel, "PowTyt"):
            w.setStyleSheet(f"#PowTyt {{ color:{txt}; font-family:'Segoe UI'; font-size:12px; font-weight:700; background:transparent; border:none; }}")
        for w in self.wnetrze.findChildren(QLabel, "PowCzas"):
            w.setStyleSheet(f"#PowCzas {{ color:{txt_mut}; font-family:'Segoe UI'; font-size:10px; background:transparent; border:none; }}")
        for w in self.wnetrze.findChildren(QLabel, "PowOpis"):
            w.setStyleSheet(f"#PowOpis {{ color:{txt_mut}; font-family:'Segoe UI'; font-size:11px; background:transparent; border:none; }}")
        for w in self.wnetrze.findChildren(QLabel, "PowPusty"):
            w.setStyleSheet(f"#PowPusty {{ color:{txt_mut}; font-family:'Segoe UI'; font-size:12px; background:transparent; border:none; padding:30px; }}")


class DzwonekPowiadomien(QPushButton):
    """Ikona dzwoneczka w topbarze z czerwonym licznikiem nieprzeczytanych.
    Klik otwiera/zamyka PanelPowiadomien z historią ostatnich komunikatów."""
    def __init__(self, parent=None):
        super().__init__(parent)
        # 44px — mniejszy rozmiar ucinał plakietkę z licznikiem
        self.setFixedSize(44, 44)
        self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self._licznik = 0
        self.is_dark = True

    def ustaw_licznik(self, n):
        self._licznik = n; self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        rect = self.rect()
        txt = QColor("#F8FAFC") if self.is_dark else QColor("#0F172A")
        # ikona lekko w lewo-dół, żeby plakietka miała miejsce w prawym górnym rogu
        cx = rect.center().x() - 1
        cy = rect.center().y() + 1

        # --- korpus dzwonka (wypełniony kształt: kopuła + rozszerzona podstawa) ---
        korpus = QPainterPath()
        korpus.moveTo(cx - 8.5, cy + 5.0)                       # lewy skraj podstawy
        korpus.quadTo(cx - 6.0, cy + 3.6, cx - 6.0, cy + 0.5)   # podjazd do boku
        korpus.lineTo(cx - 6.0, cy - 1.5)                       # lewy bok
        korpus.quadTo(cx - 6.0, cy - 8.0, cx, cy - 8.0)         # kopuła (lewa połowa)
        korpus.quadTo(cx + 6.0, cy - 8.0, cx + 6.0, cy - 1.5)   # kopuła (prawa połowa)
        korpus.lineTo(cx + 6.0, cy + 0.5)                       # prawy bok
        korpus.quadTo(cx + 6.0, cy + 3.6, cx + 8.5, cy + 5.0)   # rozszerzenie podstawy
        korpus.closeSubpath()                                   # domknięcie podstawy
        p.setPen(Qt.PenStyle.NoPen)
        p.setBrush(txt)
        p.drawPath(korpus)

        # uchwyt na czubku
        p.drawEllipse(QPointF(cx, cy - 9.2), 1.7, 1.7)
        # serce dzwonka (klapa pod podstawą)
        p.drawEllipse(QPointF(cx, cy + 7.6), 2.1, 2.1)

        # --- plakietka z licznikiem (mieści się w obrysie) ---
        if self._licznik > 0:
            akcent = QColor("#EF4444")
            r = 8.0
            bx = rect.right() - r - 2
            by = rect.top() + r + 2
            # obwódka w kolorze tła, żeby plakietka odcinała się od ikony
            p.setBrush(QColor("#0B1320") if self.is_dark else QColor("#FFFFFF"))
            p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(QPointF(bx, by), r + 1.6, r + 1.6)
            p.setBrush(akcent)
            p.drawEllipse(QPointF(bx, by), r, r)
            p.setPen(QColor("#FFFFFF"))
            p.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
            tekst = str(self._licznik) if self._licznik < 10 else "9+"
            p.drawText(QRectF(bx - r, by - r, 2 * r, 2 * r),
                       Qt.AlignmentFlag.AlignCenter, tekst)


class StageTimeline(QWidget):
    """Oś czasu generowania: ●───●───●───●───● z etapami zapalającymi się
    po kolei. Linia płynnie 'nalewa się' między węzłami, a etapy wchodzą
    w spokojnym tempie, żeby całość była przyjemna dla oka."""
    ETAPY = ["Walidacja", "Trasy", "PDF", "Mapa", "Gotowe"]

    def __init__(self, is_dark=True, parent=None):
        super().__init__(parent)
        self.is_dark = is_dark
        self._aktywny = -1        # index ostatniego ukończonego etapu
        self._fill = 0.0          # płynny postęp 0..(n-1) — pozycja czoła "nalewania"
        self._glow = 0.0
        self.setFixedHeight(52)
        self.setMinimumWidth(360)
        self._t = QTimer(self); self._t.setInterval(30); self._t.timeout.connect(self._tick)
        self._faza = 0.0
        # animacja płynnego nalewania linii do docelowego etapu
        self._fill_anim = QPropertyAnimation(self, b"_fill_prop", self)
        self._fill_anim.setEasingCurve(QEasingCurve.Type.InOutCubic)

    def _get_fill(self): return self._fill
    def _set_fill(self, v): self._fill = v; self.update()
    _fill_prop = pyqtProperty(float, _get_fill, _set_fill)

    def set_theme(self, is_dark):
        self.is_dark = is_dark; self.update()

    def reset(self):
        self._aktywny = -1; self._fill = 0.0
        self._fill_anim.stop(); self._t.stop(); self.update()

    def ustaw_etap(self, idx, czas_ms=650):
        """Zapala etap idx i płynnie nalewa linię do niego przez czas_ms."""
        self._aktywny = idx
        if not self._t.isActive():
            self._t.start()          # pulsowanie poświaty aktywnego węzła
        self._fill_anim.stop()
        self._fill_anim.setDuration(czas_ms)
        self._fill_anim.setStartValue(self._fill)
        self._fill_anim.setEndValue(float(idx))
        self._fill_anim.start()
        if idx >= len(self.ETAPY) - 1:
            # po dojechaniu do "Gotowe" zatrzymaj pulsowanie z lekkim opóźnieniem
            QTimer.singleShot(czas_ms + 200, self._t.stop)
        self.update()

    def _tick(self):
        self._faza = (self._faza + 0.06) % (2 * math.pi)
        self._glow = (math.sin(self._faza) + 1) / 2
        self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        n = len(self.ETAPY)
        akcent = QColor("#00E4A1") if self.is_dark else QColor("#0D9488")
        nieakt = QColor(255, 255, 255, 40) if self.is_dark else QColor(15, 23, 42, 40)
        txt_akt = QColor("#F8FAFC") if self.is_dark else QColor("#0F172A")
        txt_nieakt = QColor(148, 163, 184, 160)

        w = self.width(); pad = 46
        y = 16; r = 6
        span = w - 2 * pad
        xs = [pad + span * i / (n - 1) for i in range(n)]

        fill = self._fill    # płynna pozycja czoła (0..n-1)

        # linie łączące — każda wypełniana proporcjonalnie do przejścia czoła
        for i in range(n - 1):
            x0, x1 = xs[i] + r, xs[i+1] - r
            # tło segmentu
            pen = QPen(nieakt); pen.setWidthF(2.5)
            p.setPen(pen); p.drawLine(QPointF(x0, y), QPointF(x1, y))
            # wypełniona część: ile czoła "fill" wpadło w ten segment [i, i+1]
            seg = max(0.0, min(1.0, fill - i))
            if seg > 0:
                xf = x0 + (x1 - x0) * seg
                penf = QPen(akcent); penf.setWidthF(2.5); penf.setCapStyle(Qt.PenCapStyle.RoundCap)
                p.setPen(penf); p.drawLine(QPointF(x0, y), QPointF(xf, y))

        # węzły + etykiety
        for i, x in enumerate(xs):
            done = fill >= i + 0.55                      # węzeł "zaliczony" gdy czoło go wyraźnie minęło
            active = (not done) and (abs(fill - i) < 0.75) and i <= self._aktywny
            p.setPen(Qt.PenStyle.NoPen)
            if active:
                gr = QRadialGradient(QPointF(x, y), 15)
                gr.setColorAt(0, QColor(akcent.red(), akcent.green(), akcent.blue(), int(150 * self._glow)))
                gr.setColorAt(1, QColor(akcent.red(), akcent.green(), akcent.blue(), 0))
                p.setBrush(QBrush(gr)); p.drawEllipse(QPointF(x, y), 15, 15)
                p.setBrush(akcent); p.drawEllipse(QPointF(x, y), r + 2, r + 2)
            elif done:
                p.setBrush(akcent); p.drawEllipse(QPointF(x, y), r, r)
                pen = QPen(QColor("#04160F") if self.is_dark else QColor("#FFFFFF")); pen.setWidthF(1.8)
                pen.setCapStyle(Qt.PenCapStyle.RoundCap); pen.setJoinStyle(Qt.PenJoinStyle.RoundJoin); p.setPen(pen)
                p.drawLine(QPointF(x-2.5, y), QPointF(x-0.5, y+2)); p.drawLine(QPointF(x-0.5, y+2), QPointF(x+3, y-2.5))
                p.setPen(Qt.PenStyle.NoPen)
            else:
                p.setBrush(nieakt); p.drawEllipse(QPointF(x, y), r, r)

            podswietl = done or active
            f = QFont("Segoe UI", 8, QFont.Weight.DemiBold if podswietl else QFont.Weight.Normal)
            p.setFont(f)
            p.setPen(txt_akt if podswietl else txt_nieakt)
            p.drawText(QRectF(x - 40, y + 12, 80, 16), Qt.AlignmentFlag.AlignCenter, self.ETAPY[i])


class AssistantPanel(QFrame):
    """Panel po prawej — trzy sekcje: szacunki na żywo, checklista walidacji
    (AI Asystent) i karta bieżącej sesji. Wszystko aktualizowane w locie
    z pól formularza, jeszcze przed kliknięciem Generuj."""
    def __init__(self, is_dark=True, parent=None):
        super().__init__(parent)
        self.is_dark = is_dark
        self.setFixedWidth(250)
        self.setObjectName("AssistantPanel")
        lay = QVBoxLayout(self); lay.setContentsMargins(20, 22, 20, 20); lay.setSpacing(10)

        # ---- ASYSTENT KONTROLI (checklista walidacji) ----
        self.lbl_sec2 = QLabel("ASYSTENT KONTROLI")
        self.lbl_sec2.setMinimumHeight(24)
        lay.addWidget(self.lbl_sec2)
        lay.addSpacing(4)

        self.check_pesel = self._check_row("PESEL")
        self.check_kod   = self._check_row("Kod pocztowy")
        self.check_woj   = self._check_row("Województwo")
        self.check_kwota = self._check_row("Kwota realna")
        self.check_dni   = self._check_row("Dni robocze")
        for r in [self.check_pesel, self.check_kod, self.check_woj, self.check_kwota, self.check_dni]:
            lay.addWidget(r)

        lay.addStretch()

        self.lbl_gotowy = QLabel("Uzupełnij dane…")
        self.lbl_gotowy.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_gotowy.setWordWrap(True)
        self.lbl_gotowy.setMinimumHeight(44)
        lay.addWidget(self.lbl_gotowy)

        self.apply_theme(is_dark)

    def _wys_wiersza(self):
        """Wysokość wiersza wyliczona z realnej metryki fontu — rośnie razem
        ze skalowaniem systemowym (DPI/powiększenie w Windows), więc tekst
        nigdy się nie przycina niezależnie od ustawień komputera."""
        fm = QFontMetrics(QFont("Segoe UI", 12))
        return max(24, fm.height() + 10)

    def _metric_row(self, etykieta, wartosc):
        w = QWidget(); w.setStyleSheet("background: transparent;")
        w.setMinimumHeight(self._wys_wiersza())
        w.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        h = QHBoxLayout(w); h.setContentsMargins(0, 0, 0, 0)
        l = QLabel(etykieta); l.setObjectName("metricLabel")
        v = QLabel(wartosc); v.setObjectName("metricValue"); v.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        h.addWidget(l); h.addWidget(v, 1)
        w._val = v
        return w

    def _check_row(self, etykieta):
        w = QWidget(); w.setStyleSheet("background: transparent;")
        w.setMinimumHeight(self._wys_wiersza())
        w.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        h = QHBoxLayout(w); h.setContentsMargins(0, 0, 0, 0); h.setSpacing(8)
        icon = QLabel("○"); icon.setFixedWidth(16); icon.setObjectName("checkIcon")
        l = QLabel(etykieta); l.setObjectName("checkLabel")
        h.addWidget(icon); h.addWidget(l, 1)
        w._icon = icon
        return w

    def _separator(self):
        s = QFrame(); s.setFixedHeight(1); s.setObjectName("panelSep")
        return s

    def set_theme(self, is_dark):
        self.apply_theme(is_dark)

    def apply_theme(self, is_dark):
        self.is_dark = is_dark
        if is_dark:
            self.setStyleSheet("""
                QFrame#AssistantPanel { background-color: rgba(10, 18, 30, 0.35); border: 1px solid rgba(0,240,255,0.15); border-radius: 14px; }
                QLabel { background: transparent; border: none; }
                QLabel#metricLabel, QLabel#checkLabel { color: #94A3B8; font-family:'Segoe UI'; font-size: 12px; }
                QLabel#metricValue { color: #00E4A1; font-family:'Segoe UI'; font-size: 13px; font-weight: 700; }
                QLabel#checkIcon { color: #475569; font-size: 14px; }
                QFrame#panelSep { background-color: rgba(255,255,255,0.08); border: none; }
            """)
            self._sec_color = "#00F0FF"
        else:
            self.setStyleSheet("""
                QFrame#AssistantPanel { background-color: rgba(255,255,255,0.55); border: 1px solid rgba(148,163,184,0.45); border-radius: 14px; }
                QLabel { background: transparent; border: none; }
                QLabel#metricLabel, QLabel#checkLabel { color: #475569; font-family:'Segoe UI'; font-size: 12px; }
                QLabel#metricValue { color: #0D9488; font-family:'Segoe UI'; font-size: 13px; font-weight: 700; }
                QLabel#checkIcon { color: #94A3B8; font-size: 14px; }
                QFrame#panelSep { background-color: rgba(15,23,42,0.10); border: none; }
            """)
            self._sec_color = "#0D9488"
        sec_style = f"color: {self._sec_color}; font-family:'Segoe UI'; font-size: 11px; font-weight: 800; letter-spacing: 1.5px; background: transparent; border: none;"
        self.lbl_sec2.setStyleSheet(sec_style)
        self._odswiez_gotowy_label()

    def set_szacunki(self, km, dni, km_dzien, dokumenty):
        # Sekcja SZACOWANA TRASA została usunięta — metoda zostaje jako
        # bezpieczna zaślepka, żeby istniejące wywołania nie rzucały błędu.
        pass

    def set_check(self, ktory, stan):
        """stan: True=OK(zielony ✓), False=błąd(czerwony ✕), None=pusty(○)."""
        mapa = {"pesel": self.check_pesel, "kod": self.check_kod, "woj": self.check_woj,
                "kwota": self.check_kwota, "dni": self.check_dni}
        row = mapa.get(ktory)
        if not row: return
        ic = row._icon
        if stan is True:
            ic.setText("✓"); ic.setStyleSheet("color: #10B981; font-size: 14px; font-weight: bold; background: transparent; border: none;")
        elif stan is False:
            ic.setText("✕"); ic.setStyleSheet("color: #EF4444; font-size: 14px; font-weight: bold; background: transparent; border: none;")
        else:
            ic.setText("○"); ic.setStyleSheet(f"color: {'#475569' if self.is_dark else '#94A3B8'}; font-size: 14px; background: transparent; border: none;")

    def _stany(self):
        out = []
        for row in [self.check_pesel, self.check_kod, self.check_woj, self.check_kwota, self.check_dni]:
            out.append(row._icon.text())
        return out

    def _odswiez_gotowy_label(self):
        pass

    def set_gotowy(self, gotowy, komunikat=None):
        if gotowy:
            self.lbl_gotowy.setText("✓  Gotowy do generowania")
            self.lbl_gotowy.setStyleSheet("color: #10B981; font-family:'Segoe UI'; font-size: 12px; font-weight: 700; background: rgba(16,185,129,0.12); border: 1px solid rgba(16,185,129,0.35); border-radius: 8px; padding: 8px;")
        else:
            self.lbl_gotowy.setText(komunikat or "Uzupełnij dane…")
            c = "#94A3B8" if self.is_dark else "#475569"
            self.lbl_gotowy.setStyleSheet(f"color: {c}; font-family:'Segoe UI'; font-size: 11px; background: transparent; border: none; padding: 4px;")


class FramelessTitleBar(QWidget):
    def __init__(self, parent):
        super().__init__(parent); self.parent = parent; self.setFixedHeight(40); self._drag_pos = None
        l = QHBoxLayout(self); l.setContentsMargins(15, 0, 15, 0); l.setSpacing(10)
        self.t = QLabel("PMT Planer"); self.t.setStyleSheet("color: #64748B; font-family: 'Segoe UI', sans-serif; font-size: 11px; font-weight: 600; background: transparent;"); l.addWidget(self.t); l.addStretch()
        self.b_min = QPushButton("—", self); self.b_min.setFixedSize(24, 24); self.b_min.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.b_max = QPushButton("🗖", self); self.b_max.setFixedSize(24, 24); self.b_max.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.b_close = QPushButton("✕", self); self.b_close.setFixedSize(24, 24); self.b_close.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.b_min.clicked.connect(self.parent.showMinimized); self.b_max.clicked.connect(self.toggle_maximize); self.b_close.clicked.connect(self.parent.close)
        l.addWidget(self.b_min); l.addWidget(self.b_max); l.addWidget(self.b_close)
    def toggle_maximize(self):
        # Okno jest BEZRAMKOWE — standardowe showMaximized() rozciąga je na cały
        # ekran fizyczny i chowa dół za paskiem zadań. Dlatego maksymalizujemy
        # ręcznie do DOSTĘPNEGO obszaru ekranu (availableGeometry = bez paska zadań).
        if getattr(self.parent, "_recznie_zmaks", False):
            self.parent.showNormal()
            geo = getattr(self.parent, "_geo_przed_maks", None)
            if geo is not None:
                self.parent.setGeometry(geo)
            self.parent._recznie_zmaks = False
        else:
            self.parent._geo_przed_maks = self.parent.geometry()
            ekran = self.parent.screen() or QApplication.primaryScreen()
            dostepny = ekran.availableGeometry()   # obszar bez paska zadań
            self.parent.setGeometry(dostepny)
            self.parent._recznie_zmaks = True
    def mouseDoubleClickEvent(self, event): self.toggle_maximize()
    def update_theme(self, is_dark):
        color = "#94A3B8" if is_dark else "#475569"; self.t.setStyleSheet(f"color: {color}; font-family: 'Segoe UI', sans-serif; font-size: 11px; font-weight: 600; background: transparent;")
        btn_style = f"QPushButton {{ background: transparent; color: {color}; border: none; font-size: 14px; border-radius: 12px; }} QPushButton:hover {{ background: rgba(100,100,100,0.2); }}"
        self.b_min.setStyleSheet(btn_style); self.b_max.setStyleSheet(btn_style); self.b_close.setStyleSheet(f"QPushButton {{ background: transparent; color: {color}; border: none; font-size: 14px; border-radius: 12px; }} QPushButton:hover {{ background: #EF4444; color: white; }}")
    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton: self._drag_pos = event.globalPosition().toPoint()
    def mouseMoveEvent(self, event):
        if self._drag_pos and not self.parent.isMaximized(): self.parent.move(self.parent.pos() + event.globalPosition().toPoint() - self._drag_pos); self._drag_pos = event.globalPosition().toPoint()
    def mouseReleaseEvent(self, event): self._drag_pos = None

class SiatkaMiesiaca(QWidget):
    """Autorska siatka miesiąca rysowana w całości QPainterem — pełna kontrola
    nad wyglądem (QCalendarWidget jest sztywny i nieczytelny w ciemnym motywie).
    Każdy dzień to kafelek; dni z rozliczeniem świecą na zielono, z planem —
    bursztynowo, wolne — czerwono. Klik wybiera dzień (callback on_klik)."""
    def __init__(self, parent=None, on_klik=None):
        super().__init__(parent)
        self.on_klik = on_klik
        self.is_dark = True
        self.rok = datetime.datetime.now().year
        self.mies = datetime.datetime.now().month
        self.wybrany_dzien = datetime.datetime.now().day
        self.rozliczenia = {}   # {iso: wpis}
        self.wpisy = {}         # {iso: {notatka,plan,wolne}}
        self.plan_dni = {}      # {iso: {"wizyt": n, "zrobione": k, "km": x}}
        # WARSTWY — dwa silniki, dwa języki wizualne. Można je oglądać osobno.
        self.warstwa_wizyty = True      # trasy z planera (wypełnienie + odznaka)
        self.warstwa_delegacje = True   # delegacje z archiwum (wstążka w rogu)
        self.setMinimumSize(520, 420)
        self.setMouseTracking(True)
        self._hover_dzien = None
        self._rects = {}        # dzień -> QRectF (do trafień myszą)

    def ustaw_warstwy(self, wizyty=True, delegacje=True):
        self.warstwa_wizyty = bool(wizyty)
        self.warstwa_delegacje = bool(delegacje)
        self.update()

    def ustaw_miesiac(self, rok, mies):
        self.rok, self.mies = rok, mies; self.update()

    def ustaw_dane(self, rozliczenia, wpisy, plan_dni=None):
        self.rozliczenia = rozliczenia or {}
        self.wpisy = wpisy or {}
        if plan_dni is not None:
            self.plan_dni = plan_dni
        self.update()

    def _sciezka_kafelka(self, rect):
        """Zaokrąglony kształt kafelka — do przycinania wstążki delegacji,
        żeby nie wystawała poza róg."""
        sc = QPainterPath()
        sc.addRoundedRect(rect, 8, 8)
        return sc

    def _iso(self, dzien):
        return f"{self.rok:04d}-{self.mies:02d}-{dzien:02d}"

    def _dni_w_miesiacu(self):
        return calendar.monthrange(self.rok, self.mies)[1]

    def _pierwszy_dzien_tygodnia(self):
        # 0=poniedziałek
        return calendar.monthrange(self.rok, self.mies)[0]

    def mousePressEvent(self, e):
        for dzien, rect in self._rects.items():
            if rect.contains(QPointF(e.position())):
                self.wybrany_dzien = dzien
                if self.on_klik: self.on_klik(self.rok, self.mies, dzien)
                self.update()
                return

    def mouseMoveEvent(self, e):
        stary = self._hover_dzien
        self._hover_dzien = None
        for dzien, rect in self._rects.items():
            if rect.contains(QPointF(e.position())):
                self._hover_dzien = dzien; break
        if stary != self._hover_dzien: self.update()

    def leaveEvent(self, e):
        self._hover_dzien = None; self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        W, H = self.width(), self.height()

        if self.is_dark:
            txt = QColor("#F8FAFC"); txt_mut = QColor("#64748B"); txt_wknd = QColor("#F87171")
            kafel = QColor(255,255,255,10); kafel_hover = QColor(0,240,255,28)
            ramka_sel = QColor("#00F0FF"); dow_col = QColor("#94A3B8")
            akc_rozl = QColor("#00E4A1"); akc_plan = QColor("#F59E0B"); akc_wolne = QColor("#EF4444")
            dzis_ring = QColor(0,240,255,120)
        else:
            txt = QColor("#0F172A"); txt_mut = QColor("#CBD5E1"); txt_wknd = QColor("#DC2626")
            kafel = QColor(15,23,42,8); kafel_hover = QColor(13,148,136,26)
            ramka_sel = QColor("#0D9488"); dow_col = QColor("#475569")
            akc_rozl = QColor("#0D9488"); akc_plan = QColor("#D97706"); akc_wolne = QColor("#DC2626")
            dzis_ring = QColor(13,148,136,120)

        # nagłówki dni tygodnia
        dni_tyg = ["Pon","Wt","Śr","Czw","Pt","Sob","Ndz"]
        margin = 6
        naglowek_h = 26
        cols = 7
        cell_w = (W - 2*margin) / cols
        p.setFont(QFont("Segoe UI", 9, QFont.Weight.DemiBold))
        for c, nazwa in enumerate(dni_tyg):
            p.setPen(txt_wknd if c >= 5 else dow_col)
            r = QRectF(margin + c*cell_w, 0, cell_w, naglowek_h)
            p.drawText(r, Qt.AlignmentFlag.AlignCenter, nazwa)

        # siatka dni
        offset = self._pierwszy_dzien_tygodnia()
        ile = self._dni_w_miesiacu()
        rows = math.ceil((offset + ile) / 7)
        cell_h = (H - naglowek_h - margin) / rows
        self._rects = {}
        dzis = datetime.datetime.now()

        for dzien in range(1, ile + 1):
            idx = offset + dzien - 1
            col = idx % 7; row = idx // 7
            x = margin + col*cell_w; y = naglowek_h + row*cell_h
            pad = 3
            rect = QRectF(x+pad, y+pad, cell_w-2*pad, cell_h-2*pad)
            self._rects[dzien] = rect
            iso = self._iso(dzien)
            weekend = col >= 5

            # tło kafelka
            if dzien == self._hover_dzien:
                p.setBrush(kafel_hover)
            else:
                p.setBrush(kafel)
            p.setPen(Qt.PenStyle.NoPen)
            p.drawRoundedRect(rect, 8, 8)

            # === WARSTWA 1: TRASY WIZYT (planer) — wypełnienie kafelka ===
            wpis = self.wpisy.get(iso, {})
            plan_d = self.plan_dni.get(iso) if self.warstwa_wizyty else None
            ma_rozl = (iso in self.rozliczenia) if self.warstwa_delegacje else False
            if plan_d:
                komplet = plan_d["zrobione"] >= plan_d["wizyt"] > 0
                kol_p = akc_rozl if komplet else (QColor("#00F0FF") if self.is_dark else QColor("#0D9488"))
                gl = QColor(kol_p); gl.setAlpha(38)
                p.setBrush(gl); p.setPen(QPen(kol_p, 1.2))
                p.drawRoundedRect(rect, 8, 8)
            elif wpis.get("wolne"):
                gl = QColor(akc_wolne); gl.setAlpha(32); p.setBrush(gl); p.setPen(Qt.PenStyle.NoPen)
                p.drawRoundedRect(rect, 8, 8)
            elif wpis.get("plan"):
                gl = QColor(akc_plan); gl.setAlpha(32); p.setBrush(gl); p.setPen(Qt.PenStyle.NoPen)
                p.drawRoundedRect(rect, 8, 8)

            # === WARSTWA 2: DELEGACJE (Moje szlaki) — wstążka w prawym górnym rogu ===
            # Osobny obszar kafelka, więc dzień z trasą I delegacją pokazuje OBA.
            if ma_rozl:
                wst = QPainterPath()
                bok = 20.0
                rx, ry = rect.right(), rect.top()
                wst.moveTo(rx - bok, ry)
                wst.lineTo(rx, ry)
                wst.lineTo(rx, ry + bok)
                wst.closeSubpath()
                p.setBrush(akc_rozl); p.setPen(Qt.PenStyle.NoPen)
                p.setClipPath(self._sciezka_kafelka(rect))
                p.drawPath(wst)
                p.setClipping(False)
                p.setPen(QColor("#04121A"))
                p.setFont(QFont("Segoe UI", 6, QFont.Weight.Black))
                p.drawText(QRectF(rx - bok + 3, ry + 1, bok - 4, 11),
                           Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop, "zł")

            # dzisiejszy dzień — pierścień
            if dzien == dzis.day and self.mies == dzis.month and self.rok == dzis.year:
                p.setBrush(Qt.BrushStyle.NoBrush); p.setPen(QPen(dzis_ring, 2))
                p.drawRoundedRect(rect, 8, 8)

            # wybrany dzień — wyraźna ramka
            if dzien == self.wybrany_dzien:
                p.setBrush(Qt.BrushStyle.NoBrush); p.setPen(QPen(ramka_sel, 2.5))
                p.drawRoundedRect(rect, 8, 8)

            # numer dnia
            p.setPen(txt_wknd if weekend else txt)
            p.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold if (ma_rozl or dzien==self.wybrany_dzien) else QFont.Weight.Normal))
            num_r = QRectF(rect.x(), rect.y()+4, rect.width(), 22)
            p.drawText(num_r, Qt.AlignmentFlag.AlignCenter, str(dzien))

            # ODZNAKA WIZYT — ile zaplanowano tego dnia i ile już zrobione
            if plan_d:
                komplet = plan_d["zrobione"] >= plan_d["wizyt"] > 0
                kol_b = akc_rozl if komplet else (QColor("#00F0FF") if self.is_dark else QColor("#0D9488"))
                etykieta = ("✓" if komplet
                            else (f"{plan_d['zrobione']}/{plan_d['wizyt']}"
                                  if plan_d["zrobione"] else f"{plan_d['wizyt']}"))
                p.setFont(QFont("Segoe UI", 7, QFont.Weight.Black))
                szer = max(18, p.fontMetrics().horizontalAdvance(etykieta) + 10)
                bh = 13
                br = QRectF(rect.center().x() - szer/2, rect.bottom() - bh - 5, szer, bh)
                p.setBrush(kol_b); p.setPen(Qt.PenStyle.NoPen)
                p.drawRoundedRect(br, bh/2, bh/2)
                p.setPen(QColor("#04121A"))
                p.drawText(br, Qt.AlignmentFlag.AlignCenter, etykieta)

            # znaczniki na dole kafelka (kropki) — tylko gdy nie ma odznaki wizyt
            kropki = []
            if wpis.get("plan"): kropki.append(akc_plan)
            if wpis.get("wolne"): kropki.append(akc_wolne)
            if wpis.get("notatka") and not kropki: kropki.append(txt_mut)
            if kropki and not plan_d:
                total_w = len(kropki)*10
                sx = rect.center().x() - total_w/2 + 5
                sy = rect.bottom() - 9
                for kolor in kropki:
                    p.setBrush(kolor); p.setPen(Qt.PenStyle.NoPen)
                    p.drawEllipse(QPointF(sx, sy), 3, 3)
                    sx += 10


class CalendarOverlay(QFrame):
    """Kalendarz Wypraw — autorski, kreatywny terminarz. Po lewej własna siatka
    miesiąca (świecące kafelki), po prawej panel dnia ze szczegółami rozliczenia
    oraz notatką / planem / dniem wolnym. Dane prywatne per pracownik."""
    def __init__(self, parent=None, on_wybor_miesiaca=None, on_dane_uzytkownika=None):
        super().__init__(parent); self.hide()
        self.is_dark = True
        self._on_wybor_miesiaca = on_wybor_miesiaca
        self._on_dane = on_dane_uzytkownika
        self._on_otworz_plan = None      # callback: przejdź do Planu Wizyt
        self._wpisy = {}
        self._rozliczenia = {}
        self._plan = None
        self._plan_dni = {}
        self._warstwa = "all"
        self._wybrana_iso = None

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        self.karta = QFrame(self); self.karta.setObjectName("KalKarta")
        self.karta.setMinimumSize(880, 560)
        root.addWidget(self.karta, alignment=Qt.AlignmentFlag.AlignCenter)
        kl = QVBoxLayout(self.karta); kl.setContentsMargins(30, 26, 30, 26); kl.setSpacing(16)

        # nagłówek z nawigacją miesięcy
        gora = QHBoxLayout()
        tyt = QVBoxLayout(); tyt.setSpacing(2)
        self.tytul = QLabel("Kalendarz Wypraw")
        self.podtytul = QLabel("Twój terminarz — wyjazdy, notatki i dni wolne w jednym miejscu")
        tyt.addWidget(self.tytul); tyt.addWidget(self.podtytul)
        gora.addLayout(tyt); gora.addStretch()

        self.btn_prev = QPushButton("‹"); self.btn_prev.setFixedSize(38, 38)
        self.btn_next = QPushButton("›"); self.btn_next.setFixedSize(38, 38)
        self.lbl_miesiac = QLabel(""); self.lbl_miesiac.setFixedWidth(180)
        self.lbl_miesiac.setAlignment(Qt.AlignmentFlag.AlignCenter)
        for b in [self.btn_prev, self.btn_next]:
            b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_prev.clicked.connect(lambda: self._zmien_miesiac(-1))
        self.btn_next.clicked.connect(lambda: self._zmien_miesiac(1))
        gora.addWidget(self.btn_prev); gora.addWidget(self.lbl_miesiac); gora.addWidget(self.btn_next)
        gora.addSpacing(10)
        self.btn_x = QPushButton("✕"); self.btn_x.setFixedSize(38, 38)
        self.btn_x.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_x.clicked.connect(self.hide)
        gora.addWidget(self.btn_x)
        kl.addLayout(gora)

        # ciało: siatka (lewa) + panel dnia (prawa)
        cialo = QHBoxLayout(); cialo.setSpacing(20)
        # === PRZEŁĄCZNIK WARSTW ===
        # Dwa silniki = dwa rodzaje dni. Można oglądać razem albo osobno.
        warstwy = QHBoxLayout(); warstwy.setSpacing(8)
        self.lbl_warstwy = QLabel("Pokaż:")
        warstwy.addWidget(self.lbl_warstwy)
        self.btn_w_all = QPushButton("Wszystko"); self.btn_w_all.setCheckable(True); self.btn_w_all.setChecked(True)
        self.btn_w_wiz = QPushButton("▣  Wizyty"); self.btn_w_wiz.setCheckable(True)
        self.btn_w_del = QPushButton("◤  Delegacje"); self.btn_w_del.setCheckable(True)
        for b in (self.btn_w_all, self.btn_w_wiz, self.btn_w_del):
            b.setFixedHeight(28)
            b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            warstwy.addWidget(b)
        self.btn_w_all.clicked.connect(lambda: self._ustaw_warstwe("all"))
        self.btn_w_wiz.clicked.connect(lambda: self._ustaw_warstwe("wizyty"))
        self.btn_w_del.clicked.connect(lambda: self._ustaw_warstwe("delegacje"))
        warstwy.addStretch()
        kl.addLayout(warstwy)

        self.siatka = SiatkaMiesiaca(on_klik=self._dzien_klikniety)
        cialo.addWidget(self.siatka, 1)

        self.panel = QFrame(); self.panel.setObjectName("KalPanel"); self.panel.setFixedWidth(280)
        pl = QVBoxLayout(self.panel); pl.setContentsMargins(18, 18, 18, 18); pl.setSpacing(11)
        self.lbl_data = QLabel("Wybierz dzień"); self.lbl_data.setObjectName("KalData")
        pl.addWidget(self.lbl_data)
        self.karta_rozl = QFrame(); self.karta_rozl.setObjectName("KalRozlKarta")
        krl = QVBoxLayout(self.karta_rozl); krl.setContentsMargins(12, 10, 12, 10); krl.setSpacing(3)
        self.lbl_rozl_tyt = QLabel("✓ Delegacja wygenerowana"); self.lbl_rozl_tyt.setObjectName("KalRozlTyt")
        self.lbl_rozl_szcz = QLabel(""); self.lbl_rozl_szcz.setObjectName("KalRozlSzcz"); self.lbl_rozl_szcz.setWordWrap(True)
        krl.addWidget(self.lbl_rozl_tyt); krl.addWidget(self.lbl_rozl_szcz)
        pl.addWidget(self.karta_rozl); self.karta_rozl.hide()

        # --- KARTA TRASY: wizyty zaplanowane na ten dzień (z planera) ---
        self.karta_trasa = QFrame(); self.karta_trasa.setObjectName("KalTrasaKarta")
        ktl = QVBoxLayout(self.karta_trasa); ktl.setContentsMargins(12, 10, 12, 10); ktl.setSpacing(6)
        self.lbl_trasa_tyt = QLabel("Trasa dnia"); self.lbl_trasa_tyt.setObjectName("KalTrasaTyt")
        ktl.addWidget(self.lbl_trasa_tyt)
        self.lbl_trasa_meta = QLabel(""); self.lbl_trasa_meta.setObjectName("KalTrasaMeta")
        self.lbl_trasa_meta.setWordWrap(True)
        ktl.addWidget(self.lbl_trasa_meta)
        self.pasek_trasy = PasekPostepuDnia(0, 1, self.is_dark)
        ktl.addWidget(self.pasek_trasy)
        self.lbl_trasa_lista = QLabel(""); self.lbl_trasa_lista.setObjectName("KalTrasaLista")
        self.lbl_trasa_lista.setWordWrap(True)
        ktl.addWidget(self.lbl_trasa_lista)
        self.btn_otworz_plan = QPushButton("Otwórz w Planie Wizyt  →")
        self.btn_otworz_plan.setFixedHeight(32)
        self.btn_otworz_plan.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_otworz_plan.clicked.connect(self._otworz_plan_dnia)
        ktl.addWidget(self.btn_otworz_plan)
        pl.addWidget(self.karta_trasa); self.karta_trasa.hide()

        self.lbl_notatka = QLabel("NOTATKA"); self.lbl_notatka.setObjectName("KalSekcja")
        pl.addWidget(self.lbl_notatka)
        self.pole_notatka = QPlainTextEdit(); self.pole_notatka.setFixedHeight(110)
        self.pole_notatka.setPlaceholderText("Plan dnia, adresy klientów, uwagi…")
        pl.addWidget(self.pole_notatka)
        self.chk_plan = QCheckBox("Zaplanowany wyjazd")
        self.chk_wolne = QCheckBox("Dzień wolny")
        pl.addWidget(self.chk_plan); pl.addWidget(self.chk_wolne)
        pl.addStretch()
        self.btn_zapisz = QPushButton("Zapisz wpis"); self.btn_zapisz.setFixedHeight(42)
        self.btn_zapisz.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_zapisz.clicked.connect(self._zapisz_wpis)
        pl.addWidget(self.btn_zapisz)
        cialo.addWidget(self.panel)
        kl.addLayout(cialo, 1)

        # stopka: legenda + użyj miesiąca
        stopka = QHBoxLayout()
        self.legenda = QLabel("▣  trasa wizyt (planer)      ◤  delegacja (Moje szlaki)      ●  plan      ●  dzień wolny")
        stopka.addWidget(self.legenda); stopka.addStretch()
        self.btn_miesiac = QPushButton("Użyj tego miesiąca w generatorze  →")
        self.btn_miesiac.setFixedHeight(40); self.btn_miesiac.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_miesiac.clicked.connect(self._uzyj_miesiaca)
        stopka.addWidget(self.btn_miesiac)
        kl.addLayout(stopka)

        self.panel.setEnabled(False)
        self._odswiez_naglowek()

    _MIES_PL = ["", "Styczeń","Luty","Marzec","Kwiecień","Maj","Czerwiec",
                "Lipiec","Sierpień","Wrzesień","Październik","Listopad","Grudzień"]

    def _odswiez_naglowek(self):
        self.lbl_miesiac.setText(f"{self._MIES_PL[self.siatka.mies]} {self.siatka.rok}")

    def _zmien_miesiac(self, delta):
        m = self.siatka.mies + delta; r = self.siatka.rok
        if m < 1: m = 12; r -= 1
        elif m > 12: m = 1; r += 1
        self.siatka.ustaw_miesiac(r, m)
        self._odswiez_naglowek()

    def _ustaw_warstwe(self, tryb):
        """Wszystko / tylko wizyty (planer) / tylko delegacje (Moje szlaki)."""
        self.btn_w_all.setChecked(tryb == "all")
        self.btn_w_wiz.setChecked(tryb == "wizyty")
        self.btn_w_del.setChecked(tryb == "delegacje")
        self._warstwa = tryb
        self.siatka.ustaw_warstwy(
            wizyty=(tryb in ("all", "wizyty")),
            delegacje=(tryb in ("all", "delegacje")))
        if self._wybrana_iso:      # panel dnia też respektuje warstwę
            try:
                y, m, d = map(int, self._wybrana_iso.split("-"))
                self._dzien_klikniety(y, m, d)
            except Exception:
                pass

    def resizeEvent(self, e):
        """Karta rośnie z oknem (jak planer i plan wizyt)."""
        super().resizeEvent(e)
        M = 26
        w = max(880, self.width() - 2 * M)
        h = max(560, self.height() - 2 * M)
        self.karta.setFixedSize(w, h)

    def odswiez_dane(self):
        imie = pesel = ""
        if self._on_dane: imie, pesel = self._on_dane()
        self._imie, self._pesel = imie, pesel
        self._wpisy = wczytaj_kalendarz(imie, pesel)
        self._rozliczenia = kalendarz_z_historii(imie, pesel)
        # PLAN WIZYT — synchronizacja z planerem: każdy dzień planu trafia
        # do kalendarza z liczbą wizyt i postępem odhaczenia
        self._plan = wczytaj_plan()
        self._plan_dni = {}
        if self._plan:
            for d in self._plan.get("dni", []):
                zrobione = sum(1 for wz in d.wizyty
                               if czy_odwiedzona(d.data, wz.adres or wz.nazwa))
                self._plan_dni[d.data.isoformat()] = {
                    "wizyt": len(d.wizyty), "zrobione": zrobione,
                    "km": d.km, "minuty": d.minuty, "dzien": d,
                }
        self.siatka.ustaw_dane(self._rozliczenia, self._wpisy, self._plan_dni)
        # skocz na DZIŚ, jeśli plan obejmuje bieżący miesiąc; inaczej na
        # miesiąc ostatniej delegacji
        dzis = datetime.date.today()
        if self._plan_dni:
            self.siatka.ustaw_miesiac(dzis.year, dzis.month)
            self._odswiez_naglowek()
        elif self._rozliczenia:
            iso = sorted(self._rozliczenia.keys())[-1]
            try:
                y, m, _ = map(int, iso.split("-"))
                self.siatka.ustaw_miesiac(y, m); self._odswiez_naglowek()
            except Exception: pass

    def _dzien_klikniety(self, rok, mies, dzien):
        self._wybrana_iso = f"{rok:04d}-{mies:02d}-{dzien:02d}"
        self.panel.setEnabled(True)
        self.lbl_data.setText(f"{dzien} {self._MIES_PL[mies]} {rok}")

        # --- TRASA DNIA (z planu wizyt) ---
        warstwa = getattr(self, "_warstwa", "all")
        pd = getattr(self, "_plan_dni", {}).get(self._wybrana_iso)
        if warstwa == "delegacje":
            pd = None
        if pd:
            d = pd["dzien"]
            h = int(pd["minuty"] // 60); m = int(pd["minuty"] % 60)
            self.lbl_trasa_meta.setText(
                f"{pd['wizyt']} wizyt  •  {pd['km']:.0f} km  •  ok. {h}h {m}min")
            self.pasek_trasy.zrobione = pd["zrobione"]
            self.pasek_trasy.wszystkie = max(1, pd["wizyt"])
            self.pasek_trasy.is_dark = self.is_dark
            self.pasek_trasy.update()
            # lista punktów (odhaczone przekreślone)
            linie = []
            for i, wz in enumerate(d.wizyty[:8], 1):
                nazwa = wz.nazwa or wz.miasto
                if czy_odwiedzona(d.data, wz.adres or wz.nazwa):
                    linie.append(f"✓ {i}. {nazwa}")
                else:
                    linie.append(f"○ {i}. {nazwa}")
            if len(d.wizyty) > 8:
                linie.append(f"… i {len(d.wizyty) - 8} więcej")
            self.lbl_trasa_lista.setText("\n".join(linie))
            self.karta_trasa.show()
        else:
            self.karta_trasa.hide()

        r = self._rozliczenia.get(self._wybrana_iso)
        if warstwa == "wizyty":
            r = None
        if r:
            self.lbl_rozl_szcz.setText(f"{r.get('kwota','')} zł  •  {r.get('dokumenty','?')} dok.  •  {r.get('woj','')}")
            self.karta_rozl.show()
        else:
            self.karta_rozl.hide()
        w = self._wpisy.get(self._wybrana_iso, {})
        self.pole_notatka.setPlainText(w.get("notatka", ""))
        self.chk_plan.setChecked(bool(w.get("plan")))
        self.chk_wolne.setChecked(bool(w.get("wolne")))

    def _otworz_plan_dnia(self):
        """Przechodzi z kalendarza do Planu Wizyt na wybranym dniu."""
        if self._on_otworz_plan and self._wybrana_iso:
            self._on_otworz_plan(self._wybrana_iso)

    def _zapisz_wpis(self):
        if not self._wybrana_iso: return
        if not (getattr(self, "_imie", "") and getattr(self, "_pesel", "")):
            self.lbl_data.setText("⚠ Wpisz imię i PESEL w formularzu")
            return
        zapisz_wpis_kalendarza(self._imie, self._pesel, self._wybrana_iso,
            self.pole_notatka.toPlainText().strip(),
            self.chk_wolne.isChecked(), self.chk_plan.isChecked())
        self._wpisy = wczytaj_kalendarz(self._imie, self._pesel)
        self.siatka.ustaw_dane(self._rozliczenia, self._wpisy)

    def _uzyj_miesiaca(self):
        if self._on_wybor_miesiaca:
            from PyQt6.QtCore import QDate
            self._on_wybor_miesiaca(QDate(self.siatka.rok, self.siatka.mies, 1))
        self.hide()

    def update_theme(self, is_dark):
        self.is_dark = is_dark
        self.siatka.is_dark = is_dark; self.siatka.update()
        if is_dark:
            tlo="rgba(5,10,20,0.90)"; karta="rgba(11,19,32,0.98)"; ramka="rgba(0,240,255,0.25)"
            akc="#00F0FF"; txt="#F8FAFC"; txt_mut="#94A3B8"; pole="rgba(5,10,20,0.5)"; sec="rgba(255,255,255,0.08)"
            rozl_bg="rgba(0,228,161,0.12)"; rozl_br="rgba(0,228,161,0.4)"
        else:
            tlo="rgba(226,232,240,0.92)"; karta="#FFFFFF"; ramka="rgba(13,148,136,0.30)"
            akc="#0D9488"; txt="#0F172A"; txt_mut="#475569"; pole="rgba(255,255,255,0.85)"; sec="rgba(15,23,42,0.08)"
            rozl_bg="rgba(13,148,136,0.10)"; rozl_br="rgba(13,148,136,0.35)"

        self.setStyleSheet(f"CalendarOverlay {{ background-color:{tlo}; }}")
        self.karta.setStyleSheet(f"#KalKarta {{ background-color:{karta}; border:1px solid {ramka}; border-radius:18px; }}")
        self.tytul.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:22px; font-weight:800; background:transparent; border:none;")
        self.podtytul.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:12px; background:transparent; border:none;")
        self.lbl_miesiac.setStyleSheet(f"color:{txt}; font-family:'Segoe UI'; font-size:15px; font-weight:700; background:transparent; border:none;")
        nav_css=(f"QPushButton {{ color:{txt}; background:{sec}; border:1px solid {ramka}; border-radius:10px; font-size:18px; font-weight:bold; }} QPushButton:hover {{ background:{ramka}; }}")
        self.btn_prev.setStyleSheet(nav_css); self.btn_next.setStyleSheet(nav_css)
        self.btn_x.setStyleSheet(f"QPushButton {{ color:{txt_mut}; background:transparent; border:none; font-size:18px; }} QPushButton:hover {{ color:{txt}; }}")
        self.panel.setStyleSheet(f"#KalPanel {{ background:{pole}; border:1px solid {ramka}; border-radius:14px; }}")
        self.lbl_data.setStyleSheet(f"#KalData {{ color:{akc}; font-family:'Segoe UI'; font-size:16px; font-weight:800; background:transparent; border:none; }}")
        self.karta_rozl.setStyleSheet(f"#KalRozlKarta {{ background:{rozl_bg}; border:1px solid {rozl_br}; border-radius:10px; }}")
        # karta trasy (wizyty z planera)
        akc_t = "#00F0FF" if is_dark else "#0D9488"
        zie_t = "#00E4A1" if is_dark else "#059669"
        txt_t = "#F8FAFC" if is_dark else "#0F172A"
        mut_t = "#94A3B8" if is_dark else "#64748B"
        tra_bg = "rgba(0,240,255,0.08)" if is_dark else "rgba(13,148,136,0.07)"
        tra_br = "rgba(0,240,255,0.30)" if is_dark else "rgba(13,148,136,0.30)"
        self.karta_trasa.setStyleSheet(f"#KalTrasaKarta {{ background:{tra_bg}; border:1px solid {tra_br}; border-radius:10px; }}")
        self.lbl_trasa_tyt.setStyleSheet(f"#KalTrasaTyt {{ color:{akc_t}; font-family:'Segoe UI'; font-size:12px; font-weight:800; background:transparent; border:none; }}")
        self.lbl_trasa_meta.setStyleSheet(f"#KalTrasaMeta {{ color:{mut_t}; font-family:'Segoe UI'; font-size:10px; background:transparent; border:none; }}")
        self.lbl_trasa_lista.setStyleSheet(f"#KalTrasaLista {{ color:{txt_t}; font-family:'Segoe UI'; font-size:10px; background:transparent; border:none; }}")
        self.btn_otworz_plan.setStyleSheet(
            f"QPushButton {{ color:#04121A; background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc_t}, stop:1 {zie_t}); "
            f"border:none; border-radius:8px; font-family:'Segoe UI'; font-size:10px; font-weight:800; }}")
        self.pasek_trasy.is_dark = is_dark
        self.pasek_trasy.update()
        # przełącznik warstw
        ramka_w = "rgba(0,240,255,0.30)" if is_dark else "rgba(13,148,136,0.35)"
        self.lbl_warstwy.setStyleSheet(
            f"color:{mut_t}; font-family:'Segoe UI'; font-size:11px; font-weight:600; background:transparent;")
        w_css = (
            f"QPushButton {{ color:{txt_t}; background:transparent; border:1px solid {ramka_w}; "
            f"border-radius:8px; padding:0 14px; font-family:'Segoe UI'; font-size:11px; font-weight:600; }} "
            f"QPushButton:hover {{ border-color:{akc_t}; color:{akc_t}; }} "
            f"QPushButton:checked {{ color:#04121A; background:{akc_t}; border-color:{akc_t}; font-weight:800; }}")
        d_css = (
            f"QPushButton {{ color:{txt_t}; background:transparent; border:1px solid {ramka_w}; "
            f"border-radius:8px; padding:0 14px; font-family:'Segoe UI'; font-size:11px; font-weight:600; }} "
            f"QPushButton:hover {{ border-color:{zie_t}; color:{zie_t}; }} "
            f"QPushButton:checked {{ color:#04121A; background:{zie_t}; border-color:{zie_t}; font-weight:800; }}")
        self.btn_w_all.setStyleSheet(w_css)
        self.btn_w_wiz.setStyleSheet(w_css)
        self.btn_w_del.setStyleSheet(d_css)
        self.lbl_rozl_tyt.setStyleSheet(f"#KalRozlTyt {{ color:{'#00E4A1' if is_dark else '#0D9488'}; font-family:'Segoe UI'; font-size:12px; font-weight:800; background:transparent; border:none; }}")
        self.lbl_rozl_szcz.setStyleSheet(f"#KalRozlSzcz {{ color:{txt}; font-family:'Segoe UI'; font-size:12px; background:transparent; border:none; }}")
        self.lbl_notatka.setStyleSheet(f"#KalSekcja {{ color:{txt_mut}; font-family:'Segoe UI'; font-size:11px; font-weight:700; letter-spacing:1px; background:transparent; border:none; }}")
        self.pole_notatka.setStyleSheet(f"QPlainTextEdit {{ background:{'rgba(0,0,0,0.3)' if is_dark else 'rgba(255,255,255,0.9)'}; color:{txt}; border:1px solid {ramka}; border-radius:10px; padding:8px; font-family:'Segoe UI'; font-size:13px; }}")
        chk_css=(f"QCheckBox {{ color:{txt}; font-family:'Segoe UI'; font-size:13px; spacing:8px; background:transparent; }}"
                 f"QCheckBox::indicator {{ width:18px; height:18px; border:1px solid {ramka}; border-radius:5px; background:{pole}; }}"
                 f"QCheckBox::indicator:checked {{ background:{akc}; }}")
        self.chk_plan.setStyleSheet(chk_css); self.chk_wolne.setStyleSheet(chk_css)
        self.btn_zapisz.setStyleSheet(f"QPushButton {{ background:{akc}; color:#04160F; border:none; border-radius:12px; font-weight:800; font-size:13px; }}")
        self.btn_miesiac.setStyleSheet(f"QPushButton {{ background:{sec}; color:{txt}; border:1px solid {ramka}; border-radius:10px; font-weight:600; font-size:12px; padding:0 16px; }} QPushButton:hover {{ background:{ramka}; }}")
        self.legenda.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:11px; background:transparent; border:none;")


class _PunktLekki:
    """Adapter: pozycja z listy planera (słownik) → obiekt czytelny dla
    cykl_dla_punktu (potrzebuje pól adres/siec/miasto)."""
    __slots__ = ("adres", "siec", "miasto", "nazwa")
    def __init__(self, poz):
        self.adres = (poz.get("adres", "") or "") if isinstance(poz, dict) else str(poz)
        self.siec = (poz.get("siec", "") or "") if isinstance(poz, dict) else ""
        self.miasto = (poz.get("miasto", "") or "") if isinstance(poz, dict) else ""
        self.nazwa = self.adres


class PlanerOverlay(QFrame):
    """Planer Nowej Wyprawy — elegancka, pełnoekranowa nakładka do ręcznego
    układania trasy. Pozwala dodawać przystanki, widzieć je jako oś podróży,
    liczyć szacunkowy dystans i czas, a docelowo wczytać listę z Excela.
    Wynik można przekazać do generatora jako punkt wyjścia."""
    def __init__(self, parent=None, on_zamknij=None, on_import=None):
        super().__init__(parent); self.hide()
        self._on_zamknij = on_zamknij
        self._on_import = on_import
        self._on_odswiez_plik = None    # callback: odśwież listę z pliku (rozpoznaj zmiany)
        self.is_dark = True
        self._przystanki = []      # lista dict: {adres, siec, miasto}
        self._sort = None          # None | 'siec' | 'miasto'
        self._on_zaplanuj = None   # callback: otwiera Plan Wizyt
        self._on_ostatni = None    # callback: otwiera ostatni zapisany plan
        self._on_toast = None      # callback: komunikat
        self._szukaj = ""          # fraza wyszukiwarki (filtr na żywo)
        self._sieci_akt = set()    # aktywne chipy sieci (puste = wszystkie)
        self._chipy_sieci = []     # widgety chipów (do przebudowy)

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        # panel-karta — responsywna (rozmiar w resizeEvent), by przy dużych
        # listach wykorzystać całą szerokość okna
        self.karta = QFrame(self); self.karta.setObjectName("PlanerKarta")
        self.karta.setMinimumWidth(680)
        root.addWidget(self.karta, alignment=Qt.AlignmentFlag.AlignCenter)

        kl = QVBoxLayout(self.karta); kl.setContentsMargins(34, 30, 34, 30); kl.setSpacing(18)

        # --- Nagłówek ---
        gora = QHBoxLayout()
        tytul_box = QVBoxLayout(); tytul_box.setSpacing(2)
        self.tytul = QLabel("Planer Nowej Wyprawy")
        self.podtytul = QLabel("Ułóż trasę krok po kroku albo wczytaj listę lokalizacji")
        tytul_box.addWidget(self.tytul); tytul_box.addWidget(self.podtytul)
        gora.addLayout(tytul_box); gora.addStretch()
        self.btn_x = QPushButton("✕"); self.btn_x.setFixedSize(38, 38)
        self.btn_x.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_x.clicked.connect(self._zamknij)
        gora.addWidget(self.btn_x, alignment=Qt.AlignmentFlag.AlignTop)
        kl.addLayout(gora)

        # --- Pasek dodawania przystanku ---
        # pasek dodawania — budowany tu, montowany w PRAWEJ kolumnie
        self.pole = GrubyKursorEdit(); self.pole.setPlaceholderText("Wpisz miejscowość i naciśnij Enter…")
        self.pole.setFixedHeight(40); self.pole.returnPressed.connect(self._dodaj_przystanek)
        self.btn_dodaj = QPushButton("＋ Dodaj"); self.btn_dodaj.setFixedHeight(38)
        self.btn_dodaj.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_dodaj.clicked.connect(self._dodaj_przystanek)
        self.btn_import = QPushButton("📄 Excel"); self.btn_import.setFixedHeight(38)
        self.btn_import.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_import.clicked.connect(self._importuj_excel)

        # ===== ZWIJANY PANEL OPCJI =====
        # Adres startowy, cykle i ustawienia zjadały ~100 px wysokości karty,
        # przez co lista punktów pokazywała ledwie kilka wierszy. Teraz są
        # schowane pod jednym przyciskiem — domyślnie zwinięte.
        # ===== UKŁAD DWUKOLUMNOWY =====
        # Pionowy układ się wyczerpał: każda nowa opcja odbierała wiersze liście.
        # Teraz LEWA kolumna = lista punktów (pełna wysokość), PRAWA = opcje.
        cialo = QHBoxLayout(); cialo.setSpacing(16)
        lewa = QVBoxLayout(); lewa.setSpacing(10)
        self.panel_opcje = QFrame(); self.panel_opcje.setObjectName("PanelOpcje")
        # panel jest PRZEWIJALNY — przy wielu sieciach/opcjach zawartość bywa
        # wyższa niż karta i bez tego elementy nachodziły na siebie
        self.panel_scroll = QScrollArea()
        self.panel_scroll.setWidgetResizable(True)
        self.panel_scroll.setFixedWidth(376)
        self.panel_scroll.setObjectName("PanelScroll")
        self.panel_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.panel_scroll.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Ignored)
        self.panel_scroll.setWidget(self.panel_opcje)
        po = QVBoxLayout(self.panel_opcje)
        po.setContentsMargins(16, 16, 16, 16); po.setSpacing(10)

        # --- SEKCJA: dodawanie punktów ---
        self.lbl_sek_dodaj = QLabel("Dodaj punkty")
        po.addWidget(self.lbl_sek_dodaj)
        po.addWidget(self.pole)
        rzad_dodaj = QHBoxLayout(); rzad_dodaj.setSpacing(8)
        rzad_dodaj.addWidget(self.btn_dodaj, 1)
        rzad_dodaj.addWidget(self.btn_import, 1)
        po.addLayout(rzad_dodaj)

        # "Odśwież z pliku" — rozpoznaje zmiany (nowe/zamknięte punkty)
        # względem obecnej listy, zamiast po cichu dokładać
        self.btn_odswiez_plik = QPushButton("🔄  Odśwież z pliku (rozpoznaj zmiany)")
        self.btn_odswiez_plik.setFixedHeight(32)
        self.btn_odswiez_plik.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_odswiez_plik.clicked.connect(self._klik_odswiez_plik)
        po.addWidget(self.btn_odswiez_plik)

        self._sep_a = QFrame(); self._sep_a.setFixedHeight(1)
        po.addWidget(self._sep_a)

        poo = QVBoxLayout(); poo.setSpacing(10)
        self.lbl_opcje_tyt = QLabel("Opcje trasy")
        poo.addWidget(self.lbl_opcje_tyt)

        # --- Punkt startowy (adres domowy) ---
        self.lbl_baza = QLabel("🏠  Startuję z:")
        poo.addWidget(self.lbl_baza)
        self.pole_baza = GrubyKursorEdit()
        self.pole_baza.setPlaceholderText("np. ul. Kwiatowa 5, Radom")
        self.pole_baza.setFixedHeight(34)
        self.pole_baza.setText(ustawienie("adres_bazy", ""))
        self.pole_baza.editingFinished.connect(self._zapisz_baze)
        poo.addWidget(self.pole_baza)

        self.btn_ustawienia = QPushButton("⚙  Ustawienia planowania")
        self.btn_ustawienia.setFixedHeight(34)
        self.btn_ustawienia.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_ustawienia.clicked.connect(self._otworz_ustawienia)
        poo.addWidget(self.btn_ustawienia)
        self.lbl_ustaw_skrot = QLabel("")
        self.lbl_ustaw_skrot.setWordWrap(True)
        poo.addWidget(self.lbl_ustaw_skrot)

        self._sep_opcje = QFrame(); self._sep_opcje.setFixedHeight(1)
        poo.addWidget(self._sep_opcje)

        # --- Cykliczność ---
        self.chk_cykl = QCheckBox("🔁  Powtarzaj cyklicznie")
        self.chk_cykl.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.chk_cykl.setChecked(bool(ustawienie("cykl_wlaczony", False)))
        self.chk_cykl.toggled.connect(self._zmien_cykl)
        poo.addWidget(self.chk_cykl)
        self.lbl_cykl_info = QLabel("")
        self.lbl_cykl_info.setWordWrap(True)
        poo.addWidget(self.lbl_cykl_info)

        self.btn_cykle = QPushButton("Ustaw cykle sieci…"); self.btn_cykle.setFixedHeight(32)
        self.btn_cykle.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_cykle.clicked.connect(self._otworz_cykle)
        poo.addWidget(self.btn_cykle)
        self.btn_cykl_widoczne = QPushButton("Cykl dla widocznych…")
        self.btn_cykl_widoczne.setFixedHeight(32)
        self.btn_cykl_widoczne.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_cykl_widoczne.clicked.connect(self._cykl_dla_widocznych)
        poo.addWidget(self.btn_cykl_widoczne)

        # SUGESTIE Z HISTORII — program zauważa naturalny rytm wizyt
        self.btn_sugestie_cyklu = QPushButton("💡  Sugestie z historii…")
        self.btn_sugestie_cyklu.setFixedHeight(32)
        self.btn_sugestie_cyklu.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_sugestie_cyklu.clicked.connect(self._pokaz_sugestie_cyklu)
        poo.addWidget(self.btn_sugestie_cyklu)

        hor = QHBoxLayout(); hor.setSpacing(8)
        self.lbl_horyzont = QLabel("Horyzont:")
        hor.addWidget(self.lbl_horyzont)
        self.combo_horyzont = QComboBox()
        for m in (3, 6, 12):
            self.combo_horyzont.addItem(f"{m} miesięcy", m)
        zapis_h = ustawienie("horyzont_mies", 3)
        idx_h = {3: 0, 6: 1, 12: 2}.get(int(zapis_h) if str(zapis_h).isdigit() else 3, 0)
        self.combo_horyzont.setCurrentIndex(idx_h)
        self.combo_horyzont.currentIndexChanged.connect(
            lambda: zapisz_ustawienie("horyzont_mies", self.combo_horyzont.currentData()))
        hor.addWidget(self.combo_horyzont, 1)
        poo.addLayout(hor)
        poo.addStretch()
        self._odswiez_cykl_info()

        # --- Wyszukiwarka na żywo (filtruje listę podczas pisania) ---
        self.pole_szukaj = GrubyKursorEdit()
        self.pole_szukaj.setPlaceholderText("🔎  Szukaj: sieć, miasto lub ulica…")
        self.pole_szukaj.setFixedHeight(40)
        self.pole_szukaj.textChanged.connect(self._zmien_szukaj)
        self.lbl_sek_filtr = QLabel("Filtruj")
        po.addWidget(self.lbl_sek_filtr)
        po.addWidget(self.pole_szukaj)

        # --- Chipy sieci (klik = pokaż tylko tę sieć; wielokrotny wybór) ---
        self.pasek_sieci = QWidget()
        self.pasek_sieci_l = QVBoxLayout(self.pasek_sieci)
        self.pasek_sieci_l.setContentsMargins(0, 0, 6, 0); self.pasek_sieci_l.setSpacing(6)
        self.pasek_sieci_l.setAlignment(Qt.AlignmentFlag.AlignLeft)
        self.scroll_sieci = QScrollArea()
        self.scroll_sieci.setWidgetResizable(True)
        # w wąskiej kolumnie chipy układają się PIONOWO i przewijają w pionie.
        # Minimum = ~4 chipy widoczne od razu (bez tego layout ściskał je do 1,5).
        self.scroll_sieci.setMinimumHeight(126)
        self.scroll_sieci.setMaximumHeight(150)
        self.scroll_sieci.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.scroll_sieci.setObjectName("PaskSieciScroll")
        self.scroll_sieci.setWidget(self.pasek_sieci)
        po.addWidget(self.scroll_sieci)

        # --- Pasek sortowania ---
        sortbar = QHBoxLayout(); sortbar.setSpacing(8)
        self.lbl_sort = QLabel("Sortuj:")
        sortbar.addWidget(self.lbl_sort)
        self.btn_sort_siec = QPushButton("wg sieci"); self.btn_sort_siec.setCheckable(True)
        self.btn_sort_miasto = QPushButton("wg miasta"); self.btn_sort_miasto.setCheckable(True)
        for b in (self.btn_sort_siec, self.btn_sort_miasto):
            b.setFixedHeight(30); b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_sort_siec.clicked.connect(lambda: self._ustaw_sort("siec"))
        self.btn_sort_miasto.clicked.connect(lambda: self._ustaw_sort("miasto"))
        sortbar.addWidget(self.btn_sort_siec); sortbar.addWidget(self.btn_sort_miasto)
        sortbar.addStretch()
        po.addLayout(sortbar)
        # PLANUJ TYLKO PRZEFILTROWANE — filtr działał dotąd wyłącznie na widok
        # listy; teraz może realnie zawęzić to, co trafi do planu.
        self.chk_tylko_widoczne = QCheckBox("Planuj tylko przefiltrowane")
        self.chk_tylko_widoczne.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.chk_tylko_widoczne.setChecked(bool(ustawienie("planuj_tylko_widoczne", False)))
        self.chk_tylko_widoczne.toggled.connect(self._zmien_tylko_widoczne)
        po.addWidget(self.chk_tylko_widoczne)
        self.lbl_zakres_planu = QLabel("")
        self.lbl_zakres_planu.setWordWrap(True)
        po.addWidget(self.lbl_zakres_planu)
        self._sep_b = QFrame(); self._sep_b.setFixedHeight(1)
        po.addWidget(self._sep_b)
        po.addLayout(poo)          # sekcja "Opcje trasy" (baza, ustawienia, cykl)
        po.addStretch()

        # --- Oś podróży (lista przystanków) ---
        self.lista_scroll = QScrollArea(); self.lista_scroll.setWidgetResizable(True)
        self.lista_scroll.setMinimumHeight(120); self.lista_scroll.setObjectName("PlanerScroll")
        # KLUCZOWE: QScrollArea z widgetResizable bierze rozmiar od zawartości
        # (przy setkach punktów — ogromny) i rozpychała layout, zasłaniając
        # stopkę. Polityka Ignored w pionie sprawia, że lista bierze dokładnie
        # tyle miejsca, ile zostanie — ani piksela więcej.
        self.lista_scroll.setSizePolicy(QSizePolicy.Policy.Expanding,
                                        QSizePolicy.Policy.Ignored)
        self.lista_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.lista_wnetrze = QWidget(); self.lista_wnetrze.setObjectName("PlanerListaWnetrze")
        self.lista_l = QVBoxLayout(self.lista_wnetrze); self.lista_l.setContentsMargins(6, 6, 6, 6); self.lista_l.setSpacing(5)
        self.lista_l.addStretch()
        self.lista_scroll.setWidget(self.lista_wnetrze)
        lewa.addWidget(self.lista_scroll, 1)

        # --- Pasek podsumowania ---
        self.podsumowanie = QLabel("Brak przystanków — dodaj pierwszą miejscowość powyżej.")
        self.podsumowanie.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lewa.addWidget(self.podsumowanie)

        # złóż kolumny: lista (elastyczna) + opcje (stała szerokość)
        cialo.addLayout(lewa, 1)
        cialo.addWidget(self.panel_scroll)
        kl.addLayout(cialo, 1)

        # --- Przyciski akcji ---
        akcje = QHBoxLayout(); akcje.setSpacing(12)
        self.btn_wyczysc = QPushButton("Wyczyść"); self.btn_wyczysc.setFixedHeight(44)
        self.btn_wyczysc.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_wyczysc.clicked.connect(self._wyczysc)
        akcje.addWidget(self.btn_wyczysc)
        # powrót do ostatniego planu (bez ponownego planowania) — widoczny
        # tylko wtedy, gdy jakiś plan jest zapisany na dysku
        self.btn_ostatni = QPushButton("📋  Ostatni plan"); self.btn_ostatni.setFixedHeight(44)
        self.btn_ostatni.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_ostatni.clicked.connect(lambda: self._on_ostatni() if self._on_ostatni else None)
        self.btn_ostatni.setVisible(False)
        akcje.addWidget(self.btn_ostatni)
        akcje.addStretch()
        self.btn_uzyj = QPushButton("Zaplanuj wizyty  →"); self.btn_uzyj.setFixedHeight(44)
        self.btn_uzyj.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_uzyj.clicked.connect(self._klik_zaplanuj)
        akcje.addWidget(self.btn_uzyj)
        kl.addLayout(akcje)

        self._puste_info()

    # ---------- logika ----------
    def _dodaj_przystanek(self):
        nazwa = self.pole.text().strip()
        if not nazwa: return
        # nie dubluj istniejącego adresu
        istnieje = any((p.get("adres", "") or "").strip().lower() == nazwa.lower()
                       for p in self._przystanki)
        if istnieje:
            if self._on_toast:
                self._on_toast("Już na liście", f"„{nazwa}” jest już wśród punktów.")
            self.pole.clear(); self.pole.setFocus()
            return
        self._przystanki.append({"adres": nazwa, "siec": "", "miasto": self._miasto_z_adresu(nazwa)})
        self.pole.clear(); self.pole.setFocus()
        self._zapisz_punkty()
        self._przerysuj()

    def _miasto_z_adresu(self, adres):
        """Wyciąga miasto z adresu 'ul. X 1, KOD Miasto' lub ostatni człon."""
        a = str(adres).strip()
        if "," in a:
            ogon = a.split(",")[-1].strip()
            # usuń kod pocztowy jeśli jest
            import re as _re
            ogon = _re.sub(r"^\d{2}-\d{3}\s*", "", ogon).strip()
            return ogon or a
        return a

    def _usun_przystanek(self, idx):
        if 0 <= idx < len(self._przystanki):
            self._przystanki.pop(idx)
            self._zapisz_punkty()
            self._przerysuj()

    def _wyczysc(self):
        self._przystanki = []; self._sort = None
        self.btn_sort_siec.setChecked(False); self.btn_sort_miasto.setChecked(False)
        self._zapisz_punkty()
        self._przerysuj()

    def _ustaw_sort(self, klucz):
        # przełącznik — ponowne kliknięcie wyłącza sortowanie
        if self._sort == klucz:
            self._sort = None
        else:
            self._sort = klucz
        self.btn_sort_siec.setChecked(self._sort == "siec")
        self.btn_sort_miasto.setChecked(self._sort == "miasto")
        if self._sort:
            self._przystanki.sort(key=lambda p: (str(p.get(self._sort, "")).lower(), str(p.get("adres", "")).lower()))
        self._przerysuj()

    # ---------- FILTROWANIE (wyszukiwarka + chipy sieci) ----------
    def _zmien_szukaj(self, txt):
        self._szukaj = str(txt or "").strip().lower()
        self._przerysuj()

    def _toggle_siec(self, siec):
        """Klik w chip sieci — dodaje/usuwa z filtra (wielokrotny wybór)."""
        if siec in self._sieci_akt:
            self._sieci_akt.discard(siec)
        else:
            self._sieci_akt.add(siec)
        self._przerysuj()

    def _wyczysc_filtry(self):
        self._szukaj = ""; self._sieci_akt = set()
        self.pole_szukaj.blockSignals(True)
        self.pole_szukaj.clear()
        self.pole_szukaj.blockSignals(False)
        self._przerysuj()

    def _widoczne(self):
        """Zwraca listę (indeks_oryginalny, pozycja) po zastosowaniu filtrów."""
        out = []
        for i, p in enumerate(self._przystanki):
            siec = str(p.get("siec", "") or "")
            # filtr chipów sieci (pusty zbiór = wszystkie)
            if self._sieci_akt and siec not in self._sieci_akt:
                continue
            # filtr wyszukiwarki — szuka w adresie, sieci i mieście
            if self._szukaj:
                hay = f"{p.get('adres','')} {siec} {p.get('miasto','')}".lower()
                if self._szukaj not in hay:
                    continue
            out.append((i, p))
        return out

    def _przebuduj_chipy_sieci(self):
        """Buduje pasek chipów z sieciami obecnymi w danych (+ liczniki)."""
        # wyczyść stare
        while self.pasek_sieci_l.count():
            it = self.pasek_sieci_l.takeAt(0)
            w = it.widget()
            if w: w.deleteLater()
        self._chipy_sieci = []
        # policz sieci
        licznik = {}
        for p in self._przystanki:
            s = str(p.get("siec", "") or "").strip()
            if s:
                licznik[s] = licznik.get(s, 0) + 1
        if not licznik:
            self.scroll_sieci.setVisible(False)
            return
        self.scroll_sieci.setVisible(True)
        akc = "#00E4A1" if self.is_dark else "#0D9488"
        txt = "#E2E8F0" if self.is_dark else "#1E293B"
        tlo_off = "rgba(255,255,255,0.06)" if self.is_dark else "rgba(15,23,42,0.05)"
        br_off = "rgba(255,255,255,0.15)" if self.is_dark else "rgba(15,23,42,0.15)"
        # sieci alfabetycznie
        for siec in sorted(licznik.keys(), key=lambda s: s.lower()):
            n = licznik[siec]
            aktywny = siec in self._sieci_akt
            b = QPushButton(f"{siec}  {n}")
            b.setCheckable(True); b.setChecked(aktywny)
            b.setFixedHeight(30)
            b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            if aktywny:
                b.setStyleSheet(f"QPushButton {{ color:#04121A; background:{akc}; border:1px solid {akc}; border-radius:15px; padding:0 14px; font-family:'Segoe UI'; font-size:11px; font-weight:800; }}")
            else:
                b.setStyleSheet(f"QPushButton {{ color:{txt}; background:{tlo_off}; border:1px solid {br_off}; border-radius:15px; padding:0 14px; font-family:'Segoe UI'; font-size:11px; font-weight:600; }} QPushButton:hover {{ border-color:{akc}; color:{akc}; }}")
            b.clicked.connect(lambda _=None, s=siec: self._toggle_siec(s))
            self.pasek_sieci_l.addWidget(b)
            self._chipy_sieci.append(b)
        # przycisk czyszczenia filtrów (gdy jakiś aktywny)
        if self._sieci_akt or self._szukaj:
            bc = QPushButton("✕ wyczyść filtry"); bc.setFixedHeight(30)
            bc.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            bc.setStyleSheet(f"QPushButton {{ color:#F87171; background:transparent; border:1px solid rgba(248,113,113,0.5); border-radius:15px; padding:0 14px; font-family:'Segoe UI'; font-size:11px; font-weight:700; }} QPushButton:hover {{ background:rgba(248,113,113,0.15); }}")
            bc.clicked.connect(self._wyczysc_filtry)
            self.pasek_sieci_l.addWidget(bc)
        self.pasek_sieci_l.addStretch()

    def _importuj_excel(self):
        if self._on_import:
            self._on_import(self)

    def _klik_odswiez_plik(self):
        if self._on_odswiez_plik:
            self._on_odswiez_plik(self)

    def dodaj_wiele(self, pozycje):
        """Wejście z importu — lista dict {adres, siec, miasto} albo stringów.
        DUPLIKATY (ten sam adres) są pomijane — inaczej sklep trafiłby do planu
        dwa razy. Zwraca liczbę pominiętych duplikatów."""
        istniejace = {(p.get("adres", "") or "").strip().lower()
                      for p in self._przystanki}
        dodane = 0
        duplikaty = 0
        for p in pozycje:
            if isinstance(p, dict):
                adres = str(p.get("adres", "")).strip()
                siec = str(p.get("siec", "")).strip()
                miasto = str(p.get("miasto", "")).strip()
            else:
                adres = str(p).strip(); siec = ""; miasto = ""
            if not adres:
                continue
            klucz = adres.lower()
            if klucz in istniejace:
                duplikaty += 1
                continue
            istniejace.add(klucz)
            self._przystanki.append({
                "adres": adres,
                "siec": siec,
                "miasto": miasto or self._miasto_z_adresu(adres)})
            dodane += 1
        if self._sort:
            self._przystanki.sort(key=lambda p: (str(p.get(self._sort, "")).lower(), str(p.get("adres", "")).lower()))
        self._zapisz_punkty()
        self._przerysuj()
        return duplikaty

    def _zapisz_punkty(self):
        """Utrwala listę punktów — przetrwa zamknięcie programu."""
        zapisz_punkty(self._przystanki)

    def wczytaj_zapisane(self):
        """Wczytuje listę punktów zapisaną w poprzedniej sesji."""
        zapisane = wczytaj_punkty()
        if zapisane:
            self._przystanki = zapisane
            self._przerysuj()
        return len(zapisane)

    def _zmien_tylko_widoczne(self, stan):
        zapisz_ustawienie("planuj_tylko_widoczne", bool(stan))
        self._odswiez_zakres_planu()

    def _do_zaplanowania(self):
        """Punkty, które trafią do planu — wszystkie albo tylko przefiltrowane."""
        if self.chk_tylko_widoczne.isChecked() and (self._sieci_akt or self._szukaj):
            return [poz for _, poz in self._widoczne()]
        return list(self._przystanki)

    def _odswiez_zakres_planu(self):
        """Mówi WPROST, co zostanie zaplanowane — żeby nie było niespodzianek."""
        wszystkie = len(self._przystanki)
        filtr_akt = bool(self._sieci_akt or self._szukaj)
        tylko = self.chk_tylko_widoczne.isChecked()
        ile = len(self._do_zaplanowania())
        zielony = "#00E4A1" if self.is_dark else "#059669"
        bursz = "#FBBF24" if self.is_dark else "#B45309"
        mut = "#94A3B8" if self.is_dark else "#64748B"
        if tylko and filtr_akt:
            self.lbl_zakres_planu.setText(f"▸ zaplanuję {ile} z {wszystkie} punktów (filtr aktywny)")
            self.lbl_zakres_planu.setStyleSheet(
                f"color:{bursz}; font-family:'Segoe UI'; font-size:10px; font-weight:700; background:transparent;")
        elif tylko:
            self.lbl_zakres_planu.setText(f"▸ brak filtra — zaplanuję wszystkie {wszystkie}")
            self.lbl_zakres_planu.setStyleSheet(
                f"color:{mut}; font-family:'Segoe UI'; font-size:10px; background:transparent;")
        else:
            self.lbl_zakres_planu.setText(f"▸ zaplanuję wszystkie {wszystkie} punktów")
            self.lbl_zakres_planu.setStyleSheet(
                f"color:{zielony}; font-family:'Segoe UI'; font-size:10px; background:transparent;")
        # przycisk też mówi, co zrobi
        if tylko and filtr_akt:
            self.btn_uzyj.setText(f"Zaplanuj {ile} wybranych  →")
        else:
            self.btn_uzyj.setText("Zaplanuj wizyty  →")

    def _klik_zaplanuj(self):
        """Otwiera Plan Wizyt na podstawie listy przystanków (nowy silnik)."""
        pozycje = self._do_zaplanowania()
        if self._on_zaplanuj and pozycje:
            self._on_zaplanuj(pozycje)

    def _przerysuj(self):
        # usuń stare wiersze (zostaw końcowy stretch)
        while self.lista_l.count() > 1:
            it = self.lista_l.takeAt(0)
            w = it.widget()
            if w: w.deleteLater()
        # odśwież chipy sieci
        self._przebuduj_chipy_sieci()

        if not self._przystanki:
            self._puste_info(); self._odswiez_podsumowanie(); return

        widoczne = self._widoczne()
        if not widoczne:
            info = QLabel("Brak punktów pasujących do filtrów.")
            info.setAlignment(Qt.AlignmentFlag.AlignCenter)
            kol = "#94A3B8" if self.is_dark else "#64748B"
            info.setStyleSheet(f"color:{kol}; font-family:'Segoe UI'; font-size:13px; background:transparent; padding:20px;")
            self.lista_l.insertWidget(0, info)
            self._odswiez_podsumowanie(); return

        # Przy bardzo długich listach rysujemy pierwsze N wierszy — reszta jest
        # dostępna po zawężeniu filtrem (inaczej tworzenie tysięcy widgetów
        # spowalnia interfejs).
        LIMIT = 300
        for licznik, (idx, poz) in enumerate(widoczne[:LIMIT]):
            self.lista_l.insertWidget(self.lista_l.count() - 1,
                                      self._wiersz(idx, poz, licznik, len(widoczne)))
        if len(widoczne) > LIMIT:
            reszta = len(widoczne) - LIMIT
            info = QLabel(f"…i jeszcze {reszta} — zawęź wyszukiwanie, aby je zobaczyć")
            info.setAlignment(Qt.AlignmentFlag.AlignCenter)
            kol = "#94A3B8" if self.is_dark else "#64748B"
            info.setStyleSheet(f"color:{kol}; font-family:'Segoe UI'; font-size:12px; background:transparent; padding:12px;")
            self.lista_l.insertWidget(self.lista_l.count() - 1, info)
        self._odswiez_podsumowanie()
        self._odswiez_zakres_planu()

    def _wiersz(self, idx, poz, poz_widoczna=0, ile_widocznych=0):
        adres = poz.get("adres", "") if isinstance(poz, dict) else str(poz)
        siec = poz.get("siec", "") if isinstance(poz, dict) else ""
        w = QFrame(); w.setObjectName("PlanerWiersz"); w.setFixedHeight(38)
        h = QHBoxLayout(w); h.setContentsMargins(14, 0, 10, 0); h.setSpacing(10)
        numer = QLabel(f"{idx+1}"); numer.setObjectName("PlanerNumer"); numer.setFixedSize(26, 26)
        numer.setAlignment(Qt.AlignmentFlag.AlignCenter)
        h.addWidget(numer)
        # sieć (chip) NA LEWO — zaczyna wiersz, stała szerokość by nie ucinać
        if siec:
            chip = QLabel(siec); chip.setObjectName("PlanerChip")
            chip.setAlignment(Qt.AlignmentFlag.AlignCenter)
            chip.setFixedWidth(96)
            akc = "#00E4A1" if self.is_dark else "#0D9488"
            tlo = "rgba(0,228,161,0.14)" if self.is_dark else "rgba(13,148,136,0.12)"
            chip.setStyleSheet(f"#PlanerChip {{ color:{akc}; background:{tlo}; border:1px solid {akc}; border-radius:9px; padding:3px 6px; font-family:'Segoe UI'; font-size:11px; font-weight:700; }}")
            h.addWidget(chip)
        # adres (główny)
        lbl = QLabel(adres); lbl.setObjectName("PlanerNazwa")
        h.addWidget(lbl, 1)
        # miasto — osobna kolumna (wykorzystuje przestrzeń przy szerokim oknie)
        miasto = poz.get("miasto", "") if isinstance(poz, dict) else ""
        if miasto:
            lm = QLabel(miasto); lm.setObjectName("PlanerMiasto")
            lm.setFixedWidth(150)
            lm.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
            kol_m = "#94A3B8" if self.is_dark else "#64748B"
            duze = czy_duze_miasto(miasto)
            lm.setText(("🏙  " if duze else "") + miasto)
            lm.setStyleSheet(f"#PlanerMiasto {{ color:{kol_m}; font-family:'Segoe UI'; font-size:11px; background:transparent; border:none; }}")
            h.addWidget(lm)
        # efektywny cykl (gdy tryb cykliczny włączony) — od razu widać rytm
        if self.chk_cykl.isChecked():
            _p = _PunktLekki(poz)
            tyg = cykl_dla_punktu(
                _p,
                cykl_domyslny=int(ustawienie("cykl_domyslny", 4) or 4),
                cykle_sieci=ustawienie("cykle_sieci", {}) or {},
                cykle_sieci_duze=ustawienie("cykle_sieci_duze", {}) or {},
                cykle_punktow=ustawienie("cykle_punktow", {}) or {})
            indywid = (poz.get("adres", "") or "").strip().lower() in (ustawienie("cykle_punktow", {}) or {})
            na_mies = 4 / tyg
            txt_c = f"{na_mies:.0f}×/mies." if na_mies >= 1 else f"co {tyg} tyg."
            lc = QLabel(("★ " if indywid else "") + txt_c)
            lc.setFixedWidth(84)
            lc.setAlignment(Qt.AlignmentFlag.AlignCenter)
            kol_c = "#00E4A1" if self.is_dark else "#059669"
            lc.setStyleSheet(
                f"color:{kol_c}; font-family:'Segoe UI'; font-size:10px; font-weight:700; "
                f"background:transparent; border:none;")
            lc.setToolTip("★ = cykl ustawiony indywidualnie dla tego punktu" if indywid
                          else "Cykl wynikający z reguły sieci")
            h.addWidget(lc)
        # ręczne przesuwanie: góra / dół (tylko gdy brak filtrów — inaczej
        # kolejność wizualna nie odpowiada faktycznej i przesuwanie myli)
        bez_filtrow = not self._szukaj and not self._sieci_akt
        if bez_filtrow:
            gora_btn = QPushButton("▲"); gora_btn.setFixedSize(24, 22); gora_btn.setObjectName("PlanerMove")
            dol_btn = QPushButton("▼"); dol_btn.setFixedSize(24, 22); dol_btn.setObjectName("PlanerMove")
            for b in (gora_btn, dol_btn):
                b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            gora_btn.setEnabled(idx > 0)
            dol_btn.setEnabled(idx < len(self._przystanki) - 1)
            gora_btn.clicked.connect(lambda _=None, i=idx: self._przesun(i, -1))
            dol_btn.clicked.connect(lambda _=None, i=idx: self._przesun(i, 1))
            akc_mv = "#00F0FF" if self.is_dark else "#0D9488"
            mv_css = (f"#PlanerMove {{ color:{akc_mv}; background:transparent; border:none; font-size:11px; }} "
                      f"#PlanerMove:hover {{ color:#FFFFFF; }} "
                      f"#PlanerMove:disabled {{ color:rgba(148,163,184,0.35); }}")
            gora_btn.setStyleSheet(mv_css); dol_btn.setStyleSheet(mv_css)
            h.addWidget(gora_btn); h.addWidget(dol_btn)
        usun = QPushButton("✕"); usun.setFixedSize(26, 26); usun.setObjectName("PlanerUsun")
        usun.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        usun.clicked.connect(lambda _=None, i=idx: self._usun_przystanek(i))
        h.addWidget(usun)
        return w

    def _przesun(self, idx, kier):
        """Przesuwa przystanek w górę (-1) lub w dół (+1). Ręczne sortowanie."""
        j = idx + kier
        if 0 <= idx < len(self._przystanki) and 0 <= j < len(self._przystanki):
            self._przystanki[idx], self._przystanki[j] = self._przystanki[j], self._przystanki[idx]
            # ręczne przesunięcie wyłącza sortowanie automatyczne
            self._sort = None
            self.btn_sort_siec.setChecked(False); self.btn_sort_miasto.setChecked(False)
            self._zapisz_punkty()
            self._przerysuj()

    def _puste_info(self):
        pusty = QLabel("🧭  Zacznij planować — wpisz pierwszą miejscowość powyżej\nlub wczytaj gotową listę z pliku Excel.")
        pusty.setObjectName("PlanerPusty"); pusty.setAlignment(Qt.AlignmentFlag.AlignCenter)
        pusty.setWordWrap(True)
        self.lista_l.insertWidget(0, pusty)

    def _odswiez_podsumowanie(self):
        n = len(self._przystanki)
        if n == 0:
            self.podsumowanie.setText("Brak przystanków — dodaj pierwszą miejscowość powyżej.")
            return
        widocznych = len(self._widoczne())
        filtr_akt = bool(self._szukaj or self._sieci_akt)
        if filtr_akt:
            # licznik filtrowania — użytkownik zawsze wie, gdzie jest
            self.podsumowanie.setText(
                f"Pokazuję {widocznych} z {n}  •  „Zaplanuj wizyty” ułoży trasy dla WSZYSTKICH {n} punktów")
        else:
            self.podsumowanie.setText(
                f"Wczytano {n} {'punkt' if n==1 else ('punkty' if 2<=n<=4 else 'punktów')}  •  kliknij „Zaplanuj wizyty”, aby ułożyć trasy")

    # ---------- CYKLICZNOŚĆ ----------
    def _zmien_cykl(self, stan):
        zapisz_ustawienie("cykl_wlaczony", bool(stan))
        self._odswiez_cykl_info()
        self._odswiez_skrot_opcji()
        self._przerysuj()          # pokaż/ukryj kolumnę cyklu przy punktach

    def _sieci_w_danych(self):
        """Lista sieci obecnych na liście punktów (alfabetycznie)."""
        s = set()
        for p in self._przystanki:
            n = str(p.get("siec", "") or "").strip()
            if n:
                s.add(n)
        return sorted(s, key=lambda x: x.lower())

    def _zapisz_baze(self):
        zapisz_ustawienie("adres_bazy", self.pole_baza.text().strip())

    def _odswiez_skrot_opcji(self):
        """Skrót aktualnych parametrów planowania (pod przyciskiem ustawień)."""
        UP = ustawienia_planowania()
        txt = (f"{UP['limit_minut']/60:.0f} h dziennie · {UP['czas_wizyty']:.0f} min na wizytę · "
               f"{UP['predkosc']:.0f} km/h")
        if UP["powrot"]:
            txt += " · z powrotem do bazy"
        self.lbl_ustaw_skrot.setText(txt)

    def _otworz_ustawienia(self):
        """Godziny pracy, czas wizyty, prędkość, powrót do bazy."""
        dlg = DialogUstawieniaPlanowania(self, is_dark=self.is_dark)
        if dlg.pokaz():
            _ustawienia_reset()
            UP = ustawienia_planowania()
            self._odswiez_cykl_info()
            self._odswiez_skrot_opcji()
            if self._on_toast:
                self._on_toast(
                    "Ustawienia zapisane",
                    f"{UP['limit_minut']/60:.0f} h dziennie · {UP['czas_wizyty']:.0f} min na wizytę · "
                    f"{UP['predkosc']:.0f} km/h"
                    + (" · z powrotem do bazy" if UP["powrot"] else ""))

    def _otworz_cykle(self):
        dlg = DialogCykle(self, sieci=self._sieci_w_danych(),
                          cykle=ustawienie("cykle_sieci", {}) or {},
                          cykle_duze=ustawienie("cykle_sieci_duze", {}) or {},
                          cykl_domyslny=int(ustawienie("cykl_domyslny", 4) or 4),
                          is_dark=self.is_dark)
        wynik = dlg.pokaz()
        if wynik is None:
            return
        dom, cykle, cykle_duze = wynik
        zapisz_ustawienie("cykl_domyslny", dom)
        zapisz_ustawienie("cykle_sieci", cykle)
        zapisz_ustawienie("cykle_sieci_duze", cykle_duze)
        self._odswiez_cykl_info()
        self._przerysuj()

    def _pokaz_sugestie_cyklu(self):
        """Analizuje dziennik odwiedzin i proponuje dopasowanie cyklu do
        rzeczywistego, zaobserwowanego rytmu."""
        sugestie = wykryj_sugestie_cyklu()
        if not sugestie:
            if self._on_toast:
                self._on_toast(
                    "Brak sugestii",
                    "Twój rytm już pasuje do ustawień, albo dziennik ma za mało "
                    "danych (potrzeba min. 3 odhaczonych wizyt w tym samym punkcie).")
            return
        dlg = DialogSugestieCyklu(self, sugestie, is_dark=self.is_dark)
        wynik = dlg.pokaz()
        if not wynik:
            return
        cykle_p = dict(ustawienie("cykle_punktow", {}) or {})
        for s in wynik:
            cykle_p[s["adres"]] = int(s["sugerowany_cykl_tyg"])
        zapisz_ustawienie("cykle_punktow", cykle_p)
        self._odswiez_cykl_info()
        self._przerysuj()
        if self._on_toast:
            n = len(wynik)
            slowo = "sugestię" if n == 1 else ("sugestie" if 2 <= n <= 4 else "sugestii")
            self._on_toast("Cykle zaktualizowane", f"Zastosowano {n} {slowo} z historii wizyt.")

    def _cykl_dla_widocznych(self):
        """Nadaje cykl WSZYSTKIM aktualnie widocznym (przefiltrowanym) punktom.
        Pozwala np. odfiltrować „Biedronka + Warszawa” i nadać im 4×/mies."""
        widoczne = self._widoczne()
        if not widoczne:
            return
        dlg = DialogCyklPunktow(self, len(widoczne), is_dark=self.is_dark)
        wynik = dlg.pokaz()
        if wynik is None:
            return
        cykle_p = dict(ustawienie("cykle_punktow", {}) or {})
        if wynik == "__USUN__":
            for _, poz in widoczne:
                cykle_p.pop((poz.get("adres", "") or "").strip().lower(), None)
            komunikat = f"Usunięto indywidualny cykl dla {len(widoczne)} punktów."
        else:
            for _, poz in widoczne:
                klucz = (poz.get("adres", "") or "").strip().lower()
                if klucz:
                    cykle_p[klucz] = int(wynik)
            na_mies = 4 / int(wynik)
            komunikat = (f"Ustawiono cykl dla {len(widoczne)} punktów: "
                         f"co {wynik} tyg. (~{na_mies:.0f}× w miesiącu).")
        zapisz_ustawienie("cykle_punktow", cykle_p)
        self._odswiez_cykl_info()
        self._przerysuj()
        if self._on_toast:
            self._on_toast("Cykl ustawiony", komunikat)

    def _odswiez_cykl_info(self):
        wl = self.chk_cykl.isChecked()
        self.btn_cykle.setEnabled(wl)
        self.btn_cykl_widoczne.setEnabled(wl)
        self.btn_sugestie_cyklu.setEnabled(wl)
        self.combo_horyzont.setEnabled(wl)
        if not wl:
            self.lbl_cykl_info.setText("plan jednorazowy — każdy punkt raz")
            return
        dom = int(ustawienie("cykl_domyslny", 4) or 4)
        cykle = ustawienie("cykle_sieci", {}) or {}
        if cykle:
            czesci = [f"{s}: co {t} tyg." for s, t in sorted(cykle.items())][:3]
            reszta = f" +{len(cykle)-3}" if len(cykle) > 3 else ""
            self.lbl_cykl_info.setText(f"domyślnie co {dom} tyg.  •  " + ", ".join(czesci) + reszta)
        else:
            self.lbl_cykl_info.setText(f"wszystkie punkty co {dom} tyg.")

    def _zamknij(self):
        if self._on_zamknij: self._on_zamknij(self._przystanki)
        self.hide()

    def resizeEvent(self, e):
        """Karta rośnie razem z oknem — przy 1000+ punktach każdy piksel się liczy."""
        super().resizeEvent(e)
        M = 24
        w = max(680, self.width() - 2 * M)
        h = max(520, self.height() - 2 * M)
        self.karta.setFixedSize(w, h)

    # ---------- motyw ----------
    def update_theme(self, is_dark):
        self.is_dark = is_dark
        if is_dark:
            tlo = "rgba(5, 10, 20, 0.90)"; karta = "rgba(11, 19, 32, 0.98)"
            ramka = "rgba(0, 240, 255, 0.25)"; akcent = "#00F0FF"; akcent2 = "#00E4A1"
            txt = "#F8FAFC"; txt_mut = "#94A3B8"; pole_bg = "rgba(5,10,20,0.5)"; wiersz_bg = "rgba(0,240,255,0.06)"
            btn_sec = "rgba(255,255,255,0.08)"; btn_sec_txt = "#F8FAFC"
        else:
            tlo = "rgba(226, 232, 240, 0.92)"; karta = "#FFFFFF"
            ramka = "rgba(13, 148, 136, 0.30)"; akcent = "#0D9488"; akcent2 = "#059669"
            txt = "#0F172A"; txt_mut = "#475569"; pole_bg = "rgba(255,255,255,0.8)"; wiersz_bg = "rgba(13,148,136,0.07)"
            btn_sec = "rgba(15,23,42,0.08)"; btn_sec_txt = "#0F172A"

        self.setStyleSheet(f"PlanerOverlay {{ background-color: {tlo}; }}")
        self.karta.setStyleSheet(f"#PlanerKarta {{ background-color: {karta}; border: 1px solid {ramka}; border-radius: 18px; }}")
        self.tytul.setStyleSheet(f"color:{akcent}; font-family:'Segoe UI'; font-size:22px; font-weight:800; background:transparent; border:none;")
        self.podtytul.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:12px; background:transparent; border:none;")
        self.btn_x.setStyleSheet(f"QPushButton {{ color:{txt_mut}; background:transparent; border:none; font-size:18px; }} QPushButton:hover {{ color:{txt}; }}")

        pole_css = (f"QLineEdit {{ background:{pole_bg}; border:1px solid {ramka}; border-radius:10px; "
                    f"padding:0 14px; color:{txt}; font-size:14px; font-family:'Segoe UI'; }}")
        self.pole.setStyleSheet(pole_css)
        if hasattr(self.pole, "set_caret_kolor"): self.pole.set_caret_kolor(akcent)
        # wyszukiwarka — nieco subtelniejsza od pola dodawania
        self.pole_szukaj.setStyleSheet(
            f"QLineEdit {{ background:{pole_bg}; border:1px solid {ramka}; border-radius:10px; "
            f"padding:0 14px; color:{txt}; font-size:13px; font-family:'Segoe UI'; }}")
        if hasattr(self.pole_szukaj, "set_caret_kolor"): self.pole_szukaj.set_caret_kolor(akcent)
        self.pole_baza.setStyleSheet(
            f"QLineEdit {{ background:{pole_bg}; border:1px solid {ramka}; border-radius:9px; "
            f"padding:0 12px; color:{txt}; font-size:12px; font-family:'Segoe UI'; }}")
        if hasattr(self.pole_baza, "set_caret_kolor"): self.pole_baza.set_caret_kolor(akcent)
        # panel opcji (prawa kolumna)
        self.btn_opcje = None
        panel_bg = "rgba(0,240,255,0.05)" if self.is_dark else "rgba(13,148,136,0.05)"
        self.panel_opcje.setStyleSheet(f"#PanelOpcje {{ background:transparent; border:none; }}")
        _uch = "rgba(0,240,255,0.35)" if self.is_dark else "rgba(13,148,136,0.35)"
        self.panel_scroll.setStyleSheet(
            f"#PanelScroll {{ background:{panel_bg}; border:1px solid {ramka}; border-radius:12px; }} "
            f"QScrollBar:vertical {{ background:transparent; width:8px; margin:6px 3px 6px 0; border:none; }} "
            f"QScrollBar::handle:vertical {{ background:{_uch}; border-radius:4px; min-height:30px; }} "
            f"QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height:0px; border:none; }} "
            f"QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{ background:transparent; }}")
        sek_css = f"color:{akcent}; font-family:'Segoe UI'; font-size:13px; font-weight:800; background:transparent;"
        self.lbl_opcje_tyt.setStyleSheet(sek_css)
        self.lbl_sek_dodaj.setStyleSheet(sek_css)
        self.lbl_sek_filtr.setStyleSheet(sek_css)
        for _s in (self._sep_a, self._sep_b):
            _s.setStyleSheet(f"background:{ramka}; border:none;")
        self.lbl_ustaw_skrot.setStyleSheet(
            f"color:{txt_mut}; font-family:'Segoe UI'; font-size:10px; background:transparent;")
        self._odswiez_skrot_opcji()
        self._sep_opcje.setStyleSheet(f"background:{ramka}; border:none;")

        # PASEK SORTOWANIA — wcześniej BEZ stylu, przez co czcionka była
        # systemowa (ciemna na ciemnym tle) i praktycznie niewidoczna.
        self.lbl_sort.setStyleSheet(
            f"color:{txt_mut}; font-family:'Segoe UI'; font-size:11px; font-weight:600; background:transparent;")
        sort_css = (
            f"QPushButton {{ color:{txt}; background:transparent; border:1px solid {ramka}; "
            f"border-radius:8px; padding:0 14px; font-family:'Segoe UI'; font-size:11px; font-weight:600; }} "
            f"QPushButton:hover {{ border-color:{akcent}; color:{akcent}; }} "
            f"QPushButton:checked {{ color:#04121A; background:{akcent}; border-color:{akcent}; }}")
        self.btn_sort_siec.setStyleSheet(sort_css)
        self.btn_sort_miasto.setStyleSheet(sort_css)

        self.lbl_baza.setStyleSheet(
            f"color:{txt}; font-family:'Segoe UI'; font-size:11px; font-weight:600; background:transparent;")
        self.btn_ustawienia.setStyleSheet(
            f"QPushButton {{ color:{akcent}; background:transparent; border:1px solid {ramka}; border-radius:9px; "
            f"padding:0 14px; font-family:'Segoe UI'; font-size:11px; font-weight:600; }} "
            f"QPushButton:hover {{ border-color:{akcent}; }}")
        # cykliczność
        self.chk_tylko_widoczne.setStyleSheet(
            f"QCheckBox {{ color:{txt}; font-family:'Segoe UI'; font-size:11px; font-weight:600; background:transparent; spacing:8px; }} "
            f"QCheckBox::indicator {{ width:15px; height:15px; border:1px solid {ramka}; border-radius:4px; background:{pole_bg}; }} "
            f"QCheckBox::indicator:checked {{ background:{akcent}; border-color:{akcent}; }}")
        self.chk_cykl.setStyleSheet(
            f"QCheckBox {{ color:{txt}; font-family:'Segoe UI'; font-size:12px; font-weight:600; background:transparent; spacing:8px; }} "
            f"QCheckBox::indicator {{ width:16px; height:16px; border:1px solid {ramka}; border-radius:5px; background:{pole_bg}; }} "
            f"QCheckBox::indicator:checked {{ background:{akcent}; border-color:{akcent}; }}")
        _btn_cyk_css = (
            f"QPushButton {{ color:{akcent}; background:transparent; border:1px solid {ramka}; border-radius:8px; "
            f"padding:0 12px; font-family:'Segoe UI'; font-size:11px; font-weight:600; }} "
            f"QPushButton:hover {{ border-color:{akcent}; }} "
            f"QPushButton:disabled {{ color:{txt_mut}; border-color:{ramka}; }}")
        self.btn_cykle.setStyleSheet(_btn_cyk_css)
        self.btn_cykl_widoczne.setStyleSheet(_btn_cyk_css)
        self.lbl_cykl_info.setStyleSheet(
            f"color:{txt_mut}; font-family:'Segoe UI'; font-size:11px; font-style:italic; background:transparent;")
        self.lbl_horyzont.setStyleSheet(
            f"color:{txt_mut}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
        self.combo_horyzont.setStyleSheet(
            f"QComboBox {{ background:{pole_bg}; border:1px solid {ramka}; border-radius:8px; padding:2px 8px; "
            f"color:{txt}; font-family:'Segoe UI'; font-size:11px; }} "
            f"QComboBox QAbstractItemView {{ background:{pole_bg}; color:{txt}; "
            f"selection-background-color:{akcent}; selection-color:#04121A; }} "
            f"QComboBox:disabled {{ color:{txt_mut}; }}")
        self.scroll_sieci.setStyleSheet(
            "#PaskSieciScroll { background:transparent; border:none; } "
            "QScrollBar:horizontal { height:0px; }")
        self.pasek_sieci.setStyleSheet("background:transparent;")

        btn_akc = (f"QPushButton {{ background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akcent}, stop:1 {akcent2}); "
                   f"color:#04160F; border:none; border-radius:10px; font-weight:800; font-size:13px; }} "
                   f"QPushButton:hover {{ opacity:0.9; }}")
        self.btn_dodaj.setStyleSheet(btn_akc)
        self.btn_uzyj.setStyleSheet(btn_akc.replace("border-radius:10px", "border-radius:12px; padding:0 20px"))
        btn_sec_css = (f"QPushButton {{ background:{btn_sec}; color:{btn_sec_txt}; border:1px solid {ramka}; "
                       f"border-radius:10px; font-weight:600; font-size:13px; }} QPushButton:hover {{ background:{ramka}; }}")
        self.btn_import.setStyleSheet(btn_sec_css)
        self.btn_odswiez_plik.setStyleSheet(
            f"QPushButton {{ color:{akcent}; background:transparent; border:1px solid {ramka}; border-radius:8px; "
            f"padding:0 10px; font-family:'Segoe UI'; font-size:10px; font-weight:600; text-align:left; }} "
            f"QPushButton:hover {{ border-color:{akcent}; }}")
        # Przyciski stopki — mają być WIDOCZNE (wcześniej ledwo się odcinały
        # od tła karty). Wypełnione tło + wyraźna ramka w kolorze akcentu.
        wyp = "rgba(0,240,255,0.10)" if self.is_dark else "rgba(13,148,136,0.10)"
        wyp_h = "rgba(0,240,255,0.22)" if self.is_dark else "rgba(13,148,136,0.20)"
        btn_akcja_css = (
            f"QPushButton {{ background:{wyp}; color:{akcent}; border:1.5px solid {akcent}; "
            f"border-radius:12px; padding:0 22px; font-family:'Segoe UI'; font-size:13px; font-weight:700; }} "
            f"QPushButton:hover {{ background:{wyp_h}; }}")
        self.btn_wyczysc.setStyleSheet(btn_akcja_css)
        self.btn_ostatni.setStyleSheet(btn_akcja_css)
        # Chipy sieci i wiersze mają kolory "wypalone" w chwili budowy — po
        # zmianie motywu trzeba je przebudować, inaczej zostają jasne litery
        # na jasnym tle (i odwrotnie).
        self._przerysuj()

        uchwyt = "rgba(0,240,255,0.35)" if self.is_dark else "rgba(13,148,136,0.35)"
        uchwyt_hover = "rgba(0,240,255,0.65)" if self.is_dark else "rgba(13,148,136,0.6)"
        self.lista_scroll.setStyleSheet(
            f"#PlanerScroll {{ background:transparent; border:1px solid {ramka}; border-radius:12px; }} "
            f"QScrollBar:vertical {{ background:transparent; width:10px; margin:6px 3px 6px 0; border:none; }} "
            f"QScrollBar::handle:vertical {{ background:{uchwyt}; border-radius:5px; min-height:36px; }} "
            f"QScrollBar::handle:vertical:hover {{ background:{uchwyt_hover}; }} "
            f"QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height:0px; border:none; background:transparent; }} "
            f"QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{ background:transparent; }}")
        self.lista_wnetrze.setStyleSheet("#PlanerListaWnetrze { background: transparent; }")
        self.podsumowanie.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:12px; font-weight:600; background:transparent; border:none;")

        # przerysuj wiersze z aktualnym motywem
        self._akcent = akcent; self._txt = txt; self._txt_mut = txt_mut; self._wiersz_bg = wiersz_bg; self._ramka = ramka
        self.setProperty("_style_wiersz", True)
        self._css_wiersz = (f"#PlanerWiersz {{ background:{wiersz_bg}; border:1px solid {ramka}; border-radius:10px; }}")
        self._css_numer = (f"#PlanerNumer {{ background:{akcent}; color:#04160F; border-radius:13px; font-weight:800; font-size:12px; }}")
        self._css_nazwa = (f"#PlanerNazwa {{ color:{txt}; font-family:'Segoe UI'; font-size:14px; background:transparent; border:none; }}")
        self._css_usun = (f"#PlanerUsun {{ color:{txt_mut}; background:transparent; border:none; font-size:14px; }} #PlanerUsun:hover {{ color:#EF4444; }}")
        self._css_pusty = (f"#PlanerPusty {{ color:{txt_mut}; font-family:'Segoe UI'; font-size:13px; background:transparent; border:none; padding:40px; }}")
        self._zastosuj_css_wierszy()

    def _zastosuj_css_wierszy(self):
        laczny = "\n".join([getattr(self, k, "") for k in
                            ["_css_wiersz", "_css_numer", "_css_nazwa", "_css_usun", "_css_pusty"]])
        self.lista_wnetrze.setStyleSheet("#PlanerListaWnetrze { background: transparent; }\n" + laczny)


class PobieranieAktualizacjiThread(QThread):
    def _raportuj(self, pobrano, calosc, czy_zip):
        """Pasek postępu — rozpakowanie zajmuje ostatnie ~15%."""
        gorna = 0.80 if czy_zip else 0.95
        if calosc:
            frakcja = 0.10 + (gorna - 0.10) * (pobrano / calosc)
            self.postep.emit(min(gorna, frakcja),
                             f"Pobieram… {pobrano/1048576:.1f} / {calosc/1048576:.1f} MB")
        else:
            self.postep.emit(0.35, f"Pobieram… {pobrano/1048576:.1f} MB")

    """Pobiera nową wersję do katalogu tymczasowego, raportując postęp.
    Obsługuje zarówno .zip (preferowane — patrz znajdz_plik_wydania) jak
    i goły .exe (zgodność wsteczna ze starszymi wydaniami)."""
    postep = pyqtSignal(float, str)      # 0..1, opis
    sukces = pyqtSignal(str)             # ścieżka gotowego pliku .exe
    blad = pyqtSignal(str)

    def run(self):
        try:
            self.postep.emit(0.05, "Szukam najnowszego wydania…")
            url, czy_zip = znajdz_plik_wydania()
            if not url:
                self.blad.emit(
                    "Nie znalazłem pliku programu w najnowszym wydaniu na GitHubie."
                    if CZY_WINDOWS else
                    "Najnowsze wydanie nie zawiera jeszcze pliku dla tego systemu"
                    " (macOS/Linux).")
                return

            pobrany = os.path.join(
                tempfile.gettempdir(),
                "PMT_Planer_pobrany.zip" if czy_zip
                else ("PMT_Planer_nowy.exe" if CZY_WINDOWS else "PMT_Planer_nowy"))
            if os.path.exists(pobrany):
                try: os.remove(pobrany)
                except Exception: pass

            self.postep.emit(0.10, "Łączę z serwerem…")
            # POBIERANIE Z WZNAWIANIEM. Plik ma ~48 MB, a firmowe sieci, VPN-y
            # i słabe łącza potrafią zerwać takie połączenie w połowie
            # ("Remote end closed connection without response"). Zamiast
            # zaczynać od zera, dociągamy brakującą część nagłówkiem Range —
            # do pięciu podejść.
            calosc = 0
            pobrano = 0
            ostatni_blad = None
            for podejscie in range(1, 6):
                try:
                    naglowki = {"User-Agent": "PMT-Planer",
                                "Accept-Encoding": "identity",   # bez kompresji: znany rozmiar
                                "Connection": "close"}
                    if pobrano:
                        naglowki["Range"] = f"bytes={pobrano}-"
                        self.postep.emit(min(0.75, 0.10 + 0.65 * (pobrano / max(calosc, 1))),
                                         f"Wznawiam pobieranie… (próba {podejscie})")
                    req = urllib.request.Request(url, headers=naglowki)
                    with urllib.request.urlopen(req, timeout=60) as resp:
                        dlugosc = int(resp.headers.get("Content-Length") or 0)
                        if not calosc:
                            calosc = dlugosc
                        tryb_pliku = "ab" if (pobrano and resp.status == 206) else "wb"
                        if tryb_pliku == "wb":
                            pobrano = 0
                        with open(pobrany, tryb_pliku) as f:
                            while True:
                                kawalek = resp.read(262144)     # 256 kB
                                if not kawalek:
                                    break
                                f.write(kawalek)
                                pobrano += len(kawalek)
                                self._raportuj(pobrano, calosc, czy_zip)
                    if calosc and pobrano < calosc:
                        raise IOError("połączenie przerwane w trakcie pobierania")
                    ostatni_blad = None
                    break
                except Exception as e:
                    ostatni_blad = e
                    if podejscie < 5:
                        time.sleep(2 * podejscie)
            if ostatni_blad is not None:
                raise ostatni_blad
            if os.path.getsize(pobrany) < 50 * 1024:      # sanity check
                self.blad.emit("Pobrany plik wygląda na uszkodzony (za mały).")
                return

            if czy_zip:
                self.postep.emit(0.85, "Rozpakowuję…")
                # Najpierw sprawdzamy, czy to nowe wydanie FOLDEROWE.
                folder_nowej = self._rozpakuj_folder(pobrany)
                if folder_nowej:
                    self.gotowe.emit(folder_nowej)
                    return
                cel = self._rozpakuj_exe(pobrany)
                if not cel:
                    self.blad.emit("W pobranym archiwum nie znalazłem pliku programu.")
                    return
            else:
                cel = pobrany

            if os.path.getsize(cel) < 500 * 1024:
                self.blad.emit("Rozpakowany plik programu wygląda na uszkodzony (za mały).")
                return

            if not CZY_WINDOWS:
                # binarka PyInstallera na uniksach musi mieć prawo wykonywania
                try: os.chmod(cel, 0o755)
                except Exception: pass
            self.postep.emit(1.0, "Gotowe — instaluję…")
            self.sukces.emit(cel)
        except Exception as e:
            self.blad.emit(f"Nie udało się pobrać aktualizacji: {e}")

    def _rozpakuj_folder(self, zip_path: str):
        """Nowe wydania to FOLDER z programem (plik .exe + katalog _internal).
        Zwraca sciezke do rozpakowanego folderu albo None, gdy archiwum jest
        starego typu (pojedynczy plik)."""
        try:
            with zipfile.ZipFile(zip_path, "r") as z:
                nazwy = [i.filename for i in z.infolist() if not i.is_dir()]
                # cecha wersji folderowej: katalog _internal obok programu
                if not any("_internal/" in n or "_internal\\" in n for n in nazwy):
                    return None
                cel = os.path.join(tempfile.gettempdir(), "PMT_Planer_nowa_wersja")
                if os.path.isdir(cel):
                    shutil.rmtree(cel, ignore_errors=True)
                os.makedirs(cel, exist_ok=True)
                z.extractall(cel)
            # jesli w archiwum byl jeden folder nadrzedny — wchodzimy do niego
            wpisy = [w for w in os.listdir(cel) if not w.startswith(".")]
            if len(wpisy) == 1 and os.path.isdir(os.path.join(cel, wpisy[0])):
                cel = os.path.join(cel, wpisy[0])
            nazwa_exe = os.path.basename(sys.executable)
            if not os.path.exists(os.path.join(cel, nazwa_exe)):
                # w paczce program moze nazywac sie inaczej niz plik uzytkownika
                for w in os.listdir(cel):
                    if w.lower().endswith(".exe") or (not CZY_WINDOWS and "." not in w):
                        try:
                            shutil.copy2(os.path.join(cel, w), os.path.join(cel, nazwa_exe))
                        except Exception:
                            pass
                        break
            try:
                os.remove(zip_path)
            except Exception:
                pass
            return cel
        except Exception:
            return None

    def _rozpakuj_exe(self, zip_path: str) -> str:
        """Wyciąga plik .exe z archiwum do katalogu tymczasowego. Jeśli
        archiwum zawiera WIĘCEJ NIŻ JEDEN plik .exe (np. przypadkiem dołączony
        inny plik), bierze NAJWIĘKSZY — prawdziwy program (Python + Qt w
        środku) waży dziesiątki MB, więc przypadkowy mniejszy plik nie ma
        szans go przebić. Archiwum po rozpakowaniu jest kasowane."""
        cel = ""
        try:
            with zipfile.ZipFile(zip_path, "r") as z:
                if CZY_WINDOWS:
                    kandydaci = [info for info in z.infolist()
                                if not info.is_dir() and info.filename.lower().endswith(".exe")]
                else:
                    # macOS/Linux: program to binarka bez rozszerzenia — bierzemy
                    # pliki poza oczywistymi nie-programami (reguła „największy
                    # plik wygrywa” niżej załatwia resztę)
                    pomijane = (".txt", ".md", ".png", ".jpg", ".ico", ".icns",
                                ".pdf", ".json", ".zip", ".exe")
                    kandydaci = [info for info in z.infolist()
                                if not info.is_dir()
                                and not info.filename.lower().endswith(pomijane)]
                if kandydaci:
                    najwiekszy = max(kandydaci, key=lambda i: i.file_size)
                    cel = os.path.join(tempfile.gettempdir(),
                                       "PMT_Planer_nowy.exe" if CZY_WINDOWS else "PMT_Planer_nowy")
                    with z.open(najwiekszy) as src, open(cel, "wb") as dst:
                        dst.write(src.read())
                    if not CZY_WINDOWS:
                        try: os.chmod(cel, 0o755)
                        except Exception: pass
        except Exception:
            cel = ""
        finally:
            try: os.remove(zip_path)
            except Exception: pass
        return cel


class AureolaAktualizacji(QWidget):
    """Tło okna aktualizacji — pulsujące kręgi i unoszące się iskry.
    Ma przyciągać wzrok i budzić chęć kliknięcia, bez krzykliwości."""
    def __init__(self, parent=None, is_dark=True):
        super().__init__(parent)
        self.is_dark = is_dark
        self._faza = 0.0
        self._iskry = []
        for i in range(14):
            self._iskry.append({
                "x": random.uniform(0.05, 0.95),
                "y": random.uniform(0.1, 1.0),
                "v": random.uniform(0.0016, 0.005),
                "r": random.uniform(1.2, 2.8),
                "a": random.uniform(0.25, 0.75),
            })
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._tick)
        self._timer.start(33)
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents)

    def _tick(self):
        self._faza += 0.022
        for it in self._iskry:
            it["y"] -= it["v"]
            if it["y"] < -0.05:
                it["y"] = 1.05
                it["x"] = random.uniform(0.05, 0.95)
        self.update()

    def paintEvent(self, e):
        p = QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        W, H = self.width(), self.height()
        akc = QColor("#00F0FF") if self.is_dark else QColor("#0D9488")
        zie = QColor("#00E4A1") if self.is_dark else QColor("#059669")

        # pulsujące kręgi wokół górnej części karty
        cx, cy = W / 2, H * 0.24
        for i in range(3):
            faza = self._faza + i * 0.75
            postep = (math.sin(faza) + 1) / 2          # 0..1
            promien = 42 + postep * 58 + i * 12
            alpha = int(46 * (1 - postep))
            kol = QColor(akc); kol.setAlpha(max(0, alpha))
            p.setPen(QPen(kol, 1.6))
            p.setBrush(Qt.BrushStyle.NoBrush)
            p.drawEllipse(QPointF(cx, cy), promien, promien)

        # unoszące się iskry
        for it in self._iskry:
            kol = QColor(zie if it["r"] > 2.0 else akc)
            migot = 0.6 + 0.4 * math.sin(self._faza * 2 + it["x"] * 8)
            kol.setAlpha(int(255 * it["a"] * migot * 0.5))
            p.setPen(Qt.PenStyle.NoPen); p.setBrush(kol)
            p.drawEllipse(QPointF(it["x"] * W, it["y"] * H), it["r"], it["r"])


class OknoAktualizacji(QDialog):
    """Powitanie nowej wersji — pokazywane PRZY STARCIE, gdy jest aktualizacja.
    Zamiast dyskretnej ikonki: pełne, efektowne okno, które od razu proponuje
    jedno kliknięcie („Zaktualizuj teraz”) i samo podmienia plik programu."""
    def __init__(self, parent=None, wersja_stara="", wersja_nowa="", opis="",
                 is_dark=True, on_instaluj=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setModal(True)
        self._on_instaluj = on_instaluj
        self._wersja_nowa = wersja_nowa
        self.is_dark = is_dark
        self._thread = None

        if is_dark:
            karta="#0A1220"; ramka="rgba(0,240,255,0.38)"; akc="#00F0FF"; zie="#00E4A1"
            txt="#F8FAFC"; mut="#94A3B8"; pole="rgba(5,10,20,0.55)"
        else:
            karta="#FFFFFF"; ramka="rgba(13,148,136,0.40)"; akc="#0D9488"; zie="#059669"
            txt="#0F172A"; mut="#64748B"; pole="rgba(241,245,249,0.9)"
        self._akc, self._zie, self._mut = akc, zie, mut

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        self.box = QFrame(); self.box.setObjectName("UpdBox")
        self.box.setStyleSheet(
            f"#UpdBox {{ background:{karta}; border:1.5px solid {ramka}; border-radius:20px; }}")
        self.box.setFixedSize(520, 500)
        root.addWidget(self.box)

        # animowane tło (pod treścią)
        self.aureola = AureolaAktualizacji(self.box, is_dark)
        self.aureola.setGeometry(0, 0, 520, 500)
        self.aureola.lower()

        bl = QVBoxLayout(self.box); bl.setContentsMargins(34, 30, 34, 28); bl.setSpacing(0)

        # --- plakietka „NOWA WERSJA” ---
        plak = QLabel("✦   NOWA WERSJA JEST GOTOWA   ✦")
        plak.setAlignment(Qt.AlignmentFlag.AlignCenter)
        plak.setStyleSheet(
            f"color:{akc}; font-family:'Segoe UI'; font-size:11px; font-weight:800; "
            f"letter-spacing:2px; background:transparent;")
        bl.addWidget(plak)
        bl.addSpacing(18)

        # --- kompas (logo) jako bohater ---
        logo_wrap = QHBoxLayout()
        logo_wrap.addStretch()
        self.logo = LogoKompas(self.box, srednica=88, is_dark=is_dark)
        logo_wrap.addWidget(self.logo)
        logo_wrap.addStretch()
        bl.addLayout(logo_wrap)
        bl.addSpacing(16)

        # --- przejście wersji: stara → nowa ---
        wers = QHBoxLayout(); wers.setSpacing(12)
        wers.addStretch()
        l_stara = QLabel(f"v{wersja_stara}")
        l_stara.setStyleSheet(
            f"color:{mut}; font-family:'Segoe UI'; font-size:15px; font-weight:600; "
            f"background:transparent; text-decoration:line-through;")
        wers.addWidget(l_stara)
        l_strz = QLabel("→")
        l_strz.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:16px; background:transparent;")
        wers.addWidget(l_strz)
        l_nowa = QLabel(f"v{wersja_nowa}")
        l_nowa.setStyleSheet(
            f"color:{akc}; font-family:'Segoe UI'; font-size:26px; font-weight:900; background:transparent;")
        wers.addWidget(l_nowa)
        wers.addStretch()
        bl.addLayout(wers)
        bl.addSpacing(16)

        # --- co nowego ---
        self.lbl_opis = QLabel(opis or "Dostępna jest nowsza wersja programu.")
        self.lbl_opis.setWordWrap(True)
        self.lbl_opis.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_opis.setStyleSheet(
            f"color:{txt}; font-family:'Segoe UI'; font-size:12px; background:{pole}; "
            f"border:1px solid {ramka}; border-radius:12px; padding:14px 16px;")
        bl.addWidget(self.lbl_opis)
        bl.addSpacing(6)

        # --- pasek postępu (ukryty do czasu pobierania) ---
        self.pasek = PasekPostepuDnia(0, 100, is_dark)
        self.pasek.setFixedHeight(8)
        self.pasek.hide()
        bl.addWidget(self.pasek)
        self.lbl_status = QLabel("")
        self.lbl_status.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_status.setStyleSheet(
            f"color:{mut}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
        self.lbl_status.hide()
        bl.addWidget(self.lbl_status)
        bl.addStretch()

        # --- przyciski ---
        akcje = QVBoxLayout(); akcje.setSpacing(8)
        self.btn_akt = QPushButton("⬇   Zaktualizuj teraz")
        self.btn_akt.setFixedHeight(48)
        self.btn_akt.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_akt.setStyleSheet(
            f"QPushButton {{ color:#04121A; background:qlineargradient(x1:0,y1:0,x2:1,y2:0, "
            f"stop:0 {akc}, stop:1 {zie}); border:none; border-radius:14px; "
            f"font-family:'Segoe UI'; font-size:15px; font-weight:900; letter-spacing:0.5px; }} "
            f"QPushButton:hover {{ background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {zie}, stop:1 {akc}); }} "
            f"QPushButton:disabled {{ color:{mut}; background:{pole}; }}")
        self.btn_akt.clicked.connect(self._instaluj)
        akcje.addWidget(self.btn_akt)

        self.btn_pozniej = QPushButton("Przypomnij przy następnym uruchomieniu")
        self.btn_pozniej.setFixedHeight(34)
        self.btn_pozniej.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_pozniej.setStyleSheet(
            f"QPushButton {{ color:{mut}; background:transparent; border:none; "
            f"font-family:'Segoe UI'; font-size:11px; }} "
            f"QPushButton:hover {{ color:{txt}; }}")
        self.btn_pozniej.clicked.connect(self.reject)
        akcje.addWidget(self.btn_pozniej)
        bl.addLayout(akcje)

    # ---- instalacja ----
    def _instaluj(self):
        if not czy_zamrozony():
            # wersja uruchomiona ze źródeł — nie ma czego podmieniać
            webbrowser.open(URL_POBIERANIA)
            self.accept()
            return
        self.btn_akt.setEnabled(False)
        self.btn_akt.setText("Pobieram…")
        self.btn_pozniej.setEnabled(False)
        self.pasek.show(); self.lbl_status.show()
        self._thread = PobieranieAktualizacjiThread()
        self._thread.postep.connect(self._postep)
        self._thread.sukces.connect(self._pobrano)
        self._thread.blad.connect(self._blad)
        self._thread.start()

    def _postep(self, frakcja, opis):
        self.pasek.zrobione = int(frakcja * 100)
        self.pasek.wszystkie = 100
        self.pasek.update()
        self.lbl_status.setText(opis)

    def _pobrano(self, sciezka):
        self.lbl_status.setText("Zamykam program i podmieniam plik…\nZa chwilę uruchomi się nowa wersja.")
        self.btn_akt.setText("Instaluję…")
        QApplication.processEvents()
        if self._on_instaluj:
            self._on_instaluj(sciezka)      # tu program się kończy
        self.accept()

    def _blad(self, komunikat):
        self.pasek.hide()
        self.lbl_status.setText(komunikat + "\nMożesz pobrać ręcznie ze strony wydania.")
        self.lbl_status.setStyleSheet(
            "color:#F87171; font-family:'Segoe UI'; font-size:11px; background:transparent;")
        self.btn_akt.setEnabled(True)
        self.btn_akt.setText("Otwórz stronę pobierania")
        try:
            self.btn_akt.clicked.disconnect()
        except Exception:
            pass
        self.btn_akt.clicked.connect(lambda: (webbrowser.open(URL_POBIERANIA), self.accept()))
        self.btn_pozniej.setEnabled(True)


class UpdateThread(QThread):
    """Sprawdza dostępność nowej wersji w tle, by nie blokować startu programu."""
    wynik = pyqtSignal(bool, str, str)   # (jest_nowsza, wersja, opis)
    def run(self):
        try:
            jest, wersja, opis = sprawdz_aktualizacje()
            self.wynik.emit(jest, wersja, opis)
        except Exception:
            self.wynik.emit(False, "", "")


class WykresSlupkowy(QWidget):
    """Autorski wykres słupkowy rysowany QPainterem — miesięczne koszty/km.
    Słupki z gradientem, animowane wysokości, wartości nad słupkami."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.is_dark = True
        self.dane = []          # [(etykieta, wartosc), ...]
        self.tytul = ""
        self.jednostka = ""
        self.kolor1 = "#00F0FF"; self.kolor2 = "#00E4A1"
        self._anim = 0.0
        self._postep = 0.0
        self.setMinimumHeight(200)
        self._timer = QTimer(self); self._timer.timeout.connect(self._krok_anim)

    def ustaw_dane(self, dane, tytul="", jednostka="", kolor1=None, kolor2=None):
        self.dane = dane; self.tytul = tytul; self.jednostka = jednostka
        if kolor1: self.kolor1 = kolor1
        if kolor2: self.kolor2 = kolor2
        self._anim = 0.0; self._postep = 0.0; self._timer.start(16); self.update()

    def _krok_anim(self):
        # Wolniejszy, płynny wzrost słupków. _postep rośnie liniowo (0→1),
        # a faktyczną wysokość liczymy z easingu (easeOutCubic) — słupki
        # startują żwawo i miękko wyhamowują na końcu.
        self._postep = min(1.0, self._postep + 0.018)
        t = self._postep
        self._anim = 1 - pow(1 - t, 3)      # easeOutCubic
        if self._postep >= 1.0: self._timer.stop()
        self.update()

    def paintEvent(self, e):
        p = QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        W, H = self.width(), self.height()
        txt = QColor("#F8FAFC") if self.is_dark else QColor("#0F172A")
        txt_mut = QColor("#94A3B8") if self.is_dark else QColor("#64748B")
        grid = QColor(255,255,255,18) if self.is_dark else QColor(15,23,42,12)

        # Rezerwujemy miejsce: pod tytułem (góra) osobny pas na WARTOŚCI nad
        # słupkami, żeby nigdy nie nachodziły na tytuł ani nie były ucinane.
        h_tytul = 30 if self.tytul else 0       # miejsce na tytuł
        h_wartosci = 24                          # pas na liczby nad słupkami
        pad_top = h_tytul + h_wartosci
        pad_bot = 30; pad_left = 12; pad_right = 12
        if self.tytul:
            p.setPen(txt); p.setFont(QFont("Segoe UI", 11, QFont.Weight.DemiBold))
            p.drawText(QRectF(0, 6, W, h_tytul - 6), Qt.AlignmentFlag.AlignHCenter, self.tytul)
        if not self.dane:
            p.setPen(txt_mut); p.setFont(QFont("Segoe UI", 10))
            p.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, "Brak danych do pokazania")
            return

        maxv = max((v for _, v in self.dane), default=1) or 1
        obszar_h = H - pad_top - pad_bot
        obszar_w = W - pad_left - pad_right
        n = len(self.dane)
        slot = obszar_w / n
        szer = min(slot * 0.6, 46)

        # linie siatki (3 poziomy)
        p.setPen(QPen(grid, 1))
        for i in range(1, 4):
            y = pad_top + obszar_h * i / 4
            p.drawLine(int(pad_left), int(y), int(W - pad_right), int(y))

        for i, (ety, val) in enumerate(self.dane):
            cx = pad_left + slot * i + slot / 2
            h = (val / maxv) * obszar_h * self._anim
            x = cx - szer / 2; y = pad_top + obszar_h - h
            grad = QLinearGradient(x, y, x, y + h)
            grad.setColorAt(0, QColor(self.kolor1)); grad.setColorAt(1, QColor(self.kolor2))
            p.setBrush(grad); p.setPen(Qt.PenStyle.NoPen)
            p.drawRoundedRect(QRectF(x, y, szer, h), 6, 6)
            # wartość nad słupkiem — w zarezerwowanym pasie. Prostokąt tekstu jest
            # SZERSZY niż słupek (żeby długie liczby jak "8506 zł" się mieściły) i
            # przycięty do obszaru wykresu, by skrajne słupki nie ucinały cyfr.
            if self._anim > 0.7 and val > 0:
                p.setPen(txt); p.setFont(QFont("Segoe UI", 8, QFont.Weight.DemiBold))
                szer_txt = max(slot, 56)         # min. 56 px na liczbę z jednostką
                lx = cx - szer_txt / 2
                # nie wychodź poza obszar wykresu (skrajne słupki)
                lx = max(2.0, min(lx, W - szer_txt - 2.0))
                y_txt = max(h_tytul, y - 22)     # tuż nad słupkiem, poniżej tytułu
                p.drawText(QRectF(lx, y_txt, szer_txt, 20),
                           Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignBottom,
                           f"{int(val)}{self.jednostka}")
            # etykieta pod spodem
            p.setPen(txt_mut); p.setFont(QFont("Segoe UI", 8))
            p.drawText(QRectF(cx - slot/2, H - pad_bot + 4, slot, 20), Qt.AlignmentFlag.AlignCenter, str(ety))


class WykresDonut(QWidget):
    """Autorski wykres pierścieniowy (donut) — udział województw.
    Segmenty z różnymi kolorami, legenda, wartość w środku."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.is_dark = True
        self.dane = []          # [(nazwa, wartosc), ...]
        self.tytul = ""
        self._anim = 0.0
        self.setMinimumHeight(200)
        self._timer = QTimer(self); self._timer.timeout.connect(self._krok)
        self._paleta = ["#00F0FF","#00E4A1","#F59E0B","#EC4899","#8B5CF6","#3B82F6","#EF4444","#10B981"]

    def ustaw_dane(self, dane, tytul=""):
        self.dane = sorted(dane, key=lambda x: -x[1])[:8]
        self.tytul = tytul
        self._anim = 0.0; self._postep = 0.0; self._timer.start(16); self.update()

    def _krok(self):
        # Wolniejszy, płynny wzrost (spójnie ze słupkami) — miękkie wyhamowanie.
        self._postep = min(1.0, getattr(self, "_postep", 0.0) + 0.02)
        self._anim = 1 - pow(1 - self._postep, 3)     # easeOutCubic
        if self._postep >= 1.0: self._timer.stop()
        self.update()

    def paintEvent(self, e):
        p = QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        W, H = self.width(), self.height()
        txt = QColor("#F8FAFC") if self.is_dark else QColor("#0F172A")
        txt_mut = QColor("#94A3B8") if self.is_dark else QColor("#64748B")

        if self.tytul:
            p.setPen(txt); p.setFont(QFont("Segoe UI", 11, QFont.Weight.DemiBold))
            p.drawText(QRectF(0, 6, W, 22), Qt.AlignmentFlag.AlignHCenter, self.tytul)
        if not self.dane:
            p.setPen(txt_mut); p.setFont(QFont("Segoe UI", 10))
            p.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, "Brak danych do pokazania")
            return

        total = sum(v for _, v in self.dane) or 1
        # pierścień po lewej, legenda po prawej
        R = min(H - 60, W * 0.42) / 2
        cx = 20 + R; cy = 34 + R + 6
        gruby = R * 0.42
        start = 90.0
        for i, (nazwa, val) in enumerate(self.dane):
            kat = 360.0 * (val / total) * self._anim
            kolor = QColor(self._paleta[i % len(self._paleta)])
            p.setPen(QPen(kolor, gruby, cap=Qt.PenCapStyle.FlatCap))
            rect = QRectF(cx - R + gruby/2, cy - R + gruby/2, 2*R - gruby, 2*R - gruby)
            p.drawArc(rect, int(-start * 16), int(-kat * 16))
            start += 360.0 * (val / total) * self._anim
        # środek — liczba wypraw
        p.setPen(txt); p.setFont(QFont("Segoe UI", 18, QFont.Weight.Bold))
        p.drawText(QRectF(cx - R, cy - 16, 2*R, 22), Qt.AlignmentFlag.AlignCenter, str(int(total)))
        p.setPen(txt_mut); p.setFont(QFont("Segoe UI", 8))
        p.drawText(QRectF(cx - R, cy + 6, 2*R, 16), Qt.AlignmentFlag.AlignCenter, "wypraw")

        # legenda
        lx = cx + R + 24; ly = 44
        p.setFont(QFont("Segoe UI", 9))
        for i, (nazwa, val) in enumerate(self.dane):
            kolor = QColor(self._paleta[i % len(self._paleta)])
            p.setBrush(kolor); p.setPen(Qt.PenStyle.NoPen)
            p.drawRoundedRect(QRectF(lx, ly + i*24, 12, 12), 3, 3)
            p.setPen(txt); 
            proc = 100.0 * val / total
            p.drawText(QRectF(lx + 20, ly + i*24 - 2, W - lx - 24, 16),
                       Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter,
                       f"{nazwa}  ·  {int(val)} ({proc:.0f}%)")


class SzlakiOverlay(QFrame):
    """Moje Szlaki — nakładka łącząca historię wypraw z modułem analitycznym.
    Górą karty-statystyki (rok), pod nimi autorskie wykresy: koszty miesięczne,
    kilometry, udział województw. Dane prywatne per pracownik."""
    def __init__(self, parent=None, on_dane_uzytkownika=None):
        super().__init__(parent); self.hide()
        self.is_dark = True
        self._on_dane = on_dane_uzytkownika

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        self.karta = QFrame(self); self.karta.setObjectName("SzlakiKarta")
        self.karta.setFixedSize(920, 640)
        root.addWidget(self.karta, alignment=Qt.AlignmentFlag.AlignCenter)
        kl = QVBoxLayout(self.karta); kl.setContentsMargins(30, 26, 30, 26); kl.setSpacing(16)

        # nagłówek
        gora = QHBoxLayout()
        tyt = QVBoxLayout(); tyt.setSpacing(2)
        self.tytul = QLabel("Moje Szlaki")
        self.podtytul = QLabel("Twoja aktywność w liczbach — podsumowanie roku")
        tyt.addWidget(self.tytul); tyt.addWidget(self.podtytul)
        gora.addLayout(tyt); gora.addStretch()
        # przełącznik roku
        self.btn_rok_prev = QPushButton("‹"); self.btn_rok_prev.setFixedSize(34, 34)
        self.lbl_rok = QLabel(""); self.lbl_rok.setFixedWidth(80); self.lbl_rok.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.btn_rok_next = QPushButton("›"); self.btn_rok_next.setFixedSize(34, 34)
        for b in (self.btn_rok_prev, self.btn_rok_next): b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_rok_prev.clicked.connect(lambda: self._zmien_rok(-1))
        self.btn_rok_next.clicked.connect(lambda: self._zmien_rok(1))
        gora.addWidget(self.btn_rok_prev); gora.addWidget(self.lbl_rok); gora.addWidget(self.btn_rok_next)
        gora.addSpacing(8)
        self.btn_x = QPushButton("✕"); self.btn_x.setFixedSize(36, 36)
        self.btn_x.setCursor(QCursor(Qt.CursorShape.PointingHandCursor)); self.btn_x.clicked.connect(self.hide)
        gora.addWidget(self.btn_x)
        kl.addLayout(gora)

        # karty statystyk (4 kafelki)
        self.staty = QHBoxLayout(); self.staty.setSpacing(12)
        self.karty_stat = []
        for _ in range(4):
            k = QFrame(); k.setObjectName("StatKarta"); k.setFixedHeight(76)
            kv = QVBoxLayout(k); kv.setContentsMargins(14, 10, 14, 10); kv.setSpacing(2)
            wart = QLabel("—"); wart.setObjectName("StatWart")
            opis = QLabel(""); opis.setObjectName("StatOpis")
            kv.addWidget(wart); kv.addWidget(opis)
            self.staty.addWidget(k); self.karty_stat.append((wart, opis))
        kl.addLayout(self.staty)

        # wykresy: słupki kosztów + km (obok siebie), donut województw pod nimi
        wykresy_gora = QHBoxLayout(); wykresy_gora.setSpacing(16)
        self.wyk_koszty = WykresSlupkowy(); self.wyk_koszty.setObjectName("WykBox")
        self.wyk_km = WykresSlupkowy(); self.wyk_km.setObjectName("WykBox")
        wykresy_gora.addWidget(self.wyk_koszty); wykresy_gora.addWidget(self.wyk_km)
        kl.addLayout(wykresy_gora, 1)

        # dwa donuty obok siebie: udział województw + najczęstsze bazy (top starty)
        donuty = QHBoxLayout(); donuty.setSpacing(16)
        self.wyk_woj = WykresDonut(); self.wyk_woj.setObjectName("WykBox")
        self.wyk_bazy = WykresDonut(); self.wyk_bazy.setObjectName("WykBox")
        donuty.addWidget(self.wyk_woj); donuty.addWidget(self.wyk_bazy)
        kl.addLayout(donuty, 1)

        self._rok = datetime.datetime.now().year

    def odswiez_dane(self):
        imie = pesel = ""
        if self._on_dane: imie, pesel = self._on_dane()
        self._historia = wczytaj_historie(imie, pesel)
        self._przelicz()

    def _zmien_rok(self, d):
        self._rok += d; self._przelicz()

    def _przelicz(self):
        self.lbl_rok.setText(str(self._rok))
        hist = getattr(self, "_historia", [])
        # filtruj po roku (z pola 'rok' albo z daty)
        wpisy = []
        for h in hist:
            r = h.get("rok")
            if r is None:
                try: r = int(h.get("data","").split(".")[2].split(" ")[0])
                except Exception: r = None
            if r == self._rok: wpisy.append(h)

        # miesięczne sumy
        koszty_mies = [0.0]*12
        km_mies = [0.0]*12
        woj_licznik = {}
        bazy_licznik = {}
        suma_km = 0.0; suma_kwota = 0.0
        for h in wpisy:
            m = h.get("miesiac")
            if m is None:
                try: m = int(h.get("data","").split(".")[1])
                except Exception: m = None
            kwota = _parsuj_kwote(h.get("kwota", 0))
            km = float(h.get("km", 0) or 0)
            suma_km += km; suma_kwota += kwota
            if m and 1 <= m <= 12:
                koszty_mies[m-1] += kwota; km_mies[m-1] += km
            w_wizyty = h.get("woj_wizyty")
            if isinstance(w_wizyty, dict) and w_wizyty:
                # nowy format: realny udział wszystkich odwiedzonych regionów
                for wnaz, ile in w_wizyty.items():
                    try: ile = int(ile)
                    except Exception: ile = 1
                    woj_licznik[wnaz] = woj_licznik.get(wnaz, 0) + ile
            else:
                # stary wpis (sprzed tej wersji): tylko województwo bazy
                w = h.get("woj", "—")
                woj_licznik[w] = woj_licznik.get(w, 0) + 1
            # odwiedzane miejscowości (bez bazy) — do wykresu top miejscowości
            m_wizyty = h.get("miejsc_wizyty")
            if isinstance(m_wizyty, dict):
                for mnaz, ile in m_wizyty.items():
                    try: ile = int(ile)
                    except Exception: ile = 1
                    bazy_licznik[mnaz] = bazy_licznik.get(mnaz, 0) + ile

        # karty statystyk
        self.karty_stat[0][0].setText(f"{len(wpisy)}"); self.karty_stat[0][1].setText("wygenerowanych wypraw")
        self.karty_stat[1][0].setText(f"{int(suma_km):,}".replace(","," ")+" km"); self.karty_stat[1][1].setText("łącznie w roku")
        self.karty_stat[2][0].setText(f"{suma_kwota:,.0f} zł".replace(","," ")); self.karty_stat[2][1].setText("suma rozliczeń")
        naj = max(woj_licznik.items(), key=lambda x: x[1])[0] if woj_licznik else "—"
        self.karty_stat[3][0].setText(naj); self.karty_stat[3][1].setText("najczęstszy region")

        skrot = ["Sty","Lut","Mar","Kwi","Maj","Cze","Lip","Sie","Wrz","Paź","Lis","Gru"]
        self.wyk_koszty.ustaw_dane(list(zip(skrot, koszty_mies)), "Koszty miesięczne", " zł",
                                   "#00F0FF" if self.is_dark else "#0D9488",
                                   "#00E4A1" if self.is_dark else "#059669")
        self.wyk_km.ustaw_dane(list(zip(skrot, km_mies)), "Kilometry miesięcznie", "",
                               "#F59E0B", "#EC4899")
        self.wyk_woj.ustaw_dane(list(woj_licznik.items()), "Udział województw")
        # Top 3 najczęściej odwiedzane miejscowości (bez bazy startowej)
        top_miejsc = sorted(bazy_licznik.items(), key=lambda x: -x[1])[:3]
        self.wyk_bazy.ustaw_dane(top_miejsc, "Najczęściej odwiedzane")

    def update_theme(self, is_dark):
        self.is_dark = is_dark
        for w in (self.wyk_koszty, self.wyk_km, self.wyk_woj, self.wyk_bazy):
            w.is_dark = is_dark; w.update()
        if is_dark:
            tlo="rgba(5,10,20,0.90)"; karta="rgba(11,19,32,0.98)"; ramka="rgba(0,240,255,0.25)"
            akc="#00F0FF"; txt="#F8FAFC"; txt_mut="#94A3B8"; box="rgba(5,10,20,0.4)"; sec="rgba(255,255,255,0.08)"
            statbg="rgba(0,240,255,0.06)"
        else:
            tlo="rgba(226,232,240,0.92)"; karta="#FFFFFF"; ramka="rgba(13,148,136,0.30)"
            akc="#0D9488"; txt="#0F172A"; txt_mut="#475569"; box="rgba(255,255,255,0.6)"; sec="rgba(15,23,42,0.08)"
            statbg="rgba(13,148,136,0.07)"
        self.setStyleSheet(f"SzlakiOverlay {{ background-color:{tlo}; }}")
        self.karta.setStyleSheet(f"#SzlakiKarta {{ background-color:{karta}; border:1px solid {ramka}; border-radius:18px; }}")
        self.tytul.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:22px; font-weight:800; background:transparent; border:none;")
        self.podtytul.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:12px; background:transparent; border:none;")
        self.lbl_rok.setStyleSheet(f"color:{txt}; font-family:'Segoe UI'; font-size:15px; font-weight:700; background:transparent; border:none;")
        nav=(f"QPushButton {{ color:{txt}; background:{sec}; border:1px solid {ramka}; border-radius:9px; font-size:16px; font-weight:bold; }} QPushButton:hover {{ background:{ramka}; }}")
        self.btn_rok_prev.setStyleSheet(nav); self.btn_rok_next.setStyleSheet(nav)
        self.btn_x.setStyleSheet(f"QPushButton {{ color:{txt_mut}; background:transparent; border:none; font-size:18px; }} QPushButton:hover {{ color:{txt}; }}")
        for wart, opis in self.karty_stat:
            wart.parent().setStyleSheet(f"#StatKarta {{ background:{statbg}; border:1px solid {ramka}; border-radius:12px; }}")
            wart.setStyleSheet(f"#StatWart {{ color:{akc}; font-family:'Segoe UI'; font-size:18px; font-weight:800; background:transparent; border:none; }}")
            opis.setStyleSheet(f"#StatOpis {{ color:{txt_mut}; font-family:'Segoe UI'; font-size:10px; background:transparent; border:none; }}")
        for w in (self.wyk_koszty, self.wyk_km, self.wyk_woj, self.wyk_bazy):
            w.setStyleSheet(f"#WykBox {{ background:{box}; border:1px solid {ramka}; border-radius:12px; }}")


class PierscienPostepu(QWidget):
    """Pierścień pokazujący postęp dnia — ile wizyt z dzisiejszych już zrobione.
    Wypełnia się płynnie po pojawieniu się karty (animacja 'nalewania')."""
    def __init__(self, parent=None, is_dark=True, rozmiar=96):
        super().__init__(parent)
        self.is_dark = is_dark
        self.zrobione = 0
        self.wszystkie = 0
        self._postep = 0.0        # 0..1 — animowane wypełnienie
        self._cel = 0.0
        self.setFixedSize(rozmiar, rozmiar)
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._krok)

    def ustaw(self, zrobione, wszystkie):
        self.zrobione = zrobione
        self.wszystkie = max(0, wszystkie)
        self._cel = (zrobione / wszystkie) if wszystkie else 0.0
        self._postep = 0.0
        self._timer.start(16)
        self.update()

    def _krok(self):
        if abs(self._postep - self._cel) < 0.005:
            self._postep = self._cel
            self._timer.stop()
        else:
            self._postep += (self._cel - self._postep) * 0.10
        self.update()

    def paintEvent(self, e):
        p = QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        akc = QColor("#00F0FF") if self.is_dark else QColor("#0D9488")
        zie = QColor("#00E4A1") if self.is_dark else QColor("#059669")
        tlo = QColor(148, 163, 184, 55)
        txt = QColor("#F8FAFC") if self.is_dark else QColor("#0F172A")
        mut = QColor("#94A3B8") if self.is_dark else QColor("#64748B")

        skala = self.width() / 96.0
        gruba = max(4.0, 8.0 * skala)
        r = QRectF(9 * skala, 9 * skala, self.width() - 18 * skala, self.height() - 18 * skala)
        # tor
        p.setPen(QPen(tlo, gruba, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap))
        p.drawArc(r, 0, 360 * 16)
        # wypełnienie — gradient wzdłuż łuku
        if self._postep > 0:
            grad = QConicalGradient(r.center(), 90)
            grad.setColorAt(0.0, akc)
            grad.setColorAt(0.5, zie)
            grad.setColorAt(1.0, akc)
            p.setPen(QPen(QBrush(grad), gruba, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap))
            p.drawArc(r, 90 * 16, -int(360 * 16 * self._postep))

        komplet = self.wszystkie > 0 and self.zrobione >= self.wszystkie
        if komplet:
            # ptaszek zamiast liczb
            p.setPen(QPen(zie, max(2.0, 4.0 * skala), Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap))
            c = r.center()
            p.drawLine(QPointF(c.x() - 11*skala, c.y() + 1*skala), QPointF(c.x() - 3*skala, c.y() + 9*skala))
            p.drawLine(QPointF(c.x() - 3*skala, c.y() + 9*skala), QPointF(c.x() + 12*skala, c.y() - 8*skala))
        else:
            p.setPen(txt)
            p.setFont(QFont("Segoe UI", max(10, round(19 * skala)), QFont.Weight.Black))
            p.drawText(QRectF(0, 26 * skala, self.width(), 30 * skala),
                       Qt.AlignmentFlag.AlignCenter, str(self.zrobione))
            p.setPen(mut)
            p.setFont(QFont("Segoe UI", max(7, round(9 * skala)), QFont.Weight.DemiBold))
            p.drawText(QRectF(0, 54 * skala, self.width(), 18 * skala),
                       Qt.AlignmentFlag.AlignCenter, f"z {self.wszystkie}")


class KartaDzisiaj(QFrame):
    """„Dziś w trasie" — kokpit dnia na ekranie powitalnym.
    Po uruchomieniu programu od razu widzisz: ile masz dziś wizyt, ile już
    zrobione i DOKĄD jedziesz następnie. Bez klikania po zakładkach."""
    _DNI = ["poniedziałek", "wtorek", "środa", "czwartek", "piątek", "sobota", "niedziela"]
    _MIES = ["", "stycznia", "lutego", "marca", "kwietnia", "maja", "czerwca",
             "lipca", "sierpnia", "września", "października", "listopada", "grudnia"]

    def __init__(self, parent=None, is_dark=True, on_trasa=None, on_planer=None):
        super().__init__(parent)
        self.is_dark = is_dark
        self._on_trasa = on_trasa      # otwórz Plan Wizyt na dziś
        self._on_planer = on_planer    # otwórz planer (gdy brak planu)
        self.setObjectName("KartaDzis")
        self.setFixedHeight(138)

        h = QHBoxLayout(self); h.setContentsMargins(22, 16, 22, 16); h.setSpacing(20)
        self.pierscien = PierscienPostepu(self, is_dark)
        h.addWidget(self.pierscien)

        srodek = QVBoxLayout(); srodek.setSpacing(3)
        self.lbl_naglowek = QLabel("")
        self.lbl_naglowek.setObjectName("KDNagl")
        srodek.addWidget(self.lbl_naglowek)
        self.lbl_meta = QLabel("")
        self.lbl_meta.setObjectName("KDMeta")
        srodek.addWidget(self.lbl_meta)
        srodek.addSpacing(4)
        self.lbl_nastepny = QLabel("")
        self.lbl_nastepny.setObjectName("KDNast")
        self.lbl_nastepny.setWordWrap(True)
        srodek.addWidget(self.lbl_nastepny)
        srodek.addStretch()
        h.addLayout(srodek, 1)

        self.btn_trasa = QPushButton("Wejdź w trasę  →")
        self.btn_trasa.setFixedHeight(44)
        self.btn_trasa.setMinimumWidth(190)
        self.btn_trasa.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_trasa.clicked.connect(self._klik)
        h.addWidget(self.btn_trasa)

        self._tryb = "brak"        # brak | trasa | wolne | komplet | zalegle
        self._dzien_dzis_obj = None
        self._on_tryb_trasy = None  # callback(dzien, data) — startuje Tryb Trasy
        self.update_theme(is_dark)

    def _klik(self):
        # "Dziś w trasie" z wizytami do zrobienia → prosto w Tryb Trasy
        # (immersyjny widok jazdy). Pozostałe stany → zwykły ekran planu.
        if self._tryb == "trasa" and self._on_tryb_trasy and self._dzien_dzis_obj:
            self._on_tryb_trasy(self._dzien_dzis_obj, datetime.date.today())
        elif self._tryb == "brak" and self._on_planer:
            self._on_planer()
        elif self._on_trasa:
            self._on_trasa()

    def odswiez(self):
        """Czyta zapisany plan i pokazuje sytuację na DZIŚ."""
        plan = wczytaj_plan()
        dzis = datetime.date.today()
        dzien_dzis = None
        nastepny_dzien = None
        if plan:
            for d in plan.get("dni", []):
                if d.data == dzis:
                    dzien_dzis = d
                elif d.data > dzis and nastepny_dzien is None:
                    nastepny_dzien = d

        etyk_data = f"{self._DNI[dzis.weekday()]}, {dzis.day} {self._MIES[dzis.month]}"

        if dzien_dzis:
            self._dzien_dzis_obj = dzien_dzis
            zrobione = sum(1 for w in dzien_dzis.wizyty
                           if czy_odwiedzona(dzis, w.adres or w.nazwa))
            ile = len(dzien_dzis.wizyty)
            self.pierscien.ustaw(zrobione, ile)
            h = int(dzien_dzis.minuty // 60); m = int(dzien_dzis.minuty % 60)
            if zrobione >= ile:
                self._tryb = "komplet"
                self.lbl_naglowek.setText("Dzień zamknięty  ✓")
                self.lbl_meta.setText(f"{etyk_data}  •  {ile} wizyt  •  {dzien_dzis.km:.0f} km")
                self.lbl_nastepny.setText("Wszystkie dzisiejsze wizyty odhaczone. Dobra robota.")
                self.btn_trasa.setText("Zobacz dzień  →")
            else:
                self._tryb = "trasa"
                self.lbl_naglowek.setText("Dziś w trasie")
                self.lbl_meta.setText(
                    f"{etyk_data}  •  {ile} wizyt  •  {dzien_dzis.km:.0f} km  •  ok. {h}h {m}min")
                # następny nieodhaczony punkt
                nast = None
                for w in dzien_dzis.wizyty:
                    if not czy_odwiedzona(dzis, w.adres or w.nazwa):
                        nast = w; break
                if nast:
                    adres = nast.adres or ""
                    self.lbl_nastepny.setText(f"▸  Następnie:  {nast.nazwa}"
                                              + (f"   ·   {adres}" if adres else ""))
                self.btn_trasa.setText(f"Wejdź w trasę  →")
        elif nastepny_dzien:
            self._tryb = "wolne"
            self._dzien_dzis_obj = None
            self.pierscien.ustaw(0, 0)
            dni_do = (nastepny_dzien.data - dzis).days
            kiedy = "jutro" if dni_do == 1 else f"za {dni_do} dni"
            self.lbl_naglowek.setText("Dziś bez wizyt")
            self.lbl_meta.setText(etyk_data)
            self.lbl_nastepny.setText(
                f"▸  Najbliższa trasa {kiedy} — {nastepny_dzien.data.strftime('%d.%m')}"
                f"   ·   {len(nastepny_dzien.wizyty)} wizyt, {nastepny_dzien.km:.0f} km")
            self.btn_trasa.setText("Zobacz plan  →")
        else:
            # Plan może istnieć, ale mieć wszystkie dni już za nami — to nie to
            # samo co brak planu. Wtedy pokazujemy zaległości.
            zal = zalegle_wizyty(plan) if plan else []
            if plan and plan.get("dni"):
                self._tryb = "zalegle" if zal else "komplet"
                self._dzien_dzis_obj = None
                self.pierscien.ustaw(0, 0)
                self.lbl_meta.setText(etyk_data)
                if zal:
                    self.lbl_naglowek.setText("Plan wymaga odświeżenia")
                    self.lbl_nastepny.setText(
                        f"▸  {len(zal)} wizyt z minionych dni bez odhaczenia — "
                        f"mogę przenieść je na kolejne dni robocze.")
                    self.btn_trasa.setText("Otwórz plan  →")
                else:
                    self.lbl_naglowek.setText("Plan zakończony  ✓")
                    self.lbl_nastepny.setText("▸  Wszystkie zaplanowane wizyty odhaczone. Czas ułożyć nowy plan.")
                    self.btn_trasa.setText("Zobacz plan  →")
            else:
                self._tryb = "brak"
                self._dzien_dzis_obj = None
                self.pierscien.ustaw(0, 0)
                self.lbl_naglowek.setText("Brak planu wizyt")
                self.lbl_meta.setText(etyk_data)
                self.lbl_nastepny.setText("▸  Wczytaj listę sklepów i ułóż trasy — zajmie chwilę.")
                self.btn_trasa.setText("Ułóż plan  →")
        self.update_theme(self.is_dark)

    def update_theme(self, is_dark):
        self.is_dark = is_dark
        self.pierscien.is_dark = is_dark
        self.pierscien.update()
        if is_dark:
            karta = "rgba(11,19,32,0.82)"; ramka = "rgba(0,240,255,0.30)"
            akc = "#00F0FF"; zie = "#00E4A1"; txt = "#F8FAFC"; mut = "#94A3B8"
        else:
            karta = "rgba(255,255,255,0.92)"; ramka = "rgba(13,148,136,0.32)"
            akc = "#0D9488"; zie = "#059669"; txt = "#0F172A"; mut = "#64748B"
        kol_nagl = zie if self._tryb == "komplet" else akc
        self.setStyleSheet(
            f"#KartaDzis {{ background:{karta}; border:1px solid {ramka}; border-radius:18px; }}")
        self.lbl_naglowek.setStyleSheet(
            f"#KDNagl {{ color:{kol_nagl}; font-family:'Segoe UI'; font-size:20px; font-weight:900; background:transparent; border:none; }}")
        self.lbl_meta.setStyleSheet(
            f"#KDMeta {{ color:{mut}; font-family:'Segoe UI'; font-size:12px; background:transparent; border:none; }}")
        self.lbl_nastepny.setStyleSheet(
            f"#KDNast {{ color:{txt}; font-family:'Segoe UI'; font-size:13px; font-weight:600; background:transparent; border:none; }}")
        if self._tryb == "brak":
            self.btn_trasa.setStyleSheet(
                f"QPushButton {{ color:{akc}; background:transparent; border:1.5px solid {akc}; "
                f"border-radius:12px; font-family:'Segoe UI'; font-size:13px; font-weight:800; padding:0 18px; }} "
                f"QPushButton:hover {{ background:rgba(0,240,255,0.12); }}")
        else:
            self.btn_trasa.setStyleSheet(
                f"QPushButton {{ color:#04121A; background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc}, stop:1 {zie}); "
                f"border:none; border-radius:12px; font-family:'Segoe UI'; font-size:13px; font-weight:800; padding:0 18px; }} "
                f"QPushButton:hover {{ background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {zie}, stop:1 {akc}); }}")


class TrybTrasyOverlay(QFrame):
    """TRYB TRASY — pełnoekranowy widok „na czas jazdy”.

    Filozofia: to nie jest lista do czytania, tylko towarzysz jazdy. Jeden
    punkt na raz, ogromna czcionka czytelna kątem oka, jeden wielki przycisk.
    Po odhaczeniu program SAM przeskakuje do następnego nieodwiedzonego —
    zero przewijania, zero szukania na liście.
    """
    def __init__(self, parent=None, on_zamknij=None):
        super().__init__(parent); self.hide()
        self.is_dark = True
        self._on_zamknij = on_zamknij
        self._dzien = None
        self._data = None
        self._kolejnosc = []
        self._idx = 0

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        self.karta = QFrame(self); self.karta.setObjectName("TrybKarta")
        self.karta.setMinimumSize(640, 480)
        root.addWidget(self.karta, alignment=Qt.AlignmentFlag.AlignCenter)
        kl = QVBoxLayout(self.karta); kl.setContentsMargins(40, 30, 40, 34); kl.setSpacing(0)

        # --- nagłówek: tytuł + zamknij ---
        gora = QHBoxLayout()
        tyt = QVBoxLayout(); tyt.setSpacing(2)
        self.tytul = QLabel("🚗  Tryb Trasy")
        self.podtytul = QLabel("")
        tyt.addWidget(self.tytul); tyt.addWidget(self.podtytul)
        gora.addLayout(tyt); gora.addStretch()
        self.btn_x = QPushButton("✕  Zamknij"); self.btn_x.setFixedHeight(36)
        self.btn_x.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_x.clicked.connect(self._zamknij)
        gora.addWidget(self.btn_x)
        kl.addLayout(gora)
        kl.addSpacing(18)

        # --- środek: pierścień + aktualny punkt (widoczne, gdy jest co jechać) ---
        self.widok_jazdy = QWidget()
        wj = QVBoxLayout(self.widok_jazdy); wj.setContentsMargins(0, 0, 0, 0); wj.setSpacing(0)

        srodek = QHBoxLayout(); srodek.setSpacing(28)
        self.pierscien = PierscienPostepu(self, self.is_dark, rozmiar=118)
        srodek.addWidget(self.pierscien)
        info = QVBoxLayout(); info.setSpacing(4)
        self.lbl_siec = QLabel(""); self.lbl_siec.setObjectName("TrybSiec")
        info.addWidget(self.lbl_siec)
        self.lbl_nazwa = QLabel(""); self.lbl_nazwa.setObjectName("TrybNazwa")
        self.lbl_nazwa.setWordWrap(True)
        info.addWidget(self.lbl_nazwa)
        self.lbl_adres = QLabel(""); self.lbl_adres.setObjectName("TrybAdres")
        self.lbl_adres.setWordWrap(True)
        info.addWidget(self.lbl_adres)
        srodek.addLayout(info, 1)
        wj.addLayout(srodek)
        wj.addSpacing(28)

        self.btn_dojechalem = QPushButton("✓   DOJECHAŁEM — ODHACZ WIZYTĘ")
        self.btn_dojechalem.setFixedHeight(76)
        self.btn_dojechalem.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_dojechalem.clicked.connect(self._dojechalem)
        wj.addWidget(self.btn_dojechalem)
        wj.addSpacing(18)

        self.lbl_nastepny = QLabel(""); self.lbl_nastepny.setObjectName("TrybNastepny")
        self.lbl_nastepny.setWordWrap(True)
        wj.addWidget(self.lbl_nastepny)
        wj.addStretch()

        dolny = QHBoxLayout(); dolny.setSpacing(10)
        self.btn_poprzedni = QPushButton("←  Poprzedni"); self.btn_poprzedni.setFixedHeight(40)
        self.btn_notatka = QPushButton("📝  Notatka"); self.btn_notatka.setFixedHeight(40)
        self.btn_nastepny_reczny = QPushButton("Następny  →"); self.btn_nastepny_reczny.setFixedHeight(40)
        for b in (self.btn_poprzedni, self.btn_notatka, self.btn_nastepny_reczny):
            b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_poprzedni.clicked.connect(lambda: self._przesun_recznie(-1))
        self.btn_nastepny_reczny.clicked.connect(lambda: self._przesun_recznie(1))
        self.btn_notatka.clicked.connect(self._notatka)
        dolny.addWidget(self.btn_poprzedni); dolny.addStretch()
        dolny.addWidget(self.btn_notatka); dolny.addStretch()
        dolny.addWidget(self.btn_nastepny_reczny)
        wj.addLayout(dolny)
        kl.addWidget(self.widok_jazdy)

        # --- widok ukończenia dnia (celebracja) ---
        self.widok_koniec = QWidget()
        wk = QVBoxLayout(self.widok_koniec); wk.setContentsMargins(0, 30, 0, 0); wk.setSpacing(14)
        wk.addStretch()
        self.lbl_konfetti = QLabel("🎉")
        self.lbl_konfetti.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_konfetti.setStyleSheet("font-size:56px; background:transparent;")
        wk.addWidget(self.lbl_konfetti)
        self.lbl_koniec_tyt = QLabel("Dzień ukończony!")
        self.lbl_koniec_tyt.setObjectName("TrybKoniecTyt")
        self.lbl_koniec_tyt.setAlignment(Qt.AlignmentFlag.AlignCenter)
        wk.addWidget(self.lbl_koniec_tyt)
        self.lbl_koniec_meta = QLabel("")
        self.lbl_koniec_meta.setObjectName("TrybKoniecMeta")
        self.lbl_koniec_meta.setAlignment(Qt.AlignmentFlag.AlignCenter)
        wk.addWidget(self.lbl_koniec_meta)
        wk.addSpacing(10)
        self.btn_zakoncz = QPushButton("Zakończ  ✓"); self.btn_zakoncz.setFixedHeight(52)
        self.btn_zakoncz.setMinimumWidth(220)
        self.btn_zakoncz.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_zakoncz.clicked.connect(self._zamknij)
        wk_btn_wrap = QHBoxLayout(); wk_btn_wrap.addStretch(); wk_btn_wrap.addWidget(self.btn_zakoncz); wk_btn_wrap.addStretch()
        wk.addLayout(wk_btn_wrap)
        wk.addStretch()
        self.widok_koniec.setVisible(False)
        kl.addWidget(self.widok_koniec)

        self.update_theme(self.is_dark)

    # ---------- API ----------
    def ustaw_dzien(self, dzien, data):
        """Wchodzi w tryb trasy dla danego dnia — startuje od pierwszego
        NIEODWIEDZONEGO punktu (jeśli część już odhaczona wcześniej ręcznie)."""
        self._dzien = dzien
        self._data = data
        self._kolejnosc = list(dzien.wizyty)
        self._idx = 0
        for i, wz in enumerate(self._kolejnosc):
            if not czy_odwiedzona(data, wz.adres or wz.nazwa):
                self._idx = i
                break
        else:
            self._idx = len(self._kolejnosc)   # wszystko już zrobione
        self.podtytul.setText(self._etykieta_daty(data))
        self._odswiez()

    def _etykieta_daty(self, data):
        dni = ["poniedziałek","wtorek","środa","czwartek","piątek","sobota","niedziela"]
        mies = ["","stycznia","lutego","marca","kwietnia","maja","czerwca",
                "lipca","sierpnia","września","października","listopada","grudnia"]
        return f"{dni[data.weekday()]}, {data.day} {mies[data.month]} {data.year}"

    def _odswiez(self):
        n = len(self._kolejnosc)
        zrobione = sum(1 for wz in self._kolejnosc if czy_odwiedzona(self._data, wz.adres or wz.nazwa))
        self.pierscien.ustaw(zrobione, n)

        if self._idx >= n:
            self.widok_jazdy.setVisible(False)
            self.widok_koniec.setVisible(True)
            self.lbl_koniec_meta.setText(f"{n} {'wizyta' if n==1 else 'wizyt'} • {self._dzien.km:.0f} km • {self._etykieta_daty(self._data)}")
            return

        self.widok_jazdy.setVisible(True)
        self.widok_koniec.setVisible(False)
        wz = self._kolejnosc[self._idx]
        self.lbl_siec.setText(wz.siec or "—")
        self.lbl_nazwa.setText(wz.nazwa or wz.miasto or "")
        adres_txt = wz.adres or ""
        if wz.miasto and wz.miasto not in adres_txt:
            adres_txt = f"{adres_txt}, {wz.miasto}" if adres_txt else wz.miasto
        self.lbl_adres.setText(adres_txt)

        if self._idx + 1 < n:
            nast = self._kolejnosc[self._idx + 1]
            self.lbl_nastepny.setText(f"▸  Następnie:  {nast.nazwa or nast.miasto}"
                                      + (f"   ·   {nast.adres}" if nast.adres else ""))
        else:
            self.lbl_nastepny.setText("▸  To ostatni punkt na dziś.")

        juz_odwiedzona = czy_odwiedzona(self._data, wz.adres or wz.nazwa)
        self.btn_dojechalem.setText("✓   ODWIEDZONE — PRZEJDŹ DALEJ" if juz_odwiedzona
                                     else "✓   DOJECHAŁEM — ODHACZ WIZYTĘ")
        self.btn_poprzedni.setEnabled(self._idx > 0)
        self.btn_nastepny_reczny.setEnabled(self._idx + 1 <= n)

    def _dojechalem(self):
        if self._idx >= len(self._kolejnosc):
            return
        wz = self._kolejnosc[self._idx]
        if not czy_odwiedzona(self._data, wz.adres or wz.nazwa):
            ustaw_odwiedzona(self._data, wz.adres or wz.nazwa, True)
        # przeskocz do KOLEJNEGO nieodwiedzonego (pomija już odhaczone ręcznie)
        nowy = self._idx + 1
        while nowy < len(self._kolejnosc) and czy_odwiedzona(self._data, self._kolejnosc[nowy].adres or self._kolejnosc[nowy].nazwa):
            nowy += 1
        self._idx = nowy
        self._odswiez()

    def _przesun_recznie(self, kier):
        """Ręczne przeglądanie (bez odhaczania) — np. żeby zerknąć wstecz."""
        nowy = self._idx + kier
        if 0 <= nowy <= len(self._kolejnosc):
            self._idx = nowy
            self._odswiez()

    def _notatka(self):
        if self._idx >= len(self._kolejnosc):
            return
        wz = self._kolejnosc[self._idx]
        adres = wz.adres or wz.nazwa
        obecna = wpis_wizyty(self._data, adres).get("notatka", "")
        dlg = DialogNotatki(self, obecna, self.is_dark)
        wynik = dlg.pokaz()
        if wynik is None:
            return
        if wynik == "__USUN__":
            ustaw_odwiedzona(self._data, adres, False)
        else:
            if not czy_odwiedzona(self._data, adres):
                ustaw_odwiedzona(self._data, adres, True)
            ustaw_notatke(self._data, adres, wynik)
        self._odswiez()

    def _zamknij(self):
        if self._on_zamknij:
            self._on_zamknij()
        else:
            self.hide()

    def resizeEvent(self, e):
        super().resizeEvent(e)
        M = 40
        w = max(640, min(820, self.width() - 2 * M))
        h = max(480, min(700, self.height() - 2 * M))
        self.karta.setFixedSize(w, h)

    def update_theme(self, is_dark):
        self.is_dark = is_dark
        self.pierscien.is_dark = is_dark
        self.pierscien.update()
        if is_dark:
            tlo = "rgba(3,7,14,0.94)"; karta = "rgba(11,19,32,0.98)"; ramka = "rgba(0,240,255,0.25)"
            akc = "#00F0FF"; zie = "#00E4A1"; txt = "#F8FAFC"; mut = "#94A3B8"
            pole_sec = "rgba(255,255,255,0.06)"; ramka_sec = "rgba(255,255,255,0.15)"
        else:
            tlo = "rgba(226,232,240,0.96)"; karta = "#FFFFFF"; ramka = "rgba(13,148,136,0.30)"
            akc = "#0D9488"; zie = "#059669"; txt = "#0F172A"; mut = "#475569"
            pole_sec = "rgba(15,23,42,0.05)"; ramka_sec = "rgba(15,23,42,0.15)"
        self.setStyleSheet(f"TrybTrasyOverlay {{ background-color:{tlo}; }}")
        self.karta.setStyleSheet(f"#TrybKarta {{ background-color:{karta}; border:1px solid {ramka}; border-radius:20px; }}")
        self.tytul.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:18px; font-weight:900; background:transparent; border:none;")
        self.podtytul.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:12px; background:transparent; border:none;")
        self.btn_x.setStyleSheet(
            f"QPushButton {{ color:{txt}; background:transparent; border:1px solid {ramka_sec}; border-radius:9px; "
            f"padding:0 14px; font-family:'Segoe UI'; font-size:12px; }} QPushButton:hover {{ border-color:{akc}; color:{akc}; }}")

        self.lbl_siec.setStyleSheet(
            f"color:{zie}; font-family:'Segoe UI'; font-size:13px; font-weight:800; letter-spacing:1px; background:transparent;")
        self.lbl_nazwa.setStyleSheet(
            f"color:{txt}; font-family:'Segoe UI'; font-size:30px; font-weight:900; background:transparent;")
        self.lbl_adres.setStyleSheet(
            f"color:{mut}; font-family:'Segoe UI'; font-size:15px; background:transparent;")
        self.lbl_nastepny.setStyleSheet(
            f"color:{txt}; font-family:'Segoe UI'; font-size:13px; font-weight:600; background:transparent;")

        self.btn_dojechalem.setStyleSheet(
            f"QPushButton {{ color:#04121A; background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc}, stop:1 {zie}); "
            f"border:none; border-radius:16px; font-family:'Segoe UI'; font-size:18px; font-weight:900; }} "
            f"QPushButton:hover {{ background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {zie}, stop:1 {akc}); }}")

        sec_css = (
            f"QPushButton {{ color:{txt}; background:{pole_sec}; border:1px solid {ramka_sec}; border-radius:10px; "
            f"padding:0 14px; font-family:'Segoe UI'; font-size:12px; font-weight:600; }} "
            f"QPushButton:hover {{ border-color:{akc}; }} QPushButton:disabled {{ color:{mut}; }}")
        for b in (self.btn_poprzedni, self.btn_notatka, self.btn_nastepny_reczny):
            b.setStyleSheet(sec_css)

        self.lbl_koniec_tyt.setStyleSheet(
            f"color:{zie}; font-family:'Segoe UI'; font-size:26px; font-weight:900; background:transparent;")
        self.lbl_koniec_meta.setStyleSheet(
            f"color:{mut}; font-family:'Segoe UI'; font-size:13px; background:transparent;")
        self.btn_zakoncz.setStyleSheet(
            f"QPushButton {{ color:#04121A; background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc}, stop:1 {zie}); "
            f"border:none; border-radius:14px; font-family:'Segoe UI'; font-size:15px; font-weight:800; padding:0 26px; }}")


class EkranPowitalny(QFrame):
    """Ekran powitalny 'Kokpit nawigatora' — pokazuje się po uruchomieniu
    programu i zasłania generator. Autorska animacja QPainter: obracający się
    pierścień kompasu, świecące trasy rysujące się ku centrum, logo z poświatą.
    Znika po kliknięciu 'Bilans Miesiąca' w menu (odsłania generator)."""
    def __init__(self, parent=None, on_dane_uzytkownika=None):
        super().__init__(parent)
        self.is_dark = True
        # Przezroczyste tło — prześwituje tło aplikacji (góry, świecąca droga),
        # dzięki czemu ekran powitalny współgra z resztą programu.
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground, True)
        self.setStyleSheet("background: transparent;")
        self._on_dane = on_dane_uzytkownika
        self._kat = 0.0                # obrót kompasu
        self._t = 0.0                  # czas dla animacji tras
        self._logo_pix = None
        sciezka = znajdz_logo()
        if sciezka:
            pix = QPixmap(sciezka)
            if not pix.isNull():
                self._logo_pix = pix
        # świecące trasy: kilka linii od krawędzi ku centrum (kąt, długość, faza)
        import random as _r
        rng = _r.Random(7)
        self._trasy = []
        for _ in range(9):
            self._trasy.append({
                "kat": rng.uniform(0, 2 * math.pi),
                "faza": rng.uniform(0, 1),
                "predkosc": rng.uniform(0.004, 0.010),
                "dlugosc": rng.uniform(0.55, 0.92),
            })
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._krok)
        self._powitanie = "Witaj w PMT Planer"

        # KOKPIT DNIA — od razu widać, co dziś w trasie
        self.karta_dzis = KartaDzisiaj(self, self.is_dark)
        self.karta_dzis.hide()
        self._karta_anim = None

    def podepnij_akcje(self, on_trasa, on_planer, on_tryb_trasy=None):
        self.karta_dzis._on_trasa = on_trasa
        self.karta_dzis._on_planer = on_planer
        self.karta_dzis._on_tryb_trasy = on_tryb_trasy

    def _ustaw_karte(self):
        """Karta na dole, wyśrodkowana — szeroka, ale nie na całą szerokość."""
        szer = max(560, min(880, int(self.width() * 0.62)))
        x = (self.width() - szer) // 2
        y = self.height() - self.karta_dzis.height() - 52
        self.karta_dzis.setGeometry(x, max(10, y), szer, self.karta_dzis.height())

    def resizeEvent(self, e):
        super().resizeEvent(e)
        self._ustaw_karte()

    def start(self):
        # personalizacja powitania — jeśli jest zapisany profil, przywitaj imieniem
        try:
            imie = pesel = ""
            if self._on_dane:
                imie, pesel = self._on_dane()
            if imie:
                pierwsze = imie.strip().split()[0]
                self._powitanie = f"Witaj ponownie, {pierwsze}"
            else:
                self._powitanie = "Witaj w PMT Planer"
        except Exception:
            self._powitanie = "Witaj w PMT Planer"
        self._timer.start(16)
        self.show()
        self.raise_()
        # kokpit dnia: odśwież dane i wjedź od dołu
        try:
            self.karta_dzis.odswiez()
            self._ustaw_karte()
            self.karta_dzis.show()
            self.karta_dzis.raise_()
            koniec = self.karta_dzis.geometry()
            start = QRect(koniec.x(), koniec.y() + 60, koniec.width(), koniec.height())
            self.karta_dzis.setGeometry(start)
            efekt = QGraphicsOpacityEffect(self.karta_dzis)
            self.karta_dzis.setGraphicsEffect(efekt)
            efekt.setOpacity(0.0)
            self._anim_poz = QPropertyAnimation(self.karta_dzis, b"geometry", self)
            self._anim_poz.setDuration(620)
            self._anim_poz.setStartValue(start)
            self._anim_poz.setEndValue(koniec)
            self._anim_poz.setEasingCurve(QEasingCurve.Type.OutCubic)
            self._anim_op = QPropertyAnimation(efekt, b"opacity", self)
            self._anim_op.setDuration(620)
            self._anim_op.setStartValue(0.0)
            self._anim_op.setEndValue(1.0)
            self._karta_anim = QParallelAnimationGroup(self)
            self._karta_anim.addAnimation(self._anim_poz)
            self._karta_anim.addAnimation(self._anim_op)
            QTimer.singleShot(380, self._karta_anim.start)
        except Exception:
            pass

    def stop(self):
        self._timer.stop()
        self.hide()

    def _krok(self):
        self._kat = (self._kat + 0.25) % 360
        self._t += 1
        for tr in self._trasy:
            tr["faza"] += tr["predkosc"]
            if tr["faza"] > 1.0:
                tr["faza"] -= 1.0
        self.update()

    def update_theme(self, is_dark):
        self.is_dark = is_dark
        if hasattr(self, "karta_dzis"):
            self.karta_dzis.update_theme(is_dark)
        self.update()

    def paintEvent(self, e):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        W, H = self.width(), self.height()
        # karta "Dziś w trasie" zajmuje dół — kompas i powitanie idą wyżej,
        # żeby nic na siebie nie nachodziło
        przesun = 78 if (hasattr(self, "karta_dzis") and self.karta_dzis.isVisible()) else 20
        cx, cy = W / 2, H / 2 - przesun

        if self.is_dark:
            akc = QColor(0, 240, 255)
            txt = QColor(248, 250, 252); txt_mut = QColor(203, 213, 225)
            ring = QColor(0, 240, 255, 90)
            przyciem = QColor(4, 8, 16, 130)     # delikatne przyciemnienie pod treścią
        else:
            akc = QColor(13, 148, 136)
            txt = QColor(248, 250, 252); txt_mut = QColor(226, 232, 240)
            ring = QColor(255, 255, 255, 120)
            przyciem = QColor(10, 20, 30, 90)

        # NIE malujemy pełnego tła — widget jest przezroczysty, więc prześwituje
        # tło aplikacji (góry + świecąca droga). Dajemy tylko miękką, owalną
        # poświatę pod kompasem, żeby tekst i kompas były czytelne na jasnych
        # fragmentach tła (np. zachód słońca).
        poswiata = QRadialGradient(cx, cy, min(W, H) * 0.55)
        poswiata.setColorAt(0.0, przyciem)
        k_kraw = QColor(przyciem); k_kraw.setAlpha(0)
        poswiata.setColorAt(1.0, k_kraw)
        p.setBrush(poswiata); p.setPen(Qt.PenStyle.NoPen)
        p.drawEllipse(QPointF(cx, cy), W * 0.55, H * 0.55)

        # --- świecące trasy: linie od krawędzi ku centrum, rysujące się i gasnące ---
        maxr = math.hypot(W, H) / 2
        for tr in self._trasy:
            faza = tr["faza"]
            r_od = maxr * tr["dlugosc"] * (1.0 - faza)
            r_do = r_od - maxr * 0.18
            if r_do < 40:
                r_do = 40
            x1 = cx + math.cos(tr["kat"]) * r_od
            y1 = cy + math.sin(tr["kat"]) * r_od
            x2 = cx + math.cos(tr["kat"]) * r_do
            y2 = cy + math.sin(tr["kat"]) * r_do
            # jasność zależna od fazy (najjaśniej w środku ruchu)
            alpha = int(180 * math.sin(faza * math.pi))
            kolor = QColor(akc); kolor.setAlpha(max(0, min(alpha, 180)))
            gr = QLinearGradient(x1, y1, x2, y2)
            k0 = QColor(kolor); k0.setAlpha(0)
            gr.setColorAt(0, k0); gr.setColorAt(1, kolor)
            pen = QPen(QBrush(gr), 2.2)
            pen.setCapStyle(Qt.PenCapStyle.RoundCap)
            p.setPen(pen)
            p.drawLine(QPointF(x1, y1), QPointF(x2, y2))
            # świecący punkt na końcu (bliżej centrum)
            p.setBrush(kolor); p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(QPointF(x2, y2), 2.5, 2.5)

        # --- pierścień kompasu (obracający się) ---
        R = min(W, H) * 0.16
        p.save()
        p.translate(cx, cy)
        # zewnętrzny pierścień
        p.setBrush(Qt.BrushStyle.NoBrush)
        p.setPen(QPen(ring, 1.5))
        p.drawEllipse(QPointF(0, 0), R + 26, R + 26)
        p.setPen(QPen(ring, 1))
        p.drawEllipse(QPointF(0, 0), R + 40, R + 40)
        # znaczniki kompasu, obracające się
        p.rotate(self._kat)
        for i in range(72):
            duzy = (i % 9 == 0)
            dl = 12 if duzy else 6
            kol = QColor(akc if duzy else ring)
            if not duzy:
                kol.setAlpha(90)
            p.setPen(QPen(kol, 2 if duzy else 1))
            p.drawLine(QPointF(0, -(R + 40)), QPointF(0, -(R + 40 - dl)))
            p.rotate(5)
        p.restore()

        # kierunki świata (nieruchome, N na górze)
        p.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
        for etyk, dx, dy in [("N", 0, -(R + 58)), ("E", R + 58, 0), ("S", 0, R + 58), ("W", -(R + 58), 0)]:
            p.setPen(txt_mut)
            p.drawText(QRectF(cx + dx - 12, cy + dy - 10, 24, 20),
                       Qt.AlignmentFlag.AlignCenter, etyk)

        # --- logo z poświatą w centrum ---
        if self._logo_pix:
            rozmiar = int(R * 1.35)      # nieco mniejsze — napis mieści się w kuli
            r_kolo = rozmiar / 2
            # poświata pod logo
            pos = QRadialGradient(cx, cy, rozmiar * 0.95)
            g0 = QColor(akc); g0.setAlpha(70); g1 = QColor(akc); g1.setAlpha(0)
            pos.setColorAt(0, g0); pos.setColorAt(1, g1)
            p.setBrush(pos); p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(QPointF(cx, cy), rozmiar * 0.95, rozmiar * 0.95)

            # Białe koło-podkład — gdy logo jest prostokątne (szersze niż wyższe),
            # po wpisaniu całego w koło zostają puste pasy u góry/dołu. Podkład
            # w kolorze tła logo (biały) wypełnia je spójnie, więc widać pełne koło.
            p.setBrush(QColor(255, 255, 255)); p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(QPointF(cx, cy), r_kolo, r_kolo)

            # Logo mieści się CAŁE w kole (KeepAspectRatio) — nic nie obcinamy,
            # więc litery P i T po bokach są w całości widoczne. Zajmuje ~88%
            # średnicy koła. Renderujemy w 2x dla ostrości.
            dpr = 2
            cel = int(rozmiar * 0.88)
            skala = self._logo_pix.scaled(int(cel * dpr), int(cel * dpr),
                                          Qt.AspectRatioMode.KeepAspectRatio,
                                          Qt.TransformationMode.SmoothTransformation)
            skala.setDevicePixelRatio(dpr)
            p.save()
            klip = QPainterPath()
            klip.addEllipse(QPointF(cx, cy), r_kolo, r_kolo)
            p.setClipPath(klip)
            szer_log = skala.width() / dpr; wys_log = skala.height() / dpr
            p.drawPixmap(int(cx - szer_log / 2), int(cy - wys_log / 2),
                         int(szer_log), int(wys_log), skala)
            p.restore()

            # subtelna obwódka koła w kolorze akcentu
            p.setBrush(Qt.BrushStyle.NoBrush)
            obw = QColor(akc); obw.setAlpha(160)
            p.setPen(QPen(obw, 2.5))
            p.drawEllipse(QPointF(cx, cy), r_kolo, r_kolo)
        else:
            # brak logo — narysuj tekst PMT
            p.setPen(akc); p.setFont(QFont("Segoe UI", 34, QFont.Weight.Black))
            p.drawText(QRectF(cx - R, cy - R, 2 * R, 2 * R),
                       Qt.AlignmentFlag.AlignCenter, "PMT")

        # --- powitanie ---
        p.setPen(txt); p.setFont(QFont("Segoe UI", 22, QFont.Weight.Bold))
        p.drawText(QRectF(0, cy + R + 46, W, 40),
                   Qt.AlignmentFlag.AlignHCenter, self._powitanie)

        # --- pasek na dole: wersja + data ---
        p.setPen(txt_mut); p.setFont(QFont("Segoe UI", 9))
        stopka = f"PMT Planer  •  wersja {WERSJA_PROGRAMU}  •  {datetime.datetime.now().strftime('%d.%m.%Y')}"
        p.drawText(QRectF(0, H - 34, W, 20), Qt.AlignmentFlag.AlignHCenter, stopka)


class DialogWyboru(QFrame):
    """Autorski dialog wyboru w stylu programu (zamiast systemowego QMessageBox).
    Półprzezroczysta przesłona + karta z pytaniem i dwoma przyciskami. Zwraca
    wybór przez exec_wybor() → 'a' (pierwszy przycisk) lub 'b' (drugi)."""
    def __init__(self, parent, tytul, pytanie, opis, tekst_a, tekst_b, is_dark=True):
        super().__init__(parent)
        self._wynik = None
        self.is_dark = is_dark
        self.setGeometry(parent.rect())
        # przesłona
        self.setStyleSheet("background: rgba(0,0,0,0.55);")

        if is_dark:
            karta_bg = "#0B1320"; ramka = "rgba(0,240,255,0.30)"
            txt = "#F8FAFC"; txt_mut = "#94A3B8"
            akc1, akc2 = "#00F0FF", "#00E4A1"
            btn2_bg = "rgba(255,255,255,0.06)"; btn2_br = "rgba(255,255,255,0.18)"
        else:
            karta_bg = "#FFFFFF"; ramka = "rgba(13,148,136,0.30)"
            txt = "#0F172A"; txt_mut = "#64748B"
            akc1, akc2 = "#0D9488", "#059669"
            btn2_bg = "rgba(15,23,42,0.05)"; btn2_br = "rgba(15,23,42,0.15)"

        # karta na środku
        self.karta = QFrame(self)
        self.karta.setObjectName("KartaDialog")
        self.karta.setStyleSheet(
            f"#KartaDialog {{ background:{karta_bg}; border:1px solid {ramka}; border-radius:18px; }}")
        cien = QGraphicsDropShadowEffect(self.karta); cien.setBlurRadius(40)
        cien.setColor(QColor(0,0,0,160)); cien.setOffset(0,12); self.karta.setGraphicsEffect(cien)

        lay = QVBoxLayout(self.karta); lay.setContentsMargins(28, 26, 28, 24); lay.setSpacing(14)
        l_tyt = QLabel(tytul)
        l_tyt.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:11px; font-weight:700; letter-spacing:1px; background:transparent;")
        l_pyt = QLabel(pytanie); l_pyt.setWordWrap(True)
        l_pyt.setStyleSheet(f"color:{txt}; font-family:'Segoe UI'; font-size:16px; font-weight:700; background:transparent;")
        lay.addWidget(l_tyt); lay.addWidget(l_pyt)
        if opis:
            l_opis = QLabel(opis); l_opis.setWordWrap(True)
            l_opis.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:12px; background:transparent;")
            lay.addWidget(l_opis)
        lay.addSpacing(4)

        rzad = QHBoxLayout(); rzad.setSpacing(12)
        self.btn_a = QPushButton(tekst_a); self.btn_a.setFixedHeight(44)
        self.btn_a.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_a.setStyleSheet(
            f"QPushButton {{ color:#04121A; font-family:'Segoe UI'; font-size:13px; font-weight:800;"
            f" border:none; border-radius:11px;"
            f" background: qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc1}, stop:1 {akc2}); }}"
            f"QPushButton:hover {{ background:{akc1}; }}")
        self.btn_b = QPushButton(tekst_b); self.btn_b.setFixedHeight(44)
        self.btn_b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_b.setStyleSheet(
            f"QPushButton {{ color:{txt}; font-family:'Segoe UI'; font-size:13px; font-weight:700;"
            f" border:1px solid {btn2_br}; border-radius:11px; background:{btn2_bg}; }}"
            f"QPushButton:hover {{ background:{btn2_br}; }}")
        rzad.addWidget(self.btn_a, 1); rzad.addWidget(self.btn_b, 1)
        lay.addLayout(rzad)

        self.btn_a.clicked.connect(lambda: self._zamknij('a'))
        self.btn_b.clicked.connect(lambda: self._zamknij('b'))

        self.karta.setFixedWidth(440)
        self.karta.adjustSize()
        self._wysrodkuj()

        self._loop = None

    def _wysrodkuj(self):
        kw, kh = self.karta.width(), self.karta.height()
        self.karta.move((self.width()-kw)//2, (self.height()-kh)//2)

    def _zamknij(self, w):
        self._wynik = w
        if self._loop is not None:
            self._loop.quit()
        self.hide(); self.deleteLater()

    def exec_wybor(self):
        """Pokazuje dialog i BLOKUJE do wyboru. Zwraca 'a' lub 'b'."""
        from PyQt6.QtCore import QEventLoop
        self.show(); self.raise_()
        # animacja karty: lekki wjazd od dołu
        start = self.karta.pos()
        self.karta.move(start.x(), start.y() + 30)
        anim = QPropertyAnimation(self.karta, b"pos", self)
        anim.setDuration(220); anim.setStartValue(self.karta.pos())
        anim.setEndValue(start); anim.setEasingCurve(QEasingCurve.Type.OutCubic)
        anim.start()
        self._loop = QEventLoop()
        self._loop.exec()
        return self._wynik or 'a'


class LogoKompas(QWidget):
    """Małe okrągłe logo PMT z obracającym się wokół niego pierścieniem kompasu
    — mini-wersja ekranu powitalnego, do topbara. Spójne wizualnie z powitalnym."""
    def __init__(self, parent=None, srednica=76, is_dark=True):
        super().__init__(parent)
        self.is_dark = is_dark
        self._srednica = srednica
        self.setFixedSize(srednica, srednica)
        self._kat = 0.0
        self._logo_pix = None
        sciezka = znajdz_logo()
        if sciezka:
            pix = QPixmap(sciezka)
            if not pix.isNull():
                self._logo_pix = pix
        self._timer = QTimer(self)
        self._timer.timeout.connect(self._krok)
        self._timer.start(33)          # ~30 fps, spokojny obrót

    def _krok(self):
        self._kat = (self._kat + 0.25) % 360
        self.update()

    def update_theme(self, is_dark):
        self.is_dark = is_dark
        self.update()

    def paintEvent(self, e):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        p.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        W = H = self._srednica
        cx, cy = W / 2, H / 2
        akc = QColor(0, 240, 255) if self.is_dark else QColor(13, 148, 136)
        ring = QColor(akc); ring.setAlpha(90)

        R = W * 0.34          # promień koła logo
        # obracający się pierścień znaczników
        p.save(); p.translate(cx, cy); p.rotate(self._kat)
        for i in range(48):
            duzy = (i % 6 == 0)
            dl = 5 if duzy else 3
            kol = QColor(akc) if duzy else QColor(ring)
            if not duzy: kol.setAlpha(70)
            p.setPen(QPen(kol, 1.4 if duzy else 1))
            p.drawLine(QPointF(0, -(R + 8)), QPointF(0, -(R + 8 - dl)))
            p.rotate(7.5)
        p.restore()
        # cienki pierścień statyczny
        p.setBrush(Qt.BrushStyle.NoBrush)
        p.setPen(QPen(ring, 1))
        p.drawEllipse(QPointF(cx, cy), R + 8, R + 8)

        # poświata
        pos = QRadialGradient(cx, cy, R * 1.5)
        g0 = QColor(akc); g0.setAlpha(60); g1 = QColor(akc); g1.setAlpha(0)
        pos.setColorAt(0, g0); pos.setColorAt(1, g1)
        p.setBrush(pos); p.setPen(Qt.PenStyle.NoPen)
        p.drawEllipse(QPointF(cx, cy), R * 1.5, R * 1.5)

        if self._logo_pix:
            # białe koło-podkład (wypełnia puste pasy gdy logo prostokątne)
            p.setBrush(QColor(255, 255, 255)); p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(QPointF(cx, cy), R, R)
            # logo mieści się CAŁE w kole (KeepAspectRatio) — P i T nieobcięte
            dpr = 2
            d = int(R * 2 * 0.90)
            skala = self._logo_pix.scaled(int(d * dpr), int(d * dpr),
                                          Qt.AspectRatioMode.KeepAspectRatio,
                                          Qt.TransformationMode.SmoothTransformation)
            skala.setDevicePixelRatio(dpr)
            p.save()
            klip = QPainterPath(); klip.addEllipse(QPointF(cx, cy), R, R)
            p.setClipPath(klip)
            sw = skala.width() / dpr; sh = skala.height() / dpr
            p.drawPixmap(int(cx - sw/2), int(cy - sh/2), int(sw), int(sh), skala)
            p.restore()
            # obwódka
            obw = QColor(akc); obw.setAlpha(160)
            p.setBrush(Qt.BrushStyle.NoBrush); p.setPen(QPen(obw, 1.8))
            p.drawEllipse(QPointF(cx, cy), R, R)
        else:
            p.setPen(akc); p.setFont(QFont("Segoe UI", 14, QFont.Weight.Black))
            p.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, "PMT")


class WykresMiesieczny(QWidget):
    """Autorski wykres słupkowy (QPainter) — wizyty w ostatnich miesiącach.
    Bez zewnętrznych bibliotek wykresów, w stylu wizualnym reszty programu."""
    def __init__(self, parent=None, is_dark=True):
        super().__init__(parent)
        self.is_dark = is_dark
        self.dane = []       # [(etykieta, wartość), ...]
        self.setMinimumHeight(180)

    def ustaw_dane(self, dane):
        self.dane = dane or []
        self.update()

    def paintEvent(self, event):
        p = QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        W, H = self.width(), self.height()
        if not self.dane:
            return
        akc = QColor("#00F0FF") if self.is_dark else QColor("#0D9488")
        zie = QColor("#00E4A1") if self.is_dark else QColor("#059669")
        mut = QColor("#94A3B8") if self.is_dark else QColor("#64748B")
        txt = QColor("#F8FAFC") if self.is_dark else QColor("#0F172A")

        margines_dol = 26
        margines_gora = 22
        n = len(self.dane)
        maks = max((v for _, v in self.dane), default=0) or 1
        szer_slupka = min(56, (W - 20) / n * 0.55)
        odstep = (W - 20) / n

        for i, (etyk, wart) in enumerate(self.dane):
            cx = 10 + odstep * i + odstep / 2
            wys = (H - margines_dol - margines_gora) * (wart / maks) if maks else 0
            wys = max(3, wys) if wart > 0 else 0
            gora_y = H - margines_dol - wys
            rect = QRectF(cx - szer_slupka/2, gora_y, szer_slupka, wys)
            if i == n - 1:
                grad = QLinearGradient(0, gora_y, 0, H - margines_dol)
                grad.setColorAt(0, zie); grad.setColorAt(1, akc)
                p.setBrush(QBrush(grad))
            else:
                kol = QColor(akc); kol.setAlpha(110)
                p.setBrush(kol)
            p.setPen(Qt.PenStyle.NoPen)
            p.drawRoundedRect(rect, 5, 5)

            # wartość nad słupkiem
            if wart > 0:
                p.setPen(txt)
                p.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
                p.drawText(QRectF(cx - odstep/2, gora_y - 16, odstep, 14),
                           Qt.AlignmentFlag.AlignCenter, str(wart))
            # etykieta miesiąca
            p.setPen(mut)
            p.setFont(QFont("Segoe UI", 9))
            p.drawText(QRectF(cx - odstep/2, H - margines_dol + 6, odstep, 16),
                       Qt.AlignmentFlag.AlignCenter, etyk)


class StatystykiOverlay(QFrame):
    """'Twoja praca w liczbach' — osobisty panel analityczny. Dwie zakładki:
    Wizyty (PRAWDZIWE dane z dziennika — ile wizyt, jaki rytm, sieci) i
    Delegacje (dawne 'Moje Szlaki' — koszty, km, województwa ze starego
    silnika). Jeden panel na WSZYSTKIE osobiste liczby, zamiast dwóch
    osobnych pozycji w menu bocznym."""
    def __init__(self, parent=None, on_dane_uzytkownika=None):
        super().__init__(parent)
        self.is_dark = True
        self.hide()
        self._on_dane_uzytkownika = on_dane_uzytkownika
        self._zakladka = "wizyty"
        self._rok_del = datetime.datetime.now().year
        self.setStyleSheet("background: rgba(3, 7, 14, 0.90);")
        lay = QVBoxLayout(self); lay.setContentsMargins(36, 28, 36, 28); lay.setSpacing(18)

        gora = QHBoxLayout()
        naglowek = QVBoxLayout(); naglowek.setSpacing(2)
        self.l_tyt = QLabel("Twoja praca w liczbach")
        self.l_tyt.setStyleSheet("color:#F8FAFC; font-family:'Segoe UI'; font-size:22px; font-weight:800; background:transparent;")
        self.l_pod = QLabel("Zbudowane z rzeczywistej historii odhaczonych wizyt")
        self.l_pod.setStyleSheet("color:#94A3B8; font-family:'Segoe UI'; font-size:12px; background:transparent;")
        naglowek.addWidget(self.l_tyt); naglowek.addWidget(self.l_pod)
        gora.addLayout(naglowek); gora.addStretch()

        # zakładki — Wizyty (nowy silnik) / Delegacje (dawne "Moje Szlaki")
        self.btn_zak_wizyty = QPushButton("📆  Wizyty"); self.btn_zak_wizyty.setCheckable(True); self.btn_zak_wizyty.setChecked(True)
        self.btn_zak_delegacje = QPushButton("💰  Delegacje"); self.btn_zak_delegacje.setCheckable(True)
        for b in (self.btn_zak_wizyty, self.btn_zak_delegacje):
            b.setFixedHeight(34); b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_zak_wizyty.clicked.connect(lambda: self._przelacz_zakladke("wizyty"))
        self.btn_zak_delegacje.clicked.connect(lambda: self._przelacz_zakladke("delegacje"))
        gora.addWidget(self.btn_zak_wizyty); gora.addWidget(self.btn_zak_delegacje)
        gora.addSpacing(10)

        # przełącznik roku — TYLKO w zakładce Delegacje
        self.btn_rok_prev = QPushButton("‹"); self.btn_rok_prev.setFixedSize(30, 30)
        self.lbl_rok = QLabel(""); self.lbl_rok.setFixedWidth(60); self.lbl_rok.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.btn_rok_next = QPushButton("›"); self.btn_rok_next.setFixedSize(30, 30)
        for b in (self.btn_rok_prev, self.btn_rok_next): b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_rok_prev.clicked.connect(lambda: self._zmien_rok_del(-1))
        self.btn_rok_next.clicked.connect(lambda: self._zmien_rok_del(1))
        self._rok_widgets = [self.btn_rok_prev, self.lbl_rok, self.btn_rok_next]
        for w in self._rok_widgets:
            gora.addWidget(w)
            w.setVisible(False)
        gora.addSpacing(8)

        self.btn_x = QPushButton("✕"); self.btn_x.setFixedSize(38, 38)
        self.btn_x.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_x.setStyleSheet(
            "QPushButton { color:#F8FAFC; background:rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.15); border-radius:10px; font-size:15px; }"
            "QPushButton:hover { background:rgba(239,68,68,0.85); }")
        self.btn_x.clicked.connect(self.hide)
        gora.addWidget(self.btn_x)
        lay.addLayout(gora)

        # === ZAKŁADKA: WIZYTY (istniejąca zawartość, opakowana w panel) ===
        self.panel_wizyty = QWidget()
        pw = QVBoxLayout(self.panel_wizyty); pw.setContentsMargins(0, 0, 0, 0); pw.setSpacing(18)

        self.kafle = QHBoxLayout(); self.kafle.setSpacing(14)
        self._kafle_lbl = {}
        for klucz, tytul in [("laczna", "wizyt łącznie"), ("mies", "w tym miesiącu"),
                             ("passa", "dni z rzędu w 100%"), ("km", "km z ukończonych dni")]:
            box = QFrame(); box.setObjectName("StatKafel")
            box.setStyleSheet("#StatKafel { background:rgba(11,19,32,0.95); border:1px solid rgba(0,240,255,0.22); border-radius:14px; }")
            bl = QVBoxLayout(box); bl.setContentsMargins(18, 14, 18, 14); bl.setSpacing(4)
            l_val = QLabel("0"); l_val.setStyleSheet("color:#00F0FF; font-family:'Segoe UI'; font-size:26px; font-weight:800; background:transparent;")
            l_op = QLabel(tytul); l_op.setStyleSheet("color:#94A3B8; font-family:'Segoe UI'; font-size:11px; background:transparent;")
            bl.addWidget(l_val); bl.addWidget(l_op)
            self._kafle_lbl[klucz] = l_val
            self.kafle.addWidget(box, 1)
        pw.addLayout(self.kafle)

        srodek = QHBoxLayout(); srodek.setSpacing(18)

        lewa_karta = QFrame(); lewa_karta.setObjectName("StatKarta")
        lewa_karta.setStyleSheet("#StatKarta { background:rgba(11,19,32,0.85); border:1px solid rgba(0,240,255,0.18); border-radius:14px; }")
        ll = QVBoxLayout(lewa_karta); ll.setContentsMargins(18, 16, 18, 16); ll.setSpacing(10)
        lt = QLabel("Wizyty w ostatnich miesiącach")
        lt.setStyleSheet("color:#00F0FF; font-family:'Segoe UI'; font-size:13px; font-weight:800; background:transparent;")
        ll.addWidget(lt)
        self.wykres = WykresMiesieczny(self, is_dark=True)
        ll.addWidget(self.wykres, 1)
        srodek.addWidget(lewa_karta, 3)

        prawa_karta = QFrame(); prawa_karta.setObjectName("StatKarta2")
        prawa_karta.setStyleSheet("#StatKarta2 { background:rgba(11,19,32,0.85); border:1px solid rgba(0,240,255,0.18); border-radius:14px; }")
        pl = QVBoxLayout(prawa_karta); pl.setContentsMargins(18, 16, 18, 16); pl.setSpacing(8)
        pt = QLabel("Podział na sieci")
        pt.setStyleSheet("color:#00F0FF; font-family:'Segoe UI'; font-size:13px; font-weight:800; background:transparent;")
        pl.addWidget(pt)
        self.wiersze_siec_wrap = QVBoxLayout(); self.wiersze_siec_wrap.setSpacing(8)
        pl.addLayout(self.wiersze_siec_wrap)
        pl.addSpacing(6)
        sep = QFrame(); sep.setFixedHeight(1); sep.setStyleSheet("background:rgba(0,240,255,0.15);")
        pl.addWidget(sep)
        pl.addSpacing(6)
        rt = QLabel("🏆  Twój rekord dnia")
        rt.setStyleSheet("color:#00E4A1; font-family:'Segoe UI'; font-size:12px; font-weight:700; background:transparent;")
        pl.addWidget(rt)
        self.l_rekord = QLabel("—")
        self.l_rekord.setWordWrap(True)
        self.l_rekord.setStyleSheet("color:#F8FAFC; font-family:'Segoe UI'; font-size:12px; background:transparent;")
        pl.addWidget(self.l_rekord)
        pl.addStretch()
        srodek.addWidget(prawa_karta, 2)

        pw.addLayout(srodek, 1)
        lay.addWidget(self.panel_wizyty, 1)

        # === ZAKŁADKA: DELEGACJE (dawne "Moje Szlaki", scalone tutaj) ===
        self.panel_delegacje = QWidget()
        self.panel_delegacje.setVisible(False)
        pd = QVBoxLayout(self.panel_delegacje); pd.setContentsMargins(0, 0, 0, 0); pd.setSpacing(14)

        self.staty_del = QHBoxLayout(); self.staty_del.setSpacing(12)
        self.karty_stat_del = []
        for _ in range(4):
            k = QFrame(); k.setObjectName("StatKarDel"); k.setFixedHeight(76)
            kv = QVBoxLayout(k); kv.setContentsMargins(14, 10, 14, 10); kv.setSpacing(2)
            wart = QLabel("—"); wart.setObjectName("StatWartDel")
            opis = QLabel(""); opis.setObjectName("StatOpisDel")
            kv.addWidget(wart); kv.addWidget(opis)
            self.staty_del.addWidget(k); self.karty_stat_del.append((wart, opis))
        pd.addLayout(self.staty_del)

        wykresy_gora = QHBoxLayout(); wykresy_gora.setSpacing(16)
        self.wyk_koszty = WykresSlupkowy(); self.wyk_koszty.setObjectName("WykBoxDel")
        self.wyk_km = WykresSlupkowy(); self.wyk_km.setObjectName("WykBoxDel")
        wykresy_gora.addWidget(self.wyk_koszty); wykresy_gora.addWidget(self.wyk_km)
        pd.addLayout(wykresy_gora, 1)

        donuty = QHBoxLayout(); donuty.setSpacing(16)
        self.wyk_woj = WykresDonut(); self.wyk_woj.setObjectName("WykBoxDel")
        self.wyk_bazy = WykresDonut(); self.wyk_bazy.setObjectName("WykBoxDel")
        donuty.addWidget(self.wyk_woj); donuty.addWidget(self.wyk_bazy)
        pd.addLayout(donuty, 1)

        lay.addWidget(self.panel_delegacje, 1)

    def _przelacz_zakladke(self, tryb):
        self._zakladka = tryb
        self.btn_zak_wizyty.setChecked(tryb == "wizyty")
        self.btn_zak_delegacje.setChecked(tryb == "delegacje")
        self.panel_wizyty.setVisible(tryb == "wizyty")
        self.panel_delegacje.setVisible(tryb == "delegacje")
        for w in self._rok_widgets:
            w.setVisible(tryb == "delegacje")
        if tryb == "delegacje" and not getattr(self, "_delegacje_wczytane", False):
            self._wczytaj_delegacje()

    def _wczytaj_delegacje(self):
        imie = pesel = ""
        if self._on_dane_uzytkownika:
            imie, pesel = self._on_dane_uzytkownika()
        self._historia_del = wczytaj_historie(imie, pesel)
        self._delegacje_wczytane = True
        self._przelicz_delegacje()

    def _zmien_rok_del(self, d):
        self._rok_del += d
        self._przelicz_delegacje()

    def _przelicz_delegacje(self):
        self.lbl_rok.setText(str(self._rok_del))
        hist = getattr(self, "_historia_del", [])
        wpisy = []
        for h in hist:
            r = h.get("rok")
            if r is None:
                try: r = int(h.get("data", "").split(".")[2].split(" ")[0])
                except Exception: r = None
            if r == self._rok_del:
                wpisy.append(h)

        koszty_mies = [0.0] * 12
        km_mies = [0.0] * 12
        woj_licznik = {}
        bazy_licznik = {}
        suma_km = 0.0; suma_kwota = 0.0
        for h in wpisy:
            m = h.get("miesiac")
            if m is None:
                try: m = int(h.get("data", "").split(".")[1])
                except Exception: m = None
            kwota = _parsuj_kwote(h.get("kwota", 0))
            km = float(h.get("km", 0) or 0)
            suma_km += km; suma_kwota += kwota
            if m and 1 <= m <= 12:
                koszty_mies[m - 1] += kwota; km_mies[m - 1] += km
            w_wizyty = h.get("woj_wizyty")
            if isinstance(w_wizyty, dict) and w_wizyty:
                for wnaz, ile in w_wizyty.items():
                    try: ile = int(ile)
                    except Exception: ile = 1
                    woj_licznik[wnaz] = woj_licznik.get(wnaz, 0) + ile
            else:
                w = h.get("woj", "—")
                woj_licznik[w] = woj_licznik.get(w, 0) + 1
            m_wizyty = h.get("miejsc_wizyty")
            if isinstance(m_wizyty, dict):
                for mnaz, ile in m_wizyty.items():
                    try: ile = int(ile)
                    except Exception: ile = 1
                    bazy_licznik[mnaz] = bazy_licznik.get(mnaz, 0) + ile

        self.karty_stat_del[0][0].setText(f"{len(wpisy)}"); self.karty_stat_del[0][1].setText("wygenerowanych wypraw")
        self.karty_stat_del[1][0].setText(f"{int(suma_km):,}".replace(",", " ") + " km"); self.karty_stat_del[1][1].setText("łącznie w roku")
        self.karty_stat_del[2][0].setText(f"{suma_kwota:,.0f} zł".replace(",", " ")); self.karty_stat_del[2][1].setText("suma rozliczeń")
        naj = max(woj_licznik.items(), key=lambda x: x[1])[0] if woj_licznik else "—"
        self.karty_stat_del[3][0].setText(naj); self.karty_stat_del[3][1].setText("najczęstszy region")

        skrot = ["Sty", "Lut", "Mar", "Kwi", "Maj", "Cze", "Lip", "Sie", "Wrz", "Paź", "Lis", "Gru"]
        self.wyk_koszty.ustaw_dane(list(zip(skrot, koszty_mies)), "Koszty miesięczne", " zł",
                                   "#00F0FF" if self.is_dark else "#0D9488",
                                   "#00E4A1" if self.is_dark else "#059669")
        self.wyk_km.ustaw_dane(list(zip(skrot, km_mies)), "Kilometry miesięcznie", "", "#F59E0B", "#EC4899")
        self.wyk_woj.ustaw_dane(list(woj_licznik.items()), "Udział województw")
        top_miejsc = sorted(bazy_licznik.items(), key=lambda x: -x[1])[:3]
        self.wyk_bazy.ustaw_dane(top_miejsc, "Najczęściej odwiedzane")

    def update_theme(self, is_dark):
        self.is_dark = is_dark
        self.wykres.is_dark = is_dark
        for w in (self.wyk_koszty, self.wyk_km, self.wyk_woj, self.wyk_bazy):
            w.is_dark = is_dark; w.update()
        if is_dark:
            self.setStyleSheet("background: rgba(3, 7, 14, 0.92);")
            self.l_tyt.setStyleSheet("color:#F8FAFC; font-family:'Segoe UI'; font-size:22px; font-weight:800; background:transparent;")
            akc = "#00F0FF"; karta_bg = "rgba(11,19,32,0.95)"; karta_br = "rgba(0,240,255,0.20)"
            txt_kol = "#F8FAFC"; mut = "#94A3B8"; ramka = "rgba(0,240,255,0.25)"; sec = "rgba(255,255,255,0.08)"
            box = "rgba(5,10,20,0.4)"; statbg = "rgba(0,240,255,0.06)"
        else:
            self.setStyleSheet("background: rgba(226, 232, 240, 0.96);")
            self.l_tyt.setStyleSheet("color:#0F172A; font-family:'Segoe UI'; font-size:22px; font-weight:800; background:transparent;")
            akc = "#0D9488"; karta_bg = "rgba(255,255,255,0.95)"; karta_br = "rgba(13,148,136,0.25)"
            txt_kol = "#0F172A"; mut = "#64748B"; ramka = "rgba(13,148,136,0.30)"; sec = "rgba(15,23,42,0.08)"
            box = "rgba(255,255,255,0.6)"; statbg = "rgba(13,148,136,0.07)"
        self.l_pod.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:12px; background:transparent;")
        for l in self._kafle_lbl.values():
            l.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:26px; font-weight:800; background:transparent;")
        for nazwa_obj in ("StatKafel", "StatKarta", "StatKarta2"):
            for w in self.findChildren(QFrame, nazwa_obj):
                w.setStyleSheet(f"#{nazwa_obj} {{ background:{karta_bg}; border:1px solid {karta_br}; border-radius:14px; }}")
        self.l_rekord.setStyleSheet(f"color:{txt_kol}; font-family:'Segoe UI'; font-size:12px; background:transparent;")
        self.wykres.update()
        self._odswiez_wiersze_siec(getattr(self, "_ostatnie_per_siec", {}), akc, txt_kol, mut)

        # --- zakładki + zawartość Delegacje ---
        zak_css = (f"QPushButton {{ color:{txt_kol}; background:transparent; border:1px solid {ramka}; border-radius:9px; "
                  f"font-family:'Segoe UI'; font-size:12px; font-weight:700; padding:0 12px; }} "
                  f"QPushButton:checked {{ background:{akc}; color:#04121A; border-color:{akc}; }} "
                  f"QPushButton:hover {{ border-color:{akc}; }}")
        self.btn_zak_wizyty.setStyleSheet(zak_css); self.btn_zak_delegacje.setStyleSheet(zak_css)
        self.lbl_rok.setStyleSheet(f"color:{txt_kol}; font-family:'Segoe UI'; font-size:14px; font-weight:700; background:transparent;")
        nav_css = f"QPushButton {{ color:{txt_kol}; background:{sec}; border:1px solid {ramka}; border-radius:8px; font-size:14px; font-weight:bold; }} QPushButton:hover {{ background:{ramka}; }}"
        self.btn_rok_prev.setStyleSheet(nav_css); self.btn_rok_next.setStyleSheet(nav_css)
        for wart, opis in self.karty_stat_del:
            wart.parent().setStyleSheet(f"#StatKarDel {{ background:{statbg}; border:1px solid {ramka}; border-radius:12px; }}")
            wart.setStyleSheet(f"#StatWartDel {{ color:{akc}; font-family:'Segoe UI'; font-size:18px; font-weight:800; background:transparent; }}")
            opis.setStyleSheet(f"#StatOpisDel {{ color:{mut}; font-family:'Segoe UI'; font-size:10px; background:transparent; }}")
        for w in (self.wyk_koszty, self.wyk_km, self.wyk_woj, self.wyk_bazy):
            w.setStyleSheet(f"#WykBoxDel {{ background:{box}; border:1px solid {ramka}; border-radius:12px; }}")
        if getattr(self, "_delegacje_wczytane", False):
            self._przelicz_delegacje()

    def _odswiez_wiersze_siec(self, per_siec, akc, txt_kol, mut):
        while self.wiersze_siec_wrap.count():
            it = self.wiersze_siec_wrap.takeAt(0)
            w = it.widget()
            if w: w.deleteLater()
        if not per_siec:
            pusty = QLabel("Brak jeszcze danych.")
            pusty.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
            self.wiersze_siec_wrap.addWidget(pusty)
            return
        maks = max(per_siec.values())
        for siec, n in list(per_siec.items())[:6]:
            wiersz = QVBoxLayout(); wiersz.setSpacing(2)
            gora = QHBoxLayout()
            l1 = QLabel(siec); l1.setStyleSheet(f"color:{txt_kol}; font-family:'Segoe UI'; font-size:11px; font-weight:600; background:transparent;")
            l2 = QLabel(str(n)); l2.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:11px; font-weight:800; background:transparent;")
            gora.addWidget(l1); gora.addStretch(); gora.addWidget(l2)
            wiersz.addLayout(gora)
            pasek_tlo = QFrame(); pasek_tlo.setFixedHeight(6)
            pasek_tlo.setStyleSheet("background:rgba(148,163,184,0.25); border-radius:3px;")
            pasek_l = QHBoxLayout(pasek_tlo); pasek_l.setContentsMargins(0,0,0,0)
            pasek = QFrame(); pasek.setFixedHeight(6)
            proc = max(4, int(160 * n / maks))
            pasek.setFixedWidth(proc)
            pasek.setStyleSheet(f"background:{akc}; border-radius:3px;")
            pasek_l.addWidget(pasek); pasek_l.addStretch()
            wiersz.addWidget(pasek_tlo)
            self.wiersze_siec_wrap.addLayout(wiersz)

    def odswiez_dane(self):
        s = oblicz_statystyki_osobiste()
        self._kafle_lbl["laczna"].setText(str(s["lacznie_wizyt"]))
        self._kafle_lbl["mies"].setText(str(s["w_tym_miesiacu"]))
        self._kafle_lbl["passa"].setText(str(s["passa_dni"]))
        self._kafle_lbl["km"].setText(f"{s['suma_km_ukonczone']:.0f}")
        self.wykres.ustaw_dane(s["per_miesiac"])
        self._ostatnie_per_siec = s["per_siec"]
        akc = "#00F0FF" if self.is_dark else "#0D9488"
        txt_kol = "#F8FAFC" if self.is_dark else "#0F172A"
        mut = "#94A3B8" if self.is_dark else "#64748B"
        self._odswiez_wiersze_siec(s["per_siec"], akc, txt_kol, mut)
        if s["rekord_dzien"]:
            data, n = s["rekord_dzien"]
            dni_pl = ["poniedziałek","wtorek","środa","czwartek","piątek","sobota","niedziela"]
            self.l_rekord.setText(f"{n} {'wizyta' if n==1 else 'wizyty' if 2<=n<=4 else 'wizyt'} — "
                                  f"{dni_pl[data.weekday()]}, {data.strftime('%d.%m.%Y')}")
        else:
            self.l_rekord.setText("Jeszcze za mało danych.")


class PanelAdminaOverlay(QFrame):
    """Panel administratora — wgląd we WSZYSTKICH użytkowników programu na tym
    komputerze: ilu ich jest, kto (imię+nazwisko), ile dokumentów utworzyli,
    ile wypraw, km, od kiedy korzystają. Tylko lokalnie (dane z ~/.pmt_uzytkownicy.json)."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.is_dark = True
        self.hide()
        self.setStyleSheet("background: rgba(3, 7, 14, 0.90);")
        lay = QVBoxLayout(self); lay.setContentsMargins(36, 28, 36, 28); lay.setSpacing(18)

        # nagłówek + zamknięcie
        gora = QHBoxLayout()
        naglowek = QVBoxLayout(); naglowek.setSpacing(2)
        self.l_tyt = QLabel("Panel administratora")
        self.l_tyt.setStyleSheet("color:#F8FAFC; font-family:'Segoe UI'; font-size:22px; font-weight:800; background:transparent;")
        self.l_pod = QLabel("Wszyscy użytkownicy programu na tym komputerze")
        self.l_pod.setStyleSheet("color:#94A3B8; font-family:'Segoe UI'; font-size:12px; background:transparent;")
        naglowek.addWidget(self.l_tyt); naglowek.addWidget(self.l_pod)
        gora.addLayout(naglowek); gora.addStretch()
        # Wylogowanie — kasuje zapamiętany kod użytkownika, żeby przy następnym
        # uruchomieniu program znów zapytał o kod (np. gdy komputer zmienia właściciela).
        self.btn_wyloguj = QPushButton("⎋  Wyloguj")
        styl_wyloguj(self.btn_wyloguj, True)
        self.btn_wyloguj.clicked.connect(self._wyloguj)
        gora.addWidget(self.btn_wyloguj)
        gora.addSpacing(8)
        self.btn_x = QPushButton("✕"); self.btn_x.setFixedSize(38, 38)
        self.btn_x.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_x.setStyleSheet(
            "QPushButton { color:#F8FAFC; background:rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.15); border-radius:10px; font-size:15px; }"
            "QPushButton:hover { background:rgba(239,68,68,0.85); }")
        self.btn_x.clicked.connect(self.hide)
        gora.addWidget(self.btn_x)
        lay.addLayout(gora)

        # kafelki podsumowania
        self.kafle = QHBoxLayout(); self.kafle.setSpacing(14)
        self._kafle_lbl = {}
        for klucz, tytul in [("uz", "użytkowników"), ("wy", "wypraw"),
                             ("dok", "dokumentów"), ("km", "kilometrów"),
                             ("kw", "suma rozliczeń")]:
            box = QFrame(); box.setObjectName("AdmKafel")
            box.setStyleSheet("#AdmKafel { background:rgba(11,19,32,0.95); border:1px solid rgba(0,240,255,0.22); border-radius:14px; }")
            bl = QVBoxLayout(box); bl.setContentsMargins(18, 14, 18, 14); bl.setSpacing(4)
            l_val = QLabel("0"); l_val.setStyleSheet("color:#00F0FF; font-family:'Segoe UI'; font-size:24px; font-weight:800; background:transparent;")
            l_op = QLabel(tytul); l_op.setStyleSheet("color:#94A3B8; font-family:'Segoe UI'; font-size:11px; background:transparent;")
            bl.addWidget(l_val); bl.addWidget(l_op)
            self._kafle_lbl[klucz] = l_val
            self.kafle.addWidget(box, 1)
        lay.addLayout(self.kafle)

        # tabela użytkowników
        self.tabela = QTableWidget()
        self.tabela.setColumnCount(7)
        self.tabela.setHorizontalHeaderLabels(
            ["Imię i nazwisko", "Stanowisko", "Wyprawy", "Dokumenty", "Kilometry", "Od kiedy", "Ostatnio"])
        self.tabela.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.tabela.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.tabela.verticalHeader().setVisible(False)
        self.tabela.horizontalHeader().setStretchLastSection(True)
        self.tabela.setColumnWidth(0, 200); self.tabela.setColumnWidth(1, 140)
        self.tabela.setColumnWidth(2, 80); self.tabela.setColumnWidth(3, 90)
        self.tabela.setColumnWidth(4, 90); self.tabela.setColumnWidth(5, 130)
        lay.addWidget(self.tabela, 1)

        # --- pole aktywacyjne — bez fanfar, na samym dole panelu ---
        akt = QHBoxLayout(); akt.setSpacing(8)
        akt.addStretch()
        self.pole_kod_dostepu = QLineEdit()
        self.pole_kod_dostepu.setPlaceholderText("Kod dostępu")
        self.pole_kod_dostepu.setFixedWidth(160)
        self.pole_kod_dostepu.setFixedHeight(28)
        self.pole_kod_dostepu.returnPressed.connect(self._zatwierdz_kod_dostepu)
        akt.addWidget(self.pole_kod_dostepu)
        self.btn_kod_dostepu = QPushButton("Zatwierdź")
        self.btn_kod_dostepu.setFixedHeight(28)
        self.btn_kod_dostepu.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_kod_dostepu.clicked.connect(self._zatwierdz_kod_dostepu)
        akt.addWidget(self.btn_kod_dostepu)
        self.lbl_kod_status = QLabel("")
        akt.addWidget(self.lbl_kod_status)
        lay.addLayout(akt)

    def _wyloguj(self):
        """Kasuje zapamiętany kod użytkownika i zamyka program.
        Przy kolejnym uruchomieniu pojawi się okno logowania."""
        self.hide()                    # schowaj panel Ustawień
        okno = self.window()
        if hasattr(okno, "_wyloguj_uzytkownika"):
            okno._wyloguj_uzytkownika()    # jedna wspólna ścieżka wylogowania
            return
        try:
            online_synchronizuj()
        except Exception:
            pass
        online_wyloguj()
        QApplication.quit()

    def _zatwierdz_kod_dostepu(self):
        wpisany = self.pole_kod_dostepu.text().strip()
        self.pole_kod_dostepu.clear()
        if wpisany and wpisany == KOD_AKTYWACYJNY:
            odblokuj_licencje_na_stale()
            self.lbl_kod_status.setText("✓")
            self.lbl_kod_status.setStyleSheet(
                "color:#00E4A1; font-family:'Segoe UI'; font-size:14px; font-weight:800; background:transparent;")
        elif wpisany:
            self.lbl_kod_status.setText("✕")
            self.lbl_kod_status.setStyleSheet(
                "color:#F87171; font-family:'Segoe UI'; font-size:14px; font-weight:800; background:transparent;")

    def update_theme(self, is_dark):
        self.is_dark = is_dark
        if is_dark:
            self.setStyleSheet("background: rgba(3, 7, 14, 0.92);")
            self.l_tyt.setStyleSheet("color:#F8FAFC; font-family:'Segoe UI'; font-size:22px; font-weight:800; background:transparent;")
            akc = "#00F0FF"; naglo = "rgba(0,240,255,0.10)"; txt_kol = "#E2E8F0"; ramka = "rgba(0,240,255,0.18)"
        else:
            self.setStyleSheet("background: rgba(226, 232, 240, 0.96);")
            self.l_tyt.setStyleSheet("color:#0F172A; font-family:'Segoe UI'; font-size:22px; font-weight:800; background:transparent;")
            akc = "#0D9488"; naglo = "rgba(13,148,136,0.12)"; txt_kol = "#1E293B"; ramka = "rgba(13,148,136,0.20)"
        for l in self._kafle_lbl.values():
            l.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:24px; font-weight:800; background:transparent;")
        self.tabela.setStyleSheet(
            f"QTableWidget {{ background:transparent; color:{txt_kol}; gridline-color:{ramka}; border:1px solid {ramka}; border-radius:10px; font-family:'Segoe UI'; font-size:12px; }}"
            f"QHeaderView::section {{ background:{naglo}; color:{akc}; border:none; padding:8px; font-weight:700; }}"
            f"QTableWidget::item {{ padding:6px; }}")
        # celowo dyskretne — bez zwracania uwagi
        mut = "#4B5563" if is_dark else "#94A3B8"
        self.pole_kod_dostepu.setStyleSheet(
            f"QLineEdit {{ background:transparent; border:1px solid {ramka}; border-radius:6px; "
            f"padding:0 8px; color:{mut}; font-family:'Segoe UI'; font-size:10px; }}")
        self.btn_kod_dostepu.setStyleSheet(
            f"QPushButton {{ color:{mut}; background:transparent; border:1px solid {ramka}; border-radius:6px; "
            f"padding:0 10px; font-family:'Segoe UI'; font-size:10px; }} "
            f"QPushButton:hover {{ color:{akc}; border-color:{akc}; }}")

    def odswiez_dane(self):
        dane = statystyki_administratora()
        self._kafle_lbl["uz"].setText(str(dane["liczba_uzytkownikow"]))
        self._kafle_lbl["wy"].setText(str(dane["suma_wypraw"]))
        self._kafle_lbl["dok"].setText(str(dane["suma_dokumentow"]))
        self._kafle_lbl["km"].setText(f"{dane['suma_km']:,}".replace(",", " "))
        self._kafle_lbl["kw"].setText(f"{dane['suma_kwota']:,.0f} zł".replace(",", " "))
        uzytk = dane["uzytkownicy"]
        self.tabela.setRowCount(len(uzytk))
        for r, u in enumerate(uzytk):
            wart = [u["imie"], u["stanowisko"] or "—", str(u["wyprawy"]),
                    str(u["dokumenty"]), f"{u['km']:,}".replace(",", " "),
                    u["pierwsza"] or "—", u["ostatnia"] or "—"]
            for c, w in enumerate(wart):
                it = QTableWidgetItem(str(w))
                if c >= 2 and c <= 4:
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.tabela.setItem(r, c, it)


class DialogUstawieniaPlanowania(QDialog):
    """Parametry, wokół których układa się CAŁY plan: ile godzin dziennie
    pracujesz, ile trwa wizyta, jak szybko jedziesz. Wcześniej sztywne."""
    def __init__(self, parent=None, is_dark=True):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self._wynik = None
        if is_dark:
            karta="#0B1320"; ramka="rgba(0,240,255,0.30)"; akc="#00F0FF"; zielony="#00E4A1"
            txt="#F8FAFC"; mut="#94A3B8"; pole="rgba(5,10,20,0.6)"
        else:
            karta="#FFFFFF"; ramka="rgba(13,148,136,0.35)"; akc="#0D9488"; zielony="#059669"
            txt="#0F172A"; mut="#64748B"; pole="rgba(241,245,249,0.9)"
        self._akc = akc; self._mut = mut

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        box = QFrame(); box.setObjectName("UstBox")
        box.setStyleSheet(f"#UstBox {{ background:{karta}; border:1px solid {ramka}; border-radius:16px; }}")
        box.setFixedWidth(520)
        root.addWidget(box)
        bl = QVBoxLayout(box); bl.setContentsMargins(26, 24, 26, 24); bl.setSpacing(6)

        t = QLabel("Ustawienia planowania")
        t.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:17px; font-weight:800; background:transparent;")
        bl.addWidget(t)
        p = QLabel("Wokół tych wartości układa się cały plan — ile wizyt zmieści się "
                   "w dniu, ile zajmie i ile wyjedziesz kilometrów.")
        p.setWordWrap(True)
        p.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
        bl.addWidget(p)
        bl.addSpacing(10)

        UP = ustawienia_planowania()
        suw_css = (
            f"QSlider::groove:horizontal {{ height:5px; border-radius:3px; background:rgba(148,163,184,0.28); }} "
            f"QSlider::sub-page:horizontal {{ height:5px; border-radius:3px; "
            f"background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc}, stop:1 {zielony}); }} "
            f"QSlider::handle:horizontal {{ width:16px; height:16px; margin:-6px 0; border-radius:8px; "
            f"background:{akc}; border:2px solid {karta}; }}")

        def suwak(tytul, opis, mini, maks, wartosc, formatuj):
            wrap = QVBoxLayout(); wrap.setSpacing(2)
            gora = QHBoxLayout()
            lt = QLabel(tytul)
            lt.setStyleSheet(f"color:{txt}; font-family:'Segoe UI'; font-size:12px; font-weight:700; background:transparent;")
            gora.addWidget(lt); gora.addStretch()
            lw = QLabel(formatuj(wartosc))
            lw.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:13px; font-weight:800; background:transparent;")
            gora.addWidget(lw)
            wrap.addLayout(gora)
            s = QSlider(Qt.Orientation.Horizontal)
            s.setMinimum(mini); s.setMaximum(maks); s.setValue(int(wartosc))
            s.setStyleSheet(suw_css)
            s.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            s.valueChanged.connect(lambda v, l=lw, f=formatuj: l.setText(f(v)))
            wrap.addWidget(s)
            lo = QLabel(opis)
            lo.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:10px; background:transparent;")
            wrap.addWidget(lo)
            bl.addLayout(wrap)
            bl.addSpacing(12)
            return s

        self.s_godz = suwak(
            "Godziny pracy dziennie", "Ile maksymalnie godzin trwa dzień w trasie (dojazdy + wizyty).",
            1, 16, round(UP["limit_minut"] / 60), lambda v: f"{v} h")
        self.s_wiz = suwak(
            "Czas jednej wizyty", "Ile średnio spędzasz w sklepie.",
            5, 120, round(UP["czas_wizyty"]), lambda v: f"{v} min")
        self.s_pred = suwak(
            "Średnia prędkość", "Do szacowania czasu przejazdu między punktami.",
            20, 120, round(UP["predkosc"]), lambda v: f"{v} km/h")

        self.chk_powrot = QCheckBox("Wliczaj powrót do bazy na koniec dnia")
        self.chk_powrot.setChecked(bool(UP["powrot"]))
        self.chk_powrot.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.chk_powrot.setStyleSheet(
            f"QCheckBox {{ color:{txt}; font-family:'Segoe UI'; font-size:12px; font-weight:600; background:transparent; spacing:8px; }} "
            f"QCheckBox::indicator {{ width:16px; height:16px; border:1px solid {ramka}; border-radius:5px; background:{pole}; }} "
            f"QCheckBox::indicator:checked {{ background:{akc}; border-color:{akc}; }}")
        bl.addWidget(self.chk_powrot)
        lp = QLabel("Doliczy kilometry i czas powrotu z ostatniego punktu do domu.")
        lp.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:10px; background:transparent;")
        bl.addWidget(lp)
        bl.addSpacing(14)

        akcje = QHBoxLayout(); akcje.setSpacing(10)
        b_dom = QPushButton("Przywróć domyślne"); b_dom.setFixedHeight(36)
        b_dom.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_dom.setStyleSheet(f"QPushButton {{ color:{mut}; background:transparent; border:1px solid {ramka}; "
                            f"border-radius:9px; padding:0 12px; font-family:'Segoe UI'; font-size:11px; }}")
        b_dom.clicked.connect(self._domyslne)
        akcje.addWidget(b_dom)
        akcje.addStretch()
        b_anu = QPushButton("Anuluj"); b_anu.setFixedHeight(36)
        b_anu.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_anu.setStyleSheet(f"QPushButton {{ color:{txt}; background:transparent; border:1px solid {ramka}; "
                            f"border-radius:9px; padding:0 16px; font-family:'Segoe UI'; font-size:12px; }}")
        b_anu.clicked.connect(self.reject)
        akcje.addWidget(b_anu)
        b_ok = QPushButton("Zapisz"); b_ok.setFixedHeight(36)
        b_ok.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_ok.setStyleSheet(f"QPushButton {{ color:#04121A; background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc}, stop:1 {zielony}); "
                           f"border:none; border-radius:9px; padding:0 22px; font-family:'Segoe UI'; font-size:12px; font-weight:800; }}")
        b_ok.clicked.connect(self._zapisz)
        akcje.addWidget(b_ok)
        bl.addLayout(akcje)

    def _domyslne(self):
        self.s_godz.setValue(8)
        self.s_wiz.setValue(20)
        self.s_pred.setValue(50)
        self.chk_powrot.setChecked(False)

    def _zapisz(self):
        zapisz_ustawienie("plan_godziny", int(self.s_godz.value()))
        zapisz_ustawienie("plan_minuty_wizyta", int(self.s_wiz.value()))
        zapisz_ustawienie("plan_predkosc", int(self.s_pred.value()))
        zapisz_ustawienie("plan_powrot_do_bazy", bool(self.chk_powrot.isChecked()))
        self._wynik = True
        self.accept()

    def pokaz(self):
        if self.exec() == QDialog.DialogCode.Accepted:
            return self._wynik
        return None


class DialogCykle(QDialog):
    """Ustawienie cyklu wizyt: osobno dla DUŻYCH MIAST i poza nimi.
    Np. Biedronki w miastach 4×/mies., poza miastami 1×/mies."""
    # (tygodnie, etykieta) — etykiety w przeliczeniu na miesiąc, bo tak
    # myśli użytkownik ("4 razy w miesiącu"), a silnik liczy w tygodniach
    OPCJE = [(1, "4× w mies. (co tydzień)"),
             (2, "2× w mies. (co 2 tyg.)"),
             (3, "co 3 tyg."),
             (4, "1× w mies. (co 4 tyg.)"),
             (6, "co 6 tyg."),
             (8, "co 2 mies."),
             (12, "co 3 mies.")]

    def __init__(self, parent=None, sieci=None, cykle=None, cykle_duze=None,
                 cykl_domyslny=4, is_dark=True):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self._wynik = None
        sieci = sieci or []
        cykle = cykle or {}
        cykle_duze = cykle_duze or {}
        if is_dark:
            karta="#0B1320"; ramka="rgba(0,240,255,0.30)"; akc="#00F0FF"; zielony="#00E4A1"
            txt="#F8FAFC"; mut="#94A3B8"; pole="rgba(5,10,20,0.6)"
        else:
            karta="#FFFFFF"; ramka="rgba(13,148,136,0.35)"; akc="#0D9488"; zielony="#059669"
            txt="#0F172A"; mut="#64748B"; pole="rgba(241,245,249,0.9)"

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        box = QFrame(); box.setObjectName("CykBox")
        box.setStyleSheet(f"#CykBox {{ background:{karta}; border:1px solid {ramka}; border-radius:16px; }}")
        box.setFixedWidth(680)
        root.addWidget(box)
        bl = QVBoxLayout(box); bl.setContentsMargins(24, 22, 24, 22); bl.setSpacing(12)

        t = QLabel("Cykl wizyt")
        t.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:17px; font-weight:800; background:transparent;")
        bl.addWidget(t)
        p = QLabel("Jak często wracać do każdej sieci — osobno w dużych miastach i poza nimi. "
                   "Np. Biedronki w miastach 4× w miesiącu, poza miastami 1×.")
        p.setWordWrap(True)
        p.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
        bl.addWidget(p)

        combo_css = (f"QComboBox {{ background:{pole}; border:1px solid {ramka}; border-radius:8px; "
                     f"padding:4px 10px; color:{txt}; font-family:'Segoe UI'; font-size:11px; }} "
                     f"QComboBox QAbstractItemView {{ background:{karta}; color:{txt}; "
                     f"selection-background-color:{akc}; selection-color:#04121A; }}")

        def nowy_combo(z_domyslnym=True):
            cb = QComboBox(); cb.setStyleSheet(combo_css); cb.setFixedWidth(190)
            if z_domyslnym:
                cb.addItem("jak domyślnie", None)
            for tyg, etykieta in self.OPCJE:
                cb.addItem(etykieta, tyg)
            return cb

        # nagłówki kolumn
        naglo = QHBoxLayout()
        h0 = QLabel("Sieć"); h0.setFixedWidth(130)
        h1 = QLabel("🏙  Duże miasta"); h1.setFixedWidth(190)
        h2 = QLabel("🌾  Poza dużymi miastami"); h2.setFixedWidth(190)
        for h in (h0, h1, h2):
            h.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:10px; font-weight:700; background:transparent;")
        naglo.addWidget(h0); naglo.addStretch(); naglo.addWidget(h1); naglo.addWidget(h2)
        bl.addLayout(naglo)

        # lista sieci — dwa combo na sieć
        self.combos = {}        # sieć -> (combo_duze, combo_poza)
        if sieci:
            scroll = QScrollArea(); scroll.setWidgetResizable(True)
            scroll.setMaximumHeight(260)
            scroll.setStyleSheet("QScrollArea { background:transparent; border:none; }")
            wrap = QWidget(); wl = QVBoxLayout(wrap)
            wl.setContentsMargins(0, 0, 8, 0); wl.setSpacing(8)
            wrap.setStyleSheet("background:transparent;")
            for s in sieci:
                r = QHBoxLayout()
                ls = QLabel(s); ls.setFixedWidth(130)
                ls.setStyleSheet(f"color:{zielony}; font-family:'Segoe UI'; font-size:12px; font-weight:700; background:transparent;")
                r.addWidget(ls); r.addStretch()
                cb_d = nowy_combo()
                cb_p = nowy_combo()
                wd = cykle_duze.get(s); wp = cykle.get(s)
                tygodnie = [o[0] for o in self.OPCJE]
                if wd in tygodnie:
                    cb_d.setCurrentIndex(1 + tygodnie.index(wd))
                if wp in tygodnie:
                    cb_p.setCurrentIndex(1 + tygodnie.index(wp))
                r.addWidget(cb_d); r.addWidget(cb_p)
                self.combos[s] = (cb_d, cb_p)
                wl.addLayout(r)
            wl.addStretch()
            scroll.setWidget(wrap)
            bl.addWidget(scroll)

        sep = QFrame(); sep.setFixedHeight(1)
        sep.setStyleSheet(f"background:{ramka}; border:none;")
        bl.addWidget(sep)

        # cykl domyślny
        dom = QHBoxLayout()
        ld = QLabel("Domyślnie (sieci bez własnej reguły):")
        ld.setStyleSheet(f"color:{txt}; font-family:'Segoe UI'; font-size:12px; font-weight:600; background:transparent;")
        dom.addWidget(ld); dom.addStretch()
        self.combo_dom = nowy_combo(z_domyslnym=False)
        tygodnie = [o[0] for o in self.OPCJE]
        self.combo_dom.setCurrentIndex(tygodnie.index(cykl_domyslny) if cykl_domyslny in tygodnie else 3)
        dom.addWidget(self.combo_dom)
        bl.addLayout(dom)

        akcje = QHBoxLayout(); akcje.setSpacing(10)
        akcje.addStretch()
        b_anu = QPushButton("Anuluj"); b_anu.setFixedHeight(38)
        b_anu.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_anu.setStyleSheet(f"QPushButton {{ color:{txt}; background:transparent; border:1px solid {ramka}; "
                            f"border-radius:10px; padding:0 18px; font-family:'Segoe UI'; font-size:12px; }}")
        b_anu.clicked.connect(self.reject)
        akcje.addWidget(b_anu)
        b_ok = QPushButton("Zastosuj"); b_ok.setFixedHeight(38)
        b_ok.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_ok.setStyleSheet(f"QPushButton {{ color:#04121A; background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc}, stop:1 {zielony}); "
                           f"border:none; border-radius:10px; padding:0 24px; font-family:'Segoe UI'; font-size:12px; font-weight:800; }}")
        b_ok.clicked.connect(self._zapisz)
        akcje.addWidget(b_ok)
        bl.addLayout(akcje)

    def _zapisz(self):
        cykle = {}; cykle_duze = {}
        for s, (cb_d, cb_p) in self.combos.items():
            vd = cb_d.currentData()
            vp = cb_p.currentData()
            if vd:
                cykle_duze[s] = int(vd)
            if vp:
                cykle[s] = int(vp)
        self._wynik = (int(self.combo_dom.currentData()), cykle, cykle_duze)
        self.accept()

    def pokaz(self):
        """Zwraca (cykl_domyslny, {sieć: tyg. poza}, {sieć: tyg. w dużych}) albo None."""
        if self.exec() == QDialog.DialogCode.Accepted:
            return self._wynik
        return None


class DialogCyklPunktow(QDialog):
    """Nadanie cyklu grupie punktów (aktualnie przefiltrowanych)."""
    def __init__(self, parent=None, ile=0, is_dark=True):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self._wynik = None
        if is_dark:
            karta="#0B1320"; ramka="rgba(0,240,255,0.30)"; akc="#00F0FF"; zielony="#00E4A1"
            txt="#F8FAFC"; mut="#94A3B8"; pole="rgba(5,10,20,0.6)"
        else:
            karta="#FFFFFF"; ramka="rgba(13,148,136,0.35)"; akc="#0D9488"; zielony="#059669"
            txt="#0F172A"; mut="#64748B"; pole="rgba(241,245,249,0.9)"

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        box = QFrame(); box.setObjectName("CykPBox")
        box.setStyleSheet(f"#CykPBox {{ background:{karta}; border:1px solid {ramka}; border-radius:16px; }}")
        box.setFixedWidth(460)
        root.addWidget(box)
        bl = QVBoxLayout(box); bl.setContentsMargins(24, 22, 24, 22); bl.setSpacing(12)

        t = QLabel(f"Cykl dla {ile} widocznych punktów")
        t.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:16px; font-weight:800; background:transparent;")
        bl.addWidget(t)
        p = QLabel("Ten cykl nadpisze regułę sieci dla tych konkretnych punktów. "
                   "Przydatne, gdy np. wybrane sklepy wymagają częstszych wizyt.")
        p.setWordWrap(True)
        p.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
        bl.addWidget(p)

        self.combo = QComboBox(); self.combo.setFixedHeight(36)
        self.combo.setStyleSheet(
            f"QComboBox {{ background:{pole}; border:1px solid {ramka}; border-radius:8px; "
            f"padding:4px 12px; color:{txt}; font-family:'Segoe UI'; font-size:12px; }} "
            f"QComboBox QAbstractItemView {{ background:{karta}; color:{txt}; "
            f"selection-background-color:{akc}; selection-color:#04121A; }}")
        for tyg, etykieta in DialogCykle.OPCJE:
            self.combo.addItem(etykieta, tyg)
        self.combo.setCurrentIndex(0)
        bl.addWidget(self.combo)

        akcje = QHBoxLayout(); akcje.setSpacing(10)
        b_usun = QPushButton("Usuń indywidualny cykl"); b_usun.setFixedHeight(36)
        b_usun.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_usun.setStyleSheet(
            "QPushButton { color:#F87171; background:transparent; border:1px solid rgba(248,113,113,0.5); "
            "border-radius:9px; padding:0 12px; font-family:'Segoe UI'; font-size:11px; }")
        b_usun.clicked.connect(self._usun)
        akcje.addWidget(b_usun)
        akcje.addStretch()
        b_anu = QPushButton("Anuluj"); b_anu.setFixedHeight(36)
        b_anu.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_anu.setStyleSheet(f"QPushButton {{ color:{txt}; background:transparent; border:1px solid {ramka}; "
                            f"border-radius:9px; padding:0 16px; font-family:'Segoe UI'; font-size:12px; }}")
        b_anu.clicked.connect(self.reject)
        akcje.addWidget(b_anu)
        b_ok = QPushButton("Zastosuj"); b_ok.setFixedHeight(36)
        b_ok.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_ok.setStyleSheet(f"QPushButton {{ color:#04121A; background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc}, stop:1 {zielony}); "
                           f"border:none; border-radius:9px; padding:0 22px; font-family:'Segoe UI'; font-size:12px; font-weight:800; }}")
        b_ok.clicked.connect(self._zapisz)
        akcje.addWidget(b_ok)
        bl.addLayout(akcje)

    def _zapisz(self):
        self._wynik = int(self.combo.currentData())
        self.accept()

    def _usun(self):
        self._wynik = "__USUN__"
        self.accept()

    def pokaz(self):
        if self.exec() == QDialog.DialogCode.Accepted:
            return self._wynik
        return None


class DialogRoznicaImportu(QDialog):
    """Pokazuje rozpoznane zmiany między plikiem a obecną listą punktów —
    nowe (zostaną dodane) i zniknione (możesz je usunąć). Zamiast po cichu
    dokładać do listy przy każdym ponownym imporcie z centrali."""
    def __init__(self, parent=None, dodane=None, zniknely=None, is_dark=True):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self._wynik = None
        self._zniknely = zniknely or []
        dodane = dodane or []
        if is_dark:
            karta="#0B1320"; ramka="rgba(0,240,255,0.30)"; akc="#00F0FF"; zielony="#00E4A1"
            txt="#F8FAFC"; mut="#94A3B8"; pole="rgba(5,10,20,0.6)"; czerwony="#F87171"
        else:
            karta="#FFFFFF"; ramka="rgba(13,148,136,0.35)"; akc="#0D9488"; zielony="#059669"
            txt="#0F172A"; mut="#64748B"; pole="rgba(241,245,249,0.9)"; czerwony="#DC2626"

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        box = QFrame(); box.setObjectName("RozBox")
        box.setStyleSheet(f"#RozBox {{ background:{karta}; border:1px solid {ramka}; border-radius:16px; }}")
        box.setFixedWidth(540)
        root.addWidget(box)
        bl = QVBoxLayout(box); bl.setContentsMargins(26, 22, 26, 22); bl.setSpacing(8)

        t = QLabel("🔄  Rozpoznane zmiany")
        t.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:16px; font-weight:800; background:transparent;")
        bl.addWidget(t)
        p = QLabel("Porównałem plik z obecną listą punktów.")
        p.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
        bl.addWidget(p)
        bl.addSpacing(6)

        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(340)
        scroll.setStyleSheet("QScrollArea { background:transparent; border:none; }")
        wrap = QWidget(); wl = QVBoxLayout(wrap)
        wl.setContentsMargins(0, 0, 8, 0); wl.setSpacing(6)
        wrap.setStyleSheet("background:transparent;")

        def etyk(p):
            opis = p.get("adres", "") or "(bez adresu)"
            if p.get("miasto") and p["miasto"].lower() not in opis.lower():
                opis += f", {p['miasto']}"
            return f"[{p['siec']}]  {opis}" if p.get("siec") else opis

        if dodane:
            nagl = QLabel(f"✓  Nowe — zostaną dodane ({len(dodane)})")
            nagl.setStyleSheet(f"color:{zielony}; font-family:'Segoe UI'; font-size:12px; font-weight:800; background:transparent;")
            wl.addWidget(nagl)
            for p in dodane[:100]:
                l = QLabel("  " + etyk(p))
                l.setStyleSheet(f"color:{txt}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
                wl.addWidget(l)
            if len(dodane) > 100:
                wiecej = QLabel(f"  … i {len(dodane) - 100} więcej")
                wiecej.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:10px; background:transparent;")
                wl.addWidget(wiecej)
            wl.addSpacing(10)

        self._checki = []
        if self._zniknely:
            nagl2 = QLabel(f"✕  Zniknęły z pliku — zaznaczone zostaną usunięte ({len(self._zniknely)})")
            nagl2.setStyleSheet(f"color:{czerwony}; font-family:'Segoe UI'; font-size:12px; font-weight:800; background:transparent;")
            wl.addWidget(nagl2)
            for p in self._zniknely:
                chk = QCheckBox(etyk(p))
                chk.setChecked(True)     # domyślnie: zniknięte = do usunięcia
                chk.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
                chk.setStyleSheet(
                    f"QCheckBox {{ color:{txt}; font-family:'Segoe UI'; font-size:11px; background:transparent; spacing:8px; }} "
                    f"QCheckBox::indicator {{ width:15px; height:15px; border:1px solid {ramka}; border-radius:4px; background:{pole}; }} "
                    f"QCheckBox::indicator:checked {{ background:{czerwony}; border-color:{czerwony}; }}")
                wl.addWidget(chk)
                self._checki.append((chk, p))
        wl.addStretch()
        scroll.setWidget(wrap)
        bl.addWidget(scroll)
        bl.addSpacing(8)

        akcje = QHBoxLayout(); akcje.setSpacing(10)
        b_anu = QPushButton("Anuluj"); b_anu.setFixedHeight(38)
        b_anu.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_anu.setStyleSheet(f"QPushButton {{ color:{txt}; background:transparent; border:1px solid {ramka}; "
                            f"border-radius:10px; padding:0 18px; font-family:'Segoe UI'; font-size:12px; }}")
        b_anu.clicked.connect(self.reject)
        akcje.addWidget(b_anu)
        akcje.addStretch()
        b_ok = QPushButton(f"Zastosuj ({len(dodane)} nowych)"); b_ok.setFixedHeight(38)
        b_ok.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_ok.setStyleSheet(f"QPushButton {{ color:#04121A; background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc}, stop:1 {zielony}); "
                           f"border:none; border-radius:10px; padding:0 22px; font-family:'Segoe UI'; font-size:12px; font-weight:800; }}")
        b_ok.clicked.connect(self._zapisz)
        akcje.addWidget(b_ok)
        bl.addLayout(akcje)

    def _zapisz(self):
        self._wynik = [p for chk, p in self._checki if chk.isChecked()]
        self.accept()

    def pokaz(self):
        """Zwraca listę punktów DO USUNIĘCIA (zaznaczonych) albo None gdy anulowano."""
        if self.exec() == QDialog.DialogCode.Accepted:
            return self._wynik
        return None


class DialogBezLokalizacji(QDialog):
    """Pokazuje KTÓRE konkretnie punkty nie mają lokalizacji (nie tylko ich
    liczbę) — z opcją usunięcia wybranych z listy punktów. Adresy, których
    program nie potrafi zgeokodować, są martwym balastem w planowaniu."""
    def __init__(self, parent=None, punkty=None, is_dark=True):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self._wynik = None
        self._punkty = punkty or []
        if is_dark:
            karta="#0B1320"; ramka="rgba(251,191,36,0.35)"; akc="#FBBF24"
            txt="#F8FAFC"; mut="#94A3B8"; pole="rgba(5,10,20,0.6)"
        else:
            karta="#FFFFFF"; ramka="rgba(180,83,9,0.35)"; akc="#B45309"
            txt="#0F172A"; mut="#64748B"; pole="rgba(241,245,249,0.9)"

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        box = QFrame(); box.setObjectName("BezLokBox")
        box.setStyleSheet(f"#BezLokBox {{ background:{karta}; border:1px solid {ramka}; border-radius:16px; }}")
        box.setFixedWidth(520)
        root.addWidget(box)
        bl = QVBoxLayout(box); bl.setContentsMargins(26, 22, 26, 22); bl.setSpacing(8)

        t = QLabel(f"⚠️  {len(self._punkty)} punktów bez lokalizacji")
        t.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:16px; font-weight:800; background:transparent;")
        bl.addWidget(t)
        p = QLabel("Programowi nie udało się ustalić współrzędnych tych adresów — "
                   "sprawdź literówki albo doprecyzuj miejscowość. Możesz je też "
                   "usunąć z listy, jeśli już nie są aktualne.")
        p.setWordWrap(True)
        p.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
        bl.addWidget(p)
        bl.addSpacing(6)

        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(280)
        scroll.setStyleSheet("QScrollArea { background:transparent; border:none; }")
        wrap = QWidget(); wl = QVBoxLayout(wrap)
        wl.setContentsMargins(0, 0, 8, 0); wl.setSpacing(6)
        wrap.setStyleSheet("background:transparent;")

        self._checki = []
        for pt in self._punkty:
            opis = pt.adres or pt.nazwa or "(bez adresu)"
            if pt.miasto and pt.miasto.lower() not in opis.lower():
                opis += f", {pt.miasto}"
            etykieta = f"[{pt.siec}]  {opis}" if pt.siec else opis
            chk = QCheckBox(etykieta)
            chk.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            chk.setStyleSheet(
                f"QCheckBox {{ color:{txt}; font-family:'Segoe UI'; font-size:11px; background:transparent; spacing:8px; }} "
                f"QCheckBox::indicator {{ width:15px; height:15px; border:1px solid {ramka}; border-radius:4px; background:{pole}; }} "
                f"QCheckBox::indicator:checked {{ background:{akc}; border-color:{akc}; }}")
            wl.addWidget(chk)
            self._checki.append((chk, pt))
        wl.addStretch()
        scroll.setWidget(wrap)
        bl.addWidget(scroll)
        bl.addSpacing(8)

        akcje = QHBoxLayout(); akcje.setSpacing(10)
        b_zamknij = QPushButton("Zamknij"); b_zamknij.setFixedHeight(38)
        b_zamknij.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_zamknij.setStyleSheet(f"QPushButton {{ color:{txt}; background:transparent; border:1px solid {ramka}; "
                                f"border-radius:10px; padding:0 18px; font-family:'Segoe UI'; font-size:12px; }}")
        b_zamknij.clicked.connect(self.reject)
        akcje.addWidget(b_zamknij)
        akcje.addStretch()
        b_usun = QPushButton("Usuń zaznaczone z listy punktów"); b_usun.setFixedHeight(38)
        b_usun.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_usun.setStyleSheet(
            f"QPushButton {{ color:#04121A; background:{akc}; border:none; border-radius:10px; "
            f"padding:0 18px; font-family:'Segoe UI'; font-size:12px; font-weight:800; }}")
        b_usun.clicked.connect(self._usun)
        akcje.addWidget(b_usun)
        bl.addLayout(akcje)

    def _usun(self):
        self._wynik = [pt for chk, pt in self._checki if chk.isChecked()]
        if not self._wynik:
            self.reject()
            return
        self.accept()

    def pokaz(self):
        """Zwraca listę usuniętych punktów (PunktWizyty) albo None."""
        if self.exec() == QDialog.DialogCode.Accepted:
            return self._wynik
        return None


class DialogSugestieCyklu(QDialog):
    """Sugestie cyklu wykryte z HISTORII wizyt — program zauważył Twój
    naturalny rytm i proponuje dopasować do niego ustawienia. Każda sugestia
    ma checkbox; zaznaczone zostaną zapisane jako cykl indywidualny punktu."""
    def __init__(self, parent=None, sugestie=None, is_dark=True):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self._wynik = None
        self._sugestie = sugestie or []
        if is_dark:
            karta="#0B1320"; ramka="rgba(0,240,255,0.30)"; akc="#00F0FF"; zielony="#00E4A1"
            txt="#F8FAFC"; mut="#94A3B8"; pole="rgba(5,10,20,0.6)"
        else:
            karta="#FFFFFF"; ramka="rgba(13,148,136,0.35)"; akc="#0D9488"; zielony="#059669"
            txt="#0F172A"; mut="#64748B"; pole="rgba(241,245,249,0.9)"

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        box = QFrame(); box.setObjectName("SugBox")
        box.setStyleSheet(f"#SugBox {{ background:{karta}; border:1px solid {ramka}; border-radius:16px; }}")
        box.setFixedWidth(560)
        root.addWidget(box)
        bl = QVBoxLayout(box); bl.setContentsMargins(26, 22, 26, 22); bl.setSpacing(6)

        t = QLabel("💡  Sugestie cyklu z historii")
        t.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:17px; font-weight:800; background:transparent;")
        bl.addWidget(t)
        p = QLabel("Na podstawie faktycznych dat odhaczonych wizyt zauważyłem naturalny "
                   "rytm, który różni się od obecnych ustawień. Zaznacz, co chcesz zastosować.")
        p.setWordWrap(True)
        p.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
        bl.addWidget(p)
        bl.addSpacing(8)

        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        scroll.setMaximumHeight(320)
        scroll.setStyleSheet("QScrollArea { background:transparent; border:none; }")
        wrap = QWidget(); wl = QVBoxLayout(wrap)
        wl.setContentsMargins(0, 0, 8, 0); wl.setSpacing(10)
        wrap.setStyleSheet("background:transparent;")

        def etyk_cykl(tyg):
            na_mies = 4 / tyg
            return f"{na_mies:.0f}×/mies." if na_mies >= 1 else f"co {tyg} tyg."

        self._checki = []
        for s in self._sugestie:
            wiersz = QFrame(); wiersz.setObjectName("SugWiersz")
            wiersz.setStyleSheet(
                f"#SugWiersz {{ background:{pole}; border:1px solid {ramka}; border-radius:10px; }}")
            wl_row = QVBoxLayout(wiersz); wl_row.setContentsMargins(12, 10, 12, 10); wl_row.setSpacing(3)
            gora_row = QHBoxLayout()
            chk = QCheckBox(s["nazwa"])
            chk.setChecked(True)
            chk.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            chk.setStyleSheet(
                f"QCheckBox {{ color:{txt}; font-family:'Segoe UI'; font-size:13px; font-weight:700; "
                f"background:transparent; spacing:8px; }} "
                f"QCheckBox::indicator {{ width:16px; height:16px; border:1px solid {ramka}; border-radius:5px; "
                f"background:{pole}; }} QCheckBox::indicator:checked {{ background:{akc}; border-color:{akc}; }}")
            gora_row.addWidget(chk); gora_row.addStretch()
            wl_row.addLayout(gora_row)

            zmiana = QLabel(f"{etyk_cykl(s['obecny_cykl_tyg'])}  →  {etyk_cykl(s['sugerowany_cykl_tyg'])}")
            zmiana.setStyleSheet(f"color:{zielony}; font-family:'Segoe UI'; font-size:12px; font-weight:800; background:transparent;")
            wl_row.addWidget(zmiana)

            szczegoly = QLabel(f"na podstawie {s['liczba_wizyt']} wizyt — średnio co {s['sredni_odstep_dni']:.0f} dni")
            szczegoly.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:10px; background:transparent;")
            wl_row.addWidget(szczegoly)

            wl.addWidget(wiersz)
            self._checki.append((chk, s))
        wl.addStretch()
        scroll.setWidget(wrap)
        bl.addWidget(scroll)
        bl.addSpacing(10)

        pasek_zaznacz = QHBoxLayout()
        b_wszystkie = QPushButton("Zaznacz wszystkie")
        b_zadne = QPushButton("Odznacz wszystkie")
        for b in (b_wszystkie, b_zadne):
            b.setFixedHeight(28); b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            b.setStyleSheet(
                f"QPushButton {{ color:{mut}; background:transparent; border:1px solid {ramka}; "
                f"border-radius:8px; padding:0 10px; font-family:'Segoe UI'; font-size:10px; }} "
                f"QPushButton:hover {{ color:{akc}; border-color:{akc}; }}")
        b_wszystkie.clicked.connect(lambda: [c.setChecked(True) for c, _ in self._checki])
        b_zadne.clicked.connect(lambda: [c.setChecked(False) for c, _ in self._checki])
        pasek_zaznacz.addWidget(b_wszystkie); pasek_zaznacz.addWidget(b_zadne); pasek_zaznacz.addStretch()
        bl.addLayout(pasek_zaznacz)

        akcje = QHBoxLayout(); akcje.setSpacing(10)
        akcje.addStretch()
        b_anu = QPushButton("Anuluj"); b_anu.setFixedHeight(38)
        b_anu.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_anu.setStyleSheet(f"QPushButton {{ color:{txt}; background:transparent; border:1px solid {ramka}; "
                            f"border-radius:10px; padding:0 18px; font-family:'Segoe UI'; font-size:12px; }}")
        b_anu.clicked.connect(self.reject)
        akcje.addWidget(b_anu)
        b_ok = QPushButton("Zastosuj zaznaczone"); b_ok.setFixedHeight(38)
        b_ok.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_ok.setStyleSheet(f"QPushButton {{ color:#04121A; background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc}, stop:1 {zielony}); "
                           f"border:none; border-radius:10px; padding:0 22px; font-family:'Segoe UI'; font-size:12px; font-weight:800; }}")
        b_ok.clicked.connect(self._zapisz)
        akcje.addWidget(b_ok)
        bl.addLayout(akcje)

    def _zapisz(self):
        self._wynik = [s for chk, s in self._checki if chk.isChecked()]
        self.accept()

    def pokaz(self):
        """Zwraca listę zaznaczonych sugestii (do zastosowania) albo None."""
        if self.exec() == QDialog.DialogCode.Accepted:
            return self._wynik
        return None


class DialogKopiaZapasowa(QDialog):
    """Kopia zapasowa jednym kliknięciem — pakuje punkty, plan, dziennik
    i ustawienia do jednego pliku .zip. Chroni przed utratą danych przy
    awarii dysku albo przy przenoszeniu programu na inny komputer."""
    def __init__(self, parent=None, is_dark=True):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.is_dark = is_dark
        if is_dark:
            karta="#0B1320"; ramka="rgba(0,240,255,0.30)"; akc="#00F0FF"; zielony="#00E4A1"
            txt="#F8FAFC"; mut="#94A3B8"; pole="rgba(5,10,20,0.6)"
        else:
            karta="#FFFFFF"; ramka="rgba(13,148,136,0.35)"; akc="#0D9488"; zielony="#059669"
            txt="#0F172A"; mut="#64748B"; pole="rgba(241,245,249,0.9)"
        self._akc, self._zie, self._txt, self._mut, self._pole, self._ramka = akc, zielony, txt, mut, pole, ramka

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        box = QFrame(); box.setObjectName("KopBox")
        box.setStyleSheet(f"#KopBox {{ background:{karta}; border:1px solid {ramka}; border-radius:16px; }}")
        box.setFixedWidth(480)
        root.addWidget(box)
        bl = QVBoxLayout(box); bl.setContentsMargins(26, 24, 26, 24); bl.setSpacing(10)

        t = QLabel("🛡️  Kopia zapasowa")
        t.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:17px; font-weight:800; background:transparent;")
        bl.addWidget(t)
        p = QLabel("Wszystko, co masz w programie — punkty, plan, dziennik wizyt, cykle "
                   "i ustawienia — żyje wyłącznie na tym komputerze. Ta kopia to Twoje "
                   "zabezpieczenie na wypadek awarii albo zmiany komputera.")
        p.setWordWrap(True)
        p.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
        bl.addWidget(p)
        bl.addSpacing(8)

        self.btn_eksport = QPushButton("⬇   Eksportuj kopię zapasową")
        self.btn_eksport.setFixedHeight(50)
        self.btn_eksport.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_eksport.setStyleSheet(
            f"QPushButton {{ color:#04121A; background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc}, stop:1 {zielony}); "
            f"border:none; border-radius:13px; font-family:'Segoe UI'; font-size:14px; font-weight:800; }} "
            f"QPushButton:hover {{ background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {zielony}, stop:1 {akc}); }}")
        self.btn_eksport.clicked.connect(self._eksportuj)
        bl.addWidget(self.btn_eksport)

        self.btn_przywroc = QPushButton("⬆   Przywróć z kopii")
        self.btn_przywroc.setFixedHeight(46)
        self.btn_przywroc.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_przywroc.setStyleSheet(
            f"QPushButton {{ color:{txt}; background:transparent; border:1.5px solid {ramka}; "
            f"border-radius:13px; font-family:'Segoe UI'; font-size:13px; font-weight:700; }} "
            f"QPushButton:hover {{ border-color:{akc}; color:{akc}; }}")
        self.btn_przywroc.clicked.connect(self._przywroc)
        bl.addWidget(self.btn_przywroc)

        self.lbl_status = QLabel("")
        self.lbl_status.setWordWrap(True)
        self.lbl_status.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
        bl.addWidget(self.lbl_status)
        bl.addSpacing(6)

        akcje = QHBoxLayout(); akcje.addStretch()
        b_zamknij = QPushButton("Zamknij"); b_zamknij.setFixedHeight(36)
        b_zamknij.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_zamknij.setStyleSheet(f"QPushButton {{ color:{mut}; background:transparent; border:1px solid {ramka}; "
                                f"border-radius:9px; padding:0 18px; font-family:'Segoe UI'; font-size:12px; }}")
        b_zamknij.clicked.connect(self.accept)
        akcje.addWidget(b_zamknij)
        bl.addLayout(akcje)

    def _eksportuj(self):
        domyslna = f"PMT_Kopia_{datetime.date.today().strftime('%Y-%m-%d')}.zip"
        sciezka, _ = QFileDialog.getSaveFileName(self, "Zapisz kopię zapasową", domyslna, "Archiwum ZIP (*.zip)")
        if not sciezka:
            return
        if not sciezka.lower().endswith(".zip"):
            sciezka += ".zip"
        try:
            n = eksportuj_kopie_zapasowa(sciezka)
            if n == 0:
                self.lbl_status.setText("⚠ Brak danych do zapisania — program jest jeszcze pusty.")
            else:
                self.lbl_status.setText(f"✓ Zapisano kopię ({n} plików) w:\n{sciezka}")
                self.lbl_status.setStyleSheet(f"color:{self._zie}; font-family:'Segoe UI'; font-size:11px; font-weight:600; background:transparent;")
        except Exception as e:
            self.lbl_status.setText(f"✗ Nie udało się zapisać kopii: {e}")
            self.lbl_status.setStyleSheet("color:#F87171; font-family:'Segoe UI'; font-size:11px; background:transparent;")

    def _przywroc(self):
        sciezka, _ = QFileDialog.getOpenFileName(self, "Wybierz kopię zapasową", "", "Archiwum ZIP (*.zip)")
        if not sciezka:
            return
        manifest = sprawdz_kopie_zapasowa(sciezka)
        if not manifest:
            self.lbl_status.setText("✗ To nie jest prawidłowa kopia zapasowa PMT Planer.")
            self.lbl_status.setStyleSheet("color:#F87171; font-family:'Segoe UI'; font-size:11px; background:transparent;")
            return
        data_kopii = manifest.get("data_eksportu", "")[:10]
        dlg = DialogWyboru(
            self, tytul="POTWIERDZENIE",
            pytanie="Przywrócić dane z kopii zapasowej?",
            opis=f"Kopia z dnia {data_kopii} (wersja {manifest.get('wersja', '?')}). "
                 f"To NADPISZE obecne punkty, plan, dziennik i ustawienia.",
            tekst_a="Przywróć", tekst_b="Anuluj", is_dark=self.is_dark)
        if dlg.exec_wybor() != "a":
            return
        try:
            n = przywroc_z_kopii(sciezka)
            self.lbl_status.setText(f"✓ Przywrócono {n} plików. Zamknij i uruchom program ponownie, "
                                    f"żeby wszystkie widoki na pewno zobaczyły świeże dane.")
            self.lbl_status.setStyleSheet(f"color:{self._zie}; font-family:'Segoe UI'; font-size:11px; font-weight:600; background:transparent;")
        except Exception as e:
            self.lbl_status.setText(f"✗ Nie udało się przywrócić kopii: {e}")
            self.lbl_status.setStyleSheet("color:#F87171; font-family:'Segoe UI'; font-size:11px; background:transparent;")


class DialogNotatki(QDialog):
    """Notatka z wizyty — co zastałem, co zrobiłem. To ona czyni dziennik
    wiarygodnym: konkretny opis nie powstaje zza biurka."""
    def __init__(self, parent=None, tekst="", is_dark=True):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self._wynik = None
        if is_dark:
            karta="#0B1320"; ramka="rgba(0,240,255,0.30)"; akc="#00F0FF"; zielony="#00E4A1"
            txt="#F8FAFC"; mut="#94A3B8"; pole="rgba(5,10,20,0.6)"
        else:
            karta="#FFFFFF"; ramka="rgba(13,148,136,0.35)"; akc="#0D9488"; zielony="#059669"
            txt="#0F172A"; mut="#64748B"; pole="rgba(241,245,249,0.9)"

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        box = QFrame(); box.setObjectName("NotBox")
        box.setStyleSheet(f"#NotBox {{ background:{karta}; border:1px solid {ramka}; border-radius:16px; }}")
        box.setFixedWidth(440)
        root.addWidget(box)
        bl = QVBoxLayout(box); bl.setContentsMargins(24, 22, 24, 22); bl.setSpacing(12)

        t = QLabel("Notatka z wizyty")
        t.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:17px; font-weight:800; background:transparent;")
        bl.addWidget(t)
        p = QLabel("Co zastałeś? Co zrobiłeś? Konkretny opis to najlepszy dowód rzetelnej pracy.")
        p.setWordWrap(True)
        p.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
        bl.addWidget(p)

        self.edit = QPlainTextEdit()
        self.edit.setPlainText(tekst or "")
        self.edit.setPlaceholderText("np. Uzupełniono ekspozycję, brak Coca-Coli 1,5 l, kierownik prosi o dostawę w środę…")
        self.edit.setFixedHeight(110)
        self.edit.setStyleSheet(
            f"QPlainTextEdit {{ background:{pole}; border:1px solid {ramka}; border-radius:10px; "
            f"padding:8px; color:{txt}; font-family:'Segoe UI'; font-size:12px; }}")
        bl.addWidget(self.edit)

        akcje = QHBoxLayout(); akcje.setSpacing(10)
        b_cof = QPushButton("Cofnij odhaczenie")
        b_cof.setFixedHeight(38); b_cof.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_cof.setStyleSheet(
            "QPushButton { color:#F87171; background:transparent; border:1px solid rgba(248,113,113,0.5); "
            "border-radius:10px; padding:0 14px; font-family:'Segoe UI'; font-size:11px; font-weight:600; } "
            "QPushButton:hover { background:rgba(248,113,113,0.15); }")
        b_cof.clicked.connect(self._cofnij)
        akcje.addWidget(b_cof)
        akcje.addStretch()
        b_anu = QPushButton("Anuluj"); b_anu.setFixedHeight(38)
        b_anu.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_anu.setStyleSheet(
            f"QPushButton {{ color:{txt}; background:transparent; border:1px solid {ramka}; "
            f"border-radius:10px; padding:0 18px; font-family:'Segoe UI'; font-size:12px; }}")
        b_anu.clicked.connect(self.reject)
        akcje.addWidget(b_anu)
        b_ok = QPushButton("Zapisz"); b_ok.setFixedHeight(38)
        b_ok.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_ok.setStyleSheet(
            f"QPushButton {{ color:#04121A; background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc}, stop:1 {zielony}); "
            f"border:none; border-radius:10px; padding:0 24px; font-family:'Segoe UI'; font-size:12px; font-weight:800; }}")
        b_ok.clicked.connect(self._zapisz)
        akcje.addWidget(b_ok)
        bl.addLayout(akcje)

    def _zapisz(self):
        self._wynik = self.edit.toPlainText().strip()
        self.accept()

    def _cofnij(self):
        self._wynik = "__USUN__"
        self.accept()

    def pokaz(self):
        """Zwraca tekst notatki, '__USUN__' (cofnięcie) albo None (anulowano)."""
        if self.exec() == QDialog.DialogCode.Accepted:
            return self._wynik
        return None


class DialogNotatkaDnia(QDialog):
    """Notatka dnia + 'dzień wolny' — OGÓLNE (nie powiązane z konkretnym
    silnikiem), widoczne z poziomu kalendarza w Planie Wizyt. Zaznaczenie
    'dzień wolny' sprawia, że planer wizyt REALNIE omija ten dzień przy
    kolejnym układaniu tras — nie tylko oznacza go wizualnie."""
    def __init__(self, parent=None, data: datetime.date = None, tekst="", wolne=False, is_dark=True):
        super().__init__(parent)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Dialog)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self._wynik = None
        self._data = data
        if is_dark:
            karta="#0B1320"; ramka="rgba(0,240,255,0.30)"; akc="#00F0FF"; zielony="#00E4A1"
            txt="#F8FAFC"; pole="rgba(5,10,20,0.6)"
        else:
            karta="#FFFFFF"; ramka="rgba(13,148,136,0.35)"; akc="#0D9488"; zielony="#059669"
            txt="#0F172A"; pole="rgba(241,245,249,0.9)"

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        box = QFrame(); box.setObjectName("NotDniaBox")
        box.setStyleSheet(f"#NotDniaBox {{ background:{karta}; border:1px solid {ramka}; border-radius:16px; }}")
        box.setFixedWidth(420)
        root.addWidget(box)
        bl = QVBoxLayout(box); bl.setContentsMargins(24, 22, 24, 22); bl.setSpacing(12)

        dni_pl = ["poniedziałek", "wtorek", "środa", "czwartek", "piątek", "sobota", "niedziela"]
        etyk = f"{dni_pl[data.weekday()]}, {data.strftime('%d.%m.%Y')}" if data else ""
        t = QLabel(f"Notatka — {etyk}")
        t.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:16px; font-weight:800; background:transparent;")
        bl.addWidget(t)

        self.edit = QPlainTextEdit()
        self.edit.setPlainText(tekst or "")
        self.edit.setPlaceholderText("Uwagi, przypomnienia…")
        self.edit.setFixedHeight(100)
        self.edit.setStyleSheet(
            f"QPlainTextEdit {{ background:{pole}; border:1px solid {ramka}; border-radius:10px; "
            f"padding:8px; color:{txt}; font-family:'Segoe UI'; font-size:12px; }}")
        bl.addWidget(self.edit)

        self.chk_wolne = QCheckBox("Dzień wolny — planer wizyt będzie go omijał")
        self.chk_wolne.setChecked(bool(wolne))
        self.chk_wolne.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.chk_wolne.setStyleSheet(
            f"QCheckBox {{ color:{txt}; font-family:'Segoe UI'; font-size:12px; background:transparent; spacing:8px; }} "
            f"QCheckBox::indicator {{ width:16px; height:16px; border:1px solid {ramka}; border-radius:5px; background:{pole}; }} "
            f"QCheckBox::indicator:checked {{ background:#EF4444; border-color:#EF4444; }}")
        bl.addWidget(self.chk_wolne)

        akcje = QHBoxLayout(); akcje.setSpacing(10)
        akcje.addStretch()
        b_anu = QPushButton("Anuluj"); b_anu.setFixedHeight(38)
        b_anu.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_anu.setStyleSheet(
            f"QPushButton {{ color:{txt}; background:transparent; border:1px solid {ramka}; "
            f"border-radius:10px; padding:0 18px; font-family:'Segoe UI'; font-size:12px; }}")
        b_anu.clicked.connect(self.reject)
        akcje.addWidget(b_anu)
        b_ok = QPushButton("Zapisz"); b_ok.setFixedHeight(38)
        b_ok.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b_ok.setStyleSheet(
            f"QPushButton {{ color:#04121A; background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc}, stop:1 {zielony}); "
            f"border:none; border-radius:10px; padding:0 24px; font-family:'Segoe UI'; font-size:12px; font-weight:800; }}")
        b_ok.clicked.connect(self._zapisz)
        akcje.addWidget(b_ok)
        bl.addLayout(akcje)

    def _zapisz(self):
        self._wynik = (self.edit.toPlainText().strip(), self.chk_wolne.isChecked())
        self.accept()

    def pokaz(self):
        """Zwraca (notatka, wolne) albo None (anulowano)."""
        if self.exec() == QDialog.DialogCode.Accepted:
            return self._wynik
        return None


class WezelSladu(QWidget):
    """Węzeł na 'Śladzie Dnia' — kropka na trasie z odcinkami linii w górę/dół.
    Odwiedzone: pełne koło z ptaszkiem i poświatą. Następne w kolejce: pulsujące
    obramowanie. Dalsze: przygaszone. Rysowane QPainterem."""
    def __init__(self, zrobiony=False, nastepny=False, pierwszy=False, ostatni=False,
                 is_dark=True, parent=None):
        super().__init__(parent)
        self.zrobiony = zrobiony
        self.nastepny = nastepny
        self.pierwszy = pierwszy
        self.ostatni = ostatni
        self.is_dark = is_dark
        self.setFixedWidth(34)
        self._puls = 0.0
        if nastepny:
            self._t = QTimer(self); self._t.timeout.connect(self._krok); self._t.start(40)

    def _krok(self):
        self._puls = (self._puls + 0.05) % (2 * math.pi)
        self.update()

    def paintEvent(self, e):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        W, H = self.width(), self.height()
        cx, cy = W / 2, H / 2
        akc = QColor(0, 240, 255) if self.is_dark else QColor(13, 148, 136)
        zielony = QColor(0, 228, 161) if self.is_dark else QColor(5, 150, 105)
        mglisty = QColor(148, 163, 184, 90)

        # odcinki trasy (linia szlaku)
        kol_gora = zielony if self.zrobiony else mglisty
        p.setPen(QPen(kol_gora if not self.pierwszy else Qt.GlobalColor.transparent, 2))
        if not self.pierwszy:
            p.drawLine(QPointF(cx, 0), QPointF(cx, cy - 9))
        # dolny odcinek — zielony tylko gdy ten węzeł zrobiony (ślad "przebyty")
        p.setPen(QPen(zielony if self.zrobiony else mglisty, 2))
        if not self.ostatni:
            p.drawLine(QPointF(cx, cy + 9), QPointF(cx, H))

        # kropka węzła
        if self.zrobiony:
            # poświata
            pos = QRadialGradient(cx, cy, 14)
            g0 = QColor(zielony); g0.setAlpha(110); g1 = QColor(zielony); g1.setAlpha(0)
            pos.setColorAt(0, g0); pos.setColorAt(1, g1)
            p.setBrush(pos); p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(QPointF(cx, cy), 14, 14)
            p.setBrush(zielony); p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(QPointF(cx, cy), 8, 8)
            # ptaszek
            p.setPen(QPen(QColor(4, 18, 26), 2))
            p.drawLine(QPointF(cx - 3.2, cy), QPointF(cx - 0.8, cy + 2.6))
            p.drawLine(QPointF(cx - 0.8, cy + 2.6), QPointF(cx + 3.4, cy - 2.8))
        elif self.nastepny:
            r = 7 + math.sin(self._puls) * 1.6      # pulsuje = "tu jesteś"
            aura = QColor(akc); aura.setAlpha(60)
            p.setBrush(aura); p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(QPointF(cx, cy), r + 5, r + 5)
            p.setBrush(Qt.BrushStyle.NoBrush); p.setPen(QPen(akc, 2))
            p.drawEllipse(QPointF(cx, cy), r, r)
        else:
            p.setBrush(Qt.BrushStyle.NoBrush)
            p.setPen(QPen(mglisty, 1.6))
            p.drawEllipse(QPointF(cx, cy), 6, 6)


class Pieczatka(QWidget):
    """Pieczątka z godziną odhaczenia — jak stempel w papierowym dzienniku.
    Przy pierwszym pokazaniu 'przybija się' (animacja uderzenia)."""
    def __init__(self, czas_txt, is_dark=True, animuj=False, parent=None):
        super().__init__(parent)
        self.czas_txt = czas_txt or ""
        self.is_dark = is_dark
        # pole większe niż sam stempel — obrót o -4° wysuwa rogi poza obrys,
        # więc zostawiamy zapas, żeby nic się nie ucinało
        self.setFixedSize(140, 42)
        self._skala = 1.9 if animuj else 1.0
        self._alfa = 0.0 if animuj else 1.0
        if animuj:
            self._t = QTimer(self); self._t.timeout.connect(self._krok); self._t.start(16)

    def _krok(self):
        # szybkie "uderzenie": skala opada do 1.0, przezroczystość rośnie
        self._skala += (1.0 - self._skala) * 0.28
        self._alfa = min(1.0, self._alfa + 0.10)
        if abs(self._skala - 1.0) < 0.01 and self._alfa >= 1.0:
            self._skala = 1.0; self._alfa = 1.0
            self._t.stop()
        self.update()

    def paintEvent(self, e):
        if not self.czas_txt:
            return
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        W, H = self.width(), self.height()
        tusz = QColor(0, 228, 161) if self.is_dark else QColor(5, 150, 105)
        tusz.setAlphaF(max(0.0, min(1.0, self._alfa)) * 0.95)

        p.translate(W / 2, H / 2)
        p.scale(self._skala, self._skala)
        p.rotate(-4)                       # lekki przekos — jak ręczna pieczątka
        p.translate(-W / 2, -H / 2)

        # ramka pieczątki (przerywana, jak stempel) — z marginesem, by obrót
        # nie wypychał rogów poza widget
        M = 8
        pen = QPen(tusz, 1.6)
        pen.setStyle(Qt.PenStyle.DashLine)
        p.setPen(pen); p.setBrush(Qt.BrushStyle.NoBrush)
        p.drawRoundedRect(QRectF(M, M, W - 2*M, H - 2*M), 6, 6)
        # tekst
        p.setPen(tusz)
        f = QFont("Segoe UI", 8, QFont.Weight.Black)
        p.setFont(f)
        p.drawText(QRectF(M, M, W - 2*M, H - 2*M), Qt.AlignmentFlag.AlignCenter,
                   f"✓ {self.czas_txt}")


class PasekPostepuDnia(QWidget):
    """Autorski pasek postępu wizyt w dniu (QPainter). Wypełnia się na zielono
    w miarę odznaczania wizyt jako odwiedzone."""
    def __init__(self, zrobione, wszystkie, is_dark=True, parent=None):
        super().__init__(parent)
        self.zrobione = zrobione
        self.wszystkie = max(1, wszystkie)
        self.is_dark = is_dark
        self.setFixedHeight(8)

    def paintEvent(self, e):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        W, H = self.width(), self.height()
        r = H / 2
        # tło toru
        tor = QColor(255, 255, 255, 28) if self.is_dark else QColor(15, 23, 42, 28)
        p.setBrush(tor); p.setPen(Qt.PenStyle.NoPen)
        p.drawRoundedRect(QRectF(0, 0, W, H), r, r)
        # wypełnienie
        frakcja = min(1.0, self.zrobione / self.wszystkie)
        if frakcja > 0:
            szer = max(H, W * frakcja)      # minimum = kółko, żeby było widać
            grad = QLinearGradient(0, 0, szer, 0)
            if self.is_dark:
                grad.setColorAt(0, QColor(0, 240, 255)); grad.setColorAt(1, QColor(0, 228, 161))
            else:
                grad.setColorAt(0, QColor(13, 148, 136)); grad.setColorAt(1, QColor(5, 150, 105))
            p.setBrush(grad)
            p.drawRoundedRect(QRectF(0, 0, szer, H), r, r)


class SiatkaMiesiacaPlan(QWidget):
    """Siatka kalendarza dla WIDOKU MIESIĄCA w Planie Wizyt. Prawdziwy
    kalendarz zamiast listy 20 kart do przescrollowania — i JEDYNY kalendarz
    w programie (Kalendarz Wypraw został tu w całości scalony): wizyty,
    delegacje, notatki dnia i dni wolne w jednym miejscu.

    Kolor kafelka = STATUS względem dzisiaj, nie wolumen wizyt. Mapa cieplna
    po liczbie wizyt wypadła słabo w praktyce — przy cyklicznym planowaniu
    większość dni ma bardzo zbliżoną liczbę wizyt, więc intensywność koloru
    prawie się nie różniła. Status (zaległe/dziś/zaplanowane/zrobione) różni
    się NAPRAWDĘ z dnia na dzień — i to właśnie warto widzieć jednym rzutem
    oka: gdzie mam zaległości, a co mam już z głowy.

    Delegacje (stary silnik) pokazane jak dawniej — wstążką w rogu kafelka,
    więc dzień z wizytami I delegacją pokazuje oba naraz, bez konfliktu."""
    _DNI_NAGLOWEK = ["Pon", "Wt", "Śr", "Czw", "Pt", "Sob", "Ndz"]

    def __init__(self, parent=None, is_dark=True, on_klik=None, on_notatka=None):
        super().__init__(parent)
        self.is_dark = is_dark
        self.on_klik = on_klik
        self.on_notatka = on_notatka     # callback(data) — otwiera DialogNotatkaDnia
        self.rok = datetime.date.today().year
        self.mies = datetime.date.today().month
        self._dni_planu = {}     # {dzien_liczba: DzienPlanu}
        self._delegacje = {}     # {iso: wpis_rozliczenia} — z historii (Moje Szlaki)
        self.warstwa_wizyty = True
        self.warstwa_delegacje = True
        self._rects = {}
        self._rects_notatka = {}
        self._hover = None
        self.setMouseTracking(True)
        self.setMinimumHeight(360)

    def ustaw_miesiac(self, rok, mies, dni_planu):
        self.rok, self.mies = rok, mies
        self._dni_planu = dni_planu or {}
        self.update()

    def ustaw_delegacje(self, delegacje):
        self._delegacje = delegacje or {}
        self.update()

    def ustaw_warstwy(self, wizyty=True, delegacje=True):
        self.warstwa_wizyty = bool(wizyty)
        self.warstwa_delegacje = bool(delegacje)
        self.update()

    def _dni_w_miesiacu(self):
        return calendar.monthrange(self.rok, self.mies)[1]

    def _pierwszy_dzien_tygodnia(self):
        return calendar.monthrange(self.rok, self.mies)[0]   # 0=poniedziałek

    def _sciezka_kafelka(self, rect):
        sc = QPainterPath(); sc.addRoundedRect(rect, 8, 8)
        return sc

    def mousePressEvent(self, e):
        pt = QPointF(e.position())
        # notatka ma pierwszeństwo — to mały obszar w rogu, sprawdzamy najpierw
        for dzien, rect in self._rects_notatka.items():
            if rect.contains(pt):
                if self.on_notatka:
                    data_kom = datetime.date(self.rok, self.mies, dzien)
                    self.on_notatka(data_kom)
                return
        for dzien, rect in self._rects.items():
            if rect.contains(pt):
                if self.on_klik:
                    self.on_klik(dzien)
                return

    def mouseMoveEvent(self, e):
        stary = self._hover
        self._hover = None
        pt = QPointF(e.position())
        for dzien, rect in self._rects.items():
            if rect.contains(pt):
                self._hover = dzien
                break
        if stary != self._hover:
            self.update()

    def leaveEvent(self, e):
        self._hover = None
        self.update()

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        W, H = self.width(), self.height()
        self._rects = {}
        self._rects_notatka = {}

        akc = QColor("#00F0FF") if self.is_dark else QColor("#0D9488")
        zie = QColor("#00E4A1") if self.is_dark else QColor("#059669")
        zal = QColor("#FBBF24") if self.is_dark else QColor("#B45309")   # ten sam kolor co baner "zaległe"
        krw = QColor("#DC2626") if self.is_dark else QColor("#B91C1C")   # krwista czerwień: dzień minął bez ani jednej wizyty
        wolny_kol = QColor("#EF4444") if self.is_dark else QColor("#DC2626")
        txt = QColor("#F8FAFC") if self.is_dark else QColor("#0F172A")
        mut = QColor("#94A3B8") if self.is_dark else QColor("#64748B")
        siatka_linia = QColor(255, 255, 255, 18) if self.is_dark else QColor(15, 23, 42, 18)

        margines = 6
        naglowek_h = 24
        n_dni = self._dni_w_miesiacu()
        pierwszy_tydz = self._pierwszy_dzien_tygodnia()
        n_tygodni = math.ceil((n_dni + pierwszy_tydz) / 7)

        szer_kol = (W - 2 * margines) / 7
        wys_wiersz = (H - naglowek_h - 2 * margines) / max(1, n_tygodni)

        p.setFont(QFont("Segoe UI", 9, QFont.Weight.DemiBold))
        p.setPen(mut)
        for i, nazwa in enumerate(self._DNI_NAGLOWEK):
            rx = margines + i * szer_kol
            p.drawText(QRectF(rx, 2, szer_kol, naglowek_h - 4), Qt.AlignmentFlag.AlignCenter, nazwa)

        dzis = datetime.date.today()
        for dzien in range(1, n_dni + 1):
            idx = pierwszy_tydz + dzien - 1
            kol = idx % 7
            wiersz = idx // 7
            rx = margines + kol * szer_kol
            ry = margines + naglowek_h + wiersz * wys_wiersz
            rect = QRectF(rx + 2, ry + 2, szer_kol - 4, wys_wiersz - 4)
            self._rects[dzien] = rect

            data_kom = datetime.date(self.rok, self.mies, dzien)
            iso = data_kom.isoformat()
            dp = self._dni_planu.get(dzien) if self.warstwa_wizyty else None
            ma_del = (iso in self._delegacje) if self.warstwa_delegacje else False
            n_wpis = notatka_dnia(iso)
            wolny = bool(n_wpis.get("wolne"))
            ma_notatke = bool(n_wpis.get("notatka"))
            jest_dzis = (data_kom == dzis)
            hover = (dzien == self._hover)

            if wolny:
                # DZIEŃ WOLNY ma pierwszeństwo wizualne — planer go i tak omija
                kol_tlo = QColor(wolny_kol); kol_tlo.setAlphaF(0.16)
                p.setBrush(kol_tlo)
                p.setPen(QPen(QColor(wolny_kol.red(), wolny_kol.green(), wolny_kol.blue(), 140), 1.4))
            elif dp and dp.wizyty:
                n = len(dp.wizyty)
                zrobione = sum(1 for wz in dp.wizyty if czy_odwiedzona(data_kom, wz.adres or wz.nazwa))
                komplet = zrobione >= n

                # STATUS decyduje o kolorze — nie liczba wizyt (patrz docstring klasy)
                if komplet:
                    baza = zie; intensywnosc = 0.28
                elif data_kom < dzis and zrobione == 0:
                    # DZIEŃ NIEZREALIZOWANY (minął, zero wizyt) — krwista czerwień
                    baza = krw; intensywnosc = 0.38
                elif data_kom < dzis:
                    baza = zal; intensywnosc = 0.30
                elif jest_dzis:
                    baza = akc; intensywnosc = 0.32
                else:
                    baza = akc; intensywnosc = 0.12

                kol_tlo = QColor(baza); kol_tlo.setAlphaF(intensywnosc)
                p.setBrush(kol_tlo)
                if komplet:
                    obrys = zie
                elif hover or jest_dzis:
                    obrys = baza
                else:
                    obrys = QColor(baza.red(), baza.green(), baza.blue(), 100)
                p.setPen(QPen(obrys, 1.6 if (jest_dzis or hover or komplet or data_kom < dzis) else 1.0))
            else:
                p.setBrush(QColor(255, 255, 255, 6) if self.is_dark else QColor(15, 23, 42, 4))
                p.setPen(QPen(siatka_linia, 1.0))
            p.drawRoundedRect(rect, 8, 8)

            # WSTĄŻKA DELEGACJI — róg, przycięta do zaokrąglonego kształtu kafelka
            if ma_del:
                wst = QPainterPath()
                bok = 20.0
                rx2, ry2 = rect.right(), rect.top()
                wst.moveTo(rx2 - bok, ry2); wst.lineTo(rx2, ry2); wst.lineTo(rx2, ry2 + bok)
                wst.closeSubpath()
                p.setBrush(zie); p.setPen(Qt.PenStyle.NoPen)
                p.setClipPath(self._sciezka_kafelka(rect))
                p.drawPath(wst)
                p.setClipping(False)
                p.setPen(QColor("#04121A"))
                p.setFont(QFont("Segoe UI", 6, QFont.Weight.Black))
                p.drawText(QRectF(rx2 - bok + 3, ry2 + 1, bok - 4, 11),
                           Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignTop, "zł")

            p.setPen(txt if (dp or wolny) else mut)
            p.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold if jest_dzis else QFont.Weight.Normal))
            p.drawText(QRectF(rect.x() + 6, rect.y() + 3, rect.width() - 12, 16),
                       Qt.AlignmentFlag.AlignLeft, str(dzien))
            if jest_dzis:
                p.setBrush(akc); p.setPen(Qt.PenStyle.NoPen)
                p.drawEllipse(QPointF(rect.right() - 10, rect.y() + 10), 3, 3)

            # ODZNAKA WIZYT (n/m albo ✓) — dolny prawy róg
            if dp and dp.wizyty and not wolny:
                n = len(dp.wizyty)
                zrobione = sum(1 for wz in dp.wizyty if czy_odwiedzona(data_kom, wz.adres or wz.nazwa))
                komplet = zrobione >= n
                etyk = "✓" if komplet else f"{zrobione}/{n}"
                kol_etyk = (zie if komplet
                            else (krw if (data_kom < dzis and zrobione == 0)
                                  else (zal if data_kom < dzis else akc)))
                p.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
                p.setPen(kol_etyk)
                p.drawText(QRectF(rect.x(), rect.bottom() - 18, rect.width() - 6, 16),
                          Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter, etyk)

            # WSKAŹNIK NOTATKI — mała kropka w dolnym LEWYM rogu (klikalna,
            # osobny hit-test od reszty kafelka — otwiera DialogNotatkaDnia)
            notatka_rect = QRectF(rect.x() + 2, rect.bottom() - 16, 16, 16)
            self._rects_notatka[dzien] = notatka_rect
            if ma_notatke:
                kol_kropki = wolny_kol if wolny else mut
                p.setBrush(kol_kropki); p.setPen(Qt.PenStyle.NoPen)
                p.drawEllipse(QPointF(rect.x() + 9, rect.bottom() - 9), 3, 3)


class PrzystanekTygodnia(QFrame):
    """Jeden 'przystanek' na Torze Tygodnia — dzień jako stacja na trasie:
    nazwa+data, pierścień postępu, wizyt/km. Klik przenosi w Ślad Dnia."""
    RING = 56
    Y_LINII = 78     # stały offset środka pierścienia — używany przez LacznikTygodnia

    def __init__(self, parent, dzien, is_dark, aktywny, on_klik):
        super().__init__(parent)
        self.dzien = dzien
        self.is_dark = is_dark
        self._aktywny = aktywny
        self._on_klik = on_klik
        self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.setObjectName("Przystanek")

        l = QVBoxLayout(self); l.setContentsMargins(4, 6, 4, 8); l.setSpacing(4)
        l.setAlignment(Qt.AlignmentFlag.AlignHCenter)

        dni_pl = ["Pon", "Wt", "Śr", "Czw", "Pt", "Sob", "Ndz"]
        jest_dzis = dzien.data == datetime.date.today()
        nazwa = "DZIŚ" if jest_dzis else dni_pl[dzien.data.weekday()]
        self.l_nazwa = QLabel(nazwa); self.l_nazwa.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.l_nazwa.setFixedHeight(18)
        l.addWidget(self.l_nazwa)
        self.l_data = QLabel(dzien.data.strftime("%d.%m")); self.l_data.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.l_data.setFixedHeight(14)
        l.addWidget(self.l_data)
        l.addSpacing(4)

        wrap_ring = QHBoxLayout(); wrap_ring.addStretch()
        self.ring = PierscienPostepu(self, is_dark, rozmiar=self.RING)
        n = len(dzien.wizyty)
        zrobione = sum(1 for wz in dzien.wizyty if czy_odwiedzona(dzien.data, wz.adres or wz.nazwa))
        self.ring.ustaw(zrobione, n)
        wrap_ring.addWidget(self.ring); wrap_ring.addStretch()
        l.addLayout(wrap_ring)
        l.addSpacing(4)

        self.l_meta = QLabel(f"{n} {'wizyta' if n==1 else 'wizyty' if 2<=n<=4 else 'wizyt'}" if n else "—")
        self.l_meta.setAlignment(Qt.AlignmentFlag.AlignCenter)
        l.addWidget(self.l_meta)
        self.l_km = QLabel(f"{dzien.km:.0f} km" if n else "")
        self.l_km.setAlignment(Qt.AlignmentFlag.AlignCenter)
        l.addWidget(self.l_km)

        self.update_theme(is_dark, aktywny)

    def update_theme(self, is_dark, aktywny):
        self.is_dark = is_dark
        self._aktywny = aktywny
        self.ring.is_dark = is_dark; self.ring.update()
        akc = "#00F0FF" if is_dark else "#0D9488"
        txt = "#F8FAFC" if is_dark else "#0F172A"
        mut = "#94A3B8" if is_dark else "#64748B"
        if aktywny:
            bg = "rgba(0,240,255,0.08)" if is_dark else "rgba(13,148,136,0.08)"
            ramka = akc
        else:
            bg = "transparent"; ramka = "transparent"
        self.setStyleSheet(f"#Przystanek {{ background:{bg}; border:1.5px solid {ramka}; border-radius:12px; }}")
        self.l_nazwa.setStyleSheet(f"color:{akc if aktywny else txt}; font-family:'Segoe UI'; font-size:12px; font-weight:800; background:transparent; border:none;")
        self.l_data.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:10px; background:transparent; border:none;")
        self.l_meta.setStyleSheet(f"color:{txt}; font-family:'Segoe UI'; font-size:11px; font-weight:700; background:transparent; border:none;")
        self.l_km.setStyleSheet(f"color:{mut}; font-family:'Segoe UI'; font-size:10px; background:transparent; border:none;")

    def mousePressEvent(self, e):
        if self._on_klik:
            self._on_klik(self.dzien.data)
        super().mousePressEvent(e)


class LacznikTygodnia(QWidget):
    """Krótki odcinek 'trasy' między przystankami — cienka linia na
    wysokości środka pierścienia, żeby dni tygodnia czytały się jako
    przystanki NA JEDNEJ TRASIE, a nie osobne, niepowiązane karty."""
    def __init__(self, parent=None, is_dark=True):
        super().__init__(parent)
        self.is_dark = is_dark
        self.setFixedWidth(18)
        self.setMinimumHeight(140)

    def paintEvent(self, e):
        p = QPainter(self); p.setRenderHint(QPainter.RenderHint.Antialiasing)
        akc = QColor("#00F0FF" if self.is_dark else "#0D9488")
        akc.setAlpha(110)
        y = PrzystanekTygodnia.Y_LINII
        p.setPen(QPen(akc, 2))
        p.drawLine(QPointF(0, y), QPointF(self.width(), y))
        p.setBrush(akc); p.setPen(Qt.PenStyle.NoPen)
        p.drawEllipse(QPointF(self.width() / 2, y), 2.5, 2.5)


class TorTygodnia(QFrame):
    """Tydzień jako TRASA: dni robocze jako przystanki połączone jedną
    linią. Celowo inny kształt niż pionowy Ślad Dnia — dzięki temu tydzień
    daje prawdziwe porównanie dni jednym rzutem oka (który dzień najcięższy,
    który pusty), zamiast być kolejną pionową listą do przescrollowania."""
    def __init__(self, parent=None, is_dark=True, on_klik=None):
        super().__init__(parent)
        self.is_dark = is_dark
        self._on_klik = on_klik
        self._l = QHBoxLayout(self)
        self._l.setContentsMargins(10, 4, 10, 4)
        self._l.setSpacing(0)
        self.setObjectName("TorTyg")

    def ustaw_dni(self, dni_tygodnia, dzien_aktywny):
        while self._l.count():
            it = self._l.takeAt(0)
            w = it.widget()
            if w: w.deleteLater()
        for i, d in enumerate(dni_tygodnia):
            if i > 0:
                self._l.addWidget(LacznikTygodnia(self, self.is_dark))
            stop = PrzystanekTygodnia(self, d, self.is_dark, d.data == dzien_aktywny, self._on_klik)
            self._l.addWidget(stop, 1)
        self.update_theme(self.is_dark)

    def update_theme(self, is_dark):
        self.is_dark = is_dark
        bg = "rgba(11,19,32,0.5)" if is_dark else "rgba(255,255,255,0.6)"
        ramka = "rgba(0,240,255,0.15)" if is_dark else "rgba(13,148,136,0.18)"
        self.setStyleSheet(f"#TorTyg {{ background:{bg}; border:1px solid {ramka}; border-radius:14px; }}")
        for i in range(self._l.count()):
            w = self._l.itemAt(i).widget()
            if isinstance(w, LacznikTygodnia):
                w.is_dark = is_dark; w.update()
            elif isinstance(w, PrzystanekTygodnia):
                w.update_theme(is_dark, w._aktywny)


class PlanWizytOverlay(QFrame):
    """Nakładka prezentująca PLAN WIZYT z nowego silnika planera. Trzy widoki:
    dzień / tydzień / miesiąc. Przełącznik u góry, treść w scrollu, przycisk
    podglądu na mapie (Google Maps). Dane z planuj_wizyty()."""
    _DNI_PL = ["Poniedziałek","Wtorek","Środa","Czwartek","Piątek","Sobota","Niedziela"]
    _MIES_PL = ["","styczeń","luty","marzec","kwiecień","maj","czerwiec","lipiec",
                "sierpień","wrzesień","październik","listopad","grudzień"]

    def __init__(self, parent=None, on_mapa=None):
        super().__init__(parent); self.hide()
        self.is_dark = True
        self._plan = None
        self._widok = "dzien"          # dzien | tydzien | miesiac
        self._on_mapa = on_mapa
        self._dzien_idx = 0            # który dzień pokazujemy w widoku dziennym
        self._tydzien_idx = 0
        self._swiezo_odhaczona = None  # klucz wizyty do animacji pieczątki
        self._on_przenies = None       # callback: przenieś zaległe wizyty
        self._on_tryb_trasy = None     # callback: otwórz Tryb Trasy dla dnia
        self._on_toast = None          # callback: krótki komunikat
        self._on_punkty_zmienione = None   # callback: odśwież listę w planerze
        self._tydzien_aktywny_dzien = None  # który dzień jest "rozwinięty" pod Torem Tygodnia
        self._warstwa_kalendarza = "all"    # all | wizyty | delegacje — dla widoku Miesiąc
        self._on_dane_uzytkownika = None    # callback() -> (imię, pesel) — do wczytania delegacji

        root = QVBoxLayout(self); root.setContentsMargins(0, 0, 0, 0)
        self.karta = QFrame(self); self.karta.setObjectName("PlanKarta")
        # karta responsywna — rozmiar ustawiany w resizeEvent (wypełnia okno
        # z marginesem). Sztywny rozmiar marnował ekran przy dużych planach.
        self.karta.setMinimumSize(700, 460)
        root.addWidget(self.karta, alignment=Qt.AlignmentFlag.AlignCenter)
        kl = QVBoxLayout(self.karta); kl.setContentsMargins(30, 26, 30, 26); kl.setSpacing(16)

        # nagłówek
        gora = QHBoxLayout()
        tyt = QVBoxLayout(); tyt.setSpacing(2)
        self.tytul = QLabel("Plan Wizyt")
        self.podtytul = QLabel("Propozycja odwiedzin ułożona z głową — pon–pt, bez świąt")
        tyt.addWidget(self.tytul); tyt.addWidget(self.podtytul)
        gora.addLayout(tyt); gora.addStretch()
        self.btn_mapa = QPushButton("🗺  Podgląd na mapie"); self.btn_mapa.setFixedHeight(38)
        self.btn_mapa.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_mapa.clicked.connect(self._klik_mapa)
        gora.addWidget(self.btn_mapa)
        # szukaj: "kiedy jestem w sklepie X?" — przeszukuje CAŁY wczytany plan
        self.btn_szukaj = QPushButton("🔍"); self.btn_szukaj.setFixedSize(38, 38)
        self.btn_szukaj.setCheckable(True)
        self.btn_szukaj.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_szukaj.clicked.connect(self._przelacz_szukaj)
        gora.addWidget(self.btn_szukaj)
        # tryb trasy — pełnoekranowy widok "na czas jazdy" dla WYŚWIETLANEGO dnia
        self.btn_tryb_trasy = QPushButton("🚗  Tryb Trasy"); self.btn_tryb_trasy.setFixedHeight(38)
        self.btn_tryb_trasy.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_tryb_trasy.clicked.connect(self._klik_tryb_trasy)
        gora.addWidget(self.btn_tryb_trasy)
        # eksport — Excel (plan + dziennik) albo PDF (drukowalny wykaz)
        self.btn_eksport = QPushButton("⬇  Eksportuj"); self.btn_eksport.setFixedHeight(38)
        self.btn_eksport.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_eksport.clicked.connect(self._klik_eksport)
        gora.addWidget(self.btn_eksport)
        self.btn_x = QPushButton("✕"); self.btn_x.setFixedSize(38, 38)
        self.btn_x.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_x.clicked.connect(self.hide)
        gora.addWidget(self.btn_x)
        kl.addLayout(gora)

        # przełącznik widoków (segmentowany)
        przel = QHBoxLayout(); przel.setSpacing(0); przel.setContentsMargins(0,0,0,0)
        self.btn_dzien = QPushButton("Dzień")
        self.btn_tydzien = QPushButton("Tydzień")
        self.btn_miesiac = QPushButton("Miesiąc")
        for b in (self.btn_dzien, self.btn_tydzien, self.btn_miesiac):
            b.setCheckable(True); b.setFixedHeight(38); b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_dzien.setChecked(True)
        self.btn_dzien.clicked.connect(lambda: self._ustaw_widok("dzien"))
        self.btn_tydzien.clicked.connect(lambda: self._ustaw_widok("tydzien"))
        self.btn_miesiac.clicked.connect(lambda: self._ustaw_widok("miesiac"))
        seg = QHBoxLayout(); seg.setSpacing(8)
        seg.addWidget(self.btn_dzien); seg.addWidget(self.btn_tydzien); seg.addWidget(self.btn_miesiac)
        seg.addStretch()
        # nawigacja (dla dnia/tygodnia) + podsumowanie
        self.btn_prev = QPushButton("‹"); self.btn_prev.setFixedSize(34, 34)
        self.btn_next = QPushButton("›"); self.btn_next.setFixedSize(34, 34)
        self.lbl_zakres = QLabel(""); self.lbl_zakres.setMinimumWidth(200)
        self.lbl_zakres.setAlignment(Qt.AlignmentFlag.AlignCenter)
        for b in (self.btn_prev, self.btn_next):
            b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_prev.clicked.connect(lambda: self._nawiguj(-1))
        self.btn_next.clicked.connect(lambda: self._nawiguj(1))
        seg.addWidget(self.btn_prev); seg.addWidget(self.lbl_zakres); seg.addWidget(self.btn_next)
        kl.addLayout(seg)

        # przełącznik warstw — TYLKO w widoku Miesiąc (scalony Kalendarz Wypraw:
        # wizyty + delegacje na jednej siatce, można filtrować którą warstwę widać)
        self.warstwy_row = QHBoxLayout(); self.warstwy_row.setSpacing(8)
        self.lbl_warstwy = QLabel("Pokaż:")
        self.warstwy_row.addWidget(self.lbl_warstwy)
        self.btn_w_all = QPushButton("Wszystko"); self.btn_w_all.setCheckable(True); self.btn_w_all.setChecked(True)
        self.btn_w_wiz = QPushButton("▣  Wizyty"); self.btn_w_wiz.setCheckable(True)
        self.btn_w_del = QPushButton("◤  Delegacje"); self.btn_w_del.setCheckable(True)
        for b in (self.btn_w_all, self.btn_w_wiz, self.btn_w_del):
            b.setFixedHeight(28); b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_w_all.clicked.connect(lambda: self._ustaw_warstwe_kalendarza("all"))
        self.btn_w_wiz.clicked.connect(lambda: self._ustaw_warstwe_kalendarza("wizyty"))
        self.btn_w_del.clicked.connect(lambda: self._ustaw_warstwe_kalendarza("delegacje"))
        self.warstwy_row.addWidget(self.btn_w_all); self.warstwy_row.addWidget(self.btn_w_wiz); self.warstwy_row.addWidget(self.btn_w_del)
        self.warstwy_row.addStretch()
        self.lbl_legenda_mies = QLabel("●  wpis / notatka dnia")
        self.warstwy_row.addWidget(self.lbl_legenda_mies)
        self._warstwy_row_widgets = [self.lbl_warstwy, self.btn_w_all, self.btn_w_wiz, self.btn_w_del, self.lbl_legenda_mies]
        kl.addLayout(self.warstwy_row)

        # --- PANEL SZUKANIA: "kiedy jestem w sklepie X?" — ukryty domyślnie ---
        self.panel_szukaj = QFrame(); self.panel_szukaj.setObjectName("PlanPanelSzukaj")
        self.panel_szukaj.setVisible(False)
        psz = QVBoxLayout(self.panel_szukaj); psz.setContentsMargins(14, 12, 14, 12); psz.setSpacing(8)
        self.pole_szukaj_plan = GrubyKursorEdit()
        self.pole_szukaj_plan.setPlaceholderText("Szukaj sklepu, miasta lub sieci — np. „Biedronka Radom”…")
        self.pole_szukaj_plan.setFixedHeight(38)
        self.pole_szukaj_plan.textChanged.connect(self._szukaj_w_planie)
        psz.addWidget(self.pole_szukaj_plan)
        self.wyniki_szukaj_scroll = QScrollArea(); self.wyniki_szukaj_scroll.setWidgetResizable(True)
        self.wyniki_szukaj_scroll.setMaximumHeight(240)
        self.wyniki_szukaj_scroll.setObjectName("WynikiSzukajScroll")
        self.wyniki_szukaj_wnetrze = QWidget()
        self.wyniki_szukaj_l = QVBoxLayout(self.wyniki_szukaj_wnetrze)
        self.wyniki_szukaj_l.setContentsMargins(4, 4, 4, 4); self.wyniki_szukaj_l.setSpacing(6)
        self.wyniki_szukaj_scroll.setWidget(self.wyniki_szukaj_wnetrze)
        self.wyniki_szukaj_scroll.setVisible(False)
        psz.addWidget(self.wyniki_szukaj_scroll)
        self.lbl_szukaj_podpowiedz = QLabel("Wpisz min. 2 znaki, aby wyszukać w całym planie.")
        psz.addWidget(self.lbl_szukaj_podpowiedz)
        kl.addWidget(self.panel_szukaj)

        # --- BANER ZALEGŁYCH: wizyty z minionych dni, których nie odhaczono ---
        self.baner = QFrame(); self.baner.setObjectName("BanerZalegle")
        self.baner.setVisible(False)
        bl_b = QHBoxLayout(self.baner); bl_b.setContentsMargins(16, 10, 12, 10); bl_b.setSpacing(12)
        self.lbl_baner = QLabel("")
        self.lbl_baner.setWordWrap(True)
        bl_b.addWidget(self.lbl_baner, 1)
        self.btn_przenies = QPushButton("Przenieś na kolejne dni")
        self.btn_przenies.setFixedHeight(32)
        self.btn_przenies.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_przenies.clicked.connect(self._klik_przenies)
        bl_b.addWidget(self.btn_przenies)
        kl.addWidget(self.baner)

        # obszar treści (scroll)
        self.scroll = QScrollArea(); self.scroll.setWidgetResizable(True)
        self.scroll.setObjectName("PlanScroll")
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.tresc = QWidget(); self.tresc_l = QVBoxLayout(self.tresc)
        self.tresc_l.setContentsMargins(4, 4, 4, 4); self.tresc_l.setSpacing(12)
        self.tresc_l.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.scroll.setWidget(self.tresc)
        kl.addWidget(self.scroll, 1)

        # stopka — podsumowanie
        self.lbl_podsum = QLabel("")
        self.lbl_podsum.setAlignment(Qt.AlignmentFlag.AlignCenter)
        kl.addWidget(self.lbl_podsum)

        # "bez lokalizacji" — KLIKALNE (nie tylko liczba), pozwala zobaczyć
        # KTÓRE konkretnie adresy się nie zgeokodowały i je posprzątać
        row_bez = QHBoxLayout(); row_bez.addStretch()
        self.btn_bez_lokalizacji = QPushButton("")
        self.btn_bez_lokalizacji.setVisible(False)
        self.btn_bez_lokalizacji.setFixedHeight(26)
        self.btn_bez_lokalizacji.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_bez_lokalizacji.clicked.connect(self._pokaz_bez_lokalizacji)
        row_bez.addWidget(self.btn_bez_lokalizacji)
        row_bez.addStretch()
        kl.addLayout(row_bez)

    def ustaw_plan(self, plan: dict):
        self._plan = plan
        self._dzien_idx = 0; self._tydzien_idx = 0; self._mies_idx = 0
        self._skocz_na_dzis()
        self._przerysuj()

    def _skocz_na_dzis(self):
        """Ustawia widok na DZISIEJSZY dzień (albo najbliższy przyszły).
        Bez tego plan na 3 miesiące otwierał się zawsze na 1. dniu i trzeba
        było klikać strzałkę kilkanaście razy."""
        mm = self._miesiace()
        if not mm:
            return
        dzis = datetime.date.today()
        # 1) szukamy dnia == dziś, inaczej pierwszego dnia po dziś
        cel_m = cel_d = None
        for im, m in enumerate(mm):
            for idd, d in enumerate(m["dni"]):
                if d.data == dzis:
                    cel_m, cel_d = im, idd
                    break
                if d.data > dzis and cel_m is None:
                    cel_m, cel_d = im, idd
            if cel_m is not None and mm[cel_m]["dni"][cel_d].data == dzis:
                break
        if cel_m is None:
            # plan w całości w przeszłości — pokaż ostatni dzień
            cel_m = len(mm) - 1
            cel_d = len(mm[cel_m]["dni"]) - 1
        self._mies_idx = cel_m
        self._dzien_idx = cel_d
        # tydzień odpowiadający temu dniowi
        try:
            m = mm[cel_m]
            dni_rob = pobierz_dni_robocze(m["rok"], m["miesiac"])
            nr = _tydzien_numer(m["dni"][cel_d].data, dni_rob)
            self._tydzien_idx = max(0, nr - 1)
        except Exception:
            self._tydzien_idx = 0

    def resizeEvent(self, e):
        """Karta wypełnia okno (z marginesem) — przy dużych planach potrzeba
        miejsca; sztywny rozmiar marnował ekran."""
        super().resizeEvent(e)
        M = 28                                  # margines wokół karty
        w = max(700, self.width() - 2 * M)
        h = max(460, self.height() - 2 * M)
        self.karta.setFixedSize(w, h)

    def _odswiez_baner(self):
        """Pokazuje baner, gdy są niezrealizowane wizyty z minionych dni."""
        zal = zalegle_wizyty(self._plan) if self._plan else []
        n = len(zal)
        if not n:
            self.baner.setVisible(False)
            return
        self.baner.setVisible(True)
        slowo = "wizyta" if n == 1 else ("wizyty" if 2 <= n <= 4 else "wizyt")
        self.lbl_baner.setText(
            f"⚠  {n} {slowo} z minionych dni bez odhaczenia. "
            f"Mogę je przenieść na kolejne dni robocze — historia (to, co odhaczone) zostanie nietknięta.")
        akc = "#FBBF24" if self.is_dark else "#B45309"
        tlo = "rgba(251,191,36,0.10)" if self.is_dark else "rgba(180,83,9,0.08)"
        self.baner.setStyleSheet(
            f"#BanerZalegle {{ background:{tlo}; border:1px solid {akc}; border-radius:10px; }}")
        self.lbl_baner.setStyleSheet(
            f"color:{akc}; font-family:'Segoe UI'; font-size:11px; font-weight:600; background:transparent; border:none;")
        self.btn_przenies.setStyleSheet(
            f"QPushButton {{ color:#0B1320; background:{akc}; border:none; border-radius:8px; "
            f"padding:0 14px; font-family:'Segoe UI'; font-size:11px; font-weight:800; }} "
            f"QPushButton:hover {{ background:#FCD34D; }}")

    def _klik_przenies(self):
        if self._on_przenies:
            self._on_przenies()

    def _ustaw_widok_bez_skoku(self, w):
        """Jak _ustaw_widok, ale ZACHOWUJE wybrany dzień (wejście z kalendarza)."""
        self._widok = w
        self.btn_dzien.setChecked(w == "dzien")
        self.btn_tydzien.setChecked(w == "tydzien")
        self.btn_miesiac.setChecked(w == "miesiac")
        self._przerysuj()

    def _ustaw_widok(self, w):
        self._widok = w
        self.btn_dzien.setChecked(w == "dzien")
        self.btn_tydzien.setChecked(w == "tydzien")
        self.btn_miesiac.setChecked(w == "miesiac")
        # przy zmianie widoku trzymamy się DZISIEJSZEJ daty, a nie początku planu
        self._skocz_na_dzis()
        self._przerysuj()

    def _miesiace(self):
        """Lista planów miesięcznych (nowy format) lub jeden zawinięty (stary)."""
        if not self._plan:
            return []
        if self._plan.get("miesiace"):
            return self._plan["miesiace"]
        # zgodność wstecz: pojedynczy miesiąc
        if self._plan.get("dni"):
            return [{"rok": self._plan.get("rok"), "miesiac": self._plan.get("miesiac"),
                     "dni": self._plan["dni"], "suma_km": self._plan.get("suma_km", 0),
                     "suma_wizyt": self._plan.get("suma_wizyt", 0)}]
        return []

    def _biezacy_miesiac(self):
        mm = self._miesiace()
        if not mm:
            return None
        self._mies_idx = max(0, min(len(mm) - 1, getattr(self, "_mies_idx", 0)))
        return mm[self._mies_idx]

    def _nawiguj(self, kier):
        mm = self._miesiace()
        if not mm:
            return
        mies = self._biezacy_miesiac()
        dni = mies["dni"]
        if self._widok == "dzien":
            # nawigacja po dniach; przekroczenie granicy → zmiana miesiąca
            nowy = self._dzien_idx + kier
            if nowy < 0:
                if self._mies_idx > 0:
                    self._mies_idx -= 1; self._dzien_idx = len(self._biezacy_miesiac()["dni"]) - 1
            elif nowy >= len(dni):
                if self._mies_idx < len(mm) - 1:
                    self._mies_idx += 1; self._dzien_idx = 0
            else:
                self._dzien_idx = nowy
        elif self._widok == "tydzien":
            maxtyg = self._liczba_tygodni()
            nowy = self._tydzien_idx + kier
            if nowy < 0:
                if self._mies_idx > 0:
                    self._mies_idx -= 1; self._tydzien_idx = 0
            elif nowy >= maxtyg:
                if self._mies_idx < len(mm) - 1:
                    self._mies_idx += 1; self._tydzien_idx = 0
            else:
                self._tydzien_idx = nowy
        else:  # miesiac — nawigacja między miesiącami
            self._mies_idx = max(0, min(len(mm) - 1, self._mies_idx + kier))
        self._przerysuj()

    def _biezacy_dzien_obiekt(self):
        """DzienPlanu aktualnie pokazywany w widoku 'dzień' (albo None)."""
        if self._widok != "dzien":
            return None
        mm = self._miesiace()
        if not mm:
            return None
        mies = self._biezacy_miesiac()
        dni = mies["dni"]
        if not dni:
            return None
        idx = max(0, min(len(dni) - 1, self._dzien_idx))
        return dni[idx]

    def _klik_tryb_trasy(self):
        d = self._biezacy_dzien_obiekt()
        if d is None:
            if self._on_toast:
                self._on_toast("Wybierz dzień", "Tryb Trasy działa dla pojedynczego dnia — przełącz widok na „Dzień”.", False)
            return
        if not d.wizyty:
            if self._on_toast:
                self._on_toast("Brak wizyt", "Ten dzień nie ma zaplanowanych wizyt.", False)
            return
        if self._on_tryb_trasy:
            self._on_tryb_trasy(d, d.data)

    def _przelacz_szukaj(self, otwarte):
        self.panel_szukaj.setVisible(bool(otwarte))
        if not otwarte:
            self.pole_szukaj_plan.clear()
        else:
            self.pole_szukaj_plan.setFocus()

    def _szukaj_w_planie(self, tekst):
        """Przeszukuje CAŁY wczytany plan (nie tylko widoczny miesiąc) po
        nazwie sklepu, adresie, mieście lub sieci. Nadchodzące wizyty na
        górze (chronologicznie), minione niżej (najnowsze pierwsze) —
        odpowiada wprost na „kiedy NASTĘPNYM razem jestem w sklepie X?”."""
        while self.wyniki_szukaj_l.count():
            it = self.wyniki_szukaj_l.takeAt(0)
            w = it.widget()
            if w: w.deleteLater()

        tekst = (tekst or "").strip().lower()
        if len(tekst) < 2:
            self.wyniki_szukaj_scroll.setVisible(False)
            self.lbl_szukaj_podpowiedz.setVisible(True)
            self.lbl_szukaj_podpowiedz.setText("Wpisz min. 2 znaki, aby wyszukać w całym planie.")
            return

        dzis = datetime.date.today()
        nadchodzace, minione = [], []
        for d in (self._plan or {}).get("dni", []):
            for wz in d.wizyty:
                hay = " ".join([wz.nazwa or "", wz.adres or "", wz.miasto or "", wz.siec or ""]).lower()
                if tekst in hay:
                    (nadchodzace if d.data >= dzis else minione).append((d.data, wz))
        nadchodzace.sort(key=lambda x: x[0])
        minione.sort(key=lambda x: x[0], reverse=True)

        self.lbl_szukaj_podpowiedz.setVisible(False)
        self.wyniki_szukaj_scroll.setVisible(True)

        if not nadchodzace and not minione:
            info = QLabel("Brak wyników.")
            info.setObjectName("WynikBrak")
            self.wyniki_szukaj_l.addWidget(info)
            self._stylizuj_wynik_brak(info)
            return

        LIMIT = 25
        if nadchodzace:
            self._dodaj_naglowek_wynikow("Nadchodzące")
            for data, wz in nadchodzace[:LIMIT]:
                self.wyniki_szukaj_l.addWidget(self._wiersz_wyniku(data, wz, dzis))
        if minione:
            self._dodaj_naglowek_wynikow("Minione")
            for data, wz in minione[:LIMIT]:
                self.wyniki_szukaj_l.addWidget(self._wiersz_wyniku(data, wz, dzis))

    def _dodaj_naglowek_wynikow(self, tekst):
        l = QLabel(tekst); l.setObjectName("WynikNaglowek")
        self.wyniki_szukaj_l.addWidget(l)
        akc = "#00F0FF" if self.is_dark else "#0D9488"
        l.setStyleSheet(f"#WynikNaglowek {{ color:{akc}; font-family:'Segoe UI'; font-size:10px; "
                        f"font-weight:800; letter-spacing:1px; background:transparent; }}")

    def _stylizuj_wynik_brak(self, lbl):
        mut = "#94A3B8" if self.is_dark else "#64748B"
        lbl.setStyleSheet(f"#WynikBrak {{ color:{mut}; font-family:'Segoe UI'; font-size:12px; background:transparent; }}")

    def _wiersz_wyniku(self, data, wz, dzis):
        """Jeden klikalny wynik wyszukiwania — klik przenosi do tego dnia."""
        akc = "#00F0FF" if self.is_dark else "#0D9488"
        txt = "#F8FAFC" if self.is_dark else "#0F172A"
        ramka = "rgba(0,240,255,0.20)" if self.is_dark else "rgba(13,148,136,0.22)"
        hover = "rgba(0,240,255,0.10)" if self.is_dark else "rgba(13,148,136,0.08)"

        b = QPushButton()
        b.setObjectName("WierszWyniku")
        b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        b.setFixedHeight(44)
        dni_roznica = (data - dzis).days
        if dni_roznica == 0:
            kiedy = "DZIŚ"
        elif dni_roznica == 1:
            kiedy = "jutro"
        elif 0 < dni_roznica <= 13:
            kiedy = f"za {dni_roznica} dni"
        else:
            kiedy = data.strftime("%d.%m.%Y")
        etykieta = f"{data.strftime('%d.%m.%Y')} ({kiedy})   •   {wz.nazwa or wz.miasto}"
        if wz.adres:
            etykieta += f"   ·   {wz.adres}"
        b.setText("  " + etykieta)
        b.setStyleSheet(
            f"QPushButton#WierszWyniku {{ text-align:left; color:{txt}; background:transparent; "
            f"border:1px solid {ramka}; border-radius:8px; padding:0 10px; "
            f"font-family:'Segoe UI'; font-size:11px; }} "
            f"QPushButton#WierszWyniku:hover {{ background:{hover}; border-color:{akc}; }}")
        b.clicked.connect(lambda: self._skocz_do_wyniku(data))
        return b

    def _skocz_do_wyniku(self, data):
        """Przenosi widok na dzień z wyniku wyszukiwania."""
        mm = self._miesiace()
        for im, m in enumerate(mm):
            for idd, d in enumerate(m["dni"]):
                if d.data == data:
                    self._mies_idx = im
                    self._dzien_idx = idd
                    self.btn_szukaj.setChecked(False)
                    self._przelacz_szukaj(False)
                    self._ustaw_widok_bez_skoku("dzien")
                    return

    def _liczba_tygodni(self):
        mies = self._biezacy_miesiac()
        if not mies:
            return 1
        dni_robocze = pobierz_dni_robocze(mies["rok"], mies["miesiac"])
        tygodnie = set(_tydzien_numer(d.data, dni_robocze) for d in mies["dni"])
        return max(tygodnie) if tygodnie else 1

    def _czysc_tresc(self):
        while self.tresc_l.count():
            it = self.tresc_l.takeAt(0)
            w = it.widget()
            if w: w.deleteLater()

    def _karta_dnia(self, dzien, pokaz_wizyty=True, pokaz_pierscien=True):
        """Buduje kartę jednego dnia planu — z paskiem postępu i możliwością
        odznaczania wizyt jako odwiedzone (zapisywane na dysk). Pierścień w
        nagłówku (jak na ekranie powitalnym) — ten sam język 'postęp = ring'
        wraca też w Torze Tygodnia, więc widoki czują się jak jedna rodzina."""
        akc = "#00F0FF" if self.is_dark else "#0D9488"
        zielony = "#00E4A1" if self.is_dark else "#059669"
        txt_mut = "#94A3B8" if self.is_dark else "#64748B"
        karta_bg = "rgba(11,19,32,0.75)" if self.is_dark else "rgba(255,255,255,0.9)"
        ramka = "rgba(0,240,255,0.22)" if self.is_dark else "rgba(13,148,136,0.25)"

        # ile wizyt odwiedzonych tego dnia
        zrobione = sum(1 for w in dzien.wizyty if czy_odwiedzona(dzien.data, w.adres or w.nazwa))
        wszystkie = len(dzien.wizyty)
        komplet = wszystkie > 0 and zrobione == wszystkie

        box = QFrame(); box.setObjectName("PlanDzien")
        # dzień w pełni zrobiony — zielona ramka
        br = zielony if komplet else ramka
        box.setStyleSheet(f"#PlanDzien {{ background:{karta_bg}; border:1px solid {br}; border-radius:12px; }}")
        bl = QHBoxLayout(box); bl.setContentsMargins(16, 14, 16, 14); bl.setSpacing(14)

        # pierścień postępu — hero widoku Dnia (pomijany w rozwiniętej karcie
        # Tygodnia, gdzie ten sam ring już świeci w Torze powyżej)
        if pokaz_pierscien and wszystkie:
            ring = PierscienPostepu(box, self.is_dark, rozmiar=56)
            ring.ustaw(zrobione, wszystkie)
            bl.addWidget(ring, 0, Qt.AlignmentFlag.AlignTop)

        prawa = QVBoxLayout(); prawa.setSpacing(8)
        # nagłówek dnia + znacznik ukończenia
        gora = QHBoxLayout()
        dzien_tyg = self._DNI_PL[dzien.data.weekday()]
        jest_dzis = (dzien.data == datetime.date.today())
        etykieta = f"{dzien_tyg}, {dzien.data.strftime('%d.%m.%Y')}"
        if jest_dzis:
            etykieta = "DZIŚ  •  " + etykieta
        naglo = QLabel(etykieta)
        kol_naglo = zielony if komplet else akc
        naglo.setStyleSheet(f"color:{kol_naglo}; font-family:'Segoe UI'; font-size:14px; font-weight:800; background:transparent; border:none;")
        gora.addWidget(naglo)
        gora.addStretch()
        if wszystkie:
            znacznik = QLabel("✓ dzień zrobiony" if komplet else f"{zrobione}/{wszystkie} zrobione")
            kol_zn = zielony if komplet else txt_mut
            znacznik.setStyleSheet(f"color:{kol_zn}; font-family:'Segoe UI'; font-size:11px; font-weight:700; background:transparent; border:none;")
            gora.addWidget(znacznik)
        prawa.addLayout(gora)

        h_godz = int(dzien.minuty // 60); m_godz = int(dzien.minuty % 60)
        meta = QLabel(f"{wszystkie} wizyt  •  {dzien.km:.0f} km  •  ok. {h_godz}h {m_godz}min")
        meta.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:12px; background:transparent; border:none;")
        prawa.addWidget(meta)

        # pasek postępu dnia (szczegół pod nagłówkiem — ring daje ogólny stan,
        # pasek zostaje widoczny też gdy się przewinie niżej listę wizyt)
        if wszystkie:
            pasek = PasekPostepuDnia(zrobione, wszystkie, self.is_dark)
            prawa.addWidget(pasek)
        bl.addLayout(prawa, 1)

        if pokaz_wizyty:
            # ŚLAD DNIA — wizyty jako węzły na trasie. Pierwsza nieodhaczona
            # dostaje pulsujący węzeł ("tu jesteś").
            idx_nastepny = None
            for i, wz in enumerate(dzien.wizyty):
                if not czy_odwiedzona(dzien.data, wz.adres or wz.nazwa):
                    idx_nastepny = i
                    break
            for i, wz in enumerate(dzien.wizyty):
                adres_w = wz.adres or wz.nazwa
                odw = czy_odwiedzona(dzien.data, adres_w)
                wiersz = self._wiersz_wizyty(
                    dzien, wz, i + 1, odw,
                    pierwszy=(i == 0), ostatni=(i == len(dzien.wizyty) - 1),
                    nastepny=(i == idx_nastepny and mozna_odhaczyc(dzien.data)))
                prawa.addWidget(wiersz)
        return box

    def _wiersz_wizyty(self, dzien, wizyta, numer, odwiedzona, pierwszy=False,
                       ostatni=False, nastepny=False):
        """Wizyta jako węzeł na 'Śladzie Dnia': kropka na trasie, etykieta,
        pieczątka z godziną (po odhaczeniu) i karteczka z notatką."""
        zielony = "#00E4A1" if self.is_dark else "#059669"
        txt = "#F8FAFC" if self.is_dark else "#0F172A"
        txt_mut = "#94A3B8" if self.is_dark else "#64748B"
        akc = "#00F0FF" if self.is_dark else "#0D9488"

        adres_w = wizyta.adres or wizyta.nazwa
        wpis = wpis_wizyty(dzien.data, adres_w)
        czas = wpis.get("czas", "")
        notatka = wpis.get("notatka", "")
        przyszlosc = not mozna_odhaczyc(dzien.data)

        kontener = QFrame(); kontener.setObjectName("WizytaKontener")
        kontener.setStyleSheet("#WizytaKontener { background:transparent; border:none; }")
        kl = QHBoxLayout(kontener); kl.setContentsMargins(0, 0, 0, 0); kl.setSpacing(0)

        # --- lewa kolumna: węzeł na szlaku ---
        wezel = WezelSladu(zrobiony=odwiedzona, nastepny=nastepny,
                           pierwszy=pierwszy, ostatni=ostatni, is_dark=self.is_dark)
        kl.addWidget(wezel)

        # --- prawa kolumna: treść wizyty ---
        prawa = QVBoxLayout(); prawa.setContentsMargins(6, 6, 8, 6); prawa.setSpacing(4)
        gora = QHBoxLayout(); gora.setSpacing(8)

        ety = wizyta.nazwa if wizyta.nazwa else wizyta.miasto
        lbl = QLabel(ety)
        if odwiedzona:
            lbl.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:12px; font-weight:600; background:transparent; border:none;")
        else:
            lbl.setStyleSheet(f"color:{txt}; font-family:'Segoe UI'; font-size:12px; font-weight:600; background:transparent; border:none;")
        gora.addWidget(lbl)
        gora.addStretch()

        if odwiedzona and czas:
            # PIECZĄTKA z godziną — animowana tylko przy świeżym odhaczeniu
            swieza = (self._swiezo_odhaczona == klucz_wizyty(dzien.data, adres_w))
            gora.addWidget(Pieczatka(czas, self.is_dark, animuj=swieza))
            if swieza:
                self._swiezo_odhaczona = None
        elif przyszlosc:
            # blokada — dziennik dokumentuje przeszłość, nie przyszłość
            zam = QLabel("🔒 jeszcze przed nami")
            zam.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:10px; font-style:italic; background:transparent; border:none;")
            gora.addWidget(zam)
        else:
            # przycisk odhaczenia (dostępny tylko dla dni ≤ dziś)
            b = QPushButton("Odhacz wizytę")
            b.setFixedHeight(26); b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            b.setStyleSheet(
                f"QPushButton {{ color:{akc}; background:transparent; border:1px solid {akc}; "
                f"border-radius:13px; padding:0 12px; font-family:'Segoe UI'; font-size:10px; font-weight:700; }} "
                f"QPushButton:hover {{ background:{akc}; color:#04121A; }}")
            b.clicked.connect(lambda _=None, d=dzien.data, a=adres_w: self._odhacz(d, a))
            gora.addWidget(b)
        prawa.addLayout(gora)

        # sieć + adres (drobnym drukiem)
        pod = wizyta.siec or ""
        if wizyta.adres and wizyta.adres != ety:
            pod = f"{pod} · {wizyta.adres}" if pod else wizyta.adres
        if pod:
            l2 = QLabel(pod)
            l2.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:10px; background:transparent; border:none;")
            prawa.addWidget(l2)

        # --- karteczka z notatką (tylko dla odhaczonych) ---
        if odwiedzona:
            if notatka:
                kart = QLabel(f"📝  {notatka}")
                kart.setWordWrap(True)
                tlo_k = "rgba(0,228,161,0.10)" if self.is_dark else "rgba(5,150,105,0.10)"
                kart.setStyleSheet(
                    f"color:{txt}; background:{tlo_k}; border-left:3px solid {zielony}; "
                    f"border-radius:4px; padding:6px 10px; font-family:'Segoe UI'; font-size:11px;")
                kart.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
                kart.mousePressEvent = lambda _e, d=dzien.data, a=adres_w, n=notatka: self._edytuj_notatke(d, a, n)
                prawa.addWidget(kart)
            else:
                bn = QPushButton("＋ dodaj notatkę z wizyty")
                bn.setFixedHeight(24); bn.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
                bn.setStyleSheet(
                    f"QPushButton {{ color:{txt_mut}; background:transparent; border:1px dashed {txt_mut}; "
                    f"border-radius:6px; padding:0 10px; font-family:'Segoe UI'; font-size:10px; text-align:left; }} "
                    f"QPushButton:hover {{ color:{zielony}; border-color:{zielony}; }}")
                bn.clicked.connect(lambda _=None, d=dzien.data, a=adres_w: self._edytuj_notatke(d, a, ""))
                prawa.addWidget(bn, alignment=Qt.AlignmentFlag.AlignLeft)

        kl.addLayout(prawa, 1)
        return kontener

    def _odhacz(self, data, adres):
        """Odhacza wizytę — zapisuje znacznik czasu i uruchamia animację pieczątki."""
        if ustaw_odwiedzona(data, adres, True):
            self._swiezo_odhaczona = klucz_wizyty(data, adres)
            self._przerysuj()

    def _edytuj_notatke(self, data, adres, obecna):
        """Okienko notatki z wizyty — co zastałem, co zrobiłem."""
        dlg = DialogNotatki(self, obecna, self.is_dark)
        wynik = dlg.pokaz()
        if wynik is None:
            return
        if wynik == "__USUN__":
            ustaw_odwiedzona(data, adres, False)     # cofnięcie odhaczenia
        else:
            ustaw_notatke(data, adres, wynik)
        self._przerysuj()

    def _przerysuj(self):
        self._czysc_tresc()
        self.update_theme(self.is_dark)
        mm = self._miesiace()
        if not mm:
            pusto = QLabel("Brak zaplanowanych wizyt.\nWczytaj plik z punktami w Planerze i wygeneruj plan.")
            pusto.setAlignment(Qt.AlignmentFlag.AlignCenter)
            pusto.setStyleSheet(f"color:{'#94A3B8' if self.is_dark else '#64748B'}; font-family:'Segoe UI'; font-size:14px; background:transparent;")
            self.tresc_l.addWidget(pusto)
            self.lbl_zakres.setText(""); self.lbl_podsum.setText("")
            self.btn_prev.setVisible(False); self.btn_next.setVisible(False)
            return
        mies = self._biezacy_miesiac()
        dni = mies["dni"]
        mies_nazwa = self._MIES_PL[mies["miesiac"]].capitalize()
        wiele = len(mm) > 1
        etyk_mies = f"{mies_nazwa} {mies['rok']}"
        if wiele:
            etyk_mies += f"  ({self._mies_idx + 1}/{len(mm)})"

        if self._widok == "dzien":
            self.btn_prev.setVisible(True); self.btn_next.setVisible(True)
            self.btn_tryb_trasy.setVisible(True)
            for w in self._warstwy_row_widgets: w.setVisible(False)
            self._dzien_idx = max(0, min(len(dni) - 1, self._dzien_idx))
            d = dni[self._dzien_idx]
            self.lbl_zakres.setText(f"{etyk_mies} — dzień {self._dzien_idx + 1}/{len(dni)}")
            self.tresc_l.addWidget(self._karta_dnia(d, pokaz_wizyty=True))

        elif self._widok == "tydzien":
            self.btn_prev.setVisible(True); self.btn_next.setVisible(True)
            self.btn_tryb_trasy.setVisible(False)
            for w in self._warstwy_row_widgets: w.setVisible(False)
            dni_robocze = pobierz_dni_robocze(mies["rok"], mies["miesiac"])
            maxtyg = self._liczba_tygodni()
            self._tydzien_idx = max(0, min(maxtyg - 1, self._tydzien_idx))
            nr = self._tydzien_idx + 1
            self.lbl_zakres.setText(f"{etyk_mies} — tydzień {nr}/{maxtyg}")
            dni_tyg = [d for d in dni if _tydzien_numer(d.data, dni_robocze) == nr]
            if dni_tyg:
                dzis = datetime.date.today()
                daty_tyg = [d.data for d in dni_tyg]
                if self._tydzien_aktywny_dzien not in daty_tyg:
                    # domyślnie: dziś (jeśli w tym tygodniu) → pierwszy nieukończony → pierwszy dzień
                    aktywny = dzis if dzis in daty_tyg else None
                    if aktywny is None:
                        for d in dni_tyg:
                            if any(not czy_odwiedzona(d.data, wz.adres or wz.nazwa) for wz in d.wizyty):
                                aktywny = d.data
                                break
                    self._tydzien_aktywny_dzien = aktywny or dni_tyg[0].data

                tor = TorTygodnia(is_dark=self.is_dark, on_klik=self._wybierz_dzien_tygodnia)
                tor.ustaw_dni(dni_tyg, self._tydzien_aktywny_dzien)
                self.tresc_l.addWidget(tor)

                wybrany = next((d for d in dni_tyg if d.data == self._tydzien_aktywny_dzien), dni_tyg[0])
                self.tresc_l.addWidget(self._karta_dnia(wybrany, pokaz_wizyty=True, pokaz_pierscien=False))
            else:
                info = QLabel("Brak wizyt w tym tygodniu.")
                info.setAlignment(Qt.AlignmentFlag.AlignCenter)
                info.setStyleSheet(f"color:{'#94A3B8' if self.is_dark else '#64748B'}; font-family:'Segoe UI'; font-size:13px; background:transparent;")
                self.tresc_l.addWidget(info)

        else:  # miesiac — JEDYNY kalendarz w programie: wizyty + delegacje + notatki
            self.btn_prev.setVisible(wiele); self.btn_next.setVisible(wiele)
            self.btn_tryb_trasy.setVisible(False)
            for w in self._warstwy_row_widgets: w.setVisible(True)
            self.lbl_zakres.setText(etyk_mies)
            dni_planu_mapa = {d.data.day: d for d in dni}
            delegacje = {}
            if self._on_dane_uzytkownika:
                imie, pesel = self._on_dane_uzytkownika()
                if imie and pesel:
                    delegacje = kalendarz_z_historii(imie, pesel)
            siatka = SiatkaMiesiacaPlan(is_dark=self.is_dark, on_klik=self._klik_dzien_miesiaca,
                                       on_notatka=self._otworz_notatka_dnia)
            siatka.setMinimumHeight(430)
            siatka.ustaw_miesiac(mies["rok"], mies["miesiac"], dni_planu_mapa)
            siatka.ustaw_delegacje(delegacje)
            siatka.ustaw_warstwy(wizyty=(self._warstwa_kalendarza in ("all", "wizyty")),
                                delegacje=(self._warstwa_kalendarza in ("all", "delegacje")))
            self.tresc_l.addWidget(siatka)

        # podsumowanie (zbiorcze — całość planu) + POSTĘP odwiedzin
        p = self._plan
        # policz odwiedzone w całym planie
        odw = 0
        for d in p.get("dni", []):
            for wz in d.wizyty:
                if czy_odwiedzona(d.data, wz.adres or wz.nazwa):
                    odw += 1
        czesci = [f"{p['suma_wizyt']} wizyt", f"{len(mm)} {'miesiąc' if len(mm)==1 else 'miesięcy'}", f"{p['suma_km']:.0f} km łącznie"]
        if p.get("cykliczny"):
            dom = p.get("cykl_domyslny", 4)
            czesti_cykl = f"🔁 cykl co {dom} tyg."
            cs = p.get("cykle_sieci") or {}
            if cs:
                czesti_cykl += f" (+{len(cs)} własnych)"
            czesci.insert(1, czesti_cykl)
        if p.get("nierozplanowane"):
            czesci.append(f"{len(p['nierozplanowane'])} poza planem")
        podsum = "   •   ".join(czesci)
        if odw:
            proc = round(odw / max(p['suma_wizyt'], 1) * 100)
            podsum = f"✓ {odw} zrobionych ({proc}%)   •   " + podsum
        self.lbl_podsum.setText(podsum)
        # przycisk "bez lokalizacji" — KLIKALNY, pokazuje KTÓRE konkretnie adresy
        bez_poz = p.get("bez_pozycji") or []
        if bez_poz:
            self.btn_bez_lokalizacji.setText(f"⚠  {len(bez_poz)} bez lokalizacji — zobacz listę")
            self.btn_bez_lokalizacji.setVisible(True)
        else:
            self.btn_bez_lokalizacji.setVisible(False)
        self._odswiez_baner()
    def _pokaz_bez_lokalizacji(self):
        """Pokazuje KTÓRE konkretnie adresy się nie zgeokodowały — zamiast
        samej liczby. Pozwala je od razu usunąć z listy punktów."""
        bez_poz = (self._plan or {}).get("bez_pozycji") or []
        if not bez_poz:
            return
        dlg = DialogBezLokalizacji(self, bez_poz, is_dark=self.is_dark)
        usuniete = dlg.pokaz()
        if not usuniete:
            return
        adresy_do_usuniecia = {(p.adres or p.nazwa or "").strip().lower() for p in usuniete}
        punkty = wczytaj_punkty()
        nowe = [p for p in punkty if (p.get("adres", "") or "").strip().lower() not in adresy_do_usuniecia]
        zapisz_punkty(nowe)
        # usuń też z bieżącego planu (żeby znikły z podsumowania od razu)
        if self._plan:
            self._plan["bez_pozycji"] = [
                p for p in self._plan.get("bez_pozycji", [])
                if (p.adres or p.nazwa or "").strip().lower() not in adresy_do_usuniecia]
        if self._on_punkty_zmienione:
            self._on_punkty_zmienione()
        self._przerysuj()
        if self._on_toast:
            n = len(usuniete)
            self._on_toast("Punkty usunięte", f"Usunięto {n} {'punkt' if n==1 else 'punkty' if 2<=n<=4 else 'punktów'} bez lokalizacji z listy.")

    def _klik_eksport(self):
        """Eksportuje CAŁY wczytany plan (nie tylko widoczny miesiąc) do
        Excela (plan + dziennik w jednym pliku) albo PDF (drukowalny wykaz).
        Format wybiera się przez filtr w oknie zapisu."""
        if not self._plan or not self._plan.get("dni"):
            if self._on_toast:
                self._on_toast("Brak planu", "Nie ma czego eksportować — ułóż najpierw plan wizyt.", False)
            return
        domyslna = f"Plan_Wizyt_{datetime.date.today().strftime('%Y-%m-%d')}.xlsx"
        sciezka, filtr = QFileDialog.getSaveFileName(
            self, "Eksportuj plan wizyt", domyslna,
            "Excel — plan i dziennik (*.xlsx);;PDF — drukowalny wykaz (*.pdf)")
        if not sciezka:
            return
        chce_pdf = sciezka.lower().endswith(".pdf") or "PDF" in filtr
        if chce_pdf and not sciezka.lower().endswith(".pdf"):
            sciezka += ".pdf"
        elif not chce_pdf and not sciezka.lower().endswith(".xlsx"):
            sciezka += ".xlsx"
        try:
            if chce_pdf:
                eksportuj_plan_pdf(self._plan, sciezka)
            else:
                eksportuj_plan_excel(self._plan, sciezka)
            if self._on_toast:
                self._on_toast("Wyeksportowano", f"Zapisano w:\n{sciezka}")
        except Exception as e:
            if self._on_toast:
                self._on_toast("Błąd eksportu", f"Nie udało się zapisać pliku: {e}", False)

    def _wybierz_dzien_tygodnia(self, data):
        """Klik w przystanek Toru Tygodnia — rozwija ten dzień poniżej,
        bez przeskakiwania do widoku Dzień."""
        self._tydzien_aktywny_dzien = data
        self._przerysuj()

    def _ustaw_warstwe_kalendarza(self, tryb):
        self.btn_w_all.setChecked(tryb == "all")
        self.btn_w_wiz.setChecked(tryb == "wizyty")
        self.btn_w_del.setChecked(tryb == "delegacje")
        self._warstwa_kalendarza = tryb
        self._przerysuj()

    def _otworz_notatka_dnia(self, data):
        """Klik w kropkę notatki na siatce miesiąca — dialog notatki OGÓLNEJ
        (nie powiązanej z żadnym silnikiem) + 'dzień wolny'. Zaznaczenie dnia
        wolnego sprawia, że planer wizyt realnie go pomija przy układaniu tras."""
        wpis = notatka_dnia(data.isoformat())
        dlg = DialogNotatkaDnia(self, data, wpis.get("notatka", ""), wpis.get("wolne", False), is_dark=self.is_dark)
        wynik = dlg.pokaz()
        if wynik is None:
            return
        notatka, wolne = wynik
        ustaw_notatke_dnia(data.isoformat(), notatka, wolne)
        self._przerysuj()

    def _klik_dzien_miesiaca(self, dzien_liczba):
        """Klik w kafelek siatki miesiąca — przenosi wprost do Śladu Dnia."""
        mies = self._biezacy_miesiac()
        for idd, d in enumerate(mies["dni"]):
            if d.data.day == dzien_liczba:
                self._dzien_idx = idd
                self._ustaw_widok_bez_skoku("dzien")
                return

    def _klik_mapa(self):
        if self._on_mapa and self._plan and self._plan.get("dni"):
            self._on_mapa(self._plan)

    def update_theme(self, is_dark):
        self.is_dark = is_dark
        if is_dark:
            tlo="rgba(3,7,14,0.90)"; karta="rgba(11,19,32,0.98)"; ramka="rgba(0,240,255,0.25)"
            akc="#00F0FF"; txt="#F8FAFC"; txt_mut="#94A3B8"; seg_bg="rgba(255,255,255,0.06)"; seg_br="rgba(255,255,255,0.15)"
        else:
            tlo="rgba(226,232,240,0.94)"; karta="#FFFFFF"; ramka="rgba(13,148,136,0.30)"
            akc="#0D9488"; txt="#0F172A"; txt_mut="#475569"; seg_bg="rgba(15,23,42,0.05)"; seg_br="rgba(15,23,42,0.15)"
        self.setStyleSheet(f"PlanWizytOverlay {{ background-color:{tlo}; }}")
        self.karta.setStyleSheet(f"#PlanKarta {{ background-color:{karta}; border:1px solid {ramka}; border-radius:18px; }}")
        self.tytul.setStyleSheet(f"color:{akc}; font-family:'Segoe UI'; font-size:22px; font-weight:800; background:transparent; border:none;")
        self.podtytul.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:12px; background:transparent; border:none;")
        self.lbl_zakres.setStyleSheet(f"color:{txt}; font-family:'Segoe UI'; font-size:13px; font-weight:700; background:transparent; border:none;")
        self.lbl_podsum.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:12px; background:transparent; border:none;")
        bursz = "#FBBF24" if is_dark else "#B45309"
        self.btn_bez_lokalizacji.setStyleSheet(
            f"QPushButton {{ color:{bursz}; background:transparent; border:1px solid {bursz}; "
            f"border-radius:8px; padding:0 12px; font-family:'Segoe UI'; font-size:10px; font-weight:700; }} "
            f"QPushButton:hover {{ background:rgba(251,191,36,0.12); }}")
        # segmentowane przyciski
        for b in (self.btn_dzien, self.btn_tydzien, self.btn_miesiac):
            akt = b.isChecked()
            if akt:
                b.setStyleSheet(f"QPushButton {{ color:#04121A; font-weight:800; border:none; border-radius:10px; padding:0 20px; background: qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc}, stop:1 {'#00E4A1' if is_dark else '#059669'}); }}")
            else:
                b.setStyleSheet(f"QPushButton {{ color:{txt}; font-weight:600; border:1px solid {seg_br}; border-radius:10px; padding:0 20px; background:{seg_bg}; }} QPushButton:hover {{ background:{seg_br}; }}")
        nav_css=(f"QPushButton {{ color:{txt}; background:{seg_bg}; border:1px solid {seg_br}; border-radius:9px; font-size:16px; font-weight:bold; }} QPushButton:hover {{ background:{seg_br}; }}")
        self.btn_prev.setStyleSheet(nav_css); self.btn_next.setStyleSheet(nav_css)
        self.btn_mapa.setStyleSheet(f"QPushButton {{ color:{'#04121A' if is_dark else '#FFFFFF'}; font-weight:700; border:none; border-radius:10px; padding:0 16px; background:{akc}; }} QPushButton:hover {{ opacity:0.9; }}")
        self.btn_x.setStyleSheet(f"QPushButton {{ color:{txt_mut}; background:transparent; border:none; font-size:18px; }} QPushButton:hover {{ color:{txt}; }}")
        zielony_tt = "#00E4A1" if is_dark else "#059669"
        self.btn_szukaj.setStyleSheet(
            f"QPushButton {{ color:{txt}; background:{seg_bg}; border:1px solid {seg_br}; border-radius:9px; font-size:15px; }} "
            f"QPushButton:hover {{ background:{seg_br}; }} "
            f"QPushButton:checked {{ background:{akc}; color:#04121A; border-color:{akc}; }}")
        self.btn_tryb_trasy.setStyleSheet(
            f"QPushButton {{ color:#04121A; font-weight:800; border:none; border-radius:10px; padding:0 16px; "
            f"background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {zielony_tt}, stop:1 {akc}); }} "
            f"QPushButton:hover {{ background:qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 {akc}, stop:1 {zielony_tt}); }}")
        self.btn_eksport.setStyleSheet(
            f"QPushButton {{ color:{txt}; background:{seg_bg}; border:1px solid {seg_br}; border-radius:10px; "
            f"padding:0 16px; font-family:'Segoe UI'; font-size:12px; font-weight:700; }} "
            f"QPushButton:hover {{ border-color:{akc}; color:{akc}; }}")
        self.lbl_warstwy.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
        self.lbl_legenda_mies.setStyleSheet(f"color:{txt_mut}; font-family:'Segoe UI'; font-size:10px; background:transparent;")
        for b in (self.btn_w_all, self.btn_w_wiz, self.btn_w_del):
            b.setStyleSheet(
                f"QPushButton {{ color:{txt_mut}; background:transparent; border:1px solid {seg_br}; border-radius:8px; "
                f"padding:0 10px; font-family:'Segoe UI'; font-size:10px; font-weight:600; }} "
                f"QPushButton:hover {{ color:{akc}; border-color:{akc}; }} "
                f"QPushButton:checked {{ background:{akc}; color:#04121A; border-color:{akc}; font-weight:800; }}")
        panel_bg = "rgba(0,240,255,0.05)" if is_dark else "rgba(13,148,136,0.05)"
        self.panel_szukaj.setStyleSheet(f"#PlanPanelSzukaj {{ background:{panel_bg}; border:1px solid {seg_br}; border-radius:12px; }}")
        self.pole_szukaj_plan.setStyleSheet(
            f"QLineEdit {{ background:{'rgba(5,10,20,0.6)' if is_dark else 'rgba(255,255,255,0.9)'}; "
            f"border:1px solid {seg_br}; border-radius:9px; padding:0 12px; color:{txt}; "
            f"font-family:'Segoe UI'; font-size:12px; }} QLineEdit:focus {{ border-color:{akc}; }}")
        self.lbl_szukaj_podpowiedz.setStyleSheet(
            f"color:{txt_mut}; font-family:'Segoe UI'; font-size:11px; background:transparent;")
        uchwyt_w = "rgba(0,240,255,0.35)" if is_dark else "rgba(13,148,136,0.35)"
        self.wyniki_szukaj_scroll.setStyleSheet(
            "#WynikiSzukajScroll { background:transparent; border:none; } "
            f"QScrollBar:vertical {{ background:transparent; width:8px; margin:2px; border:none; }} "
            f"QScrollBar::handle:vertical {{ background:{uchwyt_w}; border-radius:4px; min-height:28px; }} "
            f"QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height:0px; border:none; }} "
            f"QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{ background:transparent; }}")
        uchwyt_p = "rgba(0,240,255,0.35)" if self.is_dark else "rgba(13,148,136,0.35)"
        uchwyt_p_h = "rgba(0,240,255,0.65)" if self.is_dark else "rgba(13,148,136,0.6)"
        self.scroll.setStyleSheet(
            "#PlanScroll { background:transparent; border:none; } "
            f"QScrollBar:vertical {{ background:transparent; width:10px; margin:4px 2px 4px 0; border:none; }} "
            f"QScrollBar::handle:vertical {{ background:{uchwyt_p}; border-radius:5px; min-height:36px; }} "
            f"QScrollBar::handle:vertical:hover {{ background:{uchwyt_p_h}; }} "
            f"QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height:0px; border:none; background:transparent; }} "
            f"QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{ background:transparent; }}")
        self.tresc.setStyleSheet("background:transparent;")
        if self.panel_szukaj.isVisible():
            self._szukaj_w_planie(self.pole_szukaj_plan.text())


class App(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)

        # Ikona aplikacji (pasek zadań / Alt-Tab) z pliku pmt_logo, jeśli jest
        _ikona = znajdz_ikone()
        if _ikona:
            self.setWindowIcon(QIcon(_ikona))

        # --- ROZMIAR OKNA: responsywny do ekranu użytkownika ---
        # Program służy do analizy dużych list (1000+ punktów), więc startuje
        # zmaksymalizowany do DOSTĘPNEGO obszaru ekranu (bez paska zadań).
        # Rozmiar "normalny" (po przywróceniu z maksymalizacji) to ~80% ekranu,
        # ograniczone rozsądnymi widełkami — dzięki temu działa dobrze zarówno
        # na małym laptopie, jak i na dużym monitorze.
        ekran = QApplication.primaryScreen()
        dostepny = ekran.availableGeometry() if ekran else None
        if dostepny is not None:
            szer_norm = max(1000, min(1600, int(dostepny.width() * 0.80)))
            wys_norm = max(620, min(1000, int(dostepny.height() * 0.85)))
            self.resize(szer_norm, wys_norm)
            # geometria "normalna" — do niej wróci przycisk przywracania
            self._geo_przed_maks = QRect(
                dostepny.x() + (dostepny.width() - szer_norm) // 2,
                dostepny.y() + (dostepny.height() - wys_norm) // 2,
                szer_norm, wys_norm)
            # start na pełnym dostępnym obszarze
            self.setGeometry(dostepny)
            self._recznie_zmaks = True
        else:
            self.resize(1120, 700)
        # minimum dopasowane do małych ekranów (nie blokuje laptopów 1366x768)
        self.setMinimumSize(940, 600)
        self.is_dark = True
        
        self.main_container = ImageBackgroundWidget(self)
        self.main_container.setStyleSheet("border-radius: 16px;")
        self.setCentralWidget(self.main_container)
        
        root = QVBoxLayout(self.main_container); root.setContentsMargins(0, 0, 0, 0); root.setSpacing(0)
        self.title_bar = FramelessTitleBar(self); root.addWidget(self.title_bar)
        
        main_body_layout = QHBoxLayout()
        main_body_layout.setContentsMargins(0, 0, 0, 0)
        main_body_layout.setSpacing(0)
        root.addLayout(main_body_layout, 1)
        
        self.sidebar_frame = QFrame()
        self.sidebar_frame.setFixedWidth(68)
        self.sidebar_frame.installEventFilter(self)     # hover → rozwinięcie
        self.sidebar_layout = QVBoxLayout(self.sidebar_frame)
        self.sidebar_layout.setContentsMargins(10, 24, 10, 24)
        self.sidebar_layout.setSpacing(8)
        main_body_layout.addWidget(self.sidebar_frame)
        self._sidebar_przypiety = False    # ⚙️ przypina panel otwarty

        def nav_btn(text):
            b = QPushButton("")
            b._pelna_nazwa = text
            b.setFixedHeight(50); b.setMinimumWidth(230)
            b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            return b

        # --- Grupa główna (nawigacja) — kreatywne nazwy ---
        self.btn_nav_kokpit    = nav_btn("Nowa Wyprawa")       # generator delegacji
        self.btn_nav_plan      = nav_btn("Plan Wizyt")         # bezpośredni dostęp do planu (zawiera scalony kalendarz)
        self.btn_nav_archiwum  = nav_btn("Bilans Miesiąca")    # podsumowanie/archiwum
        self.btn_nav_staty     = nav_btn("Twoja praca")        # statystyki wizyt + delegacji (zawiera dawne "Moje Szlaki")
        # --- Grupa systemowa (pod separatorem) ---
        self.btn_nav_kopia     = nav_btn("Kopia zapasowa")
        self.btn_nav_ustaw     = nav_btn("Ustawienia")
        self.btn_nav_info      = nav_btn("O programie")

        self._nav_defs = [
            (self.btn_nav_kokpit,    "kompas"),
            (self.btn_nav_plan,      "checklista"),
            (self.btn_nav_archiwum,  "wykres"),
            (self.btn_nav_staty,     "trend"),
            (self.btn_nav_kopia,     "tarcza"),
            (self.btn_nav_ustaw,     "zebatka"),
            (self.btn_nav_info,      "info"),
        ]
        self._nav_aktywny = self.btn_nav_kokpit    # dla animowanej belki

        for b in [self.btn_nav_kokpit, self.btn_nav_plan, self.btn_nav_archiwum, self.btn_nav_staty]:
            self.sidebar_layout.addWidget(b)

        self.sidebar_sep = QFrame(); self.sidebar_sep.setFixedHeight(1)
        self.sidebar_layout.addWidget(self.sidebar_sep)

        for b in [self.btn_nav_kopia, self.btn_nav_ustaw, self.btn_nav_info]:
            self.sidebar_layout.addWidget(b)
        self.sidebar_layout.addStretch()

        self.btn_nav_kokpit.clicked.connect(lambda: self._nav_klik(self.btn_nav_kokpit, self._pokaz_planer))
        self.btn_nav_plan.clicked.connect(lambda: self._nav_klik(self.btn_nav_plan, self._pokaz_ostatni_plan))
        self.btn_nav_archiwum.clicked.connect(lambda: self._nav_klik(self.btn_nav_archiwum, self._fokus_kokpit))
        self.btn_nav_staty.clicked.connect(lambda: self._nav_klik(self.btn_nav_staty, self._pokaz_statystyki))
        self.btn_nav_kopia.clicked.connect(lambda: self._nav_klik(self.btn_nav_kopia, self._pokaz_kopia_zapasowa))
        self.btn_nav_ustaw.clicked.connect(lambda: self._nav_klik(self.btn_nav_ustaw, self._pokaz_panel_admina))
        self.btn_nav_info.clicked.connect(self._pokaz_o_programie)

        right_content_container = QWidget()
        self.right_content_container = right_content_container   # ekran powitalny zasłania TYLKO ten obszar (nie menu)
        right_content_container.setStyleSheet("background: transparent;")
        right_content_layout = QVBoxLayout(right_content_container)
        right_content_layout.setContentsMargins(0, 0, 0, 0)
        right_content_layout.setSpacing(0)
        main_body_layout.addWidget(right_content_container, 1)

        # ---- Powiększony topbar z DUŻYM logo PMT (72px) ----
        self.topbar = QFrame()
        self.topbar.setFixedHeight(88)
        tb = QHBoxLayout(self.topbar); tb.setContentsMargins(24, 0, 24, 0); tb.setSpacing(18)

        # Okrągłe logo PMT z mini-kompasem (spójne z ekranem powitalnym)
        self.logo_lbl = LogoKompas(self.topbar, srednica=78, is_dark=self.is_dark)
        tb.addWidget(self.logo_lbl)

        tb_titles = QVBoxLayout(); tb_titles.setSpacing(2); tb_titles.setAlignment(Qt.AlignmentFlag.AlignVCenter)
        self.t1 = QLabel("Planer")
        tb_titles.addWidget(self.t1); tb.addLayout(tb_titles); tb.addStretch()

        self.btn_theme = OutlineButton("🌗 Motyw", self.is_dark, self.topbar)
        self.btn_theme.clicked.connect(self.toggle_theme)
        tb.addWidget(self.btn_theme)

        self.btn_dzwonek = DzwonekPowiadomien(self.topbar)
        self.btn_dzwonek.clicked.connect(self._toggle_panel_powiadomien)
        tb.addWidget(self.btn_dzwonek)

        self.btn_bug = QPushButton("⚠  Zgłoś błąd", self.topbar)
        styl_zglos_blad(self.btn_bug)
        self.btn_bug.clicked.connect(lambda: webbrowser.open("mailto:anowicki@pmt.com.pl"))
        tb.addWidget(self.btn_bug)

        # Wylogowanie zawsze pod ręką — ten sam styl co pozostałe przyciski paska.
        self.btn_wyloguj = QPushButton("⎋  Wyloguj", self.topbar)
        styl_wyloguj(self.btn_wyloguj, self.is_dark)
        self.btn_wyloguj.setToolTip("Wyloguj — program poprosi o login i hasło")
        self.btn_wyloguj.clicked.connect(self._wyloguj_uzytkownika)
        tb.addWidget(self.btn_wyloguj)
        right_content_layout.addWidget(self.topbar)
        
        body = QWidget(); body.setStyleSheet("background-color: transparent;")
        self.body = body    # generator tras — pokazywany TYLKO z "Bilans Miesiąca"
        body_l = QVBoxLayout(body); body_l.setContentsMargins(24, 16, 24, 65); body_l.setSpacing(16)
        
        self.cards_wrap = QWidget()
        cards_layout = QVBoxLayout(self.cards_wrap)
        cards_layout.setContentsMargins(0,0,0,0)
        cards_layout.setSpacing(16)
        
        def field(lbl_text, widget):
            w = QWidget(); w.setStyleSheet("background: transparent;")
            w.setMinimumHeight(56) 
            v = QVBoxLayout(w); v.setContentsMargins(0,0,0,0); v.setSpacing(4)
            l = QLabel(lbl_text)
            l.setProperty("class", "fieldLabel")
            v.addWidget(l); v.addWidget(widget)
            return w, l

        self.card_top_frame = QFrame()
        self.card_top_frame.setMinimumHeight(178)   # zapas na pole adresu (dwa wiersze)
        sh1 = QGraphicsDropShadowEffect(self.card_top_frame); sh1.setBlurRadius(30); sh1.setColor(QColor(0, 0, 0, 80)); sh1.setOffset(0, 8); self.card_top_frame.setGraphicsEffect(sh1)
        cl = QVBoxLayout(self.card_top_frame); cl.setContentsMargins(20, 14, 20, 14); cl.setSpacing(12)
        
        self.ic_user_wrap = QWidget(); self.ic_user_wrap.setStyleSheet("background: transparent;"); l_u = QHBoxLayout(self.ic_user_wrap); l_u.setContentsMargins(0,0,0,0); l_u.setSpacing(12)
        self.ic_user = SvgIconLabel("user", parent=self.ic_user_wrap, size=24)
        self.lbl_u_title = QLabel("DANE PRACOWNIKA")
        self.lbl_u_desc = QLabel("Wprowadź dane pracownika do wygenerowania.")
        tx_u = QVBoxLayout(); tx_u.setSpacing(2); tx_u.addWidget(self.lbl_u_title); tx_u.addWidget(self.lbl_u_desc)
        l_u.addWidget(self.ic_user); l_u.addLayout(tx_u); l_u.addStretch()
        cl.addWidget(self.ic_user_wrap)

        row1_t = QHBoxLayout(); row1_t.setSpacing(16)
        self.e_imie = GrubyKursorEdit(); self.e_imie.setPlaceholderText("np. Jan Kowalski")
        self.si_imie = StyledInput("user", self.e_imie, self.is_dark, self.card_top_frame)
        w_imie, self.l_imie = field("Imię i nazwisko", self.si_imie)
        # Dane osoby biorą się z konta, na które zalogowano program. Pole jest
        # zablokowane — dokument można wystawić tylko na siebie, nie na kolegę.
        _imie_konta = online_imie_uzytkownika()
        if _imie_konta:
            self.e_imie.setText(_imie_konta)
            self.e_imie.setReadOnly(True)
            self.e_imie.setToolTip("Dane z Twojego konta (kod " +
                                   str(online_kod_uzytkownika() or "?") +
                                   ") — nie można ich zmienić.")
            self.l_imie.setText("Imię i nazwisko (z konta)")
        
        self.e_pesel = GrubyKursorEdit(); self.e_pesel.setPlaceholderText("np. 85010112345")
        self.si_pesel = StyledInput("card", self.e_pesel, self.is_dark, self.card_top_frame)
        w_pesel, self.l_pesel = field("PESEL", self.si_pesel)
        row1_t.addWidget(w_imie); row1_t.addWidget(w_pesel); cl.addLayout(row1_t)
        
        row2_t = QHBoxLayout(); row2_t.setSpacing(16)
        self.e_adres = GrubyKursorEdit(); self.e_adres.setPlaceholderText("Ulica lub wieś, nr, 00-000 Miejscowość")
        self.si_adres = StyledInput("home", self.e_adres, self.is_dark, self.card_top_frame)
        w_adres, self.l_adres = field("Adres zamieszkania", self.si_adres)
        
        self.c_stan = QComboBox(); self.c_stan.addItems(["merchandiser", "KR"])
        self.si_stan = StyledInput("briefcase", self.c_stan, self.is_dark, self.card_top_frame)
        w_stan, self.l_stan = field("Stanowisko", self.si_stan)
        row2_t.addWidget(w_adres); row2_t.addWidget(w_stan)
        cl.addLayout(row2_t)
        
        cards_layout.addWidget(self.card_top_frame)
        
        self.card_bot_frame = QFrame()
        self.card_bot_frame.setMinimumHeight(118)   # jeden wiersz — wszystko widoczne bez przewijania

        sh2 = QGraphicsDropShadowEffect(self.card_bot_frame); sh2.setBlurRadius(30); sh2.setColor(QColor(0, 0, 0, 80)); sh2.setOffset(0, 8); self.card_bot_frame.setGraphicsEffect(sh2)
        cr = QVBoxLayout(self.card_bot_frame); cr.setContentsMargins(20, 12, 20, 12); cr.setSpacing(6)
        
        self.ic_map_wrap = QWidget(); self.ic_map_wrap.setStyleSheet("background: transparent;"); l_m = QHBoxLayout(self.ic_map_wrap); l_m.setContentsMargins(0,0,0,0); l_m.setSpacing(12)
        self.ic_map = SvgIconLabel("map", parent=self.ic_map_wrap, size=24)
        self.lbl_m_title = QLabel("PARAMETRY TRASY")
        self.lbl_m_desc = QLabel("Określ parametry finansowe i ramy czasowe.")
        tx_m = QVBoxLayout(); tx_m.setSpacing(2); tx_m.addWidget(self.lbl_m_title); tx_m.addWidget(self.lbl_m_desc)
        l_m.addWidget(self.ic_map); l_m.addLayout(tx_m); l_m.addStretch()
        cr.addWidget(self.ic_map_wrap)

        row1_b = QHBoxLayout(); row1_b.setSpacing(16)
        self.e_kwota = GrubyKursorEdit(); self.e_kwota.setPlaceholderText("np. 1200.00")
        self.si_kwota = StyledInput("card", self.e_kwota, self.is_dark, self.card_bot_frame)
        w_kwota, self.l_kwota = field("Kwota docelowa (PLN)", self.si_kwota)

        self.e_mies = GrubyKursorEdit(); self.e_mies.setPlaceholderText("np. 06.2026")
        # Rozliczamy miesiąc WSTECZ — delegacje wystawia się po jego zakończeniu.
        # Użytkownik może wpisać inny, ale domyślnie nie musi nic zmieniać.
        _teraz = datetime.datetime.now()
        _pierwszy = _teraz.replace(day=1)
        _poprzedni = _pierwszy - datetime.timedelta(days=1)
        self.e_mies.setText(f"{_poprzedni.month:02d}.{_poprzedni.year}")
        self.si_mies = StyledInput("calendar", self.e_mies, self.is_dark, self.card_bot_frame)
        w_mies, self.l_mies = field("Miesiąc rozliczenia", self.si_mies)

        self.c_silnik = QComboBox()
        self.c_silnik.addItems(["poniżej 900 cm³", "powyżej 900 cm³"])
        self.c_silnik.setCurrentIndex(1)   # domyślnie wyższa pojemność (typowe auto)
        self.si_silnik = StyledInput("activity", self.c_silnik, self.is_dark, self.card_bot_frame)
        w_silnik, self.l_silnik = field("Pojemność silnika", self.si_silnik)

        row1_b.addWidget(w_kwota, 2); row1_b.addWidget(w_mies, 2); row1_b.addWidget(w_silnik, 2)
        # Tryb pracy i wyłączanie dni w TYM SAMYM wierszu co kwota i miesiąc —
        # nic nie schodzi poniżej krawędzi okna, więc żadna opcja nie umknie.
        def _kolumna(etykieta_txt, zawartosc):
            kol = QVBoxLayout(); kol.setSpacing(3); kol.setContentsMargins(0, 0, 0, 0)
            lab = QLabel(etykieta_txt)
            lab.setStyleSheet("QLabel { color:%s; font-family:'Segoe UI'; font-size:11px;"
                              " font-weight:700; letter-spacing:0.4px; background:transparent;"
                              " border:none; }" % ("#94A3B8" if self.is_dark else "#475569"))
            kol.addWidget(lab)
            if isinstance(zawartosc, QHBoxLayout):
                kol.addLayout(zawartosc)
            else:
                kol.addWidget(zawartosc)
            kont = QWidget(); kont.setLayout(kol)
            return kont, lab

        self.tryb_wybrany = 0
        self.btn_tryb_tydzien  = QPushButton("Tygodniowy")
        self.btn_tryb_wieczory = QPushButton("Wieczory i weekendy")
        for _b in (self.btn_tryb_tydzien, self.btn_tryb_wieczory):
            _b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
            _b.setFixedHeight(32)
            _b.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        rzad_tryb = QHBoxLayout(); rzad_tryb.setSpacing(6); rzad_tryb.setContentsMargins(0, 0, 0, 0)
        rzad_tryb.addWidget(self.btn_tryb_tydzien); rzad_tryb.addWidget(self.btn_tryb_wieczory)
        w_tryb, self.l_tryb = _kolumna("TRYB PRACY", rzad_tryb)

        self.dni_wylaczone = set()
        self.btn_wylacz_dni = QPushButton("Wyłącz dni")
        self.btn_wylacz_dni.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_wylacz_dni.setFixedHeight(32)
        self.btn_wylacz_dni.clicked.connect(self._otworz_wylaczanie_dni)
        self._odswiez_przycisk_dni()
        w_dni, self.l_dni = _kolumna("DNI BEZ PRACY", self.btn_wylacz_dni)

        # Podtytuł zostaje neutralny — użytkownik nie musi znać godzin ani zasad
        # działania silnika. Szczegóły tylko jako podpowiedź po najechaniu.
        def _tryb_zmieniony(idx):
            self.btn_tryb_tydzien.setToolTip("Trasy w dni robocze, od poniedziałku do piątku")
            self.btn_tryb_wieczory.setToolTip("Trasy po godzinach oraz w soboty")
        def _wybierz_tryb(idx):
            self.tryb_wybrany = idx
            self._odswiez_przyciski_trybu()
            _tryb_zmieniony(idx)
        self.btn_tryb_tydzien.clicked.connect(lambda: _wybierz_tryb(0))
        self.btn_tryb_wieczory.clicked.connect(lambda: _wybierz_tryb(1))
        self._odswiez_przyciski_trybu()
        _tryb_zmieniony(0)

        row1_b.addWidget(w_tryb, 3); row1_b.addWidget(w_dni, 1)
        cr.addLayout(row1_b)

        # --- Tryb pracy: cykl tygodniowy albo wieczory i weekendy ------------

        cards_layout.addWidget(self.card_bot_frame)

        # ---- Formularz (lewa) + Panel Asystenta (prawa) obok siebie ----
        self.body_row = QHBoxLayout(); self.body_row.setContentsMargins(0, 0, 0, 0); self.body_row.setSpacing(16)
        self.body_row.addWidget(self.cards_wrap, 1)
        self.assistant = AssistantPanel(self.is_dark)
        self.body_row.addWidget(self.assistant, 0)
        body_l.addLayout(self.body_row)

        self.bot_card = QFrame()
        self.bot_card.setFixedHeight(78)
        bot = QHBoxLayout(self.bot_card); bot.setContentsMargins(20, 8, 4, 8); bot.setSpacing(12)

        # Oś czasu etapów (zamiast zwykłego paska postępu)
        self.timeline = StageTimeline(self.is_dark, self.bot_card)
        bot.addWidget(self.timeline, 1)

        self.btn = SpinnerButton("Generuj PDF", "download", self.is_dark, self.bot_card); self.btn.setFixedSize(210, 46)
        self.btn.clicked.connect(self._klik_generuj); bot.addWidget(self.btn)

        # status/pct — chowane pod overlayem, ale zostawiamy do metody status()
        self.lbl_status = QLabel("Wprowadź dane aby rozpocząć...")
        self.lbl_pct = QLabel("0%")
        self.status_icon = SvgIconLabel("activity", parent=self.bot_card)
        self.gps_prog = GpsProgressBar(self.is_dark, self.bot_card)
        for hidden in [self.lbl_status, self.lbl_pct, self.status_icon, self.gps_prog]:
            hidden.hide()

        body_l.addWidget(self.bot_card)
        right_content_layout.addWidget(body)

        # Ekran powitalny — w tym samym miejscu layoutu co generator (pod
        # topbarem). Generator i ekran powitalny ZAMIENIAJĄ się: jeden widoczny,
        # drugi ukryty. Dzięki temu topbar (Motyw, dzwonek) jest zawsze wolny,
        # a menu boczne klikalne. Na starcie widać powitalny, generator ukryty.
        self.ekran_powitalny = EkranPowitalny(
            self.right_content_container,
            on_dane_uzytkownika=lambda: (self.e_imie.text().strip(), self.e_pesel.text().strip()))
        right_content_layout.addWidget(self.ekran_powitalny, 1)
        self.ekran_powitalny.hide()
        # kokpit dnia — akcje z karty "Dziś w trasie"
        self.ekran_powitalny.podepnij_akcje(
            on_trasa=self._wejdz_w_trase,
            on_planer=self._pokaz_planer,
            on_tryb_trasy=self._otworz_tryb_trasy)

        self.btn_settings = QPushButton("⚙️", self.main_container)
        self.btn_settings.setFixedSize(76, 76)
        self.btn_settings.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.btn_settings.clicked.connect(self.toggle_sidebar)

        self.grip = QSizeGrip(self.main_container)
        self.toast = ToastNotification(self)

        # Centrum powiadomień: panel rozwijany spod dzwoneczka, pokazujący
        # historię komunikatów (nawet po zniknięciu toastu).
        self.panel_powiadomien = PanelPowiadomien(self.main_container)
        self.panel_powiadomien.podepnij_historie(self.toast.historia)
        self.toast.on_nowe_powiadomienie = self._na_nowe_powiadomienie
        self._powiadomien_nieprzeczytane = 0

        self._last_folder = None
        self._profil_zaproponowany = None    # by nie podpowiadać w kółko

        # Kalendarz Wypraw ZOSTAŁ SCALONY z Planem Wizyt → Miesiąc (jeden
        # kalendarz zamiast dwóch pokazujących te same wizyty). Delegacje,
        # notatki i dni wolne teraz żyją w overlay_plan (patrz niżej).

        # Planer Nowej Wyprawy (nakładka)
        self.overlay_planer = PlanerOverlay(self.main_container,
                                            on_zamknij=self._planer_zamkniety,
                                            on_import=self._planer_import_excel)
        self.overlay_planer._on_zaplanuj = self._zaplanuj_wizyty
        self.overlay_planer._on_ostatni = self._pokaz_ostatni_plan
        self.overlay_planer._on_toast = lambda t, o, ok=True: self.toast.show_toast(t, o, success=ok)
        self.overlay_planer._on_odswiez_plik = self._odswiez_z_pliku
        self.overlay_planer.resize(self.main_container.size())
        # wczytaj listę punktów z poprzedniej sesji (nie trzeba importować Excela od nowa)
        self._ile_punktow = self.overlay_planer.wczytaj_zapisane()
        # wczytaj ostatni plan z dysku (przetrwał zamknięcie programu)
        self._gotowy_plan = wczytaj_plan()
        self.overlay_planer.btn_ostatni.setVisible(self._gotowy_plan is not None)

        # Plan Wizyt — nakładka nowego silnika planera (dzień/tydzień/miesiąc)
        self.overlay_plan = PlanWizytOverlay(self.main_container, on_mapa=self._pokaz_mape_planu)
        self.overlay_plan._on_przenies = self._przenies_zalegle
        self.overlay_plan._on_tryb_trasy = self._otworz_tryb_trasy
        self.overlay_plan._on_toast = lambda t, o, ok=True: self.toast.show_toast(t, o, success=ok)
        self.overlay_plan._on_punkty_zmienione = lambda: self.overlay_planer.wczytaj_zapisane()
        self.overlay_plan._on_dane_uzytkownika = lambda: (self.e_imie.text().strip(), self.e_pesel.text().strip())
        self.overlay_plan.resize(self.main_container.size())
        self.overlay_plan.btn_x.clicked.connect(self._powrot_do_powitalnego)

        # TRYB TRASY — pełnoekranowy widok "na czas jazdy", wywoływany z
        # ekranu powitalnego (karta "Dziś w trasie") albo z Planu Wizyt
        self.overlay_tryb_trasy = TrybTrasyOverlay(self.main_container, on_zamknij=self._zamknij_tryb_trasy)
        self.overlay_tryb_trasy.resize(self.main_container.size())

        # Moje Szlaki ZOSTAŁO SCALONE z "Twoja praca" (StatystykiOverlay,
        # zakładka "Delegacje") — jedna pozycja w menu zamiast dwóch.

        # Panel administratora — wgląd we wszystkich użytkowników (tylko lokalnie)
        self.overlay_admin = PanelAdminaOverlay(self.main_container)
        self.overlay_staty = StatystykiOverlay(
            self.main_container,
            on_dane_uzytkownika=lambda: (self.e_imie.text().strip(), self.e_pesel.text().strip()))
        self.overlay_staty.btn_x.clicked.connect(self._powrot_do_powitalnego)
        self.overlay_admin.resize(self.main_container.size())
        self.overlay_admin.btn_x.clicked.connect(self._powrot_do_powitalnego)

        # Overlay WOW — musi być ostatni żeby był na wierzchu
        self.overlay = GeneratingOverlay(self.main_container)
        self.overlay.resize(self.main_container.size())

        # ---- ŻYWA WALIDACJA: każde pole odświeża panel asystenta ----
        self.e_imie.textChanged.connect(self._analizuj_formularz)
        self.e_imie.editingFinished.connect(self._podpowiedz_profil)
        self.e_pesel.textChanged.connect(self._analizuj_formularz)
        self.e_adres.textChanged.connect(self._analizuj_formularz)
        self.e_kwota.textChanged.connect(self._analizuj_formularz)
        self.e_mies.textChanged.connect(self._analizuj_formularz)

        # Enter w dowolnym polu tekstowym = "Generuj" (jeśli dane są kompletne)
        for pole in [self.e_imie, self.e_pesel, self.e_adres, self.e_kwota, self.e_mies]:
            pole.returnPressed.connect(self._enter_generuj)
        self.c_silnik.currentIndexChanged.connect(self._analizuj_formularz)

        self.apply_theme()

        # Ekran powitalny na starcie — zasłania generator do czasu, aż użytkownik
        # wybierze "Bilans Miesiąca". Uruchamiamy PO apply_theme (zna motyw).
        QTimer.singleShot(0, self._pokaz_ekran_powitalny)
        QTimer.singleShot(50, self._analizuj_formularz)

        # Sprawdź aktualizacje w tle (2s po starcie, żeby nie opóźniać okna)
        self._nowa_wersja = ""
        self._nowa_opis = ""
        QTimer.singleShot(2000, self._start_sprawdzania_aktualizacji)

    def _otworz_tryb_trasy(self, dzien, data):
        """Uruchamia pełnoekranowy Tryb Trasy dla danego dnia. Wywoływane z
        karty „Dziś w trasie” (ekran powitalny) albo z przycisku w Planie
        Wizyt — w obu przypadkach zapamiętujemy, dokąd wrócić po zamknięciu."""
        self._tryb_trasy_powrot = "plan" if self.overlay_plan.isVisible() else "powitalny"
        self.overlay_tryb_trasy.update_theme(self.is_dark)
        self.overlay_tryb_trasy.resize(self.main_container.size())
        self.overlay_tryb_trasy.ustaw_dzien(dzien, data)
        self.overlay_tryb_trasy.raise_()
        self.overlay_tryb_trasy.show()

    def _zamknij_tryb_trasy(self):
        """Zamyka Tryb Trasy i odświeża ekran, spod którego został wywołany —
        żeby pierścień/lista od razu pokazały świeżo odhaczone wizyty."""
        self.overlay_tryb_trasy.hide()
        powrot = getattr(self, "_tryb_trasy_powrot", "powitalny")
        if powrot == "plan" and self.overlay_plan.isVisible():
            self.overlay_plan._przerysuj()
        else:
            self.ekran_powitalny.karta_dzis.odswiez()

    def _wejdz_w_trase(self):
        """Z karty „Dziś w trasie" prosto do Planu Wizyt, otwartego na dziś."""
        plan = getattr(self, "_gotowy_plan", None) or wczytaj_plan()
        if not plan:
            self._pokaz_planer()
            return
        self._gotowy_plan = plan
        self.ekran_powitalny.stop()
        self.overlay_plan.update_theme(self.is_dark)
        self.overlay_plan.resize(self.main_container.size())
        self.overlay_plan.ustaw_plan(plan)      # sam skacze na dziś
        self.overlay_plan.raise_()
        self.overlay_plan.show()

    def _pokaz_ekran_powitalny(self):
        # Na starcie: chowamy generator, pokazujemy ekran powitalny.
        # Layout sam ustawia pozycję (pod topbarem), bez ręcznej geometrii.
        if hasattr(self, "body"):
            self.body.hide()
        self.ekran_powitalny.update_theme(self.is_dark)
        self.ekran_powitalny.show()
        self.ekran_powitalny.start()

    def _start_sprawdzania_aktualizacji(self):
        self._update_thread = UpdateThread()
        self._update_thread.wynik.connect(self._aktualizacja_wynik)
        self._update_thread.start()

    def _aktualizacja_wynik(self, jest_nowsza, wersja, opis):
        """Jest nowa wersja → pokazujemy pełne okno powitania (nie ikonkę).
        Chwila zwłoki, żeby nie nachodzić na ekran powitalny."""
        if not jest_nowsza:
            return
        self._nowa_wersja = wersja
        self._nowa_opis = opis
        QTimer.singleShot(900, self._pokaz_okno_aktualizacji)

    def _pokaz_okno_aktualizacji(self):
        dlg = OknoAktualizacji(
            self, wersja_stara=WERSJA_PROGRAMU, wersja_nowa=self._nowa_wersja,
            opis=self._nowa_opis, is_dark=self.is_dark,
            on_instaluj=self._zainstaluj_aktualizacje)
        dlg.exec()

    def _zainstaluj_aktualizacje(self, pobrany_plik):
        """Uruchamia skrypt podmiany i TWARDO kończy proces.

        Kluczowe: program musi NAPRAWDĘ się zamknąć, inaczej Windows dalej
        trzyma zablokowany plik .exe i podmiana się nie uda. QApplication.quit()
        do tego nie wystarcza — kończy tylko pętlę zdarzeń, a żyjące wątki
        (pobieranie, sprawdzanie wersji) potrafią utrzymać proces przy życiu.
        Dlatego zatrzymujemy wątki i wychodzimy przez os._exit().

        Skrypt podmiany pobieramy ŚWIEŻO z GitHuba (przygotuj_skrypt_aktualizacji) —
        gdyby w jego logice znalazł się kiedyś błąd, poprawka na GitHubie
        zadziała u wszystkich, bez potrzeby wydawania nowej wersji programu.
        """
        try:
            docelowy = sciezka_programu()
            bat, args = przygotuj_skrypt_aktualizacji(pobrany_plik, docelowy, os.getpid())

            # DETACHED_PROCESS — skrypt ma przeżyć śmierć programu.
            # UWAGA: nie łączymy z CREATE_NO_WINDOW (Windows nie pozwala
            # łączyć tych flag — proces po prostu się nie tworzy).
            # CREATE_NEW_CONSOLE — skrypt dostaje własne okno i ŻYJE DALEJ
            # po śmierci programu. Widoczne okno jest celowe: gdy coś pójdzie
            # nie tak, użytkownik zobaczy komunikat zamiast mignięcia CMD.
            if CZY_WINDOWS:
                flagi = getattr(subprocess, "CREATE_NEW_CONSOLE", 0)
                subprocess.Popen(["cmd", "/c", bat] + args,
                                 creationflags=flagi,
                                 close_fds=True)
            else:
                # macOS/Linux: skrypt powłoki odczepiony od procesu programu
                # (start_new_session — przeżyje zamknięcie przez os._exit)
                subprocess.Popen(["/bin/bash", bat] + args,
                                 start_new_session=True,
                                 close_fds=True)
        except Exception as e:
            self.toast.show_toast(
                "Nie udało się zainstalować",
                f"{e}\nSpróbuj pobrać ręcznie ze strony wydania.", success=False)
            return

        # --- twarde zamknięcie: zwalniamy plik .exe ---
        for nazwa in ("_update_thread", "_plan_thread"):
            w = getattr(self, nazwa, None)
            try:
                if w is not None and w.isRunning():
                    w.quit()
                    w.wait(1500)
            except Exception:
                pass
        try:
            QApplication.processEvents()
            QApplication.quit()
        except Exception:
            pass
        os._exit(0)          # bez tego proces potrafi zostać i blokować plik

    def toggle_sidebar(self):
        # ⚙️ "przypina" panel otwarty (klik) — hover nadal działa, ale nie zwija przypiętego
        self._sidebar_przypiety = not self._sidebar_przypiety
        self._rozwin_sidebar(self._sidebar_przypiety)

    def _rozwin_sidebar(self, rozwin):
        docelowa = 250 if rozwin else 68
        if self.sidebar_frame.width() == docelowa:
            return
        for b, _ in self._nav_defs:
            b.setText("   " + b._pelna_nazwa if rozwin else "")
        self.sidebar_anim = QVariantAnimation()
        self.sidebar_anim.setDuration(260)
        self.sidebar_anim.setEasingCurve(QEasingCurve.Type.OutCubic)
        self.sidebar_anim.setStartValue(self.sidebar_frame.width())
        self.sidebar_anim.setEndValue(docelowa)
        self.sidebar_anim.valueChanged.connect(self._animate_sidebar)
        self.sidebar_anim.start()

    def eventFilter(self, obj, event):
        # Hover nad sidebarem → rozwinięcie; opuszczenie → zwinięcie (o ile nie przypięty)
        if obj is self.sidebar_frame:
            if event.type() == QEvent.Type.Enter:
                self._rozwin_sidebar(True)
            elif event.type() == QEvent.Type.Leave:
                if not self._sidebar_przypiety:
                    self._rozwin_sidebar(False)
        return super().eventFilter(obj, event)

    def _nav_klik(self, przycisk, akcja):
        """Ustawia aktywny element (animowana belka) i wykonuje akcję."""
        # Każde wejście w menu opuszcza ekran powitalny (odsłania właściwy widok).
        if hasattr(self, "ekran_powitalny") and self.ekran_powitalny.isVisible():
            self.ekran_powitalny.stop()
        # Generator pokazuje TYLKO "Bilans Miesiąca". Dla pozostałych opcji
        # chowamy go, żeby nie wyjrzał spod zamykanej nakładki.
        if hasattr(self, "body") and przycisk is not self.btn_nav_archiwum:
            self.body.hide()
        self._nav_aktywny = przycisk
        self._maluj_nav_belke()
        akcja()

    def _maluj_nav_belke(self):
        """Podświetla aktywny przycisk nawigacji lewą belką akcentu."""
        akcent = "#00F0FF" if self.is_dark else "#0D9488"
        hover_bg = "rgba(0,240,255,0.07)" if self.is_dark else "rgba(13,148,136,0.08)"
        txt = "#94A3B8" if self.is_dark else "#475569"
        for b, _ in self._nav_defs:
            aktywny = (b is self._nav_aktywny)
            if aktywny:
                b.setStyleSheet(
                    f"QPushButton {{ text-align: left; padding-left: 12px; background: {hover_bg};"
                    f" color: {akcent}; font-family:'Segoe UI'; font-size: 13px; font-weight: 700;"
                    f" border: none; border-left: 3px solid {akcent}; border-radius: 8px; }}")
            else:
                b.setStyleSheet(
                    f"QPushButton {{ text-align: left; padding-left: 15px; background: transparent;"
                    f" color: {txt}; font-family:'Segoe UI'; font-size: 13px; font-weight: 600;"
                    f" border: none; border-radius: 8px; }}"
                    f" QPushButton:hover {{ color: {akcent}; background: {hover_bg}; }}")

    # =====================================================================
    #  ŻYWA ANALIZA FORMULARZA  (AI Asystent + szacunki na bieżąco)
    # =====================================================================
    def _analizuj_formularz(self):
        imie  = self.e_imie.text().strip()
        pesel = self.e_pesel.text().strip()
        adres = self.e_adres.text().strip()
        kwota_s = self.e_kwota.text().strip().replace(' ', '').replace(',', '.')
        mies_s  = self.e_mies.text().strip().replace('-', '.').replace('/', '.').replace(' ', '')

        # PESEL
        pesel_ok = waliduj_pesel(pesel) if pesel else None
        self.assistant.set_check("pesel", pesel_ok)
        self.si_pesel.ustaw_walidacje("ok" if pesel_ok is True else ("err" if pesel_ok is False else None))

        # Adres → kod pocztowy → województwo
        kod_ok, woj = None, None
        adres_d = None
        if adres:
            try:
                adres_d = waliduj_adres(adres)
                kod_ok = True
                woj = rozpoznaj_wojewodztwo(adres_d['kod_pocztowy'])
            except Exception:
                kod_ok = False
        self.assistant.set_check("kod", kod_ok)
        self.assistant.set_check("woj", True if woj else (False if adres else None))
        self.si_adres.ustaw_walidacje("ok" if (kod_ok and woj) else ("err" if kod_ok is False else None))

        # Kwota + dni robocze
        kwota_val, mies_ok, dni_count = None, None, 0
        try:
            if kwota_s:
                kwota_val = float(kwota_s)
        except ValueError:
            kwota_val = None
        try:
            if re.match(r"^\d{2}\.\d{4}$", mies_s):
                mm, yy = map(int, mies_s.split('.'))
                if 1 <= mm <= 12:
                    dni_count = len(pobierz_dni_robocze(yy, mm))
                    mies_ok = True
                else:
                    mies_ok = False
            elif mies_s:
                mies_ok = False
        except Exception:
            mies_ok = False

        # Kwota realna? (mieści się w limicie dni * 260)
        kwota_ok = None
        if kwota_val is not None and kwota_val >= MIN_KWOTA:
            if dni_count > 0:
                kwota_ok = kwota_val <= dni_count * 260.0
            else:
                kwota_ok = True
        elif kwota_val is not None:
            kwota_ok = False
        self.assistant.set_check("kwota", kwota_ok)
        self.assistant.set_check("dni", True if (mies_ok and dni_count > 0) else (False if mies_s else None))
        self.si_kwota.ustaw_walidacje("ok" if kwota_ok is True else ("err" if kwota_ok is False else None))
        self.si_mies.ustaw_walidacje("ok" if (mies_ok and dni_count > 0) else ("err" if mies_ok is False else None))

        # Szacunki (gdy jest kwota)
        if kwota_val and kwota_val >= MIN_KWOTA:
            stawka = 0.89 if self.c_silnik.currentIndex() == 0 else 1.15
            dni_baza = dni_count if dni_count > 0 else 21
            sz = szacuj_delegacje(kwota_val, stawka, dni_baza)
            self.assistant.set_szacunki(sz['km'], sz['dni_wyjazdowe'], sz['km_dzien'], sz['dokumenty'])
        else:
            self.assistant.set_szacunki(0, 0, 0, 0)

        # Gotowość
        wszystko = [pesel_ok, kod_ok, (woj is not None), kwota_ok, (mies_ok and dni_count > 0)]
        if all(x is True for x in wszystko):
            self.assistant.set_gotowy(True)
        else:
            brakuje = []
            if not imie: brakuje.append("imię")
            if pesel_ok is not True: brakuje.append("PESEL")
            if kod_ok is not True: brakuje.append("adres")
            if kwota_ok is not True: brakuje.append("kwota")
            if not (mies_ok and dni_count > 0): brakuje.append("miesiąc")
            self.assistant.set_gotowy(False, "Uzupełnij: " + ", ".join(brakuje) if brakuje else "Uzupełnij dane…")

    def _podpowiedz_profil(self):
        """Po wpisaniu imienia — jeśli znamy tę osobę, podpowiada resztę danych."""
        imie = self.e_imie.text().strip()
        if not imie or imie == self._profil_zaproponowany:
            return
        prof = szukaj_profilu_po_nazwisku(imie)
        if prof and not self.e_pesel.text().strip():
            self._profil_zaproponowany = imie
            self.e_pesel.setText(prof.get("pesel", ""))
            self.e_adres.setText(prof.get("adres", ""))
            idx = self.c_stan.findText(prof.get("stanowisko", ""))
            if idx >= 0: self.c_stan.setCurrentIndex(idx)
            self.c_silnik.setCurrentIndex(prof.get("silnik_idx", 1))
            self.toast.show_toast("Rozpoznano pracownika",
                                  f"Uzupełniłem dane {imie.split()[0]} z ostatniej sesji. Sprawdź i popraw w razie potrzeby.",
                                  success=True)

    def _pokaz_o_programie(self):
        # Test silnika NA ŻYWO — wylicza, ile dni wychodzi dla kontrolnej kwoty.
        # Aktualny silnik dla 2000 zł (Warszawa) daje ~10-13 dni. Jeśli pokaże
        # 3-4 dni, to znaczy, że uruchomiony plik/exe jest STARĄ wersją.
        diag = ""
        try:
            dni = pobierz_dni_robocze(2026, 9)
            t = generuj_trasy(2000, "Warszawa", 52.23, 21.01, "mazowieckie", dni, "85010112345", 1.15)
            osiag = getattr(t, "kwota_osiagnieta", sum(d.suma for d in t))
            stan = "✓ silnik OK" if len(t) >= 7 else "✗ STARA WERSJA silnika!"
            diag = f"\nTest silnika: 2000 zł → {len(t)} dni, {osiag:.0f} zł  ({stan})"
        except Exception:
            diag = ""
        self.toast.show_toast(
            f"PMT Planer — wersja {WERSJA_PROGRAMU}",
            "Generator delegacji z asystentem kontroli i historią rozliczeń.\n"
            f"Silnik: {SYGNATURA_SILNIKA}{diag}",
            success=True)

    def _animate_sidebar(self, width):
        self.sidebar_frame.setFixedWidth(width)
        self.btn_settings.move(width - 38, self.height() - 110)

    # =====================================================================
    #  AKCJE NAWIGACJI
    # =====================================================================
    def _fokus_kokpit(self):
        # Bilans Miesiąca — POKAZUJE generator tras (serce programu). Generator
        # jest domyślnie ukryty; pojawia się tylko tutaj. Zamyka nakładki i
        # ekran powitalny.
        if hasattr(self, "ekran_powitalny"):
            self.ekran_powitalny.stop()
        for nakladka in [getattr(self, "overlay_planer", None),
                         getattr(self, "overlay_admin", None),
                         getattr(self, "overlay_staty", None),
                         getattr(self, "overlay_plan", None)]:
            if nakladka is not None:
                nakladka.hide()
        if hasattr(self, "body"):
            self.body.show()
        self.e_imie.setFocus()

    def _na_nowe_powiadomienie(self, title, desc, success):
        """Każdy nowy toast — zwiększ licznik nieprzeczytanych na dzwoneczku
        i odśwież panel, jeśli akurat jest otwarty."""
        self._powiadomien_nieprzeczytane += 1
        self.btn_dzwonek.ustaw_licznik(self._powiadomien_nieprzeczytane)
        if self.panel_powiadomien.isVisible():
            self.panel_powiadomien.odswiez()

    def _toggle_panel_powiadomien(self):
        if self.panel_powiadomien.isVisible():
            self.panel_powiadomien.hide()
            return
        # przy otwarciu zeruj licznik — powiadomienia zostały "przejrzane"
        self._powiadomien_nieprzeczytane = 0
        self.btn_dzwonek.ustaw_licznik(0)
        self.panel_powiadomien.update_theme(self.is_dark)
        self.panel_powiadomien.odswiez()
        # pozycja: pod dzwoneczkiem, przy prawej krawędzi
        pt = self.btn_dzwonek.mapTo(self.main_container, QPoint(self.btn_dzwonek.width(), self.btn_dzwonek.height()))
        x = pt.x() - self.panel_powiadomien.width()
        y = pt.y() + 6
        self.panel_powiadomien.move(max(8, x), y)
        self.panel_powiadomien.raise_()
        self.panel_powiadomien.show()

    def _pokaz_planer(self):
        self.overlay_planer.update_theme(self.is_dark)
        self.overlay_planer.resize(self.main_container.size())
        self.overlay_planer.raise_()
        self.overlay_planer.show()
        self.overlay_planer.pole.setFocus()

    def _planer_zamkniety(self, przystanki):
        """Po zamknięciu planera (✕) wracamy do ekranu powitalnego."""
        self.overlay_planer.hide()
        self._powrot_do_powitalnego()

    def _zaplanuj_wizyty(self, pozycje):
        """Geokoduje punkty i układa plan w tle (wątek), z paskiem postępu i
        logo z okejką (jak generator delegacji). Po sukcesie otwiera Plan Wizyt."""
        if not pozycje:
            return
        import datetime as _dt
        teraz = _dt.date.today()
        rok, miesiac = teraz.year, teraz.month
        mies_txt = self.e_mies.text().strip()
        try:
            if re.match(r"^\d{2}\.\d{4}$", mies_txt):
                miesiac, rok = map(int, mies_txt.split("."))
        except Exception:
            pass

        # pokaż overlay z paskiem postępu (bez logo — pojawi się przy sukcesie)
        self.overlay.show_generating()
        self.overlay.update_status("Planowanie wizyt...")

        # adres bazy — czytamy WPROST z pola (użytkownik mógł go właśnie wpisać
        # i kliknąć Zaplanuj bez opuszczania pola) i od razu zapisujemy
        adres_bazy = ""
        try:
            adres_bazy = self.overlay_planer.pole_baza.text().strip()
        except Exception:
            pass
        if not adres_bazy:
            adres_bazy = ustawienie("adres_bazy", "")
        else:
            zapisz_ustawienie("adres_bazy", adres_bazy)

        # cykliczność — czytamy stan wprost z planera
        cykliczny = False
        horyzont = 3
        try:
            cykliczny = self.overlay_planer.chk_cykl.isChecked()
            horyzont = int(self.overlay_planer.combo_horyzont.currentData() or 3)
        except Exception:
            pass

        self._plan_thread = PlanerWizytThread(
            list(pozycje), rok, miesiac, _osrm_dostepny is False,
            adres_bazy=adres_bazy,
            cykliczny=cykliczny,
            cykl_domyslny=int(ustawienie("cykl_domyslny", 4) or 4),
            cykle_sieci=ustawienie("cykle_sieci", {}) or {},
            cykle_sieci_duze=ustawienie("cykle_sieci_duze", {}) or {},
            cykle_punktow=ustawienie("cykle_punktow", {}) or {},
            horyzont=horyzont)
        self._plan_thread.postep.connect(lambda txt, _v: self.overlay.update_status(txt))
        self._plan_thread.sukces.connect(self._plan_gotowy)
        self._plan_thread.blad.connect(self._plan_blad)
        self._plan_thread.start()

    def _plan_gotowy(self, plan):
        """Sukces planowania — logo z okejką, potem odsłonięcie Planu Wizyt."""
        self._gotowy_plan = plan
        zapisz_plan(plan)                 # plan przetrwa zamknięcie programu
        self.overlay_planer.btn_ostatni.setVisible(True)
        # przygotuj plan POD spodem (jeszcze niewidoczny nad overlayem)
        self.overlay_planer.hide()
        self.overlay_plan.update_theme(self.is_dark)
        self.overlay_plan.resize(self.main_container.size())
        self.overlay_plan.ustaw_plan(plan)
        self.overlay_plan.show()
        # overlay z okejką NA WIERZCHU — po jego animacji (~5.2s) sam się chowa,
        # odsłaniając gotowy plan pod spodem.
        self.overlay.show_success()
        self.overlay.raise_()

    def _pokaz_kopia_zapasowa(self):
        dlg = DialogKopiaZapasowa(self, is_dark=self.is_dark)
        dlg.exec()
        # dialog jest modalny — _nav_klik już schował pulpit ZANIM się otworzył,
        # więc po zamknięciu trzeba go jawnie przywrócić (ten sam mechanizm,
        # co przy planerze/kalendarzu — inaczej zostaje „gołe tło").
        self._powrot_do_powitalnego()

    def _pokaz_ostatni_plan(self):
        """Otwiera ostatnio zapisany plan (bez ponownego planowania)."""
        plan = getattr(self, "_gotowy_plan", None)
        if plan is None:
            plan = wczytaj_plan()
            self._gotowy_plan = plan
        if plan is None:
            self.toast.show_toast("Brak planu",
                                  "Nie ma jeszcze zapisanego planu. Wczytaj punkty i kliknij „Zaplanuj wizyty”.",
                                  success=False)
            # Ta metoda ma DWA wejścia: z menu bocznego (wtedy pulpit jest już
            # schowany przez _nav_klik — bez planu zostałoby „gołe tło") i z
            # przycisku „Ostatni plan” WEWNĄTRZ planera (wtedy planer jest
            # widoczny i nie ma czego naprawiać). Przywracamy pulpit TYLKO
            # w tym pierwszym przypadku.
            if not self.overlay_planer.isVisible():
                self._powrot_do_powitalnego()
            return
        self.overlay_planer.hide()
        self.overlay_plan.update_theme(self.is_dark)
        self.overlay_plan.resize(self.main_container.size())
        self.overlay_plan.ustaw_plan(plan)
        self.overlay_plan.raise_()
        self.overlay_plan.show()

    def _plan_blad(self, komunikat):
        self.overlay.hide_overlay()
        self.toast.show_toast("Błąd planowania", komunikat or "Nie udało się ułożyć planu.", success=False)

    def _przenies_zalegle(self):
        """Przenosi niezrealizowane wizyty na kolejne dni robocze.
        Historia (odhaczone wizyty) pozostaje nietknięta."""
        plan = getattr(self, "_gotowy_plan", None)
        if not plan:
            return
        zal = zalegle_wizyty(plan)
        if not zal:
            self.toast.show_toast("Brak zaległości", "Wszystkie minione wizyty są odhaczone.", success=True)
            return
        # baza (adres domowy) — żeby przeplanowane trasy też startowały z domu
        baza = None
        adres_b = ustawienie("adres_bazy", "").strip()
        if adres_b:
            klucz_b = adres_b.lower().strip()
            lat_b = lng_b = None
            if klucz_b in _geo_cache:
                lat_b, lng_b = tuple(_geo_cache[klucz_b])
            else:
                miasto_b = adres_b.split(",")[-1].strip() if "," in adres_b else adres_b
                miasto_b = re.sub(r"^\d{2}-\d{3}\s*", "", miasto_b).strip()
                c = coords_z_miasta(miasto_b)
                if c:
                    lat_b, lng_b = c
            if lat_b is not None:
                baza = PunktWizyty("Baza (start)", adres_b, lat=lat_b, lng=lng_b)
        nowy = przeplanuj_zalegle(plan, baza=baza, offline=(_osrm_dostepny is False))
        self._gotowy_plan = nowy
        zapisz_plan(nowy)
        self.overlay_plan.ustaw_plan(nowy)
        n = len(zal)
        self.toast.show_toast(
            "Zaległe przeniesione",
            f"{n} {'wizyta' if n==1 else ('wizyty' if 2<=n<=4 else 'wizyt')} przeniesiono na kolejne dni robocze.\n"
            f"Historia odhaczonych wizyt pozostała nietknięta.",
            success=True)

    def _pokaz_mape_planu(self, plan):
        """Generuje HTML mapy planu i otwiera w przeglądarce."""
        try:
            folder = os.path.join(tempfile.gettempdir(), "pmt_plan")
            os.makedirs(folder, exist_ok=True)
            sciezka = generuj_mape_planu_html(plan, folder, is_dark=self.is_dark)
            webbrowser.open("file://" + os.path.realpath(sciezka).replace("\\", "/"))
        except Exception as e:
            self.toast.show_toast("Błąd mapy", f"Nie udało się otworzyć mapy:\n{e}", success=False)

    def _powrot_do_powitalnego(self):
        """Wraca do ekranu powitalnego (gdy zamknięto nakładkę, a generator ma
        pozostać ukryty). Zapobiega 'gołemu tłu' po zamknięciu planera/kalendarza."""
        if hasattr(self, "body"):
            self.body.hide()
        for nakladka in [getattr(self, "overlay_planer", None),
                         getattr(self, "overlay_admin", None),
                         getattr(self, "overlay_staty", None),
                         getattr(self, "overlay_plan", None)]:
            if nakladka is not None:
                nakladka.hide()
        self._nav_aktywny = self.btn_nav_kokpit
        self._maluj_nav_belke()
        self.ekran_powitalny.update_theme(self.is_dark)
        self.ekran_powitalny.show()
        self.ekran_powitalny.raise_()
        self.ekran_powitalny.start()

    def _wczytaj_pozycje_z_pliku(self, sciezka):
        """Wspólna logika wczytywania Excel/CSV — używana zarówno przy
        zwykłym imporcie, jak i przy „odśwież z pliku” (rozpoznawanie zmian).
        Zwraca (pozycje, mapa) albo (None, None) przy błędzie — komunikat
        o błędzie jest już wtedy pokazany jako toast."""
        wiersze = []
        try:
            if sciezka.lower().endswith(".csv"):
                import csv
                with open(sciezka, newline="", encoding="utf-8-sig") as f:
                    probka = f.read(2048); f.seek(0)
                    sep = ";" if probka.count(";") > probka.count(",") else ","
                    for w in csv.reader(f, delimiter=sep):
                        wiersze.append(w)
            else:
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    self.toast.show_toast(
                        "Brak biblioteki",
                        "Aby wczytać .xlsx zainstaluj openpyxl (pip install openpyxl) lub użyj pliku .csv.",
                        success=False)
                    return None, None
                wb = load_workbook(sciezka, read_only=True, data_only=True)
                ws = wb.active
                for w in ws.iter_rows(values_only=True):
                    wiersze.append(list(w))
                wb.close()
        except Exception as e:
            self.toast.show_toast("Błąd importu", f"Nie udało się wczytać pliku:\n{e}", success=False)
            return None, None

        wiersze = [w for w in wiersze if any(str(c or "").strip() for c in w)]
        if not wiersze:
            self.toast.show_toast("Pusty plik", "Plik nie zawiera danych.", success=False)
            return None, None

        naglowki = wiersze[0]
        mapa = _dopasuj_kolumny(naglowki)
        if mapa:
            dane_wiersze = wiersze[1:]
        else:
            mapa = {"miasto": 0}
            dane_wiersze = wiersze

        pozycje = []
        for w in dane_wiersze:
            adres, siec = _wiersz_na_adres(w, mapa)
            if adres:
                i_m = mapa.get("miasto")
                miasto = ""
                if i_m is not None and i_m < len(w):
                    miasto = str(w[i_m] if w[i_m] is not None else "").strip()
                pozycje.append({"adres": adres, "siec": siec, "miasto": miasto})
        if not pozycje:
            self.toast.show_toast(
                "Nie znaleziono adresów",
                "Nie rozpoznałem kolumn z lokalizacjami. Upewnij się, że plik ma nagłówki (np. Miasto, Ulica).",
                success=False)
            return None, None
        return pozycje, mapa

    def _planer_import_excel(self, planer):
        """Inteligentny import lokalizacji z pliku Excel/CSV. Sam rozpoznaje
        kolumny (miasto, ulica, numer, kod, sieć) po nagłówkach — niezależnie
        od ich nazw i kolejności. Buduje pełne adresy i dodaje jako przystanki.
        .csv działa bez zależności; .xlsx wymaga openpyxl."""
        sciezka, _ = QFileDialog.getOpenFileName(
            self, "Wybierz plik z lokalizacjami", "",
            "Pliki z danymi (*.xlsx *.csv);;Excel (*.xlsx);;CSV (*.csv)")
        if not sciezka:
            return
        pozycje, mapa = self._wczytaj_pozycje_z_pliku(sciezka)
        if pozycje is None:
            return

        duplikaty = planer.dodaj_wiele(pozycje)
        # komunikat: co rozpoznano
        role_pl = {"miasto": "miasto", "ulica": "ulica", "numer": "nr",
                   "kod": "kod", "siec": "sieć"}
        rozpoznane = ", ".join(role_pl[r] for r in ["miasto", "ulica", "numer", "kod", "siec"] if r in mapa)
        dodanych = len(pozycje) - (duplikaty or 0)
        tresc = f"Dodano {dodanych} lokalizacji.\nRozpoznane kolumny: {rozpoznane or 'pierwsza kolumna'}."
        if duplikaty:
            tresc += f"\n⚠ Pominięto {duplikaty} {'duplikat' if duplikaty==1 else ('duplikaty' if 2<=duplikaty<=4 else 'duplikatów')} (ten sam adres)."
        self.toast.show_toast("Wczytano listę", tresc, success=True)

    def _odswiez_z_pliku(self, planer):
        """„Odśwież z pliku” — rozpoznaje zmiany względem aktualnie
        wczytanej listy (np. nowa wersja od centrali): pokazuje, co nowe
        i co zniknęło (zamknięte sklepy), zamiast po cichu dokładać do listy."""
        sciezka, _ = QFileDialog.getOpenFileName(
            self, "Wybierz zaktualizowaną listę", "",
            "Pliki z danymi (*.xlsx *.csv);;Excel (*.xlsx);;CSV (*.csv)")
        if not sciezka:
            return
        pozycje, mapa = self._wczytaj_pozycje_z_pliku(sciezka)
        if pozycje is None:
            return

        nowe_wg_adresu = {}
        for p in pozycje:
            klucz = (p["adres"] or "").strip().lower()
            if klucz and klucz not in nowe_wg_adresu:
                nowe_wg_adresu[klucz] = p

        obecne_wg_adresu = {(p.get("adres", "") or "").strip().lower(): p for p in planer._przystanki}

        dodane = [p for k, p in nowe_wg_adresu.items() if k not in obecne_wg_adresu]
        zniknely = [p for k, p in obecne_wg_adresu.items() if k not in nowe_wg_adresu]

        if not dodane and not zniknely:
            self.toast.show_toast("Bez zmian", "Lista z pliku jest identyczna z obecną — nic do zaktualizowania.", success=True)
            return

        dlg = DialogRoznicaImportu(self, dodane, zniknely, is_dark=self.is_dark)
        wynik = dlg.pokaz()
        if wynik is None:
            return

        do_usuniecia = {(p.get("adres", "") or "").strip().lower() for p in wynik}
        if dodane:
            planer.dodaj_wiele(dodane)
        if do_usuniecia:
            planer._przystanki = [p for p in planer._przystanki
                                  if (p.get("adres", "") or "").strip().lower() not in do_usuniecia]
            planer._zapisz_punkty()
            planer._przerysuj()
        self.toast.show_toast(
            "Lista zaktualizowana",
            f"Dodano {len(dodane)} nowych punktów, usunięto {len(do_usuniecia)} brakujących. "
            f"Cykle pozostałych punktów zostały bez zmian.",
            success=True)

    def _klik_generuj(self):
        from PyQt6.QtWidgets import QMessageBox
        from PyQt6.QtCore import QTimer
        """Sprawdza komplet danych PRZED uruchomieniem generowania. Puste lub
        błędne pola zapalają się na czerwono, a program mówi wprost, czego brakuje —
        zamiast wyrzucać komunikat dopiero w połowie pracy."""
        braki = []

        do_podswietlenia = []

        def _sprawdz(pole_si, wartosc, nazwa, warunek=None):
            zle = (not str(wartosc).strip()) or (warunek is not None and not warunek)
            stan = "err" if zle else "ok"
            pole_si.ustaw_walidacje(stan)
            do_podswietlenia.append((pole_si, stan))
            if zle:
                braki.append(nazwa)
            return not zle

        _sprawdz(self.si_imie, self.e_imie.text(), "imię i nazwisko")
        _pesel = self.e_pesel.text().strip()
        _sprawdz(self.si_pesel, _pesel, "PESEL", warunek=(len(_pesel) == 11 and _pesel.isdigit()))
        _sprawdz(self.si_adres, self.e_adres.text(), "adres zamieszkania")
        _kw = self.e_kwota.text().replace(",", ".").replace(" ", "")
        try:
            _kw_ok = float(_kw) > 0
        except Exception:
            _kw_ok = False
        _sprawdz(self.si_kwota, self.e_kwota.text(), "kwota docelowa", warunek=_kw_ok)
        _mies = self.e_mies.text().strip()
        _mies_ok = bool(re.fullmatch(r"\d{2}\.\d{4}", _mies)) and 1 <= int(_mies[:2]) <= 12
        _sprawdz(self.si_mies, _mies, "miesiąc rozliczenia (format MM.RRRR)", warunek=_mies_ok)

        # Asystent podpowiedzi odświeża pola z opóźnieniem i potrafi zgasić nasze
        # czerwone ramki — dlatego nakładamy je ponownie chwilę później.
        def _przypomnij():
            for _pole, _stan in do_podswietlenia:
                _pole.ustaw_walidacje(_stan)
        QTimer.singleShot(600, _przypomnij)
        QTimer.singleShot(1400, _przypomnij)

        if braki:
            mb = QMessageBox(self)
            mb.setWindowTitle("Uzupełnij dane")
            mb.setText("Nie mogę wygenerować dokumentu.")
            mb.setInformativeText("Popraw zaznaczone na czerwono pola:\n• " + "\n• ".join(braki))
            mb.setIcon(QMessageBox.Icon.Warning)
            mb.exec()
            return
        self.proces()

    def _odswiez_przycisk_dni(self):
        ile = len(getattr(self, "dni_wylaczone", ()))
        self.btn_wylacz_dni.setText(f"\U0001F4C5  Wyłączone: {ile}" if ile else "\U0001F4C5  Wyłącz dni")
        # Bez wyłączonych dni: wyraźny bursztynowy obrys (żeby przycisk nie ginął).
        # Z wyłączonymi dniami: pełne bursztynowe wypełnienie — widać na pierwszy rzut oka.
        if ile:
            self.btn_wylacz_dni.setStyleSheet(
                "QPushButton { background: qlineargradient(x1:0,y1:0,x2:1,y2:0,"
                " stop:0 #FBBF24, stop:1 #F59E0B); color:#1A1206; border:none;"
                " border-radius:8px; font-family:'Segoe UI'; font-size:13px;"
                " font-weight:800; padding:7px 14px; }"
                "QPushButton:hover { background:#FCD34D; }")
        else:
            self.btn_wylacz_dni.setStyleSheet(
                "QPushButton { background: rgba(251,191,36,0.10); color:#FBBF24;"
                " border:1.5px solid rgba(251,191,36,0.65); border-radius:8px;"
                " font-family:'Segoe UI'; font-size:13px; font-weight:800;"
                " padding:7px 14px; }"
                "QPushButton:hover { background:#FBBF24; color:#1A1206; }")

    def _otworz_wylaczanie_dni(self):
        """Kalendarz miesiąca rozliczenia: zaznacz dni, w których nie pracujesz.
        Wyłączone dni nie dostaną tras, a kwota rozłoży się na pozostałe."""
        from PyQt6.QtWidgets import (QDialog, QVBoxLayout, QHBoxLayout, QLabel,
                                     QPushButton, QGridLayout, QFrame)
        from PyQt6.QtCore import Qt as _Qt
        import calendar as _cal
        tekst = self.e_mies.text().strip()
        if not re.fullmatch(r"\d{2}\.\d{4}", tekst):
            self.si_mies.ustaw_walidacje("err")
            return
        mies, rok = int(tekst[:2]), int(tekst[3:])
        ustaw_tryb_pracy("wieczory" if self.tryb_wybrany == 1 else "tydzien")
        dostepne = set(pobierz_dni_robocze(rok, mies))

        d = QDialog(self); d.setWindowTitle("Wyłącz dni")
        d.setModal(True); d.setObjectName("PmtKalendarz")
        d.setStyleSheet("""
            #PmtKalendarz { background:#0F172A; }
            QLabel { color:#F8FAFC; font-family:'Segoe UI'; }
            QLabel#naglowek { font-size:15px; font-weight:800; }
            QLabel#pod { color:#94A3B8; font-size:11.5px; }
            QLabel#dow { color:#94A3B8; font-size:10.5px; font-weight:700; }
            QPushButton#dzien { border-radius:6px; font-family:'Segoe UI';
                font-size:12.5px; font-weight:700; min-width:38px; min-height:32px; }
            QPushButton#ok { background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                stop:0 #00F0FF, stop:1 #00E4A1); color:#050B14; border:none;
                border-radius:8px; font-weight:800; padding:8px 20px; }
            QPushButton#anuluj { background:transparent; color:#94A3B8;
                border:1px solid rgba(255,255,255,0.22); border-radius:8px; padding:8px 16px; }
        """)
        ukl = QVBoxLayout(d); ukl.setContentsMargins(20, 16, 20, 14); ukl.setSpacing(6)
        nag = QLabel(f"{_cal.month_name[mies].capitalize()} {rok}"); nag.setObjectName("naglowek")
        ukl.addWidget(nag)
        pod = QLabel("Stuknij dni, w których NIE pracujesz (urlop, choroba, szkolenie).\n"
                     "Szare dni są niedostępne w tym trybie pracy.")
        pod.setObjectName("pod"); ukl.addWidget(pod)
        siatka = QGridLayout(); siatka.setSpacing(4)
        for i, nazwa in enumerate(["Pn","Wt","Śr","Cz","Pt","So","Nd"]):
            et = QLabel(nazwa); et.setObjectName("dow")
            et.setAlignment(_Qt.AlignmentFlag.AlignCenter); siatka.addWidget(et, 0, i)
        wybrane = set(self.dni_wylaczone)
        def _styl(btn, data):
            if data not in dostepne:
                btn.setStyleSheet("QPushButton#dzien { background:transparent;"
                                  " color:#334155; border:1px solid rgba(255,255,255,0.08); }")
            elif data in wybrane:
                btn.setStyleSheet("QPushButton#dzien { background:#DC2626; color:#fff;"
                                  " border:1px solid #DC2626; }")
            else:
                btn.setStyleSheet("QPushButton#dzien { background:rgba(0,228,161,0.14);"
                                  " color:#6EE7B7; border:1px solid rgba(0,228,161,0.45); }")
        for tydzien_nr, tydzien in enumerate(_cal.Calendar().monthdatescalendar(rok, mies), start=1):
            for kol, data in enumerate(tydzien):
                if data.month != mies:
                    continue
                b = QPushButton(str(data.day)); b.setObjectName("dzien")
                if data in dostepne:
                    b.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
                    def _klik(_=False, dd=data, bb=b):
                        if dd in wybrane: wybrane.discard(dd)
                        else: wybrane.add(dd)
                        _styl(bb, dd)
                    b.clicked.connect(_klik)
                else:
                    b.setEnabled(False)
                _styl(b, data)
                siatka.addWidget(b, tydzien_nr, kol)
        ukl.addLayout(siatka)
        info = QLabel(""); info.setObjectName("pod"); ukl.addWidget(info)
        rzad = QHBoxLayout()
        b_czysc = QPushButton("Wyczyść"); b_czysc.setObjectName("anuluj")
        rzad.addWidget(b_czysc); rzad.addStretch(1)
        b_anuluj = QPushButton("Anuluj"); b_anuluj.setObjectName("anuluj")
        b_ok = QPushButton("Zapisz"); b_ok.setObjectName("ok")
        rzad.addWidget(b_anuluj); rzad.addSpacing(8); rzad.addWidget(b_ok)
        ukl.addLayout(rzad)
        def _czysc():
            wybrane.clear()
            for i in range(siatka.count()):
                w = siatka.itemAt(i).widget()
                if isinstance(w, QPushButton):
                    try: _styl(w, [dd for dd in dostepne if dd.day == int(w.text())][0])
                    except Exception: pass
        b_czysc.clicked.connect(_czysc)
        b_anuluj.clicked.connect(d.reject)
        b_ok.clicked.connect(d.accept)
        if d.exec():
            self.dni_wylaczone = set(wybrane)
            self._odswiez_przycisk_dni()

    def _odswiez_przyciski_trybu(self):
        """Wybrany tryb świeci gradientem systemu, drugi jest wygaszony."""
        akt = ("QPushButton { background: qlineargradient(x1:0,y1:0,x2:1,y2:0,"
               " stop:0 #00F0FF, stop:1 #00E4A1); color:#050B14; border:none;"
               " border-radius:8px; font-family:'Segoe UI'; font-size:13px;"
               " font-weight:800; padding:8px 14px; }")
        nieakt = ("QPushButton { background: rgba(255,255,255,0.05); color:%s;"
                  " border:1px solid rgba(255,255,255,0.28); border-radius:8px;"
                  " font-family:'Segoe UI'; font-size:13px; font-weight:600;"
                  " padding:8px 14px; }"
                  "QPushButton:hover { border-color:#00E4A1; color:#00E4A1; }"
                  % ("#94A3B8" if self.is_dark else "#475569"))
        self.btn_tryb_tydzien.setStyleSheet(akt if self.tryb_wybrany == 0 else nieakt)
        self.btn_tryb_wieczory.setStyleSheet(akt if self.tryb_wybrany == 1 else nieakt)

    def _wyloguj_uzytkownika(self):
        """Wylogowanie: kasuje zapamiętany kod i skrót hasła, zamyka program.
        Przy kolejnym uruchomieniu trzeba podać login i hasło."""
        kod = online_kod_uzytkownika() or "—"
        mb = QMessageBox(self)
        mb.setWindowTitle("Wylogowanie")
        mb.setText(f"Wylogować użytkownika {kod}?")
        mb.setInformativeText("Pojawi się ekran logowania — możesz od razu zalogować się "
                              "na inne konto. Plany i dokumenty zostają nienaruszone.")
        mb.setIcon(QMessageBox.Icon.Question)
        mb.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        mb.setDefaultButton(QMessageBox.StandardButton.No)
        mb.button(QMessageBox.StandardButton.Yes).setText("Wyloguj")
        mb.button(QMessageBox.StandardButton.No).setText("Anuluj")
        if mb.exec() != QMessageBox.StandardButton.Yes:
            return
        try:
            online_synchronizuj()      # oddaj liczniki poprzedniego użytkownika
        except Exception:
            pass
        online_wyloguj()
        self.hide()
        kod, imie = dialog_logowania()          # od razu ekran logowania
        if not kod:
            QApplication.quit()
            return
        online_zapisz_kod(kod)
        online_zdarzenie(uruchomienia=1)
        online_synchronizuj_w_tle()
        self.show()
        self.raise_(); self.activateWindow()
        try:
            self.toast.show_toast("Zalogowano",
                                  ("Witaj, " + imie.split()[0] + "!") if imie else ("Kod " + kod),
                                  success=True)
        except Exception:
            pass

    def _pokaz_panel_admina(self):
        """Panel administratora — statystyki wszystkich użytkowników programu
        na tym komputerze. Bez hasła (lokalny, tylko dla właściciela komputera)."""
        self.overlay_admin.update_theme(self.is_dark)
        self.overlay_admin.resize(self.main_container.size())
        self.overlay_admin.odswiez_dane()
        self.overlay_admin.raise_()
        self.overlay_admin.show()

    def _pokaz_statystyki(self):
        """'Twoja praca w liczbach' — osobisty panel zbudowany na
        rzeczywistej historii z dziennika (ile wizyt, jaki rytm, sieci) oraz,
        w drugiej zakładce, z historii delegacji (dawne 'Moje Szlaki')."""
        self.overlay_staty._delegacje_wczytane = False   # świeże dane przy każdym otwarciu
        self.overlay_staty._przelacz_zakladke("wizyty")  # zawsze startuj od głównej zakładki
        self.overlay_staty.update_theme(self.is_dark)
        self.overlay_staty.resize(self.main_container.size())
        self.overlay_staty.odswiez_dane()
        self.overlay_staty.raise_()
        self.overlay_staty.show()

    def _otworz_zwiad(self):
        if self._last_folder:
            hp = os.path.join(self._last_folder, "Trasy_Mapa.html")
            if os.path.exists(hp):
                webbrowser.open("file://" + os.path.realpath(hp).replace('\\', '/'))
                return
        self.toast.show_toast("Eksploracja",
                              "Mapa tras pojawi się po pierwszym wygenerowaniu rozliczenia.",
                              success=True)

    def _otworz_archiwum(self):
        """Archiwum — pokazuje historię TEGO pracownika (per-użytkownik) i otwiera
        ostatni folder. Historia jest prywatna: wymaga wpisanego imienia+PESEL-u."""
        imie = self.e_imie.text().strip()
        pesel = self.e_pesel.text().strip()
        historia = wczytaj_historie(imie, pesel) if (imie and pesel) else []
        if historia:
            ostatni = historia[0]
            self.toast.show_toast(
                "Ostatnio wygenerowano",
                f"{ostatni.get('imie','')}  •  {ostatni.get('data','')}\n"
                f"{ostatni.get('kwota','')} zł  •  {ostatni.get('dokumenty','?')} dok.  •  {ostatni.get('woj','')}",
                success=True,
                klik_akcja=(lambda f=ostatni.get('folder'): self._otworz_folder(f))
            )
        else:
            target = self._last_folder if (self._last_folder and os.path.isdir(self._last_folder)) \
                     else sciezka_pulpitu()
            self.toast.show_toast("Archiwum",
                                  "Brak historii dla tych danych.\nWpisz swoje imię i PESEL, aby zobaczyć własne rozliczenia." if not (imie and pesel) else "Brak zapisanych rozliczeń — wygeneruj pierwsze.",
                                  success=True,
                                  klik_akcja=(lambda t=target: self._otworz_folder(t)))

    def _otworz_folder(self, folder):
        if folder and os.path.isdir(folder):
            try: otworz_w_systemie(folder)
            except Exception: pass

    def set_form_enabled(self, enabled):
        self.e_imie.setEnabled(enabled); self.e_pesel.setEnabled(enabled)
        self.e_adres.setEnabled(enabled); self.c_stan.setEnabled(enabled)
        self.e_kwota.setEnabled(enabled); self.e_mies.setEnabled(enabled)
        self.c_silnik.setEnabled(enabled)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self.grip.move(self.width() - self.grip.width(), self.height() - self.grip.height())
        self.btn_settings.move(self.sidebar_frame.width() - 38, self.height() - 110)
        self.overlay.resize(self.main_container.size())
        self.overlay_planer.resize(self.main_container.size())
        if hasattr(self, "overlay_admin"):
            self.overlay_admin.resize(self.main_container.size())
        if hasattr(self, "overlay_staty"):
            self.overlay_staty.resize(self.main_container.size())
        if hasattr(self, "overlay_plan"):
            self.overlay_plan.resize(self.main_container.size())
        if hasattr(self, "overlay_tryb_trasy"):
            self.overlay_tryb_trasy.resize(self.main_container.size())
        if hasattr(self, "panel_powiadomien") and self.panel_powiadomien.isVisible():
            self.panel_powiadomien.hide()

    def mousePressEvent(self, event):
        super().mousePressEvent(event)
        # Klik poza panelem powiadomień i poza dzwoneczkiem — zamknij panel.
        if hasattr(self, "panel_powiadomien") and self.panel_powiadomien.isVisible():
            pt_panel = self.panel_powiadomien.mapFromGlobal(event.globalPosition().toPoint())
            pt_btn = self.btn_dzwonek.mapFromGlobal(event.globalPosition().toPoint())
            if not self.panel_powiadomien.rect().contains(pt_panel) and \
               not self.btn_dzwonek.rect().contains(pt_btn):
                self.panel_powiadomien.hide()

    def toggle_theme(self):
        self.is_dark = not self.is_dark
        self.apply_theme()

    def apply_theme(self):
        self.main_container.set_theme(self.is_dark)
        self.title_bar.update_theme(self.is_dark)

        self.btn_theme.update_theme(self.is_dark)
        self.btn_dzwonek.is_dark = self.is_dark; self.btn_dzwonek.update()
        if hasattr(self, "panel_powiadomien"): self.panel_powiadomien.update_theme(self.is_dark)
        self.si_imie.update_theme(self.is_dark)
        self.si_pesel.update_theme(self.is_dark)
        self.si_adres.update_theme(self.is_dark)
        self.si_stan.update_theme(self.is_dark)
        self.si_kwota.update_theme(self.is_dark)
        self.si_mies.update_theme(self.is_dark)
        self.si_silnik.update_theme(self.is_dark)
        self.btn.update_theme(self.is_dark)
        self.gps_prog.set_theme(self.is_dark)
        self.overlay_planer.update_theme(self.is_dark)
        if hasattr(self, "ekran_powitalny"):
            self.ekran_powitalny.update_theme(self.is_dark)

        if self.is_dark:
            c_text_hi = "#F8FAFC"
            c_text_med = "#94A3B8"
            c_accent = "#00F0FF"
            # Gradient tytułów sekcji = gradient tła logo PMT
            c_title_grad = "qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 #00F0FF, stop:1 #00E4A1)"
            self.topbar.setStyleSheet("border-bottom: 1px solid rgba(255, 255, 255, 0.05); background: transparent;")
            if hasattr(self, "logo_lbl") and hasattr(self.logo_lbl, "update_theme"):
                self.logo_lbl.update_theme(True)

            self.sidebar_frame.setStyleSheet("QFrame { background-color: rgba(10, 18, 30, 0.5); border-right: 2px solid rgba(0, 240, 255, 0.30); border-radius: 0px; }")
            self.btn_settings.setStyleSheet("QPushButton { background: transparent; color: #E2E8F0; border: none; font-size: 56px; } QPushButton:hover { color: #00F0FF; }")
            self.btn_bug.setStyleSheet("QPushButton { background-color: #EF4444; color: white; border: none; border-radius: 14px; padding: 6px 16px; font-family: 'Segoe UI', sans-serif; font-size: 12px; font-weight: bold; } QPushButton:hover { background-color: #DC2626; }")

            card_style = "QFrame { background-color: rgba(10, 18, 30, 0.25); border: 1px solid rgba(0, 240, 255, 0.15); border-radius: 12px; }"
            self.card_top_frame.setStyleSheet(card_style)
            self.card_bot_frame.setStyleSheet(card_style)

            self.bot_card.setStyleSheet("QFrame { background-color: rgba(10, 15, 30, 0.45); border: 1px solid rgba(0, 240, 255, 0.2); border-radius: 12px; }")
            self.ic_user.set_color(c_accent)
            self.ic_map.set_color(c_accent)
            self.status_icon.set_color("#00E4A1")
        else:
            c_text_hi = "#0F172A"
            c_text_med = "#475569"
            c_accent = "#0D9488"
            c_title_grad = "qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 #0D9488, stop:1 #059669)"
            self.topbar.setStyleSheet("border-bottom: 1px solid rgba(0, 0, 0, 0.05); background: transparent;")
            if hasattr(self, "logo_lbl") and hasattr(self.logo_lbl, "update_theme"):
                self.logo_lbl.update_theme(False)

            self.sidebar_frame.setStyleSheet("QFrame { background-color: rgba(255, 255, 255, 0.6); border-right: 2px solid rgba(148, 163, 184, 0.6); border-radius: 0px; }")
            self.btn_settings.setStyleSheet("QPushButton { background: transparent; color: #0F172A; border: none; font-size: 56px; } QPushButton:hover { color: #0D9488; }")
            self.btn_bug.setStyleSheet("QPushButton { background-color: #DC2626; color: white; border: none; border-radius: 14px; padding: 6px 16px; font-family: 'Segoe UI', sans-serif; font-size: 12px; font-weight: bold; } QPushButton:hover { background-color: #B91C1C; }")

            card_style_light = "QFrame { background-color: rgba(255, 255, 255, 0.4); border: 1px solid rgba(148, 163, 184, 0.4); border-radius: 12px; }"
            self.card_top_frame.setStyleSheet(card_style_light)
            self.card_bot_frame.setStyleSheet(card_style_light)

            self.bot_card.setStyleSheet("QFrame { background-color: rgba(255, 255, 255, 0.6); border: 1px solid rgba(200, 210, 220, 0.8); border-radius: 12px; }")
            self.ic_user.set_color(c_accent)
            self.ic_map.set_color(c_accent)
            self.status_icon.set_color("#10B981")

        # Separator sidebara, panel asystenta, oś czasu
        self.sidebar_sep.setStyleSheet(
            f"background-color: {'rgba(0,240,255,0.15)' if self.is_dark else 'rgba(15,23,42,0.12)'}; border: none; margin: 6px 8px;")
        self.assistant.set_theme(self.is_dark)
        self.timeline.set_theme(self.is_dark)

        # Ikony przycisków nawigacji + animowana belka aktywnego
        for b, icon in self._nav_defs:
            pix = QPixmap(30, 30); pix.fill(Qt.GlobalColor.transparent)
            pnt = QPainter(pix)
            draw_svg_icon(pnt, icon, QRectF(1, 1, 28, 28), QColor(c_accent), self.is_dark)
            pnt.end()
            b.setIcon(QIcon(pix)); b.setIconSize(pix.size())
        self._maluj_nav_belke()

        self.t1.setStyleSheet(f"font-family: 'Segoe UI', sans-serif; font-size: 20px; font-weight: 700; color: {c_text_hi}; border: none; background: transparent;")
        # Tytuły sekcji w gradiencie logo PMT
        title_grad_style = f"font-family: 'Segoe UI', sans-serif; font-size: 14px; font-weight: 800; color: {c_title_grad}; letter-spacing: 1.5px; background: transparent; border: none;"
        self.lbl_u_title.setStyleSheet(title_grad_style)
        self.lbl_m_title.setStyleSheet(title_grad_style)
        self.lbl_u_desc.setStyleSheet(f"font-family: 'Segoe UI', sans-serif; font-size: 11px; color: {c_text_med}; background: transparent; border: none;")
        self.lbl_m_desc.setStyleSheet(f"font-family: 'Segoe UI', sans-serif; font-size: 11px; color: {c_text_med}; background: transparent; border: none;")

        lbl_style = f"font-family: 'Segoe UI', sans-serif; font-size: 11px; font-weight: 600; color: {c_text_hi}; background: transparent; border: none;"
        self.l_imie.setStyleSheet(lbl_style); self.l_pesel.setStyleSheet(lbl_style); self.l_adres.setStyleSheet(lbl_style)
        self.l_stan.setStyleSheet(lbl_style); self.l_kwota.setStyleSheet(lbl_style); self.l_mies.setStyleSheet(lbl_style)
        self.l_silnik.setStyleSheet(lbl_style)

        self.lbl_status.setStyleSheet(f"color:{c_text_hi}; font-family:'Segoe UI', sans-serif; font-size:12px; font-weight:500; background: transparent; border: none;")
        self.lbl_pct.setStyleSheet(f"color:{c_text_med}; font-family:'Segoe UI', sans-serif; font-size:12px; font-weight:700; background: transparent; border: none;")

    def reset_ui(self):
        self.timeline.reset()
        self.set_form_enabled(True)

    def status(self, tekst, postep=None):
        # Sam tekst statusu w overlayu; etapy osi czasu sterowane własnym
        # scenariuszem (_scenariusz_timeline), nie surowym postępem wątku —
        # dzięki temu animacja ma stałe, przyjemne tempo niezależnie od
        # tego jak szybko policzył się wynik.
        self.overlay.update_status(tekst)

    def _scenariusz_timeline(self):
        """Odtwarza oś czasu w spokojnym tempie: Walidacja→Trasy→PDF→Mapa.
        Ostatni etap (Gotowe) zapala się dopiero po realnym zakończeniu wątku."""
        self._tl_step = 0
        # (docelowy_etap, ile_ms_nalewać, ile_ms_pauzy_po)
        self._tl_plan = [
            (0, 500, 250),   # Walidacja
            (1, 850, 350),   # Trasy (najdłużej — to serce liczenia)
            (2, 700, 300),   # PDF
            (3, 650, 250),   # Mapa
        ]
        self._tl_gotowe = False          # ustawiane przez wątek gdy skończy
        self._tl_wynik = None            # (finalne_dni, pracownik, folder)
        self._graj_kolejny_etap()

    def _graj_kolejny_etap(self):
        if self._tl_step < len(self._tl_plan):
            idx, czas, pauza = self._tl_plan[self._tl_step]
            self.timeline.ustaw_etap(idx, czas)
            self._tl_step += 1
            QTimer.singleShot(czas + pauza, self._graj_kolejny_etap)
        else:
            # doszliśmy do Mapy — czekamy aż wątek zgłosi sukces
            self._sprawdz_gotowosc()

    def _sprawdz_gotowosc(self):
        if self._tl_gotowe and self._tl_wynik is not None:
            self._finalizuj_sukces(*self._tl_wynik)
        else:
            QTimer.singleShot(120, self._sprawdz_gotowosc)   # dopinguj co 120ms

    def _enter_generuj(self):
        """Enter zatwierdza formularz. Jeśli wszystko OK — generuje.
        Jeśli czegoś brakuje — przenosi kursor do pierwszego niegotowego pola."""
        if self.btn.is_loading if hasattr(self.btn, 'is_loading') else False:
            return
        braki = [
            (self.e_imie,  not self.e_imie.text().strip()),
            (self.e_pesel, not waliduj_pesel(self.e_pesel.text().strip())),
            (self.e_adres, not self.e_adres.text().strip()),
            (self.e_kwota, not self.e_kwota.text().strip()),
            (self.e_mies,  not re.match(r"^\d{2}\.\d{4}$", self.e_mies.text().strip().replace('-', '.').replace('/', '.').replace(' ', ''))),
        ]
        for pole, brak in braki:
            if brak:
                pole.setFocus(); pole.selectAll()
                return
        self.proces()

    def proces(self):
        self.btn.start_loading()
        self.set_form_enabled(False)
        self.timeline.reset()

        try:
            imie  = ' '.join(w.capitalize() for w in self.e_imie.text().split())
            pesel = self.e_pesel.text().strip()
            adres = self.e_adres.text().strip()
            mies_s  = self.e_mies.text().strip().replace('-', '.').replace('/', '.').replace(' ', '')
            kwota_s = self.e_kwota.text().strip().replace(' ', '').replace(',', '.')

            if not all([imie, pesel, adres, mies_s, kwota_s]): raise ValueError("Proszę uzupełnić wszystkie pola formularza.")
            if not waliduj_pesel(pesel): raise ValueError("Wprowadzony numer PESEL jest nieprawidłowy.")

            adres_d = waliduj_adres(adres)
            # Adres niejednoznaczny (np. "Zielona Ścieżka" — ulica czy wieś?).
            # Zamiast zgadywać, pytamy użytkownika i ponawiamy z jego decyzją.
            if adres_d.get('niejednoznaczne'):
                nazwa_q = adres_d.get('nazwa_do_pytania', '')
                dlg = DialogWyboru(
                    self,
                    "DOPRECYZUJ ADRES",
                    f"„{nazwa_q}” — to nazwa ulicy czy miejscowości?",
                    "",
                    "To ulica", "To miejscowość",
                    is_dark=self.is_dark)
                wynik = dlg.exec_wybor()
                wybor = 'ulica' if wynik == 'a' else 'wies'
                adres_d = waliduj_adres(adres, wymus_typ=wybor)
            mies, rok = waliduj_miesiac(mies_s)
            kwota = waliduj_kwote(kwota_s)
            woj = rozpoznaj_wojewodztwo(adres_d['kod_pocztowy'])

            # Tryb pracy MUSI być ustawiony przed pobraniem dni — od niego
            # zależy, czy w planie znajdą się soboty i niedziele handlowe,
            # oraz jak długi jest dzień pracy.
            ustaw_tryb_pracy("wieczory" if self.tryb_wybrany == 1 else "tydzien")
            dni = pobierz_dni_robocze(rok, mies)
            # Dni oznaczone jako wolne wypadają z planu — kwota rozłoży się
            # na pozostałe dni miesiąca.
            _wyl = getattr(self, "dni_wylaczone", set())
            if _wyl:
                dni = [d for d in dni if d not in _wyl]
                if not dni:
                    raise ValueError("Wyłączyłeś wszystkie dni w tym miesiącu.\n"
                                     "Odznacz przynajmniej jeden dzień.")
            # Każdy dzień pracy = maksymalnie jeden limit delegacji (~587 zł).
            # Górna granica kwoty to liczba dni roboczych × limit delegacji.
            _max_kwota = len(dni) * MAX_KWOTA_DELEGACJI
            if kwota > _max_kwota:
                raise ValueError(
                    f"Kwota za wysoka na ten miesiąc.\n\n"
                    f"Przy {len(dni)} dniach roboczych maksymalna kwota to "
                    f"{_max_kwota:,.0f} zł (limit {MAX_KWOTA_DELEGACJI:.0f} zł na dzień).\n"
                    f"Zmniejsz kwotę lub wybierz miesiąc z większą liczbą dni.")

            stawka = 0.89 if self.c_silnik.currentIndex() == 0 else 1.15

            # Zapamiętaj profil pracownika (prywatny, kluczowany nazwisko+PESEL)
            zapisz_profil(imie, pesel, adres_d['adres_caly'], self.c_stan.currentText(), self.c_silnik.currentIndex())

            self._params = {
                'imie': imie, 'pesel': pesel, 'adres_caly': adres_d['adres_caly'],
                'adres_geo': adres_d.get('adres_geo', adres_d['adres_caly']),
                'kod_pocztowy': adres_d['kod_pocztowy'], 'baza_miasto': adres_d['baza_miasto'],
                'stanowisko': self.c_stan.currentText(), 'kwota_cel': kwota,
                'miesiac': mies, 'rok': rok, 'miesiac_slownie': MIESIACE_PL[mies - 1],
                'woj': woj, 'dni_robocze': dni, 'is_dark': self.is_dark, 'stawka': stawka,
            }

            self.overlay.show_generating()
            self._scenariusz_timeline()          # start spokojnej animacji osi

            self._thread = GeneratorThread(self._params)
            self._thread.postep.connect(self._on_progress)
            self._thread.sukces.connect(self._na_sukces)
            self._thread.blad.connect(self._na_blad)
            self._thread.start()

        except Exception as e:
            self.btn.stop_loading()
            self.reset_ui()
            self.overlay.hide_overlay()
            self.toast.show_toast("Błąd Walidacji", str(e))

    def _on_progress(self, tekst, postep):
        self.status(tekst, postep)

    def _na_blad(self, msg):
        self.overlay.hide_overlay()
        self.btn.stop_loading()
        self.reset_ui()
        self.toast.show_toast("Błąd Generowania", str(msg))

    def _na_sukces(self, finalne_dni, pracownik, folder):
        # Wątek skończył — ale finalizację odpala scenariusz osi czasu,
        # gdy tylko animacja dojedzie do etapu "Mapa". Jeśli już tam jest,
        # _sprawdz_gotowosc wychwyci to natychmiast.
        self._tl_wynik = (finalne_dni, pracownik, folder)
        self._tl_gotowe = True

    def _finalizuj_sukces(self, finalne_dni, pracownik, folder):
        self.timeline.ustaw_etap(4, 550)   # Gotowe — cała oś zapalona, płynnie
        self.btn.stop_loading()
        self._last_folder = folder

        # Zapis do historii (prywatnej dla tego pracownika)
        p = self._params
        suma = round(sum(d.suma for d in finalne_dni), 2)
        # łączne kilometry z tras (realny dystans po drogach)
        km_total = 0.0
        for d in finalne_dni:
            for e in d.etapy_surowe:
                km_total += getattr(e, 'dystans_rzeczywisty', e.d_line)
        # Zbierz WSZYSTKIE odwiedzone województwa (nie tylko bazę!) — z etapów.
        # Dzięki temu analityka pokaże realny udział regionów, także sąsiednich.
        woj_wizyty = {}
        for d in finalne_dni:
            for e in d.etapy_surowe:
                w = getattr(e, 'dokad_woj', '') or ''
                w = w.strip()
                if w:                                  # pomijamy powroty do bazy (puste)
                    klucz = w.capitalize()
                    woj_wizyty[klucz] = woj_wizyty.get(klucz, 0) + 1
        # awaryjnie: gdyby etapy nie miały województw, użyj bazy
        if not woj_wizyty:
            woj_wizyty = {p['woj'].capitalize(): 1}

        # Zbierz ODWIEDZANE MIEJSCOWOŚCI (z etapów), POMIJAJĄC bazę startową/
        # końcową — to punkty pośrednie trasy. Do wykresu "top miejscowości".
        baza_nazwa = (p.get('baza_miasto', '') or '').strip().lower()
        miejsc_wizyty = {}
        for d in finalne_dni:
            for e in d.etapy_surowe:
                cel = (getattr(e, 'dokad', '') or '').strip()
                if cel and cel.lower() != baza_nazwa:
                    miejsc_wizyty[cel] = miejsc_wizyty.get(cel, 0) + 1
        # Konkretne DATY dni wyjazdowych (ISO) — do oznaczenia w kalendarzu.
        # Delegacja obejmuje wiele dni jazdy; zapisujemy każdy z nich.
        dni_daty = []
        for d in finalne_dni:
            try:
                dni_daty.append(d.data.isoformat())     # "2026-07-01"
            except Exception:
                pass
        try:
            dodaj_do_historii(p['imie'], p['pesel'], {
                "imie": p['imie'],
                "data": datetime.datetime.now().strftime("%d.%m.%Y %H:%M"),
                "kwota": f"{suma:.2f}",
                "woj": p['woj'].capitalize(),          # baza (dla zgodności wstecz)
                "woj_wizyty": woj_wizyty,              # WSZYSTKIE odwiedzone regiony
                "baza": p.get('baza_miasto', '') or '', # miejscowość bazowa
                "miejsc_wizyty": miejsc_wizyty,        # odwiedzane miejscowości (bez bazy)
                "dokumenty": math.ceil(suma / 590.0) if suma else 0,
                "km": round(km_total),
                "miesiac": p['miesiac'],
                "rok": p['rok'],
                "dni_wyjazdowe": len(finalne_dni),
                "dni_daty": dni_daty,                  # daty dni wyjazdowych (ISO)
                "folder": folder,
            })
        except Exception as e:
            log_error(e)

        # Najpierw pasek dojeżdża do "Gotowe" (~750ms), DOPIERO POTEM wjeżdża
        # logo z rysującą się okejką — jako finałowa gratulacja.
        QTimer.singleShot(750, self.overlay.show_success)
        # 750ms pasek + ~5.2s efekt (orbita domyka pełny krąg + checkmark) przed otwarciem
        QTimer.singleShot(6100, lambda: self._zakoncz_po_sukcesie(folder))

    def _zakoncz_po_sukcesie(self, folder):
        self.reset_ui()
        # Jeśli realnymi trasami nie dało się pokryć pełnej kwoty (za wysoka na
        # ten miesiąc/rejon) — informujemy uczciwie, zamiast pompować trasy.
        if getattr(self, "_thread", None) and getattr(self._thread, "_kwota_niepelna", False):
            osiag = getattr(self._thread, "_kwota_osiagnieta", 0.0)
            cel = getattr(self._thread, "_kwota_docelowa", 0.0)
            self.toast.show_toast(
                "Kwota częściowo rozpisana",
                f"Przy realnych trasach w tym miesiącu udało się rozpisać "
                f"{osiag:,.0f} zł z {cel:,.0f} zł.\n".replace(","," ") +
                "Wszystkie dostępne dni robocze zostały wykorzystane — pozostała "
                "kwota nie mieści się w realnych trasach. Zmniejsz kwotę lub wybierz "
                "miesiąc z większą liczbą dni.",
                success=False,
                klik_akcja=(lambda f=folder: self._otworz_folder(f))
            )
        else:
            # Toast KLIKALNY — kliknięcie otwiera folder bez dodatkowych okien
            self.toast.show_toast(
                "PDF wygenerowany",
                "Rozliczenie gotowe. Mapa tras otwarła się w przeglądarce.",
                success=True,
                klik_akcja=(lambda f=folder: self._otworz_folder(f))
            )
        html_path = os.path.join(folder, "Trasy_Mapa.html")
        if os.path.exists(html_path):
            webbrowser.open("file://" + os.path.realpath(html_path).replace('\\', '/'))
        try: otworz_w_systemie(folder)
        except: pass


if __name__ == "__main__":
    app = QApplication(sys.argv)
    # Kursor rysujemy sami (GrubyKursorEdt) — systemowy zostawiamy domyślny 1px,
    # żeby cursorRect().center() wskazywał dokładne miejsce między znakami.
    font = QFont("Segoe UI", 10)
    font.setStyleStrategy(QFont.StyleStrategy.PreferAntialias)
    app.setFont(font)

    # --- Logowanie przy KAZDYM uruchomieniu (login + haslo) ---------------
    # Swiadoma zmiana wzgledem wczesniejszych wersji: kod nie jest juz
    # "zapamietywany na zawsze". Kazde wejscie do systemu jest odnotowane
    # w arkuszu (zakladka Log), a dostep mozna odebrac zdalnie - zmiana
    # hasla albo daty w kolumnie "Wazne do" dziala od nastepnego startu.
    _kod, _imie_zal = dialog_logowania()
    if not _kod:
        sys.exit(0)
    online_zapisz_kod(_kod)

    online_zdarzenie(uruchomienia=1)
    _START_PROGRAMU = datetime.datetime.now()

    # --- Kontrola sesji: najpierw arkusz (online), awaryjnie lokalne 30 dni ---
    _wazna_o, _dni_o, _imie_o = online_status_sesji()
    if _wazna_o is None:
        online_synchronizuj()          # pierwszy kontakt z serwerem (jesli jest siec)
        _wazna_o, _dni_o, _imie_o = online_status_sesji()
    elif _wazna_o is False:
        online_synchronizuj()          # moze administrator wlasnie przedluzyl sesje
        _wazna_o, _dni_o, _imie_o = online_status_sesji()
    if _wazna_o is not None:
        _wazna, _pozostalo = _wazna_o, _dni_o
    else:
        _wazna, _pozostalo = demo_status()   # nigdy nie bylo kontaktu z serwerem

    online_synchronizuj_w_tle()
    online_petla_synchronizacji()
    import atexit
    def _pmt_zamkniecie():
        try:
            _min = (datetime.datetime.now() - _START_PROGRAMU).total_seconds() / 60.0
            online_zdarzenie(minuty=_min)
            online_synchronizuj()
        except Exception:
            pass
    atexit.register(_pmt_zamkniecie)
    if not _wazna:
        from PyQt6.QtWidgets import QMessageBox
        _mb = QMessageBox()
        _ico = znajdz_ikone()
        if _ico: _mb.setWindowIcon(QIcon(_ico))
        _mb.setWindowTitle("PMT Planer — wersja DEMO")
        _mb.setIcon(QMessageBox.Icon.Warning)
        _mb.setText("Sesja dobiegła końca.")
        _mb.setInformativeText("Dziękujemy za przetestowanie PMT Planer.\n"
                               "Aby przedłużyć dostęp, skontaktuj się z administratorem — po przedłużeniu wystarczy ponownie uruchomić program (przy dostępie do internetu).")
        _mb.setStandardButtons(QMessageBox.StandardButton.Ok)
        _mb.exec()
        sys.exit(0)

    window = App()
    window._demo_pozostalo = _pozostalo    # do ewentualnego pokazania w UI
    window.show()
    sys.exit(app.exec())